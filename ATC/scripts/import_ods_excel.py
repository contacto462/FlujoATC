"""
Importa las hojas 'Ordenes de Servicio' y 'Administración' del Excel
'FLUJO Alguien Te Cuida.xlsx' a la BBDD.

Tablas destino:
  venta_comercial          ← hoja "Ordenes de Servicio"
  bbdd_clientes            ← upsert de RUT + razón social (FK requerida)
  venta_administracion     ← hoja "Administración" cols Admin
  venta_finanzas           ← hoja "Administración" cols Finanzas
  venta_servicio_tecnico   ← hoja "Administración" cols Servicio Técnico
  venta_soporte_tecnico    ← hoja "Administración" cols Soporte Técnico
  venta_operaciones        ← hoja "Administración" cols Operaciones

Uso:
  /Users/fernando/PROYECTO-ATC/.venv-backend/bin/python \
      /Volumes/PROYECTO-ATC-SERVIDOR/ATC/scripts/import_ods_excel.py
"""
from __future__ import annotations

import sys
import os
from datetime import datetime
from pathlib import Path

# ── Agrega raíz del proyecto al path ─────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

# ── Lee .env del servidor (último valor gana) ────────────────────────────────
# El servidor corre desde /Users/fernando/PROYECTO-ATC/ con su propio .env (PostgreSQL)
_env_path = Path.home() / "PROYECTO-ATC" / "ATC" / ".env"
if not _env_path.exists():
    _env_path = Path(__file__).resolve().parents[1] / ".env"
for line in _env_path.read_text(encoding="utf-8").splitlines():
    line = line.strip()
    if "=" in line and not line.startswith("#"):
        k, _, v = line.partition("=")
        os.environ[k.strip()] = v.strip().strip('"').strip("'")

EXCEL_PATH = Path.home() / "Desktop" / "FLUJO Alguien Te Cuida.xlsx"

# ── Imports del proyecto ──────────────────────────────────────────────────────
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

from ATC.app.core.config import settings
from ATC.app.models.incidencias import (
    ClienteBBDD,
    VentaODS,
    AdministracionODT,
    FinanzasODT,
    ServicioTecnicoVentaODT,
    SoporteTecnicoVentaODT,
    OperacionesVentaODT,
)

# ── Engine ────────────────────────────────────────────────────────────────────
engine = create_engine(str(settings.database_url), echo=False)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _s(v) -> str:
    """Convierte valor de celda a string limpio, o ''."""
    if v is None:
        return ""
    return str(v).strip()


def _date(v) -> datetime | None:
    """Convierte fecha de celda (string dd/mm/yyyy o datetime) a datetime, o None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _bool_date(v) -> tuple[bool, datetime | None]:
    """Si la celda tiene fecha → (True, fecha). Si vacío → (False, None)."""
    d = _date(v)
    return (True, d) if d else (False, None)


def _int(v) -> int | None:
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return None


# ── Carga el Excel ────────────────────────────────────────────────────────────

print(f"\nAbriendo {EXCEL_PATH} …")
wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

ws_ods   = wb["Ordenes de Servicio"]
ws_admin = wb["Administración"]

rows_ods   = list(ws_ods.iter_rows(values_only=True))[1:]   # sin cabecera
rows_admin = list(ws_admin.iter_rows(values_only=True))[1:]  # sin cabecera

print(f"  'Ordenes de Servicio': {len(rows_ods)} filas")
print(f"  'Administración':      {len(rows_admin)} filas")

# Construye diccionario por código desde Administración
admin_by_codigo: dict[str, tuple] = {}
for row in rows_admin:
    codigo = _s(row[0])
    if codigo:
        admin_by_codigo[codigo] = row


# ── Importación ───────────────────────────────────────────────────────────────

insertados = skipped = errores = 0

with Session(engine) as db:
    for i, row in enumerate(rows_ods, start=2):
        codigo = _s(row[0])
        if not codigo:
            continue

        # Skip si ya existe
        existe = db.query(VentaODS).filter_by(codigo=codigo).first()
        if existe:
            skipped += 1
            continue

        rut      = _s(row[3])
        razon    = _s(row[4])
        direccion = _s(row[5])

        try:
            # ── 1. Upsert bbdd_clientes ───────────────────────────────────────
            if rut:
                cliente_rec = db.query(ClienteBBDD).filter_by(rut=rut).first()
                if not cliente_rec:
                    cliente_rec = db.query(ClienteBBDD).filter_by(cliente=razon).first()
                if not cliente_rec:
                    cliente_rec = ClienteBBDD(
                        cliente=razon,
                        rut=rut,
                        direccion=direccion or None,
                    )
                    db.add(cliente_rec)
                    db.flush()
                elif not cliente_rec.rut:
                    cliente_rec.rut = rut
                    db.flush()

            # ── 2. venta_comercial ────────────────────────────────────────────
            # Columnas Ordenes de Servicio:
            # 0 Código, 1 Fecha, 2 Ejecutivo, 3 Rut, 4 Razón Social,
            # 5 Dirección Sucursal, 6 Nombre Sucursal, 7 Observación,
            # 8 Tipo de cliente, 9 Tipo Servicio, 10 Tipo de Plan,
            # 11 Nº Cámaras Instalar, 12 Días Grabación, 13 Días Monitoreo Adicional,
            # 14 Horario Monitoreo, 15 Cotización, 16-20 Contratos,
            # 21 Carpeta, 22 Ejecutivo Venta, 25 Materiales, 26 Consideraciones,
            # 27 Agua/Baño, 28 Nº Cámaras Vigilar, 29 ODC, 30 Desglose, 32 Monto

            contrato_parts = []
            for idx, label in [(16, "Televigilancia"), (17, "Alarma"), (18, "Guardia"),
                                (19, "Servicio Técnico"), (20, "Upgrade/Downgrade")]:
                val = _s(row[idx]) if len(row) > idx else ""
                if val:
                    contrato_parts.append(f"{label}: {val}")

            # Estado desde hoja Administración si existe
            admin_row = admin_by_codigo.get(codigo)
            estado_excel = _s(admin_row[47]) if admin_row and len(admin_row) > 47 else ""
            estado = estado_excel or "Registrada"

            # Carpeta Drive
            carpeta_url = _s(row[21]) if len(row) > 21 else ""
            carpeta_id = ""
            if "drive.google.com" in carpeta_url:
                # extrae el ID de la URL de la carpeta
                import re
                m = re.search(r"/folders/([A-Za-z0-9_-]+)", carpeta_url)
                if m:
                    carpeta_id = m.group(1)

            ods = VentaODS(
                codigo=codigo,
                creado_por=_s(row[2]) or None,
                rut_cliente=rut or None,
                razon_social=razon,
                direccion_sucursal=direccion,
                nombre_sucursal=_s(row[6]) or None,
                tipo_cliente=_s(row[8]) or None,
                tipo_servicio=_s(row[9]),
                tipo_plan=_s(row[10]) or None,
                observacion=_s(row[7]) or None,
                numero_camaras_instalar=_int(row[11]) if len(row) > 11 else None,
                dias_grabacion=_int(row[12]) if len(row) > 12 else None,
                dias_monitoreo_adicional=_s(row[13]) or None if len(row) > 13 else None,
                horario_monitoreo=(_s(row[14])[:19] + "…") if len(_s(row[14] if len(row) > 14 else "")) > 20 else (_s(row[14]) or None if len(row) > 14 else None),
                materiales=_s(row[25]) or None if len(row) > 25 else None,
                consideraciones=_s(row[26]) or None if len(row) > 26 else None,
                agua_bano=_s(row[27]) or None if len(row) > 27 else None,
                numero_camaras_vigilar=_int(row[28]) if len(row) > 28 else None,
                odc_path=_s(row[29]) or None if len(row) > 29 else None,
                desglose_path=_s(row[30]) or None if len(row) > 30 else None,
                contrato_path="; ".join(contrato_parts) or None,
                montos_a_cobrar=_s(row[32]) or None if len(row) > 32 else None,
                cotizacion_path=_s(row[15]) or None if len(row) > 15 else None,
                drive_folder_url=carpeta_url or None,
                drive_folder_id=carpeta_id or None,
                estado=estado,
                created_at=_date(row[1]) or datetime.now(),
            )
            db.add(ods)
            db.flush()

            # ── 3. Sub-tablas desde hoja Administración ───────────────────────
            if admin_row:
                ar = admin_row  # alias

                # Administración (cols 9-15)
                adm = AdministracionODT(
                    odt=codigo,
                    tecnico=_s(ar[28]) or None if len(ar) > 28 else None,
                    acompanante=_s(ar[48]) or None if len(ar) > 48 else None,
                    recepcion_info=bool(_date(ar[9])) if len(ar) > 9 else False,
                    fecha_recepcion_info=_date(ar[9]) if len(ar) > 9 else None,
                    registro_alpha3=bool(_date(ar[10])) if len(ar) > 10 else False,
                    fecha_registro_alpha3=_date(ar[10]) if len(ar) > 10 else None,
                    registro_intranet=bool(_date(ar[11])) if len(ar) > 11 else False,
                    fecha_registro_intranet=_date(ar[11]) if len(ar) > 11 else None,
                    envio_solicitud_instalacion=bool(_date(ar[12])) if len(ar) > 12 else False,
                    fecha_envio_solicitud_instalacion=_date(ar[12]) if len(ar) > 12 else None,
                    envio_datos_facturacion=bool(_date(ar[13])) if len(ar) > 13 else False,
                    fecha_envio_datos_facturacion=_date(ar[13]) if len(ar) > 13 else None,
                    envio_carta_bienvenida=bool(_date(ar[14])) if len(ar) > 14 else False,
                    fecha_envio_carta_bienvenida=_date(ar[14]) if len(ar) > 14 else None,
                    finalizado=bool(_date(ar[15])) if len(ar) > 15 else False,
                    fecha_cierre=_date(ar[15]) if len(ar) > 15 else None,
                )
                db.add(adm)

                # Finanzas (cols 16-21)
                fin = FinanzasODT(
                    odt=codigo,
                    fecha_inicio_servicio=_s(ar[41]) or None if len(ar) > 41 else None,
                    recepcion_datos_facturacion=bool(_date(ar[16])) if len(ar) > 16 else False,
                    fecha_recepcion_datos_facturacion=_date(ar[16]) if len(ar) > 16 else None,
                    creacion_clientes_piriod=bool(_date(ar[17])) if len(ar) > 17 else False,
                    fecha_creacion_clientes_piriod=_date(ar[17]) if len(ar) > 17 else None,
                    creacion_clientes_bd=bool(_date(ar[18])) if len(ar) > 18 else False,
                    fecha_creacion_clientes_bd=_date(ar[18]) if len(ar) > 18 else None,
                    facturacion_instalacion=bool(_date(ar[19])) if len(ar) > 19 else False,
                    fecha_facturacion_instalacion=_date(ar[19]) if len(ar) > 19 else None,
                    facturacion_servicio=bool(_date(ar[20])) if len(ar) > 20 else False,
                    fecha_facturacion_servicio=_date(ar[20]) if len(ar) > 20 else None,
                    finalizado=bool(_date(ar[21])) if len(ar) > 21 else False,
                    fecha_cierre=_date(ar[21]) if len(ar) > 21 else None,
                )
                db.add(fin)

                # Servicio Técnico (cols 22-30)
                st = ServicioTecnicoVentaODT(
                    odt=codigo,
                    recepcion_solicitud_instalacion=bool(_date(ar[22])) if len(ar) > 22 else False,
                    fecha_recepcion_solicitud_instalacion=_date(ar[22]) if len(ar) > 22 else None,
                    llamar_cliente=_s(ar[23]) or None if len(ar) > 23 else None,
                    solicitud_materiales=_s(ar[24]) or None if len(ar) > 24 else None,
                    fecha_inicio_instalacion=_s(ar[25]) or None if len(ar) > 25 else None,
                    fecha_fin_instalacion=_s(ar[26]) or None if len(ar) > 26 else None,
                    tecnico_a_cargo=_s(ar[28]) or None if len(ar) > 28 else None,
                    acompanante=_s(ar[48]) or None if len(ar) > 48 else None,
                    instalacion_finalizada=bool(_date(ar[29])) if len(ar) > 29 else False,
                    fecha_instalacion_finalizada=_date(ar[29]) if len(ar) > 29 else None,
                    finalizado=bool(_date(ar[30])) if len(ar) > 30 else False,
                    fecha_cierre=_date(ar[30]) if len(ar) > 30 else None,
                )
                db.add(st)

                # Soporte Técnico (cols 31-40)
                sop = SoporteTecnicoVentaODT(
                    odt=codigo,
                    configuracion_camaras=bool(_date(ar[31])) if len(ar) > 31 else False,
                    fecha_configuracion_camaras=_date(ar[31]) if len(ar) > 31 else None,
                    posicionamiento_imagen=bool(_date(ar[32])) if len(ar) > 32 else False,
                    fecha_posicionamiento_imagen=_date(ar[32]) if len(ar) > 32 else None,
                    enlace_servidor=bool(_date(ar[33])) if len(ar) > 33 else False,
                    fecha_enlace_servidor=_date(ar[33]) if len(ar) > 33 else None,
                    configuracion_ivs=bool(_date(ar[34])) if len(ar) > 34 else False,
                    fecha_configuracion_ivs=_date(ar[34]) if len(ar) > 34 else None,
                    plan_grabacion=bool(_date(ar[35])) if len(ar) > 35 else False,
                    fecha_plan_grabacion=_date(ar[35]) if len(ar) > 35 else None,
                    requiere_puesto_nuevo=_s(ar[36]) or None if len(ar) > 36 else None,
                    numero_central_asignado=_s(ar[37]) or None if len(ar) > 37 else None,
                    configuracion_cliente=bool(_date(ar[38])) if len(ar) > 38 else False,
                    fecha_configuracion_cliente=_date(ar[38]) if len(ar) > 38 else None,
                    vb_final_servicio=bool(_date(ar[39])) if len(ar) > 39 else False,
                    fecha_vb_final_servicio=_date(ar[39]) if len(ar) > 39 else None,
                    terminado=bool(_date(ar[40])) if len(ar) > 40 else False,
                    fecha_terminado=_date(ar[40]) if len(ar) > 40 else None,
                )
                db.add(sop)

                # Operaciones (cols 41-46)
                ops = OperacionesVentaODT(
                    odt=codigo,
                    fecha_inicio_servicio=_s(ar[41]) or None if len(ar) > 41 else None,
                    fecha_coordinacion=bool(_date(ar[42])) if len(ar) > 42 else False,
                    ts_fecha_coordinacion=_date(ar[42]) if len(ar) > 42 else None,
                    reunion_coordinacion=bool(_date(ar[43])) if len(ar) > 43 else False,
                    ts_reunion_coordinacion=_date(ar[43]) if len(ar) > 43 else None,
                    coord_apertura_puesto=bool(_date(ar[44])) if len(ar) > 44 else False,
                    ts_coord_apertura_puesto=_date(ar[44]) if len(ar) > 44 else None,
                    coord_equipo=bool(_date(ar[45])) if len(ar) > 45 else False,
                    ts_coord_equipo=_date(ar[45]) if len(ar) > 45 else None,
                    terminado=bool(_date(ar[46])) if len(ar) > 46 else False,
                    ts_terminado=_date(ar[46]) if len(ar) > 46 else None,
                )
                db.add(ops)

            db.commit()
            insertados += 1
            print(f"  ✅ {codigo} — {razon[:50]}")

        except Exception as exc:
            db.rollback()
            errores += 1
            print(f"  ❌ Fila {i} ({codigo}): {exc}")

print(f"\n{'─'*50}")
print(f"Insertados: {insertados}  |  Omitidos (ya existían): {skipped}  |  Errores: {errores}")
print("Listo.\n")
