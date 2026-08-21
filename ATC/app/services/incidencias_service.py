from __future__ import annotations

import csv
import bcrypt
import json
import logging
from decimal import Decimal
import mimetypes
import os
import re
import secrets
import smtplib
import threading
import urllib.error
import urllib.request
import uuid
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from email.message import EmailMessage
from html import escape as html_escape
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, urlencode, urlsplit
from zoneinfo import ZoneInfo
import xml.etree.ElementTree as ET

from passlib.context import CryptContext
from sqlalchemy import case, func, or_, select, text
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy.orm import Session

from ATC.app.data.materiales_catalogo import MATERIALES_CATALOGO_SEED
from ATC.app.core.incidencias_config import settings
from ATC.app.core.incidencias_db import SessionLocal, build_engine
from ATC.app.core.image_optimizer import optimize_image_bytes
from ATC.app.core.session_policy import expiracion_sesion
from ATC.app.services.incidencias_drive_report_service import (
    DriveReportError,
    download_support_drive_file_bytes,
    list_support_images_for_odt,
    resolve_odt_cierre_folder,
    retry_odt_cierre_drive_upload,
    upload_cierre_apertura_image_to_drive,
    upload_odt_cierre_images_to_drive,
    upload_odt_cierre_pdf_to_drive,
    upload_odt_cierre_to_drive,
    upload_support_images_for_odt,
)
from ATC.app.services.rendiciones_drive_service import (
    upload_rendicion_boleta_to_drive,
    upload_rendicion_informe_to_drive,
)
from ATC.app.models.prevencion import EstatusDocumentacionTecnico
from ATC.app.models.incidencias import (
    AdministracionODT,
    CierreAperturaImagen,
    ClienteBBDD,
    IncidenciaImagenTabla,
    MantencionImagenSucursal,
    LoginSession,
    ProtocoloInforme,
    PruebaSonido,
    Registro,
    RegistroCorreoCliente,
    Rendicion,
    RendicionPago,
    RendicionViaticoCap,
    SucursalBBDD,
    SucursalCamaraMonitoreo,
    SucursalContactoEmergencia,
    SucursalPersonaAutorizada,
    ServicioTecnicoVentaODT,
    SoporteTecnicoVentaODT,
    User,
    VentaODS,
    VentaODSArchivo,
)
from ATC.app.schemas.incidencias import (
    ContactoDestinoRequest,
    EnviarInformacionContactoRequest,
    FormularioRegistro,
    IncidenciaNueva,
    RendicionRequest,
)


PWD_CONTEXT = CryptContext(schemes=["bcrypt"], deprecated="auto")
_ATC_ROOT = Path(__file__).resolve().parents[2]
_INCIDENCIAS_APP_DIR = Path(__file__).resolve().parents[1]
_UPLOADS_ROOT = _ATC_ROOT / "uploads"
NON_ASSIGNABLE_TECHNICIAN_NAMES = {
    "fernando andres lubiano moraga",
    "fernando lubiano",
    "gianpiero lubiano",
    "gianpiero lubiano forno",
}
NON_SELECTABLE_SUPERVISOR_NAME_PREFIXES = tuple(sorted(NON_ASSIGNABLE_TECHNICIAN_NAMES))


def _url_to_path(url: str) -> Path:
    """Converts a server-relative URL (/uploads/... or /static/...) to an absolute Path."""
    stripped = url.lstrip("/")
    if stripped.startswith("uploads/"):
        return _ATC_ROOT / stripped
    return _INCIDENCIAS_APP_DIR / stripped


_PENDING_DRIVE_ROOT = _UPLOADS_ROOT / "_pending_drive"


def _guardar_en_cuarentena_drive(
    flujo: str,
    identificador: str,
    filename: str,
    content: bytes,
    retry_meta: dict[str, Any],
) -> str:
    """Fallback cuando falla una subida a Drive: guarda el archivo en una
    carpeta de cuarentena acotada (no la carpeta normal de uploads) y deja un
    sidecar .meta.json con los datos que necesita el job de reintento
    (automation_loop) para reintentar la subida y actualizar la fila
    correspondiente. Devuelve la URL local temporal a guardar en la BBDD."""
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "_", identificador or "sin_id")
    carpeta = _PENDING_DRIVE_ROOT / flujo / safe_id
    carpeta.mkdir(parents=True, exist_ok=True)
    destino = carpeta / filename
    destino.write_bytes(content)
    meta_path = destino.with_suffix(destino.suffix + ".meta.json")
    meta_path.write_text(json.dumps({"flujo": flujo, **retry_meta}), encoding="utf-8")
    rel = destino.relative_to(_UPLOADS_ROOT)
    LOGGER.warning(
        "Cuarentena Drive: flujo=%s id=%s archivo=%s (subida a Drive fallo, reintentara automation_loop)",
        flujo, identificador, filename,
    )
    return f"/uploads/{rel.as_posix()}"
IDENTITY_SEED_FILE = _ATC_ROOT / "app" / "data" / "users_areas_seed.csv"
AREA_DESTINOS: dict[str, str] = {
    "auto": "",
    "loginunico": "",
    "unificado": "",
    "panelselectorauto": "",
    "panelselectorsoporte": "soporte",
    "panel_selector_soporte": "soporte",
    "panelselector": "incidencias",
    "panel_selector": "incidencias",
    "panelselectorcoordinacion": "coordinacion",
    "panel_selector_coordinacion": "coordinacion",
    "cierreaperturaclientes": "incidencias",
    "controlprotocolos": "protocolos",
    "tablaprotocolos": "protocolos",
    "envioprotocolossemanales": "protocolos",
    "coordinacion": "coordinacion",
    "incidencias": "incidencias",
    "panelselectorservicio": "servicio_tecnico",
    "panel_selector_servicio": "servicio_tecnico",
    "serviciotecnico": "servicio_tecnico",
    "stventas": "servicio_tecnico",
    "tabla": "servicio_tecnico",
    "resumenequipostecnicos": "resumen_equipos_tecnicos",
    "resumenequipostécnicos": "resumen_equipos_tecnicos",
    "panelselectorventa": "venta",
    "panel_selector_venta": "venta",
    "registrocliente": "venta",
    "tablacliente": "venta",
    "panelselectorfinanzas": "finanzas",
    "panel_selector_finanzas": "finanzas",
    "tablafinanzas": "finanzas",
    "panelselectoradministracion": "administracion",
    "panel_selector_administracion": "administracion",
    "tablaadministracion": "administracion",
    "panelselectoroperaciones": "operaciones",
    "panel_selector_operaciones": "operaciones",
    "tablaoperaciones": "operaciones",
    "panelselectorsupervisores": "supervisores",
    "panel_selector_supervisores": "supervisores",
    "materiales": "materiales",
    "pendientes": "tecnicos",
    "tecnicos": "tecnicos",
}
AREA_PANEL_DESTINOS: dict[str, str] = {
    "soporte": "panelSelectorSoporte",
    "materiales": "materiales",
    "servicio_tecnico": "panelSelectorServicio",
    "resumen_equipos_tecnicos": "resumenEquiposTecnicos",
    "tecnicos": "tecnicos",
    "incidencias": "panelSelector",
    "coordinacion": "panelSelectorCoordinacion",
    "protocolos": "panelSelectorCoordinacion",
    "venta": "panelSelectorVenta",
    "finanzas": "panelSelectorFinanzas",
    "administracion": "panelSelectorAdministracion",
    "operaciones": "panelSelectorOperaciones",
    "guardia": "panelSelectorGuardia",
    "supervisores": "panelSelectorSupervisores",
    "rrhh": "panelSelectorRRHH",
    "prevencion": "panelSelectorPrevencion",
    "bitacora": "bitacora",
}
AREA_INFO: dict[str, tuple[str, str]] = {
    "soporte": ("Soporte", "Soporte"),
    "materiales": ("Materiales", "Materiales"),
    "servicio_tecnico": ("Servicio Tecnico", "Servicio Tecnico"),
    "resumen_equipos_tecnicos": ("Resumen Equipos Técnicos", "Resumen Equipos Técnicos"),
    "tecnicos": ("Tecnicos", "Tecnicos"),
    "incidencias": ("Incidencias", "Operador"),
    "coordinacion": ("Coordinacion", "Operador"),
    "protocolos": ("Control de Protocolos", "Operador"),
    "venta": ("Venta", "Comercial"),
    "finanzas": ("Finanzas", "Finanzas"),
    "administracion": ("Administracion", "Administracion"),
    "operaciones": ("Operaciones", "Operaciones"),
    "guardia": ("Guardia", "Guardia"),
    "supervisores": ("Supervisores", "Supervisores"),
    "rrhh": ("RRHH", "RRHH"),
    "prevencion": ("Prevención", "Prevención"),
    "bitacora": ("Bitácora", "Bitacora"),
}
DEPARTMENT_AREAS: dict[str, list[str]] = {
    "soporte": ["soporte"],
    "materiales": ["materiales"],
    "servicio tecnico": ["servicio_tecnico"],
    "servicio técnico": ["servicio_tecnico"],
    "resumen equipos tecnicos": ["resumen_equipos_tecnicos"],
    "resumen equipos técnicos": ["resumen_equipos_tecnicos"],
    "panel operadores": ["resumen_equipos_tecnicos"],
    "tecnicos": ["tecnicos"],
    "técnicos": ["tecnicos"],
    "tecnico externo": ["tecnicos"],
    "técnico externo": ["tecnicos"],
    "operador": ["incidencias"],
    "coordinacion": ["coordinacion"],
    "coordinación": ["coordinacion"],
    "comercial": ["venta"],
    "finanzas": ["finanzas"],
    "administracion": ["administracion"],
    "administración": ["administracion"],
    "operaciones": ["operaciones"],
    "guardia": ["guardia"],
    "supervisores": ["supervisores"],
    "rrhh": ["rrhh"],
    "prevencion": ["prevencion"],
    "prevención": ["prevencion"],
    "bitacora": ["bitacora"],
    "televigilante": ["bitacora"],
}
ADMIN_SELECTOR_AREA_CODES = [
    "soporte",
    "materiales",
    "servicio_tecnico",
    "resumen_equipos_tecnicos",
    "tecnicos",
    "incidencias",
    "coordinacion",
    "venta",
    "finanzas",
    "administracion",
    "operaciones",
    "guardia",
    "supervisores",
    "rrhh",
    "prevencion",
    "bitacora",
]


@dataclass(frozen=True)
class AreaInfo:
    code: str
    name: str
    department: str
_GEOCODE_CACHE: dict[str, tuple[str, str]] = {}

_GOOGLE_GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"


def _limpiar_nombre_region_google(nombre: str) -> str:
    n = (nombre or "").strip()
    for prefijo in ("Región del ", "Región de ", "Región "):
        if n.startswith(prefijo):
            return n[len(prefijo):].strip()
    return n


def _parsear_componentes_google(address_components: list[dict]) -> tuple[str, str]:
    region = ""
    comuna = ""
    locality = ""
    for comp in address_components or []:
        types = comp.get("types", [])
        if "administrative_area_level_1" in types:
            region = _limpiar_nombre_region_google(comp.get("long_name", ""))
        elif "administrative_area_level_3" in types:
            comuna = comp.get("long_name", "")
        elif "locality" in types and not locality:
            locality = comp.get("long_name", "")
    if not comuna:
        comuna = locality
    return region, comuna


def _llamar_google_geocoding(params: dict) -> dict:
    if not settings.google_maps_api_key:
        raise ValueError("GOOGLE_MAPS_API_KEY no está configurada.")
    query = {**params, "key": settings.google_maps_api_key, "language": "es", "region": "cl"}
    url = f"{_GOOGLE_GEOCODE_URL}?{urlencode(query)}"
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode("utf-8"))


def geocodificar_region_comuna_google(*, direccion: str = "", lat: str = "", lng: str = "") -> tuple[str, str, str]:
    """Devuelve (region, comuna, error) usando la API de Geocoding de Google.
    Prioriza coordenadas si se entregan; si no, usa la direccion."""
    try:
        if lat and lng:
            data = _llamar_google_geocoding({"latlng": f"{lat},{lng}"})
        elif direccion:
            data = _llamar_google_geocoding({"address": direccion, "components": "country:CL"})
        else:
            return "", "", "sin_direccion_ni_coordenadas"
    except Exception as exc:
        return "", "", f"error_llamada: {exc}"

    if data.get("status") != "OK" or not data.get("results"):
        return "", "", f"geocode_status={data.get('status')}"

    region, comuna = _parsear_componentes_google(data["results"][0].get("address_components", []))
    if not region and not comuna:
        return "", "", "sin_componentes_admin"
    return region, comuna, ""


def _normalizar_comparable_geo(valor: str) -> str:
    n = unicodedata.normalize("NFD", (valor or "").lower())
    return "".join(c for c in n if unicodedata.category(c) != "Mn").strip()


def validar_region_comuna_google(region: str, comuna: str) -> tuple[bool, str, str]:
    """Verifica que `comuna` efectivamente pertenezca a `region` segun Google
    Maps (independiente de coordenadas guardadas, que pueden estar mal
    cargadas). Devuelve (valido, region_sugerida_por_google, mensaje_error)."""
    region_txt = (region or "").strip()
    comuna_txt = (comuna or "").strip()
    if not comuna_txt:
        return False, "", "Debes indicar una comuna."
    if not region_txt:
        return False, "", "Debes indicar una región."

    region_real, _comuna_real, error = geocodificar_region_comuna_google(direccion=f"{comuna_txt}, Chile")
    if error:
        return False, "", f"No se pudo verificar '{comuna_txt}' en Google Maps ({error})."
    if not region_real:
        return False, "", f"Google Maps no devolvió una región para '{comuna_txt}'."

    if _normalizar_comparable_geo(region_real) != _normalizar_comparable_geo(region_txt):
        return False, region_real, (
            f"La comuna '{comuna_txt}' no pertenece a la región '{region_txt}' — "
            f"según Google Maps pertenece a la región '{region_real}'."
        )
    return True, region_real, ""


_COORD_FALLBACK_CL: list[tuple[tuple[str, ...], tuple[str, str]]] = [
    (("valparaiso", "valparaiso"), ("-33.0472", "-71.6127")),
    (("vina del mar", "vinaa del mar", "vina"), ("-33.0245", "-71.5518")),
    (("quilpue",), ("-33.0475", "-71.4425")),
    (("san bernardo",), ("-33.5922", "-70.6996")),
    (("san miguel",), ("-33.4979", "-70.6510")),
    (("maipu",), ("-33.5108", "-70.7653")),
    (("region metropolitana", "santiago"), ("-33.4489", "-70.6693")),
]
MANTENCIONES_PROGRAMADAS_QUILPUE: dict[int, list[str]] = {
    1: [
        "Imq Consistorial Nuevo",
        "Imq Juzgado",
        "Imq Teatro Municipal - Dirección Cultura",
        "IMQ Centro Cultural",
        "Imq Carozzi 2 Dideco/secpla/obras",
        "Imq Derecho - Carozzi 3",
        "Imq Oficina Niñez",
    ],
    2: [
        "Imq Estadio V. Olimpica",
        "Imq Piscina V. Olimpica",
        "Imq Pisc. Bto. Sur",
        "Imq Unco",
        "Imq Oficina persona mayores",
    ],
    3: [
        "Imq Vep",
        "Imq Operaciones",
        "Imq Pisc. Bto. Norte",
        "Imq Tránsito y Transporte público - deleg el Belloto",
        "IMQ Feria",
        "Imq Centro Prácticas",
    ],
    4: [
        "IMQ Desarrollo Vecinal",
        "Imq Desarrollo economico - ex biblioteca",
        "Imq Deportes- Gimnasio Municipal",
        "IMQ Zoologico",
        "Imq CIAM",
    ],
}
MANTENCIONES_TRIMESTRALES_QUINTERO: list[str] = [
    "MQUIN Terminal de Buses",
    "MQUIN Dirección de Seguridad Pública",
    "MQUIN Parque Municipal",
    "MQUIN Seguridad Pública Loncura",
    "MQUIN Medio Ambiente",
    "MQUIN Oficina Aseo y Ornato Loncura",
    "MQUIN Estadio Municipal y Cancha aledañas",
    "MQUIN Cementerio Municipal",
    "MQUIN EDIFICIO DAEM",
    "MQUIN Escombrera Municipal",
    "MQUIN DIDECO",
    "MQUIN Aparcadero Municipal",
    "MQUIN Juzgado de Policía Local",
    "MQUIN Edificio Consistorial Base",
    "MQUIN Edificio de Administración DESAM",
    "MQUIN Farmacia Municipal",
    "MQUIN Posta de Salud Loncura",
    "MQUIN Cesfam Quintero",
]
MANTENCIONES_TRIMESTRALES_CONCON: list[str] = [
    "MC Secplac",
    "MC Playa Amarilla",
    "MC OPD",
    "MC Museo",
    "MC DIDECO",
    "MC Turismo y Omil",
    "MC Edificio Municipal",
    "MC Carpa (Avanzada Cultural)",
    "MC Cesfam",
    "MC Biblioteca Municipal",
    "MC Juzgado Policía Local",
    "MC Tránsito",
    "MC CJAM - Centro de Juventud, Adulto Mayor y Discapacidad",
    "MC Deporte, Estadio Atlético",
    "MC Parque La Isla",
    "MC Jardín Infantil Conconcito",
    "MC Centro Cultural",
    "MC DOM",
    "MC Corrales Municipales",
]
MANTENCIONES_MENSUALES_LLAY_LLAY: list[str] = [
    "Cesfam Llay Llay",
]
MESES_MANTENCION_TRIMESTRAL = {3, 6, 9, 12}
MANTENCIONES_TRIMESTRALES_POR_COMUNA: dict[str, list[str]] = {
    "quintero": MANTENCIONES_TRIMESTRALES_QUINTERO,
    "concon": MANTENCIONES_TRIMESTRALES_CONCON,
}
MANTENCIONES_IMAGENES_POR_SUCURSAL: dict[str, list[str]] = {
    "imq consistorial nuevo": [
        "app/static/mantenciones/quilpue/imq_consistorial_nuevo_1.jpg",
        "app/static/mantenciones/quilpue/imq_consistorial_nuevo_2.jpg",
    ],
}
SUCURSALES_EXTRA_MANTENCION: list[str] = [
    *MANTENCIONES_TRIMESTRALES_QUINTERO,
    *MANTENCIONES_TRIMESTRALES_CONCON,
    "Quintero",
    "Concon",
]
MATERIALES_EXCEL_CACHE_LOCK = threading.Lock()
MATERIALES_EXCEL_CACHE: dict[str, Any] = {
    "path": None,
    "mtime": None,
    "items": None,
}


def _cell_value_from_xlsx(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> str:
    cell_type = cell.attrib.get("t", "")
    value = cell.findtext("a:v", default="", namespaces=ns) or ""
    if cell_type == "s" and value:
        try:
            return shared_strings[int(value)]
        except (ValueError, IndexError):
            return ""
    if cell_type == "inlineStr":
        return "".join(part.text or "" for part in cell.iterfind(".//a:t", ns))
    if value:
        return value
    is_el = cell.find("a:is", ns)
    if is_el is not None:
        return "".join(part.text or "" for part in is_el.iterfind(".//a:t", ns))
    return ""


def _read_xlsx_shared_strings(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    shared_strings: list[str] = []
    for si in root.findall("a:si", ns):
        shared_strings.append("".join(part.text or "" for part in si.iterfind(".//a:t", ns)))
    return shared_strings


def _resolve_xlsx_sheet_targets(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels: dict[str, str] = {}
    rels_path = "xl/_rels/workbook.xml.rels"
    if rels_path in zf.namelist():
        rels_root = ET.fromstring(zf.read(rels_path))
        rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        for rel in rels_root.findall("r:Relationship", rel_ns):
            rels[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")

    targets: list[str] = []
    sheets = workbook.find("a:sheets", ns)
    if sheets is None:
        return targets
    for sheet in sheets.findall("a:sheet", ns):
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        target = rels.get(rel_id, "")
        if not target:
            continue
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = f"xl/{target}"
        targets.append(target)
    return targets


def _inferir_unidad_material(nombre: str) -> str:
    texto = unicodedata.normalize("NFD", str(nombre or "").strip().lower())
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    if any(
        clave in texto
        for clave in (
            "cable",
            "corrugado",
            "tuberia",
            "piola",
            "cinta",
            "canaleta",
            "manguera",
            "metro",
        )
    ):
        return "metro"
    return "unidad"

KNOWN_REGISTRO_BLOCKING_QUERY = "ALTER TABLE registro DROP COLUMN IF EXISTS foto_2"
LOGGER = logging.getLogger(__name__)
TICKET_STATUS_ABIERTO = "open"
TICKET_STATUS_PENDIENTE = "pending"
TICKET_STATUS_PENDIENTE_SERVICIO = "pending_service"
TICKET_STATUS_PENDIENTE_CLIENTE = "pending_client"
TICKET_STATUS_RESUELTO = "resolved"
TICKET_STATUS_RESUELTO_SERVICIO = "resolved_service"
TICKET_STATUS_RESUELTO_CLIENTE = "resolved_client"
TICKET_STATUS_CERRADO = "closed"
TICKET_STATUS_ALIASES = {
    "open": TICKET_STATUS_ABIERTO,
    "abierto": TICKET_STATUS_ABIERTO,
    "pendiente": TICKET_STATUS_PENDIENTE,
    "pending": TICKET_STATUS_PENDIENTE,
    "pendiente_servicio": TICKET_STATUS_PENDIENTE_SERVICIO,
    "pending_service": TICKET_STATUS_PENDIENTE_SERVICIO,
    "pendiente_cliente": TICKET_STATUS_PENDIENTE_CLIENTE,
    "pending_client": TICKET_STATUS_PENDIENTE_CLIENTE,
    "resuelto": TICKET_STATUS_RESUELTO,
    "resolved": TICKET_STATUS_RESUELTO,
    "resuelto_servicio": TICKET_STATUS_RESUELTO_SERVICIO,
    "resuleto_servicio": TICKET_STATUS_RESUELTO_SERVICIO,
    "resolved_service": TICKET_STATUS_RESUELTO_SERVICIO,
    "resuelto_cliente": TICKET_STATUS_RESUELTO_CLIENTE,
    "resolved_client": TICKET_STATUS_RESUELTO_CLIENTE,
    "cerrado": TICKET_STATUS_CERRADO,
    "closed": TICKET_STATUS_CERRADO,
}
TICKET_STATUSES_PERMITIDOS = {
    *TICKET_STATUS_ALIASES.keys(),
    *TICKET_STATUS_ALIASES.values(),
}


def _normalizar_estado_ticket_soporte(estado: str | None) -> str:
    raw = " ".join(str(estado or "").strip().split())
    key = unicodedata.normalize("NFD", raw.casefold())
    key = "".join(ch for ch in key if unicodedata.category(ch) != "Mn")
    key = key.replace(" ", "_").replace("-", "_")
    return TICKET_STATUS_ALIASES.get(key, key)


def _to_ddmmyyyy(valor: datetime | None) -> str:
    if not valor:
        return ""
    return valor.strftime("%d/%m/%Y")


def _to_ddmmyyyy_hhmm(valor: datetime | None) -> str:
    if not valor:
        return ""
    return valor.strftime("%d/%m/%Y %H:%M")


def _parse_prefijo_numero(odt: str | None) -> int | None:
    if not odt:
        return None
    match = re.search(r"(\d+)$", odt.strip())
    if not match:
        return None
    return int(match.group(1))


def _build_db_write_error(exc: Exception, tabla: str = "Registro") -> ValueError:
    message = str(exc or "").lower()
    if "lock timeout" in message or "canceling statement due to lock timeout" in message:
        return ValueError(
            f"La tabla {tabla} esta bloqueada en PostgreSQL por otra sesion. "
            "Cierra la transaccion abierta y vuelve a intentar."
        )
    if "deadlock detected" in message:
        return ValueError(
            f"PostgreSQL detecto un deadlock al escribir en {tabla}. "
            "Vuelve a intentar en unos segundos."
        )
    return ValueError(f"No se pudo guardar el registro en {tabla}: {exc}")


def _is_lock_timeout_error(exc: Exception) -> bool:
    message = str(exc or "").lower()
    has_lock_hint = "lock timeout" in message or "locknotavailable" in message or "tiempo de espera" in message
    return has_lock_hint and ("lock" in message or "locks" in message)


def _normalizar_identidad(valor: Any) -> str:
    txt = str(valor or "").strip().lower()
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _username_desde_nombre(nombre: str) -> str:
    base = _normalizar_identidad(nombre)
    base = re.sub(r"[^a-z0-9]+", ".", base).strip(".")
    return base or "usuario"


def _load_identity_seed() -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    if not IDENTITY_SEED_FILE.exists():
        return [], []
    areas: list[dict[str, str]] = []
    users: list[dict[str, Any]] = []
    with IDENTITY_SEED_FILE.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            record_type = str(row.get("record_type") or "").strip().lower()
            if record_type == "area":
                areas.append(
                    {
                        "code": str(row.get("code") or "").strip(),
                        "name": str(row.get("name") or "").strip(),
                        "department": str(row.get("department") or "").strip(),
                    }
                )
            elif record_type == "user":
                users.append(
                    {
                        "name": str(row.get("name") or "").strip(),
                        "password": str(row.get("password") or "").strip(),
                        "role": str(row.get("role") or "agent").strip() or "agent",
                        "areas": [
                            item.strip()
                            for item in str(row.get("areas") or "").split(";")
                            if item.strip()
                        ],
                    }
                )
    return areas, users


def seed_default_identity_data(db: Session) -> None:
    """Carga inicial idempotente: la autorizacion operativa se lee desde BBDD."""
    default_areas, default_users = _load_identity_seed()
    department_by_area = {code: info[1] for code, info in AREA_INFO.items()}
    for area_data in default_areas:
        code = str(area_data.get("code") or "").strip()
        department = str(area_data.get("department") or "").strip()
        if code and department:
            department_by_area[code] = department

    usuarios_existentes = list(db.scalars(select(User)).all())
    usuarios_por_nombre = {_normalizar_identidad(u.name): u for u in usuarios_existentes}
    usuarios_por_username = {str(u.username or "").strip().lower(): u for u in usuarios_existentes}

    for user_data in default_users:
        nombre = str(user_data["name"]).strip()
        username = _username_desde_nombre(nombre)
        user = usuarios_por_nombre.get(_normalizar_identidad(nombre)) or usuarios_por_username.get(username)
        departments = []
        for code in user_data["areas"]:
            department = department_by_area.get(code)
            if department and department not in departments:
                departments.append(department)
        department_value = ";".join(departments) if departments else None
        if not user:
            user = User(
                name=nombre,
                username=username,
                hashed_password=PWD_CONTEXT.hash(str(user_data["password"])),
                role=str(user_data.get("role") or "agent"),
                department=department_value,
                is_active=True,
            )
            db.add(user)
            db.flush()
            usuarios_por_nombre[_normalizar_identidad(nombre)] = user
            usuarios_por_username[username] = user
        else:
            user.name = user.name or nombre
            user.role = str(user_data.get("role") or user.role or "agent")
            user.department = user.department or department_value
            user.is_active = True
            if not user.hashed_password:
                user.hashed_password = PWD_CONTEXT.hash(str(user_data["password"]))

    db.commit()


# Edicion de la ultima nota de observacion_servicio, con ventana de tiempo
# tipo "editar mensaje de WhatsApp" (pedido explicito, jul 2026): solo el
# mismo usuario que escribio la ultima linea puede modificarla, y solo
# dentro de OBSERVACION_EDIT_WINDOW_MINUTES desde que la escribio.
OBSERVACION_EDIT_WINDOW_MINUTES = 15
_OBS_ENTRY_RE = re.compile(
    r"^\[(?P<user>.+) - (?P<fecha>\d{2}/\d{2}/\d{4} \d{2}:\d{2})\]\s*(?:\(editado\)\s*)?(?P<texto>.*)$"
)


def _obs_last_entry(text: str) -> dict | None:
    lines = (text or "").splitlines()
    start_idx = None
    match = None
    for idx in range(len(lines) - 1, -1, -1):
        m = _OBS_ENTRY_RE.match(lines[idx].strip())
        if m:
            start_idx = idx
            match = m
            break
    if start_idx is None or match is None:
        return None
    try:
        fecha_dt = datetime.strptime(match.group("fecha"), "%d/%m/%Y %H:%M")
    except ValueError:
        return None
    return {
        "start_idx": start_idx,
        "user": match.group("user").strip(),
        "fecha_str": match.group("fecha"),
        "fecha_dt": fecha_dt,
    }


def _obs_can_edit_entry(entry: dict | None, usuario: str) -> bool:
    if not entry:
        return False
    if entry["user"].strip().casefold() != (usuario or "").strip().casefold():
        return False
    tz_name = (settings.timezone or "America/Santiago").strip() or "America/Santiago"
    ahora = datetime.now(ZoneInfo(tz_name)).replace(tzinfo=None)
    elapsed = ahora - entry["fecha_dt"]
    return timedelta(0) <= elapsed <= timedelta(minutes=OBSERVACION_EDIT_WINDOW_MINUTES)


def _obs_edit_last_line(current_text: str, entry: dict, usuario: str, nuevo_texto: str) -> str:
    lines = (current_text or "").splitlines()
    nueva_linea = f"[{usuario} - {entry['fecha_str']}] (editado) {nuevo_texto}"
    nuevas_lineas = lines[: entry["start_idx"]] + [nueva_linea]
    return "\n".join(nuevas_lineas).strip()


def _obs_delete_last_line(current_text: str, entry: dict) -> str:
    lines = (current_text or "").splitlines()
    return "\n".join(lines[: entry["start_idx"]]).strip()


# Cache de introspección de esquema (information_schema) a nivel de proceso.
# _schemas_con_tabla/_columnas_tabla se llaman ~20 veces por archivo en el
# camino de request más usado del sitio, y una instancia nueva de
# IncidenciasService se crea por request (Depends), así que un cache de
# instancia no serviría de nada — el esquema no cambia sin un reinicio del
# server (todo cambio de código Python ya requiere reinicio), por eso el
# cache es seguro a nivel de módulo y elimina round-trips a SQL Server
# repetidos en cada carga de página.
_SCHEMAS_CON_TABLA_CACHE: dict[str, list[str]] = {}
_COLUMNAS_TABLA_CACHE: dict[tuple[str, str], set[str]] = {}


class IncidenciasService:
    MANTENCION_CIERRE_MAX_IMAGENES = 80
    MANTENCION_CIERRE_MAX_BYTES = 10 * 1024 * 1024
    CIERRE_ODT_MAX_BYTES = 10 * 1024 * 1024

    CAUSAS_CIERRE: dict[str, set[str]] = {
        "ATC": {
            "instalacion_deficiente",
            "configuracion_incorrecta",
            "material_instalado_defectuoso",
            "mantenimiento_insuficiente",
            "diagnostico_previo_incompleto",
            "otro",
        },
        "Cliente": {
            "manipulacion_cliente",
            "problema_electrico_cliente",
            "red_internet_cliente",
            "infraestructura_cliente",
            "dano_terceros",
            "otro",
        },
        "Proveedor Externo": {
            "falla_proveedor_servicio",
            "corte_programado_proveedor",
            "equipo_proveedor_defectuoso",
            "otro",
        },
        "Internet": {
            "corte_internet_zona",
            "intermitencia_enlace",
            "falla_router_modem",
            "otro",
        },
        "Otro": {
            "fuerza_mayor",
            "vandalismo",
            "causa_no_determinada",
            "otro",
        },
    }
    ACCIONES_CIERRE = {
        "reconexion",
        "reconfiguracion",
        "reemplazo_material",
        "ajuste_fisico",
        "limpieza",
        "cambio_cableado",
        "cambio_fuente",
        "validacion_sin_intervencion",
        "otro",
    }
    RESULTADOS_CIERRE = {
        "operativo",
        "operativo_con_observacion",
        "requiere_seguimiento",
        "requiere_cotizacion_visita_adicional",
    }
    PRUEBAS_CIERRE = {"camaras_ok", "grabacion_ok", "audio_ok", "red_ok", "energia_ok"}

    # Equipos fijos (camioneta -> tecnicos asignados) que arma el tablero de
    # obtener_resumen_equipos_tecnicos_hoy(). Unica fuente de verdad: tambien
    # la usa la sugerencia de acompanante al derivar una ODT (ver
    # _sugerir_acompanante_para_tecnico), para que ambas coincidan siempre —
    # un tecnico que aparece solo en un equipo (lista de 1) nunca se sugiere
    # con acompanante.
    EQUIPOS_TECNICOS_POR_PATENTE: dict[str, list[str]] = {
        "RTXG 52": ["Cristopher Enrique Soto Diaz", "Dwait German Aros Contreras"],
        "RHPV 38": ["Emmanuel Issak Correa Ubilla", "Haxel Samir Del Carmen Saavedra Villanueva"],
        "KVTG 28": ["Omar Alejandro Triviño Silva"],
        "SRVP 17": ["Diego Antonio Moncada Sepulveda", "Ricardo Andres Vergara Guerra"],
        "SSZW 51": ["Bryan Benjamin Ibaceta Fabrega", "Rodrigo Octavio Carmona Agurto"],
        "SSZS 24": ["Michael Alejandro Herrera Navia", "Hans Reinhold Schemmel Rodriguez"],
        "RJXX 46": ["Marco Antonio Lopez Aguirre", "Bryan Alexander Rebolledo Hidalgo"],
        "VXLG 86": ["Luis Alberto Bustamante Aguilera", "Enrique Alejandro Sandoval Nunez"],
        "VXLG 93": ["Mauro Estefano Reyes Villegas"],
    }
    PATENTES_TECNICOS_FIJAS: list[str] = [
        "RTXG 52",
        "RHPV 38",
        "KVTG 28",
        "SRVP 17",
        "SSZW 51",
        "SSZS 24",
        "RJXX 46",
        "VXLG 86",
        "VXLG 93",
    ]
    ALIASES_PATENTE_TECNICO: dict[str, str] = {
        "christopher enrique soto diaz": "RTXG 52",
        "dwait german aros contreras": "RTXG 52",
        "omar alejandro trivino silva": "KVTG 28",
        "emmanuel issak correa ubilla": "RHPV 38",
        "haxel samir del carmen saavedra villanueva": "RHPV 38",
        "diego antonio moncada sepulveda": "SRVP 17",
        "diego moncada sepulveda": "SRVP 17",
        "ricardo vergara": "SRVP 17",
        "michael alejandro herrera navia": "SSZS 24",
        "hans reinhold schemmel rodriguez": "SSZS 24",
        "marco antonio lopez aguirre": "RJXX 46",
        "bryan alexander rebolledo hidalgo": "RJXX 46",
        "bryan rebolledo hidalgo": "RJXX 46",
        "bryan benjamin ibaceta fabrega": "SSZW 51",
        "rodrigo octavio carmona agurto": "SSZW 51",
        "luis alberto bustamante aguilera": "VXLG 86",
        "enrique alejandro sandoval nunez": "VXLG 86",
        "enrique alejandro sandoval": "VXLG 86",
        "mauro estefano reyes villegas": "VXLG 93",
        "javier ignacio salgado brito": "SRVP 17",
        "ricardo andres vergara guerra": "SRVP 17",
        "*por confirmar*": "SSZS 24",
    }
    MAX_FOTOS_CIERRE_ODS = 20
    MAX_FOTOS_INFORME_ODS = 2
    MATERIALES_CIERRE = {
        "cable_utp",
        "conector_rj45",
        "balun",
        "fuente_12v",
        "transformador",
        "camara",
        "nvr_dvr",
        "disco_duro",
        "switch_poe",
        "microfono",
        "parlante",
        "canaleta",
        "caja_estanca",
        "tornilleria_fijaciones",
        "otro",
    }

    def __init__(self, db: Session):
        self.db = db
        self._direcciones_csv_cache: dict[str, str] | None = None

    def _ruta_excel_materiales(self) -> Path:
        candidatos: list[Path] = []
        raw = str(getattr(settings, "materiales_excel_path", "") or "").strip()
        if raw:
            candidatos.append(Path(raw).expanduser())
        home = Path.home()
        candidatos.extend(
            [
                home / "Desktop" / "Hoja de cálculo sin título.xlsx",
                home / "Desktop" / "Hoja de calculo sin titulo.xlsx",
            ]
        )
        desktop = home / "Desktop"
        if desktop.exists():
            objetivo = self._normalizar_texto("Hoja de cálculo sin título")
            for archivo in desktop.glob("*.xlsx"):
                nombre_norm = self._normalizar_texto(archivo.stem)
                if objetivo in nombre_norm or nombre_norm in objetivo:
                    candidatos.append(archivo)

        for candidato in candidatos:
            if candidato.exists() and candidato.is_file():
                return candidato
        return candidatos[0] if candidatos else home / "Desktop" / "Hoja de cálculo sin título.xlsx"

    def _cargar_catalogo_materiales_excel(self) -> list[dict[str, str]]:
        ruta = self._ruta_excel_materiales()
        try:
            mtime = ruta.stat().st_mtime
        except FileNotFoundError as exc:
            raise ValueError(f"No se encontro el Excel de materiales en: {ruta}") from exc

        with MATERIALES_EXCEL_CACHE_LOCK:
            cache_path = MATERIALES_EXCEL_CACHE.get("path")
            cache_mtime = MATERIALES_EXCEL_CACHE.get("mtime")
            cache_items = MATERIALES_EXCEL_CACHE.get("items")
            if cache_path == str(ruta) and cache_mtime == mtime and isinstance(cache_items, list):
                return cache_items

        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        materiales: list[dict[str, str]] = []
        vistos: set[str] = set()

        try:
            with zipfile.ZipFile(ruta) as zf:
                shared_strings = _read_xlsx_shared_strings(zf, ns)
                targets = _resolve_xlsx_sheet_targets(zf, ns)
                for target in targets:
                    if target not in zf.namelist():
                        continue
                    root = ET.fromstring(zf.read(target))
                    sheet_data = root.find("a:sheetData", ns)
                    if sheet_data is None:
                        continue
                    for row in sheet_data.findall("a:row", ns):
                        celdas = []
                        for cell in row.findall("a:c", ns):
                            value = _cell_value_from_xlsx(cell, shared_strings, ns).strip()
                            if value:
                                celdas.append(value)
                        if not celdas:
                            continue
                        nombre = " ".join(celdas).strip()
                        if not nombre:
                            continue
                        if self._normalizar_texto(nombre) == "materiales":
                            continue
                        nombre_norm = self._normalizar_texto(nombre)
                        if not nombre_norm or nombre_norm in vistos:
                            continue
                        vistos.add(nombre_norm)
                        idx = len(materiales) + 1
                        materiales.append(
                            {
                                "codigo": f"mat_{idx:03d}",
                                "nombre": nombre,
                                "unidad_sugerida": _inferir_unidad_material(nombre),
                                "nombre_normalizado": nombre_norm,
                            }
                        )
        except zipfile.BadZipFile as exc:
            raise ValueError(f"El archivo de materiales no es un .xlsx valido: {ruta}") from exc
        except Exception as exc:
            raise ValueError(f"No se pudo leer el Excel de materiales: {exc}") from exc

        with MATERIALES_EXCEL_CACHE_LOCK:
            MATERIALES_EXCEL_CACHE["path"] = str(ruta)
            MATERIALES_EXCEL_CACHE["mtime"] = mtime
            MATERIALES_EXCEL_CACHE["items"] = materiales
        return materiales

    def _asegurar_catalogo_materiales_sql(self) -> None:
        self.db.execute(
            text(
                """
                IF OBJECT_ID('catalogo_materiales', 'U') IS NULL
                BEGIN
                CREATE TABLE catalogo_materiales (
                    codigo VARCHAR(40) PRIMARY KEY,
                    nombre NVARCHAR(MAX) NOT NULL,
                    unidad_sugerida VARCHAR(40) NOT NULL DEFAULT 'unidad',
                    nombre_normalizado NVARCHAR(450) NOT NULL,
                    activo BIT NOT NULL DEFAULT 1,
                    created_at DATETIME2 DEFAULT GETDATE(),
                    updated_at DATETIME2 DEFAULT GETDATE()
                )
                END
                """
            )
        )
        self.db.execute(
            text(
                """
                IF OBJECT_ID('catalogo_materiales', 'U') IS NOT NULL
                   AND EXISTS (
                        SELECT 1
                        FROM sys.columns
                        WHERE object_id = OBJECT_ID('catalogo_materiales')
                          AND name = 'nombre_normalizado'
                          AND max_length = -1
                   )
                BEGIN
                    UPDATE catalogo_materiales
                    SET nombre_normalizado = LEFT(nombre_normalizado, 450)
                    WHERE LEN(nombre_normalizado) > 450;

                    ALTER TABLE catalogo_materiales
                    ALTER COLUMN nombre_normalizado NVARCHAR(450) NOT NULL;
                END
                """
            )
        )
        self.db.execute(
            text(
                """
                IF NOT EXISTS (
                    SELECT 1
                    FROM sys.indexes
                    WHERE name = 'ux_catalogo_materiales_nombre_normalizado'
                      AND object_id = OBJECT_ID('catalogo_materiales')
                )
                CREATE UNIQUE INDEX ux_catalogo_materiales_nombre_normalizado
                ON catalogo_materiales (nombre_normalizado)
                """
            )
        )
        self.db.execute(
            text(
                """
                IF NOT EXISTS (SELECT 1 FROM sys.indexes WHERE name='ix_catalogo_materiales_activo' AND object_id=OBJECT_ID('catalogo_materiales'))
                CREATE INDEX ix_catalogo_materiales_activo ON catalogo_materiales (activo)
                """
            )
        )

        for idx, nombre in enumerate(MATERIALES_CATALOGO_SEED, start=1):
            nombre_limpio = str(nombre or "").strip()
            if not nombre_limpio:
                continue
            self.db.execute(
                text(
                    """
                    MERGE INTO catalogo_materiales AS target
                    USING (SELECT :codigo AS codigo, :nombre AS nombre, :unidad_sugerida AS unidad_sugerida, :nombre_normalizado AS nombre_normalizado) AS source
                    ON target.nombre_normalizado = source.nombre_normalizado
                    WHEN MATCHED THEN UPDATE SET
                        target.nombre           = source.nombre,
                        target.unidad_sugerida  = source.unidad_sugerida,
                        target.activo           = 1,
                        target.updated_at       = GETDATE()
                    WHEN NOT MATCHED THEN INSERT (codigo, nombre, unidad_sugerida, nombre_normalizado, activo, updated_at)
                        VALUES (source.codigo, source.nombre, source.unidad_sugerida, source.nombre_normalizado, 1, GETDATE());
                    """
                ),
                {
                    "codigo": f"mat_{idx:03d}",
                    "nombre": nombre_limpio,
                    "unidad_sugerida": _inferir_unidad_material(nombre_limpio),
                    "nombre_normalizado": self._normalizar_texto(nombre_limpio),
                },
            )
        self.db.commit()

    def _cargar_catalogo_materiales_sql(self) -> list[dict[str, str]]:
        try:
            self._asegurar_catalogo_materiales_sql()
            rows = self.db.execute(
                text(
                    """
                    SELECT codigo, nombre, unidad_sugerida, nombre_normalizado
                    FROM catalogo_materiales
                    WHERE activo = 1
                    ORDER BY nombre
                    """
                )
            ).mappings().all()
            return [
                {
                    "codigo": str(row.get("codigo") or ""),
                    "nombre": str(row.get("nombre") or ""),
                    "unidad_sugerida": str(row.get("unidad_sugerida") or "unidad"),
                    "nombre_normalizado": str(row.get("nombre_normalizado") or ""),
                }
                for row in rows
            ]
        except Exception:
            self.db.rollback()
            return [
                {
                    "codigo": f"mat_{idx:03d}",
                    "nombre": nombre_limpio,
                    "unidad_sugerida": _inferir_unidad_material(nombre_limpio),
                    "nombre_normalizado": self._normalizar_texto(nombre_limpio),
                }
                for idx, nombre in enumerate(MATERIALES_CATALOGO_SEED, start=1)
                if (nombre_limpio := str(nombre or "").strip())
            ]

    def buscar_materiales_catalogo_sql(self, consulta: str = "", limite: int = 10) -> dict[str, Any]:
        q = self._normalizar_texto(consulta)
        if not q:
            return {"query": "", "total": 0, "items": []}

        limite = max(1, min(int(limite or 10), 20))
        catalogo = self._cargar_catalogo_materiales_sql()
        tokens = [t for t in q.split() if t]

        from difflib import SequenceMatcher

        resultados: list[dict[str, Any]] = []
        for item in catalogo:
            nombre_norm = item.get("nombre_normalizado", "")
            if not nombre_norm:
                continue
            score = 0.0
            matched = False
            if nombre_norm == q:
                score += 1000
                matched = True
            if nombre_norm.startswith(q):
                score += 400
                matched = True
            if q in nombre_norm:
                score += 250
                matched = True
            if tokens:
                matches = sum(1 for token in tokens if token in nombre_norm)
                score += matches * 120
                if nombre_norm.startswith(tokens[0]):
                    score += 40
                if matches:
                    matched = True
            similarity = SequenceMatcher(None, q, nombre_norm).ratio()
            score += similarity * 100
            if similarity >= 0.55:
                matched = True
            if not matched:
                continue
            resultados.append(
                {
                    "codigo": item["codigo"],
                    "nombre": item["nombre"],
                    "unidadSugerida": item["unidad_sugerida"],
                    "score": round(score, 2),
                }
            )

        resultados.sort(key=lambda x: (-float(x.get("score") or 0), self._normalizar_texto(x.get("nombre"))))
        total = len(resultados)
        resultados = resultados[:limite]
        return {"query": consulta, "total": total, "items": resultados}

    def buscar_materiales_excel(self, consulta: str = "", limite: int = 10) -> dict[str, Any]:
        return self.buscar_materiales_catalogo_sql(consulta, limite)

    def _terminate_known_registro_lockers(self) -> int:
        if self.db.bind.dialect.name != "postgresql":
            return 0
        rows = self.db.execute(
            text(
                """
                SELECT pid
                FROM pg_stat_activity
                WHERE datname = current_database()
                  AND state = 'idle in transaction'
                  AND query = :query
                  AND pid <> pg_backend_pid()
                """
            ),
            {"query": KNOWN_REGISTRO_BLOCKING_QUERY},
        ).scalars().all()

        killed = 0
        for pid in rows:
            ok = self.db.execute(text("SELECT pg_terminate_backend(:pid)"), {"pid": int(pid)}).scalar()
            if ok:
                killed += 1
        if killed:
            self.db.commit()
        return killed

    def _run_registro_query(self, loader, operation: str):
        try:
            return loader()
        except OperationalError as exc:
            self.db.rollback()
            if not _is_lock_timeout_error(exc):
                raise

            killed = self._terminate_known_registro_lockers()
            if killed:
                try:
                    return loader()
                except OperationalError as retry_exc:
                    self.db.rollback()
                    if _is_lock_timeout_error(retry_exc):
                        raise ValueError(
                            f"No se pudo {operation} porque la tabla Registro sigue bloqueada en PostgreSQL."
                        ) from retry_exc
                    raise

            raise ValueError(
                f"No se pudo {operation} porque la tabla Registro esta bloqueada en PostgreSQL."
            ) from exc

    def _ruta_csv_registro_incidencias(self) -> Path:
        # parents[2] = carpeta ATC (raíz de la app helpdesk).
        return Path(__file__).resolve().parents[2] / "Registro Incidencias - Registro.csv"

    def _load_env_runtime(self) -> dict[str, str]:
        out: dict[str, str] = {}
        try:
            env_path = Path(__file__).resolve().parents[2] / ".env"
            if not env_path.exists():
                return out
            for raw in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = raw.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                key = k.strip()
                val = v.strip().strip('"').strip("'")
                if key:
                    out[key] = val
        except Exception:
            return out
        return out

    @staticmethod
    def _to_bool_env(value: Any, default: bool = False) -> bool:
        if value is None:
            return default
        txt = str(value).strip().lower()
        if txt in {"1", "true", "yes", "on"}:
            return True
        if txt in {"0", "false", "no", "off"}:
            return False
        return default

    def _smtp_runtime_config(self) -> dict[str, Any]:
        env_file = self._load_env_runtime()
        env_get = lambda k, d="": (os.getenv(k) or env_file.get(k) or d)

        enabled = bool(settings.smtp_enabled)
        enabled = enabled or self._to_bool_env(env_get("SMTP_ENABLED", "false"), False)

        host = str(settings.smtp_host or env_get("SMTP_HOST", "")).strip()
        port_raw = settings.smtp_port or env_get("SMTP_PORT", "587")
        try:
            port = int(port_raw)
        except Exception:
            port = 587
        username = str(settings.smtp_username or env_get("SMTP_USERNAME", "")).strip()
        password = str(settings.smtp_password or env_get("SMTP_PASSWORD", ""))
        from_email = str(settings.smtp_from_email or env_get("SMTP_FROM_EMAIL", username)).strip()
        from_name = str(settings.smtp_from_name or env_get("SMTP_FROM_NAME", "ATC Incidencias")).strip()
        use_tls = self._to_bool_env(settings.smtp_use_tls, True)
        use_tls = self._to_bool_env(env_get("SMTP_USE_TLS", str(use_tls).lower()), use_tls)
        use_ssl = self._to_bool_env(settings.smtp_use_ssl, False)
        use_ssl = self._to_bool_env(env_get("SMTP_USE_SSL", str(use_ssl).lower()), use_ssl)
        timeout_raw = settings.smtp_timeout_sec or env_get("SMTP_TIMEOUT_SEC", "20")
        try:
            timeout = int(timeout_raw)
        except Exception:
            timeout = 20

        bcc_raw = str(settings.smtp_bcc_emails or env_get("SMTP_BCC_EMAILS", "")).strip()
        bcc_emails = [e.strip() for e in bcc_raw.split(",") if e.strip()]

        return {
            "enabled": enabled,
            "host": host,
            "port": port,
            "username": username,
            "password": password,
            "from_email": from_email,
            "from_name": from_name,
            "use_tls": use_tls,
            "use_ssl": use_ssl,
            "timeout": timeout,
            "bcc_emails": bcc_emails,
        }

    def _smtp2_runtime_config(self) -> dict[str, Any]:
        env_file = self._load_env_runtime()
        env_get = lambda k, d="": (os.getenv(k) or env_file.get(k) or d)

        enabled = bool(settings.smtp2_enabled)
        enabled = enabled or self._to_bool_env(env_get("SMTP2_ENABLED", "false"), False)
        if not enabled:
            return {"enabled": False}

        host = str(settings.smtp2_host or env_get("SMTP2_HOST", "")).strip()
        port_raw = settings.smtp2_port or env_get("SMTP2_PORT", "587")
        try:
            port = int(port_raw)
        except Exception:
            port = 587
        username = str(settings.smtp2_username or env_get("SMTP2_USERNAME", "")).strip()
        password = str(settings.smtp2_password or env_get("SMTP2_PASSWORD", ""))
        from_email = str(settings.smtp2_from_email or env_get("SMTP2_FROM_EMAIL", username)).strip()
        from_name = str(settings.smtp2_from_name or env_get("SMTP2_FROM_NAME", "ATC Incidencias")).strip()
        use_tls = self._to_bool_env(settings.smtp2_use_tls, True)
        use_tls = self._to_bool_env(env_get("SMTP2_USE_TLS", str(use_tls).lower()), use_tls)
        use_ssl = self._to_bool_env(settings.smtp2_use_ssl, False)
        use_ssl = self._to_bool_env(env_get("SMTP2_USE_SSL", str(use_ssl).lower()), use_ssl)
        timeout_raw = settings.smtp2_timeout_sec or env_get("SMTP2_TIMEOUT_SEC", "20")
        try:
            timeout = int(timeout_raw)
        except Exception:
            timeout = 20

        return {
            "enabled": True,
            "host": host,
            "port": port,
            "username": username,
            "password": password,
            "from_email": from_email,
            "from_name": from_name,
            "use_tls": use_tls,
            "use_ssl": use_ssl,
            "timeout": timeout,
        }

    def _contacto_smtp_runtime_config(self) -> dict[str, Any]:
        """Cuenta contacto@alguientecuida.cl, ya provisionada en .env (misma que
        usan contrato_diario_service.py / ley_karin_service.py / compras_service.py)."""
        env_file = self._load_env_runtime()
        env_get = lambda k, d="": (os.getenv(k) or env_file.get(k) or d)

        host = str(env_get("CONTACTO_SMTP_HOST", "smtp.gmail.com")).strip()
        port_raw = env_get("CONTACTO_SMTP_PORT", "587")
        try:
            port = int(port_raw)
        except Exception:
            port = 587
        username = str(env_get("CONTACTO_SMTP_USERNAME", "")).strip()
        password = str(env_get("CONTACTO_SMTP_PASSWORD", ""))
        from_email = str(env_get("CONTACTO_SMTP_FROM_EMAIL", username)).strip()
        from_name = str(env_get("CONTACTO_SMTP_FROM_NAME", "Alguien Te Cuida")).strip()
        use_tls = self._to_bool_env(env_get("CONTACTO_SMTP_USE_TLS", "true"), True)

        return {
            "enabled": True,
            "host": host,
            "port": port,
            "username": username,
            "password": password,
            "from_email": from_email,
            "from_name": from_name,
            "use_tls": use_tls,
            "use_ssl": False,
            "timeout": 20,
        }

    def _visita_smtp_runtime_config(self) -> dict[str, Any]:
        """Cuenta catalina.silva@soporteatc.cl para el aviso de visita tecnica ATC."""
        env_file = self._load_env_runtime()
        env_get = lambda k, d="": (os.getenv(k) or env_file.get(k) or d)

        host = str(env_get("SMTP_VISITA_HOST", "mail.soporteatc.cl")).strip()
        port_raw = env_get("SMTP_VISITA_PORT", "587")
        try:
            port = int(port_raw)
        except Exception:
            port = 587
        username = str(env_get("SMTP_VISITA_USERNAME", "")).strip()
        password = str(env_get("SMTP_VISITA_PASSWORD", ""))
        from_email = str(env_get("SMTP_VISITA_FROM_EMAIL", username)).strip()
        from_name = str(env_get("SMTP_VISITA_FROM_NAME", "Alguien Te Cuida")).strip()
        use_tls = self._to_bool_env(env_get("SMTP_VISITA_USE_TLS", "true"), True)

        return {
            "enabled": bool(host and username and password),
            "host": host,
            "port": port,
            "username": username,
            "password": password,
            "from_email": from_email,
            "from_name": from_name,
            "use_tls": use_tls,
            "use_ssl": False,
            "timeout": 20,
        }

    def _logo_atc_bytes(self) -> bytes | None:
        try:
            logo_path = Path(__file__).resolve().parents[2] / "static" / "img" / "logo-atc2.jpeg"
            if not logo_path.exists():
                return None
            return logo_path.read_bytes()
        except Exception:
            return None

    @staticmethod
    def _parse_fecha_visita(fecha_raw: str | None) -> datetime | None:
        valor = str(fecha_raw or "").strip()
        if not valor:
            return None
        try:
            # Soporta date/datetime ISO: YYYY-MM-DD o YYYY-MM-DDTHH:MM
            return datetime.fromisoformat(valor)
        except Exception:
            pass
        try:
            # Soporta formato tabla: dd/mm/yyyy HH:MM
            return datetime.strptime(valor, "%d/%m/%Y %H:%M")
        except Exception:
            pass
        try:
            # Soporta formato simple: dd/mm/yyyy
            return datetime.strptime(valor, "%d/%m/%Y")
        except Exception:
            return None

    def _build_correo_visita_html(
        self,
        *,
        odt: str,
        sucursal: str,
        problema: str,
        estado: str,
        tecnico: str,
        acompanante: str,
        fecha_visita: datetime,
        observacion: str,
    ) -> tuple[str, str, str]:
        fecha_txt = fecha_visita.strftime("%d/%m/%Y")
        tecnico_txt = tecnico or "Por confirmar"
        acompanante_txt = acompanante or ""
        subject = f"Aviso de Visita Técnica ATC - {sucursal}"

        html_body = f"""\
<!doctype html>
<html>
  <body style="margin:0;padding:0;background:#f3f5f8;font-family:Segoe UI,Arial,sans-serif;color:#0f172a;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f3f5f8;padding:30px 14px;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:720px;background:#ffffff;border:1px solid #dde4ec;border-radius:14px;overflow:hidden;box-shadow:0 8px 24px rgba(2,18,36,0.08);">
            <tr>
              <td style="background:#0f3048;padding:12px 0;"></td>
            </tr>
            <tr>
              <td style="padding:24px 28px 12px;text-align:center;">
                <img src="cid:logoatc" alt="ATC" style="height:54px;width:auto;display:block;margin:0 auto 10px;" />
                <div style="font-size:34px;line-height:1.18;font-weight:800;color:#0f3048;letter-spacing:0.2px;">Aviso de Visita Técnica ATC</div>
                <div style="font-size:15px;color:#3b4b5c;margin-top:8px;">Alguien Te Cuida</div>
              </td>
            </tr>
            <tr>
              <td style="padding:8px 28px 26px;">
                <p style="margin:0 0 14px;font-size:21px;line-height:1.45;color:#1b2a3a;">
                  Estimados/as,
                </p>
                <p style="margin:0 0 16px;font-size:22px;line-height:1.5;color:#1b2a3a;">
                  Informamos que el equipo de <b>Servicio Tecnico de Alguien Te Cuida</b> el dia
                  <b>{html_escape(fecha_txt)}</b>, realizara una visita tecnica a la dependencia
                  "<b>{html_escape(sucursal)}</b>".
                </p>

                <div style="margin:0 0 14px;background:#f7f9fc;border:1px solid #d8e0ea;border-radius:10px;padding:14px 16px;">
                  <div style="font-size:20px;font-weight:800;color:#10263a;margin-bottom:8px;">Tecnicos asignados:</div>
                  <div style="font-size:21px;line-height:1.55;color:#1f3347;">{html_escape(tecnico_txt)}</div>
                  {f'<div style="font-size:21px;line-height:1.55;color:#1f3347;">{html_escape(acompanante_txt)}</div>' if acompanante_txt else ''}
                </div>
                <div style="font-size:16px;line-height:1.6;color:#4a5b6c;margin-bottom:16px;">
                  (Sujeto a modificaciones, de ser asi se le notificara por este mismo medio)
                </div>

                <div style="margin-top:14px;background:#f4f8ff;border:1px solid #cfe0ff;border-left:4px solid #2a5fa0;border-radius:10px;padding:12px 14px;color:#1f3b5f;font-size:16px;line-height:1.7;">
                  Esta Visita Tecnica se realizara entre las <b>09:00 AM</b> y las <b>18:00 PM</b>,
                  la que tiene como objetivo asegurar la continuidad y correcta operacion de los servicios contratados.
                </div>

                <p style="margin:18px 0 0;font-size:17px;line-height:1.65;color:#2a3a4a;">
                  Agradecemos mantener acceso disponible a las zonas de trabajo e informar si existiese alguna eventual reprogramacion.
                </p>

                <p style="margin:20px 0 0;font-size:17px;line-height:1.65;color:#2a3a4a;">
                  Saludos cordiales,<br />
                  <span style="font-weight:700;color:#1f5fa3;">Equipo Tecnico</span><br />
                  Alguien Te Cuida
                </p>
              </td>
            </tr>
            <tr>
              <td style="border-top:1px solid #e2e8f0;padding:14px 24px;text-align:center;font-size:12px;color:#6b7c8f;background:#fafbfd;">
                Mensaje generado automaticamente por ATC - Servicio Tecnico.
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
"""

        text_body = "\n".join(
            [
                f"Aviso de Visita Técnica ATC - {sucursal}",
                "",
                "Estimados/as,",
                "",
                f"Informamos que el equipo de Servicio Tecnico el dia {fecha_txt},",
                f"realizara una visita tecnica a la dependencia \"{sucursal}\".",
                "",
                "Tecnicos asignados:",
                tecnico_txt,
                *([acompanante_txt] if acompanante_txt else []),
                "(Sujeto a modificaciones, de ser asi se le notificara por este mismo medio)",
                "",
                "Esta Visita Tecnica se realizara entre las 09:00 AM y las 18:00 PM,",
                "la que tiene como objetivo asegurar la continuidad y correcta operacion de los servicios contratados.",
                "",
                "Agradecemos mantener acceso disponible a las zonas de trabajo para el óptimo servicio e informar si existiese alguna eventual reprogramacion.",
                "",
                "Saludos cordiales,",
                "Equipo Tecnico",
                "ATC - Alguien Te Cuida",
            ]
        )
        return subject, text_body + "\n", html_body

    def _direcciones_desde_csv(self) -> dict[str, str]:
        if self._direcciones_csv_cache is not None:
            return self._direcciones_csv_cache

        ruta = self._ruta_csv_registro_incidencias()
        out: dict[str, str] = {}
        if not ruta.exists():
            self._direcciones_csv_cache = out
            return out

        try:
            with ruta.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
                reader = csv.DictReader(fh)
                if not reader.fieldnames:
                    self._direcciones_csv_cache = out
                    return out

                headers = {self._normalizar_texto(h): h for h in reader.fieldnames if h}
                col_sucursal = next((headers.get(key) for key in ["sucursal", "cliente", "nombre sucursal", "nombre cliente"] if headers.get(key)), None)
                col_direccion = next((headers.get(key) for key in ["direccion", "direccion sucursal", "direccion trabajos", "direccion cliente"] if headers.get(key)), None)
                if not col_sucursal or not col_direccion:
                    self._direcciones_csv_cache = out
                    return out

                for row in reader:
                    sucursal = str(row.get(col_sucursal) or "").strip()
                    direccion = str(row.get(col_direccion) or "").strip()
                    if not sucursal or not direccion:
                        continue
                    key = self._normalizar_texto(sucursal)
                    if key and key not in out:
                        out[key] = direccion
        except Exception:
            out = {}

        self._direcciones_csv_cache = out
        return out

    def _obtener_tecnicos_helpdesk(self, solo_activos: bool = True) -> list[str]:
        db_url = (settings.support_db_url or "").strip()
        if not db_url:
            return []
        schema = (settings.support_db_schema or "public").strip() or "public"

        try:
            eng = build_engine(db_url, pool_pre_ping=True)
            with eng.connect() as conn:
                cols = conn.execute(
                    text(
                        """
                        SELECT column_name
                        FROM information_schema.columns
                        WHERE table_schema = :schema_name
                          AND table_name = 'incidencias_tecnicos'
                        """
                    ),
                    {"schema_name": schema},
                ).all()
                colset = {str(c[0]).strip() for c in cols if c and c[0]}
                if not colset:
                    return []

                col_nombre = next((c for c in ["nombre", "tecnico", "nombre_tecnico"] if c in colset), None)
                if not col_nombre:
                    return []
                col_activo = "activo" if "activo" in colset else None

                where = ""
                if solo_activos and col_activo:
                    where = f'WHERE "{col_activo}" = 1'

                rows = conn.execute(
                    text(
                        f'''
                        SELECT TRIM(CAST("{col_nombre}" AS NVARCHAR(MAX))) AS nombre
                        FROM "{schema}"."incidencias_tecnicos"
                        {where}
                        ORDER BY 1
                        '''
                    )
                ).all()
                out = [str(r[0]).strip() for r in rows if r and r[0] and str(r[0]).strip()]
                # Deduplicado preservando orden
                vistos: set[str] = set()
                unicos: list[str] = []
                for n in out:
                    k = self._normalizar_nombre_login(n)
                    if not k or k in vistos:
                        continue
                    vistos.add(k)
                    unicos.append(n)
                return unicos
        except Exception:
            return []

    def _schemas_con_tabla(self, table_name: str) -> list[str]:
        if table_name in _SCHEMAS_CON_TABLA_CACHE:
            return _SCHEMAS_CON_TABLA_CACHE[table_name]
        rows = self.db.execute(
            text(
                """
                SELECT DISTINCT table_schema
                FROM information_schema.columns
                WHERE table_name = :table_name
                  AND table_schema NOT IN ('pg_catalog', 'information_schema')
                ORDER BY table_schema
                """
            ),
            {"table_name": table_name},
        ).all()
        resultado = [str(r[0]).strip() for r in rows if r and r[0]]
        _SCHEMAS_CON_TABLA_CACHE[table_name] = resultado
        return resultado

    def _columnas_tabla(self, schema_name: str, table_name: str) -> set[str]:
        cache_key = (schema_name, table_name)
        if cache_key in _COLUMNAS_TABLA_CACHE:
            return _COLUMNAS_TABLA_CACHE[cache_key]
        rows = self.db.execute(
            text(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = :schema_name
                  AND table_name = :table_name
                """
            ),
            {"schema_name": schema_name, "table_name": table_name},
        ).all()
        resultado = {str(r[0]).strip() for r in rows if r and r[0]}
        _COLUMNAS_TABLA_CACHE[cache_key] = resultado
        return resultado

    def _pick_col(self, cols: set[str], opciones: list[str]) -> str | None:
        return next((c for c in opciones if c in cols), None)

    def _reparar_texto_mojibake(self, valor: Any) -> str:
        txt = str(valor or "").strip()
        if not txt:
            return ""

        def _score(s: str) -> tuple[int, int]:
            marcadores = ("Ã", "Â", "â", "ð", "�")
            raros = sum(s.count(ch) for ch in marcadores)
            return (raros, len(s))

        actual = txt
        mejor = txt
        visto: set[str] = {txt}

        for _ in range(4):
            candidatos = [actual]
            try:
                candidatos.append(actual.encode("latin-1").decode("utf-8"))
            except Exception:
                pass
            try:
                candidatos.append(actual.encode("cp1252").decode("utf-8"))
            except Exception:
                pass

            candidatos = [c.strip() for c in candidatos if str(c or "").strip()]
            mejor_paso = min(candidatos, key=_score)
            if _score(mejor_paso) < _score(mejor):
                mejor = mejor_paso

            if mejor_paso in visto or mejor_paso == actual:
                break

            visto.add(mejor_paso)
            actual = mejor_paso

        return mejor.strip()

    def _normalizar_nombre_login(self, valor: Any) -> str:
        txt = self._reparar_texto_mojibake(valor).lower()
        txt = unicodedata.normalize("NFD", txt)
        txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt

    def _es_rut_login(self, valor: Any) -> bool:
        return bool(re.fullmatch(r"\d{5,8}-[0-9Kk]", str(valor or "").strip()))

    def _extraer_nombres_desde_texto(self, valor: Any) -> list[str]:
        txt = self._reparar_texto_mojibake(valor)
        if not txt or txt in {"-", "Todos", "todos"}:
            return []
        txt = re.sub(r"(?i)acompa(?:n|ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â±)ante\s*:\s*", ";", txt)
        txt = re.sub(r"(?i)tecnic(?:o|a|os|as)\s*:\s*", ";", txt)
        partes = re.split(r"[\n,;/|]+", txt)
        salida: list[str] = []
        for p in partes:
            nombre = re.sub(r"\s+", " ", p).strip(" -\t\r")
            if not nombre:
                continue
            if nombre in {"-", "Todos", "todos"}:
                continue
            salida.append(nombre)
        return salida

    def _usuarios_login_tecnicos(self) -> list[str]:
        nombres: dict[str, str] = {}
        excluidos = {
            "nicolas alfonso bravo rain",
            "nicolas bravo",
            "jesus sebastian gonzalez aguilera",
            "jesus gonzalez",
        }

        def _add(valor: Any) -> None:
            for nombre in self._extraer_nombres_desde_texto(valor):
                key = self._normalizar_nombre_login(nombre)
                if key in excluidos:
                    continue
                if key and key not in nombres:
                    nombres[key] = nombre

        # Fuente principal pedida por operaciÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³n: helpdesk.incidencias_tecnicos
        for nombre in self._obtener_tecnicos_helpdesk(solo_activos=True):
            _add(nombre)

        try:
            for v in self.db.scalars(select(Registro.tecnicos)).all():
                _add(v)
            for v in self.db.scalars(select(Registro.acompanante)).all():
                _add(v)
        except Exception:
            self.db.rollback()
        try:
            for v in self.db.scalars(select(AdministracionODT.tecnico)).all():
                _add(v)
            for v in self.db.scalars(select(AdministracionODT.acompanante)).all():
                _add(v)
        except Exception:
            self.db.rollback()
        try:
            for v in self.db.scalars(select(ServicioTecnicoVentaODT.tecnico_a_cargo)).all():
                _add(v)
            for v in self.db.scalars(select(ServicioTecnicoVentaODT.acompanante)).all():
                _add(v)
        except Exception:
            self.db.rollback()

        # Compatibilidad: usuarios internos del area soporte, ahora desde BBDD.
        for nombre in self._usuarios_login_por_area("soporte"):
            _add(nombre)

        return sorted(nombres.values(), key=lambda x: self._normalizar_nombre_login(x))

    def _area_code_desde_destino(self, destino: str | None) -> str | None:
        key = self._normalizar_nombre_login(destino).replace(" ", "")
        return AREA_DESTINOS.get(key)

    def _departamentos_usuario(self, user: User | None) -> list[str]:
        raw = str(user.department if user else "").strip()
        if not raw:
            return []
        return [p.strip() for p in re.split(r"[;,|]+", raw) if p.strip()]

    def _area_codes_usuario(self, user: User | None) -> list[str]:
        if user and user.role == "superadmin":
            return list(ADMIN_SELECTOR_AREA_CODES)
        codes: list[str] = []
        for department in self._departamentos_usuario(user):
            key = self._normalizar_nombre_login(department)
            for code in DEPARTMENT_AREAS.get(key, []):
                if code not in codes:
                    codes.append(code)
        return codes

    def _area_por_codigo(self, area_code: str | None) -> AreaInfo | None:
        if not area_code:
            return None
        info = AREA_INFO.get(area_code)
        if not info:
            return None
        return AreaInfo(code=area_code, name=info[0], department=info[1])

    def _es_destino_auto(self, destino: str | None) -> bool:
        key = self._normalizar_nombre_login(destino).replace(" ", "")
        return key in {"", "auto", "loginunico", "unificado", "panelselectorauto"}

    def _usuarios_login_todos(self) -> list[str]:
        nombres = list(
            self.db.scalars(
                select(User.name)
                .where(User.is_active == True)
                .order_by(User.name.asc())
            ).all()
        )
        return sorted(nombres, key=lambda x: self._normalizar_nombre_login(x))

    def _membership_principal_usuario(self, user: User | None) -> tuple[AreaInfo | None, None]:
        if not user:
            return None, None
        codes = self._area_codes_usuario(user)
        return (self._area_por_codigo(codes[0]), None) if codes else (None, None)

    def _destino_principal_usuario(self, user: User | None) -> str:
        area, _membership = self._membership_principal_usuario(user)
        if area and area.code in AREA_PANEL_DESTINOS:
            return AREA_PANEL_DESTINOS[area.code]
        if user and user.role == "admin":
            return "panelSelectorSoporte"
        return "panelSelectorServicio"

    def _usuario_oculto_en_opciones_area(self, user: User | None, area_code: str | None) -> bool:
        if str(area_code or "").strip() not in {"tecnicos", "guardia"}:
            return False
        nombre_norm = self._normalizar_nombre_login(getattr(user, "name", "") or "")
        return any(
            nombre_norm == prefix or nombre_norm.startswith(f"{prefix} ")
            for prefix in NON_SELECTABLE_SUPERVISOR_NAME_PREFIXES
        )

    def _usuarios_login_por_area(self, area_code: str) -> list[str]:
        usuarios = self.db.scalars(select(User).where(User.is_active == True)).all()
        nombres = [
            u.name
            for u in usuarios
            if area_code in self._area_codes_usuario(u)
            and not self._usuario_oculto_en_opciones_area(u, area_code)
        ]
        return sorted(nombres, key=lambda x: self._normalizar_nombre_login(x))

    def obtener_usuarios_login_detalle(self, destino: str = "tecnicos") -> list[dict[str, Any]]:
        if self._es_destino_auto(destino):
            detalle: list[dict[str, Any]] = []
            for user in self.db.scalars(select(User).where(User.is_active == True).order_by(User.name.asc())).all():
                codes = self._area_codes_usuario(user)
                area = self._area_por_codigo(codes[0]) if codes else None
                detalle.append(
                    {
                        "user_id": user.id,
                        "usuario": user.name,
                        "username": user.username,
                        "area": area.name if area else "",
                        "area_code": area.code if area else "",
                        "department": (area.department if area else None) or user.department,
                        "role": user.role,
                    }
                )
            return sorted(detalle, key=lambda item: self._normalizar_nombre_login(item["usuario"]))

        area_code = self._area_code_desde_destino(destino)
        if not area_code:
            return []
        detalle = []
        for user in self.db.scalars(select(User).where(User.is_active == True)).all():
            if area_code not in self._area_codes_usuario(user):
                continue
            if self._usuario_oculto_en_opciones_area(user, area_code):
                continue
            area = self._area_por_codigo(area_code)
            detalle.append({
                "user_id": user.id,
                "usuario": user.name,
                "username": user.username,
                "area": area.name if area else "",
                "area_code": area.code if area else area_code,
                "department": (area.department if area else None) or user.department,
                "role": user.role,
            })
        return sorted(detalle, key=lambda item: self._normalizar_nombre_login(item["usuario"]))

    def _usuarios_login_tabla_servicio(self) -> list[str]:
        return self._usuarios_login_por_area("servicio_tecnico")

    def _usuarios_login_incidencias(self) -> list[str]:
        return self._usuarios_login_por_area("incidencias")

    def _usuarios_login_venta(self) -> list[str]:
        return self._usuarios_login_por_area("venta")

    def _usuarios_login_finanzas(self) -> list[str]:
        return self._usuarios_login_por_area("finanzas")

    def _usuarios_login_administracion(self) -> list[str]:
        return self._usuarios_login_por_area("administracion")

    def _usuarios_login_operaciones(self) -> list[str]:
        return self._usuarios_login_por_area("operaciones")

    def _es_usuario_tabla_servicio(self, usuario: str) -> bool:
        usuario_norm = self._normalizar_nombre_login(usuario)
        permitidos = {
            self._normalizar_nombre_login(n): n for n in self._usuarios_login_tabla_servicio()
        }
        return usuario_norm in permitidos

    def obtener_usuarios_login_tecnicos(self, destino: str = "tecnicos") -> list[str]:
        if self._es_destino_auto(destino):
            return self._usuarios_login_todos()
        area_code = self._area_code_desde_destino(destino)
        if area_code and area_code != "tecnicos":
            return self._usuarios_login_por_area(area_code)
        return self._usuarios_login_tecnicos()

    def _buscar_usuario_login(self, nombre: str) -> User | None:
        rut = str(nombre or "").strip()
        if not self._es_rut_login(rut):
            return None
        usuarios = self.db.scalars(select(User).where(User.is_active == True)).all()
        for user in usuarios:
            if str(user.username or "").strip().lower() == rut.lower():
                return user
        return None

    def _usuario_tiene_area(self, user: User, area_code: str | None) -> bool:
        if not area_code:
            return True
        if user.role == "admin":
            return True
        return area_code in self._area_codes_usuario(user)

    def _membership_usuario_area(self, user: User | None, area_code: str | None) -> None:
        return None

    def _password_usuario_ok(self, user: User, clave: str) -> bool:
        stored = str(user.hashed_password or "")
        incoming = str(clave or "")
        if stored.startswith("plain:"):
            return secrets.compare_digest(stored.removeprefix("plain:"), incoming)
        if stored.startswith("$2"):
            try:
                return bcrypt.checkpw(incoming.encode("utf-8"), stored.encode("utf-8"))
            except Exception:
                pass
        try:
            return PWD_CONTEXT.verify(incoming, stored)
        except Exception:
            return secrets.compare_digest(stored, incoming)

    # =========================
    # LOGIN
    # =========================
    def _expira_fin_dia_utc(self) -> datetime:
        tz_name = (settings.timezone or "America/Santiago").strip() or "America/Santiago"
        tz = ZoneInfo(tz_name)
        now_local = datetime.now(tz)
        midnight_local = datetime.combine(now_local.date() + timedelta(days=1), time.min, tzinfo=tz)
        return midnight_local.astimezone(timezone.utc).replace(tzinfo=None)

    def _redirect_selector_areas(self, app_url: str, token: str) -> str:
        helpdesk = (settings.helpdesk_base_url or "").rstrip("/")
        base = helpdesk if helpdesk else app_url
        return f"{base}/sso/login?token={token}&next=/seleccionar-area"

    def _redirect_panel_destino(self, app_url: str, destino_ok: str, token: str) -> str:
        if destino_ok == "panelSelectorSoporte":
            helpdesk = (settings.helpdesk_base_url or "").rstrip("/")
            base = helpdesk if helpdesk else app_url
            return f"{base}/sso/login?token={token}&next=/panel?area=soporte"
        if destino_ok == "materiales":
            helpdesk = (settings.helpdesk_base_url or "").rstrip("/")
            base_url = helpdesk or app_url
            return f"{base_url}/materiales?token={quote_plus(token)}"
        if destino_ok in {"servicioTecnico", "panelSelectorServicio", "stVentas"}:
            return f"{app_url}?form=panelSelectorServicio&token={token}&next={destino_ok}"
        if destino_ok == "coordinacion":
            return f"{app_url}?form=coordinacion&token={token}"
        if destino_ok in {"panelSelectorCoordinacion", "tablaProtocolos", "envioProtocolosSemanales"}:
            return f"{app_url}?form=panelSelectorCoordinacion&token={token}&next={destino_ok}"
        if destino_ok == "panelSelectorGerencia":
            helpdesk = (settings.helpdesk_base_url or "").rstrip("/")
            base = helpdesk if helpdesk else app_url
            return f"{base}/sso/login?token={token}&next=/gerencia"
        if destino_ok in {"panelSelectorVenta", "registroCliente", "tablaCliente"}:
            return f"{app_url}/venta/panel-selector?token={token}&next={destino_ok}"
        if destino_ok in {"panelSelectorAdministracion", "tablaAdministracion"}:
            return f"{app_url}/venta/administracion?token={token}&next={destino_ok}"
        if destino_ok in {"panelSelectorFinanzas", "tablaFinanzas"}:
            return f"{app_url}/venta/finanzas?token={token}&next={destino_ok}"
        if destino_ok in {"panelSelectorOperaciones", "tablaOperaciones"}:
            return f"{app_url}/venta/operaciones?token={token}&next={destino_ok}"
        if destino_ok == "panelSelectorGuardia":
            return f"{app_url}/guardia?token={token}&next=panelSelectorGuardia"
        if destino_ok == "panelSelectorSupervisores":
            return f"{app_url}/supervisores?token={token}&next=panelSelectorSupervisores"
        if destino_ok == "panelSelectorRRHH":
            return f"{app_url}/rrhh?token={token}&next=panelSelectorRRHH"
        if destino_ok == "panelSelectorPrevencion":
            return f"{app_url}/prevencion?token={token}&next=panelSelectorPrevencion"
        if destino_ok == "resumenEquiposTecnicos":
            return f"{app_url}/resumen-equipos-tecnicos?token={token}"
        if destino_ok == "bitacora":
            return f"{app_url}/bitacora?token={token}"
        if destino_ok in {"incidencias", "panelSelector", "cierreAperturaClientes", "controlProtocolos"}:
            return f"{app_url}?form=panelSelector&token={token}&next={destino_ok}"
        return f"{app_url}?form={destino_ok}&token={token}"

    def check_login(
        self,
        nombre_tecnico: str,
        clave: str,
        token: str,
        app_url: str,
        destino: str = "pendientes",
    ) -> dict[str, Any]:
        nombre_limpio = str(nombre_tecnico or "").strip()
        if not self._es_rut_login(nombre_limpio):
            return {"success": False, "message": "Ingresa tu RUT"}
        token_limpio = str(token or "").strip() or str(uuid.uuid4())

        destino_norm = (destino or "").strip()
        destino_auto = self._es_destino_auto(destino_norm)
        if destino_norm == "tabla":
            destino_norm = "servicioTecnico"
        if destino_norm == "STVentas":
            destino_norm = "stVentas"
        destino_ok = (
            destino_norm
            if destino_norm
            in {
                "panelSelector",
                "panelSelectorSoporte",
                "panelSelectorServicio",
                "panelSelectorCoordinacion",
                "panelSelectorVenta",
                "panelSelectorAdministracion",
                "tablaAdministracion",
                "panelSelectorFinanzas",
                "tablaFinanzas",
                "panelSelectorOperaciones",
                "tablaOperaciones",
                "panelSelectorGuardia",
                "panelSelectorSupervisores",
                "panelSelectorGerencia",
                "panelSelectorRRHH",
                "panelSelectorPrevencion",
                "resumenEquiposTecnicos",
                "registroCliente",
                "tablaCliente",
                "incidencias",
                "cierreAperturaClientes",
                "controlProtocolos",
                "tablaProtocolos",
                "envioProtocolosSemanales",
                "pendientes",
                "tecnicos",
                "coordinacion",
                "servicioTecnico",
                "stVentas",
            }
            else "auto" if destino_auto else "tecnicos"
        )

        area_code = None if destino_auto else self._area_code_desde_destino(destino_ok)
        area_login = self._area_por_codigo(area_code)
        redirect_selector_areas = False

        user_login = self._buscar_usuario_login(nombre_limpio)
        if not user_login:
            return {"success": False, "message": "Usuario invalido"}
        clave_limpia = str(clave or "").strip()
        if not self._password_usuario_ok(user_login, clave_limpia):
            return {"success": False, "message": "Clave incorrecta"}

        if destino_auto:
            nombre_sesion = user_login.name
            user_sesion = user_login
            if len(self._area_codes_usuario(user_sesion)) > 1:
                destino_ok = "seleccionar-area"
                area_code = None
                area_login = None
                redirect_selector_areas = True
            else:
                destino_ok = self._destino_principal_usuario(user_sesion)
                area_code = self._area_code_desde_destino(destino_ok)
                area_login = self._area_por_codigo(area_code)
        else:
            if area_code and not self._usuario_tiene_area(user_login, area_code):
                user_area_codes = self._area_codes_usuario(user_login)
                if len(user_area_codes) == 1:
                    area_code = user_area_codes[0]
                    area_login = self._area_por_codigo(area_code)
                    destino_ok = AREA_PANEL_DESTINOS.get(area_code, self._destino_principal_usuario(user_login))
                else:
                    scope_name = area_login.name if area_login else area_code
                    return {"success": False, "message": f"Usuario no autorizado para {scope_name}"}
            nombre_sesion = user_login.name
            user_sesion = user_login

        self.db.merge(
            LoginSession(
                token=token_limpio,
                usuario=nombre_sesion,
                user_id=user_sesion.id if user_sesion else None,
                area_code=area_code,
                department=area_login.department if area_login else None,
                expires_at=expiracion_sesion(
                    user_sesion.id if user_sesion else None,
                    self._expira_fin_dia_utc(),
                ),
            )
        )
        self.db.commit()
        return {
            "success": True,
            "redirect": (
                self._redirect_selector_areas(app_url, token_limpio)
                if redirect_selector_areas
                else self._redirect_panel_destino(app_url, destino_ok, token_limpio)
            ),
            "token": token_limpio,
            "user_id": user_sesion.id if user_sesion else None,
        }

    def usuario_logueado_por_token(self, token: str) -> bool:
        if not token:
            return False
        now = datetime.utcnow()
        stmt = select(LoginSession).where(LoginSession.token == token, LoginSession.expires_at > now)
        return self.db.scalar(stmt) is not None

    def contar_areas_para_token(self, token: str) -> int:
        """Cuenta cuántas áreas (mapping de Soporte) tiene el usuario del token."""
        if not token:
            return 0
        sesion = self.db.query(LoginSession).filter(LoginSession.token == token).first()
        if not sesion or not sesion.user_id or sesion.expires_at <= datetime.utcnow():
            return 0
        user = self.db.get(User, int(sesion.user_id))
        if not user or not user.is_active:
            return 0
        return len(self._area_codes_usuario(user))

    def obtener_selector_areas(self, token: str) -> dict[str, Any] | None:
        token_limpio = str(token or "").strip()
        if not token_limpio:
            return None
        sesion = self.db.query(LoginSession).filter(LoginSession.token == token_limpio).first()
        if not sesion or not sesion.user_id or sesion.expires_at <= datetime.utcnow():
            return None
        user = self.db.get(User, int(sesion.user_id))
        if not user or not user.is_active:
            return None
        areas = []
        for code in self._area_codes_usuario(user):
            if code == "protocolos":
                continue
            info = self._area_por_codigo(code)
            if not info:
                continue
            parts = [p for p in re.split(r"[\s/_-]+", info.name) if p]
            areas.append(
                {
                    "code": code,
                    "title": info.name,
                    "initials": "".join(p[:1].upper() for p in parts)[:2] or code[:2].upper(),
                }
            )
        return {"usuario": str(user.name or "").strip(), "areas": areas}

    def seleccionar_area_para_token(self, token: str, area_code: str) -> str | None:
        token_limpio = str(token or "").strip()
        area_code_limpio = str(area_code or "").strip()
        if not token_limpio or not area_code_limpio:
            return None
        sesion = self.db.query(LoginSession).filter(LoginSession.token == token_limpio).first()
        if not sesion or not sesion.user_id or sesion.expires_at <= datetime.utcnow():
            return None
        user = self.db.get(User, int(sesion.user_id))
        if not user or not user.is_active or not self._usuario_tiene_area(user, area_code_limpio):
            return None
        area_info = self._area_por_codigo(area_code_limpio)
        sesion.area_code = area_code_limpio
        sesion.department = area_info.department if area_info else None
        self.db.commit()
        return AREA_PANEL_DESTINOS.get(area_code_limpio)

    def obtener_resumen_equipos_tecnicos_hoy(self, *, incluir_pendientes_prioritarios: bool = False) -> dict[str, Any]:
        tz = ZoneInfo(settings.timezone or "America/Santiago")
        hoy = datetime.now(tz).date()

        def _clean(value: Any) -> str:
            return re.sub(r"\s+", " ", self._reparar_texto_mojibake(value)).strip()

        def _nombre_corto(value: str) -> str:
            return self._nombre_corto_tecnico(value)

        equipos_por_patente = self.EQUIPOS_TECNICOS_POR_PATENTE
        patentes_fijas = self.PATENTES_TECNICOS_FIJAS
        patentes_por_tecnico: dict[str, str] = {}
        for patente, miembros in equipos_por_patente.items():
            for nombre in miembros:
                for alias in (nombre, _nombre_corto(nombre)):
                    key = self._normalizar_nombre_login(alias)
                    if key:
                        patentes_por_tecnico[key] = patente
        patentes_por_tecnico.update(self.ALIASES_PATENTE_TECNICO)

        def _patente_tecnico(value: str) -> str:
            nombre = _clean(value)
            nombre_corto = _nombre_corto(nombre)
            candidatos = [
                self._normalizar_nombre_login(nombre),
                self._normalizar_nombre_login(nombre.strip("*")),
                self._normalizar_nombre_login(nombre_corto),
            ]
            return next((patentes_por_tecnico.get(c) for c in candidatos if patentes_por_tecnico.get(c)), "")

        def _miembro_equipo(value: str, *, es_tecnico: bool = False) -> dict[str, str] | None:
            nombre = _nombre_corto(value)
            if not nombre:
                return None
            return {"nombre": nombre, "patente": _patente_tecnico(value) if es_tecnico else ""}

        def _miembros_camioneta(patente: str) -> list[dict[str, str]]:
            miembros = equipos_por_patente.get(patente) or ["Sin asignacion"]
            return [
                {
                    "nombre": nombre,
                    "patente": patente if index == 0 else "",
                }
                for index, nombre in enumerate(miembros)
            ]

        def _es_equipo_jason_perez(equipo: dict[str, Any]) -> bool:
            valores = [
                str(equipo.get("titulo") or ""),
                str(equipo.get("tecnico") or ""),
                str(equipo.get("acompanante") or ""),
            ]
            for miembro in equipo.get("miembros") or []:
                if isinstance(miembro, dict):
                    valores.append(str(miembro.get("nombre") or ""))
            return any(
                self._normalizar_nombre_login(valor).startswith("jason perez")
                for valor in valores
                if valor
            )

        def _fecha_referencia(row: Registro) -> datetime | None:
            return row.fecha_derivacion_tecnico or row.fecha_derivacion_area

        def _fecha_en_hoy(value: datetime | None) -> bool:
            if not value:
                return False
            if value.tzinfo is not None:
                return value.astimezone(tz).date() == hoy
            return value.date() == hoy

        def _estado_en_proceso(value: str) -> bool:
            estado_norm = self._normalizar_texto(value)
            return estado_norm == "en proceso"

        def _team_key(tecnico: str, acompanante: str) -> tuple[str, ...]:
            patente = _patente_tecnico(tecnico) or _patente_tecnico(acompanante)
            if patente:
                return (f"patente:{patente}",)
            names = [n for n in [tecnico, acompanante] if n]
            normalized = sorted({self._normalizar_nombre_login(n) for n in names if n})
            return tuple(normalized) if normalized else ("sin-equipo",)

        def _agregar_odt(
            *,
            odt: str,
            cliente: str,
            problema: str,
            detalle: str = "",
            detalle_alt: str = "",
            direccion: str = "",
            estado: str = "Pendiente",
            prioridad: Any = None,
            fecha_ref: datetime | None,
            tecnico: str,
            acompanante: str,
            origen: str = "",
            incluir_si_activa: bool = False,
            requerir_en_proceso: bool = True,
        ) -> None:
            nonlocal total_odt
            tecnico_limpio = _clean(tecnico)
            acompanante_limpio = _clean(acompanante)
            if tecnico_limpio in {"-", "--"}:
                tecnico_limpio = ""
            if acompanante_limpio in {"-", "--"}:
                acompanante_limpio = ""
            if not tecnico_limpio and not acompanante_limpio:
                return
            if requerir_en_proceso and not _estado_en_proceso(estado):
                return
            if not _fecha_en_hoy(fecha_ref) and not incluir_si_activa:
                return

            key = _team_key(tecnico_limpio, acompanante_limpio)
            if key not in equipos:
                patente_equipo = _patente_tecnico(tecnico_limpio) or _patente_tecnico(acompanante_limpio)
                if patente_equipo:
                    miembros = _miembros_camioneta(patente_equipo)
                else:
                    miembros = []
                    miembro_tec = _miembro_equipo(tecnico_limpio, es_tecnico=True) if tecnico_limpio else None
                    if miembro_tec:
                        miembros.append(miembro_tec)
                    miembro_aco = _miembro_equipo(acompanante_limpio, es_tecnico=False) if acompanante_limpio else None
                    if miembro_aco:
                        miembros.append(miembro_aco)
                    if not miembros:
                        miembros = [{"nombre": "Sin tecnico asignado", "patente": ""}]
                equipos[key] = {
                    "titulo": " + ".join(miembro["nombre"] for miembro in miembros),
                    "miembros": miembros,
                    "patente": patente_equipo,
                    "tecnico": tecnico_limpio,
                    "acompanante": acompanante_limpio,
                    "odts": [],
                }

            cliente_limpio = _clean(cliente)
            problema_limpio = _clean(problema)
            direccion_limpia = _clean(direccion)
            estado_limpio = _clean(estado) or "Pendiente"
            fecha_texto = _to_ddmmyyyy_hhmm(fecha_ref)
            # Solo Desconexion guarda su detalle estructurado (contacto,
            # prioridad) en detalle_problema — otros tipos como Problema de
            # Parlante lo guardan en observacion en su lugar (mismo formato
            # "[quien - fecha] detalle"). Si no hay ninguno de los dos, se
            # cae al nombre del problema, nunca al estado — pedido
            # explicito, ago 2026.
            detalle_limpio = _clean(detalle) or _clean(detalle_alt) or problema_limpio

            equipos[key]["odts"].append(
                {
                    "odt": _clean(odt),
                    "cliente": cliente_limpio,
                    "problema": problema_limpio,
                    "detalle": detalle_limpio,
                    "direccion": direccion_limpia,
                    "estado": estado_limpio,
                    "prioridad": prioridad,
                    "fecha": fecha_texto,
                    "origen": origen,
                }
            )
            total_odt += 1

        rows = (
            self.db.scalars(
                select(Registro)
                .where(
                    or_(
                        Registro.fecha_derivacion_tecnico.is_not(None),
                        Registro.fecha_derivacion_area.is_not(None),
                        Registro.tecnicos.is_not(None),
                        Registro.acompanante.is_not(None),
                    )
                )
                .order_by(
                    case((Registro.fecha_derivacion_tecnico.is_(None), 1), else_=0),
                    Registro.fecha_derivacion_tecnico.desc(),
                    case((Registro.fecha_derivacion_area.is_(None), 1), else_=0),
                    Registro.fecha_derivacion_area.desc(),
                    Registro.odt.asc(),
                )
            )
            .all()
        )

        equipos: dict[tuple[str, ...], dict[str, Any]] = {}
        total_odt = 0
        odts_agregadas: set[str] = set()
        for row in rows:
            fecha_ref = _fecha_referencia(row)

            tecnico = _clean(row.tecnicos)
            acompanante = _clean(row.acompanante)
            odt_key = self._normalizar_texto(row.odt)
            antes = total_odt
            _agregar_odt(
                odt=row.odt,
                cliente=row.cliente,
                problema=row.problema,
                detalle=row.detalle_problema,
                detalle_alt=row.observacion,
                direccion=row.direccion,
                estado=row.estado,
                prioridad=row.prioridad,
                fecha_ref=fecha_ref,
                tecnico=tecnico,
                acompanante=acompanante,
                origen="registro",
                incluir_si_activa=True,
            )
            if total_odt > antes and odt_key:
                odts_agregadas.add(odt_key)

        # Ventas (ODS) derivadas a un técnico y con la instalación aún no
        # finalizada: se suman al mismo tablero de equipos que las incidencias.
        try:
            rows_venta = (
                self.db.execute(
                    select(VentaODS, ServicioTecnicoVentaODT, AdministracionODT)
                    .outerjoin(
                        ServicioTecnicoVentaODT,
                        func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                    )
                    .outerjoin(
                        AdministracionODT,
                        func.lower(func.trim(AdministracionODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                    )
                    .where(VentaODS.estado != "Anulada")
                    .order_by(VentaODS.created_at.asc(), VentaODS.id.asc())
                )
                .all()
            )
        except Exception:
            rows_venta = []

        for ods, st, adm in rows_venta:
            if bool(getattr(st, "instalacion_finalizada", False)):
                continue
            odt_key = self._normalizar_texto(ods.codigo)
            if odt_key and odt_key in odts_agregadas:
                continue
            tecnico_venta = str(getattr(st, "tecnico_a_cargo", "") or getattr(adm, "tecnico", "") or "").strip()
            acompanante_venta = str(getattr(st, "acompanante", "") or getattr(adm, "acompanante", "") or "").strip()
            if not tecnico_venta and not acompanante_venta:
                continue
            fecha_ref = getattr(st, "updated_at", None) or getattr(adm, "fecha_derivacion", None) or ods.created_at
            antes = total_odt
            _agregar_odt(
                odt=ods.codigo,
                cliente=ods.nombre_sucursal or ods.razon_social or "",
                problema=ods.tipo_servicio or "",
                detalle=ods.observacion or ods.consideraciones or "",
                direccion=ods.direccion_sucursal or "",
                estado="En Proceso",
                fecha_ref=fecha_ref,
                tecnico=tecnico_venta,
                acompanante=acompanante_venta,
                origen="venta_ods",
                incluir_si_activa=True,
                requerir_en_proceso=False,
            )
            if total_odt > antes and odt_key:
                odts_agregadas.add(odt_key)

        equipos_ordenados = []
        for patente in patentes_fijas:
            key = (f"patente:{patente}",)
            equipo = equipos.get(key)
            if not equipo:
                equipo = {
                    "titulo": patente,
                    "miembros": _miembros_camioneta(patente),
                    "patente": patente,
                    "tecnico": "",
                    "acompanante": "",
                    "odts": [],
                }
            equipos_ordenados.append(equipo)

        equipos_extra = sorted(
            [
                equipo
                for key, equipo in equipos.items()
                if not (len(key) == 1 and str(key[0]).startswith("patente:") and equipo.get("patente") in patentes_fijas)
            ],
            key=lambda item: (-len(item["odts"]), self._normalizar_nombre_login(item["titulo"])),
        )
        odts_prioritarias: list[dict[str, Any]] = []
        equipos_extra_visibles: list[dict[str, Any]] = []
        for equipo in equipos_extra:
            if _es_equipo_jason_perez(equipo):
                odts_prioritarias.extend(equipo.get("odts") or [])
                continue
            equipos_extra_visibles.append(equipo)

        equipos_ordenados.extend(equipos_extra_visibles)
        if incluir_pendientes_prioritarios:
            equipos_ordenados.append(
                {
                    "titulo": "Pendientes Prioritarios",
                    "miembros": [{"nombre": "Pendientes Prioritarios", "patente": ""}],
                    "patente": "",
                    "tecnico": "",
                    "acompanante": "",
                    "odts": odts_prioritarias,
                    "columna_prioritaria": True,
                }
            )
        # Dentro de cada equipo, las ODT se muestran por prioridad (1 = más
        # urgente, definida en la tabla de Incidencias Servicio Técnico) en
        # vez del orden de derivación con el que llegaron arriba. Las que no
        # tienen prioridad asignada quedan al final, en el mismo orden en que
        # ya venían.
        def _clave_orden_prioridad(odt: dict[str, Any]) -> tuple[bool, int]:
            prioridad = odt.get("prioridad")
            return (prioridad is None, prioridad if prioridad is not None else 0)

        for equipo in equipos_ordenados:
            equipo["odts"].sort(key=_clave_orden_prioridad)

        return {
            "fecha": hoy.strftime("%d/%m/%Y"),
            "equipos": equipos_ordenados,
            "total_equipos": len(equipos_ordenados),
            "total_odt": total_odt,
        }

    def usuario_admin_para_resumen_equipos(self, token: str) -> bool:
        token_limpio = str(token or "").strip()
        if not token_limpio:
            return False
        sesion = self.db.query(LoginSession).filter(LoginSession.token == token_limpio).first()
        if not sesion or not sesion.user_id or sesion.expires_at <= datetime.utcnow():
            return False
        user = self.db.get(User, int(sesion.user_id))
        return bool(user and user.is_active and str(user.role or "").strip().lower() in {"admin", "superadmin"})

    def get_usuario_actual(self, token: str) -> str:
        if not token:
            return "Desconocido"
        now = datetime.utcnow()
        stmt = select(LoginSession).where(LoginSession.token == token, LoginSession.expires_at > now)
        sesion = self.db.scalar(stmt)
        return sesion.usuario if sesion else "Desconocido"

    def usuario_autorizado_para_tabla(self, token: str) -> bool:
        usuario = self.get_usuario_actual(token)
        if not usuario or usuario == "Desconocido":
            return False
        return self._es_usuario_tabla_servicio(usuario)

    def usuario_autorizado_para_resumen_equipos(self, token: str) -> bool:
        token_limpio = str(token or "").strip()
        if not token_limpio:
            return False
        sesion = self.db.query(LoginSession).filter(LoginSession.token == token_limpio).first()
        if not sesion or not sesion.user_id or sesion.expires_at <= datetime.utcnow():
            return False
        user = self.db.get(User, int(sesion.user_id))
        if not user or not user.is_active:
            return False
        codes = self._area_codes_usuario(user)
        return "servicio_tecnico" in codes or "resumen_equipos_tecnicos" in codes

    def logout(self, token: str) -> bool:
        sesion = self.db.get(LoginSession, token)
        if sesion:
            self.db.delete(sesion)
            self.db.commit()
        return True

    # =========================
    # REGISTRO / INCIDENCIAS
    # =========================
    def _proximo_odt(self, prefijo: str = "I") -> str:
        odts = self._run_registro_query(
            lambda: list(self.db.scalars(select(Registro.odt))),
            "obtener el correlativo de ODT",
        )
        mayor = 0
        for odt in odts:
            n = _parse_prefijo_numero(odt)
            if n is not None:
                mayor = max(mayor, n)
        return f"{prefijo}{mayor + 1}"

    def _ticket_soporte_existe(self, conn, schema: str, ticket_id: int | None) -> int | None:
        if not ticket_id:
            return None
        exists = conn.execute(
            text(f'SELECT id FROM "{schema}"."tickets" WHERE id = :ticket_id'),
            {"ticket_id": int(ticket_id)},
        ).scalar()
        return int(exists) if exists else None

    def _resolver_ticket_id_soporte_desde_registro(self, conn, schema: str, odt: str) -> int | None:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            return None

        for schema_registro in self._schemas_con_tabla("registro"):
            cols = self._columnas_tabla(schema_registro, "registro")
            col_odt = self._pick_col(cols, ["odt", "codigo_odt", "codigo", "nro_odt"])
            col_source_row = self._pick_col(cols, ["source_row", "source_id", "origen_id"])
            col_source_file = self._pick_col(cols, ["source_file", "source", "origin", "origen"])
            if not col_odt or not col_source_row:
                continue

            select_source = f'CAST("{col_source_file}" AS NVARCHAR(MAX)) AS source_file' if col_source_file else "'' AS source_file"
            row = self.db.execute(
                text(
                    f'''
                    SELECT CAST("{col_source_row}" AS NVARCHAR(MAX)) AS source_row,
                           {select_source}
                    FROM "{schema_registro}"."registro"
                    WHERE TRIM(CAST("{col_odt}" AS NVARCHAR(MAX))) = :odt
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    '''
                ),
                {"odt": odt_limpia},
            ).mappings().first()
            if not row:
                continue

            source_file_norm = self._normalizar_texto(row.get("source_file") or "")
            if col_source_file and source_file_norm not in {"ticket", "tickets", "helpdesk", "soporte"}:
                continue
            try:
                ticket_id = int(str(row.get("source_row") or "").strip())
            except Exception:
                continue
            return self._ticket_soporte_existe(conn, schema, ticket_id)

        return None

    def _resolver_ticket_id_soporte_desde_mensajes(self, conn, schema: str, odt: str) -> int | None:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            return None

        needle = f"%{odt_limpia}%"
        rows = conn.execute(
            text(
                f'''
                SELECT m.ticket_id, CAST(m.content AS NVARCHAR(MAX)) AS content
                FROM "{schema}"."messages" m
                JOIN "{schema}"."tickets" t ON t.id = m.ticket_id
                WHERE CAST(m.content AS NVARCHAR(MAX)) LIKE :needle
                ORDER BY m.created_at DESC, m.id DESC OFFSET 0 ROWS FETCH NEXT 50 ROWS ONLY
                '''
            ),
            {"needle": needle},
        ).mappings().all()
        patron_odt = re.compile(rf"(?<![A-Za-z0-9]){re.escape(odt_limpia)}(?![A-Za-z0-9])", re.IGNORECASE)
        for row in rows:
            if patron_odt.search(str(row.get("content") or "")):
                try:
                    return int(row.get("ticket_id"))
                except Exception:
                    continue
        return None

    def _resolver_ticket_id_soporte(self, conn, schema: str, odt: str) -> int | None:
        # Importante: la ODT es correlativo de Incidencias, no ID de ticket.
        # Nunca debe mapearse T15 -> ticket #15 por el numero.
        ticket_id = self._resolver_ticket_id_soporte_desde_registro(conn, schema, odt)
        if ticket_id:
            return ticket_id
        return self._resolver_ticket_id_soporte_desde_mensajes(conn, schema, odt)

    def _insertar_nota_ticket_soporte(self, conn, schema: str, ticket_id: int, contenido: str) -> None:
        contenido_limpio = str(contenido or "").strip()
        if not contenido_limpio:
            return
        conn.execute(
            text(
                f'''
                INSERT INTO "{schema}"."messages"
                    (ticket_id, sender_type, channel, content, is_internal_note, created_at)
                VALUES
                    (:ticket_id, :sender_type, :channel, :content, :is_internal_note, :created_at)
                '''
            ),
            {
                "ticket_id": ticket_id,
                "sender_type": "system",
                "channel": "internal",
                "content": contenido_limpio,
                "is_internal_note": True,
                "created_at": datetime.now(timezone.utc),
            },
        )

    def _actualizar_estado_ticket_soporte(
        self,
        odt: str,
        estado: str,
        nota_interna: str | None = None,
    ) -> bool:
        estado_limpio = _normalizar_estado_ticket_soporte(estado)
        if estado_limpio not in TICKET_STATUSES_PERMITIDOS:
            raise ValueError("Estado de ticket no permitido.")

        db_url = str(settings.support_db_url or "").strip()
        if not db_url:
            return False

        schema = (settings.support_db_schema or "public").strip() or "public"
        engine = build_engine(db_url, pool_pre_ping=True)
        ahora = datetime.utcnow()
        with engine.begin() as conn:
            ticket_id = self._resolver_ticket_id_soporte(conn, schema, odt)
            if ticket_id is None:
                return False

            result = conn.execute(
                text(
                    f'''
                    UPDATE "{schema}"."tickets"
                    SET status = CAST(:status AS varchar),
                        updated_at = :updated_at,
                        resolved_at = CASE
                            WHEN CAST(:status AS NVARCHAR(MAX)) IN (:resuelto, :resuelto_servicio, :resuelto_cliente) THEN COALESCE(resolved_at, :updated_at)
                            ELSE resolved_at
                        END,
                        closed_at = CASE
                            WHEN CAST(:status AS NVARCHAR(MAX)) = :cerrado THEN COALESCE(closed_at, :updated_at)
                            ELSE closed_at
                        END
                    WHERE id = :ticket_id
                    '''
                ),
                {
                    "status": estado_limpio,
                    "updated_at": ahora,
                    "resuelto": TICKET_STATUS_RESUELTO,
                    "resuelto_servicio": TICKET_STATUS_RESUELTO_SERVICIO,
                    "resuelto_cliente": TICKET_STATUS_RESUELTO_CLIENTE,
                    "cerrado": TICKET_STATUS_CERRADO,
                    "ticket_id": ticket_id,
                },
            )
            if nota_interna:
                try:
                    self._insertar_nota_ticket_soporte(conn, schema, ticket_id, nota_interna)
                except Exception:
                    LOGGER.exception("No se pudo agregar nota interna al ticket soporte para ODT %s.", odt)
            return int(result.rowcount or 0) > 0

    def _sync_estado_ticket_soporte_silencioso(
        self,
        odt: str,
        estado: str,
        nota_interna: str | None = None,
    ) -> None:
        try:
            self._actualizar_estado_ticket_soporte(odt, estado, nota_interna=nota_interna)
        except Exception:
            LOGGER.exception("No se pudo actualizar estado de ticket soporte para ODT %s.", odt)

    def _build_nota_cierre_ticket_soporte(
        self,
        *,
        odt: str,
        estado_ticket: str,
        derivacion: str,
        observacion_final: str,
    ) -> str:
        estado_label = {
            TICKET_STATUS_RESUELTO_CLIENTE: "Resuelto Cliente",
            TICKET_STATUS_RESUELTO_SERVICIO: "Resuelto Servicio",
            TICKET_STATUS_RESUELTO: "Resuelto",
            TICKET_STATUS_CERRADO: "Cerrado",
        }.get(_normalizar_estado_ticket_soporte(estado_ticket), estado_ticket)
        detalle = str(observacion_final or "").strip() or "Sin observacion final registrada."
        return (
            "<strong>Actualizacion automatica desde Incidencias</strong><br>"
            f"ODT: {html_escape(str(odt or '').strip())}<br>"
            f"Estado: {html_escape(estado_label)}<br>"
            f"Derivacion/Cierre: {html_escape(str(derivacion or '').strip() or '-')}<br>"
            f"Resultado final: {html_escape(detalle).replace(chr(10), '<br>')}"
        )


    def _proximo_odt_incidencias(self, prefijo: str = "I") -> str:
        return self._proximo_odt(prefijo)

    def _normalizar_nombre_sucursal_match(self, valor: Any) -> str:
        txt = self._normalizar_texto(valor)
        txt = re.sub(r"[^a-z0-9]+", " ", txt)
        return re.sub(r"\s+", " ", txt).strip()

    def _score_nombre_sucursal_match(self, objetivo: str, candidato: str) -> int:
        obj = self._normalizar_nombre_sucursal_match(objetivo)
        cand = self._normalizar_nombre_sucursal_match(candidato)
        if not obj or not cand:
            return 0
        if obj == cand:
            return 100
        # Antes esto devolvía 75 fijo con solo que uno fuera substring del otro,
        # sin importar cuánto abarcaba — un candidato corto tipo "DIDECO" que
        # aparece por casualidad dentro de un nombre mucho más largo y distinto
        # ("Imq Carozzi 2 Dideco/secpla/obras") le ganaba a la sucursal correcta
        # ("IMQ Dideco/secpla/obras", que no matcheaba como substring por el
        # "Carozzi 2" en el medio) aunque esta última compartiera más tokens.
        # Causó que una Mantención Preventiva de Quilpué quedara con la dirección
        # de una sucursal de Quintero (ago 2026). Ahora se escala por cuánto del
        # string más largo cubre el más corto, para que un match parcial chico no
        # opaque a un match por tokens más fuerte.
        if len(obj) >= 4 and len(cand) >= 4 and (obj in cand or cand in obj):
            cobertura_substr = min(len(obj), len(cand)) / max(len(obj), len(cand))
            substr_score = int(75 * cobertura_substr)
        else:
            substr_score = 0

        stopwords = {"de", "del", "la", "el", "los", "las", "y", "ex", "n", "s", "sn", "sin"}
        obj_tokens = {t for t in obj.split() if len(t) > 1 and t not in stopwords}
        cand_tokens = {t for t in cand.split() if len(t) > 1 and t not in stopwords}
        if not obj_tokens or not cand_tokens:
            return substr_score
        inter = obj_tokens & cand_tokens
        if not inter:
            return substr_score
        cobertura_obj = len(inter) / max(1, len(obj_tokens))
        if len(obj_tokens) <= 2 and cobertura_obj < 1:
            return substr_score
        # Coeficiente de Dice (simétrico): premia que la intersección cubra bien
        # AMBOS lados, no solo el objetivo. La fórmula anterior (cobertura del
        # objetivo * 60 + cantidad de tokens en común) le daba solo 52 puntos a
        # "IMQ Dideco/secpla/obras" contra el objetivo "Imq Carozzi 2
        # Dideco/secpla/obras" (comparte 4 de 5 tokens) — por debajo del piso de
        # 60 para aceptar el match, aunque fuera claramente la sucursal correcta.
        token_score = int(200 * len(inter) / (len(obj_tokens) + len(cand_tokens)))
        return max(substr_score, token_score)

    def _direccion_cliente(self, cliente: str) -> str:
        cliente_txt = str(cliente or "").strip()
        if not cliente_txt:
            return ""

        try:
            objetivo_norm = self._normalizar_nombre_sucursal_match(cliente_txt)
            rows_sucursales = self.db.execute(
                select(SucursalBBDD.nombre_sucursal, SucursalBBDD.direccion_sucursal)
                .where(SucursalBBDD.direccion_sucursal.is_not(None))
            ).all()
            mejor_direccion = ""
            mejor_score = 0
            for nombre_sucursal, direccion_sucursal in rows_sucursales:
                direccion = str(direccion_sucursal or "").strip()
                if not direccion:
                    continue
                nombre_norm = self._normalizar_nombre_sucursal_match(nombre_sucursal)
                if nombre_norm == objetivo_norm:
                    return direccion
                score = self._score_nombre_sucursal_match(objetivo_norm, str(nombre_sucursal or ""))
                if score > mejor_score:
                    mejor_score = score
                    mejor_direccion = direccion
            if mejor_score >= 60 and mejor_direccion:
                return mejor_direccion
        except Exception:
            self.db.rollback()

        try:
            for schema_name in self._schemas_con_tabla("catalogo_clientes"):
                cols = self._columnas_tabla(schema_name, "catalogo_clientes")
                col_cliente = self._pick_col(cols, ["nombre_sucursal", "nombre_cliente", "sucursal", "cliente"])
                col_direccion = self._pick_col(cols, ["direccion", "direccion_sucursal", "direccion_trabajos", "direccion_cliente"])
                if not col_cliente or not col_direccion:
                    continue

                sql = text(
                    f"""
                    SELECT COALESCE(CAST("{col_direccion}" AS NVARCHAR(MAX)), '') AS direccion
                    FROM "{schema_name}"."catalogo_clientes"
                    WHERE lower(TRIM(CAST("{col_cliente}" AS NVARCHAR(MAX)))) = lower(:cliente)
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                )
                row = self.db.execute(sql, {"cliente": cliente_txt}).mappings().first()
                if row:
                    direccion = str(row.get("direccion") or "").strip()
                    if direccion:
                        return direccion
        except Exception:
            self.db.rollback()

        try:
            stmt = select(ClienteBBDD.direccion).where(ClienteBBDD.cliente == cliente_txt)
            direccion = self.db.scalar(stmt) or ""
            if direccion:
                return str(direccion).strip()
        except Exception:
            self.db.rollback()

        try:
            return self._direcciones_desde_csv().get(self._normalizar_texto(cliente_txt), "")
        except Exception:
            return ""

    def _normalizar_direccion(self, valor: Any) -> str:
        txt = self._normalizar_texto(valor)
        txt = re.sub(r"[^a-z0-9\s]", " ", txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt

    def _extraer_numero_direccion(self, valor: Any) -> str:
        txt = str(valor or "").strip()
        if not txt:
            return ""
        m = re.search(r"\b(\d{2,6})\b", txt)
        return m.group(1) if m else ""

    def _contacto_preferente_sucursal(self, cliente: str) -> dict[str, str]:
        salida = {"contacto": "", "telefono": "", "correo": ""}
        cliente_norm = self._normalizar_texto(cliente)
        if not cliente_norm:
            return salida

        try:
            contactos = self.obtener_contactos_por_sucursal()
            sucursal_key = next(
                (key for key in contactos.keys() if self._normalizar_texto(key) == cliente_norm),
                None,
            )
            if not sucursal_key:
                return salida

            candidatos = contactos.get(sucursal_key) or []
            if not candidatos:
                return salida

            primero = candidatos[0]
            salida["contacto"] = str(primero.get("nombre") or "").strip()
            salida["telefono"] = str(primero.get("telefono") or "").strip()
            salida["correo"] = str(primero.get("email") or "").strip()
        except Exception:
            return salida

        return salida

    def _coordenadas_por_direccion_bd(self, direccion: str) -> tuple[str, str]:
        direccion_txt = str(direccion or "").strip()
        direccion_norm = self._normalizar_direccion(direccion_txt)
        nro_ref = self._extraer_numero_direccion(direccion_txt)
        if not direccion_norm:
            return "", ""

        for table_name in ["bbdd_clientes", "catalogo_clientes"]:
            try:
                for schema_name in self._schemas_con_tabla(table_name):
                    cols = self._columnas_tabla(schema_name, table_name)
                    if not cols:
                        continue
                    col_dir = self._pick_col(cols, ["direccion", "direccion_sucursal", "direccion_trabajos", "direccion_cliente"])
                    col_lat = self._pick_col(cols, ["latitud", "lat", "latitude"])
                    col_lng = self._pick_col(cols, ["longitud", "lng", "lon", "longitude"])
                    if not col_dir or not col_lat or not col_lng:
                        continue

                    sql = text(
                        f"""
                        SELECT
                            COALESCE(CAST("{col_lat}" AS NVARCHAR(MAX)), '') AS latitud,
                            COALESCE(CAST("{col_lng}" AS NVARCHAR(MAX)), '') AS longitud
                        FROM "{schema_name}"."{table_name}"
                        WHERE lower(TRIM(CAST("{col_dir}" AS NVARCHAR(MAX))))
                              = lower(:direccion)
                          AND TRIM(CAST("{col_lat}" AS NVARCHAR(MAX))) <> ''
                          AND TRIM(CAST("{col_lng}" AS NVARCHAR(MAX))) <> ''
                        OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                        """
                    )
                    row = self.db.execute(sql, {"direccion": direccion_txt}).mappings().first()
                    if row:
                        lat = str(row.get("latitud") or "").strip()
                        lng = str(row.get("longitud") or "").strip()
                        if lat and lng:
                            return lat, lng
            except Exception:
                continue

        # Fallback BD flexible: tolera variaciones de formato en direccion
        # (parentesis, comas, espacios, etc.).
        mejor_score = -1
        mejor_coords: tuple[str, str] = ("", "")
        tokens_ref = set(direccion_norm.split())

        for table_name in ["bbdd_clientes", "catalogo_clientes"]:
            try:
                for schema_name in self._schemas_con_tabla(table_name):
                    cols = self._columnas_tabla(schema_name, table_name)
                    if not cols:
                        continue
                    col_dir = self._pick_col(cols, ["direccion", "direccion_sucursal", "direccion_trabajos", "direccion_cliente"])
                    col_lat = self._pick_col(cols, ["latitud", "lat", "latitude"])
                    col_lng = self._pick_col(cols, ["longitud", "lng", "lon", "longitude"])
                    if not col_dir or not col_lat or not col_lng:
                        continue

                    sql_all = text(
                        f"""
                        SELECT
                            COALESCE(CAST("{col_dir}" AS NVARCHAR(MAX)), '') AS direccion,
                            COALESCE(CAST("{col_lat}" AS NVARCHAR(MAX)), '') AS latitud,
                            COALESCE(CAST("{col_lng}" AS NVARCHAR(MAX)), '') AS longitud
                        FROM "{schema_name}"."{table_name}"
                        WHERE TRIM(CAST("{col_dir}" AS NVARCHAR(MAX))) <> ''
                          AND TRIM(CAST("{col_lat}" AS NVARCHAR(MAX))) <> ''
                          AND TRIM(CAST("{col_lng}" AS NVARCHAR(MAX))) <> ''
                        """
                    )
                    for row in self.db.execute(sql_all).mappings().all():
                        dir_cand = str(row.get("direccion") or "").strip()
                        lat = str(row.get("latitud") or "").strip()
                        lng = str(row.get("longitud") or "").strip()
                        if not dir_cand or not lat or not lng:
                            continue

                        dir_norm_cand = self._normalizar_direccion(dir_cand)
                        if not dir_norm_cand:
                            continue
                        tokens_cand = set(dir_norm_cand.split())
                        inter = len(tokens_ref & tokens_cand)
                        if inter <= 0:
                            continue

                        score = inter
                        if direccion_norm in dir_norm_cand or dir_norm_cand in direccion_norm:
                            score += 4
                        nro_cand = self._extraer_numero_direccion(dir_cand)
                        if nro_ref and nro_cand and nro_ref == nro_cand:
                            score += 8
                        elif nro_ref and nro_cand and nro_ref != nro_cand:
                            score -= 6

                        if score > mejor_score:
                            mejor_score = score
                            mejor_coords = (lat, lng)
            except Exception:
                continue

        # Exigir coincidencia minima razonable.
        if mejor_score >= 3 and mejor_coords[0] and mejor_coords[1]:
            return mejor_coords

        return "", ""

    def _geocodificar_direccion(self, direccion: str) -> tuple[str, str]:
        def _parse_coords(lat_raw: Any, lng_raw: Any) -> tuple[str, str]:
            try:
                lat_txt = str(lat_raw or "").strip().replace(",", ".")
                lng_txt = str(lng_raw or "").strip().replace(",", ".")
                if not lat_txt or not lng_txt:
                    return "", ""
                lat_f = float(lat_txt)
                lng_f = float(lng_txt)
                if not (-90 <= lat_f <= 90 and -180 <= lng_f <= 180):
                    return "", ""
                return f"{lat_f:.6f}", f"{lng_f:.6f}"
            except Exception:
                return "", ""

        direccion_txt = str(direccion or "").strip()
        cache_key = self._normalizar_direccion(direccion_txt)
        if not cache_key:
            return "", ""

        cached = _GEOCODE_CACHE.get(cache_key)
        if cached and cached[0] and cached[1]:
            return cached

        queries = [direccion_txt]
        dir_norm = self._normalizar_texto(direccion_txt)
        if "chile" not in dir_norm:
            queries.append(f"{direccion_txt}, Chile")

        for q in queries:
            # 1) Nominatim (OpenStreetMap)
            try:
                url = (
                    "https://nominatim.openstreetmap.org/search"
                    f"?q={quote_plus(q)}&format=jsonv2&limit=1&countrycodes=cl"
                )
                req = urllib.request.Request(
                    url,
                    headers={
                        "User-Agent": "ATC-Incidencias/1.0",
                        "Accept": "application/json",
                    },
                )
                with urllib.request.urlopen(req, timeout=4.0) as response:
                    payload = json.loads(response.read().decode("utf-8", errors="replace"))
                if isinstance(payload, list) and payload:
                    first = payload[0] or {}
                    lat, lng = _parse_coords(first.get("lat"), first.get("lon"))
                    if lat and lng:
                        _GEOCODE_CACHE[cache_key] = (lat, lng)
                        return lat, lng
            except Exception:
                pass

            # 2) maps.co
            try:
                url = f"https://geocode.maps.co/search?q={quote_plus(q)}"
                req = urllib.request.Request(
                    url,
                    headers={
                        "User-Agent": "ATC-Incidencias/1.0",
                        "Accept": "application/json",
                    },
                )
                with urllib.request.urlopen(req, timeout=4.0) as response:
                    payload = json.loads(response.read().decode("utf-8", errors="replace"))
                if isinstance(payload, list) and payload:
                    first = payload[0] or {}
                    lat, lng = _parse_coords(first.get("lat"), first.get("lon"))
                    if lat and lng:
                        _GEOCODE_CACHE[cache_key] = (lat, lng)
                        return lat, lng
            except Exception:
                pass

            # 3) Photon (komoot)
            try:
                url = f"https://photon.komoot.io/api/?q={quote_plus(q)}&limit=1"
                req = urllib.request.Request(
                    url,
                    headers={
                        "User-Agent": "ATC-Incidencias/1.0",
                        "Accept": "application/json",
                    },
                )
                with urllib.request.urlopen(req, timeout=4.0) as response:
                    payload = json.loads(response.read().decode("utf-8", errors="replace"))
                features = payload.get("features") if isinstance(payload, dict) else []
                if isinstance(features, list) and features:
                    geom = (features[0] or {}).get("geometry") or {}
                    coords = geom.get("coordinates") if isinstance(geom, dict) else None
                    if isinstance(coords, (list, tuple)) and len(coords) >= 2:
                        lng_raw, lat_raw = coords[0], coords[1]
                        lat, lng = _parse_coords(lat_raw, lng_raw)
                        if lat and lng:
                            _GEOCODE_CACHE[cache_key] = (lat, lng)
                            return lat, lng
            except Exception:
                pass

        return "", ""

    def _coordenadas_aproximadas_por_direccion(self, direccion: str) -> tuple[str, str]:
        dir_norm = self._normalizar_texto(direccion or "")
        if not dir_norm:
            return "", ""

        # Si la direccion trae numero, no usar centro de comuna para evitar
        # coordenadas incorrectas (prefiere exactitud antes que aproximacion).
        if self._extraer_numero_direccion(direccion):
            return "", ""

        for keywords, coords in _COORD_FALLBACK_CL:
            if any(k in dir_norm for k in keywords):
                return coords

        # Fallback nacional: centro de Santiago.
        return "-33.4489", "-70.6693"

    def _persistir_coordenadas_sucursal(self, cliente: str, direccion: str, latitud: str, longitud: str) -> None:
        cli = str(cliente or "").strip()
        dir_txt = str(direccion or "").strip()
        lat = str(latitud or "").strip()
        lng = str(longitud or "").strip()
        if not lat or not lng:
            return

        hubo_cambios = False
        for table_name in ["bbdd_clientes", "catalogo_clientes"]:
            try:
                for schema_name in self._schemas_con_tabla(table_name):
                    cols = self._columnas_tabla(schema_name, table_name)
                    if not cols:
                        continue
                    col_lat = self._pick_col(cols, ["latitud", "lat", "latitude"])
                    col_lng = self._pick_col(cols, ["longitud", "lng", "lon", "longitude"])
                    if not col_lat or not col_lng:
                        continue
                    col_cliente = self._pick_col(cols, ["cliente", "nombre_sucursal", "sucursal", "nombre_cliente"])
                    col_dir = self._pick_col(cols, ["direccion", "direccion_sucursal", "direccion_trabajos", "direccion_cliente"])

                    where_parts: list[str] = []
                    params: dict[str, Any] = {"lat": lat, "lng": lng}
                    if col_cliente and cli:
                        where_parts.append(f'TRIM(CAST("{col_cliente}" AS NVARCHAR(MAX))) = :cliente')
                        params["cliente"] = cli
                    if col_dir and dir_txt:
                        where_parts.append(f'TRIM(CAST("{col_dir}" AS NVARCHAR(MAX))) = :direccion')
                        params["direccion"] = dir_txt
                    if not where_parts:
                        continue

                    sql = text(
                        f"""
                        UPDATE "{schema_name}"."{table_name}"
                        SET "{col_lat}" = :lat,
                            "{col_lng}" = :lng
                        WHERE ({' OR '.join(where_parts)})
                          AND (
                            TRIM(CAST("{col_lat}" AS NVARCHAR(MAX))) = ''
                            OR TRIM(CAST("{col_lng}" AS NVARCHAR(MAX))) = ''
                          )
                        """
                    )
                    result = self.db.execute(sql, params)
                    if (result.rowcount or 0) > 0:
                        hubo_cambios = True
            except Exception:
                continue

        if hubo_cambios:
            try:
                self.db.commit()
            except Exception:
                self.db.rollback()

    def _support_sync_enabled(self) -> bool:
        # AJUSTE SOPORTE REGISTRO SQL #
        mode = (settings.support_sync_mode or "off").strip().lower()
        return mode in {"db", "api"}

    def _build_support_payload(self, odt: str, data: IncidenciaNueva, fecha: datetime) -> dict[str, Any]:
        fecha_txt = fecha.strftime("%d/%m/%Y %H:%M")
        tecnico = (data.tecnico or "").strip()
        acompanante = (data.acompanante or "").strip()
        return {
            "odt": odt,
            "fecha": fecha_txt,
            "fecha_registro": fecha.isoformat(),
            "puesto": (data.puesto or "").strip(),
            "sucursal": (data.cliente or "").strip(),
            "cliente": (data.cliente or "").strip(),
            "problema": (data.tipo_incidencia or "").strip(),
            "tipo_incidencia": (data.tipo_incidencia or "").strip(),
            "derivacion": (data.derivacion or "").strip() or "Pendiente",
            "observacion": (data.descripcion or "").strip(),
            "descripcion": (data.descripcion or "").strip(),
            "estado": (data.estado or "").strip() or "Pendiente",
            "tecnico": tecnico,
            "tecnicos": tecnico,
            "acompanante": acompanante,
            "origen": "incidencias_app",
            "source_file": "incidencias_sync",
        }

    def _sync_to_support_api(self, payload: dict[str, Any]) -> None:
        raw_url = (settings.support_sync_api_url or "").strip()
        if not raw_url:
            raise RuntimeError("SUPPORT_SYNC_API_URL no configurado.")

        def _rstrip_slash(v: str) -> str:
            return v[:-1] if v.endswith("/") else v

        def _build_payload_nueva(src: dict[str, Any]) -> dict[str, Any]:
            return {
                "puesto": (src.get("puesto") or "").strip(),
                "cliente": (src.get("cliente") or src.get("sucursal") or "").strip(),
                "tipoIncidencia": (src.get("tipo_incidencia") or src.get("problema") or "").strip(),
                "descripcion": (src.get("descripcion") or src.get("observacion") or "").strip(),
                "estado": (src.get("estado") or "Pendiente").strip(),
                "derivacion": (src.get("derivacion") or "").strip() or None,
                "tecnico": (src.get("tecnico") or src.get("tecnicos") or "").strip(),
                "acompanante": (src.get("acompanante") or "").strip(),
            }

        def _expand_api_candidates(url_value: str) -> list[tuple[str, Any]]:
            src = _rstrip_slash(url_value)
            parsed = urlsplit(src)
            origin = src
            if parsed.scheme and parsed.netloc:
                origin = f"{parsed.scheme}://{parsed.netloc}"

            payload_nueva = _build_payload_nueva(payload)
            candidates: list[tuple[str, Any]] = [
                (src, payload),
                (_rstrip_slash(origin) + "/api/incidencias/sync", payload),
                (_rstrip_slash(origin) + "/api/incidencias/nueva", payload_nueva),
                (_rstrip_slash(origin) + "/api/incidencias", payload_nueva),
                (_rstrip_slash(origin) + "/api/incidencias/multiples", [payload_nueva]),
            ]

            # Deduplicar por URL + body serializado.
            seen: set[tuple[str, str]] = set()
            deduped: list[tuple[str, Any]] = []
            for u, b in candidates:
                key = (_rstrip_slash(u), json.dumps(b, ensure_ascii=False, sort_keys=True))
                if key in seen:
                    continue
                seen.add(key)
                deduped.append((u, b))
            return deduped

        token = (settings.support_sync_api_token or "").strip()
        headers = {"Content-Type": "application/json; charset=utf-8"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        timeout = max(1, int(settings.support_sync_timeout_sec or 10))

        errores: list[str] = []
        for candidate_url, candidate_body in _expand_api_candidates(raw_url):
            req_body = json.dumps(candidate_body, ensure_ascii=False).encode("utf-8")
            req = urllib.request.Request(url=candidate_url, data=req_body, headers=headers, method="POST")
            try:
                with urllib.request.urlopen(req, timeout=timeout) as resp:  # nosec - endpoint configurable del usuario
                    status_code = getattr(resp, "status", 200)
                    if status_code >= 400:
                        raise RuntimeError(f"Sync API status {status_code}")
                    return
            except urllib.error.HTTPError as exc:
                detail = exc.read().decode("utf-8", errors="ignore")
                errores.append(f"{candidate_url} -> HTTP {exc.code}: {detail}")
                # Reintentamos solo para errores de ruta/mÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â©todo/payload no compatible.
                if exc.code in {404, 405, 422}:
                    continue
                raise RuntimeError(f"Sync API HTTPError {exc.code}: {detail}") from exc
            except urllib.error.URLError as exc:
                errores.append(f"{candidate_url} -> URLError: {exc}")
                continue

        raise RuntimeError("Sync API fallÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³ en todos los endpoints candidatos. " + " | ".join(errores))

    def _sync_to_support_db(self, payload: dict[str, Any]) -> None:
        db_url = (settings.support_db_url or "").strip()
        if not db_url:
            raise RuntimeError("SUPPORT_DB_URL no configurado.")

        schema = (settings.support_db_schema or "public").strip()
        # AJUSTE SOPORTE REGISTRO SQL #
        table = (settings.support_db_table or "registro").strip()
        engine = build_engine(db_url, pool_pre_ping=True)

        with engine.begin() as conn:
            cols_rows = conn.execute(
                text(
                    """
                    SELECT
                        column_name,
                        is_nullable,
                        column_default,
                        data_type,
                        udt_name,
                        is_identity,
                        is_generated
                    FROM information_schema.columns
                    WHERE table_schema = :schema_name
                      AND table_name = :table_name
                    """
                ),
                {"schema_name": schema, "table_name": table},
            ).all()

            columns_info: dict[str, dict[str, Any]] = {}
            for r in cols_rows:
                if not r or not r[0]:
                    continue
                col_name = str(r[0]).strip()
                columns_info[col_name] = {
                    "is_nullable": str(r[1] or "").strip().upper() == "YES",
                    "column_default": r[2],
                    "data_type": str(r[3] or "").strip().lower(),
                    "udt_name": str(r[4] or "").strip().lower(),
                    "is_identity": str(r[5] or "").strip().upper() == "YES",
                    "is_generated": str(r[6] or "").strip().upper(),
                }

            cols = set(columns_info.keys())
            if not cols:
                raise RuntimeError(f"No se encontrÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³ {schema}.{table} en SQL de soporte.")

            campo_map = {
                "odt": ["odt", "codigo", "codigo_odt", "nro_odt"],
                "fecha": ["fecha", "fecha_registro", "created_at"],
                "puesto": ["puesto", "nro_puesto", "puesto_numero"],
                "sucursal": ["sucursal", "cliente", "nombre_sucursal", "nombre_cliente"],
                "problema": ["problema", "tipo_incidencia", "tipo"],
                "derivacion": ["derivacion", "servicio", "area"],
                # AJUSTE SOPORTE REGISTRO SQL #
                "observacion": ["observacion", "detalle_problema", "descripcion", "detalle"],
                "estado": ["estado", "status", "situacion"],
            }

            insert_cols: list[str] = []
            params: dict[str, Any] = {}
            for key, opciones in campo_map.items():
                col = next((c for c in opciones if c in cols), None)
                if not col:
                    continue
                insert_cols.append(col)
                params[col] = payload.get(key) or ""

            # Campos ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âºtiles por nombre directo, si existen.
            for direct_col in ["source_file", "origen", "origin", "source"]:
                if direct_col in cols and direct_col not in insert_cols:
                    insert_cols.append(direct_col)
                    params[direct_col] = (
                        payload.get(direct_col)
                        or payload.get("origen")
                        or payload.get("source_file")
                        or "incidencias_sync"
                    )

            # AJUSTE SOPORTE REGISTRO SQL #
            for direct_col in ["tecnicos", "acompanante", "prioridad", "direccion", "observacion_final"]:
                if direct_col in cols and direct_col not in insert_cols:
                    insert_cols.append(direct_col)
                    params[direct_col] = payload.get(direct_col) or ""

            def _coerce_required_value(col_name: str, info: dict[str, Any]) -> Any:
                cname = col_name.lower()
                data_type = info.get("data_type", "")
                udt_name = info.get("udt_name", "")
                now = datetime.utcnow()
                odt_tail = _parse_prefijo_numero(str(payload.get("odt") or "").strip()) or int(now.timestamp())

                if cname in {"source_file", "source", "origin", "origen"}:
                    return payload.get("source_file") or payload.get("origen") or "incidencias_sync"
                if cname in {"source_row", "source_id", "origen_id"}:
                    raw = payload.get("source_row")
                    if raw is not None and str(raw).strip() != "":
                        try:
                            return int(str(raw).strip())
                        except Exception:
                            pass
                    return odt_tail
                if cname in {"estado", "status", "situacion"}:
                    return payload.get("estado") or "Pendiente"
                if cname in {"created_at", "updated_at", "fecha_creacion", "fecha_actualizacion"}:
                    return now
                if cname in {"uuid", "guid"} or "uuid" in udt_name:
                    return str(uuid.uuid4())
                if "bool" in data_type or udt_name == "bool":
                    return False
                if any(x in data_type for x in ["int", "numeric", "decimal", "real", "double"]) or udt_name in {
                    "int2",
                    "int4",
                    "int8",
                    "numeric",
                    "float4",
                    "float8",
                }:
                    return 0
                if "timestamp" in data_type or data_type in {"date", "time"} or udt_name in {
                    "timestamp",
                    "timestamptz",
                    "date",
                    "time",
                }:
                    return now
                if "json" in data_type or udt_name in {"json", "jsonb"}:
                    return "{}"
                return ""

            # Completar columnas NOT NULL sin default para evitar fallos por esquema destino.
            for col_name, info in columns_info.items():
                if col_name in insert_cols:
                    continue
                if info.get("is_nullable", True):
                    continue
                if info.get("column_default") is not None:
                    continue
                if info.get("is_identity", False):
                    continue
                if info.get("is_generated", "") in {"ALWAYS", "BY DEFAULT"}:
                    continue
                insert_cols.append(col_name)
                params[col_name] = _coerce_required_value(col_name, info)

            if not insert_cols:
                raise RuntimeError(f"No hay columnas compatibles para insertar en {schema}.{table}.")

            placeholders = [f":{c}" for c in insert_cols]
            quoted_cols = ", ".join(f'"{c}"' for c in insert_cols)
            sql_insert = text(
                f'INSERT INTO "{schema}"."{table}" ({quoted_cols}) '
                f'VALUES ({", ".join(placeholders)})'
            )
            conn.execute(sql_insert, params)

    def _registrar_sync_soporte_nueva(
        self,
        odt: str,
        data: IncidenciaNueva,
        fecha: datetime,
        descripcion_registro: str | None = None,
    ) -> None:
        # AJUSTE SOPORTE REGISTRO SQL #
        if not self._support_sync_enabled():
            return
        try:
            payload = self._build_support_payload(odt, data, fecha)
            if descripcion_registro is not None:
                payload["observacion"] = descripcion_registro
                payload["descripcion"] = descripcion_registro
            mode = (settings.support_sync_mode or "off").lower()
            if mode == "api":
                self._sync_to_support_api(payload)
            elif mode == "db":
                self._sync_to_support_db(payload)
        except Exception:
            self.db.rollback()

    def enviar_formulario(self, datos: FormularioRegistro) -> str:
        odt = datos.odt or self._proximo_odt("I")
        registro = Registro(
            odt=odt,
            fecha_registro=datos.fecha or datetime.now(),
            cliente=datos.cliente,
            problema=datos.problema,
            detalle_problema=datos.detalle_problema,
            derivacion=datos.derivacion,
            observacion=datos.observacion,
            tecnicos=datos.tecnicos,
            acompanante=datos.acompanante,
            estado=datos.estado,
            dias_ejecucion=datos.dias_ejecucion,
            foto_1=datos.foto,
            observacion_final=datos.observacion_final,
            fecha_cierre=datos.fecha_cierre,
            direccion=self._direccion_cliente(datos.cliente),
        )
        self.db.add(registro)
        try:
            self.db.commit()
        except SQLAlchemyError as exc:
            self.db.rollback()
            raise _build_db_write_error(exc) from exc
        return "Registro guardado en SQL"

    def _firmar_observacion_registro(
        self,
        token: str | None,
        observacion: str,
        fecha: datetime,
        usuario_fallback: str | None = None,
        area_label: str | None = None,
    ) -> str:
        texto = (observacion or "").strip()
        if not texto:
            return ""
        token_limpio = (token or "").strip()
        # Antes, si no llegaba token (ej. incidencias_soporte.html manda
        # token="" siempre), se caia directo a "Usuario no identificado" sin
        # intentar usuario_fallback (el usuario real logueado, que si esta
        # disponible) — quedaba "Usuario no identificado" aunque supieramos
        # perfectamente quien lo creo. Ahora el fallback SIEMPRE se intenta
        # cuando no hay usuario resuelto por token — pedido explicito, ago 2026.
        usuario = self.get_usuario_actual(token_limpio) if token_limpio else ""
        if not usuario or usuario == "Desconocido":
            usuario = (usuario_fallback or "").strip() or "Usuario no identificado"
        tz_name = (settings.timezone or "America/Santiago").strip() or "America/Santiago"
        fecha_local = fecha.replace(tzinfo=timezone.utc).astimezone(ZoneInfo(tz_name))
        marca = fecha_local.strftime("%d/%m/%Y %H:%M")
        area_txt = (area_label or "").strip()
        firma = f"{usuario} - {area_txt} - {marca}" if area_txt else f"{usuario} - {marca}"
        return f"[{firma}] {texto}"

    def _extraer_equipo_desde_observacion_servicio(self, texto: str) -> tuple[str, str]:
        raw = str(texto or "")
        tecnico = ""
        acompanante = ""
        match_tecnico = re.search(
            r"T[eé]cnico\s+sugerido:\s*(.*?)(?=\s*\|\s*Acompa(?:ñ|n)ante:|$)",
            raw,
            flags=re.IGNORECASE,
        )
        if match_tecnico:
            tecnico = match_tecnico.group(1).strip()
        match_acompanante = re.search(
            r"Acompa(?:ñ|n)ante:\s*(.*?)(?=\s*\||$)",
            raw,
            flags=re.IGNORECASE,
        )
        if match_acompanante:
            acompanante = match_acompanante.group(1).strip()
        return tecnico, acompanante

    def _resolver_puesto_por_sucursal(self, nombre_sucursal: str) -> str | None:
        """Dado un nombre de sucursal, devuelve el puesto (columna `central`
        de sucursal_camaras_monitoreo) mas frecuente para esa sucursal — el
        usuario ya no elige el puesto a mano en el formulario de creacion de
        ODT, se enlaza solo a partir de la sucursal elegida (para analisis
        futuros) — pedido explicito, ago 2026. Si la sucursal no tiene
        camaras registradas (ej. clientes solo-alarma) devuelve None, que es
        un valor valido para Registro.puesto (ya nullable)."""
        nombre = (nombre_sucursal or "").strip()
        if not nombre:
            return None
        try:
            rows = (
                self.db.query(SucursalCamaraMonitoreo.central, func.count().label("n"))
                .join(SucursalBBDD, SucursalCamaraMonitoreo.sucursal_id == SucursalBBDD.id)
                .filter(SucursalCamaraMonitoreo.central.isnot(None))
                .filter(func.trim(SucursalBBDD.nombre_sucursal) == nombre)
                .group_by(SucursalCamaraMonitoreo.central)
                .order_by(func.count().desc())
                .all()
            )
        except SQLAlchemyError:
            self.db.rollback()
            return None
        if not rows:
            return None
        return str(int(rows[0][0]))

    def guardar_incidencia_nueva(self, data: IncidenciaNueva, usuario_fallback: str | None = None) -> str:
        odt = self._proximo_odt("I")
        ahora = datetime.now()
        cliente = (data.cliente or "").strip()
        puesto_valor = (data.puesto or "").strip()
        if not puesto_valor:
            puesto_valor = self._resolver_puesto_por_sucursal(cliente) or ""
            data.puesto = puesto_valor or None
        descripcion = (data.descripcion or "").strip()
        area_normalizada = (data.area or "").strip().lower()
        es_servicio_tecnico = area_normalizada == "servicio_tecnico"
        es_soporte = area_normalizada == "soporte"
        # ODT creado desde "Crear ODT" en incidencias_soporte.html: va directo a
        # "Gestión Soporte" (observacion_soporte), texto plano sin firma — no debe
        # quedar en "Registro Operaciones" (observacion) — pedido explicito, ago 2026.
        observacion_soporte_creacion = descripcion or None if es_soporte else None
        observacion_registro = (
            None
            if es_soporte
            else self._firmar_observacion_registro(
                data.token,
                descripcion,
                ahora,
                usuario_fallback=usuario_fallback,
                area_label=("Servicio Técnico" if es_servicio_tecnico else ""),
            )
        )

        derivacion = (data.derivacion or "").strip() or "Pendiente"
        estado = (data.estado or "").strip() or "Pendiente"
        direccion = (data.direccion or "").strip() or self._direccion_cliente(cliente)
        tecnico = (data.tecnico or "").strip()
        acompanante = (data.acompanante or "").strip()
        if es_servicio_tecnico and (not tecnico or not acompanante):
            tecnico_obs, acompanante_obs = self._extraer_equipo_desde_observacion_servicio(descripcion)
            tecnico = tecnico or tecnico_obs
            acompanante = acompanante or acompanante_obs
        if es_servicio_tecnico and (tecnico or acompanante):
            estado = "En Proceso"
        reg = Registro(
            odt=odt,
            fecha_registro=ahora,
            puesto=(puesto_valor or None),
            cliente=cliente,
            problema=(data.tipo_incidencia or "").strip(),
            detalle_problema=(observacion_registro or observacion_soporte_creacion or None),
            derivacion=derivacion,
            observacion=(None if (es_servicio_tecnico or es_soporte) else (observacion_registro or None)),
            observacion_servicio=(observacion_registro if es_servicio_tecnico else None),
            observacion_soporte=observacion_soporte_creacion,
            estado=estado,
            fecha_derivacion_area=ahora,
            direccion=direccion,
            tecnicos=tecnico,
            acompanante=acompanante,
            fecha_derivacion_tecnico=(ahora if tecnico else None),
        )
        self.db.add(reg)
        try:
            self.db.commit()
        except SQLAlchemyError as exc:
            self.db.rollback()
            raise _build_db_write_error(exc) from exc
        # AJUSTE SOPORTE REGISTRO SQL #
        self._registrar_sync_soporte_nueva(odt, data, ahora, observacion_registro)
        return odt

    def _nombre_corto_tecnico(self, value: str) -> str:
        """Nombre corto (primer nombre + apellido representativo) para
        matchear tecnicos contra EQUIPOS_TECNICOS_POR_PATENTE sin importar
        si el dato original viene con nombre completo o ya abreviado.
        Misma logica que usaba el closure _nombre_corto de
        obtener_resumen_equipos_tecnicos_hoy, ahora compartida."""
        texto = re.sub(r"\s+", " ", self._reparar_texto_mojibake(value)).strip()
        partes = [p for p in texto.split() if p and self._normalizar_texto(p) not in {"atc"}]
        if not partes:
            return ""
        if len(partes) == 1:
            return partes[0]
        if len(partes) >= 4:
            return f"{partes[0]} {partes[-2]}"
        nombres_intermedios = {
            "alberto", "alejandro", "alfonso", "andres", "antonio", "benjamin",
            "constanza", "enrique", "estefano", "ignacio", "issak", "kevin",
            "octavio", "samir", "sebastian",
        }
        apellido = partes[-1] if len(partes) == 3 and self._normalizar_texto(partes[1]) in nombres_intermedios else partes[1]
        return f"{partes[0]} {apellido}"

    def _patente_para_tecnico(self, value: str) -> str:
        """Resuelve nombre de tecnico -> patente de su equipo fijo, con la
        misma logica (roster + alias manuales) que usa el tablero de
        obtener_resumen_equipos_tecnicos_hoy."""
        patentes_por_tecnico: dict[str, str] = {}
        for patente, miembros in self.EQUIPOS_TECNICOS_POR_PATENTE.items():
            for nombre in miembros:
                for alias in (nombre, self._nombre_corto_tecnico(nombre)):
                    key = self._normalizar_nombre_login(alias)
                    if key:
                        patentes_por_tecnico[key] = patente
        patentes_por_tecnico.update(self.ALIASES_PATENTE_TECNICO)

        nombre = re.sub(r"\s+", " ", self._reparar_texto_mojibake(value)).strip()
        candidatos = [
            self._normalizar_nombre_login(nombre),
            self._normalizar_nombre_login(nombre.strip("*")),
            self._normalizar_nombre_login(self._nombre_corto_tecnico(nombre)),
        ]
        return next((patentes_por_tecnico.get(c) for c in candidatos if patentes_por_tecnico.get(c)), "")

    def sugerir_acompanante(self, tecnico: str) -> str:
        """Wrapper publico de _sugerir_acompanante_para_tecnico, para el
        endpoint de sugerencia (GET, de solo lectura) que el frontend
        consulta antes de pedirle confirmacion al usuario."""
        return self._sugerir_acompanante_para_tecnico(tecnico)

    def _sugerir_acompanante_para_tecnico(self, tecnico: str) -> str:
        """Acompañante del equipo fijo (camioneta) de este tecnico, segun
        EQUIPOS_TECNICOS_POR_PATENTE — la misma fuente que usa el tablero de
        obtener_resumen_equipos_tecnicos_hoy, para que la sugerencia al
        derivar una ODT siempre coincida con el equipo que se ve ahi. Si el
        tecnico trabaja solo (su equipo tiene un unico integrante) o no esta
        en ningun equipo, no hay sugerencia — el frontend no pregunta nada
        en ese caso, se queda sin acompanante."""
        tecnico_txt = (tecnico or "").strip()
        if not tecnico_txt:
            return ""
        patente = self._patente_para_tecnico(tecnico_txt)
        if not patente:
            return ""
        miembros = self.EQUIPOS_TECNICOS_POR_PATENTE.get(patente) or []
        if len(miembros) < 2:
            return ""
        candidatos_propios = {
            self._normalizar_nombre_login(tecnico_txt),
            self._normalizar_nombre_login(tecnico_txt.strip("*")),
            self._normalizar_nombre_login(self._nombre_corto_tecnico(tecnico_txt)),
        }
        for miembro in miembros:
            if self._normalizar_nombre_login(miembro) not in candidatos_propios:
                return self._nombre_corto_tecnico(miembro)
        return ""

    def derivar_odt_a_tecnico(
        self,
        odt: str,
        tecnico: str = "",
        acompanante: str = "",
        derivacion: str = "Servicio Técnico",
        estado: str = "Pendiente",
    ) -> dict[str, Any] | None:
        odt_limpia = (odt or "").strip()
        tecnico_limpio = (tecnico or "").strip()
        acompanante_limpio = (acompanante or "").strip()
        derivacion_final = (derivacion or "").strip()

        if not odt_limpia:
            raise ValueError("ODT invalida.")

        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            return None

        ahora = datetime.now()
        if not derivacion_final:
            derivacion_final = str(row.derivacion or "").strip() or "Servicio Técnico"

        estado_solicitado = self._normalizar_texto(estado or "")
        dejar_pendiente = estado_solicitado == "pendiente"

        if tecnico_limpio:
            estado_final = "Pendiente" if dejar_pendiente else "En Proceso"
            row.tecnicos = tecnico_limpio
            # Ya no se auto-asigna un acompañante acá: el frontend consulta
            # la sugerencia por separado (GET /api/incidencias/sugerir-acompanante)
            # y confirma con el usuario antes de mandarla en este mismo campo.
            # Así "acompanante" siempre refleja exactamente lo que se pidió
            # guardar — incluido vacío, para poder borrarlo sin que se vuelva
            # a rellenar solo.
            row.acompanante = acompanante_limpio or None
            if not dejar_pendiente:
                row.fecha_derivacion_tecnico = ahora
        else:
            # Sin tecnico: siempre queda Pendiente y se limpia la asignacion,
            # tanto si vino de "poner sin asignar" (dejar_pendiente) como de
            # una desasignacion completa.
            estado_final = "Pendiente"
            row.tecnicos = ""
            row.acompanante = None
            row.fecha_derivacion_tecnico = None
            if not dejar_pendiente:
                # Desasignacion completa: se saca de la cola de Servicio Tecnico.
                derivacion_final = "Pendiente"

        row.derivacion = derivacion_final
        row.estado = estado_final
        if not row.fecha_derivacion_area:
            row.fecha_derivacion_area = ahora
        if estado_final == "Pendiente" and row.fecha_inicio_trabajo and not row.fecha_fin_trabajo:
            # Sacar al tecnico de la ODT (o dejarla pendiente) mientras tenia
            # el cronometro corriendo equivale a que el mismo marcara "Pendiente".
            row.fecha_fin_trabajo = ahora
        self.db.commit()
        if estado_final == "En Proceso":
            self._sync_estado_ticket_soporte_silencioso(odt_limpia, TICKET_STATUS_PENDIENTE_SERVICIO)
        else:
            self._sync_estado_ticket_soporte_silencioso(odt_limpia, TICKET_STATUS_PENDIENTE)
        return {
            "tecnico": row.tecnicos or "",
            "acompanante": row.acompanante or "",
            "estado": row.estado or "",
            "derivacion": row.derivacion or "",
        }

    def editar_incidencia_tabla(
        self,
        token: str,
        odt: str,
        derivacion: str | None = None,
        observacion: str | None = None,
        prioridad: str | None = None,
        observacion_servicio: str | None = None,
        observacion_final: str | None = None,
        repetida_odt_ref: str | None = None,
        editar_ultima_observacion_servicio: bool = False,
        observacion_coordinacion: str | None = None,
        eliminar_ultima_observacion_coordinacion: bool = False,
    ) -> dict[str, Any]:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida.")
        if not self.usuario_logueado_por_token(token):
            raise ValueError("Sesion expirada. Inicia sesion nuevamente.")

        usuario = self.get_usuario_actual(token)
        if not usuario or usuario == "Desconocido":
            raise ValueError("No se pudo identificar al usuario de la sesion.")

        derivacion_in = (derivacion or "").strip()
        observacion_in = (observacion or "").strip()
        prioridad_in = None if prioridad is None else str(prioridad or "").strip()
        observacion_servicio_in = (observacion_servicio or "").strip()
        observacion_final_in = (observacion_final or "").strip()
        repetida_ref_in = (repetida_odt_ref or "").strip()
        observacion_coordinacion_in = (observacion_coordinacion or "").strip()
        if (
            not derivacion_in
            and not observacion_in
            and prioridad_in is None
            and not observacion_servicio_in
            and not observacion_final_in
            and not observacion_coordinacion_in
            and not eliminar_ultima_observacion_coordinacion
        ):
            raise ValueError("Debes enviar derivacion, observacion o prioridad para editar.")

        opciones_derivacion = [
            "Técnico Externo",
            "Técnico Externo",
            "Cliente",
            "Soporte Técnico",
            "Soporte Técnico",
            "Servicio Técnico",
            "Coordinacion",
            "Coordinación",
            "Finalizado por Soporte",
            "Finalizado Sin VT",
            "Repetida",
        ]
        if derivacion_in:
            mapa_deriv = {self._normalizar_texto(v): v for v in opciones_derivacion}
            key = self._normalizar_texto(derivacion_in)
            if key not in mapa_deriv:
                raise ValueError("Derivacion no permitida.")
            derivacion_in = mapa_deriv[key]

        tz_name = (settings.timezone or "America/Santiago").strip() or "America/Santiago"
        ahora_local = datetime.now(ZoneInfo(tz_name))
        ahora_utc = datetime.now()
        marca = ahora_local.strftime("%d/%m/%Y %H:%M")

        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            return {"ok": False}

        if prioridad_in is not None:
            estado_norm = self._normalizar_texto(getattr(row, "estado", "") or "")
            if "proceso" not in estado_norm:
                raise ValueError("Solo puedes editar la prioridad de una ODT En Proceso.")
            if prioridad_in == "":
                row.prioridad = None
            else:
                try:
                    prioridad_num = int(prioridad_in)
                except Exception as exc:
                    raise ValueError("Prioridad invalida. Debe ser un numero del 1 al 10.") from exc
                if prioridad_num < 1 or prioridad_num > 10:
                    raise ValueError("Prioridad invalida. Debe ser un numero del 1 al 10.")
                row.prioridad = prioridad_num

        observacion_soporte_final = ""
        observacion_servicio_final = ""
        observacion_coordinacion_final = ""
        estado_ticket_soporte = ""
        nota_ticket_soporte = ""
        correo_derivacion_result: dict[str, Any] | None = None

        if derivacion_in:
            row.derivacion = derivacion_in
            if derivacion_in == "Finalizado por Soporte":
                if not observacion_final_in:
                    raise ValueError("Debes indicar que se hizo para finalizar por soporte.")
                row.estado = "Terminado"
                row.observacion_final = observacion_final_in
                row.fecha_cierre = ahora_utc
                if row.fecha_registro:
                    row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days
                estado_ticket_soporte = TICKET_STATUS_RESUELTO_SERVICIO
                nota_ticket_soporte = self._build_nota_cierre_ticket_soporte(
                    odt=odt_limpia,
                    estado_ticket=estado_ticket_soporte,
                    derivacion=derivacion_in,
                    observacion_final=observacion_final_in,
                )
            elif derivacion_in == "Finalizado Sin VT":
                row.estado = "Terminado"
                row.fecha_cierre = ahora_utc
                if row.fecha_registro:
                    row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days
                estado_ticket_soporte = TICKET_STATUS_RESUELTO_SERVICIO
                nota_ticket_soporte = self._build_nota_cierre_ticket_soporte(
                    odt=odt_limpia,
                    estado_ticket=estado_ticket_soporte,
                    derivacion=derivacion_in,
                    observacion_final=observacion_final_in,
                )
            elif derivacion_in == "Repetida":
                if not repetida_ref_in:
                    match_ref = re.search(r"\b([A-Za-z]\d+)\b", observacion_servicio_in or "")
                    repetida_ref_in = str(match_ref.group(1) if match_ref else "").strip()
                if not repetida_ref_in:
                    raise ValueError("Debes indicar la ODT con la que se repite.")
                if repetida_ref_in == odt_limpia:
                    raise ValueError("La ODT repetida no puede ser la misma ODT actual.")

                row_ref = self.db.scalar(select(Registro).where(Registro.odt == repetida_ref_in))
                if not row_ref:
                    raise ValueError(f"No se encontro la ODT de referencia {repetida_ref_in}.")

                estado_ref_norm = self._normalizar_texto(getattr(row_ref, "estado", "") or "")
                if ("pend" not in estado_ref_norm) and ("proceso" not in estado_ref_norm):
                    raise ValueError("La ODT de referencia debe estar Pendiente o En Proceso.")

                sucursal_actual = self._normalizar_texto(getattr(row, "cliente", "") or "")
                sucursal_ref = self._normalizar_texto(getattr(row_ref, "cliente", "") or "")
                if sucursal_actual != sucursal_ref:
                    raise ValueError("La ODT de referencia debe ser de la misma sucursal.")

                problema_actual = self._normalizar_texto(getattr(row, "problema", "") or "")
                problema_ref = self._normalizar_texto(getattr(row_ref, "problema", "") or "")
                if problema_actual != problema_ref:
                    raise ValueError("Solo puedes marcar como repetida con una ODT del mismo problema.")

                observacion_servicio_in = f"ODT con la que se repite {repetida_ref_in}"
                row.estado = "Repetida"
                row.fecha_cierre = ahora_utc
                if row.fecha_registro:
                    row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days
                estado_ticket_soporte = TICKET_STATUS_RESUELTO_SERVICIO
                nota_ticket_soporte = self._build_nota_cierre_ticket_soporte(
                    odt=odt_limpia,
                    estado_ticket=estado_ticket_soporte,
                    derivacion=derivacion_in,
                    observacion_final=observacion_servicio_in,
                )
            elif self._normalizar_texto(derivacion_in) == "tecnico externo":
                row.estado = "En Proceso"
                row.fecha_cierre = None
                estado_ticket_soporte = TICKET_STATUS_PENDIENTE_SERVICIO
            else:
                row.estado = "Pendiente"
                row.fecha_cierre = None
                deriv_norm = self._normalizar_texto(derivacion_in)
                if deriv_norm in {"cliente", "coordinacion"}:
                    estado_ticket_soporte = TICKET_STATUS_PENDIENTE_CLIENTE
                elif "servicio tecnico" in deriv_norm or "soporte tecnico" in deriv_norm:
                    estado_ticket_soporte = TICKET_STATUS_PENDIENTE_SERVICIO
                else:
                    estado_ticket_soporte = TICKET_STATUS_PENDIENTE

        if observacion_final_in and derivacion_in != "Finalizado por Soporte":
            row.observacion_final = observacion_final_in

        if observacion_in:
            base = str(getattr(row, "observacion_soporte", "") or "").strip()
            if not base:
                # Compatibilidad con datos guardados antes de separar columnas.
                base = str(getattr(row, "observacion_servicio", "") or "").strip()
            nuevo = observacion_in.strip()
            if base:
                if nuevo.startswith(base):
                    nuevo = nuevo[len(base):].strip()
                ultima_linea = base.splitlines()[-1].strip() if base.splitlines() else ""
                if nuevo == ultima_linea:
                    nuevo = ""
            if nuevo:
                linea = f"[{usuario} - {marca}] {nuevo}"
                row.observacion_soporte = f"{base}\n{linea}".strip() if base else linea
                observacion_soporte_final = row.observacion_soporte

        if editar_ultima_observacion_servicio:
            # Edicion de la ultima nota de observacion_servicio (ventana de
            # 15 min, solo el autor) — pedido explicito, jul 2026.
            base_servicio = str(getattr(row, "observacion_servicio", "") or "").strip()
            entry = _obs_last_entry(base_servicio)
            if not _obs_can_edit_entry(entry, usuario):
                raise ValueError("Ya no puedes editar esta observacion (limite de 15 minutos).")
            nuevo_texto = observacion_servicio_in.strip()
            if not nuevo_texto:
                raise ValueError("La observacion no puede quedar vacia.")
            row.observacion_servicio = _obs_edit_last_line(base_servicio, entry, usuario, nuevo_texto)
            observacion_servicio_final = row.observacion_servicio
        elif observacion_servicio_in:
            base_servicio = str(getattr(row, "observacion_servicio", "") or "").strip()
            nuevo_servicio = observacion_servicio_in.strip()
            if base_servicio:
                if nuevo_servicio.startswith(base_servicio):
                    nuevo_servicio = nuevo_servicio[len(base_servicio):].strip()
                ultima_linea_serv = base_servicio.splitlines()[-1].strip() if base_servicio.splitlines() else ""
                if nuevo_servicio == ultima_linea_serv:
                    nuevo_servicio = ""
            if nuevo_servicio:
                linea_servicio = f"[{usuario} - {marca}] {nuevo_servicio}"
                row.observacion_servicio = (
                    f"{base_servicio}\n{linea_servicio}".strip() if base_servicio else linea_servicio
                )
                observacion_servicio_final = row.observacion_servicio

        if eliminar_ultima_observacion_coordinacion:
            base_coord = str(getattr(row, "observacion_coordinacion", "") or "").strip()
            entry_coord = _obs_last_entry(base_coord)
            if not _obs_can_edit_entry(entry_coord, usuario):
                raise ValueError("Ya no puedes eliminar esta observacion (limite de 15 minutos).")
            row.observacion_coordinacion = _obs_delete_last_line(base_coord, entry_coord)
            observacion_coordinacion_final = row.observacion_coordinacion
        elif observacion_coordinacion_in:
            base_coord = str(getattr(row, "observacion_coordinacion", "") or "").strip()
            nuevo_coord = observacion_coordinacion_in.strip()
            if base_coord:
                if nuevo_coord.startswith(base_coord):
                    nuevo_coord = nuevo_coord[len(base_coord):].strip()
                ultima_linea_coord = base_coord.splitlines()[-1].strip() if base_coord.splitlines() else ""
                if nuevo_coord == ultima_linea_coord:
                    nuevo_coord = ""
            if nuevo_coord:
                linea_coord = f"[{usuario} - {marca}] {nuevo_coord}"
                row.observacion_coordinacion = (
                    f"{base_coord}\n{linea_coord}".strip() if base_coord else linea_coord
                )
                observacion_coordinacion_final = row.observacion_coordinacion

        self.db.commit()
        if estado_ticket_soporte:
            self._sync_estado_ticket_soporte_silencioso(
                odt_limpia,
                estado_ticket_soporte,
                nota_interna=nota_ticket_soporte or None,
            )
        if derivacion_in:
            deriv_norm_correo = self._normalizar_texto(derivacion_in)
            if deriv_norm_correo in {"cliente", "coordinacion"} or "servicio tecnico" in deriv_norm_correo:
                correo_derivacion_result = self._enviar_correo_derivacion_automatico_silencioso(
                    row=row,
                    derivacion=derivacion_in,
                    usuario=usuario,
                )
        return {
            "ok": True,
            "odt": odt_limpia,
            "derivacion": derivacion_in or None,
            "observacion": observacion_soporte_final or None,
            "observacion_soporte": observacion_soporte_final or None,
            "observacion_servicio": (
                observacion_servicio_final
                or str(getattr(row, "observacion_servicio", "") or "").strip()
                or None
            ),
            "observacion_final": str(getattr(row, "observacion_final", "") or "").strip() or None,
            "observacion_coordinacion": (
                observacion_coordinacion_final
                or str(getattr(row, "observacion_coordinacion", "") or "").strip()
                or None
            ),
            "prioridad": str(getattr(row, "prioridad", "") or "").strip() or None,
            "correo_derivacion": correo_derivacion_result,
        }

    def enviar_multiples_incidencias(
        self,
        incidencias: list[IncidenciaNueva],
        usuario_fallback: str | None = None,
    ) -> list[str]:
        odts_creadas: list[str] = []
        for inc in incidencias:
            if not inc.cliente or not inc.tipo_incidencia:
                continue
            odt = self.guardar_incidencia_nueva(inc, usuario_fallback=usuario_fallback)
            odts_creadas.append(odt)
        if not odts_creadas:
            raise ValueError("No se encontrÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³ ninguna incidencia vÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡lida para registrar.")
        return odts_creadas

    def obtener_registros(self, tecnico: str | None = None) -> list[list[Any]]:
        tecnico_filtrado = (tecnico or "").strip().lower()
        stmt = select(Registro).order_by(Registro.id.desc())
        registros = self._run_registro_query(
            lambda: self.db.scalars(stmt).all(),
            "cargar los registros de tecnicos",
        )
        salida: list[list[Any]] = []
        for r in registros:
            tecnico_match = not tecnico_filtrado or (
                tecnico_filtrado in (r.tecnicos or "").lower()
                or tecnico_filtrado in (r.acompanante or "").lower()
            )
            if not tecnico_match:
                continue
            salida.append(
                [
                    r.odt,
                    _to_ddmmyyyy_hhmm(r.fecha_derivacion_area),
                    r.cliente,
                    r.problema,
                    r.observacion,
                    r.estado,
                    r.prioridad,
                ]
            )
        return salida

    def obtener_registros_desde_administracion(self, tecnico: str | None = None) -> list[list[Any]]:
        filtro = (tecnico or "").strip().lower()
        rows = self._run_registro_query(
            lambda: self.db.scalars(select(Registro).order_by(Registro.id.desc())).all(),
            "cargar la tabla principal de incidencias",
        )
        resultado: list[list[Any]] = []
        for row in rows:
            tecnico_principal = (row.tecnicos or "").lower()
            tecnico_acom = (row.acompanante or "").lower()
            if filtro and filtro not in tecnico_principal and filtro not in tecnico_acom:
                continue

            fecha_ref = row.fecha_derivacion_area or row.fecha_registro
            detalle = (row.observacion_final or row.observacion_pendiente or row.observacion or "")
            resultado.append(
                [
                    row.odt,
                    _to_ddmmyyyy_hhmm(fecha_ref),
                    row.cliente,
                    row.problema,
                    detalle,
                    row.estado,
                ]
            )
        return resultado

    def obtener_datos_cliente(self, nombre_cliente: str) -> dict[str, str]:
        stmt = select(ClienteBBDD).where(ClienteBBDD.cliente == nombre_cliente)
        row = self.db.scalar(stmt)
        if not row:
            return {}
        return {
            "derivacion": "",
            "servicio": "",
            "soporte": "",
            "problema": "",
        }

    def obtener_datos_sucursal(self, cliente: str) -> dict[str, str]:
        stmt = select(ClienteBBDD).where(ClienteBBDD.cliente == cliente)
        row = self.db.scalar(stmt)
        if not row:
            return {}
        return {
            "direccion": row.direccion or "",
            "contacto": row.nombre_representante or "",
            "correo": row.email_representante or "",
        }

    def obtener_listas_bbdd(self) -> dict[str, list[str]]:
        rows = self.db.scalars(select(ClienteBBDD).order_by(ClienteBBDD.cliente.asc())).all()
        sucursales = sorted(
            {
                str(nombre).strip()
                for nombre in [*(r.cliente for r in rows if r.cliente), *SUCURSALES_EXTRA_MANTENCION]
                if str(nombre or "").strip()
            },
            key=self._normalizar_texto,
        )
        direccion = sorted({r.direccion for r in rows if r.direccion})
        contactos = sorted({r.nombre_representante for r in rows if r.nombre_representante})
        correos = sorted({r.email_representante for r in rows if r.email_representante})
        tecnicos_helpdesk = self._obtener_tecnicos_helpdesk(solo_activos=True)
        tecnicos_map: dict[str, str] = {}

        def add_tecnico(nombre: Any) -> None:
            nombre_limpio = str(nombre or "").strip()
            nombre_key = self._normalizar_texto(nombre_limpio)
            nombre_no_asignable = any(
                nombre_key == prefix or nombre_key.startswith(f"{prefix} ")
                for prefix in NON_SELECTABLE_SUPERVISOR_NAME_PREFIXES
            )
            if nombre_limpio and nombre_key and not nombre_no_asignable and nombre_key not in tecnicos_map:
                tecnicos_map[nombre_key] = nombre_limpio

        for nombre in tecnicos_helpdesk or []:
            add_tecnico(nombre)
        for nombre in self._usuarios_login_por_area("tecnicos"):
            add_tecnico(nombre)
        if not tecnicos_map:
            tecnicos_registro = self._run_registro_query(
                lambda: [
                    nombre
                    for row in self.db.scalars(select(Registro).order_by(Registro.id.desc())).all()
                    for nombre in [row.tecnicos, row.acompanante]
                ],
                "cargar la lista de tecnicos desde registros",
            )
            for nombre in tecnicos_registro or []:
                add_tecnico(nombre)
        tecnicos = sorted(tecnicos_map.values(), key=self._normalizar_texto)
        derivaciones: list[str] = []
        soportes: list[str] = []
        problemas = [
            "Desconexion",
            "Problema de Parlante",
            "Problema de Alarma",
            "Hora y/o Fecha Cambiada",
            "Problema de Visual",
        ]
        return {
            "sucursales": sucursales,
            "direccion": direccion,
            "contactos": contactos,
            "correos": correos,
            "tecnicos": tecnicos,
            "derivaciones": derivaciones,
            "soportes": soportes,
            "problemas": problemas,
        }

    def obtener_listas_incidencias(self) -> dict[str, list[str]]:
        clientes = self.obtener_catalogo_clientes()
        sucursales = self.obtener_catalogo_sucursales()
        problemas = [
            "DesconexiÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â³n",
            "Problema de Parlante",
            "Problema de Alarma",
            "Hora y/o Fecha Cambiada",
            "Problema de Visual",
        ]
        return {"clientes": clientes, "sucursales": sucursales, "problemas": problemas}

    def _clave_contacto_envio(self, nombre: str = "", email: str = "", telefono: str = "") -> str:
        email_limpio = str(email or "").strip().lower()
        if email_limpio in {"-", "--", "sin correo"}:
            email_limpio = ""
        if email_limpio:
            return f"email:{email_limpio}"
        telefono_limpio = re.sub(r"\D+", "", str(telefono or ""))
        if telefono_limpio:
            return f"tel:{telefono_limpio}"
        nombre_limpio = self._normalizar_texto(nombre or "")
        return f"nombre:{nombre_limpio}" if nombre_limpio else ""

    def _extraer_valor_log_envio(self, observacion: str, etiqueta: str) -> str:
        match = re.search(rf"(?:^|\|)\s*{re.escape(etiqueta)}:\s*(.*?)(?=\s*\|\s*[^|]+:|$)", observacion or "", re.IGNORECASE)
        return (match.group(1).strip() if match else "").strip()

    def _obtener_envios_informacion_contacto_por_odt(self, odts: list[Any] | None = None) -> dict[str, dict[str, Any]]:
        stmt_base = select(
            RegistroCorreoCliente.odt,
            RegistroCorreoCliente.fecha_envio,
            RegistroCorreoCliente.observacion,
        ).where(RegistroCorreoCliente.observacion.ilike("%Envio de informacion a contacto de cliente%"))

        odt_filter = sorted({str(odt or "").strip() for odt in (odts or []) if str(odt or "").strip()})
        if odts is not None and not odt_filter:
            rows = []
        elif odt_filter:
            rows = []
            for i in range(0, len(odt_filter), 900):
                chunk = odt_filter[i : i + 900]
                rows.extend(
                    self.db.execute(
                        stmt_base
                        .where(RegistroCorreoCliente.odt.in_(chunk))
                        .order_by(RegistroCorreoCliente.fecha_envio.asc(), RegistroCorreoCliente.id.asc())
                    ).all()
                )
        else:
            rows = self.db.execute(
                stmt_base.order_by(RegistroCorreoCliente.fecha_envio.asc(), RegistroCorreoCliente.id.asc())
            ).all()

        resumen: dict[str, dict[str, Any]] = {}
        for odt, fecha_envio, observacion in rows:
            odt_key = str(odt or "").strip()
            obs = str(observacion or "")
            if not odt_key or "Envio de informacion a contacto de cliente" not in obs:
                continue

            estado_correo = self._normalizar_texto(self._extraer_valor_log_envio(obs, "Estado correo"))
            estado_whatsapp = self._normalizar_texto(self._extraer_valor_log_envio(obs, "Estado WhatsApp"))
            if estado_correo != "enviado" and estado_whatsapp != "enviado":
                continue

            nombre = self._extraer_valor_log_envio(obs, "Contacto")
            telefono = self._extraer_valor_log_envio(obs, "Telefono")
            email = self._extraer_valor_log_envio(obs, "Correo")
            clave = self._clave_contacto_envio(nombre=nombre, email=email, telefono=telefono)
            if not clave:
                continue

            item = resumen.setdefault(
                odt_key,
                {"total": 0, "claves": set(), "contactos": [], "ultimo_envio": None},
            )
            if clave in item["claves"]:
                item["ultimo_envio"] = fecha_envio or item.get("ultimo_envio")
                continue
            item["claves"].add(clave)
            item["contactos"].append(
                {
                    "nombre": nombre,
                    "telefono": telefono,
                    "email": email,
                    "canal": "correo" if estado_correo == "enviado" else "whatsapp",
                    "fecha": _to_ddmmyyyy_hhmm(fecha_envio) if fecha_envio else "",
                }
            )
            item["ultimo_envio"] = fecha_envio or item.get("ultimo_envio")

        salida: dict[str, dict[str, Any]] = {}
        for odt, item in resumen.items():
            claves = sorted(item.get("claves") or [])
            salida[odt] = {
                "total": len(claves),
                "claves": claves,
                "contactos": item.get("contactos") or [],
                "ultimo_envio": _to_ddmmyyyy_hhmm(item.get("ultimo_envio")) if item.get("ultimo_envio") else "",
            }
        return salida


    def obtener_incidencias_por_puesto(
        self,
        tecnico: str | None = None,
        solo_panel_tecnico: bool = False,
        solo_servicio_tecnico: bool = False,
        incluir_ventas: bool = True,
    ) -> list[list[Any]]:
        def _fmt_fecha(v: Any) -> str:
            if isinstance(v, datetime):
                return _to_ddmmyyyy_hhmm(v)
            if v is None:
                return ""
            s = str(v).strip()
            try:
                dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
                return _to_ddmmyyyy_hhmm(dt)
            except Exception:
                return s

        def _fmt_fecha_raw(v: Any) -> str:
            # Formato parseable por el popover de tiempo (ver detalle_ticket.html),
            # a diferencia de _fmt_fecha que devuelve dd/mm/yyyy solo para mostrar.
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d %H:%M:%S")
            return ""

        def _build_direccion_cache() -> dict[str, tuple[str, str, str, str, int]]:
            # valor: (direccion, region, comuna, tabla, id) — tabla/id identifican
            # el registro real (SucursalBBDD o ClienteBBDD) para poder editar
            # region/comuna despues desde la UI (ver actualizar_region_comuna_sucursal).
            cache: dict[str, tuple[str, str, str, str, int]] = {}

            def add(nombre: Any, direccion: Any, region: Any, comuna: Any, tabla: str, entidad_id: Any) -> None:
                nombre_txt = str(nombre or "").strip()
                direccion_txt = str(direccion or "").strip()
                if not nombre_txt or not direccion_txt:
                    return
                valor = (direccion_txt, str(region or "").strip(), str(comuna or "").strip(), tabla, int(entidad_id))
                for key in {
                    self._normalizar_texto(nombre_txt),
                    self._normalizar_nombre_sucursal_match(nombre_txt),
                }:
                    if key and key not in cache:
                        cache[key] = valor

            try:
                for suc_id, nombre_sucursal, direccion_sucursal, region, comuna in self.db.execute(
                    select(SucursalBBDD.id, SucursalBBDD.nombre_sucursal, SucursalBBDD.direccion_sucursal, SucursalBBDD.region, SucursalBBDD.comuna)
                    .where(SucursalBBDD.direccion_sucursal.is_not(None))
                ).all():
                    add(nombre_sucursal, direccion_sucursal, region, comuna, "sucursal", suc_id)
            except Exception:
                self.db.rollback()

            try:
                for cli_id, cliente, direccion, region, comuna in self.db.execute(
                    select(ClienteBBDD.id, ClienteBBDD.cliente, ClienteBBDD.direccion, ClienteBBDD.region, ClienteBBDD.comuna)
                    .where(ClienteBBDD.direccion.is_not(None))
                ).all():
                    add(cliente, direccion, region, comuna, "cliente", cli_id)
            except Exception:
                self.db.rollback()

            return cache

        def _direccion_cliente_cached(cliente: str, cache: dict[str, tuple[str, str, str, str, int]]) -> tuple[str, str, str, str, int]:
            cliente_txt = str(cliente or "").strip()
            if not cliente_txt:
                return ("", "", "", "", 0)
            key_exacta = self._normalizar_texto(cliente_txt)
            if key_exacta in cache:
                return cache[key_exacta]
            objetivo = self._normalizar_nombre_sucursal_match(cliente_txt)
            if objetivo in cache:
                return cache[objetivo]

            mejor = ("", "", "", "", 0)
            mejor_score = 0
            for nombre_norm, valor in cache.items():
                score = self._score_nombre_sucursal_match(objetivo, nombre_norm)
                if score > mejor_score:
                    mejor_score = score
                    mejor = valor
            return mejor if mejor_score >= 60 else ("", "", "", "", 0)

        def _es_terminal_tecnico(row: Registro) -> bool:
            derivacion = self._normalizar_texto(getattr(row, "derivacion", "") or "")
            estado = self._normalizar_texto(getattr(row, "estado", "") or "")
            return (
                "termin" in estado
                or "final" in estado
                or "finalizado" in derivacion
                or "terminado" in derivacion
                or "repetida" in derivacion
            )

        def _fecha_ref_tecnico(row: Registro) -> datetime | None:
            for value in (
                getattr(row, "fecha_cierre", None),
                getattr(row, "fecha_derivacion_tecnico", None),
                getattr(row, "fecha_derivacion_area", None),
                getattr(row, "fecha_registro", None),
            ):
                if isinstance(value, datetime):
                    if value.tzinfo is not None:
                        return value.astimezone(timezone.utc).replace(tzinfo=None)
                    return value
            return None

        def _fila_visible_panel_tecnico(row: Registro) -> bool:
            derivacion = self._normalizar_texto(getattr(row, "derivacion", "") or "")
            if "servicio tecnico" not in derivacion and "tecnico externo" not in derivacion and not _es_terminal_tecnico(row):
                return False
            if not _es_terminal_tecnico(row):
                return True
            fecha_ref = _fecha_ref_tecnico(row)
            if fecha_ref is None:
                return False
            return fecha_ref >= datetime.now() - timedelta(days=45)

        tecnico_norm_original = self._normalizar_texto(tecnico or "")
        tecnico_norm = tecnico_norm_original
        if solo_panel_tecnico and tecnico_norm in NON_ASSIGNABLE_TECHNICIAN_NAMES:
            tecnico_norm = ""
        stmt = select(Registro)
        if solo_servicio_tecnico:
            # Mantener matching tolerante al mojibake historico de derivacion:
            # "Servicio Tecnico", "Servicio Tcnico" y "Servicio T?cnico".
            stmt = stmt.where(
                or_(
                    Registro.derivacion.ilike("%servicio t%cnico%"),
                    Registro.derivacion.ilike("%cnico externo%"),
                    Registro.derivacion.ilike("%tecnico externo%"),
                    Registro.derivacion.ilike("%finalizado sin vt%"),
                    Registro.estado.ilike("%finalizado sin vt%"),
                )
            )
        stmt = stmt.order_by(Registro.id.asc())
        rows = self._run_registro_query(
            lambda: self.db.scalars(stmt).all(),
            "cargar las incidencias por puesto",
        )
        out: list[list[Any]] = []
        odts_vistas: set[str] = set()
        envios_info = self._obtener_envios_informacion_contacto_por_odt(
            [r.odt for r in rows] if solo_servicio_tecnico else None
        )
        direccion_cache = _build_direccion_cache()
        # Memo por nombre de sucursal: muchas filas comparten la misma sucursal
        # y _direccion_cliente_cached hace un scan con fuzzy-match (costoso) del
        # cache completo cuando no hay match exacto — sin este memo se repetía
        # ese scan una vez POR FILA en vez de una vez por sucursal distinta,
        # el cuello de botella real detras de la lentitud de esta pagina.
        direccion_por_sucursal: dict[str, tuple[str, str, str, str, int]] = {}
        for r in rows:
            if solo_panel_tecnico and not _fila_visible_panel_tecnico(r):
                continue
            if tecnico_norm and not self._fila_aplica_a_tecnico(r, tecnico_norm):
                continue
            sucursal = str(r.cliente or "").strip()
            if sucursal in direccion_por_sucursal:
                direccion_bd, region, comuna, geo_tabla, geo_id = direccion_por_sucursal[sucursal]
            else:
                direccion_bd, region, comuna, geo_tabla, geo_id = _direccion_cliente_cached(sucursal, direccion_cache)
                direccion_por_sucursal[sucursal] = (direccion_bd, region, comuna, geo_tabla, geo_id)
            direccion = str(r.direccion or "").strip() or direccion_bd
            # Para incidencias_servicio_tecnico.html la observacion visible debe salir de "observacion"
            # y no de "detalle_problema".
            detalle = str(r.observacion or "").strip()
            obs_pend = str(getattr(r, "observacion_pendiente", "") or "").strip()
            obs_soporte = str(getattr(r, "observacion_soporte", "") or "").strip()
            obs_servicio = str(getattr(r, "observacion_servicio", "") or "").strip()
            estado = str(r.estado or "").strip() or "Pendiente"
            out.append([
                r.odt,
                _fmt_fecha(r.fecha_registro),
                r.puesto or "",
                sucursal,
                r.problema,
                r.derivacion,
                detalle,
                r.tecnicos or "",
                estado,
                r.id,
                r.acompanante or "",
                _fmt_fecha(r.fecha_cierre),
                str(r.prioridad or ""),
                direccion,
                obs_pend,
                obs_soporte,
                obs_servicio,
                str(getattr(r, "observacion_final", "") or "").strip(),
                envios_info.get(str(r.odt or "").strip(), {"total": 0, "claves": [], "contactos": [], "ultimo_envio": ""}),
                _fmt_fecha_raw(getattr(r, "fecha_inicio_trabajo", None)),
                _fmt_fecha_raw(getattr(r, "fecha_fin_trabajo", None)),
                region,
                comuna,
                geo_tabla,
                geo_id or "",
            ])
            odt_key = self._normalizar_texto(r.odt)
            if odt_key:
                odts_vistas.add(odt_key)

        if incluir_ventas:
            try:
                rows_venta = (
                    self.db.execute(
                        select(VentaODS, ServicioTecnicoVentaODT, AdministracionODT, SoporteTecnicoVentaODT)
                        .outerjoin(
                            ServicioTecnicoVentaODT,
                            func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                        )
                        .outerjoin(
                            AdministracionODT,
                            func.lower(func.trim(AdministracionODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                        )
                        .outerjoin(
                            SoporteTecnicoVentaODT,
                            func.lower(func.trim(SoporteTecnicoVentaODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                        )
                        .where(VentaODS.estado != "Anulada")
                        .order_by(VentaODS.created_at.asc(), VentaODS.id.asc())
                    )
                    .all()
                )
            except Exception:
                rows_venta = []

            for ods, st, adm, soporte in rows_venta:
                tecnico_venta = str(getattr(st, "tecnico_a_cargo", "") or getattr(adm, "tecnico", "") or "").strip()
                acompanante_venta = str(getattr(st, "acompanante", "") or getattr(adm, "acompanante", "") or "").strip()
                if not tecnico_venta and not acompanante_venta:
                    continue
                if tecnico_norm:
                    asignados_venta = self._normalizar_texto(f"{tecnico_venta} {acompanante_venta}")
                    if tecnico_norm not in asignados_venta:
                        continue

                odt_key = self._normalizar_texto(ods.codigo)
                if odt_key and odt_key in odts_vistas:
                    continue

                estado_venta = str(ods.estado or "").strip() or "Pendiente"
                tipo_servicio_norm = self._normalizar_texto(ods.tipo_servicio or "")
                if "televigilancia" in tipo_servicio_norm:
                    venta_finalizada = bool(
                        soporte
                        and (
                            getattr(soporte, "terminado", False)
                            or getattr(soporte, "fecha_terminado", None)
                        )
                    )
                else:
                    venta_finalizada = bool(
                        st
                        and (
                            getattr(st, "finalizado", False)
                            or getattr(st, "fecha_cierre", None)
                        )
                    )
                estado_visible = "Terminado" if venta_finalizada else "En Proceso"
                fecha_ref = (
                    getattr(st, "fecha_recepcion_solicitud_instalacion", None)
                    or getattr(st, "updated_at", None)
                    or ods.created_at
                )
                fecha_cierre = getattr(st, "fecha_cierre", None)
                if not isinstance(fecha_cierre, datetime):
                    fecha_cierre = None
                out.append([
                    ods.codigo or "",
                    _fmt_fecha(fecha_ref),
                    "",
                    ods.nombre_sucursal or ods.razon_social or "",
                    ods.tipo_servicio or "",
                    "Servicio Tecnico",
                    ods.observacion or ods.consideraciones or "",
                    tecnico_venta,
                    estado_visible,
                    ods.id,
                    acompanante_venta,
                    _fmt_fecha(fecha_cierre),
                    "",
                    ods.direccion_sucursal or "",
                    "",
                    "",
                    "",
                    estado_venta,
                    envios_info.get(str(ods.codigo or "").strip(), {"total": 0, "claves": [], "contactos": [], "ultimo_envio": ""}),
                    _fmt_fecha_raw(getattr(st, "fecha_inicio_trabajo", None)),
                    _fmt_fecha_raw(getattr(st, "fecha_fin_trabajo", None)),
                ])
        return self._filtrar_incidencias_para_tecnico(out, "" if solo_panel_tecnico else tecnico)

    def obtener_incidencias_servicio_tecnico(self) -> list[list[Any]]:
        return self.obtener_incidencias_por_puesto(
            solo_servicio_tecnico=True,
            incluir_ventas=False,
        )

    def obtener_ruta_optima_tecnico(
        self,
        tecnico: str | None = None,
        origen: str = "1 Oriente 1180, Viña del Mar, Chile",
    ) -> dict[str, Any]:
        registros = self.obtener_incidencias_por_puesto(tecnico, solo_panel_tecnico=True)

        def _coords_validas(lat: str, lng: str) -> bool:
            try:
                lat_f = float(str(lat or "").replace(",", "."))
                lng_f = float(str(lng or "").replace(",", "."))
                return -90 <= lat_f <= 90 and -180 <= lng_f <= 180
            except Exception:
                return False

        def _distancia2(a_lat: str, a_lng: str, b_lat: str, b_lng: str) -> float:
            a1 = float(str(a_lat).replace(",", "."))
            a2 = float(str(a_lng).replace(",", "."))
            b1 = float(str(b_lat).replace(",", "."))
            b2 = float(str(b_lng).replace(",", "."))
            return ((a1 - b1) ** 2) + ((a2 - b2) ** 2)

        def _coords_para_direccion(direccion: str) -> tuple[str, str]:
            lat_bd, lng_bd = self._coordenadas_por_direccion_bd(direccion)
            if _coords_validas(lat_bd, lng_bd):
                return lat_bd, lng_bd

            lat_geo, lng_geo = self._geocodificar_direccion(direccion)
            if _coords_validas(lat_geo, lng_geo):
                return lat_geo, lng_geo

            lat_apx, lng_apx = self._coordenadas_aproximadas_por_direccion(direccion)
            if _coords_validas(lat_apx, lng_apx):
                return lat_apx, lng_apx
            return "", ""

        origen_lat, origen_lng = _coords_para_direccion(origen)
        if not _coords_validas(origen_lat, origen_lng):
            origen_lat, origen_lng = "-33.024569", "-71.551831"

        paradas = []
        for fila in registros:
            direccion = str(fila[13] if len(fila) > 13 else "").strip()
            estado = str(fila[8] if len(fila) > 8 else "").strip()
            if not direccion or self._normalizar_texto(estado) != "en proceso":
                continue

            latitud, longitud = _coords_para_direccion(direccion)
            paradas.append(
                {
                    "odt": str(fila[0] if len(fila) > 0 else "").strip(),
                    "fecha": str(fila[1] if len(fila) > 1 else "").strip(),
                    "cliente": str(fila[3] if len(fila) > 3 else "").strip(),
                    "servicio": str(fila[4] if len(fila) > 4 else "").strip(),
                    "estado": estado,
                    "prioridad": str(fila[12] if len(fila) > 12 else "").strip(),
                    "direccion": direccion,
                    "latitud": latitud,
                    "longitud": longitud,
                }
            )

        geocodificadas = [p for p in paradas if _coords_validas(p["latitud"], p["longitud"])]
        no_geocodificadas = [p for p in paradas if not _coords_validas(p["latitud"], p["longitud"])]

        orden: list[dict[str, Any]] = []
        pendientes = geocodificadas[:]
        actual_lat, actual_lng = origen_lat, origen_lng
        while pendientes:
            mejor_idx = 0
            mejor_dist = float("inf")
            for idx, parada in enumerate(pendientes):
                dist = _distancia2(actual_lat, actual_lng, parada["latitud"], parada["longitud"])
                if dist < mejor_dist:
                    mejor_dist = dist
                    mejor_idx = idx
            siguiente = pendientes.pop(mejor_idx)
            orden.append(siguiente)
            actual_lat, actual_lng = siguiente["latitud"], siguiente["longitud"]

        orden.extend(no_geocodificadas)

        direcciones = [str(p.get("direccion") or "").strip() for p in orden if str(p.get("direccion") or "").strip()]
        google_url = ""
        if direcciones:
            destino = direcciones[-1]
            waypoints = "|".join(direcciones[:-1])
            google_url = (
                "https://www.google.com/maps/dir/?api=1"
                f"&origin={quote_plus(origen)}"
                f"&destination={quote_plus(destino)}"
                "&travelmode=driving"
            )
            if waypoints:
                google_url += f"&waypoints={quote_plus(waypoints)}"

        waze_url = ""
        if orden:
            primera = orden[0]
            query = ", ".join(
                part
                for part in [
                    str(primera.get("direccion") or "").strip(),
                    str(primera.get("cliente") or "").strip(),
                    "Chile",
                ]
                if part
            )
            if query:
                waze_url = f"https://www.waze.com/ul?q={quote_plus(query)}&navigate=yes"
            elif _coords_validas(primera["latitud"], primera["longitud"]):
                ll_value = f"{primera['latitud']},{primera['longitud']}"
                waze_url = f"https://www.waze.com/ul?ll={quote_plus(ll_value)}&navigate=yes"

        return {
            "origen": origen,
            "total_odts": len(orden),
            "geocodificadas": len(geocodificadas),
            "sin_geocodificar": len(no_geocodificadas),
            "paradas": orden,
            "google_maps_url": google_url,
            "waze_url": waze_url,
        }

    def obtener_incidencias_derivadas_cliente(self) -> list[list[Any]]:
        rows = self.obtener_incidencias_por_puesto()
        derivaciones_coord = {"cliente", "coordinacion"}
        return [
            row
            for row in rows
            if len(row) > 5 and self._normalizar_texto(row[5]) in derivaciones_coord
        ]

    def obtener_dashboard_coordinacion(self, desde: date | None = None, hasta: date | None = None) -> dict[str, Any]:
        """4 indicadores gerenciales de lo que Coordinacion tiene que resolver:
        incidencias derivadas a Cliente/Coordinacion (pendientes, antiguedad
        promedio y maxima) + informes semanales de protocolos sin enviar.
        Mismo criterio ILIKE que el badge de seleccion_panel_coordinacion.html
        (incidencias.py) para no reportar numeros distintos entre paginas.
        `desde`/`hasta` filtran por fecha_registro (rango inclusivo, opcional)."""
        filtro_derivacion = or_(
            Registro.derivacion.ilike("%client%"),
            Registro.derivacion.ilike("%coordinaci%"),
        )
        filtros_fecha = []
        if desde:
            filtros_fecha.append(Registro.fecha_registro >= datetime.combine(desde, time.min))
        if hasta:
            filtros_fecha.append(Registro.fecha_registro <= datetime.combine(hasta, time.max))

        pendientes = (
            self.db.query(Registro)
            .filter(
                filtro_derivacion,
                or_(Registro.estado.is_(None), ~Registro.estado.ilike("Termin%")),
                or_(Registro.estado.is_(None), Registro.estado != "Repetida"),
                *filtros_fecha,
            )
            .order_by(Registro.fecha_registro.asc())
            .all()
        )

        ahora = datetime.now()
        top_antiguas: list[dict[str, Any]] = []
        dias_totales = 0
        for r in pendientes:
            dias = (ahora - r.fecha_registro).days if r.fecha_registro else 0
            dias_totales += max(dias, 0)
            top_antiguas.append({
                "odt": r.odt or "",
                "cliente": r.cliente or "",
                "problema": r.problema or "",
                "dias": max(dias, 0),
            })
        top_antiguas.sort(key=lambda it: it["dias"], reverse=True)

        total_pendientes = len(pendientes)
        antiguedad_promedio = round(dias_totales / total_pendientes, 1) if total_pendientes else 0.0
        antiguedad_maxima = top_antiguas[0]["dias"] if top_antiguas else 0

        # Informes semanales: respeta el mismo Desde/Hasta que el resto del
        # dashboard, filtrando por la fecha en que se genero el informe.
        filtros_fecha_protocolos = []
        if desde:
            filtros_fecha_protocolos.append(ProtocoloInforme.created_at >= datetime.combine(desde, time.min))
        if hasta:
            filtros_fecha_protocolos.append(ProtocoloInforme.created_at <= datetime.combine(hasta, time.max))

        pendiente_protocolos_semanales = (
            self.db.query(ProtocoloInforme)
            .filter(
                ProtocoloInforme.tipo_informe == "SEMANAL",
                ProtocoloInforme.estado.notin_(["ENVIADO", "RECHAZADO"]),
                *filtros_fecha_protocolos,
            )
            .count()
        )
        total_protocolos_semanales = (
            self.db.query(ProtocoloInforme)
            .filter(ProtocoloInforme.tipo_informe == "SEMANAL", *filtros_fecha_protocolos)
            .count()
        )

        # Buckets de antiguedad: siempre sobre TODAS las pendientes (sin
        # aplicar el filtro de fecha), a pedido explicito, para que este
        # grafico no cambie segun el rango Desde/Hasta seleccionado arriba.
        pendientes_sin_filtro = (
            self.db.query(Registro)
            .filter(
                filtro_derivacion,
                or_(Registro.estado.is_(None), ~Registro.estado.ilike("Termin%")),
                or_(Registro.estado.is_(None), Registro.estado != "Repetida"),
            )
            .all()
        ) if filtros_fecha else pendientes

        detalle_sin_filtro = [
            {
                "cliente": str(r.cliente or "").strip() or "Sin cliente",
                "problema": str(r.problema or "").strip() or "Sin especificar",
                "dias": max((ahora - r.fecha_registro).days, 0) if r.fecha_registro else 0,
            }
            for r in pendientes_sin_filtro
        ]
        rangos = [("0-7 días", 0, 7), ("8-15 días", 8, 15), ("16-30 días", 16, 30), ("31-60 días", 31, 60), ("60+ días", 61, None)]
        buckets = []
        for label, rango_desde, rango_hasta in rangos:
            if rango_hasta is None:
                items_bucket = [it for it in detalle_sin_filtro if it["dias"] >= rango_desde]
            else:
                items_bucket = [it for it in detalle_sin_filtro if rango_desde <= it["dias"] <= rango_hasta]

            # Desglose por sucursal (y, dentro de cada una, por tipo de
            # problema) para el popup que se abre al hacer clic en la barra.
            por_sucursal_bucket: dict[str, dict[str, Any]] = {}
            for it in items_bucket:
                entry = por_sucursal_bucket.setdefault(
                    it["cliente"], {"cliente": it["cliente"], "cantidad": 0, "problemas": {}}
                )
                entry["cantidad"] += 1
                entry["problemas"][it["problema"]] = entry["problemas"].get(it["problema"], 0) + 1

            sucursales_bucket = sorted(
                (
                    {
                        "cliente": v["cliente"],
                        "cantidad": v["cantidad"],
                        "problemas": [
                            {"problema": p, "cantidad": c}
                            for p, c in sorted(v["problemas"].items(), key=lambda kv: kv[1], reverse=True)
                        ],
                    }
                    for v in por_sucursal_bucket.values()
                ),
                key=lambda x: x["cantidad"],
                reverse=True,
            )
            buckets.append({"label": label, "cantidad": len(items_bucket), "sucursales": sucursales_bucket})

        # Ranking de clientes con mas incidencias pendientes, con el desglose
        # de tipos de problema para el popup "+ Info".
        conteo_clientes: dict[str, int] = {}
        conteo_problemas: dict[str, int] = {}
        problemas_por_cliente: dict[str, dict[str, int]] = {}
        for it in top_antiguas:
            nombre = it["cliente"] or "Sin cliente"
            conteo_clientes[nombre] = conteo_clientes.get(nombre, 0) + 1
            problema = it["problema"] or "Sin especificar"
            conteo_problemas[problema] = conteo_problemas.get(problema, 0) + 1
            bucket = problemas_por_cliente.setdefault(nombre, {})
            bucket[problema] = bucket.get(problema, 0) + 1
        top_clientes = sorted(conteo_clientes.items(), key=lambda kv: kv[1], reverse=True)[:10]

        # Incidencias por mes: respeta el Desde/Hasta si viene seteado; sin
        # filtro, muestra los ultimos 6 meses (comportamiento por defecto).
        mensual = self._obtener_incidencias_por_mes_coordinacion(desde=desde, hasta=hasta)

        protocolos_enviados = total_protocolos_semanales - pendiente_protocolos_semanales
        tasa_envio_informes = (
            round(protocolos_enviados / total_protocolos_semanales * 100, 1) if total_protocolos_semanales else 0.0
        )
        odt_resueltas = sum(m["resueltas"] for m in mensual)
        odt_no_resueltas = sum(m["en_proceso"] + m["pendientes"] for m in mensual)
        tasa_resolucion_odt = (
            round(odt_resueltas / (odt_resueltas + odt_no_resueltas) * 100, 1)
            if (odt_resueltas + odt_no_resueltas) else 0.0
        )

        return {
            "incidencias_pendientes": total_pendientes,
            "antiguedad_promedio_dias": antiguedad_promedio,
            "antiguedad_maxima_dias": antiguedad_maxima,
            "protocolos_semanales_pendientes": pendiente_protocolos_semanales,
            "protocolos_semanales_enviados": protocolos_enviados,
            "tasa_envio_informes": tasa_envio_informes,
            "tasa_resolucion_odt": tasa_resolucion_odt,
            "top_antiguas": top_antiguas[:10],
            "antiguedad_buckets": buckets,
            "top_clientes": [
                {
                    "cliente": c,
                    "cantidad": n,
                    "problemas": [
                        {"problema": p, "cantidad": cnt}
                        for p, cnt in sorted(problemas_por_cliente.get(c, {}).items(), key=lambda kv: kv[1], reverse=True)
                    ],
                }
                for c, n in top_clientes
            ],
            "mensual": mensual,
            # Resumen completo (todas las pendientes, no solo el top 10 de la
            # tabla) para el popup "+ Info" de "Incidencias pendientes mas antiguas".
            "resumen_pendientes": {
                "por_sucursal": sorted(
                    ({"cliente": c, "cantidad": n} for c, n in conteo_clientes.items()),
                    key=lambda x: x["cantidad"],
                    reverse=True,
                ),
                "por_tipo": sorted(
                    (
                        {"problema": p, "cantidad": n}
                        for p, n in conteo_problemas.items()
                        if p.strip().lower() != "oficina atc"
                    ),
                    key=lambda x: x["cantidad"],
                    reverse=True,
                ),
            },
        }

    def _obtener_incidencias_por_mes_coordinacion(
        self, meses: int = 6, desde: date | None = None, hasta: date | None = None
    ) -> list[dict[str, Any]]:
        """Conteo de incidencias de Cliente/Coordinacion por estado (Resueltas
        / En proceso / Pendientes). Sin filtro: ultimos `meses` meses. Con
        desde/hasta: un bucket por DIA dentro del rango (mas legible que un
        solo bucket mensual cuando el rango es corto); si el rango supera
        ~2 meses, cae de vuelta a buckets mensuales para no saturar el grafico."""
        filtro_derivacion = or_(
            Registro.derivacion.ilike("%client%"),
            Registro.derivacion.ilike("%coordinaci%"),
        )
        meses_es = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

        if desde and hasta:
            desde_dt = datetime.combine(desde, time.min)
            hasta_dt = datetime.combine(hasta, time.max)
            if (hasta - desde).days <= 62:
                buckets_orden: list[Any] = []
                d = desde
                while d <= hasta:
                    buckets_orden.append(d)
                    d += timedelta(days=1)
                clave_fn = lambda dt: dt.date()  # noqa: E731
                label_fn = lambda k: k.strftime("%d/%m")  # noqa: E731
            else:
                y0, m0 = desde.year, desde.month
                y1, m1 = hasta.year, hasta.month
                buckets_orden = []
                y, m = y0, m0
                while (y, m) <= (y1, m1) and len(buckets_orden) < 24:
                    buckets_orden.append((y, m))
                    m += 1
                    if m == 13:
                        m, y = 1, y + 1
                clave_fn = lambda dt: (dt.year, dt.month)  # noqa: E731
                label_fn = lambda k: f"{meses_es[k[1] - 1]} {k[0]}"  # noqa: E731
        else:
            hoy = datetime.now()
            y, m = hoy.year, hoy.month
            buckets_orden = []
            for _ in range(meses):
                buckets_orden.append((y, m))
                m -= 1
                if m == 0:
                    m, y = 12, y - 1
            buckets_orden.reverse()
            desde_dt = datetime(buckets_orden[0][0], buckets_orden[0][1], 1)
            hasta_dt = None
            clave_fn = lambda dt: (dt.year, dt.month)  # noqa: E731
            label_fn = lambda k: f"{meses_es[k[1] - 1]} {k[0]}"  # noqa: E731

        filtros_fecha = [Registro.fecha_registro >= desde_dt]
        if hasta_dt:
            filtros_fecha.append(Registro.fecha_registro <= hasta_dt)

        registros = (
            self.db.query(Registro)
            .filter(
                filtro_derivacion,
                or_(Registro.estado.is_(None), Registro.estado != "Repetida"),
                *filtros_fecha,
            )
            .all()
        )

        odts = [r.odt for r in registros if r.odt]
        odts_con_mensaje: set[str] = set()
        for i in range(0, len(odts), 500):
            chunk = odts[i:i + 500]
            filas = (
                self.db.query(RegistroCorreoCliente.odt)
                .filter(
                    RegistroCorreoCliente.odt.in_(chunk),
                    RegistroCorreoCliente.observacion.ilike("%Envio de informacion a contacto de cliente%"),
                )
                .distinct()
                .all()
            )
            odts_con_mensaje.update(f[0] for f in filas)

        conteo: dict[Any, dict[str, int]] = {
            clave: {"resueltas": 0, "en_proceso": 0, "pendientes": 0} for clave in buckets_orden
        }
        for r in registros:
            estado = str(r.estado or "").strip().lower()
            es_resuelta = estado.startswith("termin")
            fecha_ref = r.fecha_cierre if (es_resuelta and r.fecha_cierre) else r.fecha_registro
            if not fecha_ref:
                continue
            clave = clave_fn(fecha_ref)
            if clave not in conteo:
                continue
            if es_resuelta:
                conteo[clave]["resueltas"] += 1
            elif r.odt in odts_con_mensaje:
                conteo[clave]["en_proceso"] += 1
            else:
                conteo[clave]["pendientes"] += 1

        return [
            {
                "mes": label_fn(clave),
                "resueltas": conteo[clave]["resueltas"],
                "en_proceso": conteo[clave]["en_proceso"],
                "pendientes": conteo[clave]["pendientes"],
            }
            for clave in buckets_orden
        ]

    def obtener_detalle_informes_semanales(self) -> dict[str, Any]:
        """Detalle para el popup "+ info" del grafico de Informes semanales
        del dashboard de Coordinacion: un listado unico por informe (uno por
        sucursal/semana), cada uno marcado enviado o pendiente, con la
        antiguedad del pendiente. Mismo criterio ENVIADO/RECHAZADO que
        obtener_dashboard_coordinacion para no reportar numeros distintos
        entre el grafico y este detalle."""
        informes = (
            self.db.query(ProtocoloInforme)
            .filter(ProtocoloInforme.tipo_informe == "SEMANAL")
            .all()
        )
        ahora = datetime.now()
        items: list[dict[str, Any]] = []

        for inf in informes:
            cliente = str(inf.cliente or "").strip() or "Sin cliente"
            sucursal = str(inf.sucursal or "").strip() or "Sin sucursal"
            es_pendiente = str(inf.estado or "").strip().upper() not in ("ENVIADO", "RECHAZADO")
            creado = inf.created_at
            actualizado = inf.updated_at or creado
            horas = max((ahora - creado).total_seconds() / 3600, 0) if (es_pendiente and creado) else 0
            items.append(
                {
                    "id": inf.id,
                    "cliente": cliente,
                    "sucursal": sucursal,
                    "pendiente": es_pendiente,
                    "creado_en": creado.isoformat() if creado else None,
                    "actualizado_en": actualizado.isoformat() if actualizado else None,
                    "horas_antiguedad": round(horas, 1),
                }
            )

        pendientes = [it for it in items if it["pendiente"]]
        enviados = [it for it in items if not it["pendiente"]]
        pendientes.sort(key=lambda it: it["creado_en"] or "")
        enviados.sort(key=lambda it: it["actualizado_en"] or "", reverse=True)

        return {
            "items": pendientes + enviados,
            "total_enviados": len(enviados),
            "total_pendientes": len(pendientes),
            "sucursales_con_pendientes": len({(it["cliente"], it["sucursal"]) for it in pendientes}),
        }

    def generar_informe_coordinacion_pdf(self, desde: date | None = None, hasta: date | None = None) -> bytes:
        """Genera el PDF "Informe de Gestion - Coordinacion con Cliente"
        (KPIs, donut de informes semanales, ranking de sucursales, tipos de
        incidencia y conclusion en lenguaje simple) para el rango desde/hasta
        seleccionado en el dashboard. Mismo estilo corporativo que el resto
        de los informes de gestion de ATC."""
        import io

        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor, white
        from reportlab.platypus import (
            BaseDocTemplate, Frame, PageTemplate, PageBreak,
            Table, TableStyle, Paragraph, Spacer, HRFlowable, Flowable,
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_JUSTIFY
        from reportlab.graphics.shapes import Drawing, Circle, String
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.charts.barcharts import HorizontalBarChart

        _C_DARK, _C_ORANGE, _C_ORDK = "#0b1424", "#f4a672", "#c2410c"
        _C_BG, _C_BORDER, _C_TEXT, _C_SOFT, _C_GREY = "#f7f8fa", "#e5e7eb", "#111827", "#4b5563", "#9ca3af"
        _C_OK, _C_WARN, _C_BAD = "#1e9c83", "#d97706", "#c0392b"
        C_DARK, C_ORANGE, C_ORDK = HexColor(_C_DARK), HexColor(_C_ORANGE), HexColor(_C_ORDK)
        C_BG, C_BORDER, C_TEXT, C_SOFT, C_GREY = (
            HexColor(_C_BG), HexColor(_C_BORDER), HexColor(_C_TEXT), HexColor(_C_SOFT), HexColor(_C_GREY)
        )
        C_OK, C_WARN, C_BAD = HexColor(_C_OK), HexColor(_C_WARN), HexColor(_C_BAD)

        datos = self.obtener_dashboard_coordinacion(desde=desde, hasta=hasta)
        total_pendientes = datos["incidencias_pendientes"]
        antiguedad_prom = datos["antiguedad_promedio_dias"]
        antiguedad_max = datos["antiguedad_maxima_dias"]
        protocolos_pend = datos["protocolos_semanales_pendientes"]
        protocolos_env = datos["protocolos_semanales_enviados"]
        top_antiguas = datos["top_antiguas"]
        top_clientes = datos["top_clientes"]
        por_tipo = datos["resumen_pendientes"]["por_tipo"]
        sucursales_afectadas = len(datos["resumen_pendientes"]["por_sucursal"])

        _MESES_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

        def _fmt(d: date | None) -> str:
            return d.strftime("%d/%m/%Y") if d else ""

        if desde and hasta:
            rango_txt = f"Período: {_fmt(desde)} al {_fmt(hasta)}"
        elif desde:
            rango_txt = f"Período: desde el {_fmt(desde)}"
        elif hasta:
            rango_txt = f"Período: hasta el {_fmt(hasta)}"
        else:
            rango_txt = "Período: histórico completo (sin filtro de fecha)"

        ahora = datetime.now()
        fecha_emision = ahora.strftime("%d/%m/%Y %H:%M")
        titulo_hdr = "INFORME DE GESTIÓN — COORDINACIÓN CON CLIENTE"
        subtitulo_hdr = rango_txt

        W, H = A4
        pad = 1.4 * cm
        HEADER_H = 2.7 * cm
        ORANGE_H = 5
        FOOTER_H = 1.0 * cm
        BODY_TOP = HEADER_H + ORANGE_H + 12
        BODY_BOT = FOOTER_H + 8
        fw = W - 2 * pad

        _atc_root = Path(__file__).resolve().parents[2]
        logo_path = _atc_root / "ATC" / "static" / "img" / "logo-atc.png"
        if not logo_path.exists():
            logo_path = _atc_root / "static" / "img" / "logo-atc.png"
        logo_w, logo_h = 2.8 * cm, 1.4 * cm

        def draw_page(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(C_DARK)
            canvas.rect(0, H - HEADER_H, W, HEADER_H, fill=1, stroke=0)
            if logo_path.exists():
                try:
                    canvas.drawImage(
                        str(logo_path),
                        pad, H - HEADER_H + (HEADER_H - logo_h) / 2,
                        width=logo_w, height=logo_h,
                        preserveAspectRatio=True, mask="auto",
                    )
                except Exception:
                    pass
            tx = pad + logo_w + 0.5 * cm
            canvas.setFillColor(white)
            canvas.setFont("Helvetica-Bold", 13)
            canvas.drawString(tx, H - HEADER_H + 1.35 * cm, titulo_hdr)
            canvas.setFillColor(HexColor("#fde68a"))
            canvas.setFont("Helvetica", 8.5)
            canvas.drawString(tx, H - HEADER_H + 0.75 * cm, subtitulo_hdr)
            canvas.setFillColor(C_ORANGE)
            canvas.rect(0, H - HEADER_H - ORANGE_H, W, ORANGE_H, fill=1, stroke=0)
            canvas.setFillColor(C_DARK)
            canvas.rect(0, 0, W, FOOTER_H, fill=1, stroke=0)
            canvas.setFillColor(C_GREY)
            canvas.setFont("Helvetica", 7)
            canvas.drawCentredString(
                W / 2, FOOTER_H / 2 - 3,
                f"Documento generado automáticamente  ·  Alguien Te Cuida  ·  {fecha_emision}",
            )
            canvas.setFont("Helvetica", 7)
            canvas.drawRightString(W - pad, FOOTER_H / 2 - 3, f"Página {doc.page}")
            canvas.restoreState()

        frame = Frame(
            pad, BODY_BOT, fw, H - BODY_TOP - BODY_BOT,
            leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
        )
        page_tmpl = PageTemplate(id="main", frames=[frame], onPage=draw_page)
        buf = io.BytesIO()
        doc = BaseDocTemplate(
            buf, pagesize=A4, pageTemplates=[page_tmpl],
            leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0,
            title=titulo_hdr, author="Alguien Te Cuida",
        )

        st_kpi_num = ParagraphStyle("kpiNum", fontName="Helvetica-Bold", fontSize=20, textColor=C_TEXT, leading=22, alignment=1)
        st_kpi_lbl = ParagraphStyle("kpiLbl", fontName="Helvetica-Bold", fontSize=7, textColor=C_SOFT, leading=9, alignment=1)
        st_sec = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=11, textColor=C_ORDK, leading=14, spaceBefore=14, spaceAfter=6)
        st_body = ParagraphStyle("body", fontName="Helvetica", fontSize=9.5, textColor=C_SOFT, leading=14.5, alignment=TA_JUSTIFY, spaceAfter=6)
        st_th = ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=white, leading=10)
        st_td = ParagraphStyle("td", fontName="Helvetica", fontSize=8, textColor=C_TEXT, leading=11)
        st_td_soft = ParagraphStyle("tdSoft", fontName="Helvetica", fontSize=7.5, textColor=C_SOFT, leading=10)

        class _CampoTextoEditable(Flowable):
            """Campo de formulario (AcroForm) para escribir directamente sobre
            el PDF con Adobe Reader / Vista Previa, sin pasar por Word (lo que
            rompia el estilo del informe al reconvertir a PDF)."""

            def __init__(self, width, height, name):
                Flowable.__init__(self)
                self.width = width
                self.height = height
                self.name = name

            def wrap(self, availWidth, availHeight):
                return self.width, self.height

            def draw(self):
                # Sin NeedAppearances varios lectores (Vista Previa de Mac,
                # algunos visores embebidos) muestran el campo pero no dejan
                # escribir en el — reportlab no lo agrega por defecto.
                self.canv.acroForm.extras["NeedAppearances"] = True
                self.canv.acroForm.textfield(
                    name=self.name,
                    tooltip="Observaciones",
                    x=0, y=0, width=self.width, height=self.height,
                    borderStyle="inset", borderWidth=1,
                    borderColor=C_BORDER, fillColor=white,
                    textColor=C_TEXT, fontSize=9,
                    fieldFlags="multiline",
                    value="",
                )

        story: list = []

        # ── KPIs ──────────────────────────────────────────────────────────
        def kpi_card(numero: str, etiqueta: str, color) -> Table:
            t = Table([[Paragraph(numero, st_kpi_num)], [Paragraph(etiqueta, st_kpi_lbl)]], colWidths=[fw / 4 - 8])
            t.setStyle(TableStyle([
                ("TOPPADDING", (0, 0), (-1, 0), 12), ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
                ("TOPPADDING", (0, 1), (-1, 1), 2), ("BOTTOMPADDING", (0, 1), (-1, 1), 12),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
                ("LINEABOVE", (0, 0), (-1, 0), 3, color),
                ("BACKGROUND", (0, 0), (-1, -1), white),
            ]))
            return t

        kpis = Table(
            [[
                kpi_card(str(total_pendientes), "INCIDENCIAS\nPENDIENTES", C_BAD),
                kpi_card(f"{antiguedad_prom}", "DÍAS PROMEDIO\nSIN GESTIONAR", C_WARN),
                kpi_card(str(antiguedad_max), "ANTIGÜEDAD\nMÁXIMA (DÍAS)", C_BAD),
                kpi_card(str(protocolos_pend), "INFORMES SEMANALES\nSIN ENVIAR", C_ORDK),
            ]],
            colWidths=[fw / 4] * 4,
        )
        kpis.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4)]))
        story.append(kpis)
        story.append(Spacer(1, 14))

        # ── Resumen general (explicado en lenguaje simple) ──────────────────
        story.append(Paragraph("RESUMEN GENERAL", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        if total_pendientes:
            frase_max = (
                f"La incidencia más antigua sin resolver lleva <b>{antiguedad_max} días</b> esperando gestión."
                if antiguedad_max else ""
            )
            story.append(Paragraph(
                f"El área de Coordinación con Cliente tiene actualmente <b>{total_pendientes} incidencias pendientes</b> "
                f"de gestión, distribuidas en <b>{sucursales_afectadas} sucursales</b>. En promedio, estas incidencias "
                f"llevan <b>{antiguedad_prom} días</b> sin recibir seguimiento. {frase_max} "
                f"Además, hay <b>{protocolos_pend} informes semanales de protocolos</b> que todavía no se han enviado "
                f"al cliente (de un total de {protocolos_env + protocolos_pend} generados).",
                st_body,
            ))
        else:
            story.append(Paragraph(
                "No hay incidencias pendientes de Coordinación con Cliente en el período seleccionado. "
                f"Hay <b>{protocolos_pend} informes semanales</b> sin enviar de un total de {protocolos_env + protocolos_pend}.",
                st_body,
            ))
        story.append(Spacer(1, 6))

        # ── Donut: informes semanales enviados vs pendientes ────────────────
        story.append(Paragraph("INFORMES SEMANALES DE PROTOCOLOS", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Cada semana se genera un informe de protocolos por sucursal. Este gráfico muestra cuántos ya se "
            "enviaron al cliente y cuántos siguen pendientes de envío.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        total_informes = protocolos_env + protocolos_pend
        dwg = Drawing(fw, 5.4 * cm)
        pie = Pie()
        pie.x, pie.y = fw / 2 - 3.4 * cm, 0.2 * cm
        pie.width, pie.height = 5.4 * cm, 5.4 * cm
        valores_pie = [protocolos_env, protocolos_pend]
        colores_pie = [C_OK, C_BAD]
        idx_no_cero = [i for i, v in enumerate(valores_pie) if v > 0]
        pie.data = [valores_pie[i] for i in idx_no_cero] or [1]
        for slot, orig_i in enumerate(idx_no_cero):
            pie.slices[slot].fillColor = colores_pie[orig_i]
            pie.slices[slot].strokeColor = white
            pie.slices[slot].strokeWidth = 1.5
        dwg.add(pie)
        centro_x, centro_y = pie.x + pie.width / 2, pie.y + pie.height / 2
        dwg.add(Circle(centro_x, centro_y, 1.85 * cm, fillColor=white, strokeColor=white))
        dwg.add(String(centro_x, centro_y + 4, str(total_informes), fontName="Helvetica-Bold", fontSize=20, fillColor=C_TEXT, textAnchor="middle"))
        dwg.add(String(centro_x, centro_y - 14, "informes totales", fontName="Helvetica", fontSize=8, fillColor=C_SOFT, textAnchor="middle"))
        story.append(dwg)

        leyenda_cells = []
        for lbl, val, col in [("Enviados", protocolos_env, C_OK), ("Pendientes", protocolos_pend, C_BAD)]:
            leyenda_cells.append(Table([[""]], colWidths=[9], rowHeights=[9], style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), col)])))
            leyenda_cells.append(Paragraph(f"{lbl} ({val})", st_td_soft))
        leyenda = Table([leyenda_cells], colWidths=None)
        leyenda.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3)]))
        story.append(Table([[leyenda]], colWidths=[fw], style=TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")])))
        story.append(Spacer(1, 4))

        # ── Top sucursales con más incidencias pendientes ───────────────────
        story.append(Paragraph("SUCURSALES CON MÁS INCIDENCIAS PENDIENTES", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Las sucursales listadas abajo son las que más incidencias tienen esperando gestión. Mientras más larga "
            "la barra, más incidencias sin resolver tiene esa sucursal.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        if top_clientes:
            nombres_suc = [c["cliente"][:38] for c in top_clientes]
            valores_suc = [c["cantidad"] for c in top_clientes]
            bar_h = max(3.2 * cm, len(nombres_suc) * 0.55 * cm)
            bdwg = Drawing(fw, bar_h)
            chart = HorizontalBarChart()
            chart.x, chart.y = 5.6 * cm, 6
            chart.width, chart.height = fw - 6.4 * cm, bar_h - 16
            chart.data = [valores_suc]
            chart.categoryAxis.categoryNames = nombres_suc
            chart.categoryAxis.labels.fontName = "Helvetica"
            chart.categoryAxis.labels.fontSize = 7
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(valores_suc + [1]) * 1.15
            chart.valueAxis.labels.fontName = "Helvetica"
            chart.valueAxis.labels.fontSize = 6.5
            chart.bars[0].fillColor = C_ORDK
            chart.barLabels.fontName = "Helvetica-Bold"
            chart.barLabels.fontSize = 7.5
            chart.barLabelFormat = "%d"
            chart.barLabels.dx = 14
            chart.categoryAxis.strokeColor = C_BORDER
            chart.valueAxis.strokeColor = C_BORDER
            bdwg.add(chart)
            story.append(bdwg)
        else:
            story.append(Paragraph("No hay incidencias pendientes en este período.", st_td_soft))

        # ── Página 2: tipos de incidencia + tabla detallada ─────────────────
        story.append(PageBreak())
        story.append(Paragraph("TIPOS DE INCIDENCIA MÁS FRECUENTES", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Qué tipo de problema es el que más se repite entre las incidencias pendientes.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        if por_tipo:
            nombres_tipo = [t["problema"][:38] for t in por_tipo]
            valores_tipo = [t["cantidad"] for t in por_tipo]
            bar_h2 = max(2.6 * cm, len(nombres_tipo) * 0.55 * cm)
            bdwg2 = Drawing(fw, bar_h2)
            chart2 = HorizontalBarChart()
            chart2.x, chart2.y = 5.6 * cm, 6
            chart2.width, chart2.height = fw - 6.4 * cm, bar_h2 - 16
            chart2.data = [valores_tipo]
            chart2.categoryAxis.categoryNames = nombres_tipo
            chart2.categoryAxis.labels.fontName = "Helvetica"
            chart2.categoryAxis.labels.fontSize = 7
            chart2.valueAxis.valueMin = 0
            chart2.valueAxis.valueMax = max(valores_tipo + [1]) * 1.15
            chart2.valueAxis.labels.fontName = "Helvetica"
            chart2.valueAxis.labels.fontSize = 6.5
            chart2.bars[0].fillColor = C_WARN
            chart2.barLabels.fontName = "Helvetica-Bold"
            chart2.barLabels.fontSize = 7.5
            chart2.barLabelFormat = "%d"
            chart2.barLabels.dx = 14
            chart2.categoryAxis.strokeColor = C_BORDER
            chart2.valueAxis.strokeColor = C_BORDER
            bdwg2.add(chart2)
            story.append(bdwg2)
        else:
            story.append(Paragraph("No hay incidencias pendientes en este período.", st_td_soft))

        story.append(Spacer(1, 10))
        story.append(Paragraph("INCIDENCIAS PENDIENTES MÁS ANTIGUAS (TOP 10)", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Las incidencias que llevan más tiempo sin gestionarse. Son las que deberían priorizarse primero.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        filas_tabla = [[
            Paragraph("ODT", st_th), Paragraph("CLIENTE / SUCURSAL", st_th),
            Paragraph("PROBLEMA", st_th), Paragraph("DÍAS SIN GESTIÓN", st_th),
        ]]
        for it in top_antiguas:
            dias = it.get("dias", 0)
            color_dias = _C_BAD if dias >= 60 else (_C_WARN if dias >= 30 else _C_OK)
            filas_tabla.append([
                Paragraph(it.get("odt", "") or "—", st_td),
                Paragraph(it.get("cliente", "") or "—", st_td),
                Paragraph(it.get("problema", "") or "—", st_td_soft),
                Paragraph(f'<font color="{color_dias}"><b>{dias}</b></font>', st_td),
            ])

        if top_antiguas:
            tabla = Table(filas_tabla, colWidths=[fw * 0.14, fw * 0.38, fw * 0.30, fw * 0.18], repeatRows=1)
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, C_BG]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(tabla)
        else:
            story.append(Paragraph("No hay incidencias pendientes registradas para este período.", st_body))

        # ── Conclusión ───────────────────────────────────────────────────
        story.append(Spacer(1, 16))
        story.append(Paragraph("CONCLUSIÓN", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))

        if total_pendientes == 0:
            texto_conclusion = (
                "El área de Coordinación con Cliente está al día: no hay incidencias pendientes de gestión en el "
                "período seleccionado."
            )
        else:
            nivel = "crítico" if antiguedad_max >= 60 else ("de atención" if antiguedad_max >= 30 else "manejable")
            frase_sucursal = ""
            if top_clientes:
                top_nombre, top_cant = top_clientes[0]["cliente"], top_clientes[0]["cantidad"]
                frase_sucursal = f" La sucursal con más incidencias pendientes es <b>{top_nombre}</b>, con {top_cant} caso(s)."
            frase_tipo = ""
            if por_tipo:
                frase_tipo = f" El tipo de problema más frecuente es <b>{por_tipo[0]['problema']}</b>, con {por_tipo[0]['cantidad']} caso(s)."
            texto_conclusion = (
                f"El estado actual del área se considera <b>{nivel}</b>: hay {total_pendientes} incidencias pendientes "
                f"con un promedio de {antiguedad_prom} días sin gestión (máximo {antiguedad_max} días).{frase_sucursal}"
                f"{frase_tipo} Se recomienda priorizar las incidencias con mayor antigüedad, listadas en la tabla "
                "de esta página, para evitar que sigan acumulando días sin respuesta."
            )
        story.append(Paragraph(texto_conclusion, st_body))

        # ── Observaciones (editable) ────────────────────────────────────────
        story.append(Spacer(1, 16))
        story.append(Paragraph("OBSERVACIONES", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Espacio editable: se puede escribir directamente sobre este PDF (Adobe Reader, Vista Previa de Mac "
            "u otro lector compatible con formularios), sin necesidad de exportar a Word.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))
        story.append(_CampoTextoEditable(fw, 6 * cm, "observaciones_coordinacion"))

        doc.build(story)
        buf.seek(0)
        return buf.getvalue()

    def generar_informe_servicio_pdf(self, desde: date | None = None, hasta: date | None = None) -> bytes:
        """Genera el PDF "Informe de Gestion - Servicio Tecnico" reflejando
        los mismos datos y criterios que dashboard_servicio.html para el
        rango desde/hasta seleccionado: KPIs, ODTs por mes (finalizadas vs
        pendientes/en proceso), clientes con mas ODTs (siempre historico,
        igual que el dashboard), instalacion de camaras (siempre estado
        actual, igual que el dashboard), y ademas dos secciones de calidad
        calculadas con datos reales (no las tarjetas "DEMO" del dashboard,
        que usan datos de ejemplo): recurrencia de visitas por cliente +
        tipo de problema, y clientes con camaras instaladas que reportaron
        desconexiones en el periodo."""
        import io
        import json as _json
        import re as _re

        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor, white
        from reportlab.platypus import (
            BaseDocTemplate, Frame, PageTemplate, PageBreak,
            Table, TableStyle, Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_JUSTIFY
        from reportlab.graphics.shapes import Drawing, Circle, String
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart

        _C_DARK, _C_ORANGE, _C_ORDK = "#0b1424", "#5a8cff", "#1d4ed8"
        _C_BG, _C_BORDER, _C_TEXT, _C_SOFT, _C_GREY = "#f7f8fa", "#e5e7eb", "#111827", "#4b5563", "#9ca3af"
        _C_OK, _C_WARN, _C_BAD = "#1e9c83", "#d97706", "#c0392b"
        C_DARK, C_ORANGE, C_ORDK = HexColor(_C_DARK), HexColor(_C_ORANGE), HexColor(_C_ORDK)
        C_BG, C_BORDER, C_TEXT, C_SOFT, C_GREY = (
            HexColor(_C_BG), HexColor(_C_BORDER), HexColor(_C_TEXT), HexColor(_C_SOFT), HexColor(_C_GREY)
        )
        C_OK, C_WARN, C_BAD = HexColor(_C_OK), HexColor(_C_WARN), HexColor(_C_BAD)

        # ── Helpers de negocio (identicos a los de dashboard_servicio.html) ──
        def _estado_norm(estado: Any) -> str:
            e = str(estado or "").lower()
            if "termin" in e or "final" in e or "cerr" in e:
                return "finalizada"
            if "proceso" in e:
                return "proceso"
            if "pend" in e:
                return "pendiente"
            return "otro"

        def _is_final(estado: Any) -> bool:
            return _estado_norm(estado) == "finalizada"

        def _dias_cierre(fecha_registro, fecha_cierre, dias_ejecucion) -> float | None:
            if fecha_registro and fecha_cierre and fecha_cierre >= fecha_registro:
                return (fecha_cierre - fecha_registro).total_seconds() / 86400
            if isinstance(dias_ejecucion, (int, float)) and dias_ejecucion >= 0:
                return float(dias_ejecucion)
            return None

        def _dias_abierta(fecha_registro, ahora) -> float | None:
            if not fecha_registro:
                return None
            return (ahora - fecha_registro).total_seconds() / 86400

        def _sla_status(r: dict, sla_dias: int, ahora) -> str | None:
            if _is_final(r.get("estado")):
                d = _dias_cierre(r.get("fecha_registro"), r.get("fecha_cierre"), r.get("dias_ejecucion"))
                if d is None:
                    return None
                return "cumplida" if d <= sla_dias else "incumplida"
            d = _dias_abierta(r.get("fecha_registro"), ahora)
            if d is None:
                return None
            return "incumplida" if d > sla_dias else "en_plazo"

        # ── Datos base: mismo filtro que /api/servicio/kpis-data ──
        registro_cols = (
            Registro.odt.label("odt"),
            Registro.fecha_registro.label("fecha_registro"),
            Registro.fecha_cierre.label("fecha_cierre"),
            Registro.cliente.label("cliente"),
            Registro.problema.label("problema"),
            Registro.estado.label("estado"),
            Registro.tecnicos.label("tecnicos"),
            Registro.dias_ejecucion.label("dias_ejecucion"),
            Registro.porcentaje_avance.label("porcentaje_avance"),
        )
        try:
            all_regs = [
                dict(r)
                for r in self.db.execute(
                    select(*registro_cols)
                    .where(func.lower(func.trim(Registro.estado)) != "repetida")
                    .where(Registro.derivacion.ilike("%servicio t%"))
                    .where(~Registro.problema.ilike("%mantenc%"))
                    .order_by(Registro.fecha_registro.desc())
                ).mappings().all()
            ]
        except Exception:
            self.db.rollback()
            all_regs = []

        ahora = datetime.now()
        sla_dias = getattr(settings, "servicio_sla_dias", None) or 7

        regs_filtrados = [
            r for r in all_regs
            if (not desde or (r["fecha_registro"] and r["fecha_registro"].date() >= desde))
            and (not hasta or (r["fecha_registro"] and r["fecha_registro"].date() <= hasta))
        ]

        # ── KPIs (identicos a kActivas/kSla/kTiempo/kTasa) ──
        fin = [r for r in regs_filtrados if _is_final(r["estado"])]
        pend = [r for r in regs_filtrados if _estado_norm(r["estado"]) == "pendiente"]
        proc = [r for r in regs_filtrados if _estado_norm(r["estado"]) == "proceso"]
        odts_activas = len(pend) + len(proc)

        sla_eval = [s for s in (_sla_status(r, sla_dias, ahora) for r in regs_filtrados) if s in ("cumplida", "incumplida")]
        sla_ok = len([s for s in sla_eval if s == "cumplida"])
        cumplimiento_sla_pct = round(sla_ok / len(sla_eval) * 100) if sla_eval else None

        tiempos = [d for d in (_dias_cierre(r["fecha_registro"], r["fecha_cierre"], r["dias_ejecucion"]) for r in fin) if d is not None]
        tiempo_prom = round(sum(tiempos) / len(tiempos), 1) if tiempos else None

        tasa_resolucion_pct = round(len(fin) / len(regs_filtrados) * 100) if regs_filtrados else None

        # ── ODTs por mes / dia (finalizadas vs pendientes-o-en-proceso) ──
        _MESES_ES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]

        def _fmt_mes(d: date) -> str:
            return f"{_MESES_ES[d.month - 1]}-{str(d.year)[2:]}"

        buckets: list[dict] = []
        modo_diario = False
        if desde and hasta:
            diff_dias = (hasta - desde).days
            if 0 <= diff_dias <= 62:
                modo_diario = True
                cursor = desde
                while cursor <= hasta:
                    buckets.append({"y": cursor.year, "m": cursor.month, "d": cursor.day, "label": cursor.strftime("%d/%m")})
                    cursor += timedelta(days=1)
            else:
                cursor = date(desde.year, desde.month, 1)
                limite = date(hasta.year, hasta.month, 1)
                while cursor <= limite and len(buckets) < 24:
                    buckets.append({"y": cursor.year, "m": cursor.month, "label": _fmt_mes(cursor)})
                    cursor = date(cursor.year + (1 if cursor.month == 12 else 0), 1 if cursor.month == 12 else cursor.month + 1, 1)
        else:
            hoy = date.today()
            cursor = date(hoy.year, hoy.month, 1)
            for i in range(11, -1, -1):
                mes = cursor.month - i
                anio = cursor.year
                while mes <= 0:
                    mes += 12
                    anio -= 1
                buckets.append({"y": anio, "m": mes, "label": _fmt_mes(date(anio, mes, 1))})

        def _en_bucket(fecha, b: dict) -> bool:
            if not fecha:
                return False
            if modo_diario:
                return fecha.year == b["y"] and fecha.month == b["m"] and fecha.day == b["d"]
            return fecha.year == b["y"] and fecha.month == b["m"]

        for b in buckets:
            b["finalizadas"] = sum(
                1 for r in all_regs
                if _is_final(r["estado"]) and _en_bucket((r["fecha_cierre"] or r["fecha_registro"]), b)
            )
            b["pendientes"] = sum(
                1 for r in all_regs
                if not _is_final(r["estado"]) and _en_bucket(r["fecha_registro"], b)
            )

        # ── Clientes con mas ODTs (SIEMPRE historico, igual que el dashboard) ──
        CLIENTES_EXCLUIDOS_ODT = {"mantenciones viña del mar"}
        por_cliente: dict[str, dict] = {}
        for r in all_regs:
            nombre = str(r["cliente"] or "Sin cliente").strip() or "Sin cliente"
            if nombre.lower() in CLIENTES_EXCLUIDOS_ODT:
                continue
            entry = por_cliente.setdefault(
                nombre,
                {"cliente": nombre, "total": 0, "tipos": {}, "estados": {"finalizada": 0, "proceso": 0, "pendiente": 0, "otro": 0}},
            )
            entry["total"] += 1
            tipo = str(r["problema"] or "Sin especificar").strip() or "Sin especificar"
            entry["tipos"][tipo] = entry["tipos"].get(tipo, 0) + 1
            entry["estados"][_estado_norm(r["estado"])] += 1
        clientes_top = sorted(
            [e for e in por_cliente.values() if e["total"] >= 2],
            key=lambda e: -e["total"],
        )[:10]

        def _semaforo(total: int) -> tuple[str, str]:
            if total >= 4:
                return "Rojo", _C_BAD
            if total == 3:
                return "Naranjo", _C_WARN
            return "Verde", _C_OK

        # ── Instalacion de camaras (SIEMPRE estado actual, igual que el dashboard) ──
        def _contar_camaras_registradas(raw: Any) -> int:
            if raw in (None, "", [], (), {}):
                return 0
            if isinstance(raw, (list, tuple, set)):
                items = list(raw)
            else:
                texto = str(raw).strip()
                if not texto:
                    return 0
                try:
                    parsed = _json.loads(texto)
                except Exception:
                    parsed = [p.strip() for p in _re.split(r"[,\n|;]+", texto) if p.strip()]
                items = parsed if isinstance(parsed, (list, tuple, set)) else ([parsed] if parsed not in (None, "") else [])
            total = 0
            for item in items:
                if isinstance(item, dict):
                    total += 1
                elif str(item or "").strip():
                    total += 1
            return total

        def _porcentaje_a_instaladas(raw: Any, total: int) -> int:
            if not raw or total <= 0:
                return 0
            texto = str(raw).strip().replace("%", "")
            if not texto:
                return 0
            try:
                pct = float(texto.replace(",", "."))
            except Exception:
                return 0
            if pct <= 0:
                return 0
            if pct > 100:
                pct = 100
            return max(0, min(total, int(round(total * pct / 100.0))))

        avance_por_odt: dict[str, dict] = {}
        for r in all_regs:
            odt_key = str(r["odt"] or "").strip().upper()
            if odt_key and odt_key not in avance_por_odt:
                avance_por_odt[odt_key] = r

        avance_por_odt_camaras: dict[str, Any] = {}
        try:
            for odt_row, pct_row in self.db.execute(
                select(Registro.odt, Registro.porcentaje_avance).where(
                    func.lower(func.trim(Registro.estado)) != "repetida"
                )
            ).all():
                odt_key = str(odt_row or "").strip().upper()
                if odt_key and odt_key not in avance_por_odt_camaras:
                    avance_por_odt_camaras[odt_key] = pct_row
        except Exception:
            self.db.rollback()

        avance_camaras: list[dict] = []
        try:
            avance_rows = self.db.execute(
                select(ServicioTecnicoVentaODT, VentaODS, SoporteTecnicoVentaODT)
                .outerjoin(VentaODS, VentaODS.codigo == ServicioTecnicoVentaODT.odt)
                .outerjoin(SoporteTecnicoVentaODT, SoporteTecnicoVentaODT.odt == VentaODS.codigo)
                .where(VentaODS.numero_camaras_instalar.is_not(None))
                .where(VentaODS.numero_camaras_instalar > 0)
                .order_by(VentaODS.created_at.desc())
            ).all()
            for st, v, sop in avance_rows:
                total = int(v.numero_camaras_instalar or 0)
                finalizada_db = bool(st and (st.instalacion_finalizada or st.finalizado))
                instaladas_registradas = _contar_camaras_registradas(getattr(sop, "camaras_registradas", None)) if sop else 0
                if not instaladas_registradas:
                    odt_key = str((st.odt if st else None) or v.codigo or "").strip().upper()
                    pct = avance_por_odt_camaras.get(odt_key)
                    if pct is None:
                        reg = avance_por_odt.get(odt_key)
                        pct = reg.get("porcentaje_avance") if reg else None
                    instaladas_registradas = _porcentaje_a_instaladas(pct, total)
                finalizada = finalizada_db or (total > 0 and instaladas_registradas >= total)
                instaladas = total if finalizada else min(total, instaladas_registradas)
                avance_camaras.append({
                    "ods": (st.odt if st else None) or v.codigo or "",
                    "cliente": v.nombre_sucursal or v.razon_social or "",
                    "camaras_total": total,
                    "camaras_instaladas": instaladas,
                    "camaras_pendientes": max(total - instaladas, 0),
                })
        except Exception:
            self.db.rollback()

        camaras_pendientes_odt = sorted(
            [c for c in avance_camaras if c["camaras_pendientes"] > 0],
            key=lambda c: -c["camaras_pendientes"],
        )
        camaras_instaladas_total = sum(c["camaras_instaladas"] for c in avance_camaras)
        camaras_pendientes_total = sum(c["camaras_pendientes"] for c in avance_camaras)

        def _normalizar_nombre(s: Any) -> str:
            return str(s or "").strip().lower()

        # ── Calidad de visita tecnica: recurrencia de visitas por cliente y
        # tipo de problema, dentro del rango del informe. No existe (todavia)
        # un identificador de camara por incidencia, por lo que se agrupa por
        # cliente + tipo de problema, el nivel de detalle real disponible.
        por_cliente_problema: dict[tuple[str, str], dict] = {}
        for r in regs_filtrados:
            cliente = str(r["cliente"] or "Sin cliente").strip() or "Sin cliente"
            problema = str(r["problema"] or "Sin especificar").strip() or "Sin especificar"
            key = (cliente, problema)
            entry = por_cliente_problema.setdefault(
                key, {"cliente": cliente, "problema": problema, "visitas": 0, "tecnicos": set()}
            )
            entry["visitas"] += 1
            if r.get("tecnicos"):
                entry["tecnicos"].add(str(r["tecnicos"]).strip())
        # Solo interesa mostrar recurrencia real: 3 visitas o mas a la misma
        # combinacion cliente+problema (mismo umbral que el dashboard). 2 o
        # menos no se considera una alerta.
        recurrencia_visitas = sorted(
            [e for e in por_cliente_problema.values() if e["visitas"] >= 3],
            key=lambda e: -e["visitas"],
        )

        def _nivel_recurrencia(n: int) -> tuple[str, str]:
            if n >= 4:
                return "Rojo", _C_BAD
            return "Naranjo", _C_WARN

        # ── Calidad de instalacion segun desconexiones: clientes con camaras
        # ya instaladas que tuvieron incidencias de "Desconexión" en el rango
        # del informe. El cruce es por nombre de cliente/sucursal (tampoco hay
        # ID de camara por incidencia todavia), no por camara individual.
        clientes_con_camaras: dict[str, dict] = {}
        for c in avance_camaras:
            if c["camaras_instaladas"] <= 0:
                continue
            nombre = str(c["cliente"] or "").strip()
            if not nombre:
                continue
            key = _normalizar_nombre(nombre)
            entry = clientes_con_camaras.setdefault(key, {"cliente": nombre, "camaras_instaladas": 0})
            entry["camaras_instaladas"] += c["camaras_instaladas"]

        desconexiones_por_cliente: dict[str, int] = {}
        for r in regs_filtrados:
            if "desconex" not in str(r["problema"] or "").lower():
                continue
            nombre = str(r["cliente"] or "").strip()
            if not nombre:
                continue
            key = _normalizar_nombre(nombre)
            desconexiones_por_cliente[key] = desconexiones_por_cliente.get(key, 0) + 1

        calidad_instalacion = sorted(
            [
                {
                    "cliente": info["cliente"],
                    "camaras_instaladas": info["camaras_instaladas"],
                    "desconexiones": desconexiones_por_cliente.get(key, 0),
                }
                for key, info in clientes_con_camaras.items()
                if desconexiones_por_cliente.get(key, 0) > 0
            ],
            key=lambda c: -c["desconexiones"],
        )

        def _nivel_desconexion(n: int) -> tuple[str, str]:
            if n >= 3:
                return "Alto", _C_BAD
            if n == 2:
                return "Medio", _C_WARN
            return "Bajo", _C_OK

        # ── Encabezado / metadatos del documento ──
        def _fmt(d: date | None) -> str:
            return d.strftime("%d/%m/%Y") if d else ""

        if desde and hasta:
            rango_txt = f"Período: {_fmt(desde)} al {_fmt(hasta)}"
        elif desde:
            rango_txt = f"Período: desde el {_fmt(desde)}"
        elif hasta:
            rango_txt = f"Período: hasta el {_fmt(hasta)}"
        else:
            rango_txt = "Período: histórico completo (sin filtro de fecha)"

        fecha_emision = ahora.strftime("%d/%m/%Y %H:%M")
        titulo_hdr = "INFORME DE GESTIÓN — SERVICIO TÉCNICO"
        subtitulo_hdr = rango_txt

        W, H = A4
        pad = 1.4 * cm
        HEADER_H = 2.7 * cm
        ORANGE_H = 5
        FOOTER_H = 1.0 * cm
        BODY_TOP = HEADER_H + ORANGE_H + 12
        BODY_BOT = FOOTER_H + 8
        fw = W - 2 * pad

        _atc_root = Path(__file__).resolve().parents[2]
        logo_path = _atc_root / "ATC" / "static" / "img" / "logo-atc.png"
        if not logo_path.exists():
            logo_path = _atc_root / "static" / "img" / "logo-atc.png"
        logo_w, logo_h = 2.8 * cm, 1.4 * cm

        def draw_page(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(C_DARK)
            canvas.rect(0, H - HEADER_H, W, HEADER_H, fill=1, stroke=0)
            if logo_path.exists():
                try:
                    canvas.drawImage(
                        str(logo_path),
                        pad, H - HEADER_H + (HEADER_H - logo_h) / 2,
                        width=logo_w, height=logo_h,
                        preserveAspectRatio=True, mask="auto",
                    )
                except Exception:
                    pass
            tx = pad + logo_w + 0.5 * cm
            canvas.setFillColor(white)
            canvas.setFont("Helvetica-Bold", 13)
            canvas.drawString(tx, H - HEADER_H + 1.35 * cm, titulo_hdr)
            canvas.setFillColor(HexColor("#bfdbfe"))
            canvas.setFont("Helvetica", 8.5)
            canvas.drawString(tx, H - HEADER_H + 0.75 * cm, subtitulo_hdr)
            canvas.setFillColor(C_ORANGE)
            canvas.rect(0, H - HEADER_H - ORANGE_H, W, ORANGE_H, fill=1, stroke=0)
            canvas.setFillColor(C_DARK)
            canvas.rect(0, 0, W, FOOTER_H, fill=1, stroke=0)
            canvas.setFillColor(C_GREY)
            canvas.setFont("Helvetica", 7)
            canvas.drawCentredString(
                W / 2, FOOTER_H / 2 - 3,
                f"Documento generado automáticamente  ·  Alguien Te Cuida  ·  {fecha_emision}",
            )
            canvas.setFont("Helvetica", 7)
            canvas.drawRightString(W - pad, FOOTER_H / 2 - 3, f"Página {doc.page}")
            canvas.restoreState()

        frame = Frame(
            pad, BODY_BOT, fw, H - BODY_TOP - BODY_BOT,
            leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
        )
        page_tmpl = PageTemplate(id="main", frames=[frame], onPage=draw_page)
        buf = io.BytesIO()
        doc = BaseDocTemplate(
            buf, pagesize=A4, pageTemplates=[page_tmpl],
            leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0,
            title=titulo_hdr, author="Alguien Te Cuida",
        )

        st_kpi_num = ParagraphStyle("kpiNumS", fontName="Helvetica-Bold", fontSize=20, textColor=C_TEXT, leading=22, alignment=1)
        st_kpi_lbl = ParagraphStyle("kpiLblS", fontName="Helvetica-Bold", fontSize=7, textColor=C_SOFT, leading=9, alignment=1)
        st_sec = ParagraphStyle("secS", fontName="Helvetica-Bold", fontSize=11, textColor=C_ORDK, leading=14, spaceBefore=14, spaceAfter=6)
        st_body = ParagraphStyle("bodyS", fontName="Helvetica", fontSize=9.5, textColor=C_SOFT, leading=14.5, alignment=TA_JUSTIFY, spaceAfter=6)
        st_th = ParagraphStyle("thS", fontName="Helvetica-Bold", fontSize=8, textColor=white, leading=10)
        st_td = ParagraphStyle("tdS", fontName="Helvetica", fontSize=8, textColor=C_TEXT, leading=11)
        st_td_soft = ParagraphStyle("tdSoftS", fontName="Helvetica", fontSize=7.5, textColor=C_SOFT, leading=10)

        story: list = []

        # ── KPIs ──────────────────────────────────────────────────────────
        def kpi_card(numero: str, etiqueta: str, color) -> Table:
            t = Table([[Paragraph(numero, st_kpi_num)], [Paragraph(etiqueta, st_kpi_lbl)]], colWidths=[fw / 4 - 8])
            t.setStyle(TableStyle([
                ("TOPPADDING", (0, 0), (-1, 0), 12), ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
                ("TOPPADDING", (0, 1), (-1, 1), 2), ("BOTTOMPADDING", (0, 1), (-1, 1), 12),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
                ("LINEABOVE", (0, 0), (-1, 0), 3, color),
                ("BACKGROUND", (0, 0), (-1, -1), white),
            ]))
            return t

        color_sla = C_OK if (cumplimiento_sla_pct or 0) >= 80 else (C_WARN if (cumplimiento_sla_pct or 0) >= 50 else C_BAD)
        color_tasa = C_OK if (tasa_resolucion_pct or 0) >= 80 else (C_WARN if (tasa_resolucion_pct or 0) >= 50 else C_BAD)
        kpis = Table(
            [[
                kpi_card(str(odts_activas), "ODTS ACTIVAS\n(PENDIENTES + EN PROCESO)", C_ORDK),
                kpi_card(f"{cumplimiento_sla_pct}%" if cumplimiento_sla_pct is not None else "—", "CUMPLIMIENTO\nSLA", color_sla),
                kpi_card(f"{tiempo_prom}" if tiempo_prom is not None else "—", "TIEMPO PROMEDIO\nDE CIERRE (DÍAS)", C_WARN),
                kpi_card(f"{tasa_resolucion_pct}%" if tasa_resolucion_pct is not None else "—", "TASA DE\nRESOLUCIÓN", color_tasa),
            ]],
            colWidths=[fw / 4] * 4,
        )
        kpis.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4)]))
        story.append(kpis)
        story.append(Spacer(1, 14))

        # ── Resumen general ──────────────────────────────────────────────
        story.append(Paragraph("RESUMEN GENERAL", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        if regs_filtrados:
            frase_sla = (
                f"El cumplimiento de SLA (respuesta dentro de {sla_dias} días) es de <b>{cumplimiento_sla_pct}%</b>."
                if cumplimiento_sla_pct is not None else
                "No hay suficientes datos para calcular el cumplimiento de SLA en este período."
            )
            story.append(Paragraph(
                f"El área de Servicio Técnico registró <b>{len(regs_filtrados)} ODTs</b> en el período seleccionado, "
                f"de las cuales <b>{len(fin)} fueron finalizadas</b> y <b>{odts_activas} siguen activas</b> "
                f"({len(pend)} pendientes y {len(proc)} en proceso). {frase_sla} "
                + (f"El tiempo promedio de cierre de las ODTs finalizadas fue de <b>{tiempo_prom} días</b>. " if tiempo_prom is not None else "")
                + f"La tasa de resolución general del período es de <b>{tasa_resolucion_pct}%</b>."
                if tasa_resolucion_pct is not None else "",
                st_body,
            ))
        else:
            story.append(Paragraph(
                "No se registraron ODTs de Servicio Técnico en el período seleccionado.",
                st_body,
            ))
        story.append(Spacer(1, 6))

        # ── ODTs por mes/dia: finalizadas vs pendientes/en proceso ──────────
        story.append(Paragraph("ODTS " + ("POR DÍA" if modo_diario else "POR MES"), st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Comparación entre ODTs finalizadas y ODTs que siguen pendientes o en proceso, agrupadas por "
            + ("día" if modo_diario else "mes") + " de registro (independiente del filtro de fecha de los KPIs: "
            "sin filtro se muestra el último año completo).",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        if buckets:
            bar_h = 6.0 * cm
            bdwg = Drawing(fw, bar_h)
            chart = VerticalBarChart()
            chart.x, chart.y = 1.4 * cm, 1.2 * cm
            chart.width, chart.height = fw - 2.0 * cm, bar_h - 2.0 * cm
            chart.data = [[b["finalizadas"] for b in buckets], [b["pendientes"] for b in buckets]]
            chart.categoryAxis.categoryNames = [b["label"] for b in buckets]
            chart.categoryAxis.labels.fontName = "Helvetica"
            chart.categoryAxis.labels.fontSize = 6
            chart.categoryAxis.labels.angle = 90 if len(buckets) > 12 else 0
            chart.categoryAxis.labels.dy = -18 if len(buckets) > 12 else -2
            chart.valueAxis.valueMin = 0
            maximo = max([b["finalizadas"] for b in buckets] + [b["pendientes"] for b in buckets] + [1])
            chart.valueAxis.valueMax = maximo * 1.15
            chart.valueAxis.labels.fontName = "Helvetica"
            chart.valueAxis.labels.fontSize = 6.5
            chart.bars[0].fillColor = C_OK
            chart.bars[1].fillColor = C_BAD
            chart.groupSpacing = 8
            chart.barSpacing = 1
            chart.categoryAxis.strokeColor = C_BORDER
            chart.valueAxis.strokeColor = C_BORDER
            bdwg.add(chart)
            story.append(bdwg)

            leyenda_cells = []
            for lbl, col in [("Finalizadas", C_OK), ("Pendientes / En proceso", C_BAD)]:
                leyenda_cells.append(Table([[""]], colWidths=[9], rowHeights=[9], style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), col)])))
                leyenda_cells.append(Paragraph(lbl, st_td_soft))
            leyenda = Table([leyenda_cells], colWidths=None)
            leyenda.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3)]))
            story.append(Table([[leyenda]], colWidths=[fw], style=TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")])))
            story.append(Spacer(1, 8))

            filas_odt_mes = [[
                Paragraph(("DÍA" if modo_diario else "MES"), st_th), Paragraph("FINALIZADAS", st_th),
                Paragraph("PENDIENTES / EN PROCESO", st_th), Paragraph("TOTAL", st_th),
            ]]
            for b in buckets:
                filas_odt_mes.append([
                    Paragraph(b["label"], st_td),
                    Paragraph(str(b["finalizadas"]), st_td),
                    Paragraph(str(b["pendientes"]), st_td),
                    Paragraph(str(b["finalizadas"] + b["pendientes"]), st_td),
                ])
            tabla_odt_mes = Table(filas_odt_mes, colWidths=[fw * 0.25, fw * 0.25, fw * 0.30, fw * 0.20], repeatRows=1)
            tabla_odt_mes.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, C_BG]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(tabla_odt_mes)
        else:
            story.append(Paragraph("No hay datos suficientes para construir este gráfico.", st_td_soft))

        # ── Página 2: Clientes con mas ODTs ─────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("CLIENTES CON MÁS ODTS (TOP 10 · HISTÓRICO)", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Ranking histórico completo (no depende del filtro de fecha, igual que en el dashboard) de los "
            "clientes/sucursales con más ODTs de servicio técnico. El semáforo indica el nivel de recurrencia: "
            "verde 2 ODTs, naranjo 3 ODTs, rojo 4 o más ODTs.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        if clientes_top:
            nombres_cli = [c["cliente"][:38] for c in clientes_top]
            valores_cli = [c["total"] for c in clientes_top]
            bar_h3 = max(3.2 * cm, len(nombres_cli) * 0.55 * cm)
            bdwg3 = Drawing(fw, bar_h3)
            chart3 = HorizontalBarChart()
            chart3.x, chart3.y = 5.6 * cm, 6
            chart3.width, chart3.height = fw - 6.4 * cm, bar_h3 - 16
            chart3.data = [valores_cli]
            chart3.categoryAxis.categoryNames = nombres_cli
            chart3.categoryAxis.labels.fontName = "Helvetica"
            chart3.categoryAxis.labels.fontSize = 7
            chart3.valueAxis.valueMin = 0
            chart3.valueAxis.valueMax = max(valores_cli + [1]) * 1.15
            chart3.valueAxis.labels.fontName = "Helvetica"
            chart3.valueAxis.labels.fontSize = 6.5
            chart3.bars[0].fillColor = C_ORDK
            chart3.barLabels.fontName = "Helvetica-Bold"
            chart3.barLabels.fontSize = 7.5
            chart3.barLabelFormat = "%d"
            chart3.barLabels.dx = 14
            chart3.categoryAxis.strokeColor = C_BORDER
            chart3.valueAxis.strokeColor = C_BORDER
            bdwg3.add(chart3)
            story.append(bdwg3)
            story.append(Spacer(1, 8))

            filas_cli = [[
                Paragraph("CLIENTE / SUCURSAL", st_th), Paragraph("TOTAL", st_th),
                Paragraph("TIPO DE INCIDENCIA PRINCIPAL", st_th), Paragraph("FINALIZADAS", st_th),
                Paragraph("ACTIVAS", st_th), Paragraph("SEMÁFORO", st_th),
            ]]
            for c in clientes_top:
                tipo_principal = max(c["tipos"].items(), key=lambda kv: kv[1]) if c["tipos"] else ("—", 0)
                activas_cli = c["estados"]["proceso"] + c["estados"]["pendiente"]
                sem_txt, sem_color_hex = _semaforo(c["total"])
                filas_cli.append([
                    Paragraph(c["cliente"], st_td),
                    Paragraph(str(c["total"]), st_td),
                    Paragraph(f"{tipo_principal[0]} ({tipo_principal[1]})", st_td_soft),
                    Paragraph(str(c["estados"]["finalizada"]), st_td),
                    Paragraph(str(activas_cli), st_td),
                    Paragraph(f'<font color="{sem_color_hex}"><b>{sem_txt}</b></font>', st_td),
                ])
            tabla_cli = Table(filas_cli, colWidths=[fw * 0.26, fw * 0.08, fw * 0.30, fw * 0.12, fw * 0.10, fw * 0.14], repeatRows=1)
            tabla_cli.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, C_BG]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(tabla_cli)
        else:
            story.append(Paragraph("No hay clientes con 2 o más ODTs registradas.", st_td_soft))

        # ── Página 3: Instalación de cámaras ─────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("INSTALACIÓN DE CÁMARAS (ESTADO ACTUAL)", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Avance de instalación de cámaras en ODS de venta con cámaras contratadas. Esta sección refleja el "
            "estado actual (no depende del filtro de fecha, igual que en el dashboard).",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        cam_kpis = Table(
            [[
                kpi_card(str(len(camaras_pendientes_odt)), "ODS CON CÁMARAS\nPENDIENTES", C_BAD),
                kpi_card(str(camaras_pendientes_total), "CÁMARAS\nPENDIENTES", C_WARN),
                kpi_card(str(camaras_instaladas_total), "CÁMARAS\nINSTALADAS", C_OK),
            ]],
            colWidths=[fw / 3] * 3,
        )
        cam_kpis.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4)]))
        story.append(cam_kpis)
        story.append(Spacer(1, 12))

        total_camaras_contratadas = camaras_instaladas_total + camaras_pendientes_total
        if total_camaras_contratadas:
            dwg2 = Drawing(fw, 5.0 * cm)
            pie2 = Pie()
            pie2.x, pie2.y = fw / 2 - 3.0 * cm, 0.2 * cm
            pie2.width, pie2.height = 4.8 * cm, 4.8 * cm
            valores_pie2 = [camaras_instaladas_total, camaras_pendientes_total]
            colores_pie2 = [C_OK, C_BAD]
            idx_no_cero2 = [i for i, v in enumerate(valores_pie2) if v > 0]
            pie2.data = [valores_pie2[i] for i in idx_no_cero2] or [1]
            for slot, orig_i in enumerate(idx_no_cero2):
                pie2.slices[slot].fillColor = colores_pie2[orig_i]
                pie2.slices[slot].strokeColor = white
                pie2.slices[slot].strokeWidth = 1.5
            dwg2.add(pie2)
            cx2, cy2 = pie2.x + pie2.width / 2, pie2.y + pie2.height / 2
            dwg2.add(Circle(cx2, cy2, 1.65 * cm, fillColor=white, strokeColor=white))
            dwg2.add(String(cx2, cy2 + 4, str(total_camaras_contratadas), fontName="Helvetica-Bold", fontSize=18, fillColor=C_TEXT, textAnchor="middle"))
            dwg2.add(String(cx2, cy2 - 13, "cámaras contratadas", fontName="Helvetica", fontSize=7.5, fillColor=C_SOFT, textAnchor="middle"))
            story.append(dwg2)

            leyenda2_cells = []
            for lbl, val, col in [("Instaladas", camaras_instaladas_total, C_OK), ("Pendientes", camaras_pendientes_total, C_BAD)]:
                leyenda2_cells.append(Table([[""]], colWidths=[9], rowHeights=[9], style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), col)])))
                leyenda2_cells.append(Paragraph(f"{lbl} ({val})", st_td_soft))
            leyenda2 = Table([leyenda2_cells], colWidths=None)
            leyenda2.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3)]))
            story.append(Table([[leyenda2]], colWidths=[fw], style=TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")])))
            story.append(Spacer(1, 8))

        story.append(Paragraph("ODS CON CÁMARAS PENDIENTES DE INSTALAR (DE MÁS A MENOS)", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        if camaras_pendientes_odt:
            filas_cam = [[
                Paragraph("ODS", st_th), Paragraph("CLIENTE / SUCURSAL", st_th),
                Paragraph("INSTALADAS", st_th), Paragraph("TOTAL", st_th), Paragraph("PENDIENTES", st_th),
            ]]
            for c in camaras_pendientes_odt:
                filas_cam.append([
                    Paragraph(c["ods"] or "—", st_td),
                    Paragraph(c["cliente"] or "—", st_td_soft),
                    Paragraph(str(c["camaras_instaladas"]), st_td),
                    Paragraph(str(c["camaras_total"]), st_td),
                    Paragraph(f'<font color="{_C_BAD}"><b>{c["camaras_pendientes"]}</b></font>', st_td),
                ])
            tabla_cam = Table(filas_cam, colWidths=[fw * 0.14, fw * 0.46, fw * 0.14, fw * 0.12, fw * 0.14], repeatRows=1)
            tabla_cam.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, C_BG]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(tabla_cam)
        else:
            story.append(Paragraph("No hay ODS con cámaras pendientes de instalar.", st_td_soft))

        # ── Página 4: Calidad de visita técnica ─────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("CALIDAD DE VISITA TÉCNICA (RECURRENCIA POR CLIENTE)", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Clientes/sucursales que recibieron 3 o más visitas por el mismo tipo de problema en el período del "
            "informe. Mientras más visitas repetidas por el mismo motivo, peor es la calidad de la solución "
            "entregada en terreno: la falla vuelve a ocurrir y el cliente vuelve a llamar. El sistema actual no "
            "identifica cámaras individuales por incidencia, por lo que la recurrencia se mide por cliente + tipo "
            "de problema (el nivel de detalle más fino disponible hoy). Semáforo: naranjo 3 visitas, rojo 4 o más.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        if recurrencia_visitas:
            filas_rec = [[
                Paragraph("CLIENTE / SUCURSAL", st_th), Paragraph("TIPO DE PROBLEMA", st_th),
                Paragraph("VISITAS", st_th), Paragraph("TÉCNICOS INVOLUCRADOS", st_th), Paragraph("NIVEL", st_th),
            ]]
            for e in recurrencia_visitas:
                nivel_txt, nivel_hex = _nivel_recurrencia(e["visitas"])
                tecnicos_txt = ", ".join(sorted(e["tecnicos"])[:3]) or "—"
                if len(e["tecnicos"]) > 3:
                    tecnicos_txt += f" (+{len(e['tecnicos']) - 3})"
                filas_rec.append([
                    Paragraph(e["cliente"], st_td),
                    Paragraph(e["problema"], st_td_soft),
                    Paragraph(str(e["visitas"]), st_td),
                    Paragraph(tecnicos_txt, st_td_soft),
                    Paragraph(f'<font color="{nivel_hex}"><b>{nivel_txt}</b></font>', st_td),
                ])
            tabla_rec = Table(filas_rec, colWidths=[fw * 0.22, fw * 0.26, fw * 0.10, fw * 0.28, fw * 0.14], repeatRows=1)
            tabla_rec.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, C_BG]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(tabla_rec)
        else:
            story.append(Paragraph("No hay clientes con 2 o más visitas por el mismo tipo de problema en este período.", st_td_soft))

        # ── Calidad de instalación según desconexiones ──────────────────────
        story.append(Spacer(1, 16))
        story.append(Paragraph("CALIDAD DE INSTALACIÓN SEGÚN DESCONEXIONES", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))
        story.append(Paragraph(
            "Clientes que ya tienen cámaras instaladas y que además reportaron incidencias de \"Desconexión\" en el "
            "período del informe. Una cámara que se desconecta poco después de instalada es señal de una instalación "
            "de baja calidad (mala fijación, mal cableado, fuente de poder insuficiente, etc.). El cruce se hace por "
            "nombre de cliente/sucursal, no por cámara individual, porque el sistema aún no asocia cada desconexión "
            "a una cámara específica. Nivel de riesgo: bajo 1 desconexión, medio 2, alto 3 o más.",
            st_td_soft,
        ))
        story.append(Spacer(1, 6))

        if calidad_instalacion:
            filas_desc = [[
                Paragraph("CLIENTE / SUCURSAL", st_th), Paragraph("CÁMARAS INSTALADAS", st_th),
                Paragraph("DESCONEXIONES EN EL PERÍODO", st_th), Paragraph("NIVEL DE RIESGO", st_th),
            ]]
            for c in calidad_instalacion:
                nivel_txt, nivel_hex = _nivel_desconexion(c["desconexiones"])
                filas_desc.append([
                    Paragraph(c["cliente"], st_td),
                    Paragraph(str(c["camaras_instaladas"]), st_td),
                    Paragraph(str(c["desconexiones"]), st_td),
                    Paragraph(f'<font color="{nivel_hex}"><b>{nivel_txt}</b></font>', st_td),
                ])
            tabla_desc = Table(filas_desc, colWidths=[fw * 0.40, fw * 0.20, fw * 0.20, fw * 0.20], repeatRows=1)
            tabla_desc.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, C_BG]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, C_BORDER),
                ("BOX", (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(tabla_desc)
        else:
            story.append(Paragraph("No hay clientes con cámaras instaladas que hayan reportado desconexiones en este período.", st_td_soft))

        # ── Conclusión ───────────────────────────────────────────────────
        story.append(Spacer(1, 16))
        story.append(Paragraph("CONCLUSIÓN", st_sec))
        story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=8))

        partes_conclusion = []
        if not regs_filtrados:
            partes_conclusion.append("No se registró actividad de Servicio Técnico en el período seleccionado.")
        else:
            if cumplimiento_sla_pct is not None:
                nivel_sla = "crítico" if cumplimiento_sla_pct < 50 else ("de atención" if cumplimiento_sla_pct < 80 else "saludable")
                partes_conclusion.append(
                    f"El cumplimiento de SLA del período es <b>{nivel_sla}</b> ({cumplimiento_sla_pct}%): "
                    f"de {len(sla_eval)} ODTs evaluadas, {sla_ok} cumplieron el plazo de {sla_dias} días."
                )
            if odts_activas:
                partes_conclusion.append(
                    f"Actualmente hay <b>{odts_activas} ODTs activas</b> ({len(pend)} pendientes, {len(proc)} en proceso) "
                    "que requieren seguimiento."
                )
            if clientes_top:
                top_cli = clientes_top[0]
                sem_txt_top, _ = _semaforo(top_cli["total"])
                partes_conclusion.append(
                    f"El cliente con más recurrencia histórica es <b>{top_cli['cliente']}</b>, con {top_cli['total']} ODTs "
                    f"(nivel de alerta: {sem_txt_top.lower()})."
                )
            if camaras_pendientes_odt:
                partes_conclusion.append(
                    f"Hay <b>{len(camaras_pendientes_odt)} ODS con cámaras pendientes de instalar</b>, sumando "
                    f"{camaras_pendientes_total} cámaras sin instalar de un total de {total_camaras_contratadas} contratadas."
                )
            elif total_camaras_contratadas:
                partes_conclusion.append("Todas las cámaras contratadas en ODS de venta ya fueron instaladas.")
            if recurrencia_visitas:
                peor_rec = recurrencia_visitas[0]
                nivel_rec_txt, _ = _nivel_recurrencia(peor_rec["visitas"])
                partes_conclusion.append(
                    f"En calidad de visita técnica, el caso con más recurrencia es <b>{peor_rec['cliente']}</b> por "
                    f"\"{peor_rec['problema']}\", con {peor_rec['visitas']} visitas en el período "
                    f"(nivel {nivel_rec_txt.lower()})."
                )
            if calidad_instalacion:
                peor_desc = calidad_instalacion[0]
                nivel_desc_txt, _ = _nivel_desconexion(peor_desc["desconexiones"])
                partes_conclusion.append(
                    f"En calidad de instalación, <b>{peor_desc['cliente']}</b> registró {peor_desc['desconexiones']} "
                    f"desconexión(es) pese a tener cámaras ya instaladas (riesgo {nivel_desc_txt.lower()})."
                )
            partes_conclusion.append(
                "Se recomienda priorizar las ODTs activas más antiguas, las cámaras pendientes de instalación y los "
                "casos de recurrencia/desconexión listados en este informe, y dar seguimiento a los clientes con "
                "nivel de alerta naranjo, rojo, medio o alto."
            )
        story.append(Paragraph(" ".join(partes_conclusion), st_body))

        doc.build(story)
        buf.seek(0)
        return buf.getvalue()

    def finalizar_odt_coordinacion(self, odt: str, observacion_final: str = "") -> dict[str, Any]:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida.")
        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            return {"ok": False}
        ahora = datetime.now()
        obs = str(observacion_final or "").strip()
        row.estado = "Terminado"
        row.fecha_cierre = ahora
        if obs:
            row.observacion_final = obs
        if row.fecha_registro:
            row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days
        self.db.commit()
        self._sync_estado_ticket_soporte_silencioso(
            odt_limpia,
            TICKET_STATUS_RESUELTO_CLIENTE,
            nota_interna=self._build_nota_cierre_ticket_soporte(
                odt=odt_limpia,
                estado_ticket=TICKET_STATUS_RESUELTO_CLIENTE,
                derivacion=str(row.derivacion or "Cliente"),
                observacion_final=row.observacion_final or "",
            ),
        )
        return {
            "ok": True,
            "odt": odt_limpia,
            "estado": row.estado,
            "fecha_cierre": _to_ddmmyyyy_hhmm(row.fecha_cierre),
            "observacion_final": row.observacion_final or "",
        }

    def actualizar_observacion_final_coordinacion(self, odt: str, observacion_final: str = "") -> dict[str, Any]:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida.")
        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            return {"ok": False}
        row.observacion_final = str(observacion_final or "").strip()
        self.db.commit()
        return {"ok": True, "odt": odt_limpia, "observacion_final": row.observacion_final or ""}

    def _normalizar_codigo_cierre(self, valor: Any) -> str:
        return (
            self._normalizar_texto(valor)
            .replace("/", "_")
            .replace("-", "_")
            .replace(" ", "_")
        )

    def _normalizar_codigos_cierre(self, valor: Any) -> list[str]:
        items = valor if isinstance(valor, (list, tuple, set)) else [valor]
        normalizados: list[str] = []
        vistos: set[str] = set()
        for item in items:
            codigo = self._normalizar_codigo_cierre(item)
            if not codigo or codigo in vistos:
                continue
            vistos.add(codigo)
            normalizados.append(codigo)
        return normalizados

    @staticmethod
    def _formatear_lista_cierre(valores: list[str]) -> str:
        return " | ".join(v for v in valores if str(v or "").strip())

    def _normalizar_responsable_cierre(self, valor: Any) -> str:
        responsable = self._normalizar_texto(valor)
        if responsable == "atc":
            return "ATC"
        if responsable == "cliente":
            return "Cliente"
        if responsable in {"proveedor externo", "externo", "proveedor_externo"}:
            return "Proveedor Externo"
        if responsable == "internet":
            return "Internet"
        if responsable == "otro":
            return "Otro"
        raise ValueError(
            "Debes seleccionar el responsable del problema: ATC, Cliente, Proveedor Externo, Internet u Otro."
        )

    def _normalizar_materiales_cierre(
        self,
        materiales: list[Any] | None,
        materiales_sin_uso: bool = False,
    ) -> list[dict[str, Any]]:
        if materiales_sin_uso:
            return []

        catalogo = self._cargar_catalogo_materiales_sql()
        catalogo_por_codigo = {item["codigo"]: item for item in catalogo}
        catalogo_por_nombre = {
            item["nombre_normalizado"]: item for item in catalogo if item.get("nombre_normalizado")
        }

        normalizados: list[dict[str, Any]] = []
        for item in materiales or []:
            if not isinstance(item, dict):
                continue
            codigo = self._normalizar_codigo_cierre(item.get("codigo"))
            nombre = str(item.get("nombre") or "").strip()
            unidad = str(item.get("unidad") or "unidad").strip() or "unidad"
            cantidad_raw = item.get("cantidad")
            try:
                cantidad = float(str(cantidad_raw).replace(",", "."))
            except (TypeError, ValueError):
                cantidad = 0
            if not codigo and not nombre and cantidad <= 0:
                continue
            if codigo in catalogo_por_codigo:
                nombre = nombre or catalogo_por_codigo[codigo]["nombre"]
                if unidad == "unidad":
                    unidad = catalogo_por_codigo[codigo]["unidad_sugerida"] or unidad
            elif not codigo and nombre:
                item_catalogo = catalogo_por_nombre.get(self._normalizar_texto(nombre))
                if item_catalogo:
                    codigo = item_catalogo["codigo"]
                    nombre = item_catalogo["nombre"]
                    if unidad == "unidad":
                        unidad = item_catalogo["unidad_sugerida"] or unidad
            elif codigo not in self.MATERIALES_CIERRE:
                raise ValueError("Hay un material no valido en el cierre.")
            if codigo == "otro" and not nombre:
                raise ValueError("Indica el nombre del material marcado como Otro.")
            if cantidad <= 0:
                raise ValueError("La cantidad de cada material debe ser mayor a 0.")
            normalizados.append(
                {
                    "codigo": codigo,
                    "nombre": nombre or codigo,
                    "cantidad": int(cantidad) if cantidad.is_integer() else cantidad,
                    "unidad": unidad,
                }
            )

        if not normalizados:
            raise ValueError("Agrega al menos un material o marca que no se usaron materiales.")
        return normalizados

    def _normalizar_diagnostico_cierre(
        self,
        *,
        responsable_cierre: Any,
        causa_cierre: Any,
        accion_cierre: Any,
        resultado_cierre: Any,
        pruebas_cierre: list[Any] | None,
        materiales: list[Any] | None,
        materiales_sin_uso: bool = False,
        requiere_seguimiento: bool = False,
    ) -> dict[str, Any]:
        responsable = self._normalizar_responsable_cierre(responsable_cierre)
        causas = self._normalizar_codigos_cierre(causa_cierre)
        acciones = self._normalizar_codigos_cierre(accion_cierre)
        resultado = self._normalizar_codigo_cierre(resultado_cierre)
        pruebas = sorted(
            {
                self._normalizar_codigo_cierre(p)
                for p in (pruebas_cierre or [])
                if self._normalizar_codigo_cierre(p)
            }
        )

        if not causas:
            raise ValueError("Selecciona al menos una causa valida para el responsable elegido.")
        if any(causa not in self.CAUSAS_CIERRE[responsable] for causa in causas):
            raise ValueError("Selecciona una causa valida para el responsable elegido.")
        if not acciones:
            raise ValueError("Selecciona al menos una accion realizada valida.")
        if any(accion not in self.ACCIONES_CIERRE for accion in acciones):
            raise ValueError("Selecciona una accion realizada valida.")
        if resultado not in self.RESULTADOS_CIERRE:
            raise ValueError("Selecciona un resultado final valido.")
        if not pruebas:
            raise ValueError("Selecciona al menos una prueba realizada.")
        if any(p not in self.PRUEBAS_CIERRE for p in pruebas):
            raise ValueError("Hay una prueba realizada no valida.")

        materiales_norm = self._normalizar_materiales_cierre(materiales, materiales_sin_uso)
        requiere_seg = bool(
            requiere_seguimiento
            or resultado in {"requiere_seguimiento", "requiere_cotizacion_visita_adicional"}
        )

        return {
            "responsable_cierre": responsable,
            "causa_cierre": causas,
            "accion_cierre": acciones,
            "resultado_cierre": resultado,
            "pruebas_cierre": pruebas,
            "materiales": materiales_norm,
            "materiales_sin_uso": bool(materiales_sin_uso),
            "requiere_seguimiento": requiere_seg,
        }

    def _aplicar_diagnostico_cierre(self, row: Registro, diagnostico: dict[str, Any]) -> None:
        row.responsable_cierre = diagnostico["responsable_cierre"]
        row.causa_cierre = self._formatear_lista_cierre(diagnostico["causa_cierre"])
        row.accion_cierre = self._formatear_lista_cierre(diagnostico["accion_cierre"])
        row.resultado_cierre = diagnostico["resultado_cierre"]
        row.pruebas_cierre = json.dumps(diagnostico["pruebas_cierre"], ensure_ascii=False)
        row.materiales = json.dumps(
            {
                "sin_uso": diagnostico["materiales_sin_uso"],
                "items": diagnostico["materiales"],
            },
            ensure_ascii=False,
        )
        row.requiere_seguimiento = diagnostico["requiere_seguimiento"]

    def _reflejar_audio_ok_en_prueba_sonido(self, row: Registro, pruebas_cierre: list[Any] | None) -> None:
        """Si el técnico marcó 'audio_ok' al cerrar esta ODT y la sucursal está
        Pendiente en Pruebas de Sonido este mes (sin registro todavía), la
        marca como 'exitoso_terreno' automáticamente (máximo 1 por sucursal
        por mes) y avisa por correo que el sistema se verificó en terreno.
        Si la sucursal ya tiene CUALQUIER resultado este mes (exitoso,
        exitoso_terreno, falla o no_coordinacion) no se toca — el caso
        'falla' ya queda reflejado como Solucionado en pantalla en cuanto
        esta incidencia se cierra (ver odts_finalizadas en
        pruebas_sonido_sucursales), sin necesidad de nada adicional acá.

        Se ejecuta después del commit del cierre de la ODT — una falla acá
        no debe afectar el cierre, solo queda en el log."""
        try:
            if "audio_ok" not in (pruebas_cierre or []):
                return

            nombre_cliente = str(row.cliente or "").strip()
            if not nombre_cliente:
                return

            suc = self.db.scalar(
                select(SucursalBBDD).where(
                    func.lower(func.trim(SucursalBBDD.nombre_sucursal)) == nombre_cliente.lower()
                )
            )
            if not suc:
                return

            now = datetime.now()
            anio, mes = now.year, now.month
            existente = self.db.scalar(
                select(PruebaSonido).where(
                    PruebaSonido.sucursal_id == suc.id,
                    PruebaSonido.anio == anio,
                    PruebaSonido.mes == mes,
                )
            )
            if existente:
                return

            operador = str(row.tecnico_cierre or row.tecnicos or "").strip()
            prueba = PruebaSonido(
                sucursal_id=suc.id,
                anio=anio,
                mes=mes,
                resultado="exitoso_terreno",
                observacion=f"Verificado en terreno por el equipo técnico al cerrar la ODT {row.odt}",
                operador=operador,
            )
            self.db.add(prueba)
            self.db.commit()

            self._enviar_correo_prueba_sonido_terreno(suc, str(row.odt or ""))
        except Exception:
            LOGGER.exception("Error reflejando audio_ok de ODT %s en prueba de sonido", row.odt)
            self.db.rollback()

    def _construir_email_prueba_sonido_terreno(self, suc: SucursalBBDD, odt: str) -> tuple[str, str, str]:
        _MESES_ES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                     "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        now = datetime.now()
        nombre_empresa = suc.nombre_empresa or suc.nombre_sucursal or ""
        nombre_suc = suc.nombre_sucursal or ""
        mes_nombre = f"{_MESES_ES[now.month - 1]} de {now.year}"
        fecha_str = f"{now.day} de {_MESES_ES[now.month - 1]} de {now.year}"

        asunto = f"Sistema de Sonido Verificado en Terreno — {nombre_suc}"

        cuerpo_txt = (
            f"Estimados,\n\n"
            f"Junto con saludar, les informamos que durante la visita técnica realizada en la "
            f"sucursal {nombre_suc} de {nombre_empresa}, nuestro equipo técnico verificó en terreno "
            f"el sistema de sonido correspondiente a {mes_nombre}, obteniendo un resultado "
            f"100% exitoso.\n\n"
            f"Esta verificación confirma el correcto funcionamiento de parlantes, amplificadores y "
            f"toda la cadena de audio del sistema, el cual se encuentra operativo y en condiciones "
            f"óptimas.\n\n"
            f"En Alguien Te Cuida realizamos estas verificaciones de forma periódica para garantizar "
            f"que usted cuente siempre con un sistema operativo al 100%.\n\n"
            f"Cualquier problema o dificultad que usted visualice en el sistema, ya sea de cámaras, "
            f"parlantes u otros componentes, le rogamos avisarnos a la brevedad posible.\n\n"
            f"Atentamente,\nEquipo Técnico — Alguien Te Cuida SpA"
        )

        cuerpo_html = f"""<!DOCTYPE html>
<html lang="es" xmlns="http://www.w3.org/1999/xhtml" style="color-scheme:light;">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <meta name="color-scheme" content="light">
  <title>{asunto}</title>
</head>
<body style="margin:0;padding:0;background-color:#f2f4f7;-webkit-text-size-adjust:100%;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
       style="background-color:#f2f4f7;min-width:320px;">
  <tr>
    <td align="center" style="padding:40px 16px;">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0"
             style="max-width:600px;width:100%;background-color:#ffffff;border-radius:6px;
                    overflow:hidden;border:1px solid #d1d5db;">

        <!-- HEADER -->
        <tr>
          <td style="background-color:#0d1f2d;padding:20px 36px;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="vertical-align:middle;">
                  <img src="cid:logoatc" alt="Alguien Te Cuida"
                       style="height:38px;width:auto;display:block;border:0;" />
                </td>
                <td align="right" style="vertical-align:middle;">
                  <span style="font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                               color:#8aabb8;letter-spacing:0.12em;text-transform:uppercase;">
                    Informe Técnico
                  </span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- FRANJA ACENTO -->
        <tr>
          <td style="background-color:#1e3a5f;padding:20px 36px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                      color:#93b4cc;letter-spacing:0.12em;text-transform:uppercase;">
              Visita técnica &nbsp;·&nbsp; {mes_nombre}
            </p>
            <p style="margin:7px 0 0;font-family:Arial,sans-serif;font-size:20px;font-weight:700;
                      color:#ffffff;letter-spacing:-0.01em;line-height:1.25;">
              Sistema de Sonido Verificado en Terreno
            </p>
            <p style="margin:5px 0 0;font-family:Arial,sans-serif;font-size:13px;
                      color:#a8c4d8;line-height:1.4;">
              {nombre_suc} &nbsp;·&nbsp; {nombre_empresa}
            </p>
          </td>
        </tr>

        <!-- BADGE RESULTADO -->
        <tr>
          <td style="background-color:#ffffff;padding:26px 36px 4px;">
            <table role="presentation" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="background-color:#f0fdf4;border:1px solid #a7f3d0;border-radius:5px;
                           padding:9px 16px;">
                  <span style="font-family:Arial,sans-serif;font-size:12px;font-weight:700;
                               color:#15803d;letter-spacing:0.02em;">VERIFICADO EN TERRENO</span>
                  <span style="font-family:Arial,sans-serif;font-size:12px;color:#6b7280;
                               margin-left:14px;">{fecha_str}</span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- CUERPO -->
        <tr>
          <td style="padding:20px 36px 8px;background-color:#ffffff;">
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">Estimados,</p>
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Junto con saludar, les informamos que durante la visita técnica realizada en la
              sucursal <strong>{nombre_suc}</strong>, nuestro equipo técnico verificó
              <strong>en terreno</strong> el sistema de sonido correspondiente a
              <strong>{mes_nombre}</strong>, obteniendo un resultado <strong>100&#37; exitoso</strong>.
            </p>
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Esta verificación confirma el correcto funcionamiento de parlantes, amplificadores y
              toda la cadena de audio del sistema, el cual se encuentra operativo y en condiciones
              óptimas.
            </p>
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              En <strong>Alguien Te Cuida</strong> entendemos que la tranquilidad de su operación
              depende de que cada componente de su sistema de seguridad funcione correctamente.
              Por eso realizamos estas verificaciones de forma periódica: para garantizar que usted
              cuente siempre con un sistema al 100&#37;, sin sorpresas ni imprevistos.
            </p>
            <p style="margin:0 0 24px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Cualquier problema o dificultad que usted visualice en el sistema, ya sea de
              <strong>cámaras</strong>, <strong>parlantes</strong> u otros componentes, le rogamos
              avisarnos a la brevedad posible.
            </p>
          </td>
        </tr>

        <!-- SEPARADOR -->
        <tr>
          <td style="padding:0 36px 18px;background-color:#ffffff;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
              <tr><td style="border-top:1px solid #e5e7eb;font-size:0;line-height:0;">&nbsp;</td></tr>
            </table>
          </td>
        </tr>

        <!-- FIRMA -->
        <tr>
          <td style="padding:0 36px 26px;background-color:#ffffff;">
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:13px;
                      font-weight:700;color:#111827;">Equipo Técnico</p>
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">Alguien Te Cuida SpA</p>
            <p style="margin:0;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">contacto@alguientecuida.cl</p>
          </td>
        </tr>

        <!-- FOOTER -->
        <tr>
          <td style="background-color:#f8fafc;border-top:1px solid #e5e7eb;
                     padding:14px 36px;border-radius:0 0 6px 6px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:11px;color:#9ca3af;line-height:1.5;">
              Este mensaje fue generado automáticamente por el sistema de Alguien Te Cuida SpA.
              Por favor no responda directamente a este correo.
            </p>
          </td>
        </tr>

      </table>
    </td>
  </tr>
</table>
</body>
</html>"""
        return asunto, cuerpo_txt, cuerpo_html

    def _enviar_correo_prueba_sonido_terreno(self, suc: SucursalBBDD, odt: str) -> None:
        contactos = [
            str(c.email).strip()
            for c in self.db.query(SucursalContactoEmergencia)
                        .filter(SucursalContactoEmergencia.sucursal_id == suc.id)
                        .order_by(SucursalContactoEmergencia.id.asc())
                        .all()
            if c.email and str(c.email).strip()
        ]
        email_destino = ", ".join(contactos)
        if not email_destino:
            return

        asunto, cuerpo_txt, cuerpo_html = self._construir_email_prueba_sonido_terreno(suc, odt)

        logo_path = _ATC_ROOT / "static" / "img" / "logo-atc.png"
        logo_bytes = logo_path.read_bytes() if logo_path.exists() else None

        def _enviar(dest=email_destino, subj=asunto, txt=cuerpo_txt, html=cuerpo_html,
                    logo=logo_bytes, sid=suc.id, bcc=["tahira.riquelme.atc@gmail.com"]):
            try:
                svc_mail = IncidenciasService(SessionLocal())
                svc_mail._enviar_correo_automatico(
                    dest, subj, txt, html_body=html, logo_bytes=logo,
                    cfg_override=svc_mail._contacto_smtp_runtime_config(),
                    bcc_emails_extra=bcc,
                )
            except Exception:
                LOGGER.exception("Error enviando email prueba sonido terreno sucursal=%s", sid)

        threading.Thread(target=_enviar, daemon=True, name=f"email-sonido-terreno-{suc.id}").start()

    def _resumen_diagnostico_cierre(self, diagnostico: dict[str, Any]) -> str:
        materiales = diagnostico.get("materiales") or []
        if diagnostico.get("materiales_sin_uso"):
            materiales_txt = "Sin materiales"
        else:
            materiales_txt = ", ".join(
                f"{m.get('nombre') or m.get('codigo')} x{m.get('cantidad')} {m.get('unidad') or ''}".strip()
                for m in materiales
            )
        return (
            f"Responsable: {diagnostico.get('responsable_cierre')}; "
            f"Causa: {self._formatear_lista_cierre(diagnostico.get('causa_cierre') or [])}; "
            f"Accion: {self._formatear_lista_cierre(diagnostico.get('accion_cierre') or [])}; "
            f"Resultado: {diagnostico.get('resultado_cierre')}; "
            f"Pruebas: {', '.join(diagnostico.get('pruebas_cierre') or [])}; "
            f"Materiales: {materiales_txt or 'Sin materiales'}"
        )

    def _observacion_drive_cierre(self, observacion: str, diagnostico: dict[str, Any]) -> str:
        observacion_txt = str(observacion or "").strip()
        resumen = self._resumen_diagnostico_cierre(diagnostico)
        return f"{observacion_txt}\n\nDiagnostico estructurado: {resumen}".strip()

    def validar_odt_mantencion_preventiva(self, odt: str) -> Registro:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida.")
        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            raise ValueError(f"No se encontro la ODT {odt_limpia}.")
        if not self._es_registro_mantencion_preventiva(row):
            raise ValueError("La carga masiva de imagenes aplica solo a Mantencion Preventiva.")
        return row

    @staticmethod
    def nombre_staging_cierre_mantencion(index: int, filename: str, mime_type: str = "") -> str:
        raw_name = str(filename or "").strip()
        _, guessed_ext = mimetypes.guess_type(raw_name)
        ext = Path(raw_name).suffix.lower()
        if not ext:
            if "png" in str(mime_type or "").lower():
                ext = ".png"
            elif "webp" in str(mime_type or "").lower():
                ext = ".webp"
            else:
                ext = ".jpg"
        ext = re.sub(r"[^.a-z0-9]+", "", ext)[:12] or ".jpg"
        return f"Imagen {max(1, int(index))}{ext}"

    @staticmethod
    def nombre_staging_cierre_odt(index: int, filename: str, mime_type: str = "") -> str:
        raw_name = str(filename or "").strip()
        ext = Path(raw_name).suffix.lower()
        if not ext:
            if "png" in str(mime_type or "").lower():
                ext = ".png"
            elif "webp" in str(mime_type or "").lower():
                ext = ".webp"
            else:
                ext = ".jpg"
        ext = re.sub(r"[^.a-z0-9]+", "", ext)[:12] or ".jpg"
        return f"imagen_{max(1, int(index)):02d}{ext}"

    @staticmethod
    def url_publica_upload(path: Path) -> str:
        rel = Path(path).resolve().relative_to(_UPLOADS_ROOT.resolve()).as_posix()
        return f"/uploads/{rel}"

    def _guardar_drive_cierre_folder(self, odt: str, folder_id: str = "", folder_url: str = "") -> None:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            return
        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            return
        if folder_id:
            row.drive_cierre_folder_id = str(folder_id or "").strip()
        if folder_url:
            row.drive_cierre_folder_url = str(folder_url or "").strip()
        self.db.commit()

    @staticmethod
    def _subir_imagenes_cierre_mantencion_worker(
        odt: str,
        foto_payloads: list[dict[str, object]],
        observacion: str,
    ) -> None:
        db = SessionLocal()
        try:
            service = IncidenciasService(db)
            result = service._generar_drive_para_cierre(odt, observacion, [], foto_payloads=foto_payloads)
            folder_id = str(result.get("folder_id") or "").strip()
            if folder_id:
                folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
                service._guardar_drive_cierre_folder(odt, folder_id, folder_url)
        except Exception:
            LOGGER.exception("Fallo la subida de imagenes y generacion de informe de mantencion para ODT %s.", odt)
        finally:
            db.close()

    def cerrar_mantencion_con_imagenes_staging(
        self,
        odt: str,
        foto_payloads: list[dict[str, object]],
        observacion: str,
        *,
        responsable_cierre: str,
        causa_cierre: Any,
        accion_cierre: Any,
        resultado_cierre: str,
        pruebas_cierre: list[Any] | None = None,
        materiales: list[Any] | None = None,
        materiales_sin_uso: bool = False,
        requiere_seguimiento: bool = False,
        token: str = "",
    ) -> dict[str, Any]:
        odt_limpia = str(odt or "").strip()
        row = self.validar_odt_mantencion_preventiva(odt_limpia)
        payloads_validos = [p for p in foto_payloads or [] if isinstance(p.get("bytes"), (bytes, bytearray)) and p.get("bytes")]
        if not payloads_validos:
            raise ValueError("Debes adjuntar al menos una imagen para cerrar una mantencion.")
        if len(payloads_validos) > self.MANTENCION_CIERRE_MAX_IMAGENES:
            raise ValueError(f"Solo puedes adjuntar hasta {self.MANTENCION_CIERRE_MAX_IMAGENES} imagenes.")

        # Mantención preventiva no requiere diagnóstico — proveer defaults
        responsable_cierre = responsable_cierre or "ATC"
        causa_cierre = causa_cierre or ["mantenimiento_insuficiente"]
        accion_cierre = accion_cierre or ["limpieza"]
        resultado_cierre = resultado_cierre or "operativo"
        pruebas_cierre = pruebas_cierre or ["camaras_ok"]
        materiales_sin_uso = materiales_sin_uso or not (materiales or [])

        result = self.registrar_finalizacion_rapida(
            odt_limpia,
            observacion,
            responsable_cierre=responsable_cierre,
            causa_cierre=causa_cierre,
            accion_cierre=accion_cierre,
            resultado_cierre=resultado_cierre,
            pruebas_cierre=pruebas_cierre,
            materiales=materiales,
            materiales_sin_uso=materiales_sin_uso,
            requiere_seguimiento=requiere_seguimiento,
        )

        drive_enabled = bool(settings.google_drive_enabled and str(settings.google_drive_root_folder_id or "").strip())
        if drive_enabled:
            worker = threading.Thread(
                target=self._subir_imagenes_cierre_mantencion_worker,
                args=(odt_limpia, payloads_validos, observacion),
                daemon=True,
                name=f"mantencion-cierre-img-{odt_limpia}",
            )
            worker.start()

        return {
            "result": result,
            "odt": odt_limpia,
            "imagenes_recibidas": len(payloads_validos),
            "drive_enabled": drive_enabled,
            "drive_queued": drive_enabled,
            "message": (
                "ODT cerrada. Las imagenes se estan subiendo a Drive en segundo plano."
                if drive_enabled
                else "ODT cerrada. Imagenes guardadas en staging local; Drive no esta habilitado."
            ),
        }

    def cerrar_incidencia(self, odt: str, fecha_cierre: datetime) -> bool:
        odt_limpia = (odt or "").strip()
        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            return False
        row.derivacion = "Finalizado Por Encargados"
        row.estado = "Terminado"
        row.fecha_cierre = fecha_cierre
        if row.fecha_registro and row.fecha_cierre:
            row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days
        self.db.commit()
        self._sync_estado_ticket_soporte_silencioso(
            odt_limpia,
            TICKET_STATUS_RESUELTO_SERVICIO,
            nota_interna=self._build_nota_cierre_ticket_soporte(
                odt=odt_limpia,
                estado_ticket=TICKET_STATUS_RESUELTO_SERVICIO,
                derivacion=row.derivacion or "Finalizado Por Encargados",
                observacion_final=row.observacion_final or "",
            ),
        )
        return True

    def registrar_finalizacion_rapida(
        self,
        odt: str,
        observacion: str,
        *,
        responsable_cierre: str,
        causa_cierre: Any,
        accion_cierre: Any,
        resultado_cierre: str,
        pruebas_cierre: list[Any] | None = None,
        materiales: list[Any] | None = None,
        materiales_sin_uso: bool = False,
        requiere_seguimiento: bool = False,
        token: str = "",
    ) -> str:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida")

        obs_cierre = str(observacion or "").strip()
        if not obs_cierre:
            raise ValueError("Debes ingresar una observacion final breve.")

        diagnostico = self._normalizar_diagnostico_cierre(
            responsable_cierre=responsable_cierre,
            causa_cierre=causa_cierre,
            accion_cierre=accion_cierre,
            resultado_cierre=resultado_cierre,
            pruebas_cierre=pruebas_cierre,
            materiales=materiales,
            materiales_sin_uso=materiales_sin_uso,
            requiere_seguimiento=requiere_seguimiento,
        )

        ahora = datetime.now()
        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            raise ValueError(f"No se encontro la ODT {odt_limpia}")

        estado_previo = str(getattr(row, "estado", "") or "").strip()
        row.estado = "Terminado"
        row.derivacion = "Servicio Técnico"
        row.observacion_final = obs_cierre
        self._aplicar_diagnostico_cierre(row, diagnostico)
        row.porcentaje_avance = "100%"
        row.fecha_cierre = ahora
        row.prioridad = None
        if row.fecha_registro:
            row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days
        if row.fecha_inicio_trabajo and not row.fecha_fin_trabajo:
            row.fecha_fin_trabajo = ahora
        if (token or "").strip():
            usuario_token = str(self.get_usuario_actual((token or "").strip()) or "").strip()
            if usuario_token and usuario_token != "Desconocido":
                # Quien realmente finalizó la ODT, aunque haya sido derivada
                # a otro técnico (row.tecnicos no se toca acá).
                row.tecnico_cierre = usuario_token
        self._marcar_instalacion_venta_finalizada(odt_limpia, ahora)
        self.db.commit()
        self._reflejar_audio_ok_en_prueba_sonido(row, diagnostico.get("pruebas_cierre"))
        if estado_previo != "Terminado":
            self._reforzar_inicio_odt_si_corresponde(
                odt=row.odt,
                tecnico=row.tecnicos,
                acompanante=row.acompanante,
                usuario_accion=row.tecnico_cierre,
                fecha_inicio_trabajo=row.fecha_inicio_trabajo,
                verbo_accion="finalizada",
            )
        self._sync_estado_ticket_soporte_silencioso(
            odt_limpia,
            TICKET_STATUS_RESUELTO_SERVICIO,
            nota_interna=self._build_nota_cierre_ticket_soporte(
                odt=odt_limpia,
                estado_ticket=TICKET_STATUS_RESUELTO_SERVICIO,
                derivacion=row.derivacion or "Servicio Tecnico",
                observacion_final=(
                    f"{row.observacion_final or obs_cierre}\n"
                    f"{self._resumen_diagnostico_cierre(diagnostico)}"
                ).strip(),
            ),
        )
        return "OK"

    def _marcar_instalacion_venta_finalizada(self, odt: str, fecha_cierre: datetime | None = None) -> None:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            return

        venta_codigo = self.db.scalar(
            select(VentaODS.codigo).where(func.lower(func.trim(VentaODS.codigo)) == odt_limpia.lower())
        )
        if not venta_codigo:
            return

        row = self.db.scalar(
            select(ServicioTecnicoVentaODT).where(
                func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == odt_limpia.lower()
            )
        )
        if not row:
            row = ServicioTecnicoVentaODT(odt=odt_limpia)
            self.db.add(row)

        row.instalacion_finalizada = True
        if not row.fecha_instalacion_finalizada:
            row.fecha_instalacion_finalizada = fecha_cierre or datetime.now()

    def cerrar_instalacion_venta(
        self,
        odt: str,
        *,
        observacion: str,
        instalacion_completa: bool,
        pruebas_cierre: list[Any] | None = None,
        fotos_base64: list[str] | None = None,
        token: str = "",
        cantidad_instalada_total: int | None = None,
        foto_payloads: list[dict[str, object]] | None = None,
    ) -> dict[str, Any]:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida")

        obs_cierre = str(observacion or "").strip()
        if not obs_cierre:
            raise ValueError("Ingresa una observacion final breve.")
        if not instalacion_completa:
            raise ValueError("Marca que se instaló todo lo contratado para cerrar la instalación.")

        venta_row = self.db.scalar(
            select(VentaODS).where(func.lower(func.trim(VentaODS.codigo)) == odt_limpia.lower())
        )
        if not venta_row:
            raise ValueError(f"No se encontro la ODS {odt_limpia}")

        pruebas = sorted(
            {
                self._normalizar_codigo_cierre(p)
                for p in (pruebas_cierre or [])
                if self._normalizar_codigo_cierre(p)
            }
        )
        if not pruebas:
            raise ValueError("Selecciona al menos una prueba realizada.")
        if any(p not in self.PRUEBAS_CIERRE for p in pruebas):
            raise ValueError("Hay una prueba realizada no valida.")

        fotos_recibidas = [str(f or "").strip() for f in (fotos_base64 or []) if str(f or "").strip()]
        if len(fotos_recibidas) > self.MAX_FOTOS_CIERRE_ODS:
            raise ValueError(f"Solo puedes adjuntar hasta {self.MAX_FOTOS_CIERRE_ODS} imagenes para esta ODS.")
        fotos = fotos_recibidas[: self.MAX_FOTOS_CIERRE_ODS]
        if not fotos:
            raise ValueError("Debes adjuntar al menos una foto para cerrar la instalación.")

        ahora = datetime.now()
        st_row = self.db.scalar(
            select(ServicioTecnicoVentaODT).where(
                func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == odt_limpia.lower()
            )
        )
        if not st_row:
            st_row = ServicioTecnicoVentaODT(odt=odt_limpia)
            self.db.add(st_row)

        usuario_token = ""
        if (token or "").strip():
            usuario_token = str(self.get_usuario_actual((token or "").strip()) or "").strip()
            if usuario_token == "Desconocido":
                usuario_token = ""

        row = self.db.scalar(select(Registro).where(func.lower(func.trim(Registro.odt)) == odt_limpia.lower()))
        row_es_nueva = row is None
        estado_previo = "" if row_es_nueva else str(getattr(row, "estado", "") or "").strip()
        if not row:
            row = Registro(
                odt=odt_limpia,
                fecha_registro=ahora,
                puesto=None,
                cliente=str(venta_row.nombre_sucursal or venta_row.razon_social or "").strip() or str(venta_row.razon_social or "").strip(),
                problema=str(venta_row.tipo_servicio or "").strip() or "Instalacion",
                detalle_problema=str(venta_row.observacion or venta_row.consideraciones or "").strip() or None,
                derivacion="Servicio Técnico",
                observacion=str(venta_row.observacion or venta_row.consideraciones or "").strip() or None,
                observacion_soporte=None,
                observacion_servicio=None,
                tecnicos=str(getattr(st_row, "tecnico_a_cargo", "") or "").strip() or usuario_token or None,
                acompanante=str(getattr(st_row, "acompanante", "") or "").strip() or None,
                estado="Terminado",
                fecha_derivacion_area=ahora,
                fecha_derivacion_tecnico=ahora,
                direccion=str(venta_row.direccion_sucursal or "").strip() or None,
            )
            self.db.add(row)

        total_camaras = max(int(getattr(venta_row, "numero_camaras_instalar", 0) or 0), 0)
        try:
            cantidad_instalada_total = int(cantidad_instalada_total) if cantidad_instalada_total is not None else None
        except (TypeError, ValueError):
            cantidad_instalada_total = None
        excede_presupuesto = bool(
            cantidad_instalada_total and total_camaras > 0 and cantidad_instalada_total > total_camaras
        )
        row.estado = "Terminado"
        row.derivacion = "Servicio Técnico"
        row.observacion_final = obs_cierre
        row.fecha_cierre = ahora
        row.prioridad = None
        row.responsable_cierre = None
        row.causa_cierre = None
        row.accion_cierre = None
        row.resultado_cierre = "instalacion_completa"
        row.pruebas_cierre = json.dumps(pruebas, ensure_ascii=False)
        row.materiales = json.dumps({"sin_uso": True, "items": []}, ensure_ascii=False)
        row.requiere_seguimiento = False
        row.porcentaje_avance = f"{total_camaras if total_camaras > 0 else 100}%"
        if not row.fecha_inicio_trabajo and getattr(st_row, "fecha_inicio_trabajo", None):
            row.fecha_inicio_trabajo = st_row.fecha_inicio_trabajo
        if row.fecha_inicio_trabajo and not row.fecha_fin_trabajo:
            row.fecha_fin_trabajo = ahora
        if usuario_token:
            # Quien realmente cerró la instalación, aunque haya sido derivada
            # a otro técnico (row.tecnicos no se toca acá).
            row.tecnico_cierre = usuario_token
        if len(fotos) >= 1:
            row.foto_1 = fotos[0]
        if len(fotos) >= 2:
            row.foto_2 = fotos[1]
        if len(fotos) >= 3:
            row.foto_3 = fotos[2]
        self._upsert_unified_images(
            odt_limpia,
            str(row.cliente or venta_row.nombre_sucursal or venta_row.razon_social or "").strip(),
            "cierre_instalacion",
            fotos,
            max_imagenes=self.MAX_FOTOS_CIERRE_ODS,
        )
        if row.fecha_registro and row.fecha_cierre:
            row.dias_ejecucion = (row.fecha_cierre.date() - row.fecha_registro.date()).days

        # Solo "Instalación Finalizada" — "Terminado" (finalizado/fecha_cierre) es un
        # paso posterior y deliberado que se marca a mano en la Tabla Servicio
        # Técnico, no algo que el cierre de ODT del técnico deba marcar solo (ver
        # update_servicio_tecnico_ventas_estado, que exige instalacion_finalizada
        # antes de permitir marcar finalizado).
        st_row.instalacion_finalizada = True
        st_row.fecha_instalacion_finalizada = st_row.fecha_instalacion_finalizada or ahora
        if excede_presupuesto:
            st_row.camaras_instaladas_reales = cantidad_instalada_total
        if not getattr(st_row, "fecha_inicio_trabajo", None) and getattr(row, "fecha_inicio_trabajo", None):
            st_row.fecha_inicio_trabajo = row.fecha_inicio_trabajo
        if getattr(st_row, "fecha_inicio_trabajo", None) and not getattr(st_row, "fecha_fin_trabajo", None):
            st_row.fecha_fin_trabajo = ahora
        self.db.commit()
        if estado_previo != "Terminado":
            self._reforzar_inicio_odt_si_corresponde(
                odt=row.odt,
                tecnico=row.tecnicos,
                acompanante=row.acompanante,
                usuario_accion=row.tecnico_cierre,
                fecha_inicio_trabajo=row.fecha_inicio_trabajo,
                verbo_accion="finalizada",
            )
        drive_enabled = bool(settings.google_drive_enabled)
        if drive_enabled:
            worker = threading.Thread(
                target=self._ejecutar_drive_en_segundo_plano,
                args=(odt_limpia, obs_cierre, fotos, self.MAX_FOTOS_INFORME_ODS, foto_payloads),
                daemon=True,
                name=f"drive-report-{odt_limpia}",
            )
            worker.start()
        if excede_presupuesto:
            self._enviar_correo_exceso_instalacion(
                odt=odt_limpia,
                cliente=str(row.cliente or venta_row.nombre_sucursal or venta_row.razon_social or "").strip(),
                tecnico=str(row.tecnicos or "").strip(),
                acompanante=str(row.acompanante or "").strip(),
                camaras_contratadas=total_camaras,
                camaras_instaladas=cantidad_instalada_total,
                creado_por=str(getattr(venta_row, "creado_por", "") or "").strip(),
            )
        return {
            "result": "OK",
            "odt": odt_limpia,
            "camaras_instaladas": cantidad_instalada_total if excede_presupuesto else total_camaras,
            "instalacion_finalizada": True,
            "imagenes_recibidas": len(fotos),
            "imagenes_informe": min(len(fotos), self.MAX_FOTOS_INFORME_ODS),
            "drive_enabled": drive_enabled,
            "drive_queued": drive_enabled,
        }

    def _normalizar_texto(self, valor: Any) -> str:
        txt = str(valor or "").strip().lower()
        txt = unicodedata.normalize("NFD", txt)
        return "".join(c for c in txt if unicodedata.category(c) != "Mn")
    def _filtrar_incidencias_para_tecnico(
        self,
        filas: list[list[Any]],
        tecnico: str | None = None,
    ) -> list[list[Any]]:
        tecnico_norm = self._normalizar_texto(tecnico or "")
        if not tecnico_norm:
            return filas

        filtradas: list[list[Any]] = []
        for fila in filas:
            derivacion = self._normalizar_texto(fila[5] if len(fila) > 5 else "")
            estado = self._normalizar_texto(fila[8] if len(fila) > 8 else "")
            es_terminada = (
                "termin" in estado
                or "final" in estado
                or "finalizado" in derivacion
                or "terminado" in derivacion
                or "repetida" in derivacion
            )
            # Mantener foco en Servicio Tecnico, pero no ocultar ODT ya finalizadas
            # que siguen perteneciendo al tecnico logueado.
            if "servicio tecnico" not in derivacion and not es_terminada:
                continue

            tecnico_txt = self._normalizar_texto(fila[7] if len(fila) > 7 else "")
            acomp_txt = self._normalizar_texto(fila[10] if len(fila) > 10 else "")
            obs_txt = self._normalizar_texto(fila[6] if len(fila) > 6 else "")
            asignados = f"{tecnico_txt} {acomp_txt} {obs_txt}".strip()
            if tecnico_norm in asignados:
                filtradas.append(fila)
        return filtradas

    def _fila_aplica_a_tecnico(self, fila: Registro, tecnico_norm: str) -> bool:
        if not tecnico_norm:
            return True

        derivacion = self._normalizar_texto(getattr(fila, "derivacion", "") or "")
        estado = self._normalizar_texto(getattr(fila, "estado", "") or "")
        es_terminada = (
            "termin" in estado
            or "final" in estado
            or "finalizado" in derivacion
            or "terminado" in derivacion
            or "repetida" in derivacion
        )
        if "servicio tecnico" not in derivacion and not es_terminada:
            return False

        tecnico_txt = self._normalizar_texto(getattr(fila, "tecnicos", "") or "")
        acomp_txt = self._normalizar_texto(getattr(fila, "acompanante", "") or "")
        obs_txt = self._normalizar_texto(getattr(fila, "observacion", "") or "")
        asignados = f"{tecnico_txt} {acomp_txt} {obs_txt}".strip()
        return tecnico_norm in asignados

    def _buscar_cliente_por_odt(self, odt: str) -> str:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            return ""

        # Fuente de verdad: ODT activa en tablas locales de la app.
        # Esto evita tomar una sucursal incorrecta cuando en tablas externas
        # existen ODT antiguas/repetidas con el mismo código.
        row_reg = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if row_reg and row_reg.cliente:
            return row_reg.cliente

        row_venta = self.db.scalar(select(VentaODS).where(VentaODS.codigo == odt_limpia))
        if row_venta:
            return str(row_venta.nombre_sucursal or row_venta.razon_social or "").strip()

        # Fallback: tabla externa de incidencias, solo si no existe en locales.
        try:
            for schema_name in self._schemas_con_tabla("incidencias"):
                cols = self._columnas_tabla(schema_name, "incidencias")
                if not cols:
                    continue
                col_odt = self._pick_col(cols, ["odt", "codigo_odt", "codigo", "nro_odt"])
                col_cliente = self._pick_col(cols, ["cliente", "nombre_sucursal", "sucursal", "nombre_cliente"])
                if not col_odt or not col_cliente:
                    continue
                sql = text(
                    f"""
                    SELECT "{col_cliente}"
                    FROM "{schema_name}"."incidencias"
                    WHERE TRIM(CAST("{col_odt}" AS NVARCHAR(MAX))) = :odt
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                )
                value = self.db.execute(sql, {"odt": odt_limpia}).scalar()
                if value and str(value).strip():
                    return str(value).strip()
        except Exception:
            self.db.rollback()

        return ""

    def obtener_datos_sucursal_con_coordenadas(self, odt: str) -> dict[str, str]:
        odt_limpia = str(odt or "").strip()
        cliente = self._buscar_cliente_por_odt(odt_limpia)
        salida = {
            "cliente": cliente or "",
            "direccion": "",
            "contacto": "",
            "telefono": "",
            "correo": "",
            "latitud": "",
            "longitud": "",
            "layout": "",
            "observacion": "",
            "camaras_total": 0,
            "camaras_instaladas": 0,
            "camaras_pendientes": 0,
            "estado_cierre": "",
        }

        def _contar_registros_camaras(raw: Any) -> int:
            if raw in (None, "", [], (), {}):
                return 0
            if isinstance(raw, (list, tuple, set)):
                return len(raw)
            texto = str(raw or "").strip()
            if not texto:
                return 0
            try:
                parsed = json.loads(texto)
            except Exception:
                parsed = [part.strip() for part in re.split(r"[,\n|;]+", texto) if part.strip()]
            if isinstance(parsed, (list, tuple, set)):
                return len(parsed)
            return 1 if str(parsed or "").strip() else 0

        def _porcentaje_a_camaras_instaladas(raw: Any, total: int) -> int:
            if total <= 0:
                return 0
            texto = str(raw or "").strip().replace(",", ".").replace("%", "")
            if not texto:
                return 0
            try:
                pct = float(texto)
            except Exception:
                return 0
            if pct <= 0:
                return 0
            if pct > 100:
                pct = 100
            return max(0, min(total, int(round(total * pct / 100.0))))

        venta_row = None
        try:
            venta_row = (
                self.db.query(VentaODS)
                .filter(func.lower(func.trim(VentaODS.codigo)) == odt_limpia.lower())
                .first()
            )
        except Exception:
            self.db.rollback()
            venta_row = None

        if venta_row:
            venta_cliente = str(venta_row.nombre_sucursal or venta_row.razon_social or "").strip()
            if venta_cliente:
                salida["cliente"] = venta_cliente
                cliente = venta_cliente
            salida["direccion"] = str(venta_row.direccion_sucursal or "").strip()
            salida["observacion"] = str(venta_row.observacion or venta_row.consideraciones or "").strip()
            try:
                archivos = self.db.query(VentaODSArchivo).filter(VentaODSArchivo.ods_id == venta_row.id).all()
                for archivo in archivos:
                    tipo = str(archivo.tipo_documento or "").strip().lower()
                    if tipo == "layout" and not salida["layout"]:
                        salida["layout"] = str(archivo.ruta_archivo or "").strip()
            except Exception:
                self.db.rollback()

            try:
                st_row = self.db.scalar(
                    select(ServicioTecnicoVentaODT).where(
                        func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == odt_limpia.lower()
                    )
                )
                total_camaras = max(int(venta_row.numero_camaras_instalar or 0), 0)
                finalizado = bool(st_row and (st_row.instalacion_finalizada or st_row.finalizado))
            except Exception:
                self.db.rollback()
                total_camaras = max(int(venta_row.numero_camaras_instalar or 0), 0)
                finalizado = False
            if finalizado:
                instaladas_camaras = total_camaras
            else:
                try:
                    avance_row = self.db.scalar(
                        select(Registro)
                        .where(func.lower(func.trim(Registro.odt)) == odt_limpia.lower())
                        .order_by(Registro.fecha_registro.desc(), Registro.id.desc())
                    )
                    instaladas_camaras = _porcentaje_a_camaras_instaladas(
                        getattr(avance_row, "porcentaje_avance", None) if avance_row else None,
                        total_camaras,
                    )
                except Exception:
                    self.db.rollback()
                    instaladas_camaras = 0
            salida["camaras_total"] = total_camaras
            salida["camaras_instaladas"] = instaladas_camaras
            salida["camaras_pendientes"] = max(total_camaras - instaladas_camaras, 0)
            salida["estado_cierre"] = "Finalizado" if finalizado else "Pendiente"

            sucursal_nombre = str(venta_row.nombre_sucursal or venta_row.razon_social or cliente or "").strip()
            if sucursal_nombre:
                try:
                    sucursal_row = self.db.scalar(
                        select(SucursalBBDD).where(
                            func.lower(func.trim(SucursalBBDD.nombre_sucursal))
                            == sucursal_nombre.lower()
                        )
                    )
                    if sucursal_row:
                        persona = self.db.scalar(
                            select(SucursalPersonaAutorizada)
                            .where(SucursalPersonaAutorizada.sucursal_id == sucursal_row.id)
                            .order_by(SucursalPersonaAutorizada.id.asc())
                        )
                        if persona:
                            if not salida["contacto"]:
                                salida["contacto"] = str(persona.nombre or "").strip()
                            if not salida["telefono"]:
                                salida["telefono"] = str(persona.telefono or "").strip()
                            if not salida["correo"]:
                                salida["correo"] = str(persona.email or "").strip()
                except Exception:
                    self.db.rollback()

        if not cliente and not venta_row:
            return salida

        row_reg = self.db.scalar(select(Registro).where(Registro.odt == (odt or "").strip()))
        if row_reg and row_reg.direccion:
            salida["direccion"] = str(row_reg.direccion or "").strip()

        row = self.db.scalar(select(ClienteBBDD).where(ClienteBBDD.cliente == cliente))
        if row:
            salida["direccion"] = salida["direccion"] or (row.direccion or "")
            if not salida["contacto"]:
                salida["contacto"] = row.nombre_representante or ""
            if not salida["correo"]:
                salida["correo"] = row.email_representante or ""

        for table_name in ["bbdd_clientes", "catalogo_clientes"]:
            try:
                for schema_name in self._schemas_con_tabla(table_name):
                    cols = self._columnas_tabla(schema_name, table_name)
                    if not cols:
                        continue
                    col_cliente = self._pick_col(cols, ["cliente", "nombre_sucursal", "sucursal", "nombre_cliente"])
                    if not col_cliente:
                        continue
                    col_lat = self._pick_col(cols, ["latitud", "lat", "latitude"])
                    col_lng = self._pick_col(cols, ["longitud", "lng", "lon", "longitude"])
                    col_layout = self._pick_col(cols, ["layout", "plano", "plano_url", "url_layout"])
                    col_obs = self._pick_col(cols, ["observacion", "observaciones", "nota"])
                    col_contacto = self._pick_col(cols, ["contacto", "nombre_contacto", "encargado"])
                    col_tel = self._pick_col(cols, ["telefono", "telefono_contacto", "celular", "fono"])
                    col_correo = self._pick_col(cols, ["correo", "email", "mail"])
                    col_dir = self._pick_col(cols, ["direccion", "direccion_sucursal"])

                    select_cols = []
                    if col_dir:
                        select_cols.append(f'COALESCE(CAST("{col_dir}" AS NVARCHAR(MAX)), \'\') AS direccion')
                    if col_contacto:
                        select_cols.append(f'COALESCE(CAST("{col_contacto}" AS NVARCHAR(MAX)), \'\') AS contacto')
                    if col_tel:
                        select_cols.append(f'COALESCE(CAST("{col_tel}" AS NVARCHAR(MAX)), \'\') AS telefono')
                    if col_correo:
                        select_cols.append(f'COALESCE(CAST("{col_correo}" AS NVARCHAR(MAX)), \'\') AS correo')
                    if col_lat:
                        select_cols.append(f'COALESCE(CAST("{col_lat}" AS NVARCHAR(MAX)), \'\') AS latitud')
                    if col_lng:
                        select_cols.append(f'COALESCE(CAST("{col_lng}" AS NVARCHAR(MAX)), \'\') AS longitud')
                    if col_layout:
                        select_cols.append(f'COALESCE(CAST("{col_layout}" AS NVARCHAR(MAX)), \'\') AS layout')
                    if col_obs:
                        select_cols.append(f'COALESCE(CAST("{col_obs}" AS NVARCHAR(MAX)), \'\') AS observacion')
                    if not select_cols:
                        continue

                    sql = text(
                        f"""
                        SELECT {", ".join(select_cols)}
                        FROM "{schema_name}"."{table_name}"
                        WHERE TRIM(CAST("{col_cliente}" AS NVARCHAR(MAX))) = :cliente
                        ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                        """
                    )
                    row_sql = self.db.execute(sql, {"cliente": cliente}).mappings().first()
                    if not row_sql:
                        continue

                    for key in ["direccion", "contacto", "telefono", "correo", "latitud", "longitud", "layout", "observacion"]:
                        val = str(row_sql.get(key) or "").strip()
                        if val and not salida.get(key):
                            salida[key] = val
            except Exception:
                self.db.rollback()
                continue

        contacto_pref = self._contacto_preferente_sucursal(cliente)
        for key in ["contacto", "telefono", "correo"]:
            if not salida.get(key):
                salida[key] = str(contacto_pref.get(key) or "").strip()

        if not salida["direccion"] and cliente:
            salida["direccion"] = self._direccion_cliente(cliente)

        if salida["direccion"] and (not salida["latitud"] or not salida["longitud"]):
            lat_bd, lng_bd = self._coordenadas_por_direccion_bd(salida["direccion"])
            if lat_bd and lng_bd:
                salida["latitud"] = lat_bd
                salida["longitud"] = lng_bd

        if salida["direccion"] and (not salida["latitud"] or not salida["longitud"]):
            lat_geo, lng_geo = self._geocodificar_direccion(salida["direccion"])
            if lat_geo and lng_geo:
                salida["latitud"] = lat_geo
                salida["longitud"] = lng_geo
                self._persistir_coordenadas_sucursal(
                    cliente=cliente,
                    direccion=salida["direccion"],
                    latitud=lat_geo,
                    longitud=lng_geo,
                )

        if salida["direccion"] and (not salida["latitud"] or not salida["longitud"]):
            lat_apx, lng_apx = self._coordenadas_aproximadas_por_direccion(salida["direccion"])
            if lat_apx and lng_apx:
                salida["latitud"] = lat_apx
                salida["longitud"] = lng_apx

        return salida

    def obtener_ultimas_incidencias_sucursal(self, nombre_sucursal: str) -> list[dict[str, str]]:
        sucursal = (nombre_sucursal or "").strip()
        if not sucursal:
            return []
        objetivo = self._normalizar_texto(sucursal)
        incidencias: list[dict[str, str]] = []

        def _fmt_fecha_texto(valor: Any) -> str:
            if isinstance(valor, datetime):
                return _to_ddmmyyyy_hhmm(valor)
            if valor is None:
                return ""
            raw = str(valor).strip()
            if not raw:
                return ""
            try:
                dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
                return _to_ddmmyyyy_hhmm(dt)
            except Exception:
                return raw

        try:
            for schema_name in self._schemas_con_tabla("incidencias"):
                cols = self._columnas_tabla(schema_name, "incidencias")
                if not cols:
                    continue
                col_cliente = self._pick_col(cols, ["cliente", "nombre_sucursal", "sucursal", "nombre_cliente"])
                col_fecha = self._pick_col(cols, ["fecha", "fecha_registro", "created_at", "createdat"])
                col_prob = self._pick_col(cols, ["tipo_incidencia", "problema", "tipo", "servicio", "incidencia"])
                col_deriv = self._pick_col(cols, ["derivacion", "servicio", "area"])
                col_obs = self._pick_col(cols, ["observacion_final", "descripcion", "detalle", "observacion", "detalle_problema"])
                col_odt = self._pick_col(cols, ["odt", "codigo_odt", "codigo", "nro_odt"])
                if not col_cliente or not col_obs:
                    continue

                select_cols = [
                    f'COALESCE(CAST("{col_cliente}" AS NVARCHAR(MAX)), \'\') AS cliente',
                    f'"{col_fecha}" AS fecha' if col_fecha else "NULL AS fecha",
                    f'COALESCE(CAST("{col_prob}" AS NVARCHAR(MAX)), \'\') AS problema' if col_prob else "'' AS problema",
                    f'COALESCE(CAST("{col_deriv}" AS NVARCHAR(MAX)), \'\') AS derivacion' if col_deriv else "'' AS derivacion",
                    f'COALESCE(CAST("{col_obs}" AS NVARCHAR(MAX)), \'\') AS texto',
                    f'COALESCE(CAST("{col_odt}" AS NVARCHAR(MAX)), \'\') AS odt' if col_odt else "'' AS odt",
                ]

                sql = text(
                    f"""
                    SELECT {", ".join(select_cols)}
                    FROM "{schema_name}"."incidencias"
                    WHERE "{col_cliente}" IS NOT NULL
                      AND TRIM(CAST("{col_cliente}" AS NVARCHAR(MAX))) <> ''
                    """
                )
                for row in self.db.execute(sql).mappings().all():
                    cli = str(row.get("cliente") or "").strip()
                    if self._normalizar_texto(cli) != objetivo:
                        continue
                    deriv = self._normalizar_texto(row.get("derivacion"))
                    if deriv and "servicio tecnico" not in deriv and "servicio tÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â©cnico" not in deriv:
                        continue
                    texto = str(row.get("texto") or "").strip()
                    if not texto:
                        continue
                    incidencias.append(
                        {
                            "fecha": _fmt_fecha_texto(row.get("fecha")),
                            "problema": str(row.get("problema") or "").strip(),
                            "texto": texto,
                            "odt": str(row.get("odt") or "").strip(),
                        }
                    )
        except Exception:
            # En PostgreSQL, una consulta fallida deja la transaccion abortada;
            # limpiamos la sesion antes de usar el fallback ORM sobre registro.
            self.db.rollback()

        if not incidencias:
            rows = self.db.scalars(select(Registro).where(Registro.cliente == sucursal).order_by(Registro.fecha_registro.asc())).all()
            for r in rows:
                deriv = self._normalizar_texto(r.derivacion)
                if deriv and "servicio tecnico" not in deriv and "servicio tÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â©cnico" not in deriv:
                    continue
                texto = (r.observacion_final or r.observacion or "").strip()
                if not texto:
                    continue
                incidencias.append(
                    {
                        "fecha": _to_ddmmyyyy_hhmm(r.fecha_registro),
                        "problema": r.problema or "",
                        "texto": texto,
                        "odt": r.odt or "",
                    }
                )

        def _sort_key(item: dict[str, str]) -> tuple[int, str]:
            fecha = item.get("fecha") or ""
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})(?:\s+(\d{1,2}):(\d{2}))?$", fecha)
            if not m:
                return (0, fecha)
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            hh = int(m.group(4) or 0)
            mm = int(m.group(5) or 0)
            return (int(datetime(y, mo, d, hh, mm).timestamp()), fecha)

        incidencias.sort(key=_sort_key)
        return incidencias

    @staticmethod
    def _parse_image_list(value: object) -> list[str]:
        parsed_images: list[str] = []
        if isinstance(value, list):
            parsed_images = [str(v or "").strip() for v in value]
        elif isinstance(value, str):
            raw = value.strip()
            if raw:
                try:
                    decoded = json.loads(raw)
                    if isinstance(decoded, list):
                        parsed_images = [str(v or "").strip() for v in decoded]
                    else:
                        parsed_images = [raw]
                except Exception:
                    parsed_images = [raw]
        else:
            parsed_images = [str(value or "").strip()]

        unique_images: list[str] = []
        for image_url in parsed_images:
            clean = str(image_url or "").strip()
            if clean and clean not in unique_images:
                unique_images.append(clean)
        return unique_images

    def _upsert_unified_images(
        self,
        odt: str,
        sucursal: str,
        usuario: str,
        imagenes: list[str],
        max_imagenes: int = 3,
    ) -> None:
        row_imgs = self.db.scalar(select(IncidenciaImagenTabla).where(IncidenciaImagenTabla.odt == odt))
        limite = max(1, int(max_imagenes or 3))
        payload = json.dumps(imagenes[:limite], ensure_ascii=False)
        if row_imgs:
            row_imgs.sucursal = sucursal or row_imgs.sucursal or None
            row_imgs.imagenes = payload
            row_imgs.created_by = usuario
            row_imgs.updated_at = datetime.now()
        else:
            self.db.add(
                IncidenciaImagenTabla(
                    odt=odt,
                    sucursal=sucursal or None,
                    imagenes=payload,
                    created_by=usuario,
                )
            )

    def _reset_unified_images_if_odt_reused(self, odt: str, sucursal: str) -> bool:
        odt_limpia = str(odt or "").strip()
        sucursal_limpia = str(sucursal or "").strip()
        if not odt_limpia:
            return False

        row_imgs = self.db.scalar(select(IncidenciaImagenTabla).where(IncidenciaImagenTabla.odt == odt_limpia))
        if not row_imgs:
            return False

        sucursal_actual = str(row_imgs.sucursal or "").strip()
        if self._normalizar_sucursal_key(sucursal_actual) == self._normalizar_sucursal_key(sucursal_limpia):
            return False

        row_imgs.sucursal = sucursal_limpia or None
        row_imgs.imagenes = "[]"
        row_imgs.created_by = "reset_odt_reutilizada"
        row_imgs.updated_at = datetime.now()
        self.db.commit()
        return True

    @staticmethod
    def _es_registro_mantencion_preventiva(row_odt: Registro | None) -> bool:
        if not row_odt:
            return False
        txt = str(getattr(row_odt, "problema", "") or "").strip().lower()
        txt = unicodedata.normalize("NFD", txt)
        txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
        return txt == "mantencion preventiva"

    def obtener_imagenes_tabla(self, odt: str) -> list[str]:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            return []
        row_odt = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        row = self.db.scalar(select(IncidenciaImagenTabla).where(IncidenciaImagenTabla.odt == odt_limpia))
        if self._es_registro_mantencion_preventiva(row_odt):
            sucursal_real = str(getattr(row_odt, "cliente", "") or "").strip()
            unified_images = [
                self._normalizar_url_imagen(u) for u in self._parse_image_list(row.imagenes if row else "[]")
            ]
            drive_images = list_support_images_for_odt(
                odt=odt_limpia,
                root_folder_id=str(settings.google_drive_support_folder_id or "").strip(),
            )
            drive_images = [self._normalizar_url_imagen(u) for u in (drive_images or [])]

            merged_odt: list[str] = []
            for img in [*drive_images, *unified_images]:
                url = str(img or "").strip()
                if not url or url in merged_odt:
                    continue
                merged_odt.append(url)
                if len(merged_odt) >= 3:
                    break

            if merged_odt:
                if merged_odt != unified_images[:3]:
                    self._upsert_unified_images(
                        odt=odt_limpia,
                        sucursal=sucursal_real,
                        usuario="sync_mantencion_por_odt",
                        imagenes=merged_odt,
                    )
                    self.db.commit()
                return merged_odt[:3]

            imagenes_sucursal = self._imagenes_programadas_para_sucursal(sucursal_real)
            imagenes_publicas = [
                self._normalizar_url_imagen(str(url or "").strip())
                for url in (imagenes_sucursal or [])
                if self._es_url_publica_imagen(str(url or "").strip())
            ][:3]
            if imagenes_publicas:
                self._upsert_unified_images(
                    odt=odt_limpia,
                    sucursal=sucursal_real,
                    usuario="sync_mantencion_por_sucursal",
                    imagenes=imagenes_publicas,
                )
                self.db.commit()
            return imagenes_publicas

        unified_images = [
            self._normalizar_url_imagen(u) for u in self._parse_image_list(row.imagenes if row else "[]")
        ]
        legacy_images = [
            self._normalizar_url_imagen(str(url or "").strip())
            for url in (
                getattr(row_odt, "foto_1", None),
                getattr(row_odt, "foto_2", None),
                getattr(row_odt, "foto_3", None),
            )
            if str(url or "").strip()
        ]
        drive_images = list_support_images_for_odt(
            odt=odt_limpia,
            root_folder_id=str(settings.google_drive_support_folder_id or "").strip(),
        )
        drive_images = [self._normalizar_url_imagen(u) for u in (drive_images or [])]

        merged: list[str] = []
        for img in [*drive_images, *unified_images, *legacy_images]:
            url = str(img or "").strip()
            if not url or url in merged:
                continue
            merged.append(url)
            if len(merged) >= 3:
                break

        if merged and merged != unified_images[:3]:
            usuario_sync = "sync_unificado"
            sucursal_sync = str(getattr(row, "sucursal", "") or "").strip()
            if not sucursal_sync:
                row_odt = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
                sucursal_sync = str(getattr(row_odt, "cliente", "") or "").strip()
            self._upsert_unified_images(odt_limpia, sucursal_sync, usuario_sync, merged)
            self.db.commit()

        return merged[:3]

    @staticmethod
    def _normalizar_url_imagen(valor: str) -> str:
        """
        Normaliza URLs historicas de Google Drive a un proxy local.

        Drive suele responder 404 cuando el archivo no es publico (politicas de dominio).
        Con el proxy local, la app descarga la imagen con las credenciales del sistema.
        """
        raw = str(valor or "").strip()
        if not raw:
            return ""
        if raw.startswith("/api/incidencias/drive-image/"):
            return raw

        lower = raw.lower()
        file_id = ""
        try:
            if "drive.google.com" in lower or "drive.usercontent.google.com" in lower:
                parts = urlsplit(raw)
                qs = parse_qs(parts.query or "")
                file_id = (qs.get("id", [""]) or [""])[0]
                if not file_id and "/file/d/" in parts.path:
                    m = re.search(r"/file/d/([^/]+)", parts.path)
                    if m:
                        file_id = m.group(1)
        except Exception:
            file_id = ""

        file_id = str(file_id or "").strip()
        if file_id:
            return f"/api/incidencias/drive-image/{file_id}"
        return raw

    def subir_imagenes_tabla(
        self,
        odt: str,
        image_payloads: list[dict[str, object]],
        token: str = "",
    ) -> dict[str, Any]:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT es obligatoria.")

        row_odt = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row_odt:
            raise ValueError(f"ODT {odt_limpia} no encontrada.")

        incoming_images: list[dict[str, object]] = []
        for payload in image_payloads or []:
            content = payload.get("bytes")
            mime_type = str(payload.get("mime_type") or "").strip().lower()
            if not isinstance(content, (bytes, bytearray)) or not content:
                continue
            if not mime_type.startswith("image/"):
                continue
            incoming_images.append(
                {
                    "filename": str(payload.get("filename") or "imagen.png").strip() or "imagen.png",
                    "mime_type": mime_type,
                    "bytes": bytes(content),
                }
            )
        if not incoming_images:
            raise ValueError("Debes adjuntar al menos una imagen valida.")

        existing_images = self.obtener_imagenes_tabla(odt_limpia)

        remaining_slots = max(0, 3 - len(existing_images))
        if remaining_slots <= 0:
            raise ValueError("Esta ODT ya tiene 3 imagenes cargadas.")
        if len(incoming_images) > remaining_slots:
            raise ValueError(f"Solo puedes subir {remaining_slots} imagen(es) adicional(es) para esta ODT.")

        try:
            drive_result = upload_support_images_for_odt(
                odt=odt_limpia,
                image_payloads=incoming_images,
                root_folder_id=str(settings.google_drive_support_folder_id or "").strip(),
                start_index=len(existing_images) + 1,
            )
        except DriveReportError as exc:
            raise ValueError(f"No se pudo subir a Drive: {exc}") from exc
        except Exception as exc:
            raise ValueError(f"Error inesperado al subir imagenes: {exc}") from exc

        new_urls = [str(url or "").strip() for url in (drive_result.get("imagenes") or []) if str(url or "").strip()]
        if not new_urls:
            raise ValueError("No se pudo obtener URL publica de las imagenes subidas.")

        merged_images = existing_images[:]
        for url in new_urls:
            if url not in merged_images:
                merged_images.append(url)
        merged_images = merged_images[:3]

        usuario = self.get_usuario_actual((token or "").strip())
        if not usuario or usuario == "Desconocido":
            usuario = "Usuario no identificado"

        sucursal_value = str(row_odt.cliente or "").strip()
        self._upsert_unified_images(odt_limpia, sucursal_value, usuario, merged_images)

        self.db.commit()
        return {
            "ok": True,
            "odt": odt_limpia,
            "imagenes": merged_images,
            "imagenes_guardadas": len(new_urls),
            "total_imagenes": len(merged_images),
            "drive_folder_id": str(drive_result.get("folder_id") or ""),
            "drive_folder_name": str(drive_result.get("folder_name") or ""),
        }

    @staticmethod
    def _url_cierre_apertura(ruta_archivo: str) -> str:
        """ruta_archivo puede ser una URL completa ya armada (Drive proxy o
        cuarentena local, ambas con "/" al inicio) para filas nuevas, o un
        path relativo tipo "cierre_apertura/{cliente}/{archivo}" para filas
        anteriores a la migracion a Drive."""
        ruta = ruta_archivo or ""
        if ruta.startswith(("/", "http")):
            return ruta
        return f"/uploads/{ruta}"

    def guardar_imagen_cierre_apertura(
        self,
        client_id: str,
        client_name: str,
        content: bytes,
        token: str = "",
        usuario_fallback: str = "",
    ) -> dict[str, Any]:
        client_id_limpio = (client_id or "").strip()
        if not client_id_limpio:
            raise ValueError("client_id es obligatorio.")
        if not content:
            raise ValueError("Debes adjuntar una imagen valida.")

        usuario = self.get_usuario_actual((token or "").strip())
        if not usuario or usuario == "Desconocido":
            usuario = (usuario_fallback or "").strip() or "Usuario no identificado"

        nombre_cliente = (client_name or client_id_limpio).strip()
        ahora = datetime.now()
        stamp = ahora.strftime("%Y%m%d_%H%M%S")
        filename = f"{stamp}_{uuid.uuid4().hex[:8]}.png"

        try:
            drive_result = upload_cierre_apertura_image_to_drive(
                client_id=client_id_limpio,
                client_name=nombre_cliente,
                content=content,
                filename=filename,
                mime_type="image/png",
            )
            ruta_archivo = drive_result["public_uri"]
        except Exception:
            LOGGER.exception("No se pudo subir foto de apertura/cierre a Drive (client_id=%s)", client_id_limpio)
            ruta_archivo = _guardar_en_cuarentena_drive(
                "cierre_apertura",
                client_id_limpio,
                filename,
                content,
                {
                    "client_id": client_id_limpio,
                    "client_name": nombre_cliente,
                    "filename": filename,
                    "mime_type": "image/png",
                },
            )

        fila = CierreAperturaImagen(
            client_id=client_id_limpio,
            client_name=nombre_cliente,
            ruta_archivo=ruta_archivo,
            created_by=usuario,
            created_at=ahora,
        )
        self.db.add(fila)
        self.db.commit()
        self.db.refresh(fila)

        LOGGER.info("Foto apertura/cierre guardada: id=%s client_id=%s -> %s", fila.id, client_id_limpio, ruta_archivo)
        return {
            "ok": True,
            "id": fila.id,
            "url": self._url_cierre_apertura(fila.ruta_archivo),
            "created_by": fila.created_by,
            "created_at": fila.created_at.isoformat(),
        }

    def listar_imagenes_cierre_apertura(self, client_id: str = "") -> list[dict[str, Any]]:
        query = self.db.query(CierreAperturaImagen)
        client_id_limpio = (client_id or "").strip()
        if client_id_limpio:
            query = query.filter(CierreAperturaImagen.client_id == client_id_limpio)
        filas = query.order_by(CierreAperturaImagen.created_at.desc()).all()
        return [
            {
                "id": fila.id,
                "client_id": fila.client_id,
                "client_name": fila.client_name,
                "url": self._url_cierre_apertura(fila.ruta_archivo),
                "created_by": fila.created_by or "",
                "created_at": fila.created_at.isoformat() if fila.created_at else None,
            }
            for fila in filas
        ]

    def obtener_informes_cierre_odt(self) -> list[dict[str, Any]]:
        """ODTs con informe de cierre (pdf_url) y/o fotos, para el navegador
        tipo Drive de "Ver informe e imagenes ODT". Solo metadatos livianos;
        las imagenes se cargan por ODT via obtener_imagenes_finalizacion."""
        filas = self.db.execute(
            select(
                Registro.odt,
                Registro.cliente,
                Registro.problema,
                Registro.fecha_cierre,
                Registro.pdf_url,
                Registro.foto_1,
            )
            .where(
                or_(
                    func.coalesce(Registro.pdf_url, "") != "",
                    func.coalesce(Registro.foto_1, "") != "",
                )
            )
            .order_by(Registro.fecha_cierre.desc(), Registro.id.desc())
        ).all()
        items: list[dict[str, Any]] = []
        for odt, cliente, problema, fecha_cierre, pdf_url, foto_1 in filas:
            items.append(
                {
                    "odt": str(odt or "").strip(),
                    "cliente": str(cliente or "").strip() or "Sin cliente",
                    "problema": str(problema or "").strip()[:120],
                    "fecha_cierre": fecha_cierre.isoformat() if fecha_cierre else None,
                    "pdf_url": str(pdf_url or "").strip(),
                    "tiene_fotos": bool(str(foto_1 or "").strip()),
                }
            )
        return items

    def obtener_imagenes_finalizacion(self, odt: str) -> list[str]:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            return []
        urls: list[str] = []

        try:
            for schema_name in self._schemas_con_tabla("incidencias"):
                cols = self._columnas_tabla(schema_name, "incidencias")
                if not cols:
                    continue
                col_odt = self._pick_col(cols, ["odt", "codigo_odt", "codigo", "nro_odt"])
                if not col_odt:
                    continue
                col_img1 = self._pick_col(cols, ["foto_1", "foto1", "imagen_1", "imagen1", "url_foto_1"])
                col_img2 = self._pick_col(cols, ["foto_2", "foto2", "imagen_2", "imagen2", "url_foto_2"])
                col_img3 = self._pick_col(cols, ["foto_3", "foto3", "imagen_3", "imagen3", "url_foto_3"])
                col_single = self._pick_col(cols, ["foto", "foto_url", "imagen_url"])
                cols_select = []
                if col_img1:
                    cols_select.append(f'COALESCE(CAST("{col_img1}" AS NVARCHAR(MAX)), \'\') AS img1')
                if col_img2:
                    cols_select.append(f'COALESCE(CAST("{col_img2}" AS NVARCHAR(MAX)), \'\') AS img2')
                if col_img3:
                    cols_select.append(f'COALESCE(CAST("{col_img3}" AS NVARCHAR(MAX)), \'\') AS img3')
                if col_single:
                    cols_select.append(f'COALESCE(CAST("{col_single}" AS NVARCHAR(MAX)), \'\') AS img_single')
                if not cols_select:
                    continue
                sql = text(
                    f"""
                    SELECT {", ".join(cols_select)}
                    FROM "{schema_name}"."incidencias"
                    WHERE TRIM(CAST("{col_odt}" AS NVARCHAR(MAX))) = :odt
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                )
                row = self.db.execute(sql, {"odt": odt_limpia}).mappings().first()
                if not row:
                    continue
                for k in ["img1", "img2", "img3", "img_single"]:
                    v = str(row.get(k) or "").strip()
                    if v and v not in urls:
                        urls.append(v)
                if urls:
                    return urls
        except Exception:
            pass

        reg = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if reg:
            for v in [reg.foto_1, reg.foto_2, reg.foto_3]:
                s = str(v or "").strip()
                if s and s not in urls:
                    urls.append(s)
        row_imgs = self.db.scalar(select(IncidenciaImagenTabla).where(IncidenciaImagenTabla.odt == odt_limpia))
        if row_imgs:
            for v in self._parse_image_list(row_imgs.imagenes):
                s = str(v or "").strip()
                if s and s not in urls:
                    urls.append(s)
        if urls:
            return urls

        return []

    def marcar_inicio_trabajo(self, odt: str, token: str = "") -> str:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida")

        ahora = datetime.now()
        usuario_token = ""
        if (token or "").strip():
            usuario_token = str(self.get_usuario_actual((token or "").strip()) or "").strip()
            if usuario_token == "Desconocido":
                usuario_token = ""

        row = self.db.scalar(select(Registro).where(func.lower(func.trim(Registro.odt)) == odt_limpia.lower()))
        venta_row = self.db.scalar(
            select(VentaODS).where(func.lower(func.trim(VentaODS.codigo)) == odt_limpia.lower())
        )
        st_row = None
        if venta_row:
            st_row = self.db.scalar(
                select(ServicioTecnicoVentaODT).where(
                    func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == odt_limpia.lower()
                )
            )
            if not st_row:
                st_row = ServicioTecnicoVentaODT(odt=odt_limpia)
                self.db.add(st_row)

            if (
                getattr(st_row, "finalizado", False)
                or getattr(st_row, "instalacion_finalizada", False)
                or getattr(st_row, "fecha_cierre", None)
                or (row and str(getattr(row, "estado", "") or "").strip().lower() in {"terminado", "finalizado"})
            ):
                raise ValueError(f"La ODT {odt_limpia} ya esta cerrada.")

            st_row.fecha_inicio_trabajo = getattr(st_row, "fecha_inicio_trabajo", None) or ahora
            st_row.fecha_fin_trabajo = None
            if not str(getattr(st_row, "tecnico_a_cargo", "") or "").strip() and usuario_token:
                st_row.tecnico_a_cargo = usuario_token

        if not row:
            if not venta_row:
                raise ValueError(f"No se encontro la ODT {odt_limpia}")

            row = Registro(
                odt=odt_limpia,
                fecha_registro=ahora,
                puesto=None,
                cliente=str(venta_row.nombre_sucursal or venta_row.razon_social or "").strip()
                or str(venta_row.razon_social or "").strip(),
                problema=str(venta_row.tipo_servicio or "").strip() or "Servicio Tecnico",
                detalle_problema=str(venta_row.observacion or venta_row.consideraciones or "").strip() or None,
                derivacion="Servicio Técnico",
                observacion=str(venta_row.observacion or venta_row.consideraciones or "").strip() or None,
                observacion_soporte=None,
                observacion_servicio=None,
                tecnicos=str(getattr(st_row, "tecnico_a_cargo", "") or "").strip() or usuario_token or None,
                acompanante=str(getattr(st_row, "acompanante", "") or "").strip() or None,
                estado="En Proceso",
                fecha_derivacion_area=ahora,
                fecha_derivacion_tecnico=ahora,
                direccion=str(venta_row.direccion_sucursal or "").strip() or None,
            )
            self.db.add(row)

        row.fecha_inicio_trabajo = row.fecha_inicio_trabajo or ahora
        row.fecha_fin_trabajo = None
        self.db.commit()
        return "OK"

    def guardar_datos_en_proceso(
        self,
        odt: str,
        avance: int,
        observacion: str,
        token: str = "",
        camaras_instaladas: int | None = None,
    ) -> str:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida")

        usuario_token = ""
        if (token or "").strip():
            usuario_token = str(self.get_usuario_actual((token or "").strip()) or "").strip()
            if usuario_token == "Desconocido":
                usuario_token = ""

        # Se busca siempre (no solo si falta el Registro shadow) porque el aviso al
        # comercial + borrado de asignación de más abajo debe aplicar cada vez que
        # se deja pendiente una ODS de venta, no únicamente la primera vez.
        venta_row = None
        if odt_limpia[:1].upper() == "V":
            venta_row = self.db.scalar(select(VentaODS).where(func.lower(func.trim(VentaODS.codigo)) == odt_limpia.lower()))

        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            if not venta_row:
                raise ValueError(f"No se encontro la ODT {odt_limpia}")
            st_row = self.db.scalar(
                select(ServicioTecnicoVentaODT).where(
                    func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == odt_limpia.lower()
                )
            )
            row = Registro(
                odt=odt_limpia,
                fecha_registro=datetime.now(),
                puesto=None,
                cliente=str(venta_row.nombre_sucursal or venta_row.razon_social or "").strip() or str(venta_row.razon_social or "").strip(),
                problema=str(venta_row.tipo_servicio or "").strip() or "Servicio Tecnico",
                detalle_problema=str(venta_row.observacion or venta_row.consideraciones or "").strip() or None,
                derivacion=str(venta_row.tipo_servicio or "").strip() or "Servicio Tecnico",
                observacion=str(venta_row.observacion or venta_row.consideraciones or "").strip() or None,
                observacion_soporte=None,
                observacion_servicio=None,
                tecnicos=str(getattr(st_row, "tecnico_a_cargo", "") or "").strip() or usuario_token or None,
                acompanante=str(getattr(st_row, "acompanante", "") or "").strip() or None,
                estado="En Proceso",
                dias_ejecucion=None,
                fecha_cierre=None,
                fecha_derivacion_area=datetime.now(),
                fecha_derivacion_tecnico=datetime.now(),
                direccion=str(venta_row.direccion_sucursal or "").strip() or None,
                observacion_final=None,
                observacion_pendiente=None,
                prioridad=None,
                materiales=None,
                responsable_cierre=None,
                causa_cierre=None,
                accion_cierre=None,
                resultado_cierre=None,
                pruebas_cierre=None,
                requiere_seguimiento=None,
                porcentaje_avance=None,
                foto_1=None,
                foto_2=None,
                foto_3=None,
                drive_cierre_folder_id=None,
                drive_cierre_folder_url=None,
                pdf_url=None,
            )
            self.db.add(row)
            self.db.flush()

        estado_previo = str(getattr(row, "estado", "") or "").strip()
        avance_num = max(0, min(100, int(avance)))
        marca = datetime.now().strftime("%d/%m/%Y %H:%M")
        usuario = (self.get_usuario_actual((token or "").strip()) if (token or "").strip() else "").strip()
        if not usuario or usuario == "Desconocido":
            usuario = str(getattr(row, "tecnicos", "") or "").strip() or str(getattr(row, "acompanante", "") or "").strip() or usuario_token
        if not usuario:
            usuario = "Usuario no identificado"
        nota = f"[{usuario} - {marca}] {observacion.strip()} (Avance: {avance_num}%)"

        row.estado = "Pendiente"
        ahora_fin = datetime.now()
        if row.fecha_inicio_trabajo and not row.fecha_fin_trabajo:
            row.fecha_fin_trabajo = ahora_fin
        if not str(getattr(row, "tecnicos", "") or "").strip() and usuario_token:
            row.tecnicos = usuario_token
        if usuario_token:
            # Quien realmente dejó la ODT en Pendiente, aunque haya sido
            # derivada a otro técnico (row.tecnicos no se toca arriba).
            row.tecnico_cierre = usuario_token
        row.porcentaje_avance = f"{avance_num}%"
        base = (getattr(row, "observacion_pendiente", "") or "").strip()
        row.observacion_pendiente = f"{base}\n{nota}".strip() if base else nota

        # Si el registro quedo con tecnico (por el fallback al usuario logueado
        # de arriba) pero el ODS de venta asociado no tiene tecnico_a_cargo,
        # se sincroniza: evita que "tabla servicio tecnico venta" muestre el
        # ODT sin tecnico mientras "resumen_equipos_tecnicos" si lo muestra.
        tecnico_actual = str(getattr(row, "tecnicos", "") or "").strip()
        acompanante_actual = str(getattr(row, "acompanante", "") or "").strip()
        fecha_inicio_trabajo_actual = getattr(row, "fecha_inicio_trabajo", None)
        st_sync = self.db.scalar(
            select(ServicioTecnicoVentaODT).where(
                func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == odt_limpia.lower()
            )
        )
        if st_sync:
            if tecnico_actual and not str(st_sync.tecnico_a_cargo or "").strip():
                st_sync.tecnico_a_cargo = tecnico_actual
            if getattr(st_sync, "fecha_inicio_trabajo", None) and not getattr(st_sync, "fecha_fin_trabajo", None):
                st_sync.fecha_fin_trabajo = ahora_fin

        tecnico_para_email = ""
        camaras_total_para_email = 0
        camaras_instaladas_para_email = 0
        if venta_row:
            tecnico_para_email = tecnico_actual
            camaras_total_para_email = int(getattr(venta_row, "numero_camaras_instalar", 0) or 0)
            if camaras_instaladas is not None:
                camaras_instaladas_para_email = max(0, int(camaras_instaladas))
            elif camaras_total_para_email:
                camaras_instaladas_para_email = round(avance_num / 100 * camaras_total_para_email)
            # Se borra la asignación para que esta ODS quede sin técnico y el
            # encargado la vuelva a derivar al día siguiente, en vez de quedar
            # "colgada" asignada a alguien que ya no la va a retomar solo.
            row.tecnicos = None
            row.acompanante = None
            if st_sync:
                st_sync.tecnico_a_cargo = None
                st_sync.acompanante = None

        self.db.commit()

        if estado_previo != "Pendiente":
            self._reforzar_inicio_odt_si_corresponde(
                odt=odt_limpia,
                tecnico=tecnico_actual,
                acompanante=acompanante_actual,
                usuario_accion=row.tecnico_cierre,
                fecha_inicio_trabajo=fecha_inicio_trabajo_actual,
                verbo_accion="dejada en pendiente",
            )

        if venta_row:
            _codigo_bg = odt_limpia
            _obs_bg = observacion.strip()
            _avance_bg = avance_num
            _camaras_instaladas_bg = camaras_instaladas_para_email
            _camaras_total_bg = camaras_total_para_email
            _tecnico_bg = tecnico_para_email

            def _bg_pendiente():
                from ATC.app.core.db import SessionLocal
                from ATC.app.services.venta_trace_email_service import notify_odt_dejada_pendiente
                _db = SessionLocal()
                try:
                    notify_odt_dejada_pendiente(
                        _db, _codigo_bg, _obs_bg, _avance_bg,
                        _camaras_instaladas_bg, _camaras_total_bg, _tecnico_bg,
                    )
                except Exception as exc:
                    LOGGER.warning("notify_odt_dejada_pendiente %s falló: %s", _codigo_bg, exc)
                finally:
                    _db.close()

            threading.Thread(target=_bg_pendiente, daemon=True).start()

        return "OK"

    def _obtener_snapshot_cierre_odt(self, odt: str) -> dict[str, str]:
        odt_limpia = (odt or "").strip()
        out = {
            "odt": odt_limpia,
            "cliente": "",
            "sucursal": "",
            "problema": "",
            "direccion": "",
            "tecnico": "",
            "acompanante": "",
            "fecha_cierre": "",
            "observacion_final": "",
            "rut_cliente": "-",
        }

        row_reg = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if row_reg:
            out["odt"] = str(row_reg.odt or odt_limpia).strip()
            out["cliente"] = str(row_reg.cliente or "").strip()
            out["sucursal"] = out["cliente"]
            out["problema"] = str(row_reg.problema or "").strip()
            out["direccion"] = str(row_reg.direccion or "").strip()
            out["tecnico"] = str(row_reg.tecnicos or "").strip()
            out["acompanante"] = str(row_reg.acompanante or "").strip()
            out["observacion_final"] = str(row_reg.observacion_final or "").strip()
            if isinstance(row_reg.fecha_cierre, datetime):
                out["fecha_cierre"] = _to_ddmmyyyy_hhmm(row_reg.fecha_cierre)

        if not out["cliente"]:
            out["cliente"] = self._buscar_cliente_por_odt(odt_limpia)
            out["sucursal"] = out["cliente"]

        # Enriquecer desde bbdd_sucursales: dirección, razón social, rut
        if out["cliente"]:
            try:
                row_suc = self.db.execute(
                    text(
                        "SELECT "
                        "  COALESCE(nombre_empresa,'') AS nombre_empresa, "
                        "  COALESCE(direccion_sucursal,'') AS direccion, "
                        "  COALESCE(rut,'') AS rut "
                        "FROM bbdd_sucursales "
                        "WHERE lower(TRIM(nombre_sucursal)) = lower(:suc) ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY"
                    ),
                    {"suc": out["cliente"]},
                ).mappings().first()
                if row_suc:
                    if str(row_suc.get("nombre_empresa") or "").strip():
                        out["razon_social"] = str(row_suc["nombre_empresa"]).strip()
                    if str(row_suc.get("direccion") or "").strip():
                        out["direccion"] = str(row_suc["direccion"]).strip()
                    if str(row_suc.get("rut") or "").strip():
                        out["rut_cliente"] = str(row_suc["rut"]).strip()
            except Exception:
                self.db.rollback()

        if not out["direccion"]:
            out["direccion"] = self._direccion_cliente(out["cliente"])

        try:
            row_venta = self.db.execute(
                select(VentaODS, ServicioTecnicoVentaODT, AdministracionODT)
                .outerjoin(
                    ServicioTecnicoVentaODT,
                    func.lower(func.trim(ServicioTecnicoVentaODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                )
                .outerjoin(
                    AdministracionODT,
                    func.lower(func.trim(AdministracionODT.odt)) == func.lower(func.trim(VentaODS.codigo)),
                )
                .where(func.lower(func.trim(VentaODS.codigo)) == odt_limpia.lower())
            ).first()
            if row_venta:
                ods_row, st_row, adm_row = row_venta
                out["cliente"] = out["cliente"] or str(ods_row.nombre_sucursal or ods_row.razon_social or "").strip()
                out["sucursal"] = out["sucursal"] or out["cliente"]
                out["problema"] = out["problema"] or str(ods_row.tipo_servicio or "").strip()
                out["direccion"] = out["direccion"] or str(ods_row.direccion_sucursal or "").strip()
                out["tecnico"] = out["tecnico"] or str(
                    getattr(st_row, "tecnico_a_cargo", None) or getattr(adm_row, "tecnico", None) or ""
                ).strip()
                out["acompanante"] = out["acompanante"] or str(
                    getattr(st_row, "acompanante", None) or getattr(adm_row, "acompanante", None) or ""
                ).strip()
                if not out["fecha_cierre"]:
                    fecha_cierre_venta = (
                        getattr(st_row, "fecha_cierre", None)
                        or getattr(adm_row, "fecha_cierre", None)
                        or getattr(ods_row, "updated_at", None)
                    )
                    if isinstance(fecha_cierre_venta, datetime):
                        out["fecha_cierre"] = _to_ddmmyyyy_hhmm(fecha_cierre_venta)
        except Exception:
            self.db.rollback()

        if not out["fecha_cierre"]:
            tz_name = (settings.timezone or "America/Santiago").strip() or "America/Santiago"
            out["fecha_cierre"] = datetime.now(ZoneInfo(tz_name)).strftime("%d/%m/%Y %H:%M")

        self._enriquecer_snapshot_desde_catalogo(out)
        return out

    def _enriquecer_snapshot_desde_catalogo(self, out: dict[str, str]) -> None:
        cliente = str(out.get("cliente") or "").strip()
        if not cliente:
            return

        try:
            row_bbdd = self.db.execute(
                text(
                    """
                    SELECT
                        COALESCE(CAST(rut AS NVARCHAR(MAX)), '') AS rut,
                        COALESCE(CAST(direccion AS NVARCHAR(MAX)), '') AS direccion
                    FROM bbdd_clientes
                    WHERE lower(TRIM(CAST(cliente AS NVARCHAR(MAX)))) = lower(:cliente)
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                ),
                {"cliente": cliente},
            ).mappings().first()
            if row_bbdd:
                rut_txt = str(row_bbdd.get("rut") or "").strip()
                dir_txt = str(row_bbdd.get("direccion") or "").strip()
                if rut_txt:
                    out["rut_cliente"] = rut_txt
                if dir_txt:
                    out["direccion"] = dir_txt

            schema_preferido = (getattr(settings, "db_schema", None) or "public").strip() or "public"
            row_catalogo_preferido = self.db.execute(
                text(
                    f"""
                    SELECT
                        COALESCE(CAST(rut_cliente AS NVARCHAR(MAX)), '') AS rut,
                        COALESCE(CAST(direccion_sucursal AS NVARCHAR(MAX)), '') AS direccion
                    FROM "{schema_preferido}"."catalogo_clientes"
                    WHERE lower(TRIM(CAST(nombre_sucursal AS NVARCHAR(MAX)))) = lower(:cliente)
                       OR lower(TRIM(CAST(nombre_cliente AS NVARCHAR(MAX)))) = lower(:cliente)
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                ),
                {"cliente": cliente},
            ).mappings().first()
            if row_catalogo_preferido:
                rut_txt = str(row_catalogo_preferido.get("rut") or "").strip()
                dir_txt = str(row_catalogo_preferido.get("direccion") or "").strip()
                if rut_txt:
                    out["rut_cliente"] = rut_txt
                if dir_txt:
                    out["direccion"] = dir_txt
                return

            for schema_name in self._schemas_con_tabla("catalogo_clientes"):
                cols_cat = self._columnas_tabla(schema_name, "catalogo_clientes")
                if not cols_cat:
                    continue

                col_cliente_cat = self._pick_col(cols_cat, ["nombre_sucursal", "nombre_cliente", "sucursal", "cliente"])
                col_rut_cat = self._pick_col(cols_cat, ["rut_cliente", "rut", "rut_empresa", "rut_sucursal"])
                col_dir_cat = self._pick_col(cols_cat, ["direccion", "direccion_sucursal", "direccion_trabajos", "direccion_cliente"])
                if not col_cliente_cat or (not col_rut_cat and not col_dir_cat):
                    continue

                select_cols: list[str] = []
                if col_rut_cat:
                    select_cols.append(f'COALESCE(CAST("{col_rut_cat}" AS NVARCHAR(MAX)), '') AS rut')
                else:
                    select_cols.append("'' AS rut")
                if col_dir_cat:
                    select_cols.append(f'COALESCE(CAST("{col_dir_cat}" AS NVARCHAR(MAX)), '') AS direccion')
                else:
                    select_cols.append("'' AS direccion")

                sql_cat = text(
                    f"""
                    SELECT {", ".join(select_cols)}
                    FROM "{schema_name}"."catalogo_clientes"
                    WHERE lower(TRIM(CAST("{col_cliente_cat}" AS NVARCHAR(MAX)))) = lower(:cliente)
                    ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                )
                row_cat = self.db.execute(sql_cat, {"cliente": cliente}).mappings().first()
                if not row_cat:
                    continue

                rut_txt = str(row_cat.get("rut") or "").strip()
                dir_txt = str(row_cat.get("direccion") or "").strip()
                if rut_txt:
                    out["rut_cliente"] = rut_txt
                if dir_txt:
                    out["direccion"] = dir_txt
                break
        except Exception:
            self.db.rollback()

    def _guardar_pdf_url_odt(self, odt: str, pdf_url: str) -> None:
        odt_limpia = (odt or "").strip()
        url = (pdf_url or "").strip()
        if not odt_limpia or not url:
            return

        try:
            row_reg = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
            if row_reg:
                row_reg.pdf_url = url
                self.db.commit()
        except Exception:
            self.db.rollback()

    def _generar_pdf_local_cierre_odt(
        self,
        snapshot: dict[str, str],
        observacion: str,
        fotos: list[str],
    ) -> bytes | None:
        """Genera un PDF corporativo de cierre de ODT con reportlab.

        Devuelve los bytes del PDF (no se escribe a disco: lo sube a Drive
        el llamador) o None si falla la generacion.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib.colors import HexColor, white
            from reportlab.platypus import (
                BaseDocTemplate, Frame, PageTemplate,
                Table, TableStyle, Paragraph, Spacer, HRFlowable,
            )
            from reportlab.platypus import Image as RLImage
            from reportlab.lib.styles import ParagraphStyle
            import io
            from PIL import Image as PILImage

            C_DARK   = HexColor("#0b1424")
            C_ORANGE = HexColor("#f4a672")
            C_ORDK   = HexColor("#c2410c")
            C_BG     = HexColor("#f7f8fa")
            C_BORDER = HexColor("#e5e7eb")
            C_TEXT   = HexColor("#111827")
            C_SOFT   = HexColor("#4b5563")
            C_YELLOW = HexColor("#fde68a")
            C_GREY   = HexColor("#9ca3af")

            W, H = A4
            pad      = 1.4 * cm
            HEADER_H = 2.6 * cm
            ORANGE_H = 5
            FOOTER_H = 1.0 * cm
            BODY_TOP = HEADER_H + ORANGE_H + 10
            BODY_BOT = FOOTER_H + 8
            fw       = W - 2 * pad

            odt_num      = str(snapshot.get("odt") or "")
            sucursal     = str(snapshot.get("cliente") or "")
            direccion    = str(snapshot.get("direccion") or "-")
            problema     = str(snapshot.get("problema") or "-")
            fecha_cierre = str(snapshot.get("fecha_cierre") or "")
            tecnico      = str(snapshot.get("tecnico") or "-")
            acomp        = str(snapshot.get("acompanante") or "").strip()

            # Razón social, RUT y dirección vienen del snapshot (ya enriquecido por bbdd_sucursales)
            razon_social = str(snapshot.get("razon_social") or sucursal)
            rut_cliente  = str(snapshot.get("rut_cliente") or "-")
            if not direccion or direccion == "-":
                # Fallback si el snapshot llegó sin dirección
                try:
                    row_sf = self.db.execute(
                        text(
                            "SELECT COALESCE(nombre_empresa,'') AS nombre_empresa, "
                            "       COALESCE(direccion_sucursal,'') AS direccion, "
                            "       COALESCE(rut,'') AS rut "
                            "FROM bbdd_sucursales "
                            "WHERE lower(TRIM(nombre_sucursal)) = lower(:suc) ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY"
                        ),
                        {"suc": sucursal},
                    ).mappings().first()
                    if row_sf:
                        if str(row_sf.get("nombre_empresa") or "").strip():
                            razon_social = str(row_sf["nombre_empresa"]).strip()
                        if str(row_sf.get("direccion") or "").strip():
                            direccion = str(row_sf["direccion"]).strip()
                        if str(row_sf.get("rut") or "").strip() and rut_cliente == "-":
                            rut_cliente = str(row_sf["rut"]).strip()
                except Exception:
                    self.db.rollback()

            titulo_hdr    = "INFORME DE CIERRE DE ODT"
            subtitulo_hdr = f"{razon_social}  ·  ODT {odt_num}"
            fecha_emision = datetime.now().strftime("%d/%m/%Y %H:%M")

            logo_path = _ATC_ROOT / "static" / "img" / "logo-atc.png"
            logo_w, logo_h = 3.0 * cm, 1.5 * cm

            st_label = ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=7.5,
                                      textColor=C_SOFT, leading=10, spaceAfter=1)
            st_value = ParagraphStyle("val", fontName="Helvetica", fontSize=10,
                                      textColor=C_TEXT, leading=13)
            st_sec   = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=9,
                                      textColor=C_ORDK, leading=12, spaceBefore=12, spaceAfter=5)
            st_body  = ParagraphStyle("body", fontName="Helvetica", fontSize=9.5,
                                      textColor=C_SOFT, leading=14, spaceAfter=4)

            def draw_page(canvas, doc):
                canvas.saveState()
                canvas.setFillColor(C_DARK)
                canvas.rect(0, H - HEADER_H, W, HEADER_H, fill=1, stroke=0)
                if logo_path.exists():
                    try:
                        canvas.drawImage(
                            str(logo_path),
                            pad, H - HEADER_H + (HEADER_H - logo_h) / 2,
                            width=logo_w, height=logo_h,
                            preserveAspectRatio=True, mask="auto",
                        )
                    except Exception:
                        pass
                tx = pad + logo_w + 0.5 * cm
                canvas.setFillColor(white)
                canvas.setFont("Helvetica-Bold", 15)
                canvas.drawString(tx, H - HEADER_H + 1.35 * cm, titulo_hdr)
                canvas.setFillColor(C_YELLOW)
                canvas.setFont("Helvetica", 9)
                canvas.drawString(tx, H - HEADER_H + 0.52 * cm, subtitulo_hdr)
                canvas.setFillColor(C_ORANGE)
                canvas.rect(0, H - HEADER_H - ORANGE_H, W, ORANGE_H, fill=1, stroke=0)
                canvas.setFillColor(C_DARK)
                canvas.rect(0, 0, W, FOOTER_H, fill=1, stroke=0)
                canvas.setFillColor(C_GREY)
                canvas.setFont("Helvetica", 7)
                canvas.drawCentredString(
                    W / 2, FOOTER_H / 2 - 3,
                    f"Documento generado automáticamente  ·  Alguien Te Cuida  ·  {fecha_emision}",
                )
                canvas.restoreState()

            frame = Frame(
                pad, BODY_BOT, fw, H - BODY_TOP - BODY_BOT,
                leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
            )
            page_tmpl = PageTemplate(id="main", frames=[frame], onPage=draw_page)
            buf = io.BytesIO()
            doc = BaseDocTemplate(
                buf, pagesize=A4, pageTemplates=[page_tmpl],
                leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0,
                title=titulo_hdr, author="Alguien Te Cuida",
            )

            story: list = []
            sep = 0.4 * cm
            cw  = (fw - sep) / 2

            def field(lbl: str, val: str):
                return [Paragraph(lbl, st_label), Paragraph(str(val or "-"), st_value)]

            def hr():
                story.append(HRFlowable(width=fw, thickness=0.75, color=C_BORDER, spaceAfter=6))

            def detail_grid(filas: list) -> None:
                flat = []
                for row in filas:
                    lft = row[0]
                    rgt = row[1] if len(row) > 1 else [Paragraph("", st_label), Paragraph("", st_value)]
                    flat.append([lft[0], lft[1], Spacer(sep, 1), rgt[0], rgt[1]])
                t = Table(
                    flat,
                    colWidths=[cw * 0.36, cw * 0.64, sep, cw * 0.36, cw * 0.64],
                )
                t.setStyle(TableStyle([
                    ("VALIGN",         (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING",    (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
                    ("TOPPADDING",     (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING",  (0, 0), (-1, -1), 7),
                    ("LEFTPADDING",    (2, 0), (2, -1), 0),
                    ("RIGHTPADDING",   (2, 0), (2, -1), 0),
                    ("TOPPADDING",     (2, 0), (2, -1), 0),
                    ("BOTTOMPADDING",  (2, 0), (2, -1), 0),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [white, C_BG]),
                    ("LINEBELOW",      (0, 0), (-1, -2), 0.5, C_BORDER),
                    ("BOX",            (0, 0), (-1, -1), 1, C_BORDER),
                ]))
                story.append(t)

            # ── Fila 1: identificación ODT ───────────────────────────────
            # Fila 2: cliente (razón social) + RUT
            # Fila 3: sucursal + dirección
            # Fila 4: técnico + acompañante (siempre par, "-" si no aplica)
            # Fila 5: tipo de servicio (ancho completo)
            detail_grid([
                [field("NÚMERO DE ODT",    odt_num),
                 field("FECHA DE CIERRE",  fecha_cierre or fecha_emision)],
                [field("CLIENTE",          razon_social),
                 field("RUT CLIENTE",      rut_cliente)],
                [field("SUCURSAL",         sucursal),
                 field("DIRECCIÓN",        direccion)],
                [field("TÉCNICO A CARGO",  tecnico),
                 field("ACOMPAÑANTE",      acomp or "-")],
            ])
            story.append(Spacer(1, 2))
            # Fila tipo de servicio a ancho completo
            ts_t = Table(
                [[Paragraph("TIPO DE SERVICIO", st_label), Paragraph(problema, st_value)]],
                colWidths=[fw * 0.22, fw * 0.78],
            )
            ts_t.setStyle(TableStyle([
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                ("TOPPADDING",    (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("BACKGROUND",    (0, 0), (-1, -1), C_BG),
                ("BOX",           (0, 0), (-1, -1), 1, C_BORDER),
            ]))
            story.append(ts_t)
            story.append(Spacer(1, 10))

            # ── Observación de cierre ────────────────────────────────────
            _LABEL_MAP = {
                # responsable
                "atc": "ATC", "cliente": "Cliente",
                "proveedor_externo": "Proveedor Externo", "internet": "Internet", "otro": "Otro",
                # causas ATC
                "instalacion_deficiente": "Instalación deficiente",
                "configuracion_incorrecta": "Configuración incorrecta",
                "material_instalado_defectuoso": "Material instalado defectuoso",
                "mantenimiento_insuficiente": "Mantención preventiva insuficiente",
                "diagnostico_previo_incompleto": "Diagnóstico previo incompleto",
                # causas cliente
                "manipulacion_cliente": "Manipulación por parte del cliente",
                "problema_electrico_cliente": "Problema eléctrico en instalación del cliente",
                "red_internet_cliente": "Red o internet del cliente",
                "infraestructura_cliente": "Infraestructura del cliente",
                "dano_terceros": "Daño por terceros",
                # causas proveedor
                "falla_proveedor_servicio": "Falla del proveedor de servicio",
                "corte_programado_proveedor": "Corte programado por proveedor",
                "equipo_proveedor_defectuoso": "Equipo del proveedor defectuoso",
                # causas internet
                "corte_internet_zona": "Corte de internet en la zona",
                "intermitencia_enlace": "Intermitencia en el enlace",
                "falla_router_modem": "Falla en router o módem",
                # causas otro
                "fuerza_mayor": "Fuerza mayor",
                "vandalismo": "Vandalismo",
                "causa_no_determinada": "Causa no determinada",
                # acciones
                "reconexion": "Reconexión de equipos",
                "reconfiguracion": "Reconfiguración del sistema",
                "reemplazo_material": "Reemplazo de material",
                "ajuste_fisico": "Ajuste físico en terreno",
                "limpieza": "Limpieza y mantención",
                "cambio_cableado": "Cambio de cableado",
                "cambio_fuente": "Cambio de fuente de poder",
                "validacion_sin_intervencion": "Validación sin intervención física",
                # resultados
                "operativo": "Operativo",
                "operativo_con_observacion": "Operativo con observación",
                "requiere_seguimiento": "Requiere seguimiento",
                "requiere_cotizacion_visita_adicional": "Requiere cotización / visita adicional",
                # pruebas
                "camaras_ok": "Cámaras verificadas", "grabacion_ok": "Grabación verificada",
                "audio_ok": "Audio verificado", "red_ok": "Red verificada",
                "energia_ok": "Energía verificada",
                # legacy / compatibilidad
                "falla_hardware": "Falla de hardware", "falla_software": "Falla de software",
                "desgaste": "Desgaste natural", "error_usuario": "Error del usuario",
                "baja_bateria": "Batería baja",
                "configuracion": "Reconfiguración", "reinicio": "Reinicio de equipo",
                "actualizacion": "Actualización de firmware", "reemplazo": "Reemplazo de componente",
            }

            _UNIDAD_MAP = {
                "m": "metros", "mt": "metros", "mts": "metros", "metro": "metros",
                "u": "unidades", "un": "unidades", "unid": "unidades", "unidad": "unidades",
                "kg": "kilogramos", "g": "gramos", "ml": "mililitros", "l": "litros",
                "mt2": "m²", "m2": "m²", "cm": "centímetros",
            }

            def _label(code: str) -> str:
                c = str(code or "").strip().lower().replace(" ", "_")
                return _LABEL_MAP.get(c) or str(code).replace("_", " ").title()

            def _bullets(items: list[str]) -> str:
                """Convierte lista de labels a texto con viñetas para Paragraph (HTML-safe)."""
                if not items:
                    return "-"
                if len(items) == 1:
                    return items[0]
                return "<br/>".join(f"• {it}" for it in items)

            def _fmt_material(txt: str) -> str:
                """'Cable UTP x4 metros' → 'Cable UTP: 4 metros'"""
                import re as _re
                txt = txt.strip()
                m = _re.match(r"^(.+?)\s+x(\d+(?:[.,]\d+)?)\s*(.*)$", txt, _re.IGNORECASE)
                if not m:
                    return txt
                nombre, cant, unidad = m.group(1).strip(), m.group(2).strip(), m.group(3).strip().lower()
                unidad_fmt = _UNIDAD_MAP.get(unidad, unidad) if unidad else "unidades"
                return f"{nombre}: {cant} {unidad_fmt}".strip()

            raw_obs = str(observacion or snapshot.get("observacion_final") or "")
            obs_libre = raw_obs
            diag_data: dict[str, str] = {}

            if "\n\nDiagnostico estructurado:" in raw_obs:
                obs_libre, diag_raw = raw_obs.split("\n\nDiagnostico estructurado:", 1)
                for par in diag_raw.strip().split(";"):
                    if ":" in par:
                        k, _, v = par.partition(":")
                        diag_data[k.strip().lower()] = v.strip()

            story.append(Paragraph("OBSERVACIÓN DE CIERRE", st_sec))
            hr()
            obs_html = html_escape(obs_libre.strip() or "-").replace("\n", "<br/>")
            story.append(Paragraph(obs_html, st_body))
            story.append(Spacer(1, 10))

            if diag_data:
                story.append(Paragraph("DIAGNÓSTICO TÉCNICO", st_sec))
                hr()

                st_bullet = ParagraphStyle("bul", fontName="Helvetica", fontSize=9.5,
                                           textColor=C_TEXT, leading=14, spaceAfter=0)

                def diag_row(lbl: str, val: str):
                    return [Paragraph(lbl, st_label), Paragraph(val or "-", st_bullet)]

                causa_items  = [_label(c) for c in diag_data.get("causa", "").split("|")   if c.strip()]
                accion_items = [_label(a) for a in diag_data.get("accion", "").split("|")  if a.strip()]
                prueba_items = [_label(p) for p in diag_data.get("pruebas", "").split(",") if p.strip()]
                mat_raw      = diag_data.get("materiales", "").strip()
                if mat_raw.lower() in ("sin materiales", ""):
                    mat_items = ["Sin materiales"]
                else:
                    mat_items = [_fmt_material(m) for m in mat_raw.split(",") if m.strip()]

                diag_filas = []
                if "responsable" in diag_data:
                    diag_filas.append(
                        [diag_row("RESPONSABLE DEL PROBLEMA", _label(diag_data["responsable"])),
                         diag_row("RESULTADO DEL SERVICIO",   _label(diag_data.get("resultado", "")))]
                    )
                if causa_items:
                    diag_filas.append(
                        [diag_row("CAUSA DEL PROBLEMA",  _bullets(causa_items)),
                         diag_row("ACCIÓN REALIZADA",    _bullets(accion_items))]
                    )
                if prueba_items:
                    diag_filas.append(
                        [diag_row("PRUEBAS REALIZADAS",    _bullets(prueba_items)),
                         diag_row("MATERIALES UTILIZADOS", _bullets(mat_items))]
                    )

                if diag_filas:
                    flat_d = []
                    for row_d in diag_filas:
                        lft = row_d[0]
                        rgt = row_d[1] if len(row_d) > 1 else [Paragraph("", st_label), Paragraph("", st_bullet)]
                        flat_d.append([lft[0], lft[1], Spacer(sep, 1), rgt[0], rgt[1]])
                    diag_t = Table(
                        flat_d,
                        colWidths=[cw * 0.36, cw * 0.64, sep, cw * 0.36, cw * 0.64],
                    )
                    diag_t.setStyle(TableStyle([
                        ("VALIGN",         (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING",    (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
                        ("TOPPADDING",     (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING",  (0, 0), (-1, -1), 7),
                        ("LEFTPADDING",    (2, 0), (2, -1), 0),
                        ("RIGHTPADDING",   (2, 0), (2, -1), 0),
                        ("TOPPADDING",     (2, 0), (2, -1), 0),
                        ("BOTTOMPADDING",  (2, 0), (2, -1), 0),
                        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [white, C_BG]),
                        ("LINEBELOW",      (0, 0), (-1, -2), 0.5, C_BORDER),
                        ("BOX",            (0, 0), (-1, -1), 1, C_BORDER),
                    ]))
                    story.append(diag_t)

            story.append(Spacer(1, 12))

            # ── Evidencia fotográfica ────────────────────────────────────
            fotos_validas = [f for f in (fotos or []) if str(f or "").strip()][:3]
            if fotos_validas:
                story.append(Paragraph("EVIDENCIA FOTOGRÁFICA", st_sec))
                hr()
                img_cells: list = []
                max_w = (fw - 0.3 * cm * max(len(fotos_validas) - 1, 0)) / len(fotos_validas)
                max_h = 6.5 * cm
                for fuente in fotos_validas:
                    try:
                        fuente_s = str(fuente).strip()
                        if fuente_s.startswith("data:"):
                            import base64 as _b64
                            _, b64data = fuente_s.split(",", 1)
                            img_buf = io.BytesIO(_b64.b64decode(b64data))
                        elif fuente_s.startswith("/api/incidencias/drive-image/"):
                            file_id = fuente_s.rsplit("/", 1)[-1]
                            content, _mime, _name = download_support_drive_file_bytes(file_id=file_id)
                            img_buf = io.BytesIO(content)
                        else:
                            p = Path(fuente_s)
                            if fuente_s.startswith("/uploads/") or fuente_s.startswith("uploads/"):
                                p = _url_to_path(fuente_s)
                            elif not p.is_absolute():
                                p = _ATC_ROOT / fuente_s.lstrip("/")
                            if not p.exists():
                                continue
                            img_buf = io.BytesIO(p.read_bytes())
                        pil = PILImage.open(io.BytesIO(img_buf.getvalue()))
                        pil.load()  # fuerza la decodificacion completa: un
                        # archivo truncado (subida cortada a mitad, etc.)
                        # pasa el open() (solo lee el header) pero falla aca
                        # — sin este chequeo, reportlab recien lo detecta
                        # durante doc.build() y tira abajo el PDF completo
                        # en vez de solo saltear esta foto puntual.
                        ow, oh = pil.size
                        ratio = ow / oh if oh else 1
                        iw = min(max_w, max_h * ratio)
                        ih = iw / ratio
                        if ih > max_h:
                            ih = max_h
                            iw = ih * ratio
                        img_cells.append(RLImage(img_buf, width=iw, height=ih))
                    except Exception:
                        pass
                if img_cells:
                    n = len(img_cells)
                    col_w = fw / n
                    foto_t = Table([img_cells], colWidths=[col_w] * n)
                    foto_t.setStyle(TableStyle([
                        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING",    (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("BOX",           (0, 0), (-1, -1), 0.5, C_BORDER),
                        ("BACKGROUND",    (0, 0), (-1, -1), C_BG),
                    ]))
                    story.append(foto_t)

            doc.build(story)
            return buf.getvalue()

        except Exception:
            LOGGER.exception("Error generando PDF local cierre ODT %s", snapshot.get("odt"))
            return None

    def _generar_drive_para_cierre(
        self,
        odt: str,
        observacion: str,
        fotos: list[str],
        max_fotos_informe: int | None = None,
        foto_payloads: list[dict[str, object]] | None = None,
    ) -> dict[str, Any]:
        """Genera el PDF de cierre (en memoria) y lo sube a Drive junto con
        las fotos, sin escribir nada a disco local salvo que Drive falle
        (ver _guardar_en_cuarentena_drive).

        foto_payloads son las fotos ya en memoria (bytes) de los flujos
        nuevos (cierre de ODT sincronico, mantencion) — si vienen, se suben
        primero y solo falta el PDF. Si no vienen, se usa el esquema viejo
        (fotos como URLs/data-uris ya resueltas por el llamador, ej. el
        fallback legacy de base64), subiendo fotos+PDF juntos como siempre.
        """
        import base64

        snapshot = self._obtener_snapshot_cierre_odt(odt)
        observacion_final = str(observacion or snapshot.get("observacion_final") or "").strip()
        sucursal = str(snapshot.get("sucursal") or snapshot.get("cliente") or "").strip()
        cliente = str(snapshot.get("cliente") or "").strip()
        odt_id = str(snapshot.get("odt") or odt).strip()

        if foto_payloads:
            fuentes_informe_src = [
                f"data:{str(p.get('mime_type') or '').strip() or 'image/jpeg'};base64,{base64.b64encode(bytes(p.get('bytes') or b'')).decode()}"
                for p in foto_payloads
                if p.get("bytes")
            ]
            fotos_legacy: list[str] = []
        else:
            fuentes_informe_src = [str(f or "").strip() for f in (fotos or self.obtener_imagenes_finalizacion(odt)) if str(f or "").strip()]
            fotos_legacy = fuentes_informe_src
        if max_fotos_informe is not None:
            fuentes_informe = fuentes_informe_src[: max(0, int(max_fotos_informe or 0))]
        else:
            fuentes_informe = fuentes_informe_src

        # 1. Generar PDF en memoria con estilo corporativo (es el informe primario)
        pdf_bytes = self._generar_pdf_local_cierre_odt(snapshot, observacion_final, fuentes_informe)

        # 2. Subir fotos (si no se subieron ya) y el PDF a la misma carpeta Drive.
        folder_id = ""
        pdf_url = ""
        drive_meta: dict = {}
        imagenes_drive: list[str] = []
        try:
            if foto_payloads:
                folder_id, _folder_name = resolve_odt_cierre_folder(odt=odt_id, sucursal=sucursal, cliente=cliente)
                if fotos:
                    # Las fotos ya se subieron sincronicamente antes de
                    # llegar aca (ver subir_fotos_cierre_odt_sync) — no
                    # volver a subirlas (eso las duplicaba en Drive), solo
                    # reusar las URLs ya obtenidas para la tabla unificada.
                    imagenes_drive = [str(u or "").strip() for u in fotos if str(u or "").strip()]
                else:
                    imagenes_drive = upload_odt_cierre_images_to_drive(
                        folder_id=folder_id, odt=odt_id, image_payloads=foto_payloads,
                    )
                if pdf_bytes:
                    pdf_meta = upload_odt_cierre_pdf_to_drive(
                        folder_id=folder_id, odt=odt_id, sucursal=sucursal, pdf_bytes=pdf_bytes,
                    )
                    drive_meta = {"folder_id": folder_id, **pdf_meta}
                    pdf_file_id = pdf_meta.get("pdf_file_id", "")
                    if pdf_file_id:
                        pdf_url = f"/api/incidencias/drive-image/{pdf_file_id}"
            elif pdf_bytes:
                drive_meta = dict(upload_odt_cierre_to_drive(
                    pdf_bytes=pdf_bytes, odt=odt_id, sucursal=sucursal, cliente=cliente,
                    image_sources=fotos_legacy,
                ))
                folder_id = drive_meta.get("folder_id", "")
                imagenes_drive = [
                    str(url or "").strip() for url in (drive_meta.get("imagenes") or []) if str(url or "").strip()
                ]
                pdf_file_id = drive_meta.get("pdf_file_id", "")
                if pdf_file_id:
                    pdf_url = f"/api/incidencias/drive-image/{pdf_file_id}"

            if folder_id:
                self._guardar_drive_cierre_folder(
                    odt, folder_id, f"https://drive.google.com/drive/folders/{folder_id}"
                )
            if imagenes_drive:
                self._upsert_unified_images(
                    odt_id, sucursal, "drive_cierre", imagenes_drive, max_imagenes=max(len(imagenes_drive), 3),
                )
                self.db.commit()
        except Exception:
            LOGGER.exception("No se pudo subir informe/fotos de ODT %s a Drive", odt)

        if not pdf_url and pdf_bytes:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            odt_slug = odt_id.replace("/", "-").replace(" ", "_") or "SIN_ODT"
            pdf_url = _guardar_en_cuarentena_drive(
                "cierre_odt", odt_id, f"cierre_{odt_slug}_{ts}.pdf", pdf_bytes,
                {"odt": odt_id, "sucursal": sucursal, "cliente": cliente},
            )
        if pdf_url:
            self._guardar_pdf_url_odt(odt, pdf_url)

        LOGGER.info(
            "Cierre Drive ODT %s: folder_id=%s imagenes=%d pdf_url=%s",
            odt, folder_id or "-", len(imagenes_drive), pdf_url or "(ninguno)",
        )
        return {
            "pdf_web_view_link": pdf_url,
            "local_pdf_url":     pdf_url,
            "folder_id":         folder_id,
            "imagenes":          imagenes_drive,
        }

    def _localizar_carpeta_drive_odt(self, drive, odt: str, cliente: str) -> str:
        """Ubica la carpeta Drive de una ODT ya cerrada, sin depender de
        pdf_url (que para cierres antiguos puede apuntar a un archivo suelto
        que no es el que se muestra en "Ver informe e imagenes ODT" — ese
        visor navega por carpetas, no por pdf_url). Prueba, en orden:
        1) carpeta "ODT {odt}" dentro de la carpeta del cliente (esquema
           nuevo, el que usa upload_odt_cierre_to_drive).
        2) carpeta con el nombre exacto del odt en cualquier nivel (esquema
           viejo, usado antes de julio 2026)."""
        from ATC.app.services.drive_base_service import _clean_filename

        root_folder_id = str(settings.google_drive_root_folder_id or "").strip()
        odt_limpio = (odt or "").strip()
        safe_odt_folder = _clean_filename(f"ODT {odt_limpio}", fallback=f"ODT_{odt_limpio}")

        if cliente:
            safe_cliente = _clean_filename(cliente, fallback="Sucursal")
            try:
                res_cliente = drive.files().list(
                    q=(
                        f"name = '{safe_cliente}' and '{root_folder_id}' in parents "
                        "and mimeType = 'application/vnd.google-apps.folder' and trashed=false"
                    ),
                    fields="files(id)",
                    pageSize=1,
                ).execute()
                cliente_folders = res_cliente.get("files", [])
                if cliente_folders:
                    res_odt = drive.files().list(
                        q=(
                            f"name = '{safe_odt_folder}' and '{cliente_folders[0]['id']}' in parents "
                            "and mimeType = 'application/vnd.google-apps.folder' and trashed=false"
                        ),
                        fields="files(id)",
                        pageSize=1,
                    ).execute()
                    odt_folders = res_odt.get("files", [])
                    if odt_folders:
                        return odt_folders[0]["id"]
            except Exception:
                pass

        try:
            res_directo = drive.files().list(
                q=f"name = '{odt_limpio}' and mimeType = 'application/vnd.google-apps.folder' and trashed=false",
                fields="files(id)",
                pageSize=1,
            ).execute()
            folders = res_directo.get("files", [])
            if folders:
                return folders[0]["id"]
        except Exception:
            pass

        return ""

    def obtener_observacion_cierre_odt(self, odt: str) -> str:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            return ""
        row_reg = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        return str(getattr(row_reg, "observacion_final", "") or "").strip() if row_reg else ""

    def regenerar_informe_cierre_odt(self, odt: str, texto_nuevo: str, token: str = "") -> dict[str, Any]:
        """Reemplaza la 'Observacion de cierre' de una ODT ya cerrada (el
        llamador manda el texto completo editado, no una nota aparte) y
        regenera el PDF (local + Drive) con el texto resultante. A
        diferencia de _generar_drive_para_cierre, NO vuelve a subir las fotos
        (ya estan en Drive) — solo reemplaza el PDF en su carpeta real."""
        odt_limpia = (odt or "").strip()
        texto_nuevo = (texto_nuevo or "").strip()
        if not odt_limpia:
            raise ValueError("odt es obligatorio")
        if not texto_nuevo:
            raise ValueError("La observacion no puede quedar vacia")

        row_reg = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row_reg:
            raise ValueError(f"ODT {odt_limpia} no encontrada")

        row_reg.observacion_final = texto_nuevo
        self.db.commit()

        snapshot = self._obtener_snapshot_cierre_odt(odt_limpia)
        fotos = self.obtener_imagenes_finalizacion(odt_limpia)

        pdf_bytes = self._generar_pdf_local_cierre_odt(snapshot, texto_nuevo, fotos)
        if not pdf_bytes:
            raise RuntimeError("No se pudo generar el PDF actualizado")

        drive_actualizado = False
        folder_id = ""
        pdf_url = ""
        try:
            if settings.google_drive_enabled:
                from ATC.app.services.drive_base_service import _build_clients, _clean_filename, _upload_bytes

                drive, _ = _build_clients()
                folder_id = str(row_reg.drive_cierre_folder_id or "").strip()
                if not folder_id:
                    folder_id = self._localizar_carpeta_drive_odt(
                        drive, odt_limpia, str(snapshot.get("cliente") or "").strip()
                    )

                if folder_id:
                    res = drive.files().list(
                        q=f"'{folder_id}' in parents and mimeType = 'application/pdf' and trashed=false",
                        fields="files(id,name)",
                    ).execute()
                    for f in res.get("files", []):
                        try:
                            drive.files().update(fileId=f["id"], body={"trashed": True}).execute()
                        except Exception:
                            LOGGER.exception("No se pudo archivar PDF viejo %s de ODT %s", f.get("id"), odt_limpia)

                    safe_sucursal = _clean_filename(
                        str(snapshot.get("sucursal") or snapshot.get("cliente") or "Sucursal"), fallback="Sucursal"
                    )
                    now_stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M")
                    pdf_name = _clean_filename(
                        f"ODT_{odt_limpia}_{safe_sucursal}_{now_stamp}.pdf",
                        fallback=f"ODT_{odt_limpia}_{now_stamp}.pdf",
                    )
                    uploaded = _upload_bytes(drive, folder_id, pdf_name, pdf_bytes, "application/pdf")
                    pdf_url = f"/api/incidencias/drive-image/{uploaded['id']}"

                    if not row_reg.drive_cierre_folder_id:
                        self._guardar_drive_cierre_folder(
                            odt_limpia, folder_id, f"https://drive.google.com/drive/folders/{folder_id}"
                        )
                    drive_actualizado = True
        except Exception:
            LOGGER.exception("No se pudo actualizar el PDF en Drive para ODT %s", odt_limpia)

        if not pdf_url:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            odt_slug = odt_limpia.replace("/", "-").replace(" ", "_") or "SIN_ODT"
            pdf_url = _guardar_en_cuarentena_drive(
                "cierre_odt", odt_limpia, f"cierre_{odt_slug}_{ts}.pdf", pdf_bytes,
                {"odt": odt_limpia, "sucursal": str(snapshot.get("sucursal") or "")},
            )
        self._guardar_pdf_url_odt(odt_limpia, pdf_url)

        return {
            "odt": odt_limpia,
            "observacion_final": texto_nuevo,
            "pdf_url": pdf_url,
            "drive_actualizado": drive_actualizado,
            "drive_folder_id": folder_id,
        }

    def reintentar_informes_drive_pendientes(self, *, limite: int = 20) -> dict[str, int]:
        """Completa cierres de ODT que generaron el informe local pero se
        quedaron sin subirlo a Drive (folder_id vacio). Pensado para correr
        periodicamente desde el loop de automatizaciones — evita depender de
        que alguien note manualmente que un informe no llego a Drive.

        Sube solo lo que falte (nunca duplica fotos ni el PDF si ya estan).
        """
        resultado = {"revisados": 0, "reparados": 0, "con_error": 0}
        try:
            if not settings.google_drive_enabled:
                return resultado
        except Exception:
            return resultado

        # pdf_url puede traer, para cierres antiguos, un link de Drive ya
        # subido por otro flujo que nunca uso drive_cierre_folder_id (columna
        # mas nueva) — esos NO son fallas, filtrar solo a "/uploads/..."
        # (ruta local) es lo que realmente identifica un cierre que genero
        # el PDF pero se quedo sin subirlo.
        pendientes = self.db.execute(
            select(Registro.odt)
            .where(
                Registro.pdf_url.like("/uploads/%"),
                or_(Registro.drive_cierre_folder_id.is_(None), Registro.drive_cierre_folder_id == ""),
            )
            .order_by(Registro.fecha_cierre.desc())
            .limit(limite)
        ).all()

        for (odt,) in pendientes:
            resultado["revisados"] += 1
            try:
                row = self.db.scalar(select(Registro).where(Registro.odt == odt))
                if not row or not row.pdf_url:
                    continue
                pdf_abs = _url_to_path(row.pdf_url)
                if not Path(pdf_abs).exists():
                    # El informe local ya no existe (limpieza, disco, etc.);
                    # no hay nada que reintentar para esta ODT.
                    continue

                snapshot = self._obtener_snapshot_cierre_odt(odt)
                fotos = self.obtener_imagenes_finalizacion(odt)

                meta = retry_odt_cierre_drive_upload(
                    pdf_local_path=str(pdf_abs),
                    odt=odt,
                    sucursal=str(snapshot.get("sucursal") or snapshot.get("cliente") or "").strip(),
                    cliente=str(snapshot.get("cliente") or "").strip(),
                    image_sources=fotos,
                )
                folder_id = meta.get("folder_id") or ""
                if folder_id:
                    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
                    self._guardar_drive_cierre_folder(odt, folder_id, folder_url)

                if meta.get("pdf_error"):
                    resultado["con_error"] += 1
                    LOGGER.warning(
                        "Reintento Drive ODT %s: carpeta OK pero informe aun fallo: %s",
                        odt, meta.get("pdf_error"),
                    )
                else:
                    resultado["reparados"] += 1
                    LOGGER.info("Reintento Drive ODT %s: completado (pdf_subido=%s)", odt, meta.get("pdf_subido"))
            except Exception:
                resultado["con_error"] += 1
                LOGGER.exception("Reintento Drive ODT %s: fallo inesperado", odt)

        return resultado

    def reintentar_uploads_pendientes_drive(self, *, limite: int = 30) -> dict[str, int]:
        """Reintenta subir a Drive los archivos que quedaron en cuarentena
        local (ATC/uploads/_pending_drive/..., ver _guardar_en_cuarentena_drive)
        porque Drive fallo en el momento del upload. Al lograrlo, actualiza la
        fila de BBDD correspondiente con la URL definitiva de Drive y borra
        el archivo local — asi la cuarentena no crece sin limite. Pensado
        para correr periodicamente desde automation_loop()."""
        resultado = {"revisados": 0, "reparados": 0, "con_error": 0}
        try:
            if not settings.google_drive_enabled or not _PENDING_DRIVE_ROOT.exists():
                return resultado
        except Exception:
            return resultado

        meta_files = sorted(_PENDING_DRIVE_ROOT.rglob("*.meta.json"))[:limite]
        for meta_path in meta_files:
            resultado["revisados"] += 1
            data_path = Path(str(meta_path)[: -len(".meta.json")])
            try:
                if not data_path.exists():
                    # Se perdio el archivo real; limpiar el sidecar huerfano.
                    meta_path.unlink(missing_ok=True)
                    continue
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
                content = data_path.read_bytes()
                flujo = str(meta.get("flujo") or "")
                quarantine_url = f"/uploads/{data_path.relative_to(_UPLOADS_ROOT).as_posix()}"

                nueva_url = self._reintentar_un_pendiente_drive(flujo, meta, content, quarantine_url)
                if nueva_url:
                    data_path.unlink(missing_ok=True)
                    meta_path.unlink(missing_ok=True)
                    resultado["reparados"] += 1
                    LOGGER.info("Cuarentena Drive reparada: flujo=%s %s -> %s", flujo, quarantine_url, nueva_url)
                else:
                    resultado["con_error"] += 1
                    LOGGER.warning(
                        "Cuarentena Drive sigue pendiente: flujo=%s %s (fila no encontrada o Drive sigue fallando)",
                        flujo, quarantine_url,
                    )
            except Exception:
                resultado["con_error"] += 1
                LOGGER.exception("Reintento de cuarentena Drive fallo para %s", meta_path)
        return resultado

    def _reintentar_un_pendiente_drive(
        self, flujo: str, meta: dict[str, Any], content: bytes, quarantine_url: str,
    ) -> str:
        """Reintenta un item puntual de cuarentena; devuelve la nueva URL de
        Drive si tuvo exito (y ya actualizo la fila en BBDD), o "" si sigue
        fallando o la fila ya no existe/coincide (ej. se borro mientras
        tanto)."""
        if flujo == "cierre_apertura":
            fila_ca = self.db.scalar(
                select(CierreAperturaImagen).where(CierreAperturaImagen.ruta_archivo == quarantine_url)
            )
            if not fila_ca:
                return ""
            drive_result = upload_cierre_apertura_image_to_drive(
                client_id=str(meta.get("client_id") or ""),
                client_name=str(meta.get("client_name") or ""),
                content=content,
                filename=str(meta.get("filename") or "imagen.png"),
                mime_type=str(meta.get("mime_type") or "image/png"),
            )
            fila_ca.ruta_archivo = drive_result["public_uri"]
            self.db.commit()
            return drive_result["public_uri"]

        if flujo == "rendicion_boleta":
            fila_rb = self.db.scalar(select(Rendicion).where(Rendicion.url_boleta == quarantine_url))
            if not fila_rb:
                return ""
            drive_result = upload_rendicion_boleta_to_drive(
                tecnico=str(meta.get("tecnico") or "") or "Sin tecnico",
                rendicion_ref=str(meta.get("rendicion_ref") or fila_rb.id),
                content=content,
                filename=str(meta.get("filename") or "boleta.jpg"),
                mime_type=str(meta.get("mime_type") or "application/octet-stream"),
            )
            fila_rb.url_boleta = drive_result["public_uri"]
            self.db.commit()
            return drive_result["public_uri"]

        if flujo == "rendicion_informe":
            fila_ri = self.db.scalar(select(Rendicion).where(Rendicion.url_informe == quarantine_url))
            if not fila_ri:
                return ""
            drive_result = upload_rendicion_informe_to_drive(
                tecnico=str(meta.get("tecnico") or "") or str(fila_ri.tecnico or "Sin tecnico"),
                rendicion_id=int(meta.get("rendicion_id") or fila_ri.id),
                pdf_bytes=content,
                filename=str(meta.get("filename") or f"informe_{fila_ri.id}.pdf"),
            )
            fila_ri.url_informe = drive_result["public_uri"]
            self.db.commit()
            return drive_result["public_uri"]

        if flujo in ("cierre_odt", "cierre_odt_foto"):
            odt = str(meta.get("odt") or "").strip()
            if not odt:
                return ""
            fila_reg = self.db.scalar(select(Registro).where(Registro.odt == odt))
            if not fila_reg:
                return ""
            sucursal = str(meta.get("sucursal") or "")
            cliente = str(meta.get("cliente") or "")
            folder_id, _folder_name = resolve_odt_cierre_folder(odt=odt, sucursal=sucursal, cliente=cliente)
            if folder_id:
                self._guardar_drive_cierre_folder(
                    odt, folder_id, f"https://drive.google.com/drive/folders/{folder_id}"
                )

            if flujo == "cierre_odt":
                if fila_reg.pdf_url != quarantine_url:
                    return ""
                pdf_meta = upload_odt_cierre_pdf_to_drive(
                    folder_id=folder_id, odt=odt, sucursal=sucursal, pdf_bytes=content,
                )
                nueva_url = f"/api/incidencias/drive-image/{pdf_meta['pdf_file_id']}"
                fila_reg.pdf_url = nueva_url
                self.db.commit()
                return nueva_url

            campo = next(
                (nombre for nombre in ("foto_1", "foto_2", "foto_3") if getattr(fila_reg, nombre, "") == quarantine_url),
                None,
            )
            if not campo:
                return ""
            urls = upload_odt_cierre_images_to_drive(
                folder_id=folder_id, odt=odt,
                image_payloads=[{
                    "filename": str(meta.get("filename") or "imagen.jpg"),
                    "mime_type": str(meta.get("mime_type") or "image/jpeg"),
                    "bytes": content,
                }],
            )
            if not urls:
                return ""
            setattr(fila_reg, campo, urls[0])
            self.db.commit()
            return urls[0]

        return ""

    def subir_fotos_cierre_odt_sync(self, odt: str, image_payloads: list[dict[str, object]]) -> list[str]:
        """Sube las fotos de cierre de ODT a Drive de forma sincronica (antes
        de responder el request), para que foto_1/2/3 queden con la URL
        definitiva de inmediato. Si Drive falla para una foto puntual, esa
        foto cae en cuarentena local (ver _guardar_en_cuarentena_drive) sin
        perder su posicion en la lista, para no desalinear foto_1/2/3."""
        if not image_payloads:
            return []

        snapshot = self._obtener_snapshot_cierre_odt(odt)
        sucursal = str(snapshot.get("sucursal") or snapshot.get("cliente") or "").strip()
        cliente = str(snapshot.get("cliente") or "").strip()
        odt_id = str(snapshot.get("odt") or odt).strip() or odt

        folder_id = ""
        try:
            folder_id, _folder_name = resolve_odt_cierre_folder(odt=odt_id, sucursal=sucursal, cliente=cliente)
        except Exception:
            LOGGER.exception("No se pudo resolver carpeta Drive para fotos de cierre ODT %s", odt)

        urls: list[str] = []
        for idx, payload in enumerate(image_payloads, start=1):
            content = payload.get("bytes") or b""
            filename = str(payload.get("filename") or f"img_{idx}")
            mime_type = str(payload.get("mime_type") or "image/jpeg")
            url = ""
            if folder_id and content:
                try:
                    subidas = upload_odt_cierre_images_to_drive(
                        folder_id=folder_id, odt=odt_id, image_payloads=[payload],
                    )
                    if subidas:
                        url = subidas[0]
                except Exception:
                    LOGGER.exception("No se pudo subir foto %s de cierre ODT %s a Drive", idx, odt)
            if not url:
                url = _guardar_en_cuarentena_drive(
                    "cierre_odt_foto", odt_id, filename, bytes(content),
                    {"odt": odt_id, "sucursal": sucursal, "cliente": cliente, "index": idx, "mime_type": mime_type},
                )
            urls.append(url)
        cuarentena = sum(1 for u in urls if u.startswith("/uploads/_pending_drive/"))
        LOGGER.info(
            "Fotos sync cierre ODT %s: %d subidas, %d en cuarentena", odt, len(urls) - cuarentena, cuarentena,
        )
        return urls

    @staticmethod
    def _ejecutar_drive_en_segundo_plano(
        odt: str,
        observacion: str,
        fotos: list[str],
        max_fotos_informe: int | None = None,
        foto_payloads: list[dict[str, object]] | None = None,
    ) -> None:
        db = SessionLocal()
        try:
            service = IncidenciasService(db)
            service._generar_drive_para_cierre(
                odt, observacion, fotos, max_fotos_informe=max_fotos_informe, foto_payloads=foto_payloads,
            )
        except Exception:
            LOGGER.exception("Fallo la generacion automatica del informe Drive para ODT %s", odt)
        finally:
            db.close()

    def continuar_finalizacion_asincrona(
        self,
        odt: str,
        fotos_base64: list[str],
        observacion: str = "",
        *,
        responsable_cierre: str,
        causa_cierre: Any,
        accion_cierre: Any,
        resultado_cierre: str,
        pruebas_cierre: list[Any] | None = None,
        materiales: list[Any] | None = None,
        materiales_sin_uso: bool = False,
        requiere_seguimiento: bool = False,
        foto_payloads: list[dict[str, object]] | None = None,
    ) -> dict[str, Any] | str:
        odt_limpia = (odt or "").strip()
        if not odt_limpia:
            raise ValueError("ODT invalida")

        fotos = [str(f or "").strip() for f in (fotos_base64 or []) if str(f or "").strip()][:3]
        obs_cierre = str(observacion or "").strip()
        if not obs_cierre:
            raise ValueError("Debes ingresar una observacion final breve.")

        diagnostico = self._normalizar_diagnostico_cierre(
            responsable_cierre=responsable_cierre,
            causa_cierre=causa_cierre,
            accion_cierre=accion_cierre,
            resultado_cierre=resultado_cierre,
            pruebas_cierre=pruebas_cierre,
            materiales=materiales,
            materiales_sin_uso=materiales_sin_uso,
            requiere_seguimiento=requiere_seguimiento,
        )

        row = self.db.scalar(select(Registro).where(Registro.odt == odt_limpia))
        if not row:
            raise ValueError(f"No se encontro la ODT {odt_limpia}")

        estado_previo = str(getattr(row, "estado", "") or "").strip()
        cierre_ya_sincronizado = (
            self._normalizar_texto(getattr(row, "estado", "") or "") == "terminado"
            and str(getattr(row, "observacion_final", "") or "").strip() == obs_cierre
            and str(getattr(row, "responsable_cierre", "") or "").strip() == diagnostico["responsable_cierre"]
            and str(getattr(row, "causa_cierre", "") or "").strip() == self._formatear_lista_cierre(diagnostico["causa_cierre"])
            and str(getattr(row, "accion_cierre", "") or "").strip() == self._formatear_lista_cierre(diagnostico["accion_cierre"])
            and str(getattr(row, "resultado_cierre", "") or "").strip() == diagnostico["resultado_cierre"]
        )

        if len(fotos) >= 1:
            row.foto_1 = fotos[0]
        if len(fotos) >= 2:
            row.foto_2 = fotos[1]
        if len(fotos) >= 3:
            row.foto_3 = fotos[2]
        if obs_cierre:
            row.observacion_final = obs_cierre
        self._aplicar_diagnostico_cierre(row, diagnostico)
        self._marcar_instalacion_venta_finalizada(odt_limpia)
        self.db.commit()
        self._reflejar_audio_ok_en_prueba_sonido(row, diagnostico.get("pruebas_cierre"))
        if estado_previo != "Terminado" and self._normalizar_texto(getattr(row, "estado", "") or "") == "terminado":
            self._reforzar_inicio_odt_si_corresponde(
                odt=row.odt,
                tecnico=row.tecnicos,
                acompanante=row.acompanante,
                usuario_accion=row.tecnico_cierre,
                fecha_inicio_trabajo=row.fecha_inicio_trabajo,
                verbo_accion="finalizada",
            )
        if not cierre_ya_sincronizado:
            self._sync_estado_ticket_soporte_silencioso(
                odt_limpia,
                TICKET_STATUS_RESUELTO_SERVICIO,
                nota_interna=self._build_nota_cierre_ticket_soporte(
                    odt=odt_limpia,
                    estado_ticket=TICKET_STATUS_RESUELTO_SERVICIO,
                    derivacion=row.derivacion or "Servicio Tecnico",
                    observacion_final=(
                        f"{row.observacion_final or obs_cierre}\n"
                        f"{self._resumen_diagnostico_cierre(diagnostico)}"
                    ).strip(),
                ),
            )

        drive_enabled = bool(settings.google_drive_enabled)
        if drive_enabled:
            worker = threading.Thread(
                target=self._ejecutar_drive_en_segundo_plano,
                args=(odt_limpia, self._observacion_drive_cierre(obs_cierre, diagnostico), fotos, None, foto_payloads),
                daemon=True,
                name=f"drive-report-{odt_limpia}",
            )
            worker.start()
            return {
                "result": "OK",
                "drive_enabled": True,
                "drive_queued": True,
                "drive_message": "Informe en generacion en segundo plano.",
            }

        return {"result": "OK", "drive_enabled": False}

    def guardar_mantencion_correctiva(self, data: dict[str, Any]) -> str:
        sucursal = str(data.get("sucursal") or "").strip()
        problema = str(data.get("problema") or "").strip()
        if not sucursal or not problema:
            raise ValueError("Sucursal y servicio son obligatorios.")

        odt = self._proximo_odt("M")
        ahora = datetime.now()
        observacion = str(data.get("observacion") or "").strip()
        observacion_servicio = str(data.get("observacion_servicio") or "").strip()
        tecnico = str(data.get("tecnico") or "").strip()
        acompanante = str(data.get("acompanante") or "").strip()
        estado = "En Proceso" if (tecnico or acompanante) else "Pendiente"

        reg = Registro(
            odt=odt,
            fecha_registro=ahora,
            puesto=None,
            cliente=sucursal,
            problema=problema,
            detalle_problema=(observacion or None),
            derivacion="Servicio Tecnico",
            observacion=(observacion or None),
            observacion_servicio=(observacion_servicio or None),
            tecnicos=tecnico or None,
            acompanante=acompanante or None,
            estado=estado,
            fecha_derivacion_area=ahora,
            fecha_derivacion_tecnico=ahora,
            direccion=self._direccion_cliente(sucursal),
            prioridad=data.get("prioridad") or None,
        )
        self.db.add(reg)
        self.db.commit()
        self._reset_unified_images_if_odt_reused(odt, sucursal)
        return "OK"

    @staticmethod
    def _semana_del_mes(fecha_ref: datetime) -> int:
        return ((fecha_ref.day - 1) // 7) + 1

    @staticmethod
    def _normalizar_sucursal_key(valor: str) -> str:
        txt = str(valor or "").strip().lower()
        if not txt:
            return ""
        txt = unicodedata.normalize("NFD", txt)
        txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt

    def obtener_plantilla_imagenes_mantencion(self, sucursal: str) -> list[str]:
        sucursal_key = self._normalizar_sucursal_key(sucursal)
        if not sucursal_key:
            return []

        row = self.db.scalar(
            select(MantencionImagenSucursal).where(MantencionImagenSucursal.sucursal_key == sucursal_key)
        )
        if row:
            return self._parse_image_list(row.imagenes)[:3]

        # Compatibilidad/migracion desde implementacion antigua (plantilla guardada por pseudo-ODT).
        odt_tpl_legacy = f"__MANT_TPL__:{sucursal_key}"
        legacy = self.db.scalar(select(IncidenciaImagenTabla).where(IncidenciaImagenTabla.odt == odt_tpl_legacy))
        legacy_urls = self._parse_image_list(legacy.imagenes if legacy else "[]")[:3]
        if not legacy_urls:
            return []

        nuevo = MantencionImagenSucursal(
            sucursal_key=sucursal_key,
            sucursal=str(sucursal or "").strip() or sucursal_key,
            imagenes=json.dumps(legacy_urls, ensure_ascii=False),
            created_by="migracion_legacy",
        )
        self.db.add(nuevo)
        self.db.commit()
        return legacy_urls

    def guardar_plantilla_imagenes_mantencion(
        self,
        sucursal: str,
        imagenes: list[str],
    ) -> dict[str, Any]:
        sucursal_limpia = str(sucursal or "").strip()
        if not sucursal_limpia:
            raise ValueError("Sucursal es obligatoria.")

        urls = [
            str(url or "").strip()
            for url in (imagenes or [])
            if self._es_url_publica_imagen(str(url or "").strip())
        ][:3]
        if not urls:
            raise ValueError("Debes indicar al menos una URL publica de imagen.")

        sucursal_key = self._normalizar_sucursal_key(sucursal_limpia)
        if not sucursal_key:
            raise ValueError("No se pudo normalizar la sucursal.")

        row = self.db.scalar(
            select(MantencionImagenSucursal).where(MantencionImagenSucursal.sucursal_key == sucursal_key)
        )
        payload = json.dumps(urls, ensure_ascii=False)
        if row:
            row.sucursal = sucursal_limpia
            row.imagenes = payload
            row.created_by = "plantilla_mantencion"
            row.updated_at = datetime.now()
        else:
            self.db.add(
                MantencionImagenSucursal(
                    sucursal_key=sucursal_key,
                    sucursal=sucursal_limpia,
                    imagenes=payload,
                    created_by="plantilla_mantencion",
                )
            )
        self.db.commit()
        return {
            "ok": True,
            "sucursal": sucursal_limpia,
            "sucursal_key": sucursal_key,
            "imagenes": urls,
            "total_imagenes": len(urls),
        }

    def guardar_plantilla_imagenes_mantencion_desde_odt(
        self,
        sucursal: str,
        odt_origen: str,
    ) -> dict[str, Any]:
        odt_limpia = str(odt_origen or "").strip()
        if not odt_limpia:
            raise ValueError("ODT de origen es obligatoria.")
        imagenes = self.obtener_imagenes_tabla(odt_limpia)
        if not imagenes:
            raise ValueError(f"La ODT {odt_limpia} no tiene imagenes cargadas.")
        result = self.guardar_plantilla_imagenes_mantencion(sucursal=sucursal, imagenes=imagenes)
        result["odt_origen"] = odt_limpia
        return result

    def _imagenes_programadas_para_sucursal(self, sucursal: str) -> list[str]:
        tpl_urls = self.obtener_plantilla_imagenes_mantencion(sucursal)
        if tpl_urls:
            return tpl_urls
        key = self._normalizar_sucursal_key(sucursal)
        if not key:
            return []
        seeds = MANTENCIONES_IMAGENES_POR_SUCURSAL.get(key, [])
        return [str(u or "").strip() for u in seeds if str(u or "").strip()][:3]

    @staticmethod
    def _es_url_publica_imagen(valor: str) -> bool:
        txt = str(valor or "").strip().lower()
        return (
            txt.startswith("http://")
            or txt.startswith("https://")
            or txt.startswith("/api/incidencias/drive-image/")
            or txt.startswith("/uploads/")
        )

    def _payloads_imagenes_programadas(self, fuentes: list[str]) -> list[dict[str, object]]:
        payloads: list[dict[str, object]] = []
        for fuente in fuentes:
            raw = str(fuente or "").strip()
            if not raw:
                continue
            if self._es_url_publica_imagen(raw):
                continue
            file_path = Path(raw)
            if not file_path.is_absolute():
                file_path = Path.cwd() / file_path
            if not file_path.exists() or not file_path.is_file():
                continue
            mime_type = mimetypes.guess_type(file_path.name)[0] or "image/jpeg"
            if not str(mime_type).lower().startswith("image/"):
                mime_type = "image/jpeg"
            payloads.append(
                {
                    "filename": file_path.name,
                    "mime_type": mime_type,
                    "bytes": file_path.read_bytes(),
                }
            )
        return payloads[:3]

    def _asignar_imagenes_programadas_a_odt(
        self,
        odt: str,
        sucursal: str,
        sobrescribir: bool = False,
    ) -> bool:
        odt_limpia = str(odt or "").strip()
        if not odt_limpia:
            return False

        fuentes = self._imagenes_programadas_para_sucursal(sucursal)
        direct_urls = [f for f in fuentes if self._es_url_publica_imagen(f)][:3]
        payloads = self._payloads_imagenes_programadas(fuentes)
        if not direct_urls and not payloads:
            LOGGER.warning(
                "No hay imagenes plantilla disponibles para sucursal programada '%s'.",
                sucursal,
            )
            return False

        row_imgs = self.db.scalar(select(IncidenciaImagenTabla).where(IncidenciaImagenTabla.odt == odt_limpia))
        existentes = self._parse_image_list(row_imgs.imagenes if row_imgs else "[]")[:3]
        if existentes and not sobrescribir and all(self._es_url_publica_imagen(url) for url in existentes):
            return False

        try:
            existentes_drive = list_support_images_for_odt(
                odt=odt_limpia,
                root_folder_id=str(settings.google_drive_support_folder_id or "").strip(),
            )
            existentes_drive = [
                str(url or "").strip()
                for url in (existentes_drive or [])
                if str(url or "").strip()
            ][:3]
        except Exception:
            existentes_drive = []

        if existentes_drive and not sobrescribir:
            self._upsert_unified_images(
                odt=odt_limpia,
                sucursal=str(sucursal or "").strip(),
                usuario="auto_mantencion_programada",
                imagenes=existentes_drive,
            )
            self.db.commit()
            return True

        if direct_urls and not payloads:
            self._upsert_unified_images(
                odt=odt_limpia,
                sucursal=str(sucursal or "").strip(),
                usuario="auto_mantencion_programada",
                imagenes=direct_urls,
            )
            self.db.commit()
            return True

        try:
            drive_result = upload_support_images_for_odt(
                odt=odt_limpia,
                image_payloads=payloads,
                root_folder_id=str(settings.google_drive_support_folder_id or "").strip(),
                start_index=1,
            )
        except Exception:
            LOGGER.exception(
                "No se pudieron subir imagenes programadas de mantencion (odt=%s, sucursal=%s).",
                odt_limpia,
                sucursal,
            )
            return False

        nuevas_urls = [
            str(url or "").strip()
            for url in (drive_result.get("imagenes") or [])
            if str(url or "").strip()
        ][:3]
        merged_urls = []
        for url in [*direct_urls, *nuevas_urls]:
            clean = str(url or "").strip()
            if clean and clean not in merged_urls:
                merged_urls.append(clean)
            if len(merged_urls) >= 3:
                break
        if not merged_urls:
            return False

        self._upsert_unified_images(
            odt=odt_limpia,
            sucursal=str(sucursal or "").strip(),
            usuario="auto_mantencion_programada",
            imagenes=merged_urls,
        )
        self.db.commit()
        return True

    def obtener_clientes_soporte(self) -> list[str]:
        clientes_base = self.obtener_catalogo_clientes()
        clientes = sorted({(r or "").strip() for r in clientes_base if (r or "").strip() and (r or "").strip().lower() != "oficina atc"})
        return ["OFICINA ATC", *clientes]

    def obtener_catalogo_clientes(self) -> list[str]:
        try:
            rows_bbdd = self.db.scalars(select(ClienteBBDD.cliente).order_by(ClienteBBDD.cliente.asc())).all()
            return [r for r in rows_bbdd if r]
        except Exception:
            self.db.rollback()
            return []

    def obtener_catalogo_sucursales(self) -> list[str]:
        try:
            rows = self.db.scalars(
                select(SucursalBBDD.nombre_sucursal).order_by(func.lower(func.trim(SucursalBBDD.nombre_sucursal)).asc())
            ).all()
            return sorted(
                {str(r).strip() for r in rows if str(r or "").strip()},
                key=self._normalizar_texto,
            )
        except Exception:
            self.db.rollback()
            return []

    def guardar_mantencion_correctiva(self, data: dict[str, Any]) -> str:
        sucursal = str(data.get("sucursal") or "").strip()
        problema = str(data.get("problema") or "").strip()
        if not sucursal or not problema:
            raise ValueError("Sucursal y servicio son obligatorios.")

        odt = self._proximo_odt("M")
        ahora = datetime.now()
        observacion = str(data.get("observacion") or "").strip()
        observacion_servicio = str(data.get("observacion_servicio") or "").strip()
        tecnico = str(data.get("tecnico") or "").strip()
        acompanante = str(data.get("acompanante") or "").strip()
        estado = "En Proceso" if (tecnico or acompanante) else "Pendiente"

        reg = Registro(
            odt=odt,
            fecha_registro=ahora,
            puesto=None,
            cliente=sucursal,
            problema=problema,
            detalle_problema=(observacion or None),
            derivacion="Servicio Tecnico",
            observacion=(observacion or None),
            observacion_servicio=(observacion_servicio or None),
            tecnicos=tecnico or None,
            acompanante=acompanante or None,
            estado=estado,
            fecha_derivacion_area=ahora,
            fecha_derivacion_tecnico=ahora,
            direccion=self._direccion_cliente(sucursal),
            prioridad=data.get("prioridad") or None,
        )
        self.db.add(reg)
        self.db.commit()
        return "OK"

    def obtener_contactos_por_sucursal(self) -> dict[str, list[dict[str, str]]]:
        data: dict[str, list[dict[str, str]]] = {}

        def _push(sucursal: str, nombre: str, telefono: str, email: str, prioridad: str) -> None:
            suc = (sucursal or "").strip()
            if not suc:
                return
            # Evitar entradas vacÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â­as inÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Âºtiles en el selector.
            if not ((nombre or "").strip() or (telefono or "").strip() or (email or "").strip()):
                return
            data.setdefault(suc, []).append(
                {
                    "nombre": (nombre or "").strip(),
                    "telefono": (telefono or "").strip(),
                    "email": (email or "").strip(),
                    "prioridad": (prioridad or "").strip(),
                }
            )

        def _tablas_disponibles(table_name: str) -> list[str]:
            rows = self.db.execute(
                text(
                    """
                    SELECT DISTINCT table_schema
                    FROM information_schema.columns
                    WHERE table_name = :table_name
                      AND table_schema NOT IN ('pg_catalog', 'information_schema')
                    ORDER BY table_schema
                    """
                ),
                {"table_name": table_name},
            ).all()
            return [str(r[0]).strip() for r in rows if r and r[0]]

        def _columnas(schema_name: str, table_name: str) -> set[str]:
            rows = self.db.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = :schema_name
                      AND table_name = :table_name
                    """
                ),
                {"schema_name": schema_name, "table_name": table_name},
            ).all()
            return {str(r[0]).strip() for r in rows if r and r[0]}

        def _pick(cols: set[str], opciones: list[str]) -> str | None:
            return next((c for c in opciones if c in cols), None)

        def _priority_key(item: dict[str, str]) -> tuple[int, str, str]:
            prio_raw = str(item.get("prioridad") or "").strip()
            match = re.search(r"\d+", prio_raw)
            prio_num = int(match.group()) if match else 999999
            return (prio_num, prio_raw.lower(), str(item.get("nombre") or "").lower())

        def _ordenar_por_prioridad() -> None:
            for sucursal in data:
                data[sucursal].sort(key=_priority_key)

        # 1) Fuente unificada: contactos de emergencia asociados a bbdd_sucursales.
        try:
            rows_emergencia = (
                self.db.query(
                    SucursalBBDD.id,
                    SucursalBBDD.nombre_sucursal,
                    SucursalContactoEmergencia.nombre,
                    SucursalContactoEmergencia.telefono,
                    SucursalContactoEmergencia.email,
                )
                .join(SucursalContactoEmergencia, SucursalContactoEmergencia.sucursal_id == SucursalBBDD.id)
                .order_by(SucursalBBDD.id.asc(), SucursalContactoEmergencia.id.asc())
                .all()
            )
            prioridades_por_sucursal: dict[int, int] = {}
            for sucursal_id, sucursal, nombre, telefono, email in rows_emergencia:
                sid = int(sucursal_id or 0)
                prioridades_por_sucursal[sid] = prioridades_por_sucursal.get(sid, 0) + 1
                _push(sucursal, nombre, telefono, email, str(prioridades_por_sucursal[sid]))
        except Exception:
            pass

        if data:
            _ordenar_por_prioridad()
            return data

        # 2) Fuente legacy: contactos_emergencia (si existe).
        try:
            for schema_name in _tablas_disponibles("contactos_emergencia"):
                cols = _columnas(schema_name, "contactos_emergencia")
                if not cols:
                    continue
                col_sucursal = _pick(cols, ["sucursal", "nombre_sucursal", "cliente", "nombre_cliente"])
                if not col_sucursal:
                    continue
                col_nombre = _pick(cols, ["nombre_empleado", "nombre", "nombre_contacto", "contacto", "contacto_nombre"])
                col_tel = _pick(cols, ["celular", "telefono", "fono", "telefono_contacto"])
                col_email = _pick(cols, ["email", "correo", "mail"])
                col_prio = _pick(cols, ["nro_emergencia", "prioridad", "nivel_prioridad"])

                select_cols = [
                    f'"{col_sucursal}" AS sucursal',
                    f'COALESCE(CAST("{col_nombre}" AS NVARCHAR(MAX)), \'\') AS nombre' if col_nombre else "'' AS nombre",
                    f'COALESCE(CAST("{col_tel}" AS NVARCHAR(MAX)), \'\') AS telefono' if col_tel else "'' AS telefono",
                    f'COALESCE(CAST("{col_email}" AS NVARCHAR(MAX)), \'\') AS email' if col_email else "'' AS email",
                    f'COALESCE(CAST("{col_prio}" AS NVARCHAR(MAX)), \'\') AS prioridad' if col_prio else "'' AS prioridad",
                ]
                sql = text(
                    f"""
                    SELECT {", ".join(select_cols)}
                    FROM "{schema_name}"."contactos_emergencia"
                    WHERE "{col_sucursal}" IS NOT NULL
                      AND TRIM(CAST("{col_sucursal}" AS NVARCHAR(MAX))) <> ''
                    ORDER BY 1
                    """
                )
                for row in self.db.execute(sql).all():
                    _push(row[0], row[1], row[2], row[3], row[4])
        except Exception:
            self.db.rollback()

        if data:
            _ordenar_por_prioridad()
            return data

        # 3) Fallback: catalogo_clientes (cuando contactos estÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¡n en la misma tabla).
        try:
            for schema_name in _tablas_disponibles("catalogo_clientes"):
                cols = _columnas(schema_name, "catalogo_clientes")
                if not cols:
                    continue
                col_sucursal = _pick(cols, ["nombre_sucursal", "sucursal", "cliente", "nombre_cliente"])
                if not col_sucursal:
                    continue
                col_nombre = _pick(cols, ["nombre_empleado", "nombre_contacto", "contacto", "contacto_emergencia", "nombre_emergencia"])
                col_tel = _pick(cols, ["celular", "telefono_contacto", "telefono", "fono"])
                col_email = _pick(cols, ["correo_contacto", "email", "correo", "mail"])
                col_prio = _pick(cols, ["nro_emergencia", "prioridad", "nivel_prioridad"])
                if not any([col_nombre, col_tel, col_email]):
                    continue

                select_cols = [
                    f'"{col_sucursal}" AS sucursal',
                    f'COALESCE(CAST("{col_nombre}" AS NVARCHAR(MAX)), \'\') AS nombre' if col_nombre else "'' AS nombre",
                    f'COALESCE(CAST("{col_tel}" AS NVARCHAR(MAX)), \'\') AS telefono' if col_tel else "'' AS telefono",
                    f'COALESCE(CAST("{col_email}" AS NVARCHAR(MAX)), \'\') AS email' if col_email else "'' AS email",
                    f'COALESCE(CAST("{col_prio}" AS NVARCHAR(MAX)), \'\') AS prioridad' if col_prio else "'' AS prioridad",
                ]
                sql = text(
                    f"""
                    SELECT {", ".join(select_cols)}
                    FROM "{schema_name}"."catalogo_clientes"
                    WHERE "{col_sucursal}" IS NOT NULL
                      AND TRIM(CAST("{col_sucursal}" AS NVARCHAR(MAX))) <> ''
                    ORDER BY 1
                    """
                )
                for row in self.db.execute(sql).all():
                    _push(row[0], row[1], row[2], row[3], row[4])
        except Exception:
            self.db.rollback()

        _ordenar_por_prioridad()
        return data

    def registrar_envio_correo(self, odt: str, sucursal: str, observacion: str, estado: str) -> None:
        self.db.add(
            RegistroCorreoCliente(
                odt=odt,
                sucursal=sucursal,
                observacion=observacion,
                estado=estado,
            )
        )
        self.db.commit()

    def _enviar_correo_derivacion_automatico(
        self,
        *,
        row: Registro,
        derivacion: str,
        usuario: str,
    ) -> dict[str, Any]:
        odt = str(getattr(row, "odt", "") or "").strip()
        sucursal = str(getattr(row, "cliente", "") or "").strip()
        problema = str(getattr(row, "problema", "") or "").strip()
        derivacion_txt = str(derivacion or "").strip()
        if not odt or not sucursal or not derivacion_txt:
            return {"ok": False, "emails_enviados": 0, "warning": "Datos incompletos para correo automatico."}

        contactos = self.obtener_contactos_por_sucursal()
        sucursal_key = next(
            (k for k in contactos.keys() if self._normalizar_texto(k) == self._normalizar_texto(sucursal)),
            "",
        )
        destinos = contactos.get(sucursal_key) or []
        emails: list[str] = []
        vistos: set[str] = set()
        for contacto in destinos:
            email = str(contacto.get("email") or "").strip()
            email_key = email.lower()
            if email and email_key not in vistos:
                vistos.add(email_key)
                emails.append(email)

        if not emails:
            warning = "No se encontraron contactos con correo para esta sucursal."
            self.db.add(
                RegistroCorreoCliente(
                    odt=odt,
                    sucursal=sucursal,
                    observacion=f"[{usuario}] Correo automatico por derivacion a {derivacion_txt}: {warning}",
                    estado=str(getattr(row, "estado", "") or ""),
                )
            )
            self.db.commit()
            return {"ok": False, "emails_enviados": 0, "warning": warning}

        obs_base = (
            str(getattr(row, "observacion_soporte", "") or "").strip()
            or str(getattr(row, "observacion", "") or "").strip()
            or "Se informa actualizacion de la incidencia."
        )
        observacion = (
            f"ODT {odt} derivada a {derivacion_txt}. "
            f"Detalle: {obs_base}"
        )
        asunto, cuerpo, cuerpo_html = self._build_correo_incidencia_cliente_html(
            sucursal=sucursal,
            problema=problema,
            observacion=observacion,
            con_imagenes=False,
        )
        asunto = f"ATC | Derivacion a {derivacion_txt} - {sucursal}"
        logo_atc = self._logo_atc_bytes()
        cfg_contacto = self._contacto_smtp_runtime_config()

        enviados: set[str] = set()
        errores: list[str] = []
        for email in emails:
            try:
                self._enviar_correo_automatico(
                    email,
                    asunto,
                    cuerpo,
                    html_body=cuerpo_html,
                    logo_bytes=logo_atc,
                    cfg_override=cfg_contacto,
                )
                enviados.add(email.lower())
            except Exception as exc:
                errores.append(str(exc))

        for contacto in destinos:
            email = str(contacto.get("email") or "").strip()
            if email and email.lower() in enviados:
                estado_correo = "enviado"
            elif email:
                estado_correo = "fallido"
            else:
                estado_correo = "sin correo"
            self.db.add(
                RegistroCorreoCliente(
                    odt=odt,
                    sucursal=sucursal,
                    observacion=(
                        f"[{usuario}] Correo automatico por derivacion a {derivacion_txt}. "
                        f"Contacto: {contacto.get('nombre') or '-'} | Correo: {email or '-'} | "
                        f"Estado correo: {estado_correo}"
                    ),
                    estado=str(getattr(row, "estado", "") or ""),
                )
            )
        self.db.commit()
        return {
            "ok": bool(enviados),
            "emails_enviados": len(enviados),
            "emails_fallidos": max(0, len(emails) - len(enviados)),
            "warning": " | ".join(errores[:3]) if errores else "",
        }

    def _enviar_correo_derivacion_automatico_silencioso(
        self,
        *,
        row: Registro,
        derivacion: str,
        usuario: str,
    ) -> dict[str, Any]:
        try:
            return self._enviar_correo_derivacion_automatico(
                row=row,
                derivacion=derivacion,
                usuario=usuario,
            )
        except Exception as exc:
            LOGGER.exception("No se pudo enviar correo automatico de derivacion para ODT %s.", getattr(row, "odt", ""))
            return {"ok": False, "emails_enviados": 0, "warning": str(exc)}

    def _enviar_correo_automatico(
        self,
        to_email: str,
        subject: str,
        body: str,
        html_body: str | None = None,
        logo_bytes: bytes | None = None,
        attachments: list[dict[str, Any]] | None = None,
        cfg_override: dict[str, Any] | None = None,
        cc_emails: list[str] | None = None,
        bcc_emails_extra: list[str] | None = None,
    ) -> None:
        cfg = cfg_override if cfg_override is not None else self._smtp_runtime_config()
        if not cfg["enabled"]:
            raise ValueError("El envio automatico de correo esta deshabilitado (SMTP_ENABLED=false).")

        host = str(cfg["host"] or "").strip()
        port = int(cfg["port"] or 0)
        username = str(cfg["username"] or "").strip()
        password = str(cfg["password"] or "")
        from_email = str(cfg["from_email"] or "").strip()
        from_name = str(cfg["from_name"] or "").strip()
        use_tls = bool(cfg["use_tls"])
        use_ssl = bool(cfg["use_ssl"])
        timeout = int(cfg["timeout"] or 20)

        if not host or not port or not from_email:
            raise ValueError("SMTP incompleto. Configura SMTP_HOST, SMTP_PORT y SMTP_FROM_EMAIL (o SMTP_USERNAME).")

        bcc_emails: list[str] = list(cfg.get("bcc_emails") or []) + list(bcc_emails_extra or [])

        msg = EmailMessage()
        msg["Subject"] = subject
        if from_name:
            msg["From"] = f"{from_name} <{from_email}>"
        else:
            msg["From"] = from_email
        msg["To"] = to_email
        if cc_emails:
            msg["Cc"] = ", ".join(cc_emails)
        if bcc_emails:
            msg["Bcc"] = ", ".join(bcc_emails)
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype="html")
            if logo_bytes:
                try:
                    logo_subtype = "png" if logo_bytes.startswith(b"\x89PNG\r\n\x1a\n") else "jpeg"
                    html_part = msg.get_payload()[-1]
                    html_part.add_related(
                        logo_bytes,
                        maintype="image",
                        subtype=logo_subtype,
                        cid="<logoatc>",
                    )
                except Exception:
                    pass
        for item in attachments or []:
            try:
                nombre = str(item.get("nombre") or "adjunto").strip() or "adjunto"
                tipo = str(item.get("tipo") or "application/octet-stream").strip()
                contenido = item.get("contenido") or b""
                if isinstance(contenido, str):
                    import base64

                    contenido = base64.b64decode(contenido)
                maintype, subtype = (tipo.split("/", 1) + ["octet-stream"])[:2] if "/" in tipo else ("application", "octet-stream")
                msg.add_attachment(contenido, maintype=maintype, subtype=subtype, filename=nombre)
            except Exception:
                continue

        try:
            if use_ssl:
                with smtplib.SMTP_SSL(host, port, timeout=timeout) as smtp:
                    if username:
                        smtp.login(username, password)
                    smtp.send_message(msg)
            else:
                with smtplib.SMTP(host, port, timeout=timeout) as smtp:
                    smtp.ehlo()
                    if use_tls:
                        smtp.starttls()
                        smtp.ehlo()
                    if username:
                        smtp.login(username, password)
                    smtp.send_message(msg)
        except Exception as exc:
            raise ValueError(f"No se pudo enviar correo automatico a {to_email}: {exc}") from exc

    def _enviar_correo_cambio_region_comuna(
        self,
        *,
        nombre: str,
        region_anterior: str,
        comuna_anterior: str,
        region_nueva: str,
        comuna_nueva: str,
        observacion: str,
        usuario: str,
    ) -> None:
        asunto = f"ATC | Cambio de Región/Comuna — {nombre}"
        anterior = f"{region_anterior or '(vacío)'} / {comuna_anterior or '(vacío)'}"
        nueva = f"{region_nueva or '(vacío)'} / {comuna_nueva or '(vacío)'}"
        cuerpo = (
            f'Se actualizó la Región/Comuna de "{nombre}".\n\n'
            f"Anterior: {anterior}\n"
            f"Nuevo: {nueva}\n\n"
            f"Motivo: {observacion}\n\n"
            f"Realizado por: {usuario or 'Desconocido'}\n"
            f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        cfg_contacto = self._contacto_smtp_runtime_config()
        self._enviar_correo_automatico(
            "contacto@alguientecuida.cl",
            asunto,
            cuerpo,
            cfg_override=cfg_contacto,
        )

    def actualizar_region_comuna_sucursal(
        self,
        *,
        tabla: str,
        entidad_id: int,
        region: str,
        comuna: str,
        observacion: str,
        usuario: str,
    ) -> dict[str, Any]:
        tabla_norm = (tabla or "").strip().lower()
        region_txt = (region or "").strip()
        comuna_txt = (comuna or "").strip()
        observacion_txt = (observacion or "").strip()

        if tabla_norm not in {"sucursal", "cliente"} or not entidad_id:
            raise ValueError("No se pudo identificar el registro a actualizar.")
        if not observacion_txt:
            raise ValueError("Debes indicar una observación explicando el cambio.")

        valido, region_google, mensaje = validar_region_comuna_google(region_txt, comuna_txt)
        if not valido:
            raise ValueError(mensaje)
        region_txt = region_google or region_txt

        modelo = SucursalBBDD if tabla_norm == "sucursal" else ClienteBBDD
        entidad = self.db.get(modelo, int(entidad_id))
        if not entidad:
            raise ValueError("No se encontró el registro a actualizar.")

        nombre = str(getattr(entidad, "nombre_sucursal", "") or getattr(entidad, "cliente", "") or "").strip()
        region_anterior = str(entidad.region or "").strip()
        comuna_anterior = str(entidad.comuna or "").strip()

        entidad.region = region_txt
        entidad.comuna = comuna_txt
        self.db.commit()

        try:
            self._enviar_correo_cambio_region_comuna(
                nombre=nombre,
                region_anterior=region_anterior,
                comuna_anterior=comuna_anterior,
                region_nueva=region_txt,
                comuna_nueva=comuna_txt,
                observacion=observacion_txt,
                usuario=usuario,
            )
            email_enviado = True
        except Exception:
            LOGGER.exception("No se pudo enviar el correo de aviso de cambio de región/comuna para %s", nombre)
            email_enviado = False

        return {"ok": True, "region": region_txt, "comuna": comuna_txt, "email_enviado": email_enviado}

    def registrar_envio_informacion_contacto(
        self, data: EnviarInformacionContactoRequest
    ) -> dict[str, Any]:
        odt = str(data.odt or "").strip()
        sucursal = str(data.sucursal or "").strip()
        if not odt or not sucursal:
            raise ValueError("Debes indicar ODT y sucursal.")

        destinos_validos: list[ContactoDestinoRequest] = []
        for d in list(data.destinos or []):
            email = str(d.email or "").strip()
            telefono = str(d.telefono or "").strip()
            if email or telefono:
                destinos_validos.append(d)
        if not destinos_validos:
            raise ValueError("Debes seleccionar al menos un contacto con correo o telefono.")

        usuario = self.get_usuario_actual(str(data.token or ""))
        usuario = usuario if usuario and usuario != "Desconocido" else "Sistema"
        registro = self.db.scalar(select(Registro).where(Registro.odt == odt).limit(1))
        problema = str(data.problema or (registro.problema if registro else "") or "").strip()
        estado = str(data.estado or (registro.estado if registro else "") or "").strip() or "En Proceso"
        obs_base = str(data.observacion or (registro.observacion if registro else "") or "").strip()
        tecnico = str(data.tecnico or (getattr(registro, "tecnicos", "") if registro else "") or "").strip()
        acompanante = str(data.acompanante or (registro.acompanante if registro else "") or "").strip()

        fecha_visita = self._parse_fecha_visita(data.fecha_visita or "")
        if not fecha_visita and registro and registro.fecha_cierre:
            fecha_visita = registro.fecha_cierre
        if not fecha_visita:
            fecha_visita = datetime.now(ZoneInfo(settings.timezone))

        envios_previos = self._obtener_envios_informacion_contacto_por_odt().get(odt, {})
        claves_previas = set(envios_previos.get("claves") or [])
        destinos_pendientes: list[ContactoDestinoRequest] = []
        destinos_omitidos = 0
        for destino in destinos_validos:
            clave = self._clave_contacto_envio(
                nombre=str(destino.nombre or "").strip(),
                email=str(destino.email or "").strip(),
                telefono=str(destino.telefono or "").strip(),
            )
            if clave and clave in claves_previas:
                destinos_omitidos += 1
                continue
            destinos_pendientes.append(destino)

        if not destinos_pendientes:
            raise ValueError("La informacion ya fue enviada a los contactos seleccionados.")

        emails_unicos: list[str] = []
        seen_emails: set[str] = set()
        total_emails = 0
        total_telefonos = 0
        total_destinos = 0
        for destino in destinos_pendientes:
            nombre = str(destino.nombre or "").strip()
            telefono = str(destino.telefono or "").strip()
            email = str(destino.email or "").strip()
            prioridad = str(destino.prioridad or "").strip()
            total_emails += 1 if email else 0
            total_telefonos += 1 if telefono else 0
            total_destinos += 1

            email_key = email.lower()
            if email and email_key not in seen_emails:
                seen_emails.add(email_key)
                emails_unicos.append(email)

        if not emails_unicos:
            raise ValueError("Por ahora solo esta habilitado correo. Selecciona al menos un contacto con email.")

        asunto, cuerpo, cuerpo_html = self._build_correo_visita_html(
            odt=odt,
            sucursal=sucursal,
            problema=problema,
            estado=estado,
            tecnico=tecnico,
            acompanante=acompanante,
            fecha_visita=fecha_visita,
            observacion=obs_base,
        )
        logo_atc = self._logo_atc_bytes()
        if not logo_atc:
            cuerpo_html = cuerpo_html.replace(
                '<img src="cid:logoatc" alt="ATC" style="height:58px;width:auto;display:block;margin:0 auto 12px;" />',
                "",
            )

        # EN PAUSA: la cuenta propia (SMTP_VISITA_* / jperez@alguientecuida.cl)
        # todavia no tiene password configurado, asi que por ahora se sigue
        # enviando por contacto@alguientecuida.cl, con jperez en copia. Cuando
        # esa cuenta este lista, cambiar cfg_override a
        # self._visita_smtp_runtime_config() y sacar jperez de cc_destinatarios.
        cfg_contacto = self._contacto_smtp_runtime_config()
        emails_enviados: set[str] = set()
        errores_email: list[str] = []
        # Un solo correo con el resto de los contactos + jperez en copia, en
        # vez de un correo individual por cada destinatario.
        to_principal = emails_unicos[0]
        cc_destinatarios = emails_unicos[1:] + ["jperez@alguientecuida.cl"]
        try:
            self._enviar_correo_automatico(
                to_principal,
                asunto,
                cuerpo,
                html_body=cuerpo_html,
                logo_bytes=logo_atc,
                cfg_override=cfg_contacto,
                cc_emails=cc_destinatarios,
            )
            emails_enviados.update(email.lower() for email in emails_unicos)
        except Exception as exc:
            errores_email.append(str(exc))

        for destino in destinos_pendientes:
            nombre = str(destino.nombre or "").strip()
            telefono = str(destino.telefono or "").strip()
            email = str(destino.email or "").strip()
            prioridad = str(destino.prioridad or "").strip()
            if email and email.lower() in emails_enviados:
                estado_correo = "enviado"
            elif email:
                estado_correo = "fallido"
            else:
                estado_correo = "sin correo"

            partes = [
                f"[{usuario}] Envio de informacion a contacto de cliente.",
                f"Problema: {problema or '-'}",
                f"Contacto: {nombre or '-'}",
                f"Telefono: {telefono or '-'}",
                f"Correo: {email or '-'}",
                f"Prioridad: {prioridad or '-'}",
                f"Estado correo: {estado_correo}",
                "WhatsApp: pendiente API",
                f"Fecha visita enviada: {fecha_visita.strftime('%d/%m/%Y')}",
                f"Tecnico: {tecnico or '-'}",
                f"Acompanante: {acompanante or '-'}",
            ]
            if obs_base:
                partes.append(f"Detalle: {obs_base}")
            observacion_log = " | ".join(partes)
            self.db.add(
                RegistroCorreoCliente(
                    odt=odt,
                    sucursal=sucursal,
                    observacion=observacion_log,
                    estado=estado,
                )
            )

        self.db.commit()
        if not emails_enviados:
            detalle = errores_email[0] if errores_email else "No se pudo enviar ningun correo."
            raise ValueError(detalle)

        return {
            "ok": True,
            "odt": odt,
            "sucursal": sucursal,
            "destinos": total_destinos,
            "emails": total_emails,
            "emails_enviados": len(emails_enviados),
            "emails_fallidos": max(0, len(emails_unicos) - len(emails_enviados)),
            "telefonos": total_telefonos,
            "omitidos": destinos_omitidos,
            "usuario": usuario,
            "whatsapp_pendiente": total_telefonos,
            "warning": " | ".join(errores_email[:3]) if errores_email else "",
        }

    def _build_correo_incidencia_cliente_html(
        self,
        *,
        sucursal: str,
        problema: str,
        observacion: str,
        con_imagenes: bool,
    ) -> tuple[str, str, str]:
        titulo = "Incidencia Tecnica"
        mensaje = "Se informa una incidencia tecnica detectada."
        problema_key = self._normalizar_texto(problema)
        if problema_key == "desconexion":
            titulo = "Incidencia por Desconexion del Sistema"
            mensaje = "Se ha detectado una desconexion del sistema de monitoreo."
        elif problema_key == "problema de parlante":
            titulo = "Incidencia en Sistema de Audio"
            mensaje = "Se ha informado un inconveniente en el sistema de parlantes."
        elif problema_key == "problema de alarma":
            titulo = "Incidencia en Sistema de Alarma"
            mensaje = "Se ha detectado un problema en el sistema de alarma."
        elif problema_key == "problema de visual":
            titulo = "Incidencia en Sistema de Visualizacion"
            mensaje = "Se ha informado un inconveniente en el sistema visual."
        elif problema_key == "hora y/o fecha cambiada":
            titulo = "Ajuste de Fecha y/u Hora del Sistema"
            mensaje = "Se ha realizado una modificacion en la configuracion de fecha y/u hora."

        detalle_imagenes = "Imagen/es adjunta/s en este correo." if con_imagenes else "Sin imágenes adjuntas."
        subject = f"ATC · Notificación: {titulo} — {sucursal}"
        text_body = (
            "Estimados/as,\n\n"
            f"{mensaje}\n\n"
            f"Sucursal: {sucursal}\n"
            f"Incidencia: {problema or titulo}\n\n"
            f"Detalle:\n{observacion}\n\n"
            "Quedamos atentos a sus comentarios.\n\n"
            "Atentamente,\nEquipo Técnico\nAlguien Te Cuida\n\n"
            "—\nEste mensaje fue generado automáticamente. Por favor, no responda a esta dirección."
        )
        html_body = f"""\
<!doctype html>
<html lang="es">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <title>{html_escape(titulo)}</title>
  </head>
  <body style="margin:0;padding:0;background:#f4f6f9;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;color:#1f2937;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f4f6f9;padding:32px 16px;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:600px;background:#ffffff;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;">
            <tr>
              <td style="height:3px;background:#0f3048;line-height:3px;font-size:0;">&nbsp;</td>
            </tr>
            <tr>
              <td style="padding:26px 32px 8px;">
                <img src="cid:logoatc" alt="Alguien Te Cuida" style="height:34px;width:auto;display:block;">
              </td>
            </tr>
            <tr>
              <td style="padding:14px 32px 4px;">
                <div style="font-size:11px;font-weight:600;letter-spacing:.08em;text-transform:uppercase;color:#6b7280;">Notificación al cliente</div>
                <div style="margin-top:6px;font-size:20px;font-weight:600;color:#111827;letter-spacing:-.01em;line-height:1.3;">{html_escape(titulo)}</div>
              </td>
            </tr>
            <tr>
              <td style="padding:18px 32px 6px;color:#374151;font-size:14.5px;line-height:1.65;">
                <p style="margin:0 0 12px;">Estimados/as,</p>
                <p style="margin:0;">{html_escape(mensaje)}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:20px 32px 4px;">
                <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border:1px solid #e5e7eb;border-radius:6px;border-collapse:separate;">
                  <tr>
                    <td style="padding:12px 14px;color:#6b7280;font-size:12px;font-weight:500;letter-spacing:.02em;text-transform:uppercase;width:40%;">Sucursal</td>
                    <td style="padding:12px 14px;color:#111827;font-size:14px;text-align:right;">{html_escape(sucursal)}</td>
                  </tr>
                  <tr>
                    <td style="padding:12px 14px;color:#6b7280;font-size:12px;font-weight:500;letter-spacing:.02em;text-transform:uppercase;border-top:1px solid #f3f4f6;">Incidencia</td>
                    <td style="padding:12px 14px;color:#111827;font-size:14px;text-align:right;border-top:1px solid #f3f4f6;">{html_escape(problema or titulo)}</td>
                  </tr>
                </table>
              </td>
            </tr>
            <tr>
              <td style="padding:18px 32px 6px;">
                <div style="color:#6b7280;font-size:12px;font-weight:500;letter-spacing:.02em;text-transform:uppercase;margin-bottom:8px;">Detalle</div>
                <div style="color:#374151;font-size:14px;line-height:1.65;padding:14px 16px;background:#fafbfc;border:1px solid #e5e7eb;border-radius:6px;">
                  {html_escape(observacion).replace(chr(10), "<br>")}
                </div>
                <div style="margin-top:10px;color:#9ca3af;font-size:12px;line-height:1.5;">{html_escape(detalle_imagenes)}</div>
              </td>
            </tr>
            <tr>
              <td style="padding:18px 32px 28px;color:#374151;font-size:14.5px;line-height:1.65;">
                <p style="margin:0 0 16px;">Quedamos atentos a sus comentarios.</p>
                <p style="margin:0;">
                  Atentamente,<br>
                  <span style="color:#111827;font-weight:600;">Equipo Técnico</span><br>
                  <span style="color:#6b7280;">Alguien Te Cuida</span>
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:16px 32px;background:#fafbfc;border-top:1px solid #e5e7eb;text-align:center;color:#9ca3af;font-size:11.5px;line-height:1.55;">
                Este mensaje fue generado automáticamente. Por favor, no responda a esta dirección.<br>
                © Alguien Te Cuida · Soporte ATC
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""
        return subject, text_body, html_body

    def registrar_envio_correo_coordinacion(
        self, data: EnviarInformacionContactoRequest
    ) -> dict[str, Any]:
        odt = str(data.odt or "").strip()
        sucursal = str(data.sucursal or "").strip()
        if not odt or not sucursal:
            raise ValueError("Debes indicar ODT y sucursal.")

        destinos = [
            d
            for d in list(data.destinos or [])
            if str(d.email or "").strip() or str(d.telefono or "").strip()
        ]
        if not destinos:
            contactos = self.obtener_contactos_por_sucursal()
            sucursal_key = next(
                (k for k in contactos.keys() if self._normalizar_texto(k) == self._normalizar_texto(sucursal)),
                "",
            )
            destinos = [
                ContactoDestinoRequest(**c)
                for c in (contactos.get(sucursal_key) or [])
                if str(c.get("email") or "").strip() or str(c.get("telefono") or "").strip()
            ]
        if not destinos:
            raise ValueError("No se encontraron contactos para esta sucursal.")

        emails_unicos: list[str] = []
        seen_emails: set[str] = set()
        for destino in destinos:
            email = str(destino.email or "").strip()
            email_key = email.lower()
            if email and email_key not in seen_emails:
                seen_emails.add(email_key)
                emails_unicos.append(email)
        if not emails_unicos:
            raise ValueError("Por ahora solo esta habilitado correo. Selecciona al menos un contacto con email.")

        registro = self.db.scalar(select(Registro).where(Registro.odt == odt).limit(1))
        problema = str(data.problema or (registro.problema if registro else "") or "").strip()
        observacion = str(data.observacion or (registro.observacion if registro else "") or "").strip()
        if not observacion:
            raise ValueError("La observacion no puede estar vacia.")

        imagenes = [
            img
            for img in list(data.imagenes or [])
            if str(img.get("contenido") or "").strip()
        ]
        asunto, cuerpo, cuerpo_html = self._build_correo_incidencia_cliente_html(
            sucursal=sucursal,
            problema=problema,
            observacion=observacion,
            con_imagenes=bool(imagenes),
        )
        logo_atc = self._logo_atc_bytes()
        cfg_contacto = self._contacto_smtp_runtime_config()
        emails_enviados: set[str] = set()
        errores: list[str] = []
        for email in emails_unicos:
            email_key = email.lower()
            if email_key in emails_enviados:
                continue
            try:
                self._enviar_correo_automatico(
                    email,
                    asunto,
                    cuerpo,
                    html_body=cuerpo_html,
                    logo_bytes=logo_atc,
                    attachments=imagenes,
                    cfg_override=cfg_contacto,
                )
                emails_enviados.add(email_key)
            except Exception as exc:
                errores.append(str(exc))

        usuario = self.get_usuario_actual(str(data.token or ""))
        usuario = usuario if usuario and usuario != "Desconocido" else "Sistema"
        for destino in destinos:
            nombre = str(destino.nombre or "").strip()
            telefono = str(destino.telefono or "").strip()
            email = str(destino.email or "").strip()
            if email and email.lower() in emails_enviados:
                estado_correo = "enviado"
            elif email:
                estado_correo = "fallido"
            else:
                estado_correo = "sin correo"
            self.db.add(
                RegistroCorreoCliente(
                    odt=odt,
                    sucursal=sucursal,
                    observacion=(
                        f"[{usuario}] Envio de incidencia a cliente. Problema: {problema or '-'} | "
                        f"Contacto: {nombre or '-'} | Telefono: {telefono or '-'} | "
                        f"Correo: {email or '-'} | Estado correo: {estado_correo} | Detalle: {observacion}"
                        " | WhatsApp: Proximamente"
                    ),
                    estado=str(data.estado or (registro.estado if registro else "") or ""),
                )
            )
        self.db.commit()
        if not emails_enviados:
            raise ValueError(errores[0] if errores else "No se pudo enviar ningun correo.")
        return {
            "ok": True,
            "odt": odt,
            "sucursal": sucursal,
            "emails_enviados": len(emails_enviados),
            "emails_fallidos": max(0, len(emails_unicos) - len(emails_enviados)),
            "telefonos": sum(1 for d in destinos if str(d.telefono or "").strip()),
            "whatsapp_pendiente": sum(1 for d in destinos if str(d.telefono or "").strip()),
            "warning": " | ".join(errores[:3]) if errores else "",
        }

    def obtener_resumen_correos_por_odt(self) -> dict[str, dict[str, Any]]:
        stmt = (
            select(
                RegistroCorreoCliente.odt,
                RegistroCorreoCliente.fecha_envio,
                RegistroCorreoCliente.observacion,
            )
            .order_by(RegistroCorreoCliente.odt, RegistroCorreoCliente.fecha_envio, RegistroCorreoCliente.id)
        )
        resumen: dict[str, dict[str, Any]] = {}
        for odt, fecha_envio, observacion in self.db.execute(stmt).all():
            odt_key = str(odt or "").strip()
            if not odt_key:
                continue
            obs = str(observacion or "")
            obs_norm = self._normalizar_texto(obs)
            estado_correo = self._normalizar_texto(self._extraer_valor_log_envio(obs, "Estado correo"))
            estado_whatsapp = self._normalizar_texto(self._extraer_valor_log_envio(obs, "Estado WhatsApp"))
            email = self._extraer_valor_log_envio(obs, "Correo")
            telefono = self._extraer_valor_log_envio(obs, "Telefono")

            es_correo_enviado = estado_correo == "enviado" and bool(str(email or "").strip())
            es_mensaje_enviado = estado_whatsapp == "enviado" or (
                "envio de mensaje" in obs_norm and "enviado" in obs_norm and bool(str(telefono or "").strip())
            )
            if not es_correo_enviado and not es_mensaje_enviado:
                continue

            item = resumen.setdefault(
                odt_key,
                {
                    "cantidad_correos": 0,
                    "cantidad_mensajes": 0,
                    "ultimo_correo": None,
                    "ultimo_mensaje": None,
                },
            )
            if es_correo_enviado:
                item["cantidad_correos"] = int(item.get("cantidad_correos") or 0) + 1
                if fecha_envio and (not item.get("ultimo_correo") or fecha_envio > item["ultimo_correo"]):
                    item["ultimo_correo"] = fecha_envio
            if es_mensaje_enviado:
                item["cantidad_mensajes"] = int(item.get("cantidad_mensajes") or 0) + 1
                if fecha_envio and (not item.get("ultimo_mensaje") or fecha_envio > item["ultimo_mensaje"]):
                    item["ultimo_mensaje"] = fecha_envio

        for item in resumen.values():
            item["total"] = int(item.get("cantidad_correos") or 0)
            item["ultimo_envio"] = item.get("ultimo_correo")
        return resumen

    def obtener_cantidad_correos_por_odt(self) -> dict[str, int]:
        return {
            odt: int(info.get("cantidad_correos") or info.get("total") or 0)
            for odt, info in self.obtener_resumen_correos_por_odt().items()
        }

    def obtener_registros_derivaciones(self) -> list[list[Any]]:
        correos = self.obtener_resumen_correos_por_odt()
        rows = self.db.scalars(select(Registro).order_by(Registro.id.desc())).all()
        out: list[list[Any]] = []
        for r in rows:
            correo_info = correos.get(r.odt, {})
            out.append(
                [
                    r.odt,
                    _to_ddmmyyyy_hhmm(r.fecha_registro),
                    r.cliente,
                    r.problema,
                    r.derivacion,
                    r.observacion,
                    r.estado,
                    r.observacion_final,
                    int(correo_info.get("cantidad_correos") or correo_info.get("total") or 0),
                    getattr(r, "observacion_soporte", "") or "",
                    getattr(r, "observacion_servicio", "") or "",
                    _to_ddmmyyyy_hhmm(correo_info.get("ultimo_correo") or correo_info.get("ultimo_envio")),
                    _to_ddmmyyyy_hhmm(correo_info.get("ultimo_mensaje")),
                    int(correo_info.get("cantidad_mensajes") or 0),
                    getattr(r, "observacion_coordinacion", "") or "",
                ]
            )
        return out

    # =========================
    # RENDICIONES
    # =========================
    def existe_nro_documento_duplicado(self, nro_documento: str) -> bool:
        if not nro_documento:
            return False
        stmt = select(Rendicion).where(
            Rendicion.nro_documento == nro_documento.strip(),
            Rendicion.estado_revision != "Rechazado",
        )
        return self.db.scalar(stmt) is not None

    def guardar_boleta_rendicion(
        self,
        *,
        content: bytes,
        filename: str,
        tecnico: str = "",
        odt: str = "",
    ) -> str:
        if not isinstance(content, (bytes, bytearray)) or not content:
            raise ValueError("Archivo de boleta vacio.")

        nombre_original = str(filename or "boleta.jpg").strip() or "boleta.jpg"
        ext = Path(nombre_original).suffix.lower()
        permitidas = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".heic", ".heif"}
        if ext not in permitidas:
            raise ValueError("Formato de imagen no permitido para boleta.")

        odt_seguro = re.sub(r"[^A-Za-z0-9_-]+", "", str(odt or "").strip().upper()) or "SIN_ODT"
        tecnico_seguro = re.sub(r"[^A-Za-z0-9_-]+", "", str(tecnico or "").strip().upper()) or "SIN_TECNICO"
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        unique = uuid.uuid4().hex[:10]
        nombre_final = f"{ts}_{odt_seguro}_{tecnico_seguro}_{unique}{ext}"
        rendicion_ref = f"{odt_seguro}_{tecnico_seguro}_{ts}_{unique}"

        mime = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
        }.get(ext)
        contenido_guardar = (
            optimize_image_bytes(bytes(content), content_type=mime, jpeg_quality=60, webp_quality=60)
            if mime
            else bytes(content)
        )
        mime_final = mime or (mimetypes.guess_type(nombre_final)[0] or "application/octet-stream")

        try:
            drive_result = upload_rendicion_boleta_to_drive(
                tecnico=tecnico or "Sin tecnico",
                rendicion_ref=rendicion_ref,
                content=contenido_guardar,
                filename=nombre_final,
                mime_type=mime_final,
            )
            LOGGER.info("Boleta rendicion subida a Drive: odt=%s tecnico=%s -> %s", odt, tecnico, drive_result["public_uri"])
            return drive_result["public_uri"]
        except Exception:
            LOGGER.exception("No se pudo subir boleta de rendicion a Drive (odt=%s, tecnico=%s)", odt, tecnico)
            return _guardar_en_cuarentena_drive(
                "rendicion_boleta",
                rendicion_ref,
                nombre_final,
                contenido_guardar,
                {
                    "tecnico": tecnico or "",
                    "rendicion_ref": rendicion_ref,
                    "filename": nombre_final,
                    "mime_type": mime_final,
                },
            )

    def _generar_informe_rendicion_pdf(self, rend: "Rendicion") -> str | None:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm, mm
            from reportlab.lib.colors import HexColor, white
            from reportlab.platypus import (
                BaseDocTemplate, Frame, PageTemplate,
                Table, TableStyle, Paragraph, Spacer, KeepTogether,
            )
            from reportlab.platypus import Image as RLImage
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            import io

            C_DARK   = HexColor("#0b1424")
            C_ORANGE = HexColor("#f4a672")
            C_ORDK   = HexColor("#c2410c")
            C_BG     = HexColor("#f7f8fa")
            C_BORDER = HexColor("#e5e7eb")
            C_TEXT   = HexColor("#111827")
            C_SOFT   = HexColor("#4b5563")
            C_YELLOW = HexColor("#fde68a")
            C_GREY   = HexColor("#9ca3af")

            W, H = A4
            pad    = 1.4 * cm    # lateral padding of body content
            HEADER_H  = 2.6 * cm
            ORANGE_H  = 5        # points
            FOOTER_H  = 1.0 * cm
            BODY_TOP  = HEADER_H + ORANGE_H + 10   # y from top of page where body starts
            BODY_BOT  = FOOTER_H + 8

            fw = W - 2 * pad     # frame / inner width

            fecha_str    = rend.fecha_documento.strftime("%d/%m/%Y") if rend.fecha_documento else "-"
            registro_str = rend.fecha_registro.strftime("%d/%m/%Y %H:%M") if rend.fecha_registro else "-"
            monto_str    = f"$ {int(rend.monto_total or 0):,}".replace(",", ".")

            logo_path = _ATC_ROOT / "static" / "img" / "logo-atc.png"
            logo_w = 3.0 * cm
            logo_h = 1.5 * cm

            # ── Styles ───────────────────────────────────────────────────
            st_label = ParagraphStyle("lbl", fontName="Helvetica-Bold",  fontSize=7.5,
                                      textColor=C_SOFT, leading=10, spaceAfter=1)
            st_value = ParagraphStyle("val", fontName="Helvetica",       fontSize=10,
                                      textColor=C_TEXT, leading=13)
            st_sec   = ParagraphStyle("sec", fontName="Helvetica-Bold",  fontSize=8.5,
                                      textColor=C_ORDK, leading=12, spaceBefore=10, spaceAfter=4)
            st_img_lbl = ParagraphStyle("il", fontName="Helvetica",      fontSize=8,
                                        textColor=C_SOFT, leading=10, alignment=TA_CENTER)

            # ── Canvas callback: draws header + orange bar + footer ───────
            def draw_page(canvas, doc):
                canvas.saveState()
                # Dark header
                canvas.setFillColor(C_DARK)
                canvas.rect(0, H - HEADER_H, W, HEADER_H, fill=1, stroke=0)
                # Logo
                if logo_path.exists():
                    try:
                        canvas.drawImage(
                            str(logo_path),
                            pad, H - HEADER_H + (HEADER_H - logo_h) / 2,
                            width=logo_w, height=logo_h,
                            preserveAspectRatio=True, mask="auto",
                        )
                    except Exception:
                        pass
                # Title
                canvas.setFillColor(white)
                canvas.setFont("Helvetica-Bold", 17)
                tx = pad + logo_w + 0.5 * cm
                canvas.drawString(tx, H - HEADER_H + 1.4 * cm, "INFORME DE RENDICIÓN")
                canvas.setFillColor(C_YELLOW)
                canvas.setFont("Helvetica", 9)
                canvas.drawString(tx, H - HEADER_H + 0.55 * cm,
                                  f"Alguien Te Cuida  ·  #{rend.id}")
                # Orange bar
                canvas.setFillColor(C_ORANGE)
                canvas.rect(0, H - HEADER_H - ORANGE_H, W, ORANGE_H, fill=1, stroke=0)
                # Footer bar
                canvas.setFillColor(C_DARK)
                canvas.rect(0, 0, W, FOOTER_H, fill=1, stroke=0)
                canvas.setFillColor(C_GREY)
                canvas.setFont("Helvetica", 7)
                canvas.drawCentredString(
                    W / 2, FOOTER_H / 2 - 3,
                    f"Documento generado automáticamente  ·  Alguien Te Cuida  ·  {registro_str}",
                )
                canvas.restoreState()

            # ── BaseDocTemplate + Frame ───────────────────────────────────
            frame = Frame(
                pad, BODY_BOT,
                fw, H - BODY_TOP - BODY_BOT,
                leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
            )
            template = PageTemplate(id="main", frames=[frame], onPage=draw_page)

            buf = io.BytesIO()
            doc = BaseDocTemplate(
                buf, pagesize=A4,
                pageTemplates=[template],
                leftMargin=0, rightMargin=0,
                topMargin=0, bottomMargin=0,
                title=f"Informe Rendición #{rend.id}",
                author="Alguien Te Cuida",
            )

            # ── Story ─────────────────────────────────────────────────────
            story = []

            def cell(label: str, value: str):
                return [Paragraph(label, st_label), Paragraph(str(value or "-"), st_value)]

            sep    = 0.4 * cm
            col_w  = (fw - sep) / 2
            grid_data = [
                [cell("TÉCNICO", rend.tecnico or "-"),                cell("ODT", rend.odt or "-")],
                [cell("CLIENTE", rend.cliente or "-"),                cell("CÓDIGO DIARIO", rend.codigo_diario or "-")],
                [cell("TIPO DE GASTO", rend.tipo_gasto or "-"),       cell("TIPO DE DOCUMENTO", rend.tipo_documento or "-")],
                [cell("NRO DE DOCUMENTO", rend.nro_documento or "-"), cell("FECHA DEL DOCUMENTO", fecha_str)],
                [cell("MONTO TOTAL", monto_str),                      cell("FECHA DE REGISTRO", registro_str)],
            ]
            if rend.descripcion and rend.descripcion.strip():
                grid_data.append([cell("DESCRIPCIÓN", rend.descripcion), ["", ""]])

            flat_rows = []
            for row in grid_data:
                lft = row[0]
                rgt = row[1] if row[1] != ["", ""] else ["", ""]
                flat_rows.append([lft[0], lft[1], Spacer(sep, 1), rgt[0], rgt[1]])

            cw_lbl = col_w * 0.37
            cw_val = col_w * 0.63
            detail = Table(
                flat_rows,
                colWidths=[cw_lbl, cw_val, sep, cw_lbl, cw_val],
            )
            detail.setStyle(TableStyle([
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING",   (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
                ("TOPPADDING",    (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING",   (2, 0), (2, -1), 0),
                ("RIGHTPADDING",  (2, 0), (2, -1), 0),
                ("TOPPADDING",    (2, 0), (2, -1), 0),
                ("BOTTOMPADDING", (2, 0), (2, -1), 0),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [white, C_BG]),
                ("LINEBELOW",     (0, 0), (-1, -2), 0.5, C_BORDER),
                ("BOX",           (0, 0), (-1, -1), 1,   C_BORDER),
            ]))
            story.append(detail)
            story.append(Spacer(1, 14))

            # Boleta
            boleta_url = str(rend.url_boleta or "").strip()
            if boleta_url:
                boleta_path: Path | None = None
                boleta_bytes: bytes | None = None
                boleta_label = "boleta"
                if boleta_url.startswith("/api/incidencias/drive-image/"):
                    file_id = boleta_url.rsplit("/", 1)[-1]
                    try:
                        boleta_bytes, _mime, boleta_label = download_support_drive_file_bytes(file_id=file_id)
                    except Exception:
                        logging.exception("No se pudo descargar boleta desde Drive (file_id=%s)", file_id)
                elif boleta_url.startswith("/"):
                    boleta_path = _url_to_path(boleta_url)

                tiene_boleta = boleta_bytes is not None or bool(boleta_path and boleta_path.exists())
                if tiene_boleta:
                    story.append(Paragraph("IMAGEN DEL DOCUMENTO", st_sec))
                    try:
                        from PIL import Image as PILImage
                        img_source = io.BytesIO(boleta_bytes) if boleta_bytes is not None else boleta_path
                        with PILImage.open(img_source) as pil_img:
                            iw, ih = pil_img.size
                        if boleta_bytes is not None:
                            img_source.seek(0)
                        ratio  = ih / iw if iw else 1
                        max_h  = 13 * cm
                        img_w  = min(fw, max_h / ratio)
                        img_h  = img_w * ratio
                        if img_h > max_h:
                            img_h = max_h
                            img_w = img_h / ratio
                        img_el = RLImage(
                            img_source if boleta_bytes is not None else str(boleta_path),
                            width=img_w, height=img_h,
                        )
                        img_table = Table([[img_el]], colWidths=[fw])
                        img_table.setStyle(TableStyle([
                            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
                            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                            ("BOX",           (0, 0), (-1, -1), 1, C_BORDER),
                            ("TOPPADDING",    (0, 0), (-1, -1), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                            ("BACKGROUND",    (0, 0), (-1, -1), C_BG),
                        ]))
                        story.append(KeepTogether([img_table]))
                    except Exception:
                        nombre_boleta = boleta_path.name if boleta_path else boleta_label
                        story.append(Paragraph(f"Boleta: {nombre_boleta}", st_value))

            doc.build(story)

            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            odt_safe = re.sub(r"[^A-Za-z0-9_-]+", "", str(rend.odt or "").upper()) or "X"
            nombre_pdf = f"informe_{rend.id}_{ts}_{odt_safe}.pdf"
            pdf_bytes = buf.getvalue()

            try:
                drive_result = upload_rendicion_informe_to_drive(
                    tecnico=str(rend.tecnico or "") or "Sin tecnico",
                    rendicion_id=rend.id,
                    pdf_bytes=pdf_bytes,
                    filename=nombre_pdf,
                )
                LOGGER.info("Informe de rendicion subido a Drive: id=%s -> %s", rend.id, drive_result["public_uri"])
                return drive_result["public_uri"]
            except Exception:
                LOGGER.exception("No se pudo subir informe de rendicion a Drive (id=%s)", rend.id)
                return _guardar_en_cuarentena_drive(
                    "rendicion_informe",
                    str(rend.id),
                    nombre_pdf,
                    pdf_bytes,
                    {
                        "tecnico": str(rend.tecnico or ""),
                        "rendicion_id": rend.id,
                        "filename": nombre_pdf,
                    },
                )

        except Exception:
            logging.exception("Error generando PDF rendicion id=%s", getattr(rend, "id", "?"))
            return None

    def _resolver_mail_tecnico(self, tecnico: str, mail_actual: str = "") -> str:
        mail_limpio = str(mail_actual or "").strip()
        if mail_limpio:
            return mail_limpio
        tecnico_limpio = str(tecnico or "").strip()
        if not tecnico_limpio:
            return ""
        tecnico_norm = self._normalizar_texto(tecnico_limpio)
        try:
            users = self.db.scalars(select(User).where(User.email.is_not(None))).all()
            for user in users:
                if self._normalizar_texto(getattr(user, "name", "") or "") == tecnico_norm:
                    return str(getattr(user, "email", "") or "").strip()
        except Exception:
            logging.exception("No se pudo resolver correo de tecnico para rendicion: %s", tecnico_limpio)
            self.db.rollback()
        return ""

    def _resolver_rut_tecnico(self, tecnico: str) -> str:
        tecnico_limpio = str(tecnico or "").strip()
        if not tecnico_limpio:
            return ""
        tecnico_norm = self._normalizar_texto(tecnico_limpio)
        try:
            filas = self.db.scalars(
                select(EstatusDocumentacionTecnico).where(EstatusDocumentacionTecnico.rut.is_not(None))
            ).all()
            for fila in filas:
                if self._normalizar_texto(getattr(fila, "nombre", "") or "") == tecnico_norm:
                    return str(getattr(fila, "rut", "") or "").strip()
        except Exception:
            logging.exception("No se pudo resolver RUT de tecnico para rendicion: %s", tecnico_limpio)
            self.db.rollback()
        return ""

    def registrar_gasto(self, data: RendicionRequest, mail_tecnico: str = "") -> dict[str, Any]:
        if self.existe_nro_documento_duplicado(data.nro_documento):
            raise ValueError(f"El Nro de Documento {data.nro_documento} ya fue registrado.")

        cliente = data.cliente.strip()
        comuna = self._resolver_comuna_rendicion(cliente)

        dia_excel = int((data.fecha_documento.date() - datetime(1899, 12, 30).date()).days)
        iniciales = "".join([w[:1].upper() for w in data.tecnico.split() if w])
        tipo_id = re.sub(r"\s+", "", data.tipo_gasto).upper()
        codigo_diario = f"{dia_excel}-{iniciales}-{tipo_id}"

        rend = Rendicion(
            codigo_diario=codigo_diario,
            tecnico=data.tecnico,
            mail=self._resolver_mail_tecnico(data.tecnico, mail_tecnico),
            odt=data.odt.upper().strip(),
            cliente=cliente,
            comuna=comuna or None,
            tipo_gasto=data.tipo_gasto,
            tipo_documento=data.tipo_documento,
            nro_documento=data.nro_documento.strip(),
            fecha_documento=data.fecha_documento,
            monto_total=data.monto_total,
            descripcion=data.descripcion or "",
            documento=f"{data.tipo_documento} {data.nro_documento}",
            url_boleta=str(data.url_boleta or "").strip() or None,
        )
        self.db.add(rend)
        self.db.commit()
        self.db.refresh(rend)

        url_informe = self._generar_informe_rendicion_pdf(rend)
        if url_informe:
            rend.url_informe = url_informe
        self.db.commit()

        return {"id": rend.id, "codigoDiario": codigo_diario}

    _COMUNAS_RENDICIONES = sorted([
        "Arica","Iquique","Calama","Antofagasta","Copiapó","La Serena","Coquimbo",
        "Ovalle","Los Andes","Llay Llay","Hijuelas","La Calera","Quillota",
        "Quintero","Concón","Viña del Mar","Valparaíso","Quilpué","Villa Alemana",
        "Limache","Olmué","Casablanca","Lampa","Colina","Huechuraba","Quilicura",
        "Pudahuel","Lo Barnechea","Vitacura","Las Condes","Providencia","La Reina",
        "Ñuñoa","Santiago","Macul","La Florida","Maipú","Cerrillos","Pedro Aguirre Cerda",
        "San Miguel","San Joaquín","Lo Espejo","El Bosque","La Pintana","San Ramón",
        "Lo Prado","Estación Central","Cerro Navia","Renca","Conchalí","Independencia",
        "Recoleta","Cartagena","San Antonio","San Bernardo","Buin","Paine","Melipilla",
        "Rancagua","Rengo","San Fernando","Curicó","Talca","Linares","Chillán",
        "Concepción","Talcahuano","Coronel","Los Ángeles","Temuco","Valdivia",
        "Osorno","Puerto Montt","Peñalolén","Puente Alto","La Cisterna","Quinta Normal",
        "Reñaca","Tiltil","Padre Hurtado","Laja","Penco","Pucón","Villarrica",
        "La Ligua","Nogales","El Melón","Petorca","Cabildo","Los Vilos",
        "San Felipe","Llaillay","Panquehue","Catemu","Putaendo","Santa María",
        "Coinco","Litueche","Navidad","Pichilemu","Requínoa","Peumo",
    ], key=len, reverse=True)

    def _normalizar_comuna_texto(self, valor: Any) -> str:
        texto = str(valor or "").strip().replace("-", " ")
        texto = unicodedata.normalize("NFD", texto.lower())
        texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
        texto = re.sub(r"[^a-z0-9 ]", "", texto)
        return " ".join(texto.split())

    def _extraer_comuna_desde_direccion(self, direccion: str) -> str:
        direccion_txt = str(direccion or "").strip()
        if not direccion_txt:
            return ""
        display_por_norm = {self._normalizar_comuna_texto(c): c for c in self._COMUNAS_RENDICIONES}
        if "," in direccion_txt:
            for parte in reversed([p.strip() for p in direccion_txt.split(",")]):
                parte_norm = self._normalizar_comuna_texto(parte)
                if parte_norm.startswith("region") or not parte_norm:
                    continue
                if parte_norm in display_por_norm:
                    return display_por_norm[parte_norm]
                if len(parte) < 40 and not re.search(r"\d{3}", parte):
                    return parte
        return ""

    def _resolver_comuna_rendicion(self, sucursal: str) -> str:
        sucursal_txt = str(sucursal or "").strip()
        if not sucursal_txt:
            return ""
        try:
            sucursal_norm = self._normalizar_texto(sucursal_txt)
            rows = self.db.execute(
                select(SucursalBBDD.nombre_sucursal, SucursalBBDD.direccion_sucursal, SucursalBBDD.comuna)
            ).all()
            for nombre, direccion, comuna in rows:
                if self._normalizar_texto(nombre) != sucursal_norm:
                    continue
                comuna_txt = str(comuna or "").strip()
                return comuna_txt or self._extraer_comuna_desde_direccion(str(direccion or ""))
        except Exception:
            logging.exception("No se pudo resolver comuna de rendicion para %s", sucursal_txt)
        return ""

    def obtener_rendiciones(
        self,
        tecnico: str = "",
        pendientes_only: bool = False,
    ) -> list[dict[str, Any]]:
        rows = self.db.scalars(select(Rendicion).order_by(Rendicion.id.desc())).all()
        out: list[dict[str, Any]] = []
        tecnico_norm = self._normalizar_texto(tecnico or "")
        comuna_cache: dict[str, str] = {}
        for r in rows:
            tecnico_row = str(r.tecnico or "").strip()
            if tecnico_norm and self._normalizar_texto(tecnico_row) != tecnico_norm:
                continue

            estado_revision = str(r.estado_revision or "").strip() or "Pendiente"
            estado_norm = self._normalizar_texto(estado_revision)
            es_pendiente = not any(x in estado_norm for x in ("acept", "aprob", "rechaz", "pag"))
            if pendientes_only and not es_pendiente:
                continue

            cliente_row = str(r.cliente or "").strip()
            comuna = str(r.comuna or "").strip()
            if not comuna and cliente_row:
                cache_key = self._normalizar_texto(cliente_row)
                if cache_key not in comuna_cache:
                    comuna_cache[cache_key] = self._resolver_comuna_rendicion(cliente_row)
                comuna = comuna_cache[cache_key]

            out.append(
                {
                    "id": r.id,
                    "codigoDiario": r.codigo_diario,
                    "fechaRegistro": _to_ddmmyyyy_hhmm(r.fecha_registro),
                    "tecnico": tecnico_row,
                    "rut": self._resolver_rut_tecnico(tecnico_row),
                    "mail": self._resolver_mail_tecnico(tecnico_row, r.mail),
                    "odt": r.odt,
                    "cliente": r.cliente,
                    "comuna": comuna,
                    "tipoGasto": r.tipo_gasto,
                    "tipoDocumento": r.tipo_documento,
                    "nroDocumento": r.nro_documento,
                    "documento": r.documento,
                    "fechaDocumento": _to_ddmmyyyy(r.fecha_documento),
                    "montoTotal": float(r.monto_total),
                    "descripcion": r.descripcion,
                    "urlBoleta": r.url_boleta,
                    "urlInforme": r.url_informe,
                    "estadoRevision": estado_revision,
                }
            )
        return out

    _TIPOS_GASTO_VL = {"materiales", "combustible"}

    def agrupar_pagos_pendientes(self) -> dict[str, list[dict[str, Any]]]:
        """Pagos pendientes ('Por pagar'), separados en 'atc' y 'vl' segun
        tipo de gasto (Materiales/Combustible van a VL).

        Los codigos diario repetidos se funden en una sola fila sumando el
        monto, para no duplicar pagos hacia la misma persona/gasto. Tanto en
        Pagos ATC como en Pagos VL se exporta el menor valor entre el total
        rendido y el tope: $5.000 por defecto o el viatico especial configurado
        para el codigo. Cada lista queda ordenada por RUT para que los codigos
        de una misma persona queden agrupados uno debajo del otro.
        """
        rows = self.db.scalars(select(Rendicion).order_by(Rendicion.id.asc())).all()
        caps_personalizados = self._viatico_caps_personalizados()
        caps_normalizados = {
            self._normalizar_codigo_diario(codigo): monto
            for codigo, monto in caps_personalizados.items()
        }

        grupos_atc: dict[str, dict[str, Any]] = {}
        grupos_vl: dict[str, dict[str, Any]] = {}
        for r in rows:
            estado_norm = self._normalizar_texto(str(r.estado_revision or ""))
            if "acept" not in estado_norm:
                continue

            codigo = str(r.codigo_diario or "").strip()
            if not codigo:
                continue

            tecnico_row = str(r.tecnico or "").strip()
            tipo_gasto_norm = self._normalizar_texto(str(r.tipo_gasto or ""))
            destino = grupos_vl if tipo_gasto_norm in self._TIPOS_GASTO_VL else grupos_atc
            monto = Decimal(str(r.monto_total or 0))

            grupo = destino.get(codigo)
            if grupo is None:
                grupo = {
                    "rut": self._resolver_rut_tecnico(tecnico_row),
                    "tecnico": tecnico_row,
                    "codigo": codigo,
                    "monto": Decimal("0"),
                    "ids": [],
                }
                destino[codigo] = grupo

            grupo["monto"] += monto
            grupo["ids"].append(r.id)

        def _ordenar(grupos: dict[str, dict[str, Any]], aplicar_tope: bool) -> list[dict[str, Any]]:
            filas: list[dict[str, Any]] = []
            for grupo in grupos.values():
                if aplicar_tope:
                    codigo_norm = self._normalizar_codigo_diario(grupo["codigo"])
                    tope = caps_normalizados.get(codigo_norm, self.VIATICO_CAP_DIARIO_DEFECTO)
                    monto_exportable = min(grupo["monto"], tope)
                else:
                    monto_exportable = grupo["monto"]
                filas.append({
                    "rut": grupo["rut"],
                    "tecnico": grupo["tecnico"],
                    "codigo": grupo["codigo"],
                    "monto": float(monto_exportable),
                    "ids": grupo["ids"],
                })
            return sorted(filas, key=lambda g: (str(g["rut"] or "").strip() or "~", g["codigo"]))

        # Pagos VL (Materiales/Combustible) no llevan tope de viatico diario:
        # ese tope es un concepto de viaticos (ATC), no aplica a reembolsos
        # de materiales/combustible — pedido explicito, jul 2026.
        return {"atc": _ordenar(grupos_atc, aplicar_tope=True), "vl": _ordenar(grupos_vl, aplicar_tope=False)}

    def generar_excel_pagos_rendiciones(self) -> "BytesIO":
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        pagos = self.agrupar_pagos_pendientes()

        def _crear_hoja(wb, titulo: str, filas: list[dict[str, Any]], primera: bool = False):
            ws = wb.active if primera else wb.create_sheet(titulo)
            ws.title = titulo
            ws.append(["RUT", "Codigo diario", "Monto a pagar"])
            for celda in ws[1]:
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")
            for fila_dato in filas:
                ws.append([fila_dato["rut"], fila_dato["codigo"], fila_dato["monto"]])
            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 26
            ws.column_dimensions["C"].width = 18
            for fila in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for celda in fila:
                    celda.number_format = "#,##0"

        wb = Workbook()
        _crear_hoja(wb, "Pagos ATC", pagos["atc"], primera=True)
        _crear_hoja(wb, "Pagos VL", pagos["vl"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _formatear_clp(self, monto) -> str:
        try:
            valor = int(round(float(monto or 0)))
        except (TypeError, ValueError):
            valor = 0
        texto = f"{valor:,}".replace(",", ".")
        return f"${texto}"

    def _construir_email_rendicion(self, rend: "Rendicion", aprobada: bool) -> tuple[str, str, str]:
        """Arma asunto/texto/html del correo de aprobacion o rechazo de una
        rendicion — mismo estilo (header azul marino + logo) que el correo
        de pruebas de sonido en routes/incidencias.py."""
        estado_label = "APROBADA" if aprobada else "RECHAZADA"
        color_badge_bg = "#f0fdf4" if aprobada else "#fef2f2"
        color_badge_border = "#a7f3d0" if aprobada else "#fecaca"
        color_badge_text = "#15803d" if aprobada else "#b91c1c"
        verbo = "aprobada" if aprobada else "rechazada"

        tecnico = str(rend.tecnico or "").strip() or "Tecnico"
        monto_fmt = self._formatear_clp(rend.monto_total)
        fecha_doc = rend.fecha_documento.strftime("%d-%m-%Y") if rend.fecha_documento else "-"

        asunto = f"Rendición {estado_label} - Folio {rend.id}"

        detalle_filas = [
            ("Código", rend.codigo_diario or "-"),
            ("ODT", rend.odt or "-"),
            ("Cliente", rend.cliente or "-"),
            ("Tipo de gasto", rend.tipo_gasto or "-"),
            ("N° documento", rend.nro_documento or "-"),
            ("Fecha documento", fecha_doc),
            ("Monto", monto_fmt),
        ]

        texto_lineas = [f"Hola {tecnico},", "", f"Tu rendición ha sido {verbo}.", ""]
        texto_lineas += [f"{label}: {valor}" for label, valor in detalle_filas]
        texto_lineas += ["", "Equipo de Rendiciones", "Alguien Te Cuida SpA"]
        cuerpo_txt = "\n".join(texto_lineas)

        filas_html = "".join(
            f'''
              <tr>
                <td style="padding:7px 0;font-family:Arial,sans-serif;font-size:13px;color:#6b7280;border-bottom:1px solid #f1f5f9;width:150px;">{label}</td>
                <td style="padding:7px 0;font-family:Arial,sans-serif;font-size:13px;color:#111827;font-weight:600;border-bottom:1px solid #f1f5f9;">{valor}</td>
              </tr>'''
            for label, valor in detalle_filas
        )

        nota_rechazo = ""
        if not aprobada:
            nota_rechazo = '''
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;line-height:1.65;color:#374151;">
              Si tienes dudas sobre el motivo del rechazo, contacta a tu jefatura.
            </p>'''

        cuerpo_html = f"""<!DOCTYPE html>
<html lang="es" xmlns="http://www.w3.org/1999/xhtml" style="color-scheme:light;">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <meta name="color-scheme" content="light">
  <title>{asunto}</title>
</head>
<body style="margin:0;padding:0;background-color:#f2f4f7;-webkit-text-size-adjust:100%;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
       style="background-color:#f2f4f7;min-width:320px;">
  <tr>
    <td align="center" style="padding:40px 16px;">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0"
             style="max-width:600px;width:100%;background-color:#ffffff;border-radius:6px;
                    overflow:hidden;border:1px solid #d1d5db;">

        <tr>
          <td style="background-color:#0d1f2d;padding:20px 36px;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="vertical-align:middle;">
                  <img src="cid:logoatc" alt="Alguien Te Cuida"
                       style="height:34px;width:auto;display:block;border:0;" />
                </td>
                <td align="right" style="vertical-align:middle;">
                  <span style="font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                               color:#8aabb8;letter-spacing:0.12em;text-transform:uppercase;">
                    Rendición de Gastos
                  </span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="background-color:#1e3a5f;padding:20px 36px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                      color:#93b4cc;letter-spacing:0.12em;text-transform:uppercase;">
              Folio {rend.id}
            </p>
            <p style="margin:7px 0 0;font-family:Arial,sans-serif;font-size:20px;font-weight:700;
                      color:#ffffff;letter-spacing:-0.01em;line-height:1.25;">
              Rendición {estado_label.capitalize()}
            </p>
            <p style="margin:5px 0 0;font-family:Arial,sans-serif;font-size:13px;
                      color:#a8c4d8;line-height:1.4;">
              {tecnico}
            </p>
          </td>
        </tr>

        <tr>
          <td style="background-color:#ffffff;padding:26px 36px 4px;">
            <table role="presentation" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="background-color:{color_badge_bg};border:1px solid {color_badge_border};border-radius:5px;
                           padding:9px 16px;">
                  <span style="font-family:Arial,sans-serif;font-size:12px;font-weight:700;
                               color:{color_badge_text};letter-spacing:0.02em;">{estado_label}</span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:20px 36px 8px;background-color:#ffffff;">
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">Hola <strong>{tecnico}</strong>,</p>
            <p style="margin:0 0 20px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Tu rendición de gastos ha sido <strong>{verbo}</strong>. Te dejamos el detalle a continuación:
            </p>
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:8px;">
              {filas_html}
            </table>
            {nota_rechazo}
          </td>
        </tr>

        <tr>
          <td style="padding:16px 36px 18px;background-color:#ffffff;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
              <tr><td style="border-top:1px solid #e5e7eb;font-size:0;line-height:0;">&nbsp;</td></tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:0 36px 26px;background-color:#ffffff;">
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:13px;
                      font-weight:700;color:#111827;">Equipo de Rendiciones</p>
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">Alguien Te Cuida SpA</p>
            <p style="margin:0;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">contacto@alguientecuida.cl</p>
          </td>
        </tr>

        <tr>
          <td style="background-color:#f8fafc;border-top:1px solid #e5e7eb;
                     padding:14px 36px;border-radius:0 0 6px 6px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:11px;color:#9ca3af;line-height:1.5;">
              Este mensaje fue generado automáticamente por el sistema de Alguien Te Cuida SpA.
              Por favor no responda directamente a este correo.
            </p>
          </td>
        </tr>

      </table>
    </td>
  </tr>
</table>
</body>
</html>"""
        return asunto, cuerpo_txt, cuerpo_html

    def _enviar_correo_rendicion(self, rend: "Rendicion", aprobada: bool) -> None:
        dest = str(rend.mail or "").strip()
        if not dest:
            LOGGER.warning("Rendición id=%s sin correo asociado, no se notifica.", rend.id)
            return
        asunto, cuerpo_txt, cuerpo_html = self._construir_email_rendicion(rend, aprobada)

        logo_path = Path(__file__).resolve().parents[2] / "static" / "img" / "logo-atc.png"
        logo_bytes = logo_path.read_bytes() if logo_path.exists() else None

        def _enviar(dest=dest, subj=asunto, txt=cuerpo_txt, html=cuerpo_html, logo=logo_bytes, rid=rend.id):
            try:
                svc_mail = IncidenciasService(SessionLocal())
                svc_mail._enviar_correo_automatico(
                    dest, subj, txt, html_body=html, logo_bytes=logo,
                    cfg_override=svc_mail._contacto_smtp_runtime_config(),
                )
            except Exception:
                LOGGER.exception("Error enviando email de rendición id=%s", rid)

        threading.Thread(target=_enviar, daemon=True, name=f"email-rendicion-{rend.id}").start()

    def _construir_email_reforzar_inicio_odt(self, *, nombre: str, odt: str, verbo_accion: str) -> tuple[str, str, str]:
        """Correo de refuerzo cuando un técnico o acompañante finaliza/deja pendiente
        una ODT sin haber presionado 'Iniciar' antes — mismo estilo (header azul
        marino + logo) que el correo de rendiciones/pruebas de sonido.
        verbo_accion: 'finalizada' o 'dejada en pendiente'."""
        nombre_mostrar = nombre or "Técnico"
        asunto = f"Recordatorio obligatorio: Inicia la ODT antes de continuar — ODT {odt}"

        cuerpo_txt = (
            f"Hola {nombre_mostrar},\n\n"
            f"Notamos que la ODT {odt} fue {verbo_accion} sin haber sido marcada como iniciada "
            f"previamente en el sistema.\n\n"
            f"Te recordamos que presionar 'Iniciar' en la ODT es un paso estrictamente necesario "
            f"antes de finalizarla o dejarla pendiente, ya que permite registrar correctamente los "
            f"tiempos de trabajo y el seguimiento de la atención.\n\n"
            f"Te pedimos que, a partir de ahora, recuerdes iniciar cada ODT antes de comenzar tu "
            f"labor en terreno.\n\n"
            f"Equipo Técnico\nAlguien Te Cuida SpA"
        )

        cuerpo_html = f"""<!DOCTYPE html>
<html lang="es" xmlns="http://www.w3.org/1999/xhtml" style="color-scheme:light;">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <meta name="color-scheme" content="light">
  <title>{asunto}</title>
</head>
<body style="margin:0;padding:0;background-color:#f2f4f7;-webkit-text-size-adjust:100%;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
       style="background-color:#f2f4f7;min-width:320px;">
  <tr>
    <td align="center" style="padding:40px 16px;">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0"
             style="max-width:600px;width:100%;background-color:#ffffff;border-radius:6px;
                    overflow:hidden;border:1px solid #d1d5db;">

        <tr>
          <td style="background-color:#0d1f2d;padding:20px 36px;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="vertical-align:middle;">
                  <img src="cid:logoatc" alt="Alguien Te Cuida"
                       style="height:34px;width:auto;display:block;border:0;" />
                </td>
                <td align="right" style="vertical-align:middle;">
                  <span style="font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                               color:#8aabb8;letter-spacing:0.12em;text-transform:uppercase;">
                    Recordatorio Operativo
                  </span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="background-color:#1e3a5f;padding:20px 36px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                      color:#93b4cc;letter-spacing:0.12em;text-transform:uppercase;">
              ODT {odt}
            </p>
            <p style="margin:7px 0 0;font-family:Arial,sans-serif;font-size:20px;font-weight:700;
                      color:#ffffff;letter-spacing:-0.01em;line-height:1.25;">
              Inicio de ODT Obligatorio
            </p>
            <p style="margin:5px 0 0;font-family:Arial,sans-serif;font-size:13px;
                      color:#a8c4d8;line-height:1.4;">
              {nombre_mostrar}
            </p>
          </td>
        </tr>

        <tr>
          <td style="background-color:#ffffff;padding:26px 36px 4px;">
            <table role="presentation" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="background-color:#fffbeb;border:1px solid #fcd34d;border-radius:5px;
                           padding:9px 16px;">
                  <span style="font-family:Arial,sans-serif;font-size:12px;font-weight:700;
                               color:#b45309;letter-spacing:0.02em;">RECORDATORIO</span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:20px 36px 8px;background-color:#ffffff;">
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">Hola <strong>{nombre_mostrar}</strong>,</p>
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Notamos que la ODT <strong>{odt}</strong> fue {verbo_accion} sin haber sido marcada como
              iniciada previamente en el sistema.
            </p>
            <p style="margin:0 0 13px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Te recordamos que presionar &quot;Iniciar&quot; en la ODT es un paso
              <strong>estrictamente necesario</strong> antes de finalizarla o dejarla pendiente, ya que
              permite registrar correctamente los tiempos de trabajo y el seguimiento de la atención.
            </p>
            <p style="margin:0 0 20px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              Te pedimos que, a partir de ahora, recuerdes iniciar cada ODT antes de comenzar tu labor
              en terreno.
            </p>
          </td>
        </tr>

        <tr>
          <td style="padding:0 36px 18px;background-color:#ffffff;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
              <tr><td style="border-top:1px solid #e5e7eb;font-size:0;line-height:0;">&nbsp;</td></tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:0 36px 26px;background-color:#ffffff;">
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:13px;
                      font-weight:700;color:#111827;">Equipo Técnico</p>
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">Alguien Te Cuida SpA</p>
            <p style="margin:0;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">contacto@alguientecuida.cl</p>
          </td>
        </tr>

        <tr>
          <td style="background-color:#f8fafc;border-top:1px solid #e5e7eb;
                     padding:14px 36px;border-radius:0 0 6px 6px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:11px;color:#9ca3af;line-height:1.5;">
              Este mensaje fue generado automáticamente por el sistema de Alguien Te Cuida SpA.
              Por favor no responda directamente a este correo.
            </p>
          </td>
        </tr>

      </table>
    </td>
  </tr>
</table>
</body>
</html>"""
        return asunto, cuerpo_txt, cuerpo_html

    def _enviar_correo_reforzar_inicio_odt(self, email: str, nombre: str, odt: str, verbo_accion: str) -> None:
        if not email:
            return
        asunto, cuerpo_txt, cuerpo_html = self._construir_email_reforzar_inicio_odt(
            nombre=nombre, odt=odt, verbo_accion=verbo_accion
        )
        logo_path = Path(__file__).resolve().parents[2] / "static" / "img" / "logo-atc.png"
        logo_bytes = logo_path.read_bytes() if logo_path.exists() else None
        try:
            self._enviar_correo_automatico(
                email, asunto, cuerpo_txt, html_body=cuerpo_html, logo_bytes=logo_bytes,
                cfg_override=self._contacto_smtp_runtime_config(),
                cc_emails=["catalina.silva@soporteatc.cl"],
            )
        except Exception:
            LOGGER.exception("Error enviando email de refuerzo inicio ODT %s a %s", odt, email)

    def _reforzar_inicio_odt_si_corresponde(
        self,
        *,
        odt: str,
        tecnico: str | None,
        acompanante: str | None,
        usuario_accion: str | None = None,
        fecha_inicio_trabajo: Any,
        verbo_accion: str,
    ) -> None:
        """Si el técnico (o acompañante) finaliza o deja pendiente esta ODT sin haber
        presionado 'Iniciar' antes (fecha_inicio_trabajo vacía), les refuerza por
        correo que iniciar la ODT es estrictamente necesario. verbo_accion:
        'finalizada' o 'dejada en pendiente'. Se llama después del commit del
        cierre/pendiente — una falla acá no debe afectar esa operación.

        Recibe tecnico/acompanante como strings explícitos (no lee row.tecnicos
        directamente) porque en el flujo de venta esos campos se borran del
        Registro antes del commit para forzar la re-derivación del día
        siguiente (ver guardar_datos_en_proceso) — para esta notificación
        igual necesitamos saber a quién avisar. usuario_accion es quien
        realmente ejecutó el cierre/pendiente (row.tecnico_cierre) — puede
        diferir del técnico/acompañante asignado si otra persona lo hizo por
        ellos, y también debe recibir el aviso. Todos los envíos llevan copia
        fija a catalina.silva@soporteatc.cl."""
        try:
            if fecha_inicio_trabajo:
                return
            odt = str(odt or "").strip()
            if not odt:
                return
            nombres = {
                str(tecnico or "").strip(),
                str(acompanante or "").strip(),
                str(usuario_accion or "").strip(),
            }
            nombres.discard("")
            if not nombres:
                return
            for nombre in nombres:
                email = self._resolver_mail_tecnico(nombre, "")
                if not email:
                    continue

                def _bg(email=email, nombre=nombre, odt=odt, verbo_accion=verbo_accion):
                    _db = SessionLocal()
                    try:
                        svc_mail = IncidenciasService(_db)
                        svc_mail._enviar_correo_reforzar_inicio_odt(email, nombre, odt, verbo_accion)
                    except Exception:
                        LOGGER.exception("Error en hilo de refuerzo inicio ODT %s -> %s", odt, email)
                    finally:
                        _db.close()

                threading.Thread(target=_bg, daemon=True, name=f"email-inicio-odt-{odt}").start()
        except Exception:
            LOGGER.exception("Error evaluando refuerzo de inicio de ODT %s", odt)

    def _resolver_correo_comercial(self, creado_por: str) -> tuple[str, str]:
        """(email, nombre) del comercial a cargo de la ODS — VentaODS.creado_por
        normalmente ya guarda el email; si en cambio guarda un nombre (pasa
        en algunas filas antiguas), se resuelve contra User.name."""
        valor = str(creado_por or "").strip()
        if not valor:
            return "", ""
        if "@" in valor:
            usuario = self.db.scalar(select(User).where(func.lower(User.email) == valor.lower()))
            nombre = str(getattr(usuario, "name", "") or "").strip() or valor
            return valor, nombre
        email = self._resolver_mail_tecnico(valor, "")
        return email, valor

    def _construir_email_exceso_instalacion(
        self,
        *,
        odt: str,
        cliente: str,
        tecnico: str,
        acompanante: str,
        camaras_contratadas: int,
        camaras_instaladas: int,
        comercial_nombre: str,
    ) -> tuple[str, str, str]:
        """Correo informativo cuando una instalación (ODT de venta) cierra con
        más cámaras instaladas que las contratadas — mismo estilo (header
        azul marino + logo) que el correo de rendiciones."""
        diferencia = camaras_instaladas - camaras_contratadas
        asunto = f"Instalación ODT {odt} — {camaras_instaladas} cámaras instaladas (contratadas: {camaras_contratadas})"

        detalle_filas = [
            ("Código", odt or "-"),
            ("Cliente / Sucursal", cliente or "-"),
            ("Técnico", tecnico or "-"),
            ("Acompañante", acompanante or "-"),
            ("Comercial a cargo", comercial_nombre or "-"),
            ("Cantidad contratada", str(camaras_contratadas)),
            ("Cantidad instalada", str(camaras_instaladas)),
            ("Diferencia", f"+{diferencia}"),
        ]

        texto_lineas = [
            f"La instalación de la ODT {odt} se cerró con {camaras_instaladas} cámaras instaladas,",
            f"{diferencia} más que las {camaras_contratadas} contratadas originalmente.",
            "",
        ]
        texto_lineas += [f"{label}: {valor}" for label, valor in detalle_filas]
        texto_lineas += ["", "Equipo Servicio Técnico", "Alguien Te Cuida SpA"]
        cuerpo_txt = "\n".join(texto_lineas)

        filas_html = "".join(
            f'''
              <tr>
                <td style="padding:7px 0;font-family:Arial,sans-serif;font-size:13px;color:#6b7280;border-bottom:1px solid #f1f5f9;width:170px;">{label}</td>
                <td style="padding:7px 0;font-family:Arial,sans-serif;font-size:13px;color:#111827;font-weight:600;border-bottom:1px solid #f1f5f9;">{valor}</td>
              </tr>'''
            for label, valor in detalle_filas
        )

        cuerpo_html = f"""<!DOCTYPE html>
<html lang="es" xmlns="http://www.w3.org/1999/xhtml" style="color-scheme:light;">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <meta name="color-scheme" content="light">
  <title>{asunto}</title>
</head>
<body style="margin:0;padding:0;background-color:#f2f4f7;-webkit-text-size-adjust:100%;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0"
       style="background-color:#f2f4f7;min-width:320px;">
  <tr>
    <td align="center" style="padding:40px 16px;">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0"
             style="max-width:600px;width:100%;background-color:#ffffff;border-radius:6px;
                    overflow:hidden;border:1px solid #d1d5db;">

        <tr>
          <td style="background-color:#0d1f2d;padding:20px 36px;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="vertical-align:middle;">
                  <img src="cid:logoatc" alt="Alguien Te Cuida"
                       style="height:34px;width:auto;display:block;border:0;" />
                </td>
                <td align="right" style="vertical-align:middle;">
                  <span style="font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                               color:#8aabb8;letter-spacing:0.12em;text-transform:uppercase;">
                    Servicio Técnico
                  </span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="background-color:#1e3a5f;padding:20px 36px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:10px;font-weight:600;
                      color:#93b4cc;letter-spacing:0.12em;text-transform:uppercase;">
              ODT {odt}
            </p>
            <p style="margin:7px 0 0;font-family:Arial,sans-serif;font-size:20px;font-weight:700;
                      color:#ffffff;letter-spacing:-0.01em;line-height:1.25;">
              Instalación cerrada con cámaras adicionales
            </p>
            <p style="margin:5px 0 0;font-family:Arial,sans-serif;font-size:13px;
                      color:#a8c4d8;line-height:1.4;">
              {cliente}
            </p>
          </td>
        </tr>

        <tr>
          <td style="background-color:#ffffff;padding:26px 36px 4px;">
            <table role="presentation" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="background-color:#fff7ed;border:1px solid #fed7aa;border-radius:5px;
                           padding:9px 16px;">
                  <span style="font-family:Arial,sans-serif;font-size:12px;font-weight:700;
                               color:#c2410c;letter-spacing:0.02em;">{camaras_instaladas} INSTALADAS · {camaras_contratadas} CONTRATADAS</span>
                </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:20px 36px 8px;background-color:#ffffff;">
            <p style="margin:0 0 20px;font-family:Arial,sans-serif;font-size:14px;
                      line-height:1.65;color:#374151;">
              La instalación de la ODT <strong>{odt}</strong> en <strong>{cliente}</strong> se cerró con
              <strong>{camaras_instaladas} cámaras instaladas</strong>, {diferencia} más que las
              <strong>{camaras_contratadas}</strong> contratadas originalmente. Te dejamos el detalle a continuación.
            </p>
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:8px;">
              {filas_html}
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:16px 36px 18px;background-color:#ffffff;">
            <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
              <tr><td style="border-top:1px solid #e5e7eb;font-size:0;line-height:0;">&nbsp;</td></tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:0 36px 26px;background-color:#ffffff;">
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:13px;
                      font-weight:700;color:#111827;">Equipo Servicio Técnico</p>
            <p style="margin:0 0 1px;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">Alguien Te Cuida SpA</p>
            <p style="margin:0;font-family:Arial,sans-serif;font-size:12px;color:#6b7280;">contacto@alguientecuida.cl</p>
          </td>
        </tr>

        <tr>
          <td style="background-color:#f8fafc;border-top:1px solid #e5e7eb;
                     padding:14px 36px;border-radius:0 0 6px 6px;">
            <p style="margin:0;font-family:Arial,sans-serif;font-size:11px;color:#9ca3af;line-height:1.5;">
              Este mensaje fue generado automáticamente por el sistema de Alguien Te Cuida SpA.
              Por favor no responda directamente a este correo.
            </p>
          </td>
        </tr>

      </table>
    </td>
  </tr>
</table>
</body>
</html>"""
        return asunto, cuerpo_txt, cuerpo_html

    def _enviar_correo_exceso_instalacion(
        self,
        *,
        odt: str,
        cliente: str,
        tecnico: str,
        acompanante: str,
        camaras_contratadas: int,
        camaras_instaladas: int,
        creado_por: str,
    ) -> None:
        CZAMORA_EMAIL = "czamora@alguientecuida.cl"
        comercial_email, comercial_nombre = self._resolver_correo_comercial(creado_por)
        dest = comercial_email or CZAMORA_EMAIL
        cc = [CZAMORA_EMAIL] if dest.lower() != CZAMORA_EMAIL.lower() else None

        asunto, cuerpo_txt, cuerpo_html = self._construir_email_exceso_instalacion(
            odt=odt,
            cliente=cliente,
            tecnico=tecnico,
            acompanante=acompanante,
            camaras_contratadas=camaras_contratadas,
            camaras_instaladas=camaras_instaladas,
            comercial_nombre=comercial_nombre,
        )

        logo_path = Path(__file__).resolve().parents[2] / "static" / "img" / "logo-atc.png"
        logo_bytes = logo_path.read_bytes() if logo_path.exists() else None

        def _enviar(dest=dest, cc=cc, subj=asunto, txt=cuerpo_txt, html=cuerpo_html, logo=logo_bytes, odt_id=odt):
            try:
                svc_mail = IncidenciasService(SessionLocal())
                svc_mail._enviar_correo_automatico(
                    dest, subj, txt, html_body=html, logo_bytes=logo, cc_emails=cc,
                    cfg_override=svc_mail._contacto_smtp_runtime_config(),
                )
            except Exception:
                LOGGER.exception("Error enviando email de exceso de instalación ODT=%s", odt_id)

        threading.Thread(target=_enviar, daemon=True, name=f"email-exceso-instalacion-{odt}").start()

    def actualizar_monto_rendicion(self, rendicion_id: int, monto: float) -> dict[str, Any] | None:
        rend = self.db.scalar(select(Rendicion).where(Rendicion.id == rendicion_id))
        if not rend:
            return None
        if monto < 0:
            raise ValueError("El monto no puede ser negativo.")
        rend.monto_total = Decimal(str(monto))
        self.db.commit()
        return {"id": rend.id, "montoTotal": rend.monto_total}

    def marcar_rendicion(self, rendicion_id: int, accion: str, usuario: str = "") -> bool:
        rend = self.db.scalar(select(Rendicion).where(Rendicion.id == rendicion_id))
        if not rend:
            return False
        if accion == "aceptar":
            rend.estado_revision = "Aceptado"
        elif accion == "rechazar":
            rend.estado_revision = "Rechazado"
        elif accion == "pagar":
            estado_actual = str(rend.estado_revision or "").strip().lower()
            if estado_actual != "pagado":
                tecnico_row = str(rend.tecnico or "").strip()
                pago = RendicionPago(
                    codigo_diario=str(rend.codigo_diario or "").strip(),
                    tecnico=tecnico_row,
                    rut_tecnico=self._resolver_rut_tecnico(tecnico_row) or None,
                    tipo_pago="Transferencia",
                    fecha_pago=datetime.now(),
                    monto=Decimal(str(rend.monto_total or 0)),
                    creado_por=(usuario or None),
                )
                self.db.add(pago)
            rend.estado_revision = "Pagado"
        else:
            raise ValueError("Acción inválida. Debe ser 'aceptar', 'rechazar' o 'pagar'.")
        self.db.commit()
        if accion in ("aceptar", "rechazar"):
            self._enviar_correo_rendicion(rend, aprobada=(accion == "aceptar"))
        return True

    # =========================
    # FINANZAS RENDICIONES (Consolidado, Viatico, Pagos)
    # =========================
    VIATICO_CAP_DIARIO_DEFECTO = Decimal("5000")
    VIATICO_CAP_ESPECIAL_DEFECTO = Decimal("10000")
    PAGO_VL_TIPOS = {"materiales", "compras grandes"}

    @staticmethod
    def _normalizar_codigo_diario(valor: Any) -> str:
        texto = str(valor or "")
        for ch in ("\u00a0", "\u2007", "\u202f", "\u200b", "\u200c", "\u200d", "\ufeff"):
            texto = texto.replace(ch, " ")
        for ch in ("–", "—", "−", "‑", "‒"):
            texto = texto.replace(ch, "-")
        texto = " ".join(texto.split())
        texto = texto.replace(" - ", "-").replace("- ", "-").replace(" -", "-")
        return texto.strip().lower()

    def _tope_diario(self, tipo_gasto: str) -> Decimal:
        if str(tipo_gasto or "").strip().lower() == "viatico":
            return self.VIATICO_CAP_DIARIO_DEFECTO
        return Decimal("0")

    def _suma_pagos_por_codigo(self) -> dict[str, Decimal]:
        rows = self.db.execute(
            select(RendicionPago.codigo_diario, func.coalesce(func.sum(RendicionPago.monto), 0))
            .group_by(RendicionPago.codigo_diario)
        ).all()
        return {str(c or "").strip(): Decimal(str(s or 0)) for c, s in rows}

    def _viatico_caps_personalizados(self) -> dict[str, Decimal]:
        rows = self.db.scalars(select(RendicionViaticoCap)).all()
        out: dict[str, Decimal] = {}
        dirty_zero_codes: list[str] = []
        for r in rows:
            monto = Decimal(str(r.viatico_max or 0))
            if monto <= 0:
                dirty_zero_codes.append(str(r.codigo_diario or "").strip())
                continue
            out[str(r.codigo_diario or "").strip()] = monto
        if dirty_zero_codes:
            self.db.query(RendicionViaticoCap).filter(RendicionViaticoCap.codigo_diario.in_(dirty_zero_codes)).delete(synchronize_session=False)
            self.db.commit()
        return out

    def obtener_consolidado(self) -> list[dict[str, Any]]:
        rows = self.db.scalars(select(Rendicion).order_by(Rendicion.codigo_diario)).all()
        sumas_pagadas = self._suma_pagos_por_codigo()
        grupos: dict[str, dict[str, Any]] = {}
        for r in rows:
            codigo = str(r.codigo_diario or "").strip()
            if not codigo:
                continue
            estado = str(r.estado_revision or "").strip().lower()
            if estado.startswith("rechaz"):
                continue
            g = grupos.setdefault(codigo, {
                "codigo": codigo,
                "tecnico": str(r.tecnico or "").strip(),
                "tipo_gasto": str(r.tipo_gasto or "").strip(),
                "fecha": r.fecha_documento,
                "suma": Decimal("0"),
                "suma_pagado_estado": Decimal("0"),
                "todas_pagadas": True,
                "alguna_pendiente": False,
            })
            if estado.startswith("acept") or estado == "pagado":
                g["suma"] += Decimal(str(r.monto_total or 0))
            if estado == "pagado":
                g["suma_pagado_estado"] += Decimal(str(r.monto_total or 0))
            if estado != "pagado":
                g["todas_pagadas"] = False
            if estado.startswith("acept"):
                g["alguna_pendiente"] = True
            if not g["fecha"] or (r.fecha_documento and r.fecha_documento < g["fecha"]):
                g["fecha"] = r.fecha_documento

        out: list[dict[str, Any]] = []
        for codigo, g in grupos.items():
            tope = self._tope_diario(g["tipo_gasto"])
            suma = g["suma"]
            pagar = min(suma, tope) if tope > 0 else suma
            pagado_registrado = sumas_pagadas.get(codigo, Decimal("0"))
            pagado_bruto = max(pagado_registrado, g["suma_pagado_estado"])
            pagado = min(pagado_bruto, pagar) if pagar > 0 else pagado_bruto
            diferencia = max(Decimal("0"), pagar - pagado)
            if pagar > 0 and pagado >= pagar:
                estado_pago = "PAGADO"
            elif pagado > 0 and diferencia > 0:
                estado_pago = "PARCIAL"
            elif g["alguna_pendiente"]:
                estado_pago = "PENDIENTE"
            else:
                estado_pago = "SIN ACEPTAR"
            out.append({
                "codigo": codigo,
                "tecnico": g["tecnico"],
                "tipoGasto": g["tipo_gasto"],
                "fecha": _to_ddmmyyyy(g["fecha"]) if g["fecha"] else "",
                "suma": float(suma),
                "tope": float(tope),
                "aPagar": float(pagar),
                "pagado": float(pagado),
                "diferencia": float(diferencia),
                "estadoPago": estado_pago,
            })
        out.sort(key=lambda x: x["codigo"])
        return out

    def obtener_viatico_especial(self, codigo: str = "", personalizados: bool = False) -> list[dict[str, Any]]:
        codigo_buscado = self._normalizar_codigo_diario(codigo)
        query = (
            select(Rendicion)
            .where(func.lower(Rendicion.tipo_gasto) == "viatico")
            .order_by(Rendicion.codigo_diario)
        )
        rows = self.db.scalars(query).all()
        caps = self._viatico_caps_personalizados()
        caps_normalizados = {self._normalizar_codigo_diario(k): v for k, v in caps.items()}

        if not codigo_buscado and not personalizados:
            return []

        grupos: dict[str, dict[str, Any]] = {}
        for r in rows:
            codigo = str(r.codigo_diario or "").strip()
            if not codigo:
                continue
            codigo_norm = self._normalizar_codigo_diario(codigo)
            if codigo_buscado and codigo_norm != codigo_buscado:
                continue
            if personalizados and not codigo_buscado and codigo_norm not in caps_normalizados:
                continue
            estado = str(r.estado_revision or "").strip().lower()
            if estado.startswith("rechaz"):
                continue
            g = grupos.setdefault(codigo, {
                "codigo": codigo,
                "tecnico": str(r.tecnico or "").strip(),
                "fecha": r.fecha_documento,
                "odts": set(),
                "clientes": set(),
                "comunas": set(),
                "real": Decimal("0"),
            })
            g["real"] += Decimal(str(r.monto_total or 0))
            if r.odt:
                g["odts"].add(str(r.odt).strip())
            if r.cliente:
                g["clientes"].add(str(r.cliente).strip())
            if r.comuna:
                g["comunas"].add(str(r.comuna).strip())
            if not g["fecha"] or (r.fecha_documento and r.fecha_documento < g["fecha"]):
                g["fecha"] = r.fecha_documento

        out: list[dict[str, Any]] = []
        for codigo, g in grupos.items():
            codigo_norm = self._normalizar_codigo_diario(codigo)
            vmax = caps_normalizados.get(codigo_norm, self.VIATICO_CAP_ESPECIAL_DEFECTO)
            real = g["real"]
            a_pagar = min(real, vmax)
            out.append({
                "codigo": codigo,
                "tecnico": g["tecnico"],
                "fecha": _to_ddmmyyyy(g["fecha"]) if g["fecha"] else "",
                "odt": "\n".join(sorted(g["odts"])),
                "cliente": "\n".join(sorted(g["clientes"])),
                "comuna": "\n".join(sorted(g["comunas"])),
                "viaticoReal": float(real),
                "viaticoMax": float(vmax),
                "viaticoAPagar": float(a_pagar),
                "personalizado": codigo_norm in caps_normalizados,
            })
        out.sort(key=lambda x: x["codigo"])
        return out

    def set_viatico_cap(self, codigo: str, monto: Decimal, usuario: str = "") -> dict[str, Any]:
        codigo = str(codigo or "").strip()
        if not codigo:
            raise ValueError("Código diario requerido.")
        monto = Decimal(str(monto))
        if monto < 0:
            raise ValueError("El tope no puede ser negativo.")
        cap = self.db.scalar(select(RendicionViaticoCap).where(RendicionViaticoCap.codigo_diario == codigo))
        if monto <= 0:
            if cap:
                self.db.delete(cap)
                self.db.commit()
            return {"codigo": codigo, "viaticoMax": 0.0, "removed": True}
        if cap:
            cap.viatico_max = monto
            cap.updated_by = usuario or None
        else:
            self.db.add(RendicionViaticoCap(codigo_diario=codigo, viatico_max=monto, updated_by=usuario or None))
        self.db.commit()
        return {"codigo": codigo, "viaticoMax": float(monto)}

    def obtener_pagos(self) -> list[dict[str, Any]]:
        rows = self.db.scalars(select(RendicionPago).order_by(RendicionPago.fecha_pago.desc(), RendicionPago.id.desc())).all()
        out: list[dict[str, Any]] = []
        for r in rows:
            out.append({
                "id": r.id,
                "codigo": r.codigo_diario,
                "tecnico": r.tecnico,
                "rutTecnico": r.rut_tecnico or "",
                "tipoPago": r.tipo_pago,
                "fechaPago": _to_ddmmyyyy(r.fecha_pago) if r.fecha_pago else "",
                "monto": float(r.monto),
                "creadoPor": r.creado_por or "",
                "fechaRegistro": _to_ddmmyyyy_hhmm(r.fecha_registro) if r.fecha_registro else "",
            })
        return out

    def registrar_pago(self, data: dict[str, Any], usuario: str = "") -> dict[str, Any]:
        codigo = str(data.get("codigo") or "").strip()
        tecnico = str(data.get("tecnico") or "").strip()
        tipo_pago = str(data.get("tipoPago") or "Transferencia").strip() or "Transferencia"
        fecha_raw = data.get("fechaPago")
        monto_raw = data.get("monto")
        rut = str(data.get("rutTecnico") or "").strip() or None
        if not codigo or not tecnico:
            raise ValueError("Código diario y técnico son obligatorios.")
        try:
            monto = Decimal(str(monto_raw))
        except Exception as exc:
            raise ValueError("Monto inválido.") from exc
        if monto <= 0:
            raise ValueError("Monto debe ser mayor a 0.")
        fecha_pago: datetime
        if isinstance(fecha_raw, datetime):
            fecha_pago = fecha_raw
        else:
            fecha_str = str(fecha_raw or "").strip()
            if not fecha_str:
                raise ValueError("Fecha de pago requerida.")
            try:
                if "/" in fecha_str:
                    fecha_pago = datetime.strptime(fecha_str, "%d/%m/%Y")
                else:
                    fecha_pago = datetime.strptime(fecha_str, "%Y-%m-%d")
            except ValueError as exc:
                raise ValueError("Formato de fecha inválido. Usa AAAA-MM-DD o DD/MM/AAAA.") from exc

        pago = RendicionPago(
            codigo_diario=codigo,
            tecnico=tecnico,
            rut_tecnico=rut,
            tipo_pago=tipo_pago,
            fecha_pago=fecha_pago,
            monto=monto,
            creado_por=usuario or None,
        )
        self.db.add(pago)
        rendiciones_codigo = self.db.scalars(
            select(Rendicion).where(Rendicion.codigo_diario == codigo)
        ).all()
        for r in rendiciones_codigo:
            estado = str(r.estado_revision or "").strip().lower()
            if estado.startswith("acept"):
                r.estado_revision = "Pagado"
        self.db.commit()
        self.db.refresh(pago)
        return {"id": pago.id, "codigo": codigo, "monto": float(monto)}

    def eliminar_pago(self, pago_id: int) -> bool:
        pago = self.db.scalar(select(RendicionPago).where(RendicionPago.id == pago_id))
        if not pago:
            return False
        self.db.delete(pago)
        self.db.commit()
        return True

    def obtener_suma_pagos(self) -> list[dict[str, Any]]:
        rows = self.db.execute(
            select(
                RendicionPago.codigo_diario,
                func.coalesce(func.sum(RendicionPago.monto), 0),
                func.count(RendicionPago.id),
                func.max(RendicionPago.fecha_pago),
            ).group_by(RendicionPago.codigo_diario).order_by(RendicionPago.codigo_diario)
        ).all()
        caps_personalizados = self._viatico_caps_personalizados()
        caps_normalizados = {
            self._normalizar_codigo_diario(codigo): monto
            for codigo, monto in caps_personalizados.items()
        }

        out: list[dict[str, Any]] = []
        for codigo, suma, cantidad, ultimo_pago in rows:
            codigo = str(codigo or "").strip()
            monto_real = Decimal(str(suma or 0))
            codigo_norm = self._normalizar_codigo_diario(codigo)
            tope = caps_normalizados.get(codigo_norm, self.VIATICO_CAP_DIARIO_DEFECTO)
            out.append({
                "codigo": codigo,
                "suma": float(min(monto_real, tope)),
                "pagos": int(cantidad or 0),
                "ultimoPago": _to_ddmmyyyy(ultimo_pago) if ultimo_pago else "",
            })
        return out

    # =========================
    # PLANIFICACIÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œN
    # =========================
    def obtener_planificacion_total(
        self,
        mes: int,
        anio: int,
        estado: str = "Todos",
        tecnico: str = "Todos",
    ) -> dict[str, list[dict[str, Any]]]:
        resultado: dict[str, list[dict[str, Any]]] = {}

        stmt_inc = select(Registro).where(
            func.extract("month", Registro.fecha_derivacion_area) == mes,
            func.extract("year", Registro.fecha_derivacion_area) == anio,
        )
        if estado != "Todos":
            stmt_inc = stmt_inc.where(Registro.estado == estado)
        if tecnico != "Todos":
            stmt_inc = stmt_inc.where(
                or_(Registro.tecnicos == tecnico, Registro.acompanante == tecnico)
            )

        for row in self.db.scalars(stmt_inc).all():
            fecha = _to_ddmmyyyy(row.fecha_derivacion_area)
            resultado.setdefault(fecha, []).append(
                {
                    "odt": row.odt,
                    "cliente": row.cliente,
                    "direccion": row.direccion,
                    "servicio": row.problema,
                    "tecnico": row.tecnicos,
                    "acompanante": row.acompanante,
                    "estado": row.estado,
                    "origen": "incidencias",
                }
            )

        stmt_ven = (
            select(VentaODS, AdministracionODT, ServicioTecnicoVentaODT)
            .outerjoin(AdministracionODT, AdministracionODT.odt == VentaODS.codigo)
            .outerjoin(ServicioTecnicoVentaODT, ServicioTecnicoVentaODT.odt == VentaODS.codigo)
            .where(VentaODS.estado != "Anulada")
        )
        for ods_row, adm_row, st_row in self.db.execute(stmt_ven).all():
            fecha_base = (
                getattr(adm_row, "fecha_derivacion", None)
                or getattr(st_row, "updated_at", None)
                or ods_row.created_at
            )
            if not isinstance(fecha_base, datetime):
                continue
            if fecha_base.month != mes or fecha_base.year != anio:
                continue
            tecnico_venta = (
                (getattr(st_row, "tecnico_a_cargo", None) if st_row else None)
                or (adm_row.tecnico if adm_row else None)
            )
            if tecnico != "Todos" and tecnico_venta != tecnico:
                continue
            fecha = _to_ddmmyyyy(fecha_base)
            resultado.setdefault(fecha, []).append(
                {
                    "odt": ods_row.codigo,
                    "cliente": ods_row.nombre_sucursal or ods_row.razon_social,
                    "direccion": ods_row.direccion_sucursal,
                    "servicio": ods_row.tipo_servicio,
                    "tecnico": tecnico_venta,
                    "acompanante": (
                        (getattr(st_row, "acompanante", None) if st_row else None)
                        or (adm_row.acompanante if adm_row else None)
                    ),
                    "estado": ods_row.estado,
                    "origen": "ventas",
                }
            )
        return resultado
