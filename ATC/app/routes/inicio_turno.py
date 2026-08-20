from __future__ import annotations

import calendar
import hashlib
import hmac
import html
import logging
import math
import os
import random
import re
import secrets
import smtplib
import ssl
import threading
from collections import Counter, defaultdict
from io import BytesIO
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from email.utils import formatdate, parseaddr
from pathlib import Path
from typing import Optional
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

from sqlalchemy import extract, func, text
from sqlalchemy.exc import IntegrityError

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from sqlalchemy import or_

from ATC.app.core.config import settings
from ATC.app.core.db import get_db
from ATC.app.models.incidencias import SucursalBBDD
from ATC.app.models.inicio_turno import (
    GuardiaJustificacion,
    InicioTurnoGuardia,
    InicioTurnoRegistro,
    RecintoQrGenerado,
    RondaRegistro,
    SupervisorRegistro,
    TurnoEstipulado,
)
from ATC.app.models.user import User
from ATC.app.services.incidencias_service import IncidenciasService
from ATC.app.services.venta_trace_email_service import _email_html, _plain_from_html, _table
from ATC.app.routes.web import COOKIE_NAME as _COOKIE_NAME, _decode_cookie_token as _decode_web_token
from ATC.app.services.user_service import UserService as _UserService
from ATC.app.services import contrato_diario_service


router = APIRouter(tags=["inicio-turno"])
templates = Jinja2Templates(directory=str(Path(__file__).resolve().parents[1] / "templates"))
_log = logging.getLogger(__name__)


def _require_login(request: Request, db: Session = Depends(get_db)) -> User:
    """Exige sesion activa (misma cookie access_token que _usuario_sesion,
    definida mas abajo en este archivo — la referencia se resuelve recien
    cuando esta dependencia corre, no al definir esta funcion, asi que el
    orden no importa). Usar en endpoints de gestion (BBDD Guardias, aprobar
    rondas) que antes no verificaban ninguna sesion (hallazgo de auditoria
    de seguridad, ago 2026) — no se aplica al flujo de escaneo de QR de
    rondas, que es publico por diseno (el guardia no tiene cuenta, solo el
    codigo fisico)."""
    user = _usuario_sesion(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="No autenticado.")
    return user

TIPOS_TURNO = ("Dia", "Noche", "Extra", "Contrato Diario")
_TIPOS_MANUALES = ("Extra", "Contrato Diario")
_TURNOS_INFORME = ("Dia", "Noche", "Extra", "Contrato Diario")
_TURNOS_NOTIFICACION = {"Extra", "Contrato Diario"}
_TURNOS_NOTIFICACION_DESTINOS = [
    "glubiano@alguientecuida.cl",
    "jefe.seguridadfisica@alguientecuida.cl",
]
_CONTACTO_EMAIL = "contacto@alguientecuida.cl"
_ATC_ENV_PATH = Path(__file__).resolve().parents[2] / ".env"
_UBICACION_INICIO_TURNO_HABILITADA = False

_VENTANA_DIA   = (7 * 60 + 30, 8 * 60)
_VENTANA_NOCHE = (19 * 60 + 30, 20 * 60)
RADIO_MAXIMO_METROS = 200.0
DEFAULT_GUARDIAS = (
    {"rut": "211342854", "nombre": "Fernando Lubiano"},
)
NON_SELECTABLE_GUARDIA_NAME_PREFIXES = (
    "fernando andres lubiano moraga",
    "fernando lubiano",
    "gianpiero lubiano",
    "gianpiero lubiano forno",
)
NON_SELECTABLE_GUARDIA_RUTS = {"211342854", "165762878"}
DEPARTAMENTOS_GUARDIA = {
    "guardia",
    "guardiafulltime", "guardiasfulltime",
    "guardiaparttime", "guardiasparttime",
}


class InicioTurnoCreate(BaseModel):
    rut: str = Field(min_length=1, max_length=40)
    tipo_turno: Optional[str] = Field(default="", max_length=80)
    recinto: Optional[str] = Field(default=None, max_length=255)
    sucursal_id: Optional[int] = None
    latitud: Optional[float] = None
    longitud: Optional[float] = None
    precision_metros: Optional[float] = None
    ubicacion_estado: Optional[str] = Field(default=None, max_length=80)


def _normalizar_rut(value: object) -> str:
    text = str(value or "").strip().upper()
    return "".join(ch for ch in text if ch not in ". ")


def _normalizar_texto(value: object) -> str:
    import unicodedata

    text = unicodedata.normalize("NFD", str(value or "").casefold().strip())
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def _normalizar_rut_opciones(value: object) -> str:
    text = str(value or "").strip().upper()
    return "".join(ch for ch in text if ch not in ". -")


def _guardia_oculto_en_opciones(guardia: InicioTurnoGuardia | None) -> bool:
    if not guardia:
        return False
    rut_key = _normalizar_rut_opciones(getattr(guardia, "rut", "") or "")
    if rut_key in NON_SELECTABLE_GUARDIA_RUTS:
        return True
    nombre_norm = _normalizar_texto(getattr(guardia, "nombre", "") or "")
    return any(
        nombre_norm == prefix or nombre_norm.startswith(f"{prefix} ")
        for prefix in NON_SELECTABLE_GUARDIA_NAME_PREFIXES
    )


def _turno_informe(value: object) -> str:
    turno_norm = _normalizar_texto(value)
    for turno in _TURNOS_INFORME:
        if _normalizar_texto(turno) == turno_norm:
            return turno
    return ""


def _tipo_guardia_label(value: object) -> str:
    key = _normalizar_texto(value).replace(" ", "")
    if "parttime" in key:
        return "Part Time"
    if "fulltime" in key:
        return "Full Time"
    return ""


def _es_departamento_guardia(value: object) -> bool:
    partes = [
        _normalizar_texto(parte).replace(" ", "")
        for parte in str(value or "").split(";")
        if str(parte or "").strip()
    ]
    return any(parte in DEPARTAMENTOS_GUARDIA for parte in partes)


def _tipo_guardia_lookup(db: Session) -> dict[str, dict[str, str]]:
    lookup = {"rut": {}, "nombre": {}}
    users = db.query(User.username, User.name, User.department).filter(User.department.isnot(None)).all()
    for username, name, department in users:
        tipo = _tipo_guardia_label(department)
        if not tipo:
            continue
        rut_key = _normalizar_rut(username)
        nombre_key = _normalizar_texto(name)
        if rut_key:
            lookup["rut"].setdefault(rut_key, tipo)
        if nombre_key:
            lookup["nombre"].setdefault(nombre_key, tipo)
    return lookup


def _tipo_guardia_para(nombre: object, rut: object, lookup: dict[str, dict[str, str]]) -> str:
    rut_key = _normalizar_rut(rut)
    if rut_key:
        tipo = lookup.get("rut", {}).get(rut_key)
        if tipo:
            return tipo
    nombre_key = _normalizar_texto(nombre)
    if nombre_key:
        return lookup.get("nombre", {}).get(nombre_key, "")
    return ""


def _resolver_rut_guardia_por_nombre(db: Session, nombre: str) -> str:
    nombre_norm = _normalizar_texto(nombre)
    if not nombre_norm:
        return ""
    guardias = db.query(InicioTurnoGuardia).all()
    for g in guardias:
        if _normalizar_texto(g.nombre) == nombre_norm:
            return _normalizar_rut(g.rut)
    return ""


def _usuarios_guardia_activos(db: Session) -> list[User]:
    usuarios = (
        db.query(User)
        .filter(User.is_active == True, User.department.isnot(None))  # noqa: E712 - .is_(True) genera "IS 1", invalido en T-SQL
        .order_by(User.name.asc())
        .all()
    )
    return [u for u in usuarios if _es_departamento_guardia(u.department)]


def _usuario_guardia_activo_por_rut(db: Session, rut: object) -> User | None:
    rut_norm = _normalizar_rut(rut)
    if not rut_norm:
        return None
    for user in _usuarios_guardia_activos(db):
        if _normalizar_rut(user.username) == rut_norm:
            return user
    return None


def _usuario_guardia_activo_por_nombre(db: Session, nombre: object) -> User | None:
    nombre_norm = _normalizar_texto(nombre)
    if not nombre_norm:
        return None
    coincidencias = [
        user for user in _usuarios_guardia_activos(db)
        if _normalizar_texto(user.name) == nombre_norm
    ]
    if len(coincidencias) == 1:
        return coincidencias[0]
    return None


def _guardia_desde_usuario(user: User) -> InicioTurnoGuardia:
    return InicioTurnoGuardia(
        id=int(user.id or 0),
        rut=str(user.username or "").strip(),
        nombre=str(user.name or "").strip(),
    )


def _resolver_guardia_activo_payload(db: Session, rut: object, nombre: object) -> User:
    user = _usuario_guardia_activo_por_rut(db, rut)
    if not user:
        user = _usuario_guardia_activo_por_nombre(db, nombre)
    if not user:
        raise HTTPException(
            status_code=422,
            detail="Selecciona un guardia válido desde la BBDD (GuardiaFulltime o GuardiaParttime).",
        )
    return user


def _parse_float(value: object) -> float | None:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _sucursal_coords(sucursal: SucursalBBDD | None) -> tuple[float | None, float | None]:
    if not sucursal:
        return None, None
    lat = _parse_float(sucursal.latitud)
    lng = _parse_float(sucursal.longitud)
    if lat is not None and lng is not None:
        return lat, lng

    raw = str(sucursal.latitud_longitud or "").strip()
    if "," not in raw:
        return None, None
    parts = raw.split(",", 1)
    return _parse_float(parts[0]), _parse_float(parts[1])


def _obtener_o_geocodificar_sucursal(db: Session, sucursal: SucursalBBDD) -> tuple[float | None, float | None]:
    lat, lng = _sucursal_coords(sucursal)
    if lat is not None and lng is not None:
        return lat, lng

    direccion = str(sucursal.direccion_sucursal or "").strip()
    comuna = str(sucursal.comuna or "").strip() or "Quintero"
    query = ", ".join(part for part in [direccion, comuna, "Chile"] if part)
    if not query:
        return None, None

    lat_txt, lng_txt = IncidenciasService(db)._geocodificar_direccion(query)
    lat = _parse_float(lat_txt)
    lng = _parse_float(lng_txt)
    if lat is None or lng is None:
        return None, None

    sucursal.latitud = f"{lat:.6f}"
    sucursal.longitud = f"{lng:.6f}"
    sucursal.latitud_longitud = f"{lat:.6f}, {lng:.6f}"
    db.commit()
    return lat, lng


def _distancia_metros(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    earth_radius_m = 6371000.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lng2 - lng1)
    a = (
        math.sin(delta_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    )
    return earth_radius_m * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _recinto_label(sucursal: SucursalBBDD | None) -> str:
    if not sucursal:
        return ""
    # Ojo: antes se mostraba "Empresa - Sucursal", pero para los QR (rondas y
    # recintos) solo se quiere el nombre de la sucursal — con empresa/direccion
    # como respaldo si la sucursal viene vacia.
    nombre_sucursal = str(sucursal.nombre_sucursal or "").strip()
    if nombre_sucursal:
        return nombre_sucursal
    nombre_empresa = str(sucursal.nombre_empresa or "").strip()
    if nombre_empresa:
        return nombre_empresa
    return str(sucursal.direccion_sucursal or "").strip()


def _env_inicio_turno() -> dict[str, str]:
    values: dict[str, str] = {}
    try:
        for line in _ATC_ENV_PATH.read_text(encoding="utf-8-sig").splitlines():
            raw = line.strip()
            if not raw or raw.startswith("#") or "=" not in raw:
                continue
            key, value = raw.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    except Exception:
        pass
    return values


def _contacto_smtp_config() -> dict[str, object]:
    env_file = _env_inicio_turno()

    def env_get(*keys: str, default: str = "") -> str:
        for key in keys:
            value = str(os.getenv(key) or env_file.get(key) or "").strip()
            if value:
                return value
        return default

    username = env_get("CONTACTO_SMTP_USERNAME", "CONTACTO_SMTP_USER")
    password = env_get("CONTACTO_SMTP_PASSWORD")
    from_email = parseaddr(env_get("CONTACTO_SMTP_FROM_EMAIL", default=username))[1].strip().lower()

    if not username and not password:
        smtp2_from = parseaddr(env_get("SMTP2_FROM_EMAIL", default=env_get("SMTP2_USERNAME")))[1].strip().lower()
        if smtp2_from == _CONTACTO_EMAIL:
            username = env_get("SMTP2_USERNAME")
            password = env_get("SMTP2_PASSWORD")
            from_email = smtp2_from

    if from_email != _CONTACTO_EMAIL:
        return {"enabled": False, "reason": f"SMTP de contacto no configurado para {_CONTACTO_EMAIL}"}

    host = env_get("CONTACTO_SMTP_HOST", default=env_get("SMTP2_HOST", "SMTP_HOST", default="smtp.gmail.com"))
    port_raw = env_get("CONTACTO_SMTP_PORT", default=env_get("SMTP2_PORT", "SMTP_PORT", default="587"))
    try:
        port = int(port_raw)
    except Exception:
        port = 587
    use_tls = env_get("CONTACTO_SMTP_USE_TLS", default="true").lower() not in {"0", "false", "no", "off"}
    use_ssl = env_get("CONTACTO_SMTP_USE_SSL", default="false").lower() in {"1", "true", "yes", "on"}
    timeout_raw = env_get("CONTACTO_SMTP_TIMEOUT_SEC", default="20")
    try:
        timeout = int(timeout_raw)
    except Exception:
        timeout = 20
    from_name = env_get("CONTACTO_SMTP_FROM_NAME", default="Alguien Te Cuida")
    return {
        "enabled": bool(host and username and password),
        "reason": "" if host and username and password else "SMTP de contacto incompleto",
        "host": host,
        "port": port,
        "username": username,
        "password": password,
        "from_email": from_email,
        "from_name": from_name,
        "use_tls": use_tls,
        "use_ssl": use_ssl,
        "timeout": timeout,
    }


def _send_contacto_inicio_turno(to: str | list[str], subject: str, html_body: str) -> None:
    cfg = _contacto_smtp_config()
    if not cfg.get("enabled"):
        raise ValueError(str(cfg.get("reason") or "SMTP de contacto no configurado"))

    from_email = str(cfg["from_email"])
    from_name = str(cfg.get("from_name") or "Alguien Te Cuida").strip()
    recipients = [item.strip() for item in ([to] if isinstance(to, str) else to) if str(item or "").strip()]
    msg = EmailMessage()
    msg["From"] = f"{from_name} <{from_email}>" if from_name else from_email
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = (subject or "").replace("\r", " ").replace("\n", " ").strip()
    msg["Date"] = formatdate(localtime=True)
    msg.set_content(_plain_from_html(html_body), subtype="plain", charset="utf-8")
    msg.add_alternative(html_body, subtype="html", charset="utf-8")

    if bool(cfg.get("use_ssl")):
        with smtplib.SMTP_SSL(str(cfg["host"]), int(cfg["port"]), timeout=int(cfg["timeout"]), context=ssl.create_default_context()) as server:
            server.login(str(cfg["username"]), str(cfg["password"]))
            server.send_message(msg, to_addrs=recipients)
        return

    with smtplib.SMTP(str(cfg["host"]), int(cfg["port"]), timeout=int(cfg["timeout"])) as server:
        server.ehlo()
        if bool(cfg.get("use_tls")):
            server.starttls(context=ssl.create_default_context())
            server.ehlo()
        server.login(str(cfg["username"]), str(cfg["password"]))
        server.send_message(msg, to_addrs=recipients)


def _enviar_notificacion_turno_async(
    *,
    tipo_turno: object,
    nombre_guardia: object,
    rut: object,
    recinto: object,
    fecha_registro: object = None,
    fecha_turno: object = None,
    supervisor: object = None,
    nota: object = None,
    origen: str = "inicio de turno",
    registro_id: object = "",
) -> None:
    tipo_turno = str(tipo_turno or "").strip()
    if tipo_turno not in _TURNOS_NOTIFICACION:
        return

    tipo_label = "Turno Extra" if tipo_turno == "Extra" else tipo_turno
    subject = f"Nuevo {tipo_label} registrado"
    rows = [
        ("Tipo de turno", tipo_label),
        ("Guardia", str(nombre_guardia or "").strip()),
        ("RUT", str(rut or "").strip() or "-"),
        ("Recinto", str(recinto or "").strip()),
    ]
    if fecha_turno:
        rows.append(("Fecha turno", str(fecha_turno)))
    if supervisor:
        rows.append(("Registrado por supervisor", str(supervisor).strip()))
    if nota:
        rows.append(("Nota supervisor", str(nota).strip()))
    if fecha_registro:
        if hasattr(fecha_registro, "strftime"):
            fecha_registro = fecha_registro.strftime("%d/%m/%Y %H:%M")
        rows.append(("Fecha registro", str(fecha_registro)))

    body = _email_html(
        title=f"Nuevo {tipo_label}",
        sections=[
            (
                '<p style="margin:0 0 12px 0;color:#334155;font-family:Arial,sans-serif;'
                'font-size:14px;line-height:1.72;">'
                f"Se registró un nuevo <strong>{tipo_label}</strong> desde {origen}."
                "</p>"
            ),
            _table(rows),
        ],
    )

    def _worker() -> None:
        try:
            _send_contacto_inicio_turno(_TURNOS_NOTIFICACION_DESTINOS, subject, body)
        except Exception as exc:
            _log.warning(
                "No se pudo enviar notificacion de inicio de turno a %s: %s",
                ", ".join(_TURNOS_NOTIFICACION_DESTINOS),
                exc,
            )

    thread_id = str(registro_id or "").strip() or "manual"
    threading.Thread(target=_worker, daemon=True, name=f"inicio-turno-email-{thread_id}").start()


_DOBLE_MARCAJE_DESTINOS = [
    "jefe.seguridadfisica@alguientecuida.cl",
]


def _enviar_alerta_doble_marcaje_async(
    *,
    nombre_guardia: object,
    rut: object,
    marcajes: list[dict],
    registro_id: object = "",
) -> None:
    """Alerta cuando un guardia marca inicio de turno 2+ veces el mismo día
    (independiente del tipo de turno). Mismo formato que la notificación de
    turno extra."""
    nombre = str(nombre_guardia or "").strip()
    subject = f"Alerta: doble marcaje de inicio de turno — {nombre}"
    rows = [
        ("Guardia", nombre),
        ("RUT", str(rut or "").strip() or "-"),
        ("Marcajes de hoy", str(len(marcajes))),
    ]
    for idx, m in enumerate(marcajes, 1):
        rows.append((
            f"Marcaje {idx}",
            f"{m.get('hora', '')} · {m.get('recinto', '')} · {m.get('turno', '')}",
        ))

    body = _email_html(
        title="Doble marcaje de inicio de turno",
        sections=[
            (
                '<p style="margin:0 0 12px 0;color:#334155;font-family:Arial,sans-serif;'
                'font-size:14px;line-height:1.72;">'
                f"El guardia <strong>{nombre}</strong> registró "
                f"<strong>{len(marcajes)} inicios de turno el mismo día</strong>."
                "</p>"
            ),
            _table(rows),
        ],
    )

    def _worker() -> None:
        try:
            _send_contacto_inicio_turno(_DOBLE_MARCAJE_DESTINOS, subject, body)
        except Exception as exc:
            _log.warning(
                "No se pudo enviar alerta de doble marcaje a %s: %s",
                ", ".join(_DOBLE_MARCAJE_DESTINOS),
                exc,
            )

    thread_id = str(registro_id or "").strip() or "doble"
    threading.Thread(target=_worker, daemon=True, name=f"doble-marcaje-email-{thread_id}").start()


def _alertar_doble_marcaje_si_corresponde(db: Session, registro: InicioTurnoRegistro) -> None:
    """Avisa si el guardia ya tiene 2+ inicios en el mismo dia.

    La regla es por guardia y fecha calendario, sin filtrar por turno,
    sucursal ni recinto. Se normaliza el RUT para cubrir registros antiguos
    con formatos distintos.
    """
    rut_registro = _normalizar_rut(registro.rut)
    nombre_registro = str(registro.nombre_guardia or "").strip().casefold()
    if not rut_registro and not nombre_registro:
        return

    dia = registro.registrado_at.date()
    dia_inicio = datetime(dia.year, dia.month, dia.day)
    dia_fin = dia_inicio + timedelta(days=1)
    registros_dia = (
        db.query(InicioTurnoRegistro)
        .filter(
            InicioTurnoRegistro.registrado_at >= dia_inicio,
            InicioTurnoRegistro.registrado_at < dia_fin,
        )
        .order_by(InicioTurnoRegistro.registrado_at.asc())
        .all()
    )
    marcajes = [
        item
        for item in registros_dia
        if str(item.estado or "activo") != "archivado" and (
            (
                rut_registro
                and _normalizar_rut(item.rut) == rut_registro
            ) or (
                not rut_registro
                and str(item.nombre_guardia or "").strip().casefold() == nombre_registro
            )
        )
    ]
    if len(marcajes) < 2:
        return

    # Farmacia Municipal <-> DESAM: el mismo guardia cubriendo ambos puntos
    # el mismo dia no es un error de doble marcaje (ver
    # _PAR_FARMACIA_DESAM_SIN_ALERTA) — se suprime el aviso solo si TODOS
    # los marcajes del dia caen dentro de este par, con ambos representados.
    # Muchos marcajes de DESAM se cargan manualmente con sucursal_id NULL
    # (solo `recinto` como texto), por eso se resuelve por texto tambien.
    es_farmacia_o_desam = [_es_recinto_farmacia_o_desam(m.sucursal_id, m.recinto) for m in marcajes]
    es_farmacia = any(_normalizar_texto(m.recinto) == _normalizar_texto("Farmacia Municipal") for m in marcajes)
    es_desam = any(
        _normalizar_texto(m.recinto) == _normalizar_texto("MQUIN Edificio de Administración DESAM")
        or m.sucursal_id == _SUCURSAL_DESAM
        for m in marcajes
    )
    if all(es_farmacia_o_desam) and es_farmacia and es_desam:
        return

    _enviar_alerta_doble_marcaje_async(
        nombre_guardia=registro.nombre_guardia,
        rut=registro.rut,
        marcajes=[
            {
                "hora": m.registrado_at.strftime("%H:%M"),
                "recinto": m.recinto,
                "turno": m.tipo_turno,
            }
            for m in marcajes
        ],
        registro_id=registro.id,
    )


def _fusionar_automatico_si_corresponde(db: Session, registro: InicioTurnoRegistro) -> None:
    """Fusion automatica de marcajes duplicados por traslado (ago 2026 —
    fin de la marcha blanca manual, ver _detectar_candidatos_fusion). Si
    este marcaje calza en el mismo dia, mismo guardia, y mismo grupo de
    _GRUPOS_FUSION_RECINTOS con otro(s) marcaje(s) activo(s) — sin importar
    tipo_turno, salvo "Noche" (excluido) — se archiva automaticamente el o
    los mas antiguos y queda activo solo el mas tardio (el recinto donde el
    guardia termino). No requiere aprobacion manual."""
    if _normalizar_texto(registro.tipo_turno) == _normalizar_texto("Noche"):
        return
    idx_grupo = _grupo_fusion_de(registro.sucursal_id)
    if idx_grupo is None:
        return

    rut_registro = _normalizar_rut(registro.rut)
    nombre_registro = str(registro.nombre_guardia or "").strip().casefold()
    if not rut_registro and not nombre_registro:
        return

    dia = registro.registrado_at.date()
    dia_inicio = datetime(dia.year, dia.month, dia.day)
    dia_fin = dia_inicio + timedelta(days=1)
    registros_dia = (
        db.query(InicioTurnoRegistro)
        .filter(
            InicioTurnoRegistro.registrado_at >= dia_inicio,
            InicioTurnoRegistro.registrado_at < dia_fin,
        )
        .all()
    )

    grupo_regs = []
    for item in registros_dia:
        if str(item.estado or "activo") == "archivado":
            continue
        if _normalizar_texto(item.tipo_turno) == _normalizar_texto("Noche"):
            continue
        if _grupo_fusion_de(item.sucursal_id) != idx_grupo:
            continue
        item_rut = _normalizar_rut(item.rut)
        item_nombre = str(item.nombre_guardia or "").strip().casefold()
        mismo_guardia = (
            (rut_registro and item_rut == rut_registro)
            or (not rut_registro and item_nombre == nombre_registro)
        )
        if mismo_guardia:
            grupo_regs.append(item)

    if len(grupo_regs) < 2:
        return

    grupo_regs.sort(key=lambda r: r.registrado_at)
    sobreviviente = grupo_regs[-1]
    tz = ZoneInfo(settings.timezone or "America/Santiago")
    ahora = datetime.now(tz).replace(tzinfo=None)
    motivo = f"Fusion automatica con marcaje en {sobreviviente.recinto} ({sobreviviente.registrado_at.strftime('%H:%M')})"
    for reg in grupo_regs[:-1]:
        reg.estado = "archivado"
        reg.fusionado_con_id = sobreviviente.id
        reg.archivado_motivo = motivo
        reg.archivado_en = ahora
        reg.archivado_por = "sistema (fusion automatica)"
    db.commit()


_SIN_TURNO_DESTINOS = [
    "jefe.seguridadfisica@alguientecuida.cl",
]


def _email_alerta_sucursal_sin_turno(*, dependencia: str, fecha_str: str, estipulado: int) -> str:
    """Alerta de alto impacto (pedido explicito: 'profesional, unico y
    llamativo de advertencia, sin emoji, con presencia') para una sucursal
    que quedo un dia entero sin ningun turno registrado. Deliberadamente NO
    reusa el estilo generico de _email_html (venta_trace_email_service) —
    esto tiene que destacar por sobre las demas alertas de turnos."""
    estipulado_txt = str(estipulado) if estipulado else "—"
    return f"""<!doctype html>
<html>
<body style="margin:0;padding:0;background:#eef1f5;font-family:Arial,Helvetica,sans-serif;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#eef1f5;padding:36px 16px;">
<tr><td align="center">
<table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;background:#ffffff;border-radius:10px;overflow:hidden;box-shadow:0 10px 32px rgba(15,23,42,.22);">

<tr><td style="background:#111318;padding:0;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
    <tr><td style="background:#b91c1c;height:6px;line-height:6px;font-size:0;">&nbsp;</td></tr>
    <tr><td style="padding:30px 32px 26px;">
      <div style="font-size:11px;font-weight:700;letter-spacing:.16em;color:#f87171;text-transform:uppercase;margin:0 0 12px;">Alerta operacional &mdash; cobertura de guardia</div>
      <div style="font-size:27px;font-weight:800;letter-spacing:-.01em;color:#ffffff;line-height:1.18;margin:0;">Sucursal sin guardia registrado</div>
    </td></tr>
  </table>
</td></tr>

<tr><td style="padding:28px 32px 8px;border-bottom:1px solid #eef0f2;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td style="width:100px;vertical-align:top;">
        <div style="font-size:46px;font-weight:800;color:#b91c1c;line-height:1;">0</div>
        <div style="font-size:10px;font-weight:700;letter-spacing:.08em;color:#9ca3af;text-transform:uppercase;margin-top:5px;">Turnos</div>
      </td>
      <td style="vertical-align:top;padding-left:6px;">
        <div style="font-size:18px;font-weight:800;color:#111827;line-height:1.3;">{dependencia}</div>
        <div style="font-size:13px;color:#4b5563;margin-top:6px;line-height:1.6;">No se registro ningun inicio de turno en esta sucursal durante el {fecha_str}. La sucursal quedó sin cobertura.</div>
      </td>
    </tr>
  </table>
</td></tr>

<tr><td style="padding:24px 32px 8px;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;">
    <tr style="background:#f8fafc;">
      <td style="padding:12px 16px;font-size:11px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:#6b7280;border-bottom:1px solid #e5e7eb;width:46%;">Sucursal</td>
      <td style="padding:12px 16px;font-size:13.5px;font-weight:700;color:#111827;border-bottom:1px solid #e5e7eb;">{dependencia}</td>
    </tr>
    <tr>
      <td style="padding:12px 16px;font-size:11px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:#6b7280;border-bottom:1px solid #e5e7eb;">Fecha</td>
      <td style="padding:12px 16px;font-size:13.5px;font-weight:700;color:#111827;border-bottom:1px solid #e5e7eb;">{fecha_str}</td>
    </tr>
    <tr style="background:#f8fafc;">
      <td style="padding:12px 16px;font-size:11px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:#6b7280;border-bottom:1px solid #e5e7eb;">Turnos estipulados</td>
      <td style="padding:12px 16px;font-size:13.5px;font-weight:700;color:#111827;border-bottom:1px solid #e5e7eb;">{estipulado_txt}</td>
    </tr>
    <tr>
      <td style="padding:12px 16px;font-size:11px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:#6b7280;">Estado</td>
      <td style="padding:12px 16px;font-size:13.5px;font-weight:800;color:#b91c1c;">Sin cobertura</td>
    </tr>
  </table>
</td></tr>

<tr><td style="padding:24px 32px 30px;">
  <div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:17px 19px;font-size:13.5px;color:#7f1d1d;line-height:1.65;font-weight:700;">
    Requiere revision inmediata &mdash; confirmar con el guardia asignado y regularizar el registro, o gestionar un reemplazo.
  </div>
</td></tr>

<tr><td style="background:#f8fafc;padding:16px 32px;border-top:1px solid #e5e7eb;">
  <div style="font-size:11px;color:#9ca3af;font-weight:600;">Alguien Te Cuida &mdash; Sistema de Control de Turnos</div>
</td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""


def _enviar_alerta_sucursal_sin_turno(*, dependencia: str, fecha_str: str, estipulado: int) -> None:
    subject = f"ALERTA: {dependencia} sin guardia registrado — {fecha_str}"
    body = _email_alerta_sucursal_sin_turno(dependencia=dependencia, fecha_str=fecha_str, estipulado=estipulado)
    _send_contacto_inicio_turno(_SIN_TURNO_DESTINOS, subject, body)


def _dias_trabajados_guardia(db: Session, nombre_guardia: str, centro: date | None = None) -> set:
    """Fechas (QR + supervisor, combinadas) en que este guardia tiene un
    registro, buscando por nombre normalizado. Usa una ventana de 30 días
    alrededor de `centro` (por defecto hoy) para no traer el historico
    completo — `centro` debe ser la fecha que se está registrando, no
    necesariamente hoy, porque los supervisores pueden cargar turnos de
    meses futuros o pasados."""
    nombre_norm = _normalizar_texto(nombre_guardia)
    if not nombre_norm:
        return set()
    centro = centro or date.today()
    desde = centro - timedelta(days=15)
    hasta = centro + timedelta(days=15)
    fechas: set = set()
    for r in db.query(InicioTurnoRegistro).filter(
        InicioTurnoRegistro.registrado_at >= datetime(desde.year, desde.month, desde.day),
        InicioTurnoRegistro.registrado_at <= datetime(hasta.year, hasta.month, hasta.day, 23, 59),
    ):
        if _normalizar_texto(r.nombre_guardia) == nombre_norm:
            fechas.add(r.registrado_at.date())
    for r in db.query(SupervisorRegistro).filter(
        SupervisorRegistro.fecha >= desde,
        SupervisorRegistro.fecha <= hasta,
    ):
        if _normalizar_texto(r.nombre_guardia) == nombre_norm:
            fechas.add(r.fecha)
    return fechas


def _racha_consecutiva_incluyendo(db: Session, nombre_guardia: str, fecha: date) -> int:
    """Cuantos dias seguidos (terminando en `fecha`, inclusive) tiene
    trabajados este guardia, contando `fecha` como trabajado aunque el
    registro todavia no se haya guardado (para poder advertir ANTES de
    guardar)."""
    fechas = _dias_trabajados_guardia(db, nombre_guardia, centro=fecha)
    fechas.add(fecha)
    dias = 1
    cursor = fecha - timedelta(days=1)
    while cursor in fechas:
        dias += 1
        cursor -= timedelta(days=1)
    return dias


_JORNADA_EXTENDIDA_DESTINOS = _DOBLE_MARCAJE_DESTINOS


def _enviar_alerta_jornada_extendida_async(
    *,
    nombre_guardia: object,
    rut: object,
    recinto: object,
    dias_consecutivos: int,
    fecha: object,
    origen: str,
    registro_id: object = "",
) -> None:
    """Alerta a jefatura cuando un guardia queda registrado (por el
    supervisor) con 6 o mas dias seguidos trabajados (QR + supervisor
    combinados)."""
    nombre = str(nombre_guardia or "").strip()
    subject = f"Alerta: {dias_consecutivos} días seguidos trabajados — {nombre}"
    rows = [
        ("Guardia", nombre),
        ("RUT", str(rut or "").strip() or "-"),
        ("Recinto", str(recinto or "").strip()),
        ("Días consecutivos", str(dias_consecutivos)),
        ("Última fecha", str(fecha)),
        ("Origen del registro", origen),
    ]
    body = _email_html(
        title="Jornada extendida",
        sections=[
            (
                '<p style="margin:0 0 12px 0;color:#334155;font-family:Arial,sans-serif;'
                'font-size:14px;line-height:1.72;">'
                f"El guardia <strong>{nombre}</strong> quedó registrado con "
                f"<strong>{dias_consecutivos} días seguidos trabajados</strong>."
                "</p>"
            ),
            _table(rows),
        ],
    )

    def _worker() -> None:
        try:
            _send_contacto_inicio_turno(_JORNADA_EXTENDIDA_DESTINOS, subject, body)
        except Exception as exc:
            _log.warning(
                "No se pudo enviar alerta de jornada extendida a %s: %s",
                ", ".join(_JORNADA_EXTENDIDA_DESTINOS),
                exc,
            )

    thread_id = str(registro_id or "").strip() or "jornada"
    threading.Thread(target=_worker, daemon=True, name=f"jornada-extendida-email-{thread_id}").start()


def _listar_recintos(db: Session) -> list[dict[str, str | int]]:
    rows = (
        db.query(SucursalBBDD)
        .order_by(SucursalBBDD.nombre_empresa.asc(), SucursalBBDD.nombre_sucursal.asc())
        .all()
    )
    return [
        {
            "id": row.id,
            "label": _recinto_label(row),
            "direccion": str(row.direccion_sucursal or "").strip(),
        }
        for row in rows
        if _recinto_label(row)
    ]


# Lista cerrada de recintos de Quintero (pedido explícito): solo estos 19
# aparecen en el QR de guardias y en el registro del supervisor. Cuando había
# una entrada duplicada en bbdd_sucursales ("MQUIN X" y "X"), se eligió la
# que ya tenía historial real de marcaciones. Posta Loncura, Cesfam, Farmacia
# y Desam van al final (pedido explícito), en ese orden.
_QUINTERO_SUCURSAL_IDS = [
    137,   # Cementerio Municipal
    251,   # Escombrera Municipal
    439,   # Terminal de Buses (Nuevo Terminal de Buses)
    255,   # Estadio Municipal
    666,   # Estadio Municipal Cancha Sintética 1
    667,   # Estadio Municipal Cancha Sintética 2
    234,   # Edificio Consistorial Base
    774,   # CEMCO - Operadores de cámaras de seguridad
    668,   # Nueva Biblioteca Municipal
    217,   # DIDECO
    380,   # Juzgado Policía Local
    459,   # Parque Municipal
    238,   # Medio Ambiente
    669,   # Albergue Municipal
    35,    # Aparcadero Municipal
    478,   # Posta de Salud Loncura   -> al final
    438,   # Cesfam Quintero          -> al final
    257,   # Farmacia Municipal       -> al final
    236,   # Edificio de Administración DESAM -> al final
]


# Grupos de recintos entre los que un guardia se traslada dentro del MISMO
# turno (pedido explicito, ago 2026): ej. 4 marcan en Consistorial a las 8am
# y 1 de ellos se traslada a Juzgado a las 10am porque abre a esa hora — hoy
# eso genera 2 filas en inicio_turno_registros para ese guardia ese dia, e
# infla el conteo mensual de "Total Turnos". Solo dentro de estos grupos se
# sugiere fusionar (ver _detectar_candidatos_fusion) — independiente del
# tipo_turno de cada marcaje, salvo turno "Noche" (excluido a proposito).
_GRUPOS_FUSION_RECINTOS: list[set[int]] = [
    {234, 380, 668},   # Edificio Consistorial Base <-> Juzgado Policia Local <-> Nueva Biblioteca Municipal
    {439, 666, 667},   # Terminal de Buses <-> Cancha Sintetica 1 <-> Cancha Sintetica 2
    # Farmacia Municipal <-> DESAM (257, 236) NO entra aca a proposito — ver
    # _PAR_FARMACIA_DESAM_SIN_ALERTA mas abajo: ese par no se fusiona (los 2
    # marcajes se necesitan vivos para pago/contabilidad de turnos), solo se
    # evita que infle el conteo del informe y se suprime el mail de doble
    # marcaje — pedido explicito, ago 2026.
]

# Farmacia Municipal <-> DESAM: el mismo guardia de Farmacia se traslada a
# cubrir DESAM de 17 a 20hrs. A diferencia de _GRUPOS_FUSION_RECINTOS, ese
# segundo marcaje NO se archiva (sigue vivo, sigue contando turno real en
# Cumplimiento de Turnos) — solo se excluye del conteo de "Dia" en el
# informe/vista previa de "tabla de asistencia" (_agrupar_turnos_por_guardia,
# no debe pagarse como turno aparte) y del mail de doble marcaje
# (_alertar_doble_marcaje_si_corresponde) — pedido explicito, ago 2026.
_SUCURSAL_FARMACIA = 257
_SUCURSAL_DESAM = 236
_PAR_FARMACIA_DESAM_SIN_ALERTA = {_SUCURSAL_FARMACIA, _SUCURSAL_DESAM}
# Muchos marcajes de DESAM se cargan manualmente desde el panel y solo
# guardan `recinto` como texto, con sucursal_id NULL (mismo problema que ya
# se resolvio para el conteo real de Cumplimiento de Turnos) — se matchea
# tambien por texto normalizado, no solo por sucursal_id.
_RECINTOS_FARMACIA_DESAM_NORM = {
    _normalizar_texto("Farmacia Municipal"),
    _normalizar_texto("MQUIN Edificio de Administración DESAM"),
}


def _es_recinto_farmacia_o_desam(sucursal_id: int | None, recinto: object) -> bool:
    if sucursal_id in _PAR_FARMACIA_DESAM_SIN_ALERTA:
        return True
    return _normalizar_texto(recinto) in _RECINTOS_FARMACIA_DESAM_NORM


def _grupo_fusion_de(sucursal_id: int | None) -> int | None:
    if sucursal_id is None:
        return None
    for idx, grupo in enumerate(_GRUPOS_FUSION_RECINTOS):
        if sucursal_id in grupo:
            return idx
    return None


def _detectar_candidatos_fusion(
    registros_qr: list[InicioTurnoRegistro],
    rut_lookup: dict[str, str],
) -> list[dict]:
    """Candidatos a fusion: mismo guardia, mismo dia, 2+ marcajes activos en
    recintos del mismo grupo de traslado (_GRUPOS_FUSION_RECINTOS),
    independiente del tipo_turno de cada marcaje — salvo turno "Noche", que
    queda afuera de esta deteccion (pedido explicito, ago 2026). El
    "sobreviviente" propuesto es siempre el marcaje mas tardio del dia (el
    recinto donde el guardia termino)."""
    grupos: dict[tuple, list[InicioTurnoRegistro]] = defaultdict(list)
    for r in registros_qr:
        if str(r.estado or "activo") == "archivado":
            continue
        if _normalizar_texto(r.tipo_turno) == _normalizar_texto("Noche"):
            continue
        idx_grupo = _grupo_fusion_de(r.sucursal_id)
        if idx_grupo is None:
            continue
        nombre_key = _normalizar_texto(r.nombre_guardia)
        rut = rut_lookup.get(nombre_key, "") or _normalizar_rut(r.rut)
        identidad = rut or nombre_key
        if not identidad:
            continue
        dia = r.registrado_at.date()
        grupos[(identidad, dia, idx_grupo)].append(r)

    candidatos: list[dict] = []
    for (_identidad, dia, _idx_grupo), regs in grupos.items():
        if len(regs) < 2:
            continue
        regs_ordenados = sorted(regs, key=lambda item: item.registrado_at)
        sobreviviente = regs_ordenados[-1]
        archivar = regs_ordenados[:-1]
        candidatos.append({
            "nombre": sobreviviente.nombre_guardia,
            "fecha": dia.strftime("%d/%m/%Y"),
            "sobreviviente": {
                "id": sobreviviente.id,
                "recinto": sobreviviente.recinto,
                "hora": sobreviviente.registrado_at.strftime("%H:%M"),
                "turno": sobreviviente.tipo_turno,
            },
            "archivar": [
                {
                    "id": r.id,
                    "recinto": r.recinto,
                    "hora": r.registrado_at.strftime("%H:%M"),
                    "turno": r.tipo_turno,
                }
                for r in archivar
            ],
        })
    candidatos.sort(key=lambda c: c["fecha"])
    return candidatos


def _listar_recintos_qr(db: Session) -> list[dict[str, str | int]]:
    rows = (
        db.query(SucursalBBDD)
        .filter(SucursalBBDD.id.in_(_QUINTERO_SUCURSAL_IDS))
        .order_by(SucursalBBDD.nombre_sucursal.asc())
        .all()
    )
    por_id = {row.id: row for row in rows}
    return [
        {
            "id": por_id[sid].id,
            "label": _recinto_label(por_id[sid]),
            "direccion": str(por_id[sid].direccion_sucursal or "").strip(),
        }
        for sid in _QUINTERO_SUCURSAL_IDS
        if sid in por_id
    ]


_PRIVADOS_SUCURSAL_IDS = [
    411,   # MACH Camino I.
    545,   # Total Transport Placilla
    529,   # Soprodi
    650,   # Atlas Copco Renca
    533,   # Storage Viña
    674,   # Edificio Velázquez
]


# Recintos de Concón: todos vinculados a su fila real en bbdd_sucursales
# (columna "MC ...") — pedido explícito, jul 2026.
_CONCON_RECINTOS_FIJOS: list[dict[str, object]] = []

# Orden de aparición en el registro del supervisor: sucursal_id en
# bbdd_sucursales, respetando el mismo orden que tenía la tabla original.
_CONCON_ORDEN: list[int | dict[str, object]] = [
    239,   # CONSISTORIAL -> MC Edificio Municipal
    458,   # PARQUE LA ISLA -> MC Parque La Isla
    195,   # APARCADERO -> MC Corrales Municipales
    705,   # JUZGADO DE POLICIA LOCAL -> MC Juzgado Policía Local
    218,   # DIDECO -> MC DIDECO
    779,   # SAR -> MC SAR
    142,   # CESFAM -> MC Cesfam
    75,    # CARPA CULTURAL -> MC Carpa (Avanzada Cultural)
    139,   # CENTRO COMUNITARIO -> MC CJAM - Centro de Juventud, Adulto Mayor y Discapacidad
    445,   # OPERACIONES Y DIMAO -> MC OPD
    707,   # PARQUE ALTA VISTA -> MC Parque Vista al mar
    704,   # TRANSITO Y FINANZAS -> MC Edificio Finanzas
    222,   # OBRAS PUBLICAS -> MC DOM - Direccion de Obras Municipales
    83,    # (nuevo) MC Biblioteca Municipal
    434,   # (nuevo) MC Museo
]


def _listar_recintos_privados(db: Session) -> list[dict[str, str | int]]:
    rows = (
        db.query(SucursalBBDD)
        .filter(SucursalBBDD.id.in_(_PRIVADOS_SUCURSAL_IDS))
        .order_by(SucursalBBDD.nombre_empresa.asc(), SucursalBBDD.nombre_sucursal.asc())
        .all()
    )
    return [
        {
            "id": row.id,
            "label": _recinto_label(row),
            "direccion": str(row.direccion_sucursal or "").strip(),
        }
        for row in rows
        if _recinto_label(row)
    ]


def _listar_recintos_concon(db: Session) -> list[dict[str, object]]:
    ids = [item for item in _CONCON_ORDEN if isinstance(item, int)]
    rows = db.query(SucursalBBDD).filter(SucursalBBDD.id.in_(ids)).all()
    por_id = {row.id: row for row in rows}
    recintos: list[dict[str, object]] = []
    for item in _CONCON_ORDEN:
        if isinstance(item, dict):
            recintos.append(dict(item))
        elif item in por_id:
            recintos.append({
                "id": por_id[item].id,
                "label": _recinto_label(por_id[item]),
                "direccion": str(por_id[item].direccion_sucursal or "").strip(),
            })
    return recintos


def _coords_recinto_estatico(label: object) -> tuple[float | None, float | None]:
    label_norm = _normalizar_texto(label)
    if not label_norm:
        return None, None
    for recinto in _CONCON_RECINTOS_FIJOS:
        if _normalizar_texto(recinto["label"]) == label_norm:
            return float(recinto["latitud"]), float(recinto["longitud"])
    return None, None


def _recintos_para_grupo(
    db: Session,
    grupo: str,
    registros_qr: list[InicioTurnoRegistro] | None = None,
    registros_sv: list[SupervisorRegistro] | None = None,
) -> tuple[str, list[str]]:
    grupo = str(grupo or "").lower().strip()
    registros_qr = registros_qr or []
    registros_sv = registros_sv or []

    if grupo == "privados":
        recintos = [str(s["label"]) for s in _listar_recintos_privados(db)]
        return "privados", recintos
    if grupo == "concon":
        recintos = [str(s["label"]) for s in _listar_recintos_concon(db)]
        return "concon", recintos

    qr_sucursales = _listar_recintos_qr(db)
    sucursal_labels = [str(s["label"]) for s in qr_sucursales]
    privados_labels = {str(s["label"]) for s in _listar_recintos_privados(db)}
    concon_labels = {str(s["label"]) for s in _listar_recintos_concon(db)}
    all_recintos = {str(r.recinto or "").strip() for r in registros_qr}
    all_recintos.update(str(r.recinto or "").strip() for r in registros_sv)
    extra_recintos = sorted(
        r for r in all_recintos
        if r and r not in sucursal_labels and r not in privados_labels and r not in concon_labels
    )
    return "quintero", sucursal_labels + extra_recintos


def _rut_lookup_guardias(
    db: Session,
    registros_qr: list[InicioTurnoRegistro],
    registros_sv: list[SupervisorRegistro],
) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for r in registros_qr:
        nombre_key = _normalizar_texto(r.nombre_guardia)
        rut = _normalizar_rut(r.rut)
        if nombre_key and rut:
            lookup.setdefault(nombre_key, rut)

    nombres_sv = {_normalizar_texto(r.nombre_guardia) for r in registros_sv if r.nombre_guardia}
    if nombres_sv:
        for user in _usuarios_guardia_activos(db):
            nombre_key = _normalizar_texto(user.name)
            rut = _normalizar_rut(user.username)
            if nombre_key in nombres_sv and rut:
                lookup.setdefault(nombre_key, rut)
        for guardia in db.query(InicioTurnoGuardia).all():
            nombre_key = _normalizar_texto(guardia.nombre)
            rut = _normalizar_rut(guardia.rut)
            if nombre_key in nombres_sv and rut:
                lookup.setdefault(nombre_key, rut)
    return lookup


def _agrupar_turnos_por_guardia(
    registros: list[InicioTurnoRegistro] | list[SupervisorRegistro],
    rut_lookup: dict[str, str],
    tipo_guardia_lookup: dict[str, dict[str, str]],
    *,
    usa_rut_registro: bool,
) -> list[dict]:
    rows: dict[str, dict] = {}
    # Farmacia Municipal (257) <-> DESAM (236): el mismo guardia de Farmacia
    # se traslada a cubrir DESAM de 17 a 20hrs — ambos marcajes se conservan
    # intactos en la BBDD (no se fusionan/archivan, siguen sirviendo para
    # pago/contabilidad de turnos), pero en ESTE informe (descarga y vista
    # previa de "tabla de asistencia") cuentan como 1 solo turno "Dia" por
    # guardia y dia, no 2 — pedido explicito, ago 2026.
    farmacia_desam_contados: set[tuple[str, object]] = set()
    for reg in registros:
        nombre = str(reg.nombre_guardia or "").strip()
        if not nombre:
            continue
        nombre_key = _normalizar_texto(nombre)
        rut = rut_lookup.get(nombre_key, "")
        if not rut and usa_rut_registro:
            rut = _normalizar_rut(getattr(reg, "rut", ""))
        key = rut or nombre_key
        if not key:
            continue
        row = rows.setdefault(
            key,
            {
                "nombre": nombre,
                "rut": rut,
                "tipo_guardia": _tipo_guardia_para(nombre, rut, tipo_guardia_lookup),
                "turnos": Counter(),
            },
        )
        if rut and not row["rut"]:
            row["rut"] = rut
        if not row.get("tipo_guardia"):
            row["tipo_guardia"] = _tipo_guardia_para(nombre, row["rut"], tipo_guardia_lookup)
        turno = _turno_informe(reg.tipo_turno)
        if turno == "Dia" and _es_recinto_farmacia_o_desam(getattr(reg, "sucursal_id", None), getattr(reg, "recinto", None)):
            dia_marcaje = getattr(reg, "registrado_at", None)
            dia_marcaje = dia_marcaje.date() if dia_marcaje is not None else getattr(reg, "fecha", None)
            dedup_key = (key, dia_marcaje)
            if dedup_key in farmacia_desam_contados:
                continue
            farmacia_desam_contados.add(dedup_key)
        if turno:
            row["turnos"][turno] += 1

    return sorted(rows.values(), key=lambda item: (item["nombre"].casefold(), item["rut"]))


def _clave_cruce_guardia(reg, rut_lookup: dict[str, str]) -> tuple:
    nombre = str(reg.nombre_guardia or "").strip()
    nombre_key = _normalizar_texto(nombre)
    rut = rut_lookup.get(nombre_key, "") or _normalizar_rut(getattr(reg, "rut", ""))
    identidad = rut or nombre_key
    fecha = reg.registrado_at.date() if hasattr(reg, "registrado_at") else reg.fecha
    return (
        identidad,
        fecha.isoformat(),
        _normalizar_texto(reg.recinto),
        _turno_informe(reg.tipo_turno),
    )


def _agrupar_cruce_guardias(
    registros_qr: list[InicioTurnoRegistro],
    registros_sv: list[SupervisorRegistro],
    rut_lookup: dict[str, str],
    tipo_guardia_lookup: dict[str, dict[str, str]],
) -> list[dict]:
    qr_counter = Counter(_clave_cruce_guardia(reg, rut_lookup) for reg in registros_qr)
    sv_counter = Counter(_clave_cruce_guardia(reg, rut_lookup) for reg in registros_sv)

    identidad_info: dict[str, dict] = {}
    for reg in list(registros_qr) + list(registros_sv):
        nombre = str(reg.nombre_guardia or "").strip()
        nombre_key = _normalizar_texto(nombre)
        rut = rut_lookup.get(nombre_key, "") or _normalizar_rut(getattr(reg, "rut", ""))
        identidad = rut or nombre_key
        if identidad and identidad not in identidad_info:
            identidad_info[identidad] = {
                "nombre": nombre,
                "rut": rut,
                "tipo_guardia": _tipo_guardia_para(nombre, rut, tipo_guardia_lookup),
            }

    rows: dict[str, dict] = {}
    for key in set(qr_counter) | set(sv_counter):
        identidad, _fecha, _recinto, turno = key
        if not identidad:
            continue
        row = rows.setdefault(
            identidad,
            {
                "nombre": identidad_info.get(identidad, {}).get("nombre", ""),
                "rut": identidad_info.get(identidad, {}).get("rut", ""),
                "tipo_guardia": identidad_info.get(identidad, {}).get("tipo_guardia", ""),
                "turnos": Counter(),
                "coincidencias": 0,
                "solo_registro": 0,
                "solo_supervisor": 0,
            },
        )
        coincidencias = min(qr_counter[key], sv_counter[key])
        solo_registro = max(qr_counter[key] - sv_counter[key], 0)
        solo_supervisor = max(sv_counter[key] - qr_counter[key], 0)

        if turno:
            row["turnos"][turno] += coincidencias
        row["coincidencias"] += coincidencias
        row["solo_registro"] += solo_registro
        row["solo_supervisor"] += solo_supervisor

    return sorted(rows.values(), key=lambda item: (item["nombre"].casefold(), item["rut"]))


def _permisos_sin_goce_lookup(db: Session, year: int, month: int) -> dict[str, list[tuple[date, date]]]:
    """RUT normalizado -> lista de rangos (desde, hasta) con 'Permiso sin Goce'
    vigente que se superponen con el mes del informe. Se usa para marcar, dentro
    de las faltas ya detectadas, cuales corresponden a un permiso sin goce
    autorizado (no un no-show real) con un distintivo aparte."""
    primer_dia = date(year, month, 1)
    ultimo_dia = date(year, month, calendar.monthrange(year, month)[1])
    filas = (
        db.query(GuardiaJustificacion)
        .filter(
            GuardiaJustificacion.motivo == "Permiso sin Goce",
            GuardiaJustificacion.fecha_desde.isnot(None),
            GuardiaJustificacion.fecha_hasta.isnot(None),
            GuardiaJustificacion.fecha_desde <= ultimo_dia,
            GuardiaJustificacion.fecha_hasta >= primer_dia,
        )
        .all()
    )
    lookup: dict[str, list[tuple[date, date]]] = defaultdict(list)
    for j in filas:
        lookup[_normalizar_rut(j.rut)].append((j.fecha_desde, j.fecha_hasta))
    return lookup


def _justificaciones_lookup(
    db: Session, year: int, month: int, tipo_guardia_lookup: dict[str, dict[str, str]]
) -> list[dict]:
    """Guardias con una justificación (Licencia Médica, Vacaciones, Permiso
    con/sin Goce, Falta con fecha, Desvinculado, Renuncia — ver
    MOTIVOS_JUSTIFICACION) vigente durante el mes del informe. Mismo criterio
    de superposición de fechas que _permisos_sin_goce_lookup, pero sin
    acotar a un solo motivo y devolviendo las filas completas para su propia
    sección del informe, en vez de solo un lookup rut->rangos."""
    primer_dia = date(year, month, 1)
    ultimo_dia = date(year, month, calendar.monthrange(year, month)[1])
    filas = (
        db.query(GuardiaJustificacion)
        .filter(
            GuardiaJustificacion.fecha_desde.isnot(None),
            GuardiaJustificacion.fecha_desde <= ultimo_dia,
            or_(
                GuardiaJustificacion.fecha_hasta.is_(None),
                GuardiaJustificacion.fecha_hasta >= primer_dia,
            ),
        )
        .order_by(GuardiaJustificacion.fecha_desde)
        .all()
    )
    rows = []
    for j in filas:
        nombre = str(j.nombre_guardia or "").strip()
        rut = _normalizar_rut(j.rut)
        rows.append(
            {
                "nombre": nombre,
                "rut": j.rut or "",
                "tipo_guardia": _tipo_guardia_para(nombre, rut, tipo_guardia_lookup),
                "motivo": j.motivo or "",
                "fecha_desde": j.fecha_desde,
                "fecha_hasta": j.fecha_hasta,
                "notas": j.notas or "",
            }
        )
    return sorted(rows, key=lambda item: (item["nombre"].casefold(), item["fecha_desde"] or date.min))


def _clave_falta_guardia(reg, rut_lookup: dict[str, str]) -> tuple:
    nombre = str(reg.nombre_guardia or "").strip()
    nombre_key = _normalizar_texto(nombre)
    rut = rut_lookup.get(nombre_key, "") or _normalizar_rut(getattr(reg, "rut", ""))
    identidad = rut or nombre_key
    fecha = reg.registrado_at.date() if hasattr(reg, "registrado_at") else reg.fecha
    return (
        identidad,
        fecha.isoformat(),
        _normalizar_texto(reg.recinto),
    )


def _agrupar_faltas_guardias(
    registros_qr: list[InicioTurnoRegistro],
    registros_sv: list[SupervisorRegistro],
    rut_lookup: dict[str, str],
    tipo_guardia_lookup: dict[str, dict[str, str]],
    permisos_sin_goce: dict[str, list[tuple[date, date]]] | None = None,
) -> list[dict]:
    permisos_sin_goce = permisos_sin_goce or {}
    qr_counter = Counter(_clave_falta_guardia(reg, rut_lookup) for reg in registros_qr)
    sv_counter = Counter(_clave_falta_guardia(reg, rut_lookup) for reg in registros_sv)

    rows: dict[str, dict] = {}
    for reg in registros_sv:
        nombre = str(reg.nombre_guardia or "").strip()
        nombre_key = _normalizar_texto(nombre)
        rut = rut_lookup.get(nombre_key, "")
        identidad = rut or nombre_key
        if not identidad:
            continue

        clave = _clave_falta_guardia(reg, rut_lookup)
        if qr_counter[clave] >= sv_counter[clave]:
            continue

        row = rows.setdefault(
            identidad,
            {
                "nombre": nombre,
                "rut": rut,
                "tipo_guardia": _tipo_guardia_para(nombre, rut, tipo_guardia_lookup),
                "faltas": 0,
                "detalle": [],
            },
        )
        row["faltas"] += 1
        fecha = reg.fecha
        turno = _turno_informe(reg.tipo_turno) or str(reg.tipo_turno or "").strip()
        rangos = permisos_sin_goce.get(_normalizar_rut(rut), [])
        es_permiso_sin_goce = any(desde <= fecha <= hasta for desde, hasta in rangos)
        texto = f"{fecha.day:02d}/{fecha.month:02d} · {turno} · {_recinto_display(reg.recinto)}"
        row["detalle"].append((fecha, texto, es_permiso_sin_goce))
        qr_counter[clave] += 1

    salida = []
    for row in rows.values():
        row["detalle"] = [(texto, es_permiso) for _, texto, es_permiso in sorted(row["detalle"], key=lambda t: t[0])]
        salida.append(row)
    return sorted(salida, key=lambda item: (-int(item["faltas"]), item["nombre"].casefold(), item["rut"]))


def _agrupar_domingos_guardias(
    registros_qr: list[InicioTurnoRegistro],
    rut_lookup: dict[str, str],
    tipo_guardia_lookup: dict[str, dict[str, str]],
) -> list[dict]:
    """Domingos trabajados por guardia, según su propio registro (QR) — a
    pedido del jefe de guardias, esto se saca del registro de guardia, no
    del registro de supervisor. Los turnos "Contrato Diario" no se cuentan
    (a pedido explícito): son gente externa contratada por día, no dotación
    regular a la que se le hace seguimiento de domingos trabajados."""
    rows: dict[str, dict] = {}
    for reg in registros_qr:
        fecha = reg.registrado_at.date()
        if fecha.weekday() != 6:  # domingo
            continue
        if _turno_informe(reg.tipo_turno) == "Contrato Diario":
            continue
        nombre = str(reg.nombre_guardia or "").strip()
        if not nombre:
            continue
        nombre_key = _normalizar_texto(nombre)
        rut = rut_lookup.get(nombre_key, "") or _normalizar_rut(reg.rut)
        identidad = rut or nombre_key
        if not identidad:
            continue

        row = rows.setdefault(
            identidad,
            {
                "nombre": nombre,
                "rut": rut,
                "tipo_guardia": _tipo_guardia_para(nombre, rut, tipo_guardia_lookup),
                "fechas": set(),
                "detalle": [],
            },
        )
        turno = _turno_informe(reg.tipo_turno) or str(reg.tipo_turno or "").strip()
        texto = f"{fecha.day:02d}/{fecha.month:02d} · {turno} · {_recinto_display(reg.recinto)}"
        row["fechas"].add(fecha)
        row["detalle"].append((fecha, texto))

    salida = []
    for row in rows.values():
        row["domingos"] = len(row.pop("fechas"))
        row["detalle"] = [texto for _, texto in sorted(row["detalle"], key=lambda t: t[0])]
        salida.append(row)
    return sorted(salida, key=lambda item: (-int(item["domingos"]), item["nombre"].casefold(), item["rut"]))


_FALTA_DIA_RE = re.compile(r"^(\d{2})/(\d{2})")


def _recinto_permiso_sin_goce_info(
    db: Session,
    nombre_guardia: str,
    dia: date,
    primer_dia: date,
    ultimo_dia: date,
) -> tuple[str, list[str]]:
    nombre_norm = _normalizar_texto(nombre_guardia)
    if not nombre_norm:
        return "Lugar no informado", []

    margen = timedelta(days=45)
    registros = (
        db.query(
            SupervisorRegistro.fecha,
            SupervisorRegistro.nombre_guardia,
            SupervisorRegistro.tipo_turno,
            SupervisorRegistro.recinto,
            SupervisorRegistro.notas,
        )
        .filter(
            SupervisorRegistro.fecha >= primer_dia - margen,
            SupervisorRegistro.fecha <= ultimo_dia + margen,
        )
        .all()
    )
    candidatos_por_nota = [
        r
        for r in registros
        if _normalizar_texto(r.notas).find(nombre_norm) >= 0 and str(r.recinto or "").strip()
    ]
    candidatos_por_nombre = [
        r
        for r in registros
        if _normalizar_texto(r.nombre_guardia) == nombre_norm and str(r.recinto or "").strip()
    ]
    candidatos = candidatos_por_nota or candidatos_por_nombre
    if not candidatos:
        return "Lugar no informado", []
    usar_turno = not bool(candidatos_por_nota)

    distancia_min = min(abs((r.fecha - dia).days) for r in candidatos)
    cercanos = [r for r in candidatos if abs((r.fecha - dia).days) == distancia_min]
    recintos: list[str] = []
    turnos: list[str] = []
    for reg in sorted(cercanos, key=lambda r: (_recinto_display(r.recinto).casefold(), str(r.tipo_turno or "").casefold())):
        recinto = _recinto_display(reg.recinto)
        if recinto and recinto not in recintos:
            recintos.append(recinto)
        turno = _turno_informe(reg.tipo_turno) or str(reg.tipo_turno or "").strip()
        if usar_turno and turno and turno not in turnos:
            turnos.append(turno)

    turno_txt = " / ".join(turnos)
    recinto_txt = " / ".join(recintos) if recintos else "Lugar no informado"
    texto = f"{turno_txt} · {recinto_txt}" if turno_txt else recinto_txt
    return texto, recintos


def _agregar_faltas_por_permiso_sin_goce(
    db: Session,
    rows_faltas: list[dict],
    rut_lookup: dict[str, str],
    tipo_guardia_lookup: dict[str, dict[str, str]],
    year: int,
    month: int,
    fecha_limite: date,
    recintos_permitidos: set[str] | None = None,
) -> list[dict]:
    """Un guardia con "Permiso sin Goce" vigente no queda agendado en ningun
    lado, asi que el cruce QR/supervisor de _agrupar_faltas_guardias nunca lo
    detecta (no hay registro de supervisor del que "falte" el QR). Esta
    funcion agrega un dia de falta por cada dia de permiso vigente dentro del
    periodo del informe, aunque el guardia no tenga ningun registro."""
    primer_dia = date(year, month, 1)
    ultimo_dia = min(date(year, month, calendar.monthrange(year, month)[1]), fecha_limite)
    if ultimo_dia < primer_dia:
        return rows_faltas

    permisos = (
        db.query(GuardiaJustificacion)
        .filter(
            GuardiaJustificacion.motivo == "Permiso sin Goce",
            GuardiaJustificacion.fecha_desde.isnot(None),
            GuardiaJustificacion.fecha_hasta.isnot(None),
            GuardiaJustificacion.fecha_desde <= ultimo_dia,
            GuardiaJustificacion.fecha_hasta >= primer_dia,
        )
        .all()
    )
    if not permisos:
        return rows_faltas

    recintos_permitidos_norm: set[str] = set()
    for recinto in recintos_permitidos or set():
        raw = str(recinto or "").strip()
        if raw:
            recintos_permitidos_norm.add(_normalizar_texto(raw))
            recintos_permitidos_norm.add(_normalizar_texto(_recinto_display(raw)))

    def _parse_dia(texto: str) -> date | None:
        m = _FALTA_DIA_RE.match(texto)
        if not m:
            return None
        try:
            return date(year, int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None

    rows_por_identidad: dict[str, dict] = {}
    dias_por_identidad: dict[str, dict[date, tuple[str, bool]]] = defaultdict(dict)
    for row in rows_faltas:
        identidad = _normalizar_rut(row.get("rut") or "") or _normalizar_texto(row.get("nombre") or "")
        rows_por_identidad[identidad] = row
        for texto, es_permiso in row.get("detalle", []):
            dia = _parse_dia(texto)
            if dia:
                dias_por_identidad[identidad][dia] = (texto, es_permiso)
    identidades_en_salida = set(rows_por_identidad)

    tocados: set[str] = set()
    for permiso in permisos:
        rut = _normalizar_rut(permiso.rut)
        nombre = str(permiso.nombre_guardia or "").strip()
        if not rut and not nombre:
            continue
        nombre_key = _normalizar_texto(nombre)
        rut_final = rut or rut_lookup.get(nombre_key, "")
        identidad = rut_final or nombre_key
        if not identidad:
            continue

        desde = max(permiso.fecha_desde, primer_dia)
        hasta = min(permiso.fecha_hasta, ultimo_dia)
        if desde > hasta:
            continue

        row = rows_por_identidad.get(identidad)
        if row is None:
            row = {
                "nombre": nombre,
                "rut": rut_final,
                "tipo_guardia": _tipo_guardia_para(nombre, rut_final, tipo_guardia_lookup),
                "faltas": 0,
                "detalle": [],
            }
            rows_por_identidad[identidad] = row

        dia = desde
        while dia <= hasta:
            if dia not in dias_por_identidad[identidad]:
                lugar, recintos_permiso = _recinto_permiso_sin_goce_info(db, nombre, dia, primer_dia, ultimo_dia)
                if recintos_permitidos_norm:
                    recintos_permiso_norm: set[str] = set()
                    for recinto_permiso in recintos_permiso:
                        raw = str(recinto_permiso or "").strip()
                        if raw:
                            recintos_permiso_norm.add(_normalizar_texto(raw))
                            recintos_permiso_norm.add(_normalizar_texto(_recinto_display(raw)))
                    if not recintos_permiso_norm or not (recintos_permiso_norm & recintos_permitidos_norm):
                        dia += timedelta(days=1)
                        continue
                texto = f"{dia.day:02d}/{dia.month:02d} · {lugar}"
                dias_por_identidad[identidad][dia] = (texto, True)
                if identidad not in identidades_en_salida:
                    rows_faltas.append(row)
                    identidades_en_salida.add(identidad)
                tocados.add(identidad)
            dia += timedelta(days=1)

    for identidad in tocados:
        row = rows_por_identidad[identidad]
        dias = dias_por_identidad[identidad]
        row["detalle"] = [dias[d] for d in sorted(dias)]
        row["faltas"] = len(row["detalle"])

    return sorted(rows_faltas, key=lambda item: (-int(item["faltas"]), item["nombre"].casefold(), item["rut"]))


def _alertas_permiso_sin_goce(rows_faltas: list[dict], month: int) -> list[dict]:
    alertas: list[dict] = []
    for row in rows_faltas:
        nombre = str(row.get("nombre") or "").strip()
        for texto, es_permiso in row.get("detalle", []):
            if not es_permiso:
                continue
            dia = ""
            recinto = "Lugar no informado"
            partes = str(texto or "").split(" · ", 1)
            if partes:
                dia = partes[0].strip()
            if len(partes) > 1 and partes[1].strip():
                recinto = partes[1].strip()
            alertas.append({
                "fecha": dia or f"--/{month:02d}",
                "recinto": recinto,
                "detalle": nombre,
                "permiso_sin_goce": True,
            })
    return sorted(alertas, key=lambda x: (x["fecha"], x["recinto"], x["detalle"]))


def _recinto_display(valor: object) -> str:
    """Solo el nombre de la sucursal (último segmento de 'Empresa - Sucursal')."""
    txt = str(valor or "").strip()
    return txt.split(" - ")[-1].strip() or txt


def _agrupar_turnos_extra(
    registros_qr: list[InicioTurnoRegistro],
    registros_sv: list[SupervisorRegistro],
    rut_lookup: dict[str, str],
    tipo_guardia_lookup: dict[str, dict[str, str]],
    *,
    tipo_turno: str = "Extra",
) -> list[dict]:
    """Turnos Extra (o Contrato Diario) por guardia, cruzando registro de
    guardia (QR) y supervisor.

    Evento = (guardia, fecha, recinto). Se cuenta el turno aunque figure en una
    sola fuente, pero en ese caso se marca la inconsistencia indicando en qué
    registro falta.
    """
    def _identidad(reg) -> tuple[str, str, str]:
        nombre = str(reg.nombre_guardia or "").strip()
        nombre_key = _normalizar_texto(nombre)
        rut = rut_lookup.get(nombre_key, "") or _normalizar_rut(getattr(reg, "rut", ""))
        return (rut or nombre_key, nombre, rut)

    qr_eventos: Counter = Counter()
    sv_eventos: Counter = Counter()
    labels: dict[tuple, tuple] = {}
    identidad_info: dict[str, dict] = {}

    def _procesar(registros, contador, es_qr: bool):
        for reg in registros:
            if _turno_informe(reg.tipo_turno) != tipo_turno:
                continue
            identidad, nombre, rut = _identidad(reg)
            if not identidad:
                continue
            fecha = reg.registrado_at.date() if es_qr else reg.fecha
            recinto = str(reg.recinto or "").strip()
            key = (identidad, fecha.isoformat(), _normalizar_texto(recinto))
            contador[key] += 1
            labels.setdefault(key, (fecha, recinto))
            identidad_info.setdefault(identidad, {
                "nombre": nombre,
                "rut": rut,
                "tipo_guardia": _tipo_guardia_para(nombre, rut, tipo_guardia_lookup),
            })

    _procesar(registros_qr, qr_eventos, True)
    _procesar(registros_sv, sv_eventos, False)

    rows: dict[str, dict] = {}
    for key in set(qr_eventos) | set(sv_eventos):
        identidad = key[0]
        fecha, recinto = labels[key]
        n_qr, n_sv = qr_eventos[key], sv_eventos[key]
        row = rows.setdefault(identidad, {
            **identidad_info.get(identidad, {"nombre": "", "rut": "", "tipo_guardia": ""}),
            "total": 0,
            "lugares": [],       # (fecha, "dd/mm · recinto")
            "inconsistencias": [],
        })
        row["total"] += max(n_qr, n_sv)
        etiqueta = f"{fecha.day:02d}/{fecha.month:02d} · {_recinto_display(recinto)}"
        row["lugares"].append((fecha, etiqueta))
        if n_qr > n_sv:
            row["inconsistencias"].append((fecha, f"{etiqueta}: falta en registro del supervisor"))
        elif n_sv > n_qr:
            row["inconsistencias"].append((fecha, f"{etiqueta}: falta en registro de guardia (QR)"))

    salida = []
    for row in rows.values():
        row["lugares"] = [texto for _, texto in sorted(row["lugares"])]
        row["inconsistencias"] = [texto for _, texto in sorted(row["inconsistencias"])]
        salida.append(row)
    return sorted(salida, key=lambda item: (item["nombre"].casefold(), item["rut"]))


def _crear_hoja_turno_extra(ws, periodo: str, rows: list[dict], *, tipo_turno: str = "Extra") -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    titulo_hoja = f"Turnos {tipo_turno}"
    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = ["Nombre", "RUT", "Tipo de Contrato", f"Turnos {tipo_turno}", "Dónde (fecha · recinto)", "Consistencia", "Detalle inconsistencia"]

    ws.append([f"ATC - {titulo_hoja}"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for row in rows:
        inconsistencias = row["inconsistencias"]
        ws.append([
            row["nombre"],
            row["rut"],
            row.get("tipo_guardia", ""),
            int(row["total"]),
            "\n".join(row["lugares"]),
            "OK" if not inconsistencias else "Con inconsistencias",
            "\n".join(inconsistencias) if inconsistencias else "Registro de guardia y supervisor coinciden",
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column in (3, 4, 6):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif cell.column in (5, 7):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        estado = fila[5].value
        if estado == "Con inconsistencias":
            fila[5].font = Font(bold=True, color="B91C1C")
            fila[5].fill = PatternFill("solid", fgColor="FDF2F2")

    widths = [34, 16, 16, 13, 44, 20, 52]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _crear_hoja_justificaciones(ws, periodo: str, rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = ["Nombre", "RUT", "Tipo de Contrato", "Motivo", "Desde", "Hasta", "Notas"]

    ws.append(["ATC - Justificaciones"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for row in rows:
        desde = row["fecha_desde"].strftime("%d/%m/%Y") if row["fecha_desde"] else "—"
        hasta = row["fecha_hasta"].strftime("%d/%m/%Y") if row["fecha_hasta"] else "En curso"
        ws.append([
            row["nombre"],
            row["rut"],
            row.get("tipo_guardia", ""),
            row["motivo"],
            desde,
            hasta,
            row["notas"],
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column in (3, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif cell.column == 7:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    widths = [34, 16, 16, 20, 14, 14, 44]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _crear_hoja_supervisor_calendario(
    ws,
    periodo: str,
    month: int,
    recintos: list[str],
    days_info: list[dict],
    matrix_sv: dict,
    *,
    titulo: str = "Registro Supervisor",
    solo_nombres: bool = False,
) -> None:
    """Hoja calendario (recinto x fecha) solo con lo que registra el
    supervisor — misma vista que 'Cruce vs Supervisor' en pantalla, pero
    excluyendo el bloque de Registro Guardia (QR). Cada nombre se colorea
    segun el tipo de turno, igual que los chips en la tabla web.

    Nota: se usa UNA FILA POR PERSONA (en vez de varios nombres con distinto
    color dentro de una sola celda) porque Apple Numbers no respeta el color
    por-run de los "rich text" de OOXML — con una fila por persona el color
    es a nivel de celda completa, que sí es compatible con Numbers/Excel/
    Google Sheets/LibreOffice por igual.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    n_dias = len(days_info)
    n_cols = 1 + n_dias

    # Color de RELLENO de la celda por tipo de turno (no de la letra). Los
    # tonos se separan bien entre si (Dia=azul, Noche=violeta) para que no
    # se confundan como pasaba con los textos --dia-text/--noche-text del HTML.
    fill_turno = {
        "dia": "CFE0F3",
        "noche": "E6D6F0",
        "extra": "F5E6C4",
        "contrato diario": "D3EEDA",
    }
    texto_fila = "1F2937"

    def _fill_turno(turno: str) -> str:
        return fill_turno.get(str(turno or "").strip().lower(), "E5E7EB")

    ws.append([f"ATC - {titulo} (por recinto y fecha)"])
    ws.append([periodo])
    ws.append([])
    ws.append(["Recinto"] + [f"{d['day']:02d}/{month:02d} {d['dow']}" for d in days_info])

    for recinto in recintos:
        entradas_por_dia = {
            d["day"]: [
                e for e in matrix_sv.get(recinto, {}).get(d["day"], [])
                if str(e.get("nombre") or "").strip()
            ]
            for d in days_info
        }
        max_filas = max([1, *(len(v) for v in entradas_por_dia.values())])
        fila_inicio = ws.max_row + 1

        for sub in range(max_filas):
            row_idx = fila_inicio + sub
            for col_offset, d in enumerate(days_info, start=2):
                entradas = entradas_por_dia[d["day"]]
                if sub >= len(entradas):
                    continue
                e = entradas[sub]
                nombre = str(e.get("nombre") or "").strip()
                turno = str(e.get("turno") or "").strip()
                notas = str(e.get("notas") or "").strip()
                valor = nombre if solo_nombres else (f"{nombre}\n({notas})" if notas else nombre)
                cell = ws.cell(row=row_idx, column=col_offset, value=valor)
                if solo_nombres:
                    cell.fill = PatternFill("solid", fgColor=_fill_turno(turno))
                    cell.font = Font(bold=True, color=texto_fila)
                    cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
                elif notas:
                    cell.fill = PatternFill("solid", fgColor=_fill_turno(turno))
                    cell.font = Font(bold=True, color=texto_fila, size=9)
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                    altura_actual = ws.row_dimensions[row_idx].height or 15
                    ws.row_dimensions[row_idx].height = max(altura_actual, 60)
                else:
                    cell.fill = PatternFill("solid", fgColor=_fill_turno(turno))
                    cell.font = Font(bold=True, color=texto_fila)

        ws.cell(row=fila_inicio, column=1, value=recinto)
        if max_filas > 1:
            ws.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio + max_filas - 1, end_column=1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=n_cols):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column == 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.font is None or not cell.font.b:
                    cell.font = Font(bold=True)
            elif cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")

    ws.column_dimensions["A"].width = 30
    for idx in range(2, n_cols + 1):
        col_letter = ws.cell(row=4, column=idx).column_letter
        ws.column_dimensions[col_letter].width = 22
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = "B5"


def _crear_hoja_informe_turnos(
    ws,
    titulo: str,
    periodo: str,
    rows: list[dict],
    *,
    faltas_por_guardia: dict[str, int] | None = None,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    if faltas_por_guardia is not None:
        headers = ["Nombre", "RUT", "Tipo de Contrato", "Turno Normal", "Extra", "Contrato Diario", "Faltas"]
        widths = [34, 16, 14, 16, 10, 18, 12]
    else:
        headers = ["Nombre", "RUT", "Tipo de Contrato", "Dia", "Noche", "Turno Normal", "Extra", "Contrato Diario", "Total Turnos"]
        widths = [34, 16, 14, 10, 10, 16, 10, 18, 14]

    ws.append(["ATC - " + titulo])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for row in rows:
        turnos = row["turnos"]
        dia = int(turnos["Dia"])
        noche = int(turnos["Noche"])
        extra = int(turnos["Extra"])
        contrato = int(turnos["Contrato Diario"])
        if faltas_por_guardia is not None:
            identidad = row["rut"] or _normalizar_texto(row["nombre"])
            faltas = int(faltas_por_guardia.get(identidad, 0))
            ws.append([row["nombre"], row["rut"], row.get("tipo_guardia", ""), dia + noche, extra, contrato, faltas])
        else:
            total = dia + noche + extra + contrato
            ws.append([row["nombre"], row["rut"], row.get("tipo_guardia", ""), dia, noche, dia + noche, extra, contrato, total])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column >= 3:
                cell.alignment = Alignment(horizontal="center")

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _crear_hoja_cruce(ws, periodo: str, rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = [
        "Nombre",
        "RUT",
        "Tipo de Contrato",
        "Dia",
        "Noche",
        "Turno Normal",
        "Extra",
        "Contrato Diario",
        "Coincidencias",
        "Disyuntivas",
        "Resultado",
        "Detalle",
    ]

    ws.append(["ATC - Cruce Registro Guardia vs Supervisor"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for row in rows:
        turnos = row["turnos"]
        dia = int(turnos["Dia"])
        noche = int(turnos["Noche"])
        extra = int(turnos["Extra"])
        contrato = int(turnos["Contrato Diario"])
        disyuntivas = int(row["solo_registro"]) + int(row["solo_supervisor"])
        resultado = "Coincide" if disyuntivas == 0 else "Con disyuntivas"
        detalle = "Coincidencia completa"
        if disyuntivas:
            detalle = f"Solo registro guardia: {row['solo_registro']} | Solo supervisor: {row['solo_supervisor']}"
        ws.append([
            row["nombre"],
            row["rut"],
            row.get("tipo_guardia", ""),
            dia,
            noche,
            dia + noche,
            extra,
            contrato,
            row["coincidencias"],
            disyuntivas,
            resultado,
            detalle,
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if 3 <= cell.column <= 10:
                cell.alignment = Alignment(horizontal="center")

    widths = [34, 16, 14, 10, 10, 16, 10, 18, 14, 12, 18, 46]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _crear_hoja_faltas(ws, periodo: str, rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = ["Nombre", "RUT", "Tipo de Contrato", "Nro de faltas", "Cuándo y dónde tenía que estar"]

    ws.append(["ATC - Faltas"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for row in rows:
        lineas = [
            f"{texto} (Permiso sin Goce)" if es_permiso else texto
            for texto, es_permiso in row["detalle"]
        ]
        ws.append([
            row["nombre"],
            row["rut"],
            row.get("tipo_guardia", ""),
            int(row["faltas"]),
            "\n".join(lineas),
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column in (3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif cell.column == 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        fila[3].font = Font(bold=True, color="B91C1C")

    widths = [34, 16, 16, 14, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _crear_hoja_domingos(ws, periodo: str, rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = ["Nombre", "RUT", "Tipo de Contrato", "Domingos trabajados", "Fecha y lugar"]

    ws.append(["ATC - Conteo días domingo"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for row in rows:
        ws.append([
            row["nombre"],
            row["rut"],
            row.get("tipo_guardia", ""),
            int(row["domingos"]),
            "\n".join(row["detalle"]),
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column in (3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif cell.column == 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
        fila[3].font = Font(bold=True, color="166534")

    widths = [34, 16, 16, 18, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


def _recinto_qr_token(recinto_id: object) -> str:
    secret = str(settings.JWT_SECRET or "inicio-turno").encode("utf-8")
    message = f"inicio-turno-recinto:{recinto_id}".encode("utf-8")
    return hmac.new(secret, message, hashlib.sha256).hexdigest()[:24]


def _inicio_turno_base_url(request: Request) -> str:
    configured = str(settings.INICIO_TURNO_PUBLIC_BASE_URL or "").strip().rstrip("/")
    if configured:
        return configured
    return str(request.base_url).rstrip("/")


def _resolver_recinto_por_qr(db: Session, qr: str) -> SucursalBBDD | dict[str, object] | None:
    token = str(qr or "").strip()
    if not token:
        return None
    all_recintos = _listar_recintos_qr(db) + _listar_recintos_privados(db) + _listar_recintos_concon(db)
    for recinto in all_recintos:
        if hmac.compare_digest(_recinto_qr_token(recinto["id"]), token):
            recinto_id = recinto["id"]
            if str(recinto_id).isdigit():
                return db.get(SucursalBBDD, int(recinto_id))
            return recinto

    generado = db.query(RecintoQrGenerado).filter(RecintoQrGenerado.token == token).first()
    if generado:
        recinto_id = generado.recinto_id
        if str(recinto_id).isdigit():
            sucursal = db.get(SucursalBBDD, int(recinto_id))
            if sucursal:
                return sucursal
        for recinto in all_recintos:
            if str(recinto["id"]) == str(recinto_id):
                return recinto
        return {"id": recinto_id, "label": generado.recinto_label, "direccion": ""}
    return None


def _listar_recintos_para_generador(db: Session) -> list[dict[str, object]]:
    combinados: list[dict[str, object]] = []
    for recinto in _listar_recintos_qr(db):
        combinados.append({**recinto, "grupo": "quintero", "grupo_label": "Quintero"})
    for recinto in _listar_recintos_privados(db):
        combinados.append({**recinto, "grupo": "privados", "grupo_label": "Privados"})
    for recinto in _listar_recintos_concon(db):
        combinados.append({**recinto, "grupo": "concon", "grupo_label": "Concón"})
    return combinados


_PREFIJOS_QUINTERO = [
    "municipalidad de quintero - mquin ",
    "municipalidad de quintero - ",
    "municipalidad de quintero -",
    "mquin ",
]


def _limpiar_label_quintero(label: str) -> str:
    limpio = str(label or "").strip()
    plano = _normalizar_texto(limpio)
    for prefijo in _PREFIJOS_QUINTERO:
        if plano.startswith(prefijo):
            return limpio[len(prefijo):].strip() or limpio
    return limpio


def _limpiar_label_privado(label: str) -> str:
    """Los recintos privados vienen como 'Razon Social - Nombre Sucursal';
    nos quedamos solo con lo que hay despues del ultimo ' - '."""
    limpio = str(label or "").strip()
    partes = [p.strip() for p in limpio.split(" - ") if p.strip()]
    return partes[-1] if partes else limpio


def _buscar_guardia_por_rut(
    db: Session,
    rut: str,
    sucursal_id: int | None = None,
) -> InicioTurnoGuardia | None:
    rut_norm = _normalizar_rut(rut)
    if not rut_norm:
        return None
    user = _usuario_guardia_activo_por_rut(db, rut_norm)
    if user:
        return _guardia_desde_usuario(user)

    for guardia in db.query(InicioTurnoGuardia).all():
        if _normalizar_rut(guardia.rut) == rut_norm:
            return guardia
    for guardia in _buscar_guardias_legacy(db):
        if _normalizar_rut(guardia.rut) == rut_norm:
            return guardia
    return None


def _buscar_guardia_por_nombre(db: Session, nombre: str) -> InicioTurnoGuardia | None:
    nombre_norm = _normalizar_texto(nombre)
    if not nombre_norm:
        return None
    user = _usuario_guardia_activo_por_nombre(db, nombre)
    if user:
        return _guardia_desde_usuario(user)

    for guardia in db.query(InicioTurnoGuardia).all():
        if _normalizar_texto(guardia.nombre) == nombre_norm:
            return guardia

    for guardia in _buscar_guardias_legacy(db, q=nombre, limit=20):
        if _normalizar_texto(guardia.nombre) == nombre_norm:
            return guardia
    return None


def _schemas_con_tabla(db: Session, table_name: str) -> list[str]:
    try:
        rows = db.execute(
            text(
                """
                SELECT DISTINCT table_schema
                FROM information_schema.columns
                WHERE table_name = :table_name
                  AND table_schema NOT IN ('pg_catalog', 'information_schema')
                ORDER BY table_schema
                """
            ),
            {"table_name": table_name},
        ).all()
        return [str(row[0]).strip() for row in rows if row and row[0]]
    except Exception:
        db.rollback()
        return []


def _buscar_guardias_legacy(db: Session, q: str = "", limit: int = 0) -> list[InicioTurnoGuardia]:
    query_norm = _normalizar_texto(q)
    query_rut = _normalizar_rut(q)
    guardias_sql = []
    for user in _usuarios_guardia_activos(db):
        nombre = str(user.name or "").strip()
        rut = str(user.username or "").strip()
        if query_norm and query_norm not in _normalizar_texto(nombre) and query_rut not in _normalizar_rut(rut):
            continue
        guardia = _guardia_desde_usuario(user)
        if not _guardia_oculto_en_opciones(guardia):
            guardias_sql.append(guardia)
        if limit > 0 and len(guardias_sql) >= limit:
            return guardias_sql
    if guardias_sql or query_norm:
        return guardias_sql

    guardias: list[InicioTurnoGuardia] = []
    query = str(q or "").strip()
    top_clause = f"TOP ({max(1, min(int(limit), 100))})" if limit > 0 else ""
    for table_name in ("bbdd_guardias", "inicio_turno_guardias"):
        for schema in _schemas_con_tabla(db, table_name):
            try:
                where = ""
                params: dict[str, object] = {}
                if query:
                    where = "WHERE LOWER(CAST(nombre AS NVARCHAR(MAX))) LIKE LOWER(:q) OR LOWER(CAST(rut AS NVARCHAR(MAX))) LIKE LOWER(:q)"
                    params["q"] = f"%{query}%"
                rows = db.execute(
                    text(
                        f"""
                        SELECT {top_clause} id, rut, nombre
                        FROM "{schema}"."{table_name}"
                        {where}
                        ORDER BY nombre
                        """
                    ),
                    params,
                ).mappings().all()
                for row in rows:
                    guardia = InicioTurnoGuardia(
                        id=int(row.get("id") or 0),
                        rut=str(row.get("rut") or "").strip(),
                        nombre=str(row.get("nombre") or "").strip(),
                    )
                    if not _guardia_oculto_en_opciones(guardia):
                        guardias.append(guardia)
            except Exception:
                db.rollback()
    return guardias


def seed_default_inicio_turno_guardias(db: Session) -> None:
    for item in DEFAULT_GUARDIAS:
        rut = _normalizar_rut(item["rut"])
        existing = _buscar_guardia_por_rut(db, rut)
        if existing:
            if str(existing.nombre or "").strip() != item["nombre"]:
                existing.nombre = item["nombre"]
            continue
        db.add(InicioTurnoGuardia(rut=rut, nombre=item["nombre"]))
    db.commit()


@router.get("/inicio-turno", response_class=HTMLResponse)
def inicio_turno_page(
    request: Request,
    qr: str = Query(default=""),
    recinto_id: int | None = Query(default=None),
    recinto: str = Query(default=""),
    db: Session = Depends(get_db),
):
    recinto_resuelto = _resolver_recinto_por_qr(db, qr) if qr else None
    sucursal = recinto_resuelto if isinstance(recinto_resuelto, SucursalBBDD) else None
    recinto_estatico = recinto_resuelto if isinstance(recinto_resuelto, dict) else None
    if not sucursal and recinto_id:
        sucursal = db.get(SucursalBBDD, recinto_id)
    recinto_qr = (
        str(recinto_estatico.get("label") or "").strip()
        if recinto_estatico
        else _recinto_label(sucursal) if sucursal else str(recinto or "").strip()
    )
    return templates.TemplateResponse(
        request,
        "inicio_turno.html",
        {
            "request": request,
            "recinto_id": sucursal.id if sucursal else None,
            "recinto_qr": recinto_qr,
            "recintos": _listar_recintos(db) if not recinto_qr else [],
            "tipos_turno": TIPOS_TURNO,
        },
    )


@router.get("/inicio-turno/qr-recintos", response_class=HTMLResponse)
def inicio_turno_qr_recintos_page(
    request: Request,
    grupo: str = Query(default="quintero"),
    db: Session = Depends(get_db),
):
    grupo = (grupo or "quintero").strip().lower()
    if grupo not in ("quintero", "privados", "concon"):
        grupo = "quintero"

    if grupo == "privados":
        recintos = _listar_recintos_privados(db)
        titulo = "QR Privados"
    elif grupo == "concon":
        recintos = _listar_recintos_concon(db)
        titulo = "QR Concón"
    else:
        recintos = _listar_recintos_qr(db)
        titulo = "QR Quintero"

    base_url = _inicio_turno_base_url(request)
    items = []
    for recinto in recintos:
        target = f"{base_url}/inicio-turno?{urlencode({'qr': _recinto_qr_token(recinto['id'])})}"
        items.append({**recinto, "target": target})
    return templates.TemplateResponse(
        request,
        "inicio_turno_qr_recintos.html",
        {
            "request": request,
            "recintos": items,
            "titulo": titulo,
            "grupo": grupo,
        },
    )


@router.get("/inicio-turno/qr-recintos-privados", response_class=HTMLResponse)
def inicio_turno_qr_recintos_privados_page(request: Request):
    return RedirectResponse(url="/inicio-turno/qr-recintos?grupo=privados", status_code=307)


@router.get("/inicio-turno/qr-recintos-concon", response_class=HTMLResponse)
def inicio_turno_qr_recintos_concon_page(request: Request):
    return RedirectResponse(url="/inicio-turno/qr-recintos?grupo=concon", status_code=307)


# ── Generador de QR por recinto (historial persistente, sin límite) ────────────

@router.get("/inicio-turno/generador-qr", response_class=HTMLResponse)
def inicio_turno_generador_qr_page(
    request: Request,
    db: Session = Depends(get_db),
):
    return templates.TemplateResponse(
        request,
        "inicio_turno_generador_qr.html",
        {
            "request": request,
            "recintos": _listar_recintos_para_generador(db),
        },
    )


class QrGeneradoCreate(BaseModel):
    recinto_id: str = Field(min_length=1, max_length=80)
    recinto_label: str = Field(min_length=1, max_length=255)


def _qr_generado_dict(request: Request, item: RecintoQrGenerado) -> dict[str, object]:
    base_url = _inicio_turno_base_url(request)
    target = f"{base_url}/inicio-turno/ronda?{urlencode({'qr': item.token})}"
    return {
        "id": item.id,
        "recinto_id": item.recinto_id,
        "recinto_label": item.recinto_label,
        "token": item.token,
        "numero": item.numero,
        "verificador": item.verificador,
        "target": target,
        "created_at": item.created_at.strftime("%d-%m-%Y %H:%M") if item.created_at else "",
    }


@router.get("/api/inicio-turno/qr-generados")
def listar_qr_generados(
    request: Request,
    recinto_id: str = Query(default=""),
    db: Session = Depends(get_db),
):
    recinto_id = str(recinto_id or "").strip()
    if not recinto_id:
        raise HTTPException(status_code=422, detail="Falta recinto_id")
    items = (
        db.query(RecintoQrGenerado)
        .filter(RecintoQrGenerado.recinto_id == recinto_id)
        .order_by(RecintoQrGenerado.created_at.desc(), RecintoQrGenerado.id.desc())
        .all()
    )
    return {"items": [_qr_generado_dict(request, item) for item in items]}


def _generar_codigo_verificador(db: Session) -> int:
    """Código de 6 dígitos para la franja vertical junto al QR: los 6
    dígitos son todos distintos entre sí y el número completo es único en
    toda la tabla — sirve para el ingreso manual cuando el QR no lee (un
    solo campo, sin tener que elegir antes el recinto, porque es único en
    todo el sistema). Es independiente del N° de ronda (secuencial por
    recinto, 1/2/3/4…) que se sigue mostrando bajo el QR."""
    existentes = {n for (n,) in db.query(RecintoQrGenerado.verificador).all() if n is not None}
    for _ in range(500):
        primero = random.randint(1, 9)
        resto = random.sample([d for d in range(10) if d != primero], 5)
        codigo = int("".join(str(d) for d in [primero, *resto]))
        if codigo not in existentes:
            return codigo
    raise HTTPException(status_code=500, detail="No se pudo generar un código de verificación único.")


@router.post("/api/inicio-turno/qr-generados")
def crear_qr_generado(
    payload: QrGeneradoCreate,
    request: Request,
    db: Session = Depends(get_db),
):
    recinto_id = payload.recinto_id.strip()
    ultimo_numero = (
        db.query(func.max(RecintoQrGenerado.numero))
        .filter(RecintoQrGenerado.recinto_id == recinto_id)
        .scalar()
    )
    token = secrets.token_hex(16)
    item = RecintoQrGenerado(
        recinto_id=recinto_id,
        recinto_label=payload.recinto_label.strip(),
        token=token,
        numero=(ultimo_numero or 0) + 1,
        verificador=_generar_codigo_verificador(db),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _qr_generado_dict(request, item)


@router.delete("/api/inicio-turno/qr-generados/{item_id}")
def borrar_qr_generado(
    item_id: int,
    db: Session = Depends(get_db),
):
    item = db.get(RecintoQrGenerado, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="No encontrado")
    db.delete(item)
    db.commit()
    return {"ok": True}


# ── Rondas: escaneo de checkpoints por el guardia ──────────────────────────────

@router.get("/inicio-turno/ronda-manual", response_class=HTMLResponse)
def inicio_turno_ronda_manual_page(
    request: Request,
):
    """Entrada alternativa a /inicio-turno/ronda para cuando el QR físico no
    se puede leer: el guardia ingresa el código verificador de 6 dígitos
    (franja vertical junto al QR) y queda redirigido a la misma pantalla de
    registro de siempre. El código es único en todo el sistema, así que no
    hace falta elegir antes el recinto."""
    return templates.TemplateResponse(
        request,
        "inicio_turno_ronda_manual.html",
        {"request": request},
    )


@router.get("/api/inicio-turno/rondas/resolver")
def resolver_ronda_manual(
    verificador: int = Query(default=0),
    db: Session = Depends(get_db),
):
    if verificador < 1:
        raise HTTPException(status_code=422, detail="Ingresa el código verificador.")
    item = db.query(RecintoQrGenerado).filter(RecintoQrGenerado.verificador == verificador).first()
    if not item:
        raise HTTPException(
            status_code=404,
            detail="No existe ningún checkpoint con ese código. Verifica con tu supervisor.",
        )
    return {"qr": item.token, "recinto_label": item.recinto_label, "numero": item.numero}


@router.get("/inicio-turno/ronda", response_class=HTMLResponse)
def inicio_turno_ronda_page(
    request: Request,
    qr: str = Query(default=""),
    db: Session = Depends(get_db),
):
    token = str(qr or "").strip()
    item = db.query(RecintoQrGenerado).filter(RecintoQrGenerado.token == token).first() if token else None
    if not item:
        return templates.TemplateResponse(
            request,
            "inicio_turno_ronda.html",
            {"request": request, "valido": False},
            status_code=404,
        )
    return templates.TemplateResponse(
        request,
        "inicio_turno_ronda.html",
        {
            "request": request,
            "valido": True,
            "qr": token,
            "recinto_label": item.recinto_label,
            "numero": item.numero,
        },
    )


class RondaRegistrarRequest(BaseModel):
    qr: str = Field(min_length=1)
    rut: str = Field(default="", max_length=40)
    nota: str = Field(default="", max_length=2000)


@router.post("/api/inicio-turno/rondas/registrar")
def registrar_ronda(
    payload: RondaRegistrarRequest,
    db: Session = Depends(get_db),
):
    token = payload.qr.strip()
    item = db.query(RecintoQrGenerado).filter(RecintoQrGenerado.token == token).first()
    if not item:
        raise HTTPException(status_code=404, detail="QR no válido")

    rut_norm = _normalizar_rut(payload.rut) if payload.rut else ""
    nombre_guardia = None
    if rut_norm:
        guardia = _buscar_guardia_por_rut(db, rut_norm)
        if not guardia:
            raise HTTPException(status_code=404, detail="RUT no encontrado en BBDD Guardias")
        nombre_guardia = guardia.nombre

    nota = payload.nota.strip()
    registro = RondaRegistro(
        qr_generado_id=item.id,
        recinto_id=item.recinto_id,
        recinto_label=item.recinto_label,
        numero=item.numero,
        rut_guardia=rut_norm or None,
        nombre_guardia=nombre_guardia,
        nota=nota or None,
        aprobado=not bool(nota),
        registrado_at=datetime.now(),
    )
    db.add(registro)
    db.commit()
    return {
        "ok": True,
        "recinto_label": item.recinto_label,
        "numero": item.numero,
        "nombre_guardia": nombre_guardia or "",
        "con_nota": bool(nota),
        "registrado_at": registro.registrado_at.strftime("%d-%m-%Y %H:%M") if registro.registrado_at else "",
    }


@router.get("/api/inicio-turno/rondas")
def listar_rondas_dia(
    checkpoint_id: int = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    day: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_login),
):
    registros = (
        db.query(RondaRegistro)
        .filter(
            RondaRegistro.qr_generado_id == checkpoint_id,
            extract("year",  RondaRegistro.registrado_at) == year,
            extract("month", RondaRegistro.registrado_at) == month,
            extract("day",   RondaRegistro.registrado_at) == day,
        )
        .order_by(RondaRegistro.registrado_at.asc())
        .all()
    )
    return {
        "items": [
            {
                "id": r.id,
                "recinto_label": r.recinto_label,
                "numero": r.numero,
                "rut_guardia": r.rut_guardia or "",
                "nombre_guardia": r.nombre_guardia or "",
                "nota": r.nota or "",
                "aprobado": r.aprobado,
                "aprobado_por": r.aprobado_por or "",
                "registrado_at": r.registrado_at.strftime("%d-%m-%Y %H:%M") if r.registrado_at else "",
            }
            for r in registros
        ]
    }


class RondaAprobarRequest(BaseModel):
    supervisor: str = Field(default="", max_length=255)


@router.patch("/api/inicio-turno/rondas/{registro_id}/aprobar")
def aprobar_ronda(
    registro_id: int,
    payload: RondaAprobarRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_login),
):
    registro = db.get(RondaRegistro, registro_id)
    if not registro:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    registro.aprobado = True
    registro.aprobado_por = payload.supervisor.strip() or None
    registro.aprobado_at = datetime.now()
    db.commit()
    return {"ok": True, "id": registro.id, "aprobado": True}


def _construir_filas_rondas(
    db: Session,
    grupo: str,
    year: int,
    month: int,
) -> tuple[list[dict], list[dict]]:
    """Arma las filas (checkpoint x dia) de la Tabla de Rondas para un grupo.
    Compartido entre la vista web (guardia_tabla_rondas_page) y el informe
    Excel (descargar_informes_rondas)."""
    today = date.today()
    ids_recintos = {str(r["id"]) for r in _listar_recintos_para_generador(db) if r["grupo"] == grupo}

    checkpoints = (
        db.query(RecintoQrGenerado)
        .filter(RecintoQrGenerado.recinto_id.in_(ids_recintos))
        .order_by(RecintoQrGenerado.recinto_label.asc(), RecintoQrGenerado.numero.asc())
        .all()
    )

    registros = (
        db.query(RondaRegistro)
        .filter(
            RondaRegistro.recinto_id.in_(ids_recintos),
            extract("year",  RondaRegistro.registrado_at) == year,
            extract("month", RondaRegistro.registrado_at) == month,
        )
        .all()
    )

    conteo: dict[tuple[int, int], int] = defaultdict(int)
    pendiente: dict[tuple[int, int], bool] = defaultdict(bool)
    detalle: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for r in registros:
        clave = (r.qr_generado_id, r.registrado_at.day)
        conteo[clave] += 1
        if not r.aprobado:
            pendiente[clave] = True
        detalle[clave].append(
            {
                "hora": r.registrado_at.strftime("%H:%M"),
                "nombre": str(r.nombre_guardia or "Sin identificar").strip() or "Sin identificar",
                "aprobado": r.aprobado,
                "nota": str(r.nota or "").strip(),
            }
        )
    for lista in detalle.values():
        lista.sort(key=lambda e: e["hora"])

    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {
            "day": d,
            "dow": _DAY_NAMES[date(year, month, d).weekday()],
            "weekend": date(year, month, d).weekday() >= 5,
            "pasado": date(year, month, d) <= today,
            "hoy": date(year, month, d) == today,
        }
        for d in range(1, days_in_month + 1)
    ]

    if grupo == "quintero":
        _limpiar_label = _limpiar_label_quintero
    elif grupo == "privados":
        _limpiar_label = _limpiar_label_privado
    else:
        _limpiar_label = lambda v: v

    filas = [
        {
            "id": cp.id,
            "recinto_label": _limpiar_label(cp.recinto_label),
            "numero": cp.numero,
            "dias": {
                d: {
                    "cantidad": conteo.get((cp.id, d), 0),
                    "pendiente": pendiente.get((cp.id, d), False),
                    "detalle": detalle.get((cp.id, d), []),
                }
                for d in range(1, days_in_month + 1)
            },
        }
        for cp in checkpoints
    ]

    # Agrupar filas consecutivas del mismo recinto: la primera fila de cada
    # grupo lleva el nombre (con rowspan) y marca el inicio para el separador visual.
    for idx, fila in enumerate(filas):
        es_inicio = idx == 0 or filas[idx - 1]["recinto_label"] != fila["recinto_label"]
        fila["es_inicio_grupo"] = es_inicio
        if es_inicio:
            tamano = 1
            for siguiente in filas[idx + 1:]:
                if siguiente["recinto_label"] == fila["recinto_label"]:
                    tamano += 1
                else:
                    break
            fila["tamano_grupo"] = tamano
        else:
            fila["tamano_grupo"] = 0

    return filas, days_info


def _crear_hoja_rondas_calendario(
    ws,
    periodo: str,
    month: int,
    days_info: list[dict],
    filas: list[dict],
) -> None:
    """Hoja calendario (recinto x fecha) con el estado de cada checkpoint de
    ronda por dia — mismo criterio de color que la Tabla de Rondas en
    pantalla (verde=registrada, amarillo=registrada con observacion,
    rojo=no registrada)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    n_dias = len(days_info)
    n_cols = 2 + n_dias  # Recinto + Checkpoint + un col por dia

    fill_estado = {"verde": "D3EEDA", "amarillo": "F5E6C4", "rojo": "F8D0D4"}
    texto_fila = "1F2937"

    ws.append(["ATC - Registro Rondas (por recinto y checkpoint)"])
    ws.append([periodo])
    ws.append(["Verde = registrada  ·  Amarillo = registrada con observación  ·  Rojo = no registrada  ·  ⚠ = con observación pendiente de aprobar"])
    ws.append(["Recinto", "Checkpoint"] + [f"{d['day']:02d}/{month:02d} {d['dow']}" for d in days_info])

    fila_base = ws.max_row + 1
    for idx, fila in enumerate(filas):
        row_idx = fila_base + idx

        cell_num = ws.cell(row=row_idx, column=2, value=fila["numero"])
        cell_num.font = Font(bold=True, color="475569")
        cell_num.alignment = Alignment(horizontal="center", vertical="center")

        for col_offset, d in enumerate(days_info, start=3):
            info = fila["dias"].get(d["day"], {"cantidad": 0, "pendiente": False, "detalle": []})
            if info["cantidad"] <= 0:
                estado = "rojo"
            elif info["pendiente"]:
                estado = "amarillo"
            else:
                estado = "verde"
            texto = "\n".join(
                f"{e['hora']} · {e['nombre']}"
                + (f" — {e['nota']}" if e.get("nota") else "")
                + ("" if e["aprobado"] else " ⚠")
                for e in info["detalle"]
            )
            cell = ws.cell(row=row_idx, column=col_offset, value=texto or None)
            cell.font = Font(size=9, color=texto_fila)
            cell.fill = PatternFill("solid", fgColor=fill_estado[estado])
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if fila.get("es_inicio_grupo"):
            ws.cell(row=row_idx, column=1, value=fila["recinto_label"])
            tamano = fila.get("tamano_grupo") or 1
            if tamano > 1:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + tamano - 1, end_column=1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A3"].font = Font(bold=True, size=9, color="64748B")
    ws["A3"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila_cells in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=n_cols):
        for cell in fila_cells:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            if cell.column == 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.font is None or not cell.font.b:
                    cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    for idx in range(3, n_cols + 1):
        col_letter = ws.cell(row=4, column=idx).column_letter
        ws.column_dimensions[col_letter].width = 26
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = "C5"


def _preview_cell(value: object, css: str = "") -> dict:
    return {"value": value if value not in (None, "") else "—", "css": css}


def _preview_turnos_section(
    titulo: str,
    rows: list[dict],
    *,
    faltas_por_guardia: dict[str, int] | None = None,
) -> dict:
    preview_rows = []
    for row in rows:
        turnos = row["turnos"]
        dia = int(turnos["Dia"])
        noche = int(turnos["Noche"])
        extra = int(turnos["Extra"])
        contrato = int(turnos["Contrato Diario"])
        if faltas_por_guardia is not None:
            identidad = row["rut"] or _normalizar_texto(row["nombre"])
            faltas = int(faltas_por_guardia.get(identidad, 0))
            preview_rows.append([
                _preview_cell(row["nombre"]),
                _preview_cell(row["rut"]),
                _preview_cell(row.get("tipo_guardia", "")),
                _preview_cell(dia + noche, "num"),
                _preview_cell(extra, "num"),
                _preview_cell(contrato, "num"),
                _preview_cell(faltas, "num"),
            ])
        else:
            preview_rows.append([
                _preview_cell(row["nombre"]),
                _preview_cell(row["rut"]),
                _preview_cell(row.get("tipo_guardia", "")),
                _preview_cell(dia, "num"),
                _preview_cell(noche, "num"),
                _preview_cell(dia + noche, "num"),
                _preview_cell(extra, "num"),
                _preview_cell(contrato, "num"),
                _preview_cell(dia + noche + extra + contrato, "num"),
            ])
    headers = (
        ["Nombre", "RUT", "Tipo de Contrato", "Turno Normal", "Extra", "Contrato Diario", "Faltas"]
        if faltas_por_guardia is not None
        else ["Nombre", "RUT", "Tipo de Contrato", "Dia", "Noche", "Turno Normal", "Extra", "Contrato Diario", "Total Turnos"]
    )
    return {
        "title": titulo,
        "subtitle": f"{len(preview_rows)} guardia(s)",
        "headers": headers,
        "rows": preview_rows,
    }


def _preview_cruce_section(rows: list[dict]) -> dict:
    preview_rows = []
    for row in rows:
        turnos = row["turnos"]
        dia = int(turnos["Dia"])
        noche = int(turnos["Noche"])
        extra = int(turnos["Extra"])
        contrato = int(turnos["Contrato Diario"])
        disyuntivas = int(row["solo_registro"]) + int(row["solo_supervisor"])
        resultado = "Coincide" if disyuntivas == 0 else "Con disyuntivas"
        detalle = "Coincidencia completa"
        css_resultado = "ok"
        if disyuntivas:
            detalle = f"Solo registro guardia: {row['solo_registro']} | Solo supervisor: {row['solo_supervisor']}"
            css_resultado = "bad"
        preview_rows.append([
            _preview_cell(row["nombre"]),
            _preview_cell(row["rut"]),
            _preview_cell(row.get("tipo_guardia", "")),
            _preview_cell(dia, "num"),
            _preview_cell(noche, "num"),
            _preview_cell(dia + noche, "num"),
            _preview_cell(extra, "num"),
            _preview_cell(contrato, "num"),
            _preview_cell(row["coincidencias"], "num"),
            _preview_cell(disyuntivas, "num"),
            _preview_cell(resultado, css_resultado),
            _preview_cell(detalle),
        ])
    return {
        "title": "Cruce",
        "subtitle": f"{len(preview_rows)} guardia(s)",
        "headers": ["Nombre", "RUT", "Tipo de Contrato", "Dia", "Noche", "Turno Normal", "Extra", "Contrato Diario", "Coincidencias", "Disyuntivas", "Resultado", "Detalle"],
        "rows": preview_rows,
    }


def _preview_faltas_section(rows: list[dict]) -> dict:
    preview_rows = []
    for row in rows:
        lineas_html = []
        for texto, es_permiso in row["detalle"]:
            linea = html.escape(texto)
            if es_permiso:
                linea += ' <span class="tag-permiso">Permiso sin Goce</span>'
            lineas_html.append(linea)
        detalle_cell = _preview_cell("<br>".join(lineas_html))
        detalle_cell["safe"] = True
        preview_rows.append([
            _preview_cell(row["nombre"]),
            _preview_cell(row["rut"]),
            _preview_cell(row.get("tipo_guardia", "")),
            _preview_cell(int(row["faltas"]), "num bad"),
            detalle_cell,
        ])
    return {
        "title": "Faltas",
        "subtitle": f"{sum(int(row['faltas']) for row in rows)} falta(s)",
        "headers": ["Nombre", "RUT", "Tipo de Contrato", "Nro de faltas", "Cuándo y dónde tenía que estar"],
        "rows": preview_rows,
    }


def _preview_domingos_section(rows: list[dict]) -> dict:
    preview_rows = []
    for row in rows:
        detalle_cell = _preview_cell("<br>".join(html.escape(texto) for texto in row["detalle"]))
        detalle_cell["safe"] = True
        preview_rows.append([
            _preview_cell(row["nombre"]),
            _preview_cell(row["rut"]),
            _preview_cell(row.get("tipo_guardia", "")),
            _preview_cell(int(row["domingos"]), "num"),
            detalle_cell,
        ])
    return {
        "title": "Conteo días domingo",
        "subtitle": f"{sum(int(row['domingos']) for row in rows)} domingo(s) trabajado(s)",
        "headers": ["Nombre", "RUT", "Tipo de Contrato", "Domingos trabajados", "Fecha y lugar"],
        "rows": preview_rows,
    }


def _preview_turnos_extra_section(titulo: str, rows: list[dict], *, tipo_turno: str) -> dict:
    preview_rows = []
    for row in rows:
        inconsistencias = row["inconsistencias"]
        estado = "OK" if not inconsistencias else "Con inconsistencias"
        preview_rows.append([
            _preview_cell(row["nombre"]),
            _preview_cell(row["rut"]),
            _preview_cell(row.get("tipo_guardia", "")),
            _preview_cell(int(row["total"]), "num"),
            _preview_cell("\n".join(row["lugares"])),
            _preview_cell(estado, "ok" if estado == "OK" else "bad"),
            _preview_cell("\n".join(inconsistencias) if inconsistencias else "Registro de guardia y supervisor coinciden"),
        ])
    return {
        "title": titulo,
        "subtitle": f"{len(preview_rows)} guardia(s)",
        "headers": ["Nombre", "RUT", "Tipo de Contrato", f"Turnos {tipo_turno}", "Dónde (fecha · recinto)", "Consistencia", "Detalle inconsistencia"],
        "rows": preview_rows,
    }


def _preview_justificaciones_section(rows: list[dict]) -> dict:
    preview_rows = []
    for row in rows:
        desde = row["fecha_desde"].strftime("%d/%m/%Y") if row["fecha_desde"] else "—"
        hasta = row["fecha_hasta"].strftime("%d/%m/%Y") if row["fecha_hasta"] else "En curso"
        preview_rows.append([
            _preview_cell(row["nombre"]),
            _preview_cell(row["rut"]),
            _preview_cell(row.get("tipo_guardia", "")),
            _preview_cell(row["motivo"]),
            _preview_cell(desde),
            _preview_cell(hasta),
            _preview_cell(row["notas"]),
        ])
    return {
        "title": "Justificaciones",
        "subtitle": f"{len(preview_rows)} justificación(es) vigente(s) en el período (licencias, vacaciones, permisos, etc.)",
        "headers": ["Nombre", "RUT", "Tipo de Contrato", "Motivo", "Desde", "Hasta", "Notas"],
        "rows": preview_rows,
    }


_TURNO_BADGE_SLUGS = {
    "dia": "dia",
    "noche": "noche",
    "extra": "extra",
    "contrato diario": "contrato-diario",
}
def _turno_chip_html(turno: str, nombre_html: str) -> str:
    slug = _TURNO_BADGE_SLUGS.get(_normalizar_texto(turno))
    if not slug:
        return nombre_html
    return f'<span class="turno-chip turno-chip-{slug}">{nombre_html}</span>'


def _preview_supervisor_calendario_section(
    titulo: str,
    month: int,
    recintos: list[str],
    days_info: list[dict],
    matrix_sv: dict,
    *,
    solo_nombres: bool = False,
) -> dict:
    preview_rows = []
    for recinto in recintos:
        row = [_preview_cell(recinto, "recinto")]
        for d in days_info:
            entradas = [
                e for e in matrix_sv.get(recinto, {}).get(d["day"], [])
                if str(e.get("nombre") or "").strip()
            ]
            if not entradas:
                row.append(_preview_cell("", "muted"))
                continue
            if solo_nombres:
                # La celda con el nombre coloreado segun el tipo de turno, sin
                # escribir la palabra del turno (mismo esquema de colores que
                # las tarjetas de "Registro de Guardias por Recinto").
                textos = []
                for e in entradas:
                    nombre = html.escape(str(e.get("nombre") or "").strip())
                    textos.append(_turno_chip_html(str(e.get("turno") or ""), nombre))
                cell = _preview_cell("\n".join(textos))
                cell["safe"] = True
            else:
                textos = []
                for e in entradas:
                    nombre = str(e.get("nombre") or "").strip()
                    turno = str(e.get("turno") or "").strip()
                    notas = str(e.get("notas") or "").strip()
                    textos.append(f"{nombre} · {turno}" + (f"\n{notas}" if notas else ""))
                cell = _preview_cell("\n\n".join(textos))
            row.append(cell)
        preview_rows.append(row)
    return {
        "title": titulo,
        "subtitle": f"{len(recintos)} recinto(s)",
        "headers": ["Recinto"] + [f"{d['day']:02d}/{month:02d} {d['dow']}" for d in days_info],
        "rows": preview_rows,
    }


def _preview_rondas_section(titulo: str, month: int, filas: list[dict], days_info: list[dict]) -> dict:
    preview_rows = []
    for fila in filas:
        row = [
            _preview_cell(fila["recinto_label"], "recinto"),
            _preview_cell(f"QR {fila['numero']}", "num"),
        ]
        for d in days_info:
            info = fila["dias"].get(d["day"], {})
            cantidad = int(info.get("cantidad") or 0)
            detalle = info.get("detalle") or []
            pendiente = bool(info.get("pendiente"))
            if cantidad:
                texto = "\n".join(
                    f"{e['hora']} · {e['nombre']}"
                    + (f" — {e['nota']}" if e.get("nota") else "")
                    + ("" if e.get("aprobado") else " · pendiente")
                    for e in detalle
                )
                row.append(_preview_cell(texto, "cell-warn" if pendiente else "cell-ok"))
            elif d.get("pasado"):
                row.append(_preview_cell("No registrada", "cell-bad bad"))
            else:
                row.append(_preview_cell("", "muted"))
        preview_rows.append(row)
    return {
        "title": titulo,
        "subtitle": f"{len(filas)} checkpoint(s)",
        "headers": ["Recinto", "Checkpoint"] + [f"{d['day']:02d}/{month:02d} {d['dow']}" for d in days_info],
        "rows": preview_rows,
    }


def _datos_informe_guardias(db: Session, year: int, month: int, grupo: str) -> dict:
    today = date.today()
    registros_qr = (
        db.query(InicioTurnoRegistro)
        .filter(
            extract("year", InicioTurnoRegistro.registrado_at) == year,
            extract("month", InicioTurnoRegistro.registrado_at) == month,
        )
        .order_by(InicioTurnoRegistro.registrado_at, InicioTurnoRegistro.id)
        .all()
    )
    # Marcajes fusionados (ver _detectar_candidatos_fusion) quedan afuera de
    # todo conteo/informe — el turno real es el marcaje sobreviviente.
    registros_qr = [r for r in registros_qr if str(r.estado or "activo") != "archivado"]
    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year", SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    grupo, recintos = _recintos_para_grupo(db, grupo, registros_qr, registros_sv)
    recintos_set = set(recintos)
    registros_qr = [r for r in registros_qr if str(r.recinto or "").strip() in recintos_set]
    registros_sv = [r for r in registros_sv if str(r.recinto or "").strip() in recintos_set]

    rut_lookup = _rut_lookup_guardias(db, registros_qr, registros_sv)
    tipo_guardia_lookup = _tipo_guardia_lookup(db)
    rows_qr = _agrupar_turnos_por_guardia(registros_qr, rut_lookup, tipo_guardia_lookup, usa_rut_registro=True)
    rows_sv = _agrupar_turnos_por_guardia(registros_sv, rut_lookup, tipo_guardia_lookup, usa_rut_registro=False)
    registros_qr_pasado = [r for r in registros_qr if r.registrado_at.date() < today]
    registros_sv_pasado = [r for r in registros_sv if r.fecha < today]
    rows_cruce = _agrupar_cruce_guardias(registros_qr_pasado, registros_sv_pasado, rut_lookup, tipo_guardia_lookup)
    permisos_sin_goce = _permisos_sin_goce_lookup(db, year, month)
    rows_faltas = _agrupar_faltas_guardias(registros_qr_pasado, registros_sv_pasado, rut_lookup, tipo_guardia_lookup, permisos_sin_goce)
    rows_faltas = _agregar_faltas_por_permiso_sin_goce(
        db,
        rows_faltas,
        rut_lookup,
        tipo_guardia_lookup,
        year,
        month,
        today - timedelta(days=1),
        recintos_permitidos=recintos_set,
    )
    rows_extra = _agrupar_turnos_extra(registros_qr, registros_sv, rut_lookup, tipo_guardia_lookup, tipo_turno="Extra")
    rows_contrato_diario = _agrupar_turnos_extra(registros_qr, registros_sv, rut_lookup, tipo_guardia_lookup, tipo_turno="Contrato Diario")
    rows_domingos = _agrupar_domingos_guardias(registros_qr, rut_lookup, tipo_guardia_lookup)
    rows_justificaciones = _justificaciones_lookup(db, year, month, tipo_guardia_lookup)

    matrix_sv: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_sv:
        matrix_sv[r.recinto][r.fecha.day].append({
            "nombre": r.nombre_guardia,
            "turno": r.tipo_turno,
            "notas": r.notas or "",
        })
    matrix_qr: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_qr:
        matrix_qr[r.recinto][r.registrado_at.day].append({
            "id": r.id,
            "nombre": r.nombre_guardia,
            "turno": r.tipo_turno,
            "hora": r.registrado_at.strftime("%H:%M"),
        })
    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()]}
        for d in range(1, days_in_month + 1)
    ]
    return {
        "grupo": grupo,
        "periodo": f"{_MONTH_NAMES[month - 1]} {year} | Grupo: {grupo.title()}",
        "rows_qr": rows_qr,
        "rows_sv": rows_sv,
        "rows_cruce": rows_cruce,
        "rows_faltas": rows_faltas,
        "rows_extra": rows_extra,
        "rows_contrato_diario": rows_contrato_diario,
        "rows_domingos": rows_domingos,
        "rows_justificaciones": rows_justificaciones,
        "matrix_sv": matrix_sv,
        "matrix_qr": matrix_qr,
        "recintos": recintos,
        "days_info": days_info,
    }


def _datos_cumplimiento_turnos(db: Session, year: int, month: int, grupo: str) -> dict:
    """Compara la cuota EXACTA de turnos mensuales por dependencia
    (TurnoEstipulado, ver _importar_turnos_estipulados_concon.py) contra el
    conteo real de marcajes ACTIVOS (ya excluyendo fusionados/archivados,
    ver _fusionar_automatico_si_corresponde) — solo cantidades, sin
    nombres de guardias."""
    grupo = str(grupo or "concon").strip().lower()
    estipulados = (
        db.query(TurnoEstipulado)
        .filter(TurnoEstipulado.grupo == grupo)
        .order_by(TurnoEstipulado.id)
        .all()
    )

    dias_en_mes = calendar.monthrange(year, month)[1]

    def _parse_dias_semana(raw: str | None) -> set[int] | None:
        txt = str(raw or "").strip()
        if not txt:
            return None
        try:
            return {int(x) for x in txt.split(",") if x.strip() != ""}
        except ValueError:
            return None

    def _regla_aplica(dias_semana: set[int] | None, weekday: int) -> bool:
        return dias_semana is None or weekday in dias_semana

    # Varias filas de TurnoEstipulado pueden ser la misma dependencia con
    # tramos horarios/dias distintos (ej. Juzgado Lunes/Miercoles/Jueves,
    # cada uno con su propio turnos_dia) — se guardan como reglas separadas
    # y el estipulado de cada dia se recalcula sumando las reglas que
    # aplican ese dia de la semana, no con un total fijo de la planilla
    # (que asume un mes puntual de 30 o 31 dias con una combinacion de
    # dias de semana que no es la misma todos los meses).
    por_dependencia: dict[tuple, dict] = {}
    for est in estipulados:
        clave = (est.dependencia, est.sucursal_id)
        row = por_dependencia.setdefault(clave, {
            "dependencia": est.dependencia,
            "sucursal_id": est.sucursal_id,
            "reglas": [],
        })
        row["reglas"].append((est.turnos_dia or 0, _parse_dias_semana(est.dias_semana)))

    ids_sucursal = [row["sucursal_id"] for row in por_dependencia.values() if row["sucursal_id"]]

    # El texto de "dependencia" viene tal cual de la planilla del cliente
    # (a veces en minusculas, sin el prefijo "MC ..." real) — para mostrar
    # el nombre real se usa el de bbdd_sucursales via el cruce ya hecho en
    # el import (sucursal_id), con el texto de la planilla como respaldo si
    # no hay cruce.
    nombres_sucursal: dict[int, str] = {}
    if ids_sucursal:
        nombres_sucursal = {
            sid: nombre
            for sid, nombre in db.query(SucursalBBDD.id, SucursalBBDD.nombre_sucursal)
            .filter(SucursalBBDD.id.in_(ids_sucursal))
            .all()
        }
    for row in por_dependencia.values():
        if row["sucursal_id"] and nombres_sucursal.get(row["sucursal_id"]):
            row["dependencia"] = nombres_sucursal[row["sucursal_id"]]

    # El conteo "real" tiene que sumar TODOS los tipos de turno (Dia, Noche,
    # Extra, Contrato Diario — pedido explicito) y no puede depender solo de
    # sucursal_id: hay marcajes del registro de guardias con sucursal_id
    # NULL (ej. cargados manualmente desde el panel, que solo guardan
    # `recinto` como texto) que igual corresponden a estas dependencias.
    # Se matchea por sucursal_id cuando esta, y si no por el texto exacto
    # de `recinto` contra el nombre real de la sucursal.
    ids_sucursal_set = set(ids_sucursal)
    recinto_por_sucursal_norm: dict[str, int] = {
        _normalizar_texto(nombre): sid for sid, nombre in nombres_sucursal.items()
    }

    reales_por_sucursal: dict[int, int] = defaultdict(int)
    reales_diarios: dict[int, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    if ids_sucursal:
        registros_relevantes = (
            db.query(
                InicioTurnoRegistro.sucursal_id,
                InicioTurnoRegistro.recinto,
                InicioTurnoRegistro.registrado_at,
            )
            .filter(
                extract("year", InicioTurnoRegistro.registrado_at) == year,
                extract("month", InicioTurnoRegistro.registrado_at) == month,
                InicioTurnoRegistro.estado != "archivado",
                or_(
                    InicioTurnoRegistro.sucursal_id.in_(ids_sucursal),
                    InicioTurnoRegistro.recinto.in_(list(nombres_sucursal.values())),
                ),
            )
            .all()
        )
        for sid, recinto_txt, registrado_at in registros_relevantes:
            resuelto = sid if sid in ids_sucursal_set else recinto_por_sucursal_norm.get(_normalizar_texto(recinto_txt))
            if not resuelto:
                continue
            reales_por_sucursal[resuelto] += 1
            reales_diarios[resuelto][registrado_at.day] += 1

    hoy = date.today()

    filas = []
    total_estipulado = 0
    total_real = 0
    total_cumplidas = 0
    totales_dia_real = [0] * dias_en_mes
    totales_dia_estipulado = [0] * dias_en_mes
    for row in sorted(por_dependencia.values(), key=lambda r: r["dependencia"]):
        sucursal_id = row["sucursal_id"]
        reglas = row["reglas"]
        real = reales_por_sucursal.get(sucursal_id, 0) if sucursal_id else 0
        conteo_dia = reales_diarios.get(sucursal_id, {}) if sucursal_id else {}
        dias = []
        estipulado = 0
        for d in range(1, dias_en_mes + 1):
            weekday = date(year, month, d).weekday()
            estipulado_dia = sum(turnos for turnos, regla in reglas if _regla_aplica(regla, weekday))
            estipulado += estipulado_dia
            real_dia = conteo_dia.get(d, 0)
            totales_dia_real[d - 1] += real_dia
            totales_dia_estipulado[d - 1] += estipulado_dia
            if date(year, month, d) >= hoy:
                estado_dia = "futuro"
            else:
                estado_dia = "ok" if real_dia == estipulado_dia else "bajo"
            dias.append({"day": d, "real": real_dia, "estipulado": estipulado_dia, "estado": estado_dia})
        diferencia = real - estipulado
        cumplido = diferencia == 0
        total_estipulado += estipulado
        total_real += real
        if cumplido:
            total_cumplidas += 1
        filas.append({
            "dependencia": row["dependencia"],
            "sucursal_id": sucursal_id,
            "estipulado": estipulado,
            "real": real,
            "diferencia": diferencia,
            "cumplido": cumplido,
            "sin_cruce": sucursal_id is None,
            "dias": dias,
        })

    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()]}
        for d in range(1, dias_en_mes + 1)
    ]

    dias_totales = []
    for d in range(1, dias_en_mes + 1):
        real_dia = totales_dia_real[d - 1]
        estipulado_dia = totales_dia_estipulado[d - 1]
        if date(year, month, d) >= hoy:
            estado_dia = "futuro"
        else:
            estado_dia = "ok" if real_dia == estipulado_dia else "bajo"
        dias_totales.append({"day": d, "real": real_dia, "estipulado": estipulado_dia, "estado": estado_dia})

    return {
        "grupo": grupo,
        "dias_en_mes": dias_en_mes,
        "days_info": days_info,
        "filas": filas,
        "dias_totales": dias_totales,
        "total_estipulado": total_estipulado,
        "total_real": total_real,
        "total_diferencia": total_real - total_estipulado,
        "total_cumplidas": total_cumplidas,
        "total_dependencias": len(filas),
    }


@router.get("/guardia/cumplimiento-turnos", response_class=HTMLResponse)
def guardia_cumplimiento_turnos_page(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="concon"),
    db: Session = Depends(get_db),
):
    tz = ZoneInfo(settings.timezone or "America/Santiago")
    today = datetime.now(tz).date()
    year = year or today.year
    month = month or today.month
    data = _datos_cumplimiento_turnos(db, year, month, grupo)

    prev_y, prev_m = (year - 1, 12) if month == 1 else (year, month - 1)
    next_y, next_m = (year + 1, 1) if month == 12 else (year, month + 1)

    return templates.TemplateResponse(
        request,
        "guardia_cumplimiento_turnos.html",
        {
            "request": request,
            "year": year,
            "month": month,
            "month_name": _MONTH_NAMES[month - 1],
            "grupo": data["grupo"],
            "dias_en_mes": data["dias_en_mes"],
            "days_info": data["days_info"],
            "filas": data["filas"],
            "dias_totales": data["dias_totales"],
            "total_estipulado": data["total_estipulado"],
            "total_real": data["total_real"],
            "total_diferencia": data["total_diferencia"],
            "total_cumplidas": data["total_cumplidas"],
            "total_dependencias": data["total_dependencias"],
            "prev_y": prev_y, "prev_m": prev_m,
            "next_y": next_y, "next_m": next_m,
        },
    )


@router.get("/guardia/tabla-rondas/informes/preview", response_class=HTMLResponse)
def preview_informes_rondas(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: Session = Depends(get_db),
):
    today = date.today()
    year = year or today.year
    month = month or today.month

    sections = []
    for grupo, titulo in (
        ("quintero", "Rondas Quintero"),
        ("privados", "Rondas Privados"),
        ("concon", "Rondas Concón"),
    ):
        filas, days_info = _construir_filas_rondas(db, grupo, year, month)
        sections.append(_preview_rondas_section(titulo, month, filas, days_info))

    return templates.TemplateResponse(
        request,
        "guardias_informe_preview.html",
        {
            "request": request,
            "titulo": "Vista previa - Informe de rondas",
            "periodo": f"{_MONTH_NAMES[month - 1]} {year} | Quintero, Privados y Concón",
            "sections": sections,
        },
    )


@router.get("/guardia/tabla-rondas/informes")
def descargar_informes_rondas(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    preview: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook

    today = date.today()
    year = year or today.year
    month = month or today.month
    if preview:
        return preview_informes_rondas(request, year=year, month=month, db=db)

    wb = Workbook()
    wb.active.title = "Rondas Quintero"
    filas, days_info = _construir_filas_rondas(db, "quintero", year, month)
    _crear_hoja_rondas_calendario(
        wb.active, f"{_MONTH_NAMES[month - 1]} {year} | Grupo: Quintero", month, days_info, filas
    )

    filas, days_info = _construir_filas_rondas(db, "privados", year, month)
    _crear_hoja_rondas_calendario(
        wb.create_sheet("Rondas Privados"),
        f"{_MONTH_NAMES[month - 1]} {year} | Grupo: Privados",
        month,
        days_info,
        filas,
    )

    filas, days_info = _construir_filas_rondas(db, "concon", year, month)
    _crear_hoja_rondas_calendario(
        wb.create_sheet("Rondas Concón"),
        f"{_MONTH_NAMES[month - 1]} {year} | Grupo: Concón",
        month,
        days_info,
        filas,
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"informe_rondas_{year}_{month:02d}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/guardia/tabla-rondas", response_class=HTMLResponse)
def guardia_tabla_rondas_page(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    db: Session = Depends(get_db),
):
    today = date.today()
    year  = year  or today.year
    month = month or today.month
    grupo = (grupo or "quintero").strip().lower()
    if grupo not in ("quintero", "privados", "concon"):
        grupo = "quintero"

    filas, days_info = _construir_filas_rondas(db, grupo, year, month)

    prev_y, prev_m = (year - 1, 12) if month == 1  else (year, month - 1)
    next_y, next_m = (year + 1, 1)  if month == 12 else (year, month + 1)

    return templates.TemplateResponse(
        request,
        "tabla_rondas.html",
        {
            "request": request,
            "year": year, "month": month,
            "month_name": _MONTH_NAMES[month - 1],
            "grupo": grupo,
            "filas": filas,
            "days_info": days_info,
            "prev_y": prev_y, "prev_m": prev_m,
            "next_y": next_y, "next_m": next_m,
            "today_day": today.day if (today.year == year and today.month == month) else -1,
            "supervisor_nombre": _nombre_usuario_sesion(request, db),
        },
    )


_MONTH_NAMES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
_DAY_NAMES   = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]


# ── BBDD Guardias ──────────────────────────────────────────────────────────────

class GuardiaCreate(BaseModel):
    rut: str = Field(min_length=1, max_length=40)
    nombre: str = Field(min_length=1, max_length=255)
    tipo_contrato: str = Field(min_length=1, max_length=40)


def _departamento_guardia_por_contrato(value: object) -> str:
    key = _normalizar_texto(value).replace(" ", "")
    if key == "parttime":
        return "GuardiaPartTime"
    if key == "fulltime":
        return "GuardiasFullTime"
    raise HTTPException(status_code=422, detail="Tipo de contrato inválido")


@router.get("/guardia/bbdd-guardias", response_class=HTMLResponse)
def bbdd_guardias_page(request: Request, current_user: User = Depends(_require_login)):
    return templates.TemplateResponse(
        request,
        "bbdd_guardia_registro.html",
        {"user": None},
    )


@router.post("/api/inicio-turno/guardias/crear")
def crear_guardia(
    payload: GuardiaCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_login),
):
    rut_norm = _normalizar_rut(payload.rut)
    nombre = payload.nombre.strip()
    department = _departamento_guardia_por_contrato(payload.tipo_contrato)
    if not rut_norm:
        raise HTTPException(status_code=422, detail="RUT inválido")
    if not nombre:
        raise HTTPException(status_code=422, detail="Nombre inválido")

    guardias_existentes = db.query(InicioTurnoGuardia).all()
    guardia = next(
        (item for item in guardias_existentes if _normalizar_rut(item.rut) == rut_norm),
        None,
    )
    created = False
    nombre_duplicado = False
    if guardia:
        guardia.nombre = nombre
    else:
        nombre_norm = _normalizar_texto(nombre)
        nombre_duplicado = any(
            _normalizar_texto(item.nombre) == nombre_norm for item in guardias_existentes
        )
        guardia = InicioTurnoGuardia(rut=rut_norm, nombre=nombre)
        db.add(guardia)
        created = True

    user = db.query(User).filter(User.username == rut_norm).first()
    if user:
        user.name = nombre
        # No pisar el departamento de usuarios admin/superadmin: pueden tener un
        # RUT registrado como guardia (ej. jefaturas) pero su area real no es
        # la de guardias de turno — solo se autoasigna a cuentas 'agent'.
        if user.role not in ("admin", "superadmin"):
            user.department = department
        user.is_active = True
    else:
        db.add(
            User(
                username=rut_norm,
                name=nombre,
                hashed_password="",
                role="agent",
                department=department,
                is_active=True,
            )
        )

    db.commit()
    db.refresh(guardia)
    return {
        "id": guardia.id,
        "rut": guardia.rut,
        "nombre": guardia.nombre,
        "tipo_contrato": _tipo_guardia_label(department),
        "created": created,
        "nombre_duplicado": nombre_duplicado,
    }


@router.get("/api/inicio-turno/guardias/listar")
def listar_guardias(db: Session = Depends(get_db), current_user: User = Depends(_require_login)):
    guardias = db.query(InicioTurnoGuardia).order_by(InicioTurnoGuardia.nombre.asc()).all()
    lookup = _tipo_guardia_lookup(db)
    items = []
    for g in guardias:
        rut_key = _normalizar_rut(g.rut)
        items.append({
            "id": g.id,
            "rut": g.rut,
            "nombre": g.nombre,
            "tipo_contrato": lookup["rut"].get(rut_key, ""),
        })
    return {"items": items}


class GuardiaUpdate(BaseModel):
    rut: str = Field(min_length=1, max_length=40)
    nombre: str = Field(min_length=1, max_length=255)
    tipo_contrato: str = Field(min_length=1, max_length=40)


@router.put("/api/inicio-turno/guardias/{guardia_id}")
def actualizar_guardia(
    guardia_id: int,
    payload: GuardiaUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_login),
):
    guardia = db.get(InicioTurnoGuardia, guardia_id)
    if not guardia:
        raise HTTPException(status_code=404, detail="Guardia no encontrado")

    rut_norm = _normalizar_rut(payload.rut)
    nombre = payload.nombre.strip()
    department = _departamento_guardia_por_contrato(payload.tipo_contrato)
    if not rut_norm:
        raise HTTPException(status_code=422, detail="RUT inválido")
    if not nombre:
        raise HTTPException(status_code=422, detail="Nombre inválido")

    duplicado = next(
        (
            item
            for item in db.query(InicioTurnoGuardia).filter(InicioTurnoGuardia.id != guardia_id).all()
            if _normalizar_rut(item.rut) == rut_norm
        ),
        None,
    )
    if duplicado:
        raise HTTPException(status_code=409, detail="Ya existe otro guardia con ese RUT")

    rut_anterior = _normalizar_rut(guardia.rut)
    guardia.rut = rut_norm
    guardia.nombre = nombre

    user = db.query(User).filter(User.username == rut_anterior).first()
    if not user:
        user = db.query(User).filter(User.username == rut_norm).first()
    if user:
        user.username = rut_norm
        user.name = nombre
        # Ver nota en crear_guardia: no pisar el departamento de admin/superadmin.
        if user.role not in ("admin", "superadmin"):
            user.department = department
        user.is_active = True
    else:
        db.add(
            User(
                username=rut_norm,
                name=nombre,
                hashed_password="",
                role="agent",
                department=department,
                is_active=True,
            )
        )

    db.commit()
    db.refresh(guardia)
    return {
        "id": guardia.id,
        "rut": guardia.rut,
        "nombre": guardia.nombre,
        "tipo_contrato": _tipo_guardia_label(department),
    }


@router.delete("/api/inicio-turno/guardias/{guardia_id}")
def eliminar_guardia(
    guardia_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_login),
):
    guardia = db.get(InicioTurnoGuardia, guardia_id)
    if not guardia:
        raise HTTPException(status_code=404, detail="Guardia no encontrado")

    rut_norm = _normalizar_rut(guardia.rut)
    user = db.query(User).filter(User.username == rut_norm).first()

    db.delete(guardia)
    if user:
        db.delete(user)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="No se pudo eliminar: el usuario tiene tickets, mensajes u otro historial asociado.",
        )

    return {"ok": True, "id": guardia_id}


# ── BBDD Justificaciones (licencia/falta/permiso/vacaciones) ────────────────
# Registro independiente: no requiere que exista antes una fila en la tabla
# de supervisor (dia+recinto). Es solo una bitacora consultable de guardias.

MOTIVOS_JUSTIFICACION = (
    "Licencia Médica",
    "Falta",
    "Permiso con Goce",
    "Permiso sin Goce",
    "Vacaciones",
    "Desvinculado",
    "Renuncia",
)


class JustificacionCreate(BaseModel):
    rut: str = Field(min_length=1, max_length=40)
    nombre: str = Field(min_length=1, max_length=255)
    motivo: str
    fecha_desde: str | None = None  # YYYY-MM-DD
    fecha_hasta: str | None = None  # YYYY-MM-DD
    notas: str = ""


class JustificacionUpdate(JustificacionCreate):
    pass


def _validar_payload_justificacion(payload: JustificacionCreate) -> tuple[str, str, date | None, date | None]:
    rut_norm = _normalizar_rut(payload.rut)
    nombre = payload.nombre.strip()
    motivo = payload.motivo.strip()
    if not rut_norm:
        raise HTTPException(status_code=422, detail="RUT inválido")
    if not nombre:
        raise HTTPException(status_code=422, detail="Nombre inválido")
    if motivo not in MOTIVOS_JUSTIFICACION:
        raise HTTPException(status_code=422, detail="Motivo inválido")

    requiere_fechas = motivo != "Falta"
    desde = hasta = None
    if requiere_fechas:
        if not payload.fecha_desde or not payload.fecha_hasta:
            raise HTTPException(status_code=422, detail="Indica desde y hasta cuándo rige la justificación")
        try:
            desde = date.fromisoformat(payload.fecha_desde)
            hasta = date.fromisoformat(payload.fecha_hasta)
        except ValueError:
            raise HTTPException(status_code=422, detail="Fecha inválida")
        if hasta < desde:
            raise HTTPException(status_code=422, detail="'Hasta' no puede ser anterior a 'Desde'")
    return rut_norm, nombre, desde, hasta


MOTIVOS_CON_VIGENCIA = tuple(m for m in MOTIVOS_JUSTIFICACION if m != "Falta")


def _justificacion_conflicto_vigente(
    db: Session,
    rut: str,
    desde: date | None,
    hasta: date | None,
    excluir_id: int | None = None,
) -> GuardiaJustificacion | None:
    """Si el motivo tiene fechas (todo menos 'Falta'), no se puede registrar
    una nueva justificacion que se superponga con una ya vigente del mismo
    guardia — evita licencias/permisos/vacaciones duplicados o cruzados."""
    if not desde or not hasta:
        return None
    query = db.query(GuardiaJustificacion).filter(
        GuardiaJustificacion.rut == rut,
        GuardiaJustificacion.motivo.in_(MOTIVOS_CON_VIGENCIA),
        GuardiaJustificacion.fecha_desde.isnot(None),
        GuardiaJustificacion.fecha_hasta.isnot(None),
        GuardiaJustificacion.fecha_desde <= hasta,
        GuardiaJustificacion.fecha_hasta >= desde,
    )
    if excluir_id is not None:
        query = query.filter(GuardiaJustificacion.id != excluir_id)
    return query.first()


def _serializar_justificacion(j: GuardiaJustificacion) -> dict:
    return {
        "id": j.id,
        "rut": j.rut,
        "nombre": j.nombre_guardia,
        "motivo": j.motivo,
        "fecha_desde": j.fecha_desde.isoformat() if j.fecha_desde else "",
        "fecha_hasta": j.fecha_hasta.isoformat() if j.fecha_hasta else "",
        "notas": j.notas or "",
        "creado_por": j.creado_por or "",
        "created_at": j.created_at.strftime("%d/%m/%Y %H:%M") if j.created_at else "",
    }


@router.get("/guardia/bbdd-justificaciones", response_class=HTMLResponse)
def bbdd_justificaciones_page(request: Request):
    return templates.TemplateResponse(
        request,
        "bbdd_justificaciones.html",
        {"motivos": list(MOTIVOS_JUSTIFICACION)},
    )


@router.post("/api/inicio-turno/justificaciones/crear")
def crear_justificacion(
    payload: JustificacionCreate,
    request: Request,
    db: Session = Depends(get_db),
):
    rut_norm, nombre, desde, hasta = _validar_payload_justificacion(payload)
    conflicto = _justificacion_conflicto_vigente(db, rut_norm, desde, hasta)
    if conflicto:
        raise HTTPException(
            status_code=409,
            detail=f"{nombre} ya tiene una justificación vigente ({conflicto.motivo}, "
                    f"{conflicto.fecha_desde.strftime('%d/%m/%Y')}–{conflicto.fecha_hasta.strftime('%d/%m/%Y')}) "
                    f"que se superpone con este período.",
        )
    j = GuardiaJustificacion(
        rut=rut_norm,
        nombre_guardia=nombre,
        motivo=payload.motivo.strip(),
        fecha_desde=desde,
        fecha_hasta=hasta,
        notas=payload.notas.strip() or None,
        creado_por=_nombre_usuario_sesion(request, db) or None,
    )
    db.add(j)
    db.commit()
    db.refresh(j)
    return _serializar_justificacion(j)


@router.get("/api/inicio-turno/justificaciones/listar")
def listar_justificaciones(
    q: str = Query(default=""),
    db: Session = Depends(get_db),
):
    query = db.query(GuardiaJustificacion).order_by(GuardiaJustificacion.created_at.desc())
    items = [_serializar_justificacion(j) for j in query.all()]
    q_norm = _normalizar_texto(q)
    if q_norm:
        items = [
            it for it in items
            if q_norm in _normalizar_texto(it["nombre"]) or q_norm in _normalizar_texto(it["rut"])
        ]
    return {"items": items}


@router.put("/api/inicio-turno/justificaciones/{justificacion_id}")
def actualizar_justificacion(
    justificacion_id: int,
    payload: JustificacionUpdate,
    db: Session = Depends(get_db),
):
    j = db.get(GuardiaJustificacion, justificacion_id)
    if not j:
        raise HTTPException(status_code=404, detail="Justificación no encontrada")
    rut_norm, nombre, desde, hasta = _validar_payload_justificacion(payload)
    conflicto = _justificacion_conflicto_vigente(db, rut_norm, desde, hasta, excluir_id=justificacion_id)
    if conflicto:
        raise HTTPException(
            status_code=409,
            detail=f"{nombre} ya tiene una justificación vigente ({conflicto.motivo}, "
                    f"{conflicto.fecha_desde.strftime('%d/%m/%Y')}–{conflicto.fecha_hasta.strftime('%d/%m/%Y')}) "
                    f"que se superpone con este período.",
        )
    j.rut = rut_norm
    j.nombre_guardia = nombre
    j.motivo = payload.motivo.strip()
    j.fecha_desde = desde
    j.fecha_hasta = hasta
    j.notas = payload.notas.strip() or None
    db.commit()
    db.refresh(j)
    return _serializar_justificacion(j)


@router.delete("/api/inicio-turno/justificaciones/{justificacion_id}")
def eliminar_justificacion(
    justificacion_id: int,
    db: Session = Depends(get_db),
):
    j = db.get(GuardiaJustificacion, justificacion_id)
    if not j:
        raise HTTPException(status_code=404, detail="Justificación no encontrada")
    db.delete(j)
    db.commit()
    return {"ok": True}


@router.get("/guardia/tabla-registro-guardia", response_class=HTMLResponse)
def guardia_tabla_page(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    db: Session = Depends(get_db),
):
    import unicodedata as _ud

    tz = ZoneInfo(settings.timezone or "America/Santiago")
    today = datetime.now(tz).date()
    control_hasta = today - timedelta(days=1)
    year  = year  or today.year
    month = month or today.month

    registros_qr_todos = (
        db.query(InicioTurnoRegistro)
        .filter(
            extract("year",  InicioTurnoRegistro.registrado_at) == year,
            extract("month", InicioTurnoRegistro.registrado_at) == month,
        )
        .order_by(InicioTurnoRegistro.registrado_at)
        .all()
    )
    by_id_qr = {r.id: r for r in registros_qr_todos}
    registros_qr_archivados = [r for r in registros_qr_todos if str(r.estado or "activo") == "archivado"]
    # El resto de esta funcion (matriz, cruce, conflictos, conteos) usa solo
    # marcajes activos — los fusionados ya no deben generar alertas ni
    # ocupar celdas del calendario operativo; su trazabilidad vive aparte en
    # "fusionados_mes" (ver mas abajo), no mezclada en la matriz principal.
    registros_qr = [r for r in registros_qr_todos if str(r.estado or "activo") != "archivado"]

    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year",  SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    def _norm(name: str) -> str:
        n = _ud.normalize("NFD", (name or "").lower())
        return "".join(c for c in n if _ud.category(c) != "Mn").strip()

    rut_lookup = _rut_lookup_guardias(db, registros_qr, registros_sv)
    permisos_sin_goce = _permisos_sin_goce_lookup(db, year, month)

    candidatos_fusion = _detectar_candidatos_fusion(registros_qr, rut_lookup)

    fusionados_mes: list[dict] = []
    for r in registros_qr_archivados:
        destino = by_id_qr.get(r.fusionado_con_id) if r.fusionado_con_id else None
        fusionados_mes.append({
            "id": r.id,
            "fecha": f"{r.registrado_at.day}/{month:02d}",
            "nombre": r.nombre_guardia,
            "recinto_original": r.recinto,
            "hora_original": r.registrado_at.strftime("%H:%M"),
            "recinto_destino": destino.recinto if destino else "",
            "hora_destino": destino.registrado_at.strftime("%H:%M") if destino else "",
            "motivo": r.archivado_motivo or "",
            "archivado_por": r.archivado_por or "",
        })
    fusionados_mes.sort(key=lambda x: x["fecha"])

    # id del marcaje sobreviviente -> texto "de donde vino" (recinto+hora
    # del/los marcaje(s) archivado(s) que se fusionaron en el) — para
    # mostrar en la celda del calendario que este guardia paso antes por
    # otro recinto ese mismo dia.
    origenes_fusion_txt: dict[int, str] = {}
    for r in registros_qr_archivados:
        if not r.fusionado_con_id:
            continue
        pieza = f"{r.recinto} ({r.registrado_at.strftime('%H:%M')})"
        previo = origenes_fusion_txt.get(r.fusionado_con_id, "")
        origenes_fusion_txt[r.fusionado_con_id] = f"{previo}; {pieza}" if previo else pieza

    matrix_qr: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_qr:
        d = r.registrado_at.day
        matrix_qr[r.recinto][d].append({
            "id":     r.id,
            "nombre": r.nombre_guardia,
            "turno":  r.tipo_turno,
            "hora":   r.registrado_at.strftime("%H:%M"),
            "fusion_origen": origenes_fusion_txt.get(r.id, ""),
        })

    # dia -> {nombres_norm} con QR en CUALQUIER recinto (sin importar el
    # turno) — para que un guardia que el supervisor anoto en un recinto
    # pero marco QR en OTRO no salga como "Falta o Inasistencia" en el
    # primero: eso ya es un "conflicto de destino" (mas abajo), no una
    # ausencia — el guardia si se presento a trabajar. Se ignora el turno
    # a proposito: si ademas el turno registrado difiere, sigue siendo
    # conflicto de destino, no falta (pedido explicito, ago 2026).
    qr_por_dia: dict[int, set] = defaultdict(set)
    # dia -> nombre_norm -> {recintos, ...} con QR ese dia — para poder nombrar
    # el recinto real en el mensaje de "conflicto de destino" de cada celda.
    qr_por_dia_recintos: dict[int, dict[str, set]] = defaultdict(lambda: defaultdict(set))
    for r in registros_qr:
        d = r.registrado_at.day
        nn_r = _norm(r.nombre_guardia)
        qr_por_dia[d].add(nn_r)
        qr_por_dia_recintos[d][nn_r].add(r.recinto)

    matrix_sv: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_sv:
        matrix_sv[r.recinto][r.fecha.day].append({
            "id":         r.id,
            "nombre":     r.nombre_guardia,
            "turno":      r.tipo_turno,
            "supervisor": r.supervisor or "",
            "notas":      r.notas or "",
        })

    grupo = grupo.lower().strip()
    if grupo == "privados":
        qr_sucursales   = _listar_recintos_privados(db)
        sucursal_labels = [s["label"] for s in qr_sucursales]
        recintos        = [s["label"] for s in qr_sucursales]
    elif grupo == "concon":
        qr_sucursales   = _listar_recintos_concon(db)
        sucursal_labels = [s["label"] for s in qr_sucursales]
        recintos        = [s["label"] for s in qr_sucursales]
    else:
        grupo = "quintero"
        qr_sucursales    = _listar_recintos_qr(db)
        sucursal_labels  = [s["label"] for s in qr_sucursales]
        privados_labels  = {s["label"] for s in _listar_recintos_privados(db)}
        concon_labels    = {s["label"] for s in _listar_recintos_concon(db)}
        all_recintos     = set(matrix_qr) | set(matrix_sv)
        extra_recintos   = sorted(
            r for r in all_recintos
            if r not in sucursal_labels and r not in privados_labels and r not in concon_labels
        )
        recintos         = sucursal_labels + extra_recintos

    recintos_set = set(recintos)
    tipo_guardia_lookup = _tipo_guardia_lookup(db)
    rows_permiso_sin_goce = _agregar_faltas_por_permiso_sin_goce(
        db,
        [],
        rut_lookup,
        tipo_guardia_lookup,
        year,
        month,
        control_hasta,
        recintos_permitidos=recintos_set,
    )
    alertas_permiso_sin_goce = _alertas_permiso_sin_goce(rows_permiso_sin_goce, month)

    days_in_month = calendar.monthrange(year, month)[1]
    days_info     = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()], "weekend": date(year, month, d).weekday() >= 5}
        for d in range(1, days_in_month + 1)
    ]

    cruce: dict[str, dict[int, dict]] = {}
    for recinto in recintos:
        cruce[recinto] = {}
        for d in range(1, days_in_month + 1):
            fecha_celda = date(year, month, d)
            control_activo = fecha_celda <= control_hasta
            qr_entries = list(matrix_qr.get(recinto, {}).get(d, []))
            sv_entries = list(matrix_sv.get(recinto, {}).get(d, []))

            qr_por_nombre: dict[str, list] = {}
            for e in qr_entries:
                qr_por_nombre.setdefault(_norm(e["nombre"]), []).append(e)
            sv_nombres = {_norm(e["nombre"]) for e in sv_entries}

            sv_annotated = []
            for e in sv_entries:
                nn = _norm(e["nombre"])
                if nn not in qr_por_nombre:
                    # Antes esto marcaba "Falta" con solo no encontrar QR en
                    # ESTE recinto, aunque el guardia si haya marcado QR ese
                    # mismo dia y turno en otro recinto — eso ya se marca
                    # aparte como "conflicto de destino"; el guardia SI se
                    # presento a trabajar, no esta ausente.
                    tiene_qr_en_otro_recinto = nn in qr_por_dia.get(d, set())
                    if tiene_qr_en_otro_recinto:
                        sv_annotated.append({**e, "falta": False, "es_permiso": False, "turno_mismatch": False, "turno_qr": None, "conflicto_destino": True})
                        continue
                    rut_e = rut_lookup.get(nn, "")
                    rangos = permisos_sin_goce.get(_normalizar_rut(rut_e), []) if rut_e else []
                    es_permiso = control_activo and any(desde <= fecha_celda <= hasta for desde, hasta in rangos)
                    sv_annotated.append({**e, "falta": control_activo, "es_permiso": es_permiso, "turno_mismatch": False, "turno_qr": None, "conflicto_destino": False})
                else:
                    turnos_qr = [qe["turno"] for qe in qr_por_nombre[nn]]
                    turno_ok = any(_norm(e["turno"]) == _norm(t) for t in turnos_qr)
                    # Aunque el guardia SI tenga QR en este recinto, si ademas
                    # tiene QR en otro(s) recinto(s) el mismo dia tambien es
                    # un conflicto de destino — antes solo se marcaba cuando
                    # NO se encontraba localmente, y un caso como este (QR+SV
                    # calzando en dos recintos distintos el mismo dia, ej.
                    # Cemco y Parque Municipal) pasaba como "ok" sin ninguna
                    # alerta (reportado por el cliente, ago 2026).
                    otros_recintos_qr = qr_por_dia_recintos.get(d, {}).get(nn, set()) - {recinto}
                    sv_annotated.append({**e, "falta": False, "es_permiso": False, "turno_mismatch": control_activo and not turno_ok,
                                         "turno_qr": " / ".join(turnos_qr) if control_activo and not turno_ok else None,
                                         "conflicto_destino": control_activo and bool(otros_recintos_qr)})

            qr_annotated = [{**e, "sin_supervisor": control_activo and _norm(e["nombre"]) not in sv_nombres} for e in qr_entries]

            razones: list[str] = []
            for e in sv_annotated:
                if e["falta"]:
                    sufijo = " (Permiso sin Goce)" if e["es_permiso"] else ""
                    razones.append(f"Falta — {e['nombre']} ({e['turno']}): sin registro de guardia{sufijo}")
                elif e["turno_mismatch"]:
                    razones.append(f"Turno diferente — {e['nombre']}: Supervisor={e['turno']}, Registro Guardia={e['turno_qr']}")
                elif e["conflicto_destino"]:
                    otros = sorted(qr_por_dia_recintos.get(d, {}).get(_norm(e["nombre"]), set()) - {recinto})
                    destino = otros[0] if otros else "otro recinto"
                    razones.append(f"Conflicto de destino — {e['nombre']} ({e['turno']}): registrado también en {destino} ese día")
            for e in qr_annotated:
                if e["sin_supervisor"]:
                    razones.append(f"Solo en Registro Guardia — {e['nombre']} ({e['turno']}, {e['hora']}): no está en registro del supervisor")

            n_faltas        = sum(1 for e in sv_annotated if e["falta"])
            n_turno_mismatch = sum(1 for e in sv_annotated if e["turno_mismatch"])
            n_solo_qr       = sum(1 for e in qr_annotated if e["sin_supervisor"])
            n_conflicto     = sum(1 for e in sv_annotated if e["conflicto_destino"])

            if not control_activo:
                status = "ok" if qr_entries and sv_entries else "empty"
            elif not qr_entries and not sv_entries:
                status = "empty"
            elif qr_entries and not sv_entries:
                status = "solo_qr"
            elif n_faltas:
                status = "falta"
            elif n_turno_mismatch or n_solo_qr:
                status = "mismatch"
            elif n_conflicto:
                status = "conflicto"
            else:
                status = "ok"

            cruce[recinto][d] = {"qr": qr_annotated, "sv": sv_annotated, "status": status, "razones": razones}

    # ── Alertas de cruce: inasistencias y discrepancias ──────────────────────────
    alertas_inasistencia: list[dict] = []
    alertas_discrepancia: list[dict] = []

    for rec, dias in cruce.items():
        for d, cell in dias.items():
            fecha_str = f"{d}/{month:02d}"
            for razon in cell.get("razones", []):
                if razon.startswith("Falta"):
                    es_permiso = "(Permiso sin Goce)" in razon
                    detalle = razon[8:].replace(" (Permiso sin Goce)", "")
                    alertas_inasistencia.append({
                        "fecha": fecha_str,
                        "recinto": rec,
                        "detalle": detalle,
                        "permiso_sin_goce": es_permiso,
                    })
                elif razon.startswith("Turno diferente"):
                    alertas_discrepancia.append({"tipo": "Turno diferente", "fecha": fecha_str, "recinto": rec, "detalle": razon[18:]})
                elif razon.startswith("Solo en Registro Guardia"):
                    alertas_discrepancia.append({"tipo": "Sin confirmación supervisor", "fecha": fecha_str, "recinto": rec, "detalle": razon[27:]})

    def _fecha_alerta_key(valor: object) -> str:
        partes = str(valor or "").split("/")
        if len(partes) >= 2:
            try:
                return f"{int(partes[0]):02d}/{int(partes[1]):02d}"
            except ValueError:
                pass
        return str(valor or "").strip()

    alertas_normales = {
        (_fecha_alerta_key(a.get("fecha")), _normalizar_texto(str(a.get("detalle") or "").split("(", 1)[0]))
        for a in alertas_inasistencia
    }
    for alerta in alertas_permiso_sin_goce:
        key = (_fecha_alerta_key(alerta.get("fecha")), _normalizar_texto(alerta.get("detalle")))
        if key not in alertas_normales:
            alertas_inasistencia.append(alerta)

    alertas_inasistencia.sort(key=lambda x: x["fecha"])
    alertas_discrepancia.sort(key=lambda x: x["fecha"])

    # ── Conflictos de recinto: mismo guardia, distinto recinto ────────────────────
    # qr_global_any[dia][nombre_norm] = {recinto, ...} — sin filtrar por turno,
    # para que un cruce de recinto CON turno distinto tambien cuente como
    # "conflicto de destino" (y no desaparezca de toda alerta) en vez de exigir
    # que ademas coincida el turno (pedido explicito, ago 2026).
    from collections import defaultdict as _dd
    qr_global_any: dict[int, dict] = _dd(lambda: _dd(set))
    for rec in recintos:
        for d in range(1, days_in_month + 1):
            for e in matrix_qr.get(rec, {}).get(d, []):
                qr_global_any[d][_norm(e["nombre"])].add(rec)

    sv_global: dict[int, dict] = _dd(lambda: _dd(dict))
    for rec in recintos:
        for d in range(1, days_in_month + 1):
            for e in matrix_sv.get(rec, {}).get(d, []):
                key = (_norm(e["nombre"]), _norm(e["turno"]))
                sv_global[d][key][rec] = {"nombre": e["nombre"], "turno": e["turno"]}

    alertas_conflicto: list[dict] = []
    seen_conf: set = set()
    for d in range(1, days_in_month + 1):
        for (nn, nt), sv_rec_dict in sv_global[d].items():
            qr_recs = qr_global_any[d].get(nn, set())
            for sv_rec in sv_rec_dict:
                for qr_rec in qr_recs:
                    if sv_rec != qr_rec:
                        key = (d, nn, nt, sv_rec, qr_rec)
                        if key not in seen_conf:
                            seen_conf.add(key)
                            info = sv_rec_dict[sv_rec]
                            alertas_conflicto.append({
                                "fecha":      f"{d}/{month:02d}",
                                "guardia":    info["nombre"],
                                "turno":      info["turno"],
                                "recinto_sv": sv_rec,
                                "recinto_qr": qr_rec,
                            })
    alertas_conflicto.sort(key=lambda x: x["fecha"])

    # ── Alertas: guardias con 5+ días consecutivos en el mes (ventana ±4 días) ──
    _td = timedelta

    # Ampliar ventana al mes anterior/siguiente para capturar rachas que cruzan mes
    window_start = date(year, month, 1) - _td(days=6)
    window_end   = date(year, month, days_in_month) + _td(days=6)
    registros_ventana = (
        db.query(InicioTurnoRegistro)
        .filter(
            InicioTurnoRegistro.registrado_at >= datetime(window_start.year, window_start.month, window_start.day),
            InicioTurnoRegistro.registrado_at <= datetime(window_end.year, window_end.month, window_end.day, 23, 59),
        )
        .all()
    )

    guard_dates: dict[str, set] = _dd(set)
    for r in registros_ventana:
        guard_dates[r.nombre_guardia].add(r.registrado_at.date())

    STREAK_MIN = 5
    alertas: list[dict] = []
    for nombre, fechas in guard_dates.items():
        sorted_fechas = sorted(fechas)
        streak_start = sorted_fechas[0]
        streak = [streak_start]
        all_streaks = []
        for i in range(1, len(sorted_fechas)):
            if (sorted_fechas[i] - sorted_fechas[i - 1]).days == 1:
                streak.append(sorted_fechas[i])
            else:
                all_streaks.append(streak)
                streak = [sorted_fechas[i]]
        all_streaks.append(streak)

        for s in all_streaks:
            if len(s) >= STREAK_MIN:
                # Solo alertar si la racha toca el mes actual
                if s[-1] >= date(year, month, 1) and s[0] <= date(year, month, days_in_month):
                    alertas.append({
                        "nombre": nombre,
                        "dias":   len(s),
                        "desde":  f"{s[0].day}/{s[0].month}",
                        "hasta":  f"{s[-1].day}/{s[-1].month}",
                    })

    alertas.sort(key=lambda x: -x["dias"])

    prev_y, prev_m = (year - 1, 12) if month == 1  else (year, month - 1)
    next_y, next_m = (year + 1, 1)  if month == 12 else (year, month + 1)

    return templates.TemplateResponse(
        request,
        "tabla_registro_guardia.html",
        {
            "request":    request,
            "year":       year,
            "month":      month,
            "month_name": _MONTH_NAMES[month - 1],
            "recintos":   recintos,
            "days_info":  days_info,
            "matrix":     {r: dict(matrix_qr[r]) for r in recintos},
            "cruce":               cruce,
            "alertas":             alertas,
            "alertas_inasistencia": alertas_inasistencia,
            "alertas_discrepancia": alertas_discrepancia,
            "alertas_conflicto":    alertas_conflicto,
            "candidatos_fusion":    candidatos_fusion,
            "fusionados_mes":       fusionados_mes,
            "prev_y": prev_y, "prev_m": prev_m,
            "next_y": next_y, "next_m": next_m,
            "today_day": today.day if (today.year == year and today.month == month) else -1,
            "grupo": grupo,
            "tipos_turno": list(TIPOS_TURNO),
            "supervisor_nombre": _nombre_usuario_sesion(request, db),
        },
    )


@router.get("/guardia/tabla-registro-guardia/informes/preview", response_class=HTMLResponse)
def preview_informes_guardias(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    db: Session = Depends(get_db),
):
    today = date.today()
    year = year or today.year
    month = month or today.month
    data = _datos_informe_guardias(db, year, month, grupo)
    faltas_por_guardia = {
        (row["rut"] or _normalizar_texto(row["nombre"])): int(row["faltas"])
        for row in data["rows_faltas"]
    }

    sections = [
        _preview_turnos_section("Registro Guardias", data["rows_qr"], faltas_por_guardia=faltas_por_guardia),
        _preview_turnos_section("Supervisores", data["rows_sv"]),
        _preview_cruce_section(data["rows_cruce"]),
        _preview_faltas_section(data["rows_faltas"]),
        _preview_domingos_section(data["rows_domingos"]),
        _preview_supervisor_calendario_section(
            "Guardia por Recinto",
            month,
            data["recintos"],
            data["days_info"],
            data["matrix_qr"],
            solo_nombres=True,
        ),
        _preview_turnos_extra_section("Turno Extra", data["rows_extra"], tipo_turno="Extra"),
        _preview_turnos_extra_section(
            "Contrato Diario",
            data["rows_contrato_diario"],
            tipo_turno="Contrato Diario",
        ),
        _preview_justificaciones_section(data["rows_justificaciones"]),
    ]

    return templates.TemplateResponse(
        request,
        "guardias_informe_preview.html",
        {
            "request": request,
            "titulo": "Vista previa - Informes de guardias",
            "periodo": data["periodo"],
            "sections": sections,
        },
    )


@router.get("/guardia/tabla-registro-guardia/informes")
def descargar_informes_guardias(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    preview: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook

    today = date.today()
    year = year or today.year
    month = month or today.month
    if preview:
        return preview_informes_guardias(request, year=year, month=month, grupo=grupo, db=db)

    registros_qr = (
        db.query(InicioTurnoRegistro)
        .filter(
            extract("year", InicioTurnoRegistro.registrado_at) == year,
            extract("month", InicioTurnoRegistro.registrado_at) == month,
        )
        .order_by(InicioTurnoRegistro.registrado_at, InicioTurnoRegistro.id)
        .all()
    )
    # Marcajes fusionados (ver _detectar_candidatos_fusion) quedan afuera de
    # todo conteo/informe — el turno real es el marcaje sobreviviente.
    registros_qr = [r for r in registros_qr if str(r.estado or "activo") != "archivado"]
    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year", SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    grupo, recintos = _recintos_para_grupo(db, grupo, registros_qr, registros_sv)
    recintos_set = set(recintos)
    registros_qr = [r for r in registros_qr if str(r.recinto or "").strip() in recintos_set]
    registros_sv = [r for r in registros_sv if str(r.recinto or "").strip() in recintos_set]

    rut_lookup = _rut_lookup_guardias(db, registros_qr, registros_sv)
    tipo_guardia_lookup = _tipo_guardia_lookup(db)
    rows_qr = _agrupar_turnos_por_guardia(
        registros_qr,
        rut_lookup,
        tipo_guardia_lookup,
        usa_rut_registro=True,
    )
    rows_sv = _agrupar_turnos_por_guardia(
        registros_sv,
        rut_lookup,
        tipo_guardia_lookup,
        usa_rut_registro=False,
    )
    # El cruce (disyuntivas) solo tiene sentido para dias YA TERMINADOS: si el
    # dia de hoy (o el futuro) aun no se registra por completo, no es una
    # discrepancia real, es que el turno todavia no ocurre o no termina.
    registros_qr_pasado = [r for r in registros_qr if r.registrado_at.date() < today]
    registros_sv_pasado = [r for r in registros_sv if r.fecha < today]
    rows_cruce = _agrupar_cruce_guardias(registros_qr_pasado, registros_sv_pasado, rut_lookup, tipo_guardia_lookup)
    permisos_sin_goce = _permisos_sin_goce_lookup(db, year, month)
    rows_faltas = _agrupar_faltas_guardias(registros_qr_pasado, registros_sv_pasado, rut_lookup, tipo_guardia_lookup, permisos_sin_goce)
    rows_faltas = _agregar_faltas_por_permiso_sin_goce(
        db,
        rows_faltas,
        rut_lookup,
        tipo_guardia_lookup,
        year,
        month,
        today - timedelta(days=1),
        recintos_permitidos=recintos_set,
    )
    rows_extra = _agrupar_turnos_extra(registros_qr, registros_sv, rut_lookup, tipo_guardia_lookup, tipo_turno="Extra")
    rows_contrato_diario = _agrupar_turnos_extra(registros_qr, registros_sv, rut_lookup, tipo_guardia_lookup, tipo_turno="Contrato Diario")
    rows_domingos = _agrupar_domingos_guardias(registros_qr, rut_lookup, tipo_guardia_lookup)
    rows_justificaciones = _justificaciones_lookup(db, year, month, tipo_guardia_lookup)

    matrix_sv: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_sv:
        matrix_sv[r.recinto][r.fecha.day].append({
            "nombre": r.nombre_guardia,
            "turno": r.tipo_turno,
            "notas": r.notas or "",
        })
    matrix_qr: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_qr:
        matrix_qr[r.recinto][r.registrado_at.day].append({
            "id": r.id,
            "nombre": r.nombre_guardia,
            "turno": r.tipo_turno,
            "hora": r.registrado_at.strftime("%H:%M"),
        })
    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()]}
        for d in range(1, days_in_month + 1)
    ]

    faltas_por_guardia = {
        (row["rut"] or _normalizar_texto(row["nombre"])): int(row["faltas"])
        for row in rows_faltas
    }

    periodo = f"{_MONTH_NAMES[month - 1]} {year} | Grupo: {grupo.title()}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro Guardias"
    _crear_hoja_informe_turnos(ws, "Registro Guardias", periodo, rows_qr, faltas_por_guardia=faltas_por_guardia)
    _crear_hoja_informe_turnos(wb.create_sheet("Supervisores"), "Supervisores", periodo, rows_sv)
    _crear_hoja_cruce(wb.create_sheet("Cruce"), periodo, rows_cruce)
    _crear_hoja_faltas(wb.create_sheet("Faltas"), periodo, rows_faltas)
    _crear_hoja_domingos(wb.create_sheet("Conteo días domingo"), periodo, rows_domingos)
    _crear_hoja_supervisor_calendario(
        wb.create_sheet("Guardia por Recinto"),
        periodo,
        month,
        recintos,
        days_info,
        matrix_qr,
        titulo="Guardia por Recinto",
        solo_nombres=True,
    )
    _crear_hoja_turno_extra(wb.create_sheet("Turno Extra"), periodo, rows_extra, tipo_turno="Extra")
    _crear_hoja_turno_extra(wb.create_sheet("Contrato Diario"), periodo, rows_contrato_diario, tipo_turno="Contrato Diario")
    _crear_hoja_justificaciones(wb.create_sheet("Justificaciones"), periodo, rows_justificaciones)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"informes_guardias_{year}_{month:02d}_{grupo}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/api/inicio-turno/guardia")
def buscar_guardia(
    rut: str = Query(min_length=1),
    recinto_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
):
    guardia = _buscar_guardia_por_rut(db, rut, recinto_id)
    if not guardia:
        return {"found": False, "nombre": "", "rut": _normalizar_rut(rut), "tipo_contrato": ""}
    tipo_contrato = _tipo_guardia_para(guardia.nombre, guardia.rut, _tipo_guardia_lookup(db))
    return {
        "found": True,
        "nombre": str(guardia.nombre or "").strip(),
        "rut": str(guardia.rut or "").strip(),
        "sucursal_id": recinto_id,
        "tipo_contrato": tipo_contrato,
    }


@router.get("/api/inicio-turno/guardias/buscar")
def buscar_guardias_por_nombre(
    q: str = Query(default=""),
    db: Session = Depends(get_db),
):
    q = q.strip()
    if len(q) < 2:
        return []
    resultados = _buscar_guardias_legacy(db, q=q, limit=10)
    merged: dict[str, dict[str, str]] = {}
    for guardia in resultados:
        rut = str(guardia.rut or "").strip()
        nombre = str(guardia.nombre or "").strip()
        key = _normalizar_rut(rut) or nombre.casefold()
        if key and key not in merged:
            merged[key] = {"nombre": nombre, "rut": rut}
    return sorted(merged.values(), key=lambda item: item["nombre"].casefold())[:10]


@router.post("/api/inicio-turno")
def registrar_inicio_turno(
    payload: InicioTurnoCreate,
    request: Request,
    db: Session = Depends(get_db),
):
    tipo_turno = str(payload.tipo_turno or "").strip()

    if not tipo_turno:
        # Ventanas de marcaje normal del guardia. Fuera de estas franjas
        # (o si elige Extra/Contrato Diario manualmente) no se autodetecta.
        from datetime import datetime as _dt
        from datetime import time as _time
        _ahora = _dt.now().time()
        if _time(6, 0) <= _ahora <= _time(8, 30):
            tipo_turno = "Dia"
        elif _time(16, 0) <= _ahora <= _time(21, 0):
            tipo_turno = "Noche"
        else:
            _hora = _ahora.hour
            tipo_turno = "Noche" if (_hora >= 19 or _hora < 8) else "Dia"
    elif tipo_turno not in _TIPOS_MANUALES:
        raise HTTPException(status_code=400, detail="Tipo de turno invalido")

    sucursal = db.get(SucursalBBDD, payload.sucursal_id) if payload.sucursal_id else None
    recinto = _recinto_label(sucursal) if sucursal else str(payload.recinto or "").strip()
    if not recinto:
        raise HTTPException(status_code=400, detail="Recinto requerido")

    if _UBICACION_INICIO_TURNO_HABILITADA and (
        str(payload.ubicacion_estado or "").strip() != "confirmada"
        or payload.latitud is None
        or payload.longitud is None
    ):
        raise HTTPException(status_code=400, detail="Ubicacion obligatoria para iniciar turno")

    if _UBICACION_INICIO_TURNO_HABILITADA and sucursal:
        recinto_lat, recinto_lng = _obtener_o_geocodificar_sucursal(db, sucursal)
        if recinto_lat is not None and recinto_lng is not None:
            distancia = _distancia_metros(payload.latitud, payload.longitud, recinto_lat, recinto_lng)
            if distancia > RADIO_MAXIMO_METROS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Estas a {round(distancia)} metros del recinto. Maximo permitido: {round(RADIO_MAXIMO_METROS)} metros",
                )
    elif _UBICACION_INICIO_TURNO_HABILITADA and not sucursal:
        recinto_lat, recinto_lng = _coords_recinto_estatico(recinto)
        if recinto_lat is not None and recinto_lng is not None:
            distancia = _distancia_metros(payload.latitud, payload.longitud, recinto_lat, recinto_lng)
            if distancia > RADIO_MAXIMO_METROS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Estas a {round(distancia)} metros del recinto. Maximo permitido: {round(RADIO_MAXIMO_METROS)} metros",
                )

    guardia = _buscar_guardia_por_rut(db, payload.rut, payload.sucursal_id)
    if not guardia:
        raise HTTPException(status_code=404, detail="No existe un guardia registrado con ese RUT")

    registro = InicioTurnoRegistro(
        rut=_normalizar_rut(guardia.rut or payload.rut),
        nombre_guardia=str(guardia.nombre or "").strip(),
        tipo_turno=tipo_turno,
        recinto=recinto,
        sucursal_id=sucursal.id if sucursal else payload.sucursal_id,
        latitud=payload.latitud,
        longitud=payload.longitud,
        precision_metros=payload.precision_metros,
        ubicacion_estado=(
            str(payload.ubicacion_estado or "").strip() or None
            if _UBICACION_INICIO_TURNO_HABILITADA
            else "inhabilitada"
        ),
        ip_origen=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)

    # Notificación de Turno Extra / Contrato Diario (misma que la del registro
    # del supervisor; para Dia/Noche la función no envía nada).
    _enviar_notificacion_turno_async(
        tipo_turno=registro.tipo_turno,
        nombre_guardia=registro.nombre_guardia,
        rut=registro.rut,
        recinto=registro.recinto,
        fecha_turno=registro.registrado_at.strftime("%d/%m/%Y"),
        fecha_registro=registro.registrado_at,
        supervisor=None,
        nota=None,
        origen="el marcaje QR del guardia",
        registro_id=f"qr-{registro.id}",
    )

    # Fusion automatica de marcajes duplicados por traslado (ver
    # _fusionar_automatico_si_corresponde) — corre antes de la alerta de
    # doble marcaje para que esta ya no dispare por algo que se acaba de
    # resolver solo.
    try:
        _fusionar_automatico_si_corresponde(db, registro)
    except Exception as exc:
        _log.warning("No se pudo evaluar fusion automatica para registro %s: %s", registro.id, exc)

    # Alerta de doble marcaje: mismo guardia y mismo dia, sin importar turno,
    # sucursal ni recinto.
    try:
        _alertar_doble_marcaje_si_corresponde(db, registro)
    except Exception as exc:
        _log.warning("No se pudo evaluar doble marcaje para registro %s: %s", registro.id, exc)

    return {
        "ok": True,
        "id": registro.id,
        "rut": registro.rut,
        "nombre": registro.nombre_guardia,
        "tipo_turno": registro.tipo_turno,
        "recinto": registro.recinto,
        "registrado_at": registro.registrado_at.isoformat(sep=" ", timespec="seconds"),
    }


# ── Tabla supervisor ──────────────────────────────────────────────────────────

class SupervisorRegistroPayload(BaseModel):
    recinto: str
    fecha: str          # YYYY-MM-DD
    nombre_guardia: str
    rut: str = ""
    tipo_turno: str
    supervisor: str = ""
    notas: str = ""


def _usuario_sesion(request: Request, db) -> User | None:
    try:
        cookie = request.cookies.get(_COOKIE_NAME, "")
        if cookie:
            login = _decode_web_token(cookie)
            user = _UserService.find_by_login(db, login)
            if user and user.is_active:
                return user
    except Exception:
        pass
    return None


def _nombre_usuario_sesion(request: Request, db) -> str:
    user = _usuario_sesion(request, db)
    if user:
        return str(user.name or "").strip()
    return ""


def _es_solo_supervisor_quintero(user: User | None) -> bool:
    if not user or getattr(user, "is_admin", False):
        return False
    departamentos = {
        parte.strip().casefold()
        for parte in str(user.department or "").split(";")
        if parte.strip()
    }
    return "supervisorquintero" in departamentos


def _es_solo_supervisor_concon(user: User | None) -> bool:
    if not user or getattr(user, "is_admin", False):
        return False
    departamentos = {
        parte.strip().casefold()
        for parte in str(user.department or "").split(";")
        if parte.strip()
    }
    return "supervisorconcon" in departamentos


@router.get("/guardia/tabla-supervisor", response_class=HTMLResponse)
def guardia_tabla_supervisor_page(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    origen: str = Query(default="supervisores"),
    db: Session = Depends(get_db),
):
    today = date.today()
    year  = year  or today.year
    month = month or today.month

    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year",  SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    matrix: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_sv:
        matrix[r.recinto][r.fecha.day].append({
            "id":     r.id,
            "nombre": r.nombre_guardia,
            "turno":  r.tipo_turno,
            "supervisor": r.supervisor or "",
            "notas":  r.notas or "",
        })

    grupo = grupo.lower().strip()
    current_user = _usuario_sesion(request, db)
    if _es_solo_supervisor_quintero(current_user) and grupo in {"concon", "privados"}:
        return RedirectResponse(url=f"/guardia/tabla-supervisor?year={year}&month={month}&grupo=quintero", status_code=303)
    if _es_solo_supervisor_concon(current_user) and grupo in {"quintero", "privados"}:
        return RedirectResponse(url=f"/guardia/tabla-supervisor?year={year}&month={month}&grupo=concon", status_code=303)

    if grupo == "privados":
        recintos = [s["label"] for s in _listar_recintos_privados(db)]
        template_name = "registro_supervisor_privado.html"
        titulo_tabla = "Registro Supervisor — Privados"
        subtitulo_tabla = "Guardias registrados manualmente por el supervisor · Recintos privados"
    elif grupo == "concon":
        recintos = [s["label"] for s in _listar_recintos_concon(db)]
        template_name = "registro_supervisor_quintero.html"
        titulo_tabla = "Registro Supervisor — Concón"
        subtitulo_tabla = "Guardias registrados manualmente por el supervisor · Municipalidad de Concón"
    else:
        grupo = "quintero"
        qr_sucursales   = _listar_recintos_qr(db)
        sucursal_labels = [s["label"] for s in qr_sucursales]
        privados_labels = {s["label"] for s in _listar_recintos_privados(db)}
        concon_labels   = {s["label"] for s in _listar_recintos_concon(db)}
        extra_recintos  = sorted(
            r for r in matrix
            if r not in sucursal_labels and r not in privados_labels and r not in concon_labels
        )
        recintos        = sucursal_labels + extra_recintos
        template_name = "registro_supervisor_quintero.html"
        titulo_tabla = "Registro Supervisor — Quintero"
        subtitulo_tabla = "Guardias registrados manualmente por el supervisor · Municipalidad de Quintero"

    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()], "weekend": date(year, month, d).weekday() >= 5}
        for d in range(1, days_in_month + 1)
    ]

    prev_y, prev_m = (year - 1, 12) if month == 1  else (year, month - 1)
    next_y, next_m = (year + 1, 1)  if month == 12 else (year, month + 1)

    origen = origen.lower().strip()
    if origen not in ("guardia", "supervisores"):
        origen = "supervisores"

    return templates.TemplateResponse(
        request,
        template_name,
        {
            "request":    request,
            "year": year, "month": month,
            "month_name": _MONTH_NAMES[month - 1],
            "recintos":   recintos,
            "days_info":  days_info,
            "matrix":     {r: dict(matrix[r]) for r in recintos},
            "tipos_turno": list(TIPOS_TURNO),
            "prev_y": prev_y, "prev_m": prev_m,
            "next_y": next_y, "next_m": next_m,
            "today_day": today.day if (today.year == year and today.month == month) else -1,
            "supervisor_nombre": _nombre_usuario_sesion(request, db),
            "grupo": grupo,
            "origen": origen,
            "titulo_tabla": titulo_tabla,
            "subtitulo_tabla": subtitulo_tabla,
        },
    )


@router.get("/guardia/tabla-supervisor/informes/preview", response_class=HTMLResponse)
def preview_informe_tabla_supervisor(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    db: Session = Depends(get_db),
):
    today = date.today()
    year = year or today.year
    month = month or today.month

    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year", SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    grupo, recintos = _recintos_para_grupo(db, grupo, registros_sv=registros_sv)

    matrix: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    recintos_set = set(recintos)
    for r in registros_sv:
        if str(r.recinto or "").strip() not in recintos_set:
            continue
        matrix[r.recinto][r.fecha.day].append({
            "nombre": r.nombre_guardia,
            "turno": r.tipo_turno,
            "notas": r.notas or "",
        })

    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()]}
        for d in range(1, days_in_month + 1)
    ]
    periodo = f"{_MONTH_NAMES[month - 1]} {year} | Grupo: {grupo.title()}"
    sections = [
        _preview_supervisor_calendario_section(
            "Registro Supervisor",
            month,
            recintos,
            days_info,
            matrix,
        )
    ]

    return templates.TemplateResponse(
        request,
        "guardias_informe_preview.html",
        {
            "request": request,
            "titulo": "Vista previa - Registro supervisor",
            "periodo": periodo,
            "sections": sections,
        },
    )


@router.get("/guardia/tabla-supervisor/informes")
def descargar_informe_tabla_supervisor(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    grupo: str = Query(default="quintero"),
    preview: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook

    today = date.today()
    year = year or today.year
    month = month or today.month
    if preview:
        return preview_informe_tabla_supervisor(request, year=year, month=month, grupo=grupo, db=db)

    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year", SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    grupo, recintos = _recintos_para_grupo(db, grupo, registros_sv=registros_sv)

    matrix: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_sv:
        matrix[r.recinto][r.fecha.day].append({
            "nombre": r.nombre_guardia,
            "turno": r.tipo_turno,
            "notas": r.notas or "",
        })

    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()]}
        for d in range(1, days_in_month + 1)
    ]

    periodo = f"{_MONTH_NAMES[month - 1]} {year} | Grupo: {grupo.title()}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro Supervisor"
    _crear_hoja_supervisor_calendario(ws, periodo, month, recintos, days_info, matrix)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"registro_supervisor_{year}_{month:02d}_{grupo}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/guardia/cruce", response_class=HTMLResponse)
def guardia_cruce_page(
    request: Request,
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: Session = Depends(get_db),
):
    today = date.today()
    year  = year  or today.year
    month = month or today.month

    # QR registros
    registros_qr = (
        db.query(InicioTurnoRegistro)
        .filter(
            extract("year",  InicioTurnoRegistro.registrado_at) == year,
            extract("month", InicioTurnoRegistro.registrado_at) == month,
        )
        .order_by(InicioTurnoRegistro.registrado_at)
        .all()
    )

    # Supervisor registros
    registros_sv = (
        db.query(SupervisorRegistro)
        .filter(
            extract("year",  SupervisorRegistro.fecha) == year,
            extract("month", SupervisorRegistro.fecha) == month,
        )
        .order_by(SupervisorRegistro.fecha, SupervisorRegistro.id)
        .all()
    )

    def _norm(name: str) -> str:
        import unicodedata
        n = unicodedata.normalize("NFD", (name or "").lower())
        return "".join(c for c in n if unicodedata.category(c) != "Mn").strip()

    # matrix_qr[recinto][day] = [{"nombre":..,"turno":..,"hora":..}]
    matrix_qr: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_qr:
        matrix_qr[r.recinto][r.registrado_at.day].append({
            "nombre": r.nombre_guardia,
            "turno":  r.tipo_turno,
            "hora":   r.registrado_at.strftime("%H:%M"),
        })

    # matrix_sv[recinto][day] = [{"nombre":..,"turno":..,"supervisor":..}]
    matrix_sv: dict[str, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    for r in registros_sv:
        matrix_sv[r.recinto][r.fecha.day].append({
            "nombre":     r.nombre_guardia,
            "turno":      r.tipo_turno,
            "supervisor": r.supervisor or "",
        })

    qr_sucursales   = _listar_recintos_qr(db)
    sucursal_labels = [s["label"] for s in qr_sucursales]
    privados_labels = {s["label"] for s in _listar_recintos_privados(db)}
    concon_labels   = {s["label"] for s in _listar_recintos_concon(db)}
    all_recintos    = set(matrix_qr) | set(matrix_sv)
    extra_recintos  = sorted(
        r for r in all_recintos
        if r not in sucursal_labels and r not in privados_labels and r not in concon_labels
    )
    recintos        = sucursal_labels + extra_recintos

    days_in_month = calendar.monthrange(year, month)[1]
    days_info = [
        {"day": d, "dow": _DAY_NAMES[date(year, month, d).weekday()], "weekend": date(year, month, d).weekday() >= 5}
        for d in range(1, days_in_month + 1)
    ]

    # Build cruce matrix: [recinto][day] = {"qr": [...], "sv": [...], "status": ..., "razones": [...]}
    cruce: dict[str, dict[int, dict]] = {}
    for recinto in recintos:
        cruce[recinto] = {}
        for d in range(1, days_in_month + 1):
            qr_entries = list(matrix_qr.get(recinto, {}).get(d, []))
            sv_entries = list(matrix_sv.get(recinto, {}).get(d, []))

            # Índice QR por nombre normalizado → lista de entradas
            qr_por_nombre: dict[str, list] = {}
            for e in qr_entries:
                qr_por_nombre.setdefault(_norm(e["nombre"]), []).append(e)

            sv_nombres = {_norm(e["nombre"]) for e in sv_entries}

            # Anotar cada entrada SV con falta / turno_mismatch / turno_qr
            sv_annotated = []
            for e in sv_entries:
                nn = _norm(e["nombre"])
                if nn not in qr_por_nombre:
                    sv_annotated.append({**e, "falta": True, "turno_mismatch": False, "turno_qr": None})
                else:
                    turnos_qr = [qe["turno"] for qe in qr_por_nombre[nn]]
                    turno_ok = any(_norm(e["turno"]) == _norm(t) for t in turnos_qr)
                    sv_annotated.append({
                        **e,
                        "falta": False,
                        "turno_mismatch": not turno_ok,
                        "turno_qr": " / ".join(turnos_qr) if not turno_ok else None,
                    })

            # Anotar cada entrada QR: sin_supervisor si el nombre no figura en sv
            qr_annotated = [{**e, "sin_supervisor": _norm(e["nombre"]) not in sv_nombres} for e in qr_entries]

            # Construir lista de razones legibles
            razones: list[str] = []
            for e in sv_annotated:
                if e["falta"]:
                    razones.append(f"Falta — {e['nombre']} ({e['turno']}): sin registro de guardia")
                elif e["turno_mismatch"]:
                    razones.append(f"Turno diferente — {e['nombre']}: Supervisor={e['turno']}, Registro Guardia={e['turno_qr']}")
            for e in qr_annotated:
                if e["sin_supervisor"]:
                    razones.append(f"Solo en Registro Guardia — {e['nombre']} ({e['turno']}, {e['hora']}): no está en registro del supervisor")

            # Determinar status
            n_faltas = sum(1 for e in sv_annotated if e["falta"])
            n_turno_mismatch = sum(1 for e in sv_annotated if e["turno_mismatch"])
            n_solo_qr = sum(1 for e in qr_annotated if e["sin_supervisor"])

            if not qr_entries and not sv_entries:
                status = "empty"
            elif qr_entries and not sv_entries:
                status = "solo_qr"
            elif not qr_entries and sv_entries:
                status = "falta"
            elif n_faltas:
                status = "falta"
            elif n_turno_mismatch or n_solo_qr:
                status = "mismatch"
            else:
                status = "ok"

            cruce[recinto][d] = {
                "qr": qr_annotated,
                "sv": sv_annotated,
                "status": status,
                "razones": razones,
            }

    prev_y, prev_m = (year - 1, 12) if month == 1  else (year, month - 1)
    next_y, next_m = (year + 1, 1)  if month == 12 else (year, month + 1)

    return templates.TemplateResponse(
        request,
        "tabla_cruce_guardia.html",
        {
            "request":    request,
            "year": year, "month": month,
            "month_name": _MONTH_NAMES[month - 1],
            "recintos":   recintos,
            "days_info":  days_info,
            "cruce":      cruce,
            "prev_y": prev_y, "prev_m": prev_m,
            "next_y": next_y, "next_m": next_m,
            "today_day": today.day if (today.year == year and today.month == month) else -1,
        },
    )


@router.get("/api/guardia/racha")
def obtener_racha_consecutiva(
    nombre: str = Query(default=""),
    fecha: str = Query(default=""),
    db: Session = Depends(get_db),
):
    """Cuantos dias seguidos quedaria trabajando este guardia si se guarda un
    registro de supervisor en `fecha` (la incluye como si ya estuviera
    guardada). Se usa para advertir ANTES de guardar, sin bloquear el
    registro."""
    try:
        fecha_dt = date.fromisoformat(fecha)
    except ValueError:
        raise HTTPException(status_code=400, detail="Fecha invalida")
    if not nombre.strip():
        raise HTTPException(status_code=400, detail="Nombre requerido")
    dias = _racha_consecutiva_incluyendo(db, nombre, fecha_dt)
    return {"dias_consecutivos": dias}


@router.post("/api/guardia/supervisor-registro")
def crear_supervisor_registro(
    payload: SupervisorRegistroPayload,
    db: Session = Depends(get_db),
):
    from datetime import date as _date
    try:
        fecha = _date.fromisoformat(payload.fecha)
    except ValueError:
        raise HTTPException(status_code=400, detail="Fecha invalida")
    if payload.tipo_turno not in TIPOS_TURNO:
        raise HTTPException(status_code=400, detail="Tipo de turno invalido")
    if not payload.nombre_guardia.strip():
        raise HTTPException(status_code=400, detail="Nombre requerido")
    if not payload.recinto.strip():
        raise HTTPException(status_code=400, detail="Recinto requerido")

    guardia_user = _resolver_guardia_activo_payload(db, payload.rut, payload.nombre_guardia)
    nombre_guardia = str(guardia_user.name or "").strip()
    rut_guardia = _normalizar_rut(guardia_user.username)
    if rut_guardia:
        conflicto = _justificacion_conflicto_vigente(db, rut_guardia, fecha, fecha)
        if conflicto:
            raise HTTPException(
                status_code=409,
                detail=f"{nombre_guardia} tiene una justificación vigente ({conflicto.motivo}, "
                        f"{conflicto.fecha_desde.strftime('%d/%m/%Y')}–{conflicto.fecha_hasta.strftime('%d/%m/%Y')}) "
                        f"— no se puede registrar ese día.",
            )

    reg = SupervisorRegistro(
        recinto=payload.recinto.strip(),
        fecha=fecha,
        nombre_guardia=nombre_guardia,
        tipo_turno=payload.tipo_turno,
        supervisor=payload.supervisor.strip() or None,
        notas=payload.notas.strip() or None,
    )
    db.add(reg)
    db.commit()
    db.refresh(reg)
    # Sin notificación por correo: los avisos de Turno Extra / Contrato Diario /
    # doble marcaje se envían SOLO desde el registro de inicio de turno del
    # guardia, nunca desde el registro del supervisor. La alerta de jornada
    # extendida es la excepción: se registra siempre desde acá, porque este es
    # el único flujo donde el supervisor deja constancia manual de un turno.
    dias_racha = _racha_consecutiva_incluyendo(db, nombre_guardia, fecha)
    if dias_racha >= 6:
        _enviar_alerta_jornada_extendida_async(
            nombre_guardia=nombre_guardia,
            rut=rut_guardia,
            recinto=reg.recinto,
            dias_consecutivos=dias_racha,
            fecha=fecha,
            origen="registro de supervisor",
            registro_id=f"sv-{reg.id}",
        )
    return {"ok": True, "id": reg.id, "nombre": reg.nombre_guardia, "turno": reg.tipo_turno}


@router.delete("/api/guardia/supervisor-registro/{registro_id}")
def eliminar_supervisor_registro(
    registro_id: int,
    db: Session = Depends(get_db),
):
    reg = db.get(SupervisorRegistro, registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    db.delete(reg)
    db.commit()
    return {"ok": True}


# ──────────────────────────────────────────────────────────────
# Edición manual (panel de guardias): inicio de turno y supervisor
# ──────────────────────────────────────────────────────────────

class SupervisorRegistroEditPayload(BaseModel):
    nombre_guardia: str = ""
    tipo_turno: str = ""
    notas: str | None = None


@router.put("/api/guardia/supervisor-registro/{registro_id}")
def editar_supervisor_registro(
    registro_id: int,
    payload: SupervisorRegistroEditPayload,
    db: Session = Depends(get_db),
):
    reg = db.get(SupervisorRegistro, registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    if payload.tipo_turno and payload.tipo_turno not in TIPOS_TURNO:
        raise HTTPException(status_code=400, detail="Tipo de turno invalido")
    if payload.nombre_guardia.strip():
        guardia_user = _resolver_guardia_activo_payload(db, "", payload.nombre_guardia)
        reg.nombre_guardia = str(guardia_user.name or "").strip()
    if payload.tipo_turno:
        reg.tipo_turno = payload.tipo_turno
    if payload.notas is not None:
        reg.notas = payload.notas.strip() or None
    db.commit()
    return {"ok": True, "id": reg.id}


class SupervisorRegistroJustificarPayload(BaseModel):
    motivo: str
    fecha_desde: str | None = None  # YYYY-MM-DD
    fecha_hasta: str | None = None  # YYYY-MM-DD
    nombre_reemplazo: str = Field(min_length=1, max_length=255)
    tipo_turno_reemplazo: str


@router.post("/api/guardia/supervisor-registro/{registro_id}/justificar")
def justificar_supervisor_registro(
    registro_id: int,
    payload: SupervisorRegistroJustificarPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    """Reemplaza al guardia de un registro del supervisor por licencia/falta/
    permiso/vacaciones Y deja la justificacion guardada en la bitacora
    central (GuardiaJustificacion), para que ambas vias (tabla de supervisor
    y BBDD Justificaciones) alimenten el mismo registro sin duplicar."""
    reg = db.get(SupervisorRegistro, registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    motivo = payload.motivo.strip()
    if motivo not in MOTIVOS_JUSTIFICACION:
        raise HTTPException(status_code=422, detail="Motivo inválido")

    tipo_turno_reemplazo = payload.tipo_turno_reemplazo.strip()
    if tipo_turno_reemplazo not in ("Extra", "Contrato Diario"):
        raise HTTPException(status_code=422, detail="Indica si el reemplazo cumple Turno Extra o Contrato Diario")

    nombre_original = str(reg.nombre_guardia or "").strip()
    nombre_reemplazo = payload.nombre_reemplazo.strip()
    if not nombre_reemplazo:
        raise HTTPException(status_code=422, detail="Falta el nombre del reemplazo")

    requiere_fechas = motivo != "Falta"
    desde = hasta = None
    if requiere_fechas:
        if not payload.fecha_desde or not payload.fecha_hasta:
            raise HTTPException(status_code=422, detail="Indica desde y hasta cuándo rige la justificación")
        try:
            desde = date.fromisoformat(payload.fecha_desde)
            hasta = date.fromisoformat(payload.fecha_hasta)
        except ValueError:
            raise HTTPException(status_code=422, detail="Fecha inválida")
        if hasta < desde:
            raise HTTPException(status_code=422, detail="'Hasta' no puede ser anterior a 'Desde'")

    rut_original = _resolver_rut_guardia_por_nombre(db, nombre_original)
    if rut_original:
        conflicto = _justificacion_conflicto_vigente(db, rut_original, desde, hasta)
        if conflicto:
            raise HTTPException(
                status_code=409,
                detail=f"{nombre_original} ya tiene una justificación vigente ({conflicto.motivo}, "
                        f"{conflicto.fecha_desde.strftime('%d/%m/%Y')}–{conflicto.fecha_hasta.strftime('%d/%m/%Y')}) "
                        f"que se superpone con este período.",
            )

    nota = f'Por {motivo} de "{nombre_original}"'
    if requiere_fechas:
        nota += f" (desde {desde.strftime('%d/%m/%Y')} hasta {hasta.strftime('%d/%m/%Y')})"

    if rut_original:
        notas_justificacion = (
            f"Registrado desde tabla de supervisor — "
            f"{reg.fecha.strftime('%d/%m/%Y')} · {reg.tipo_turno} · {_recinto_display(reg.recinto)}; "
            f"reemplazo: {nombre_reemplazo}"
        )
        db.add(GuardiaJustificacion(
            rut=rut_original,
            nombre_guardia=nombre_original,
            motivo=motivo,
            fecha_desde=desde,
            fecha_hasta=hasta,
            notas=notas_justificacion,
            creado_por=_nombre_usuario_sesion(request, db) or None,
        ))

    reg.nombre_guardia = nombre_reemplazo
    reg.tipo_turno = tipo_turno_reemplazo
    reg.notas = nota
    db.commit()
    return {"ok": True, "id": reg.id, "notas": nota, "tipo_turno": tipo_turno_reemplazo, "justificacion_guardada": bool(rut_original)}


class RegistroGuardiaManualPayload(BaseModel):
    recinto: str
    fecha: str          # YYYY-MM-DD
    hora: str = ""      # HH:MM (opcional)
    nombre_guardia: str
    rut: str = ""
    tipo_turno: str


class RegistroGuardiaEditPayload(BaseModel):
    nombre_guardia: str = ""
    tipo_turno: str = ""
    hora: str = ""      # HH:MM (opcional)


def _parse_hora(valor: str):
    from datetime import time as _time
    txt = str(valor or "").strip()
    if not txt:
        return None
    try:
        hh, mm = txt.split(":", 1)
        return _time(int(hh), int(mm))
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Hora invalida (usa HH:MM)")


@router.post("/api/guardia/registro-guardia")
def crear_registro_guardia_manual(
    payload: RegistroGuardiaManualPayload,
    db: Session = Depends(get_db),
):
    from datetime import date as _date, time as _time
    try:
        fecha = _date.fromisoformat(payload.fecha)
    except ValueError:
        raise HTTPException(status_code=400, detail="Fecha invalida")
    if payload.tipo_turno not in TIPOS_TURNO:
        raise HTTPException(status_code=400, detail="Tipo de turno invalido")
    if not payload.nombre_guardia.strip():
        raise HTTPException(status_code=400, detail="Nombre requerido")
    if not payload.recinto.strip():
        raise HTTPException(status_code=400, detail="Recinto requerido")

    hora = _parse_hora(payload.hora) or _time(0, 0)
    guardia_user = _resolver_guardia_activo_payload(db, payload.rut, payload.nombre_guardia)
    nombre_guardia = str(guardia_user.name or "").strip()
    rut_guardia = _normalizar_rut(guardia_user.username)

    reg = InicioTurnoRegistro(
        rut=rut_guardia or "",
        nombre_guardia=nombre_guardia,
        tipo_turno=payload.tipo_turno,
        recinto=payload.recinto.strip(),
        ubicacion_estado="Manual (panel guardias)",
        registrado_at=datetime.combine(fecha, hora),
    )
    db.add(reg)
    db.commit()
    db.refresh(reg)
    _enviar_notificacion_turno_async(
        tipo_turno=reg.tipo_turno,
        nombre_guardia=reg.nombre_guardia,
        rut=rut_guardia,
        recinto=reg.recinto,
        fecha_turno=fecha.strftime("%d/%m/%Y"),
        fecha_registro=reg.registrado_at,
        supervisor=None,
        nota="Registro agregado manualmente desde el panel de guardias.",
        origen="el registro manual del panel de guardias",
        registro_id=f"manual-{reg.id}",
    )
    try:
        _fusionar_automatico_si_corresponde(db, reg)
    except Exception as exc:
        _log.warning("No se pudo evaluar fusion automatica para registro manual %s: %s", reg.id, exc)
    try:
        _alertar_doble_marcaje_si_corresponde(db, reg)
    except Exception as exc:
        _log.warning("No se pudo evaluar doble marcaje para registro manual %s: %s", reg.id, exc)
    return {"ok": True, "id": reg.id}


@router.put("/api/guardia/registro-guardia/{registro_id}")
def editar_registro_guardia_manual(
    registro_id: int,
    payload: RegistroGuardiaEditPayload,
    db: Session = Depends(get_db),
):
    reg = db.get(InicioTurnoRegistro, registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    if payload.tipo_turno and payload.tipo_turno not in TIPOS_TURNO:
        raise HTTPException(status_code=400, detail="Tipo de turno invalido")
    if payload.nombre_guardia.strip():
        guardia_user = _resolver_guardia_activo_payload(db, "", payload.nombre_guardia)
        reg.nombre_guardia = str(guardia_user.name or "").strip()
        reg.rut = _normalizar_rut(guardia_user.username)
    if payload.tipo_turno:
        reg.tipo_turno = payload.tipo_turno
    hora = _parse_hora(payload.hora)
    if hora is not None:
        reg.registrado_at = datetime.combine(reg.registrado_at.date(), hora)
    db.commit()
    return {"ok": True, "id": reg.id}


@router.delete("/api/guardia/registro-guardia/{registro_id}")
def eliminar_registro_guardia_manual(
    registro_id: int,
    db: Session = Depends(get_db),
):
    reg = db.get(InicioTurnoRegistro, registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    db.delete(reg)
    db.commit()
    return {"ok": True}


class FusionarRegistrosPayload(BaseModel):
    sobreviviente_id: int
    archivar_ids: list[int]
    motivo: str = ""


@router.post("/api/guardia/registro-guardia/fusionar")
def fusionar_registros_guardia(
    request: Request,
    payload: FusionarRegistrosPayload,
    db: Session = Depends(get_db),
):
    """Marcha blanca (ago 2026): fusion de marcajes duplicados por traslado
    entre recintos del mismo grupo (_GRUPOS_FUSION_RECINTOS). Requiere
    confirmacion manual de un supervisor por ahora — no se ejecuta sola."""
    sobreviviente = db.get(InicioTurnoRegistro, payload.sobreviviente_id)
    if not sobreviviente:
        raise HTTPException(status_code=404, detail="Registro sobreviviente no encontrado")
    if not payload.archivar_ids:
        raise HTTPException(status_code=400, detail="Debes indicar al menos un registro a archivar")

    usuario = _nombre_usuario_sesion(request, db) or "sistema"
    tz = ZoneInfo(settings.timezone or "America/Santiago")
    ahora = datetime.now(tz).replace(tzinfo=None)
    motivo_default = f"Fusionado con marcaje en {sobreviviente.recinto} ({sobreviviente.registrado_at.strftime('%H:%M')})"
    archivados: list[int] = []
    for reg_id in payload.archivar_ids:
        if reg_id == sobreviviente.id:
            continue
        reg = db.get(InicioTurnoRegistro, reg_id)
        if not reg or str(reg.estado or "activo") == "archivado":
            continue
        reg.estado = "archivado"
        reg.fusionado_con_id = sobreviviente.id
        reg.archivado_motivo = payload.motivo.strip() or motivo_default
        reg.archivado_en = ahora
        reg.archivado_por = usuario
        archivados.append(reg.id)
    db.commit()
    return {"ok": True, "sobreviviente_id": sobreviviente.id, "archivados": archivados}


@router.post("/api/guardia/registro-guardia/{registro_id}/restaurar")
def restaurar_registro_guardia(
    registro_id: int,
    db: Session = Depends(get_db),
):
    reg = db.get(InicioTurnoRegistro, registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    reg.estado = "activo"
    reg.fusionado_con_id = None
    reg.archivado_motivo = None
    reg.archivado_en = None
    reg.archivado_por = None
    db.commit()
    return {"ok": True, "id": reg.id}


# ──────────────────────────────────────────────
# Contrato Diario
# ──────────────────────────────────────────────

class ContratoDiarioOtpRequest(BaseModel):
    correo_electronico: str
    nombres: Optional[str] = ""


class ContratoDiarioEnviarRequest(BaseModel):
    rut: str
    nombres: str
    apellido_paterno: str
    apellido_materno: str
    fecha_nacimiento: Optional[str] = ""
    direccion: Optional[str] = ""
    comuna: Optional[str] = ""
    institucion_salud: Optional[str] = ""
    prevision_afp: Optional[str] = ""
    estado_civil: Optional[str] = ""
    telefono_movil: Optional[str] = ""
    telefono_fijo: Optional[str] = ""
    contacto_emergencia_nombre: Optional[str] = ""
    contacto_emergencia_telefono: Optional[str] = ""
    fecha_inicio_funciones: Optional[str] = ""
    banco: Optional[str] = ""
    tipo_cuenta: Optional[str] = ""
    numero_cuenta: Optional[str] = ""
    correo_electronico: str
    codigo_otp: str


@router.get("/guardia/contrato-diario", response_class=HTMLResponse)
def contrato_diario_page(request: Request):
    return templates.TemplateResponse(request, "contrato_diario_guardia.html", {})


@router.post("/api/guardia/contrato-diario/otp")
def contrato_diario_otp(payload: ContratoDiarioOtpRequest):
    try:
        contrato_diario_service.solicitar_otp(payload.correo_electronico, payload.nombres or "")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    except Exception:
        _log.exception("Error enviando OTP de contrato diario")
        raise HTTPException(status_code=502, detail="No se pudo enviar el código. Intenta nuevamente.")
    return {"ok": True}


@router.post("/api/guardia/contrato-diario/enviar")
def contrato_diario_enviar(payload: ContratoDiarioEnviarRequest):
    datos = payload.model_dump()
    codigo_otp = datos.pop("codigo_otp")
    try:
        contrato_diario_service.procesar_contrato_diario(datos, codigo_otp)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    except Exception:
        _log.exception("Error generando/enviando contrato diario")
        raise HTTPException(status_code=502, detail="No se pudo generar o enviar el contrato. Intenta nuevamente.")
    return {"ok": True}
