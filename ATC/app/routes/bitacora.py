from __future__ import annotations

import bisect
import html
import hmac
import json as _json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from jose import JWTError, jwt
from pydantic import BaseModel, Field
from sqlalchemy import bindparam, case, func, text
from sqlalchemy.orm import Session

from ATC.app.routes.bitacora_access import (
    can_access_bitacora,
    can_manage_bitacora_puestos,
    is_bitacora_admin,
    _normalize,
    _split_departments,
)
from ATC.app.core.config import settings
from ATC.app.core.db import get_db, get_incidencias_db
from ATC.app.models.incidencias import (
    SucursalBBDD,
    SucursalCamaraMonitoreo,
    SucursalPersonaAutorizada,
    SucursalContactoEmergencia,
)
from ATC.app.models.user import User
from ATC.app.core.security import hash_password
from ATC.app.services.user_service import UserService


router = APIRouter(tags=["bitacora"])
templates = Jinja2Templates(directory=str(Path(__file__).resolve().parents[1] / "templates"))
COOKIE_NAME = "access_token"


class NoticiaCreate(BaseModel):
    nombre_empresa: str = Field(min_length=1)
    nombre_sucursal: str = Field(min_length=1)
    fecha_fin_noticia: date
    mensaje: str = Field(min_length=1)


class SucursalEditPayload(BaseModel):
    sucursal_id: int
    nombre_sucursal: str = ""
    direccion_sucursal: str = ""
    referencia_ubicacion: str = ""
    contacto: str = ""
    email_facturas: str = ""
    horario_apertura: str = ""
    horario_cierre: str = ""
    horario_habil: str = ""
    horario_no_habil: str = ""
    plan_cuadrante: str = ""
    carabineros: str = ""
    bomberos: str = ""
    seguridad_ciudadana: str = ""
    camaras_contratadas: str = ""
    camaras_televigiladas: str = ""
    codigo_p2p: str = ""
    codigo_dss: str = ""
    telefono_porton: str = ""
    telefono_recepcion: str = ""
    internet_atc: str = ""
    compania_electricidad: str = ""
    numero_cliente_electricidad: str = ""
    proveedor_internet_cliente: str = ""
    latitud_longitud: str = ""


def _decode_cookie_token(token: str) -> str:
    try:
        payload = jwt.decode(token, settings.JWT_SECRET, algorithms=[settings.JWT_ALG])
        username = payload.get("sub")
        if not username:
            raise ValueError("Token sin sub")
        return str(username)
    except (JWTError, ValueError) as exc:
        raise HTTPException(status_code=401, detail="Token inválido") from exc


def _require_bitacora_access(user: User) -> None:
    if not can_access_bitacora(user):
        raise HTTPException(
            status_code=403,
            detail="La bitácora no está disponible para usuarios con acceso solo Técnico.",
        )


def _bitacora_users(db: Session) -> list[dict[str, str | bool | int]]:
    users = db.query(User).order_by(User.name.asc(), User.username.asc()).all()
    result = []
    for user in users:
        if not can_access_bitacora(user):
            continue
        departments = [_normalize(d) for d in _split_departments(getattr(user, "department", None))]
        result.append({
            "id": user.id,
            "name": str(user.name or "").strip(),
            "username": str(user.username or "").strip(),
            "email": str(user.email or "").strip(),
            "role": user.role,
            "user_type": "Administrador" if getattr(user, "is_admin", False) else "Operador",
            "status": "Activado" if bool(getattr(user, "is_active", False)) else "Desactivado",
            "is_active": bool(getattr(user, "is_active", False)),
            "is_televigilante": "televigilante" in departments,
        })
    return result


def _ensure_sucursal_aceptada_bitacora_column(db: Session) -> None:
    """Agrega columna aceptada_bitacora a bbdd_sucursales si no existe. Default 1 para
    no ocultar retroactivamente sucursales que ya estaban visibles en Bitácora."""
    try:
        db.execute(text("""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = 'bbdd_sucursales'
                  AND COLUMN_NAME = 'aceptada_bitacora'
            )
            BEGIN
                ALTER TABLE bbdd_sucursales
                ADD aceptada_bitacora BIT NOT NULL DEFAULT 1
            END
        """))
        db.commit()
    except Exception:
        db.rollback()


def _ensure_bitacora_noticias_table(db: Session) -> None:
    db.execute(
        text(
            """
            IF OBJECT_ID('bitacora_noticias', 'U') IS NULL
            BEGIN
            CREATE TABLE bitacora_noticias (
                id BIGINT IDENTITY(1,1) PRIMARY KEY,
                nombre_empresa NVARCHAR(MAX) NOT NULL,
                nombre_sucursal NVARCHAR(MAX) NOT NULL,
                usuario_registra NVARCHAR(MAX) NOT NULL,
                fecha_registro DATETIME2 NOT NULL DEFAULT GETDATE(),
                fecha_fin_noticia DATETIME2 NOT NULL,
                mensaje NVARCHAR(MAX) NOT NULL
            )
            END
            """
        )
    )
    db.commit()


def _serialize_noticia(row: dict) -> dict[str, str | int]:
    fecha_registro = row.get("fecha_registro")
    fecha_fin = row.get("fecha_fin_noticia")
    return {
        "id": row.get("id"),
        "nombre_empresa": str(row.get("nombre_empresa") or "").strip(),
        "nombre_sucursal": str(row.get("nombre_sucursal") or "").strip(),
        "usuario_registra": str(row.get("usuario_registra") or "").strip(),
        "mensaje": str(row.get("mensaje") or "").strip(),
        "fecha_registro": fecha_registro.isoformat(sep=" ", timespec="seconds") if isinstance(fecha_registro, datetime) else str(fecha_registro or ""),
        "fecha_fin_noticia": fecha_fin.isoformat(sep=" ", timespec="seconds") if isinstance(fecha_fin, datetime) else str(fecha_fin or ""),
    }


def _list_noticias(db: Session, estado: Literal["activas", "expiradas"]) -> list[dict[str, str | int]]:
    _ensure_bitacora_noticias_table(db)
    comparator = ">=" if estado == "activas" else "<"
    order_by = "fecha_registro DESC, id DESC" if estado == "activas" else "fecha_fin_noticia DESC, id DESC"
    rows = db.execute(
        text(
            f"""
            SELECT
                id,
                nombre_empresa,
                nombre_sucursal,
                usuario_registra,
                fecha_registro,
                fecha_fin_noticia,
                mensaje
            FROM bitacora_noticias
            WHERE fecha_fin_noticia {comparator} CURRENT_TIMESTAMP
            ORDER BY {order_by}
            """
        )
    ).mappings().all()
    return [_serialize_noticia(row) for row in rows]


def _detail_value(row: dict, key: str) -> str:
    value = row.get(key)
    return str(value or "").strip()


def _first_non_empty(*values: object) -> str:
    for value in values:
        text_value = str(value or "").strip()
        if text_value:
            return text_value
    return ""


_TODOS_LOS_DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def _first_image_url(imagenes_json: str | None) -> str:
    """Extrae la primera URL de un campo imagenes (JSON array de strings)."""
    raw = str(imagenes_json or "").strip()
    if not raw or raw == "-":
        return "-"
    try:
        urls = _json.loads(raw)
        if isinstance(urls, list) and urls:
            url = str(urls[0]).strip()
            return url if url else "-"
    except Exception:
        pass
    return "-"


def _dias_no_habiles(dias_funcionamiento: str) -> str:
    """Devuelve los días que NO están en dias_funcionamiento."""
    import unicodedata

    def _norm(s: str) -> str:
        return "".join(
            c for c in unicodedata.normalize("NFD", s.lower().strip())
            if unicodedata.category(c) != "Mn"
        )

    raw = str(dias_funcionamiento or "").strip()
    if not raw or raw == "-":
        return "-"
    habiles = {_norm(d) for d in raw.split(",")}
    no_habiles = [d for d in _TODOS_LOS_DIAS if _norm(d) not in habiles]
    return ", ".join(no_habiles) if no_habiles else "-"


def get_current_user_bitacora(
    request: Request,
    db: Session = Depends(get_db),
) -> User:
    token = request.cookies.get(COOKIE_NAME)
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")

    username = _decode_cookie_token(token)
    user = UserService.find_by_login(db, username)
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="No autenticado")
    return user


@router.get("/bitacora", response_class=HTMLResponse)
def bitacora_page(
    request: Request,
    db: Session = Depends(get_db),
    incidencias_db: Session = Depends(get_incidencias_db),
    token: str = Query(default=""),
):
    current_user: User | None = None
    token_limpio = (token or "").strip()
    if token_limpio:
        return RedirectResponse(url=f"/sso/login?token={token_limpio}&next=/bitacora", status_code=303)

    cookie_token = request.cookies.get(COOKIE_NAME)
    if cookie_token:
        try:
            username = _decode_cookie_token(cookie_token)
            current_user = UserService.find_by_login(db, username)
        except Exception:
            current_user = None

    if not current_user or not current_user.is_active:
        return RedirectResponse(url="/login", status_code=303)

    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)

    empresas_rows = incidencias_db.execute(
        text(
            """
            SELECT DISTINCT TRIM(nombre_empresa) AS empresa
            FROM bbdd_sucursales
            WHERE COALESCE(TRIM(nombre_empresa), '') <> ''
              AND aceptada_bitacora = 1
            ORDER BY TRIM(nombre_empresa) ASC
            """
        )
    ).mappings().all()
    empresas = [str(row.get("empresa") or "").strip() for row in empresas_rows if str(row.get("empresa") or "").strip()]

    # Solo para el selector de "Buscar por Empresa": empresas que tengan al
    # menos una sucursal habilitada. A diferencia de `empresas` (usado por
    # los paneles de administración, que sí deben poder llegar a sucursales
    # deshabilitadas para gestionarlas), acá una empresa sin ninguna
    # sucursal activa no debe aparecer como resultado de búsqueda.
    empresas_activas_rows = incidencias_db.execute(
        text(
            """
            SELECT DISTINCT TRIM(nombre_empresa) AS empresa
            FROM bbdd_sucursales s
            WHERE COALESCE(TRIM(nombre_empresa), '') <> ''
              AND aceptada_bitacora = 1
              AND EXISTS (
                  SELECT 1 FROM bbdd_sucursales s2
                  WHERE TRIM(s2.nombre_empresa) = TRIM(s.nombre_empresa)
                    AND (s2.habilitada = 1 OR s2.habilitada IS NULL)
              )
            ORDER BY TRIM(nombre_empresa) ASC
            """
        )
    ).mappings().all()
    empresas_activas = [str(row.get("empresa") or "").strip() for row in empresas_activas_rows if str(row.get("empresa") or "").strip()]

    # La tabla de usuarios (nombre/correo/rol/estado) solo debe llegar al HTML si el
    # usuario es admin — antes se enviaba siempre y el panel "Registro de Usuario"
    # no estaba gateado en el template, exponiendo el listado completo a cualquiera.
    bitacora_users = _bitacora_users(db) if is_bitacora_admin(current_user) else []

    # Cuentas de solo Televigilancia (sin depto Bitacora ni rol admin) no tienen
    # panel al que "volver" — en vez de Volver, solo pueden Cerrar sesión.
    role_actual = str(getattr(current_user, "role", None) or "").strip().lower()
    departamentos_actuales = [_normalize(d) for d in _split_departments(getattr(current_user, "department", None))]
    solo_televigilante = (
        "televigilante" in departamentos_actuales
        and "bitacora" not in departamentos_actuales
        and role_actual not in ("admin", "superadmin")
    )

    resp = templates.TemplateResponse(
        request,
        "bitacora.html",
        {
            "request": request,
            "user": current_user,
            "is_bitacora_admin": is_bitacora_admin(current_user),
            "can_manage_puestos": can_manage_bitacora_puestos(current_user),
            "is_operador": not bool(getattr(current_user, "is_admin", False)) and not is_bitacora_admin(current_user),
            "empresas": empresas,
            "empresas_activas": empresas_activas,
            "bitacora_users": bitacora_users,
            "solo_televigilante": solo_televigilante,
        },
    )
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/api/bitacora/sucursales")
def bitacora_sucursales_api(
    empresa: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)
    empresa_limpia = str(empresa or "").strip()

    # Sin `empresa`: listar todas las sucursales de todas las empresas —
    # usado por los buscadores que permiten elegir sucursal directo sin
    # elegir empresa antes (autocompletan la empresa dueña con el dato
    # nombre_empresa que ya viene en cada fila).
    if empresa_limpia:
        where_clause = "WHERE aceptada_bitacora = 1 AND LOWER(TRIM(nombre_empresa)) = LOWER(TRIM(:empresa))"
        params: dict[str, str] = {"empresa": empresa_limpia}
    else:
        where_clause = "WHERE aceptada_bitacora = 1"
        params = {}

    rows = incidencias_db.execute(
        text(
            f"""
            SELECT
                id,
                COALESCE(TRIM(rut), '') AS rut,
                COALESCE(TRIM(nombre_empresa), '') AS nombre_empresa,
                COALESCE(TRIM(nombre_sucursal), '') AS nombre_sucursal,
                COALESCE(TRIM(direccion_sucursal), '') AS direccion_sucursal,
                COALESCE(TRIM(comuna), '') AS comuna,
                COALESCE(TRIM(region), '') AS region
            FROM bbdd_sucursales
            {where_clause}
            ORDER BY nombre_empresa ASC, nombre_sucursal ASC, direccion_sucursal ASC, id ASC
            """
        ),
        params,
    ).mappings().all()

    sucursales = [
        {
            "id": row.get("id"),
            "rut": str(row.get("rut") or "").strip(),
            "nombre_empresa": str(row.get("nombre_empresa") or "").strip(),
            "nombre_sucursal": str(row.get("nombre_sucursal") or "").strip(),
            "direccion_sucursal": str(row.get("direccion_sucursal") or "").strip(),
            "comuna": str(row.get("comuna") or "").strip(),
            "region": str(row.get("region") or "").strip(),
        }
        for row in rows
    ]
    return {"empresa": empresa_limpia, "total": len(sucursales), "sucursales": sucursales}


@router.get("/api/bitacora/noticias")
def bitacora_noticias_api(
    estado: Literal["activas", "expiradas"] = Query(default="activas"),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    noticias = _list_noticias(incidencias_db, estado)
    return {"estado": estado, "total": len(noticias), "noticias": noticias}


@router.get("/api/bitacora/busqueda-empresa")
def bitacora_busqueda_empresa_api(
    empresa: str = Query(...),
    sucursal: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_bitacora_noticias_table(incidencias_db)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)

    empresa_limpia = str(empresa or "").strip()
    sucursal_limpia = str(sucursal or "").strip()
    if not empresa_limpia:
        raise HTTPException(status_code=400, detail="Debes indicar una empresa.")

    sucursales_rows = incidencias_db.execute(
        text(
            """
            SELECT
                id,
                COALESCE(TRIM(rut), '') AS rut,
                COALESCE(TRIM(nombre_empresa), '') AS nombre_empresa,
                COALESCE(TRIM(nombre_sucursal), '') AS nombre_sucursal,
                COALESCE(TRIM(direccion_sucursal), '') AS direccion_sucursal,
                COALESCE(TRIM(region), '') AS region,
                COALESCE(TRIM(comuna), '') AS comuna,
                COALESCE(TRIM(referencia_ubicacion), '') AS referencia_ubicacion,
                COALESCE(TRIM(latitud_longitud), '') AS latitud_longitud,
                COALESCE(TRIM(email_facturas), '') AS email_facturas,
                COALESCE(TRIM(proveedor_internet), '') AS proveedor_internet,
                COALESCE(TRIM(proveedor_electricidad), '') AS proveedor_electricidad,
                COALESCE(TRIM(nro_proveedor_electricidad), '') AS nro_proveedor_electricidad,
                COALESCE(TRIM(horario_apertura), '') AS horario_apertura,
                COALESCE(TRIM(horario_cierre), '') AS horario_cierre,
                COALESCE(TRIM(dias_funcionamiento), '') AS dias_funcionamiento,
                created_at
            FROM bbdd_sucursales
            WHERE LOWER(TRIM(nombre_empresa)) = LOWER(TRIM(:empresa))
              AND aceptada_bitacora = 1
              AND (habilitada = 1 OR habilitada IS NULL)
            ORDER BY nombre_sucursal ASC, id ASC
            """
        ),
        {"empresa": empresa_limpia},
    ).mappings().all()

    if not sucursales_rows:
        return {
            "empresa": empresa_limpia,
            "sucursal_actual": "",
            "sucursales": [],
            "detalle": {},
            "contactos_emergencia": [],
            "indicaciones_especiales": [],
            "noticias": [],
            "personas_autorizadas": [],
        }

    sucursales = [
        {
            "id": row.get("id"),
            "nombre_sucursal": _detail_value(row, "nombre_sucursal"),
        }
        for row in sucursales_rows
    ]

    selected_row = next(
        (row for row in sucursales_rows if _detail_value(row, "nombre_sucursal").casefold() == sucursal_limpia.casefold()),
        sucursales_rows[0],
    )
    selected_sucursal = _detail_value(selected_row, "nombre_sucursal")

    ficha = _ficha_sucursal(incidencias_db, selected_row, empresa_limpia)
    return {
        "empresa": empresa_limpia,
        "sucursal_id": selected_row.get("id"),
        "sucursal_actual": selected_sucursal,
        "sucursales": sucursales,
        **ficha,
    }


def _ficha_sucursal(incidencias_db: Session, selected_row: dict, empresa_limpia: str) -> dict:
    """Arma detalle + contactos de emergencia + personas autorizadas + noticias para
    una sucursal ya resuelta (selected_row). Compartido por /api/bitacora/busqueda-empresa
    (sucursales aceptadas) y /api/bitacora/sucursales-pendientes/{id}/preview (una
    sucursal pendiente, buscada directo por id, sin pasar por el filtro de aceptación)."""
    selected_sucursal = _detail_value(selected_row, "nombre_sucursal")
    venta_row = None
    venta_rows: list = []
    try:
        venta_rows = incidencias_db.execute(
            text(
                """
                SELECT
                    codigo,
                    estado,
                    tipo_plan,
                    numero_camaras_instalar,
                    numero_camaras_desinstalar,
                    numero_camaras_vigilar
                FROM venta_comercial
                WHERE LOWER(TRIM(rut_cliente)) = LOWER(TRIM(:rut))
                  AND (
                    LOWER(TRIM(direccion_sucursal)) = LOWER(TRIM(:direccion))
                    OR LOWER(TRIM(nombre_sucursal)) = LOWER(TRIM(:sucursal))
                  )
                ORDER BY id DESC
                """
            ),
            {
                "rut": _detail_value(selected_row, "rut"),
                "direccion": _detail_value(selected_row, "direccion_sucursal"),
                "sucursal": selected_sucursal,
            },
        ).mappings().all()
    except Exception:
        venta_rows = []

    venta_row = venta_rows[0] if venta_rows else None  # ultima ODS: para tipo_plan/codigo, etc.

    # Las camaras se ACUMULAN entre todas las ODS de la sucursal (mismo criterio que
    # venta_service.get_cliente_sucursal_resumen): una ODS de upgrade con 2 camaras
    # suma sobre las 4 existentes -> 6, no las pisa. Se excluyen las ODS anuladas.
    ods_validas = [o for o in venta_rows if str(o.get("estado") or "").strip().lower() != "anulada"]
    total_camaras_instalar = sum(int(o.get("numero_camaras_instalar") or 0) for o in ods_validas)
    total_camaras_desinstalar = sum(int(o.get("numero_camaras_desinstalar") or 0) for o in ods_validas)
    total_camaras_vigilar = sum(int(o.get("numero_camaras_vigilar") or 0) for o in ods_validas)
    total_camaras_contratadas = max(total_camaras_instalar - total_camaras_desinstalar, 0)

    cliente_row = None
    try:
        cliente_row = incidencias_db.execute(
            text(
                """
                SELECT telefono
                FROM bbdd_clientes
                WHERE LOWER(TRIM(cliente)) = LOWER(TRIM(:empresa))
                   OR LOWER(TRIM(rut)) = LOWER(TRIM(:rut))
                ORDER BY id DESC OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                """
            ),
            {"empresa": empresa_limpia, "rut": _detail_value(selected_row, "rut")},
        ).mappings().first()
    except Exception:
        cliente_row = None

    emergency_rows = []
    try:
        emergency_rows = incidencias_db.execute(
            text(
                """
                SELECT
                    id,
                    COALESCE(TRIM(nombre), '') AS nombre,
                    COALESCE(TRIM(telefono), '') AS celular,
                    COALESCE(TRIM(rut), '') AS rut,
                    COALESCE(TRIM(email), '') AS email,
                    orden
                FROM sucursal_contactos_emergencia
                WHERE sucursal_id = :sucursal_id
                ORDER BY COALESCE(orden, id) ASC
                """
            ),
            {"sucursal_id": selected_row.get("id")},
        ).mappings().all()
    except Exception:
        emergency_rows = []

    personas_autorizadas_rows = []
    try:
        personas_autorizadas_rows = incidencias_db.execute(
            text(
                """
                SELECT
                    id,
                    COALESCE(TRIM(nombre), '') AS nombre,
                    COALESCE(TRIM(rut), '') AS rut,
                    COALESCE(TRIM(telefono), '') AS celular,
                    COALESCE(TRIM(email), '') AS email,
                    COALESCE(TRIM(clave_verde), '') AS clave_verde,
                    COALESCE(TRIM(clave_roja), '') AS clave_roja
                FROM sucursal_personas_autorizadas
                WHERE sucursal_id = :sucursal_id
                ORDER BY id ASC
                """
            ),
            {"sucursal_id": selected_row.get("id")},
        ).mappings().all()
    except Exception:
        personas_autorizadas_rows = []

    indicaciones_rows = []
    try:
        indicaciones_rows = incidencias_db.execute(
            text(
                """
                SELECT
                    fecha,
                    mensaje
                FROM bitacora_indicaciones_bdatc
                WHERE sucursal_id = :sucursal_id
                  AND COALESCE(TRIM(mensaje), '') <> ''
                ORDER BY fecha DESC, id DESC
                """
            ),
            {"sucursal_id": selected_row.get("id")},
        ).mappings().all()
    except Exception:
        try:
            indicaciones_rows = incidencias_db.execute(
                text(
                    """
                    SELECT
                        fecha_registro AS fecha,
                        mensaje
                    FROM bitacora_noticias
                    WHERE LOWER(TRIM(nombre_empresa)) = LOWER(TRIM(:empresa))
                      AND LOWER(TRIM(nombre_sucursal)) = LOWER(TRIM(:sucursal))
                      AND LOWER(TRIM(usuario_registra)) = 'bdatc indicaciones'
                      AND COALESCE(TRIM(mensaje), '') <> ''
                    ORDER BY fecha_registro DESC, id DESC
                    """
                ),
                {"empresa": empresa_limpia, "sucursal": selected_sucursal},
            ).mappings().all()
        except Exception:
            indicaciones_rows = []

    layout_row = None
    try:
        # Prioridad 1: layout_final subido por servicio técnico
        layout_row = incidencias_db.execute(
            text(
                """
                SELECT st.layout_final AS ruta_archivo
                FROM venta_servicio_tecnico st
                JOIN venta_comercial v ON v.codigo = st.odt
                WHERE st.layout_final IS NOT NULL AND TRIM(st.layout_final) <> ''
                  AND LOWER(TRIM(v.rut_cliente)) = LOWER(TRIM(:rut))
                  AND (
                    LOWER(TRIM(v.direccion_sucursal)) = LOWER(TRIM(:direccion))
                    OR LOWER(TRIM(v.nombre_sucursal)) = LOWER(TRIM(:sucursal))
                  )
                ORDER BY st.id DESC OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                """
            ),
            {
                "rut": _detail_value(selected_row, "rut"),
                "direccion": _detail_value(selected_row, "direccion_sucursal"),
                "sucursal": selected_sucursal,
            },
        ).mappings().first()
    except Exception:
        layout_row = None

    if not layout_row:
        try:
            # Fallback: layout subido por comercial
            layout_row = incidencias_db.execute(
                text(
                    """
                    SELECT a.ruta_archivo
                    FROM venta_ods_archivos a
                    JOIN venta_comercial v ON v.codigo = a.codigo_ods
                    WHERE a.tipo_documento = 'Layout'
                      AND LOWER(TRIM(v.rut_cliente)) = LOWER(TRIM(:rut))
                      AND (
                        LOWER(TRIM(v.direccion_sucursal)) = LOWER(TRIM(:direccion))
                        OR LOWER(TRIM(v.nombre_sucursal)) = LOWER(TRIM(:sucursal))
                      )
                    ORDER BY a.id DESC OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    """
                ),
                {
                    "rut": _detail_value(selected_row, "rut"),
                    "direccion": _detail_value(selected_row, "direccion_sucursal"),
                    "sucursal": selected_sucursal,
                },
            ).mappings().first()
        except Exception:
            layout_row = None

    info_extra_row = None
    try:
        info_extra_row = incidencias_db.execute(
            text("SELECT TOP 1 * FROM sucursal_info_extra WHERE sucursal_id = :sid"),
            {"sid": selected_row.get("id")},
        ).mappings().first()
    except Exception:
        info_extra_row = None

    noticias_rows = []
    try:
        noticias_rows = incidencias_db.execute(
            text(
                """
                SELECT
                    id,
                    nombre_empresa,
                    nombre_sucursal,
                    usuario_registra,
                    fecha_registro,
                    fecha_fin_noticia,
                    mensaje
                FROM bitacora_noticias
                WHERE LOWER(TRIM(nombre_empresa)) = LOWER(TRIM(:empresa))
                  AND LOWER(TRIM(nombre_sucursal)) = LOWER(TRIM(:sucursal))
                  AND LOWER(TRIM(usuario_registra)) <> 'bdatc indicaciones'
                ORDER BY fecha_registro DESC, id DESC
                """
            ),
            {"empresa": empresa_limpia, "sucursal": selected_sucursal},
        ).mappings().all()
    except Exception:
        noticias_rows = []

    detalle = {
        "empresa": _detail_value(selected_row, "nombre_empresa"),
        "rut": _detail_value(selected_row, "rut"),
        "direccion": _detail_value(selected_row, "direccion_sucursal"),
        "latitud_longitud": _detail_value(selected_row, "latitud_longitud"),
        "referencia_ubicacion": _first_non_empty(_detail_value(selected_row, "referencia_ubicacion"), info_extra_row.get("referencia_ubicacion") if info_extra_row else None),
        "contacto": _first_non_empty(cliente_row.get("telefono") if cliente_row else None, info_extra_row.get("contacto") if info_extra_row else None),
        "correo": _first_non_empty(_detail_value(selected_row, "email_facturas"), info_extra_row.get("correo") if info_extra_row else None),
        "horario_apertura": _first_non_empty(_detail_value(selected_row, "horario_apertura")),
        "horario_cierre": _first_non_empty(_detail_value(selected_row, "horario_cierre")),
        "horario_habil": _first_non_empty(_detail_value(selected_row, "dias_funcionamiento"), info_extra_row.get("horario_habil") if info_extra_row else None),
        "horario_no_habil": _first_non_empty(_dias_no_habiles(_detail_value(selected_row, "dias_funcionamiento")), info_extra_row.get("horario_no_habil") if info_extra_row else None),
        "fecha_inicio": _first_non_empty(info_extra_row.get("fecha_inicio") if info_extra_row else None),
        "tipo_vigilancia": _first_non_empty(venta_row.get("tipo_plan") if venta_row else None, info_extra_row.get("tipo_vigilancia") if info_extra_row else None),
        "camaras_contratadas": _first_non_empty(str(total_camaras_contratadas) if ods_validas else None, info_extra_row.get("camaras_contratadas") if info_extra_row else None),
        "camaras_televigiladas": _first_non_empty(str(total_camaras_vigilar) if ods_validas else None, info_extra_row.get("camaras_televigiladas") if info_extra_row else None),
        "plano_camaras": _first_non_empty(layout_row.get("ruta_archivo") if layout_row else None),
        "plan_cuadrante": _first_non_empty(info_extra_row.get("plan_cuadrante") if info_extra_row else None),
        "carabineros": _first_non_empty(info_extra_row.get("carabineros") if info_extra_row else None),
        "bomberos": _first_non_empty(info_extra_row.get("bomberos") if info_extra_row else None),
        "seguridad_ciudadana": _first_non_empty(info_extra_row.get("seguridad_ciudadana") if info_extra_row else None),
        "codigo_p2p": _first_non_empty(info_extra_row.get("codigo_p2p") if info_extra_row else None),
        "codigo_dss": _first_non_empty(info_extra_row.get("codigo_dss") if info_extra_row else None),
        "telefono_porton": _first_non_empty(info_extra_row.get("telefono_porton") if info_extra_row else None),
        "telefono_recepcion": _first_non_empty(info_extra_row.get("telefono_recepcion") if info_extra_row else None),
        "internet_atc": _first_non_empty(info_extra_row.get("internet_atc") if info_extra_row else None),
        "compania_electricidad": _first_non_empty(_detail_value(selected_row, "proveedor_electricidad"), info_extra_row.get("compania_electricidad") if info_extra_row else None),
        "numero_cliente_electricidad": _first_non_empty(_detail_value(selected_row, "nro_proveedor_electricidad"), info_extra_row.get("numero_cliente_electricidad") if info_extra_row else None),
        "proveedor_internet_cliente": _first_non_empty(_detail_value(selected_row, "proveedor_internet"), info_extra_row.get("proveedor_internet_cliente") if info_extra_row else None),
    }

    return {
        "detalle": detalle,
        "contactos_emergencia": [
            {
                "id": row.get("id"),
                "nombre": _first_non_empty(row.get("nombre")),
                "celular": _first_non_empty(row.get("celular")),
                "rut": _first_non_empty(row.get("rut")),
                "email": _first_non_empty(row.get("email")),
                "prioridad": str(idx + 1),
            }
            for idx, row in enumerate(emergency_rows)
        ],
        "indicaciones_especiales": [
            {
                "fecha": row.get("fecha").isoformat(sep=" ", timespec="seconds")
                if isinstance(row.get("fecha"), datetime)
                else _first_non_empty(row.get("fecha")),
                "mensaje": _first_non_empty(row.get("mensaje")),
            }
            for row in indicaciones_rows
        ],
        "noticias": [_serialize_noticia(row) for row in noticias_rows],
        "personas_autorizadas": [
            {
                "id": row.get("id"),
                "nombre": _first_non_empty(row.get("nombre")),
                "rut": _first_non_empty(row.get("rut")),
                "celular": _first_non_empty(row.get("celular")),
                "email": _first_non_empty(row.get("email")),
                "clave_verde": _first_non_empty(row.get("clave_verde")),
                "clave_roja": _first_non_empty(row.get("clave_roja")),
            }
            for row in personas_autorizadas_rows
        ],
    }


@router.get("/api/bitacora/empresas-sucursales")
def bitacora_empresas_sucursales_api(
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)

    rows = incidencias_db.execute(
        text(
            """
            SELECT
                DISTINCT TRIM(nombre_empresa) AS nombre_empresa,
                TRIM(nombre_sucursal) AS nombre_sucursal
            FROM bbdd_sucursales
            WHERE COALESCE(TRIM(nombre_empresa), '') <> ''
              AND aceptada_bitacora = 1
              AND (habilitada = 1 OR habilitada IS NULL)
            ORDER BY nombre_empresa ASC, nombre_sucursal ASC
            """
        )
    ).mappings().all()

    empresas_dict = {}
    for row in rows:
        empresa = str(row.get("nombre_empresa") or "").strip()
        sucursal = str(row.get("nombre_sucursal") or "").strip()
        if empresa:
            if empresa not in empresas_dict:
                empresas_dict[empresa] = []
            if sucursal:
                empresas_dict[empresa].append(sucursal)

    empresas_list = [
        {
            "empresa": empresa,
            "sucursales": sorted(list(set(sucursales)))
        }
        for empresa, sucursales in sorted(empresas_dict.items())
    ]
    return {"empresas": empresas_list}


@router.get("/api/bitacora/personas-autorizadas")
def bitacora_personas_autorizadas_api(
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Devuelve todas las personas autorizadas con datos de sucursal y plan."""
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)
    try:
        rows = incidencias_db.execute(
            text("""
                SELECT
                    s.id AS sucursal_id,
                    p.id AS persona_id,
                    s.nombre_empresa,
                    s.nombre_sucursal,
                    p.nombre,
                    p.rut,
                    p.telefono,
                    p.email,
                    p.clave_verde,
                    p.clave_roja,
                    (
                        SELECT v.tipo_plan
                        FROM venta_comercial v
                        WHERE LOWER(TRIM(v.rut_cliente)) = LOWER(TRIM(s.rut))
                          AND (
                            LOWER(TRIM(v.direccion_sucursal)) = LOWER(TRIM(s.direccion_sucursal))
                            OR LOWER(TRIM(v.nombre_sucursal)) = LOWER(TRIM(s.nombre_sucursal))
                          )
                        ORDER BY v.id DESC OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
                    ) AS tipo_plan
                FROM sucursal_personas_autorizadas p
                JOIN bbdd_sucursales s ON s.id = p.sucursal_id
                WHERE p.habilitado = 1
                  AND s.aceptada_bitacora = 1
                  AND (s.habilitada = 1 OR s.habilitada IS NULL)
                ORDER BY p.rut, s.nombre_empresa, s.nombre_sucursal
            """)
        ).mappings().all()
    except Exception:
        rows = []
    return {
        "personas": [
            {
                "empresa": r.get("nombre_empresa") or "",
                "sucursal": r.get("nombre_sucursal") or "",
                "sucursalId": r.get("sucursal_id"),
                "personaId": r.get("persona_id"),
                "horario": _first_non_empty(r.get("tipo_plan")),
                "nombre": r.get("nombre") or "",
                "rut": r.get("rut") or "",
                "celular": r.get("telefono") or "",
                "email": r.get("email") or "",
                "claveVerde": r.get("clave_verde") or "",
                "claveRoja": r.get("clave_roja") or "",
            }
            for r in rows
        ]
    }


@router.post("/api/bitacora/noticias")
def crear_bitacora_noticia(
    payload: NoticiaCreate,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_bitacora_noticias_table(incidencias_db)

    nombre_empresa = payload.nombre_empresa.strip()
    nombre_sucursal = payload.nombre_sucursal.strip()
    mensaje = payload.mensaje.strip()
    usuario_registra = str(current_user.name or current_user.username or "").strip()
    fecha_fin = datetime.combine(payload.fecha_fin_noticia, time(23, 59, 59))

    if not nombre_empresa or not nombre_sucursal or not mensaje:
        raise HTTPException(status_code=400, detail="Debes completar todos los campos de la noticia.")

    row = incidencias_db.execute(
        text(
            """
            INSERT INTO bitacora_noticias (
                nombre_empresa,
                nombre_sucursal,
                usuario_registra,
                fecha_fin_noticia,
                mensaje
            )
            OUTPUT
                INSERTED.id,
                INSERTED.nombre_empresa,
                INSERTED.nombre_sucursal,
                INSERTED.usuario_registra,
                INSERTED.fecha_registro,
                INSERTED.fecha_fin_noticia,
                INSERTED.mensaje
            VALUES (
                :nombre_empresa,
                :nombre_sucursal,
                :usuario_registra,
                :fecha_fin_noticia,
                :mensaje
            )
            """
        ),
        {
            "nombre_empresa": nombre_empresa,
            "nombre_sucursal": nombre_sucursal,
            "usuario_registra": usuario_registra,
            "fecha_fin_noticia": fecha_fin,
            "mensaje": mensaje,
        },
    ).mappings().first()
    incidencias_db.commit()

    return {"ok": True, "noticia": _serialize_noticia(row)}


# ── OBSERVACIONES ────────────────────────────────────────────────────────────

class ObservacionCreate(BaseModel):
    nombre_empresa: str = Field(min_length=1)
    nombre_sucursal: str = Field(min_length=1)
    observacion: str = ""
    enviar_cliente: bool = False
    tipo_clave: str = "Registro Observacion"
    detalle_custom: str = ""


class ObservacionEditPayload(BaseModel):
    nombre_empresa: str = Field(min_length=1)
    nombre_sucursal: str = Field(min_length=1)
    detalle: str = ""
    observacion: str = ""


class ClaveResolverRequest(BaseModel):
    persona_id: int
    sucursal_id: int
    clave: str = ""


class ClaveRegistrarRequest(ClaveResolverRequest):
    tipo_clave: str = ""
    observacion: str = ""


def _ensure_bitacora_registros_table(db: Session) -> None:
    db.execute(
        text(
            """
            IF OBJECT_ID('bitacora_registros', 'U') IS NULL
            BEGIN
            CREATE TABLE bitacora_registros (
                id BIGINT IDENTITY(1,1) PRIMARY KEY,
                nombre_empresa NVARCHAR(MAX) NOT NULL,
                nombre_sucursal NVARCHAR(MAX) NOT NULL,
                operador NVARCHAR(MAX) NOT NULL,
                detalle NVARCHAR(MAX) NOT NULL,
                observacion NVARCHAR(MAX) NOT NULL,
                tipo_clave NVARCHAR(MAX) NOT NULL DEFAULT 'Registro Observacion',
                created_at DATETIME2 NOT NULL DEFAULT GETDATE()
            )
            END
            """
        )
    )
    db.commit()


def _serialize_registro(row: dict) -> dict:
    ts = row.get("created_at")
    return {
        "id": row.get("id"),
        "nombre_empresa": str(row.get("nombre_empresa") or "").strip(),
        "nombre_sucursal": str(row.get("nombre_sucursal") or "").strip(),
        "operador": str(row.get("operador") or "").strip(),
        "detalle": str(row.get("detalle") or "").strip(),
        "observacion": str(row.get("observacion") or "").strip(),
        "tipo_clave": str(row.get("tipo_clave") or "").strip(),
        "created_at": ts.isoformat(sep=" ", timespec="seconds") if isinstance(ts, datetime) else str(ts or ""),
    }


def _resolver_clave_persona(
    db: Session,
    *,
    persona_id: int,
    sucursal_id: int,
    clave: str,
) -> tuple[SucursalPersonaAutorizada, SucursalBBDD, str | None, bool]:
    persona = (
        db.query(SucursalPersonaAutorizada)
        .filter(
            SucursalPersonaAutorizada.id == persona_id,
            SucursalPersonaAutorizada.sucursal_id == sucursal_id,
            SucursalPersonaAutorizada.habilitado == True,  # noqa: E712
        )
        .first()
    )
    sucursal = db.get(SucursalBBDD, sucursal_id)
    if not persona or not sucursal:
        raise HTTPException(status_code=404, detail="Persona o sucursal no encontrada.")

    clave_ingresada = str(clave or "").strip()
    clave_verde = str(persona.clave_verde or "").strip()
    clave_roja = str(persona.clave_roja or "").strip()
    tiene_claves = bool(clave_verde or clave_roja)

    # Mensaje unico para todos los casos de clave invalida: si el detalle
    # distinguiera "no tiene claves registradas" de "clave incorrecta",
    # cualquiera podria usar este modal para averiguar quien tiene o no
    # clave configurada solo probando. No revelar esa informacion.
    CLAVE_INVALIDA = "Clave incorrecta."

    if not tiene_claves:
        if clave_ingresada:
            raise HTTPException(status_code=400, detail=CLAVE_INVALIDA)
        return persona, sucursal, None, True

    if not clave_ingresada:
        raise HTTPException(status_code=400, detail=CLAVE_INVALIDA)

    # Comparacion insensible a mayusculas/minusculas y tildes: al guardia le
    # puede llegar la clave hablada o escrita a mano, no debe fallar por eso.
    clave_ingresada_norm = _normalize(clave_ingresada)
    coincidencias = []
    if clave_verde and hmac.compare_digest(clave_ingresada_norm, _normalize(clave_verde)):
        coincidencias.append("Verde")
    if clave_roja and hmac.compare_digest(clave_ingresada_norm, _normalize(clave_roja)):
        coincidencias.append("Roja")
    if not coincidencias:
        raise HTTPException(status_code=400, detail=CLAVE_INVALIDA)
    if len(coincidencias) > 1:
        return persona, sucursal, None, True
    return persona, sucursal, coincidencias[0], False


@router.post("/api/bitacora/clave/resolver")
def resolver_clave_bitacora(
    payload: ClaveResolverRequest,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _persona, _sucursal, tipo, requiere_tipo = _resolver_clave_persona(
        incidencias_db,
        persona_id=payload.persona_id,
        sucursal_id=payload.sucursal_id,
        clave=payload.clave,
    )
    return {"ok": True, "tipo": tipo, "requiere_tipo": requiere_tipo}


@router.post("/api/bitacora/clave/registrar")
def registrar_clave_bitacora(
    payload: ClaveRegistrarRequest,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_bitacora_registros_table(incidencias_db)
    persona, sucursal, tipo_resuelto, requiere_tipo = _resolver_clave_persona(
        incidencias_db,
        persona_id=payload.persona_id,
        sucursal_id=payload.sucursal_id,
        clave=payload.clave,
    )

    tipo_solicitado = str(payload.tipo_clave or "").strip().casefold()
    if requiere_tipo:
        if tipo_solicitado not in {"verde", "roja"}:
            raise HTTPException(status_code=400, detail="Debes indicar si la clave es Verde o Roja.")
        tipo = "Verde" if tipo_solicitado == "verde" else "Roja"
    else:
        tipo = str(tipo_resuelto)
        if tipo_solicitado and tipo_solicitado != tipo.casefold():
            raise HTTPException(status_code=400, detail="El tipo no corresponde a la clave ingresada.")

    observacion = str(payload.observacion or "").strip()
    if tipo == "Roja" and not observacion:
        raise HTTPException(status_code=400, detail="La observación es obligatoria para una Clave Roja.")

    operador = str(current_user.name or current_user.username or "").strip()
    empresa = str(sucursal.nombre_empresa or "").strip()
    nombre_sucursal = str(sucursal.nombre_sucursal or "").strip()
    detalle = f"Clave {tipo} - {str(persona.nombre or '').strip()}".strip(" -")
    row = incidencias_db.execute(
        text(
            """
            INSERT INTO bitacora_registros
                (nombre_empresa, nombre_sucursal, operador, detalle, observacion, tipo_clave)
            OUTPUT
                INSERTED.id, INSERTED.nombre_empresa, INSERTED.nombre_sucursal,
                INSERTED.operador, INSERTED.detalle, INSERTED.observacion,
                INSERTED.tipo_clave, INSERTED.created_at
            VALUES
                (:empresa, :sucursal, :operador, :detalle, :observacion, :tipo_clave)
            """
        ),
        {
            "empresa": empresa,
            "sucursal": nombre_sucursal,
            "operador": operador,
            "detalle": detalle,
            "observacion": observacion,
            "tipo_clave": tipo,
        },
    ).mappings().first()
    incidencias_db.commit()
    return {"ok": True, "tipo": tipo, "registro": _serialize_registro(row)}


@router.post("/api/bitacora/observacion")
def crear_observacion(
    payload: ObservacionCreate,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_bitacora_registros_table(incidencias_db)

    empresa = payload.nombre_empresa.strip()
    sucursal = payload.nombre_sucursal.strip()
    observacion = payload.observacion.strip()
    operador = str(current_user.name or current_user.username or "").strip()
    tipo_clave = payload.tipo_clave.strip() or "Registro Observacion"
    detalle_custom = payload.detalle_custom.strip()
    es_clave_verde = tipo_clave.casefold() == "verde"

    if not empresa or not sucursal:
        raise HTTPException(status_code=400, detail="Debes completar todos los campos.")
    if not observacion and es_clave_verde:
        return {"ok": True, "registro": None, "skipped": True}
    if not observacion:
        raise HTTPException(status_code=400, detail="Debes completar todos los campos.")

    detalle = detalle_custom if detalle_custom else f"Observacion de la empresa {empresa} en la sucursal {sucursal}"

    row = incidencias_db.execute(
        text(
            """
            INSERT INTO bitacora_registros
                (nombre_empresa, nombre_sucursal, operador, detalle, observacion, tipo_clave)
            OUTPUT
                INSERTED.id, INSERTED.nombre_empresa, INSERTED.nombre_sucursal,
                INSERTED.operador, INSERTED.detalle, INSERTED.observacion,
                INSERTED.tipo_clave, INSERTED.created_at
            VALUES
                (:empresa, :sucursal, :operador, :detalle, :observacion, :tipo_clave)
            """
        ),
        {
            "empresa": empresa,
            "sucursal": sucursal,
            "operador": operador,
            "detalle": detalle,
            "observacion": observacion,
            "tipo_clave": tipo_clave,
        },
    ).mappings().first()
    incidencias_db.commit()

    return {"ok": True, "registro": _serialize_registro(row)}


@router.get("/api/bitacora/observaciones")
def listar_observaciones(
    empresa: str = Query(default=""),
    sucursal: str = Query(default=""),
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    fecha: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_bitacora_registros_table(incidencias_db)

    conditions = ["1=1"]
    params: dict = {}

    if empresa.strip():
        conditions.append("LOWER(TRIM(nombre_empresa)) = LOWER(TRIM(:empresa))")
        params["empresa"] = empresa.strip()
    if sucursal.strip():
        conditions.append("LOWER(TRIM(nombre_sucursal)) = LOWER(TRIM(:sucursal))")
        params["sucursal"] = sucursal.strip()
    if fecha.strip():
        conditions.append("CAST(created_at AS DATE) = :fecha")
        params["fecha"] = fecha.strip()
    else:
        if desde.strip():
            conditions.append("created_at >= :desde")
            params["desde"] = desde.strip() + " 00:00:00"
        if hasta.strip():
            conditions.append("created_at <= :hasta")
            params["hasta"] = hasta.strip() + " 23:59:59"

    where = " AND ".join(conditions)
    rows = incidencias_db.execute(
        text(
            f"""
            SELECT id, nombre_empresa, nombre_sucursal, operador, detalle, observacion, tipo_clave, created_at
            FROM bitacora_registros
            WHERE {where}
            ORDER BY created_at DESC OFFSET 0 ROWS FETCH NEXT 500 ROWS ONLY
            """
        ),
        params,
    ).mappings().all()

    return {"total": len(rows), "registros": [_serialize_registro(r) for r in rows]}


def _puede_editar_registro(current_user: User, operador_registro: str) -> bool:
    """Un operador puede editar solo sus propias observaciones (comparando el
    nombre/username guardado en `operador` contra el usuario actual, porque
    bitacora_registros no tiene FK a users); un admin de bitácora puede
    editar cualquiera."""
    if is_bitacora_admin(current_user):
        return True
    nombre_actual = str(current_user.name or current_user.username or "").strip()
    return _normalize(operador_registro) == _normalize(nombre_actual)


@router.put("/api/bitacora/observaciones/{registro_id}")
def editar_observacion(
    registro_id: int,
    payload: ObservacionEditPayload,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.execute(
        text("SELECT id, operador FROM bitacora_registros WHERE id = :id"),
        {"id": registro_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Registro no encontrado.")
    if not _puede_editar_registro(current_user, str(row.get("operador") or "")):
        raise HTTPException(status_code=403, detail="No tienes permiso para editar esta observación.")

    empresa = payload.nombre_empresa.strip()
    sucursal = payload.nombre_sucursal.strip()
    if not empresa or not sucursal:
        raise HTTPException(status_code=400, detail="Debes completar empresa y sucursal.")

    updated = incidencias_db.execute(
        text(
            """
            UPDATE bitacora_registros SET
                nombre_empresa  = :empresa,
                nombre_sucursal = :sucursal,
                detalle         = :detalle,
                observacion     = :observacion
            OUTPUT
                INSERTED.id, INSERTED.nombre_empresa, INSERTED.nombre_sucursal,
                INSERTED.operador, INSERTED.detalle, INSERTED.observacion,
                INSERTED.tipo_clave, INSERTED.created_at
            WHERE id = :id
            """
        ),
        {
            "id": registro_id,
            "empresa": empresa,
            "sucursal": sucursal,
            "detalle": payload.detalle.strip(),
            "observacion": payload.observacion.strip(),
        },
    ).mappings().first()
    incidencias_db.commit()
    return {"ok": True, "registro": _serialize_registro(updated)}


@router.delete("/api/bitacora/observaciones/{registro_id}")
def eliminar_observacion(
    registro_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo un administrador puede eliminar observaciones.")
    row = incidencias_db.execute(
        text("SELECT id FROM bitacora_registros WHERE id = :id"),
        {"id": registro_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Registro no encontrado.")
    incidencias_db.execute(text("DELETE FROM bitacora_registros WHERE id = :id"), {"id": registro_id})
    incidencias_db.commit()
    return {"ok": True}


def _parse_rango_fechas(desde: str, hasta: str) -> tuple[datetime | None, datetime | None]:
    """'YYYY-MM-DD' → (inicio, fin_exclusivo). Cualquiera puede venir vacío."""
    inicio = fin_excl = None
    try:
        if str(desde or "").strip():
            inicio = datetime.combine(date.fromisoformat(desde.strip()), time.min)
    except ValueError:
        raise HTTPException(status_code=422, detail="Fecha 'desde' inválida (usa AAAA-MM-DD).")
    try:
        if str(hasta or "").strip():
            fin_excl = datetime.combine(date.fromisoformat(hasta.strip()) + timedelta(days=1), time.min)
    except ValueError:
        raise HTTPException(status_code=422, detail="Fecha 'hasta' inválida (usa AAAA-MM-DD).")
    if inicio and fin_excl and fin_excl <= inicio:
        raise HTTPException(status_code=422, detail="'Hasta' no puede ser anterior a 'Desde'.")
    return inicio, fin_excl


def _fmt_fecha(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y %H:%M")
    return ""


def _conteo_ordenado(valores: list[str], tope: int | None = None) -> list[list]:
    """Cuenta ocurrencias y devuelve [["valor", n], ...] ordenado desc.
    Si hay tope, agrupa el resto como 'Otros'."""
    conteo: dict[str, int] = {}
    for v in valores:
        clave = str(v or "").strip() or "Sin clasificar"
        conteo[clave] = conteo.get(clave, 0) + 1
    pares = sorted(conteo.items(), key=lambda kv: (-kv[1], kv[0]))
    if tope and len(pares) > tope:
        visibles, resto = pares[:tope], pares[tope:]
        visibles.append((f"Otros ({len(resto)} tipos)", sum(n for _, n in resto)))
        pares = visibles
    return [[k, n] for k, n in pares]


_ESTADOS_INCIDENCIA_CERRADA = ("terminado", "repetida")
_MAX_DIAS_ACTIVIDAD = 30


def _agrupar_incidencias_por_dia(inc_detalle: list[dict]) -> list[dict]:
    """Une el detalle de incidencias por día calendario (más reciente primero,
    ya que inc_detalle viene ordenado DESC por fecha_registro), con cuántas
    quedaron pendientes/resueltas ese día y el promedio de días que llevan
    pendientes o que tardaron en resolverse."""
    grupos: dict[str, dict] = {}
    orden: list[str] = []
    for r in inc_detalle:
        fecha_reg = r["fecha_registro"]
        if not isinstance(fecha_reg, datetime):
            continue
        clave = fecha_reg.strftime("%d-%m-%Y")
        if clave not in grupos:
            grupos[clave] = {"cantidad": 0, "pendientes": 0, "resueltas": 0, "dias_pendiente": [], "dias_resolucion": []}
            orden.append(clave)
        g = grupos[clave]
        g["cantidad"] += 1
        if r["dias_abierta"] is not None:
            g["pendientes"] += 1
            g["dias_pendiente"].append(r["dias_abierta"])
        if r["dias_resolucion"] is not None:
            g["resueltas"] += 1
            g["dias_resolucion"].append(r["dias_resolucion"])
    resultado = []
    for clave in orden:
        g = grupos[clave]
        resultado.append({
            "fecha": clave,
            "cantidad": g["cantidad"],
            "pendientes": g["pendientes"],
            "resueltas": g["resueltas"],
            "dias_pendiente_prom": round(sum(g["dias_pendiente"]) / len(g["dias_pendiente"]), 1) if g["dias_pendiente"] else None,
            "dias_resolucion_prom": round(sum(g["dias_resolucion"]) / len(g["dias_resolucion"]), 1) if g["dias_resolucion"] else None,
        })
    return resultado


def _agrupar_protocolos_por_dia(proto_detalle: list[dict]) -> list[dict]:
    """Une el detalle de protocolos por día calendario (más reciente primero),
    con el desglose de cuántos hubo de cada tipo (Preventivo/Intrusivo/etc)."""
    grupos: dict[str, dict] = {}
    orden: list[str] = []
    for r in proto_detalle:
        fecha_reg = r["fecha_registro"]
        if not isinstance(fecha_reg, datetime):
            continue
        clave = fecha_reg.strftime("%d-%m-%Y")
        if clave not in grupos:
            grupos[clave] = {"cantidad": 0, "por_tipo": {}}
            orden.append(clave)
        g = grupos[clave]
        g["cantidad"] += 1
        tipo = r["tipo"] or "Sin clasificar"
        g["por_tipo"][tipo] = g["por_tipo"].get(tipo, 0) + 1
    resultado = []
    for clave in orden:
        g = grupos[clave]
        por_tipo = sorted(g["por_tipo"].items(), key=lambda kv: (-kv[1], kv[0]))
        resultado.append({
            "fecha": clave,
            "cantidad": g["cantidad"],
            "por_tipo": [[k, n] for k, n in por_tipo],
        })
    return resultado


def _informacion_cliente_data(
    incidencias_db: Session,
    empresa: str,
    sucursal: str,
    desde: str,
    hasta: str,
) -> dict:
    """Arma el detalle completo (Incidencias / Protocolos / Bitácora) para una
    empresa+sucursal, opcionalmente acotado a un rango de fechas. Compartido
    entre el endpoint JSON del panel y el informe Excel descargable. Todo el
    cruce entre módulos es por nombre de texto — no hay FK entre
    bbdd_sucursales e incidencias/protocolos."""
    empresa_limpia = str(empresa or "").strip()
    sucursal_limpia = str(sucursal or "").strip()
    if not sucursal_limpia:
        raise HTTPException(status_code=400, detail="Debes indicar una sucursal.")

    inicio, fin_excl = _parse_rango_fechas(desde, hasta)
    ahora = datetime.now()

    def _filtro_fecha(columna: str) -> tuple[str, dict]:
        sql, params = "", {}
        if inicio:
            sql += f" AND {columna} >= :rango_inicio"
            params["rango_inicio"] = inicio
        if fin_excl:
            sql += f" AND {columna} < :rango_fin"
            params["rango_fin"] = fin_excl
        return sql, params

    # ── Incidencias: Registro.cliente guarda el nombre de la SUCURSAL (esa
    #    tabla no tiene columna empresa, no se puede filtrar por empresa) ──
    filtro_inc, params_inc = _filtro_fecha("fecha_registro")
    inc_rows = incidencias_db.execute(
        text(
            "SELECT odt, fecha_registro, problema, derivacion, estado, fecha_cierre, tecnicos "
            "FROM incidencias "
            "WHERE LOWER(TRIM(cliente)) = LOWER(TRIM(:sucursal))" + filtro_inc +
            " ORDER BY fecha_registro DESC"
        ),
        {"sucursal": sucursal_limpia, **params_inc},
    ).mappings().all()

    inc_detalle: list[dict] = []
    pendientes_incidencias = 0
    dias_pend_acum = dias_pend_n = 0
    dias_cierre_acum = dias_cierre_n = 0
    for row in inc_rows:
        estado_norm = str(row.get("estado") or "").strip().lower()
        cerrada = estado_norm in _ESTADOS_INCIDENCIA_CERRADA
        fecha_reg = row.get("fecha_registro")
        fecha_cie = row.get("fecha_cierre")
        dias_abierta = None
        dias_resolucion = None
        if isinstance(fecha_reg, datetime):
            if cerrada and isinstance(fecha_cie, datetime):
                dias_resolucion = max((fecha_cie - fecha_reg).days, 0)
                dias_cierre_acum += dias_resolucion
                dias_cierre_n += 1
            elif not cerrada:
                dias_abierta = max((ahora - fecha_reg).days, 0)
                dias_pend_acum += dias_abierta
                dias_pend_n += 1
        if not cerrada:
            pendientes_incidencias += 1
        inc_detalle.append({
            "odt": str(row.get("odt") or ""),
            "fecha_registro": fecha_reg,
            "tipo": str(row.get("problema") or "").strip() or "Sin clasificar",
            "derivacion": str(row.get("derivacion") or "").strip(),
            "estado": str(row.get("estado") or "").strip(),
            "fecha_cierre": fecha_cie,
            "tecnicos": str(row.get("tecnicos") or "").strip(),
            "dias_resolucion": dias_resolucion,
            "dias_abierta": dias_abierta,
        })

    inc_fechas = [r["fecha_registro"] for r in inc_detalle if isinstance(r["fecha_registro"], datetime)]

    # ── Protocolos: activaciones con tipo (Preventivo/Intrusivo) y éxito ──
    filtro_proto, params_proto = _filtro_fecha("fecha_registro")
    proto_rows = incidencias_db.execute(
        text(
            "SELECT fecha_registro, tipo_protocolo, protocolo_exitoso, detectado, efectivo, "
            "sirena, voz, carabineros, alpha3, informado, puesto, operador "
            "FROM protocolos_registro "
            "WHERE LOWER(TRIM(cliente)) = LOWER(TRIM(:cliente)) AND LOWER(TRIM(sucursal)) = LOWER(TRIM(:sucursal))"
            + filtro_proto + " ORDER BY fecha_registro DESC"
        ),
        {"cliente": empresa_limpia, "sucursal": sucursal_limpia, **params_proto},
    ).mappings().all()

    proto_detalle: list[dict] = []
    proto_exitosos = proto_intrusivos = proto_preventivos = 0
    for row in proto_rows:
        tipo_norm = str(row.get("tipo_protocolo") or "").strip().lower()
        exito_norm = str(row.get("protocolo_exitoso") or "").strip().upper()
        if exito_norm == "SI":
            proto_exitosos += 1
        if tipo_norm == "intrusivo":
            proto_intrusivos += 1
        elif tipo_norm == "preventivo":
            proto_preventivos += 1
        proto_detalle.append({
            "fecha_registro": row.get("fecha_registro"),
            "tipo": str(row.get("tipo_protocolo") or "").strip() or "Sin clasificar",
            "exitoso": exito_norm or "-",
            "detectado": str(row.get("detectado") or "").strip(),
            "efectivo": str(row.get("efectivo") or "").strip(),
            "sirena": str(row.get("sirena") or "").strip(),
            "voz": str(row.get("voz") or "").strip(),
            "carabineros": str(row.get("carabineros") or "").strip(),
            "alpha3": str(row.get("alpha3") or "").strip(),
            "informado": str(row.get("informado") or "").strip(),
            "puesto": str(row.get("puesto") or "").strip(),
            "operador": str(row.get("operador") or "").strip(),
        })
    proto_fechas = [r["fecha_registro"] for r in proto_detalle if isinstance(r["fecha_registro"], datetime)]

    #    Informes de protocolo (PENDIENTE|OK|ERROR) — sin filtro de fecha: son
    #    derivados de las activaciones y su pendiente importa como estado global.
    proto_informes = incidencias_db.execute(
        text(
            "SELECT estado FROM protocolos_informes "
            "WHERE LOWER(TRIM(cliente)) = LOWER(TRIM(:cliente)) AND LOWER(TRIM(sucursal)) = LOWER(TRIM(:sucursal))"
        ),
        {"cliente": empresa_limpia, "sucursal": sucursal_limpia},
    ).mappings().all()
    total_informes = len(proto_informes)
    informes_pendientes = sum(1 for r in proto_informes if str(r.get("estado") or "").strip().upper() != "OK")

    # ── Bitácora ──
    _ensure_bitacora_registros_table(incidencias_db)
    filtro_bit, params_bit = _filtro_fecha("created_at")
    bit_rows = incidencias_db.execute(
        text(
            "SELECT created_at, tipo_clave, operador, detalle, observacion "
            "FROM bitacora_registros "
            "WHERE LOWER(TRIM(nombre_empresa)) = LOWER(TRIM(:empresa)) AND LOWER(TRIM(nombre_sucursal)) = LOWER(TRIM(:sucursal))"
            + filtro_bit + " ORDER BY created_at DESC"
        ),
        {"empresa": empresa_limpia, "sucursal": sucursal_limpia, **params_bit},
    ).mappings().all()

    hace_30_dias = ahora - timedelta(days=30)
    bit_detalle = [
        {
            "fecha": row.get("created_at"),
            "tipo": str(row.get("tipo_clave") or "").strip() or "Sin clasificar",
            "operador": str(row.get("operador") or "").strip(),
            "detalle": str(row.get("detalle") or "").strip(),
            "observacion": str(row.get("observacion") or "").strip(),
        }
        for row in bit_rows
    ]
    bit_fechas = [r["fecha"] for r in bit_detalle if isinstance(r["fecha"], datetime)]
    bitacora_30d = sum(1 for f in bit_fechas if f >= hace_30_dias)

    return {
        "empresa": empresa_limpia,
        "sucursal": sucursal_limpia,
        "desde": desde.strip() if desde else "",
        "hasta": hasta.strip() if hasta else "",
        "incidencias": {
            "total": len(inc_detalle),
            "pendientes": pendientes_incidencias,
            "cerradas": len(inc_detalle) - pendientes_incidencias,
            "antiguedad_promedio_dias": round(dias_pend_acum / dias_pend_n, 1) if dias_pend_n else 0,
            "resolucion_promedio_dias": round(dias_cierre_acum / dias_cierre_n, 1) if dias_cierre_n else 0,
            "por_tipo": _conteo_ordenado([r["tipo"] for r in inc_detalle], tope=10),
            "por_estado": _conteo_ordenado([r["estado"] or "Sin estado" for r in inc_detalle]),
            "primera_fecha": _fmt_fecha(min(inc_fechas)) if inc_fechas else "",
            "ultima_fecha": _fmt_fecha(max(inc_fechas)) if inc_fechas else "",
            "por_dia": _agrupar_incidencias_por_dia(inc_detalle)[:_MAX_DIAS_ACTIVIDAD],
            "por_dia_truncado": len(_agrupar_incidencias_por_dia(inc_detalle)) > _MAX_DIAS_ACTIVIDAD,
            "_detalle": inc_detalle,
        },
        "protocolos": {
            "total_registros": len(proto_detalle),
            "exitosos": proto_exitosos,
            "no_exitosos": len(proto_detalle) - proto_exitosos,
            "intrusivos": proto_intrusivos,
            "preventivos": proto_preventivos,
            "total_informes": total_informes,
            "informes_pendientes": informes_pendientes,
            "primera_fecha": _fmt_fecha(min(proto_fechas)) if proto_fechas else "",
            "ultima_fecha": _fmt_fecha(max(proto_fechas)) if proto_fechas else "",
            "por_dia": _agrupar_protocolos_por_dia(proto_detalle)[:_MAX_DIAS_ACTIVIDAD],
            "por_dia_truncado": len(_agrupar_protocolos_por_dia(proto_detalle)) > _MAX_DIAS_ACTIVIDAD,
            "_detalle": proto_detalle,
        },
        "bitacora": {
            "total": len(bit_detalle),
            "ultimos_30_dias": bitacora_30d,
            "por_tipo": _conteo_ordenado([r["tipo"] for r in bit_detalle], tope=10),
            "primera_fecha": _fmt_fecha(min(bit_fechas)) if bit_fechas else "",
            "ultima_fecha": _fmt_fecha(max(bit_fechas)) if bit_fechas else "",
            "_detalle": bit_detalle,
        },
    }


@router.get("/api/bitacora/informacion-cliente")
def informacion_cliente(
    empresa: str = Query(default=""),
    sucursal: str = Query(default=""),
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Resumen gerencial con desgloses (Incidencias / Protocolos / Bitácora)
    para una empresa+sucursal, con rango de fechas opcional."""
    _require_bitacora_access(current_user)
    data = _informacion_cliente_data(incidencias_db, empresa, sucursal, desde, hasta)
    # Las filas una-a-una solo van en el informe Excel, no en el JSON del panel.
    for seccion in ("incidencias", "protocolos", "bitacora"):
        data[seccion].pop("_detalle", None)
    return data


@router.get("/api/bitacora/informacion-cliente/informe")
def informacion_cliente_informe(
    empresa: str = Query(default=""),
    sucursal: str = Query(default=""),
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Informe Excel descargable por cliente: hoja de resumen + una hoja de
    detalle fila a fila por cada módulo (Incidencias / Protocolos / Bitácora)."""
    _require_bitacora_access(current_user)
    data = _informacion_cliente_data(incidencias_db, empresa, sucursal, desde, hasta)

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    azul = "0B1424"
    azul_medio = "1E3A5F"
    naranja = "DE7B36"
    borde = Side(style="thin", color="CBD5E1")
    borde_full = Border(top=borde, left=borde, right=borde, bottom=borde)

    rango_txt = "Todo el historial"
    if data["desde"] or data["hasta"]:
        rango_txt = f"Desde {data['desde'] or 'el inicio'} hasta {data['hasta'] or 'hoy'}"

    def _titulo_hoja(ws, titulo: str, n_cols: int) -> None:
        ws.append([titulo])
        ws.append([f"{data['empresa']} — {data['sucursal']}  |  {rango_txt}  |  Generado: {datetime.now().strftime('%d-%m-%Y %H:%M')}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = PatternFill("solid", fgColor=azul)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(bold=True, size=10, color="334155")
        ws["A2"].alignment = Alignment(horizontal="center")

    def _encabezados(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=azul_medio)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde_full

    def _autoancho(ws, anchos: list[int]) -> None:
        for idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = ancho

    def _bordear_datos(ws, desde_fila: int, n_cols: int) -> None:
        for fila in ws.iter_rows(min_row=desde_fila, max_row=ws.max_row, min_col=1, max_col=n_cols):
            for cell in fila:
                cell.border = borde_full
                if cell.font is None or not cell.font.b:
                    cell.font = Font(size=10)

    wb = Workbook()

    # ── Hoja Resumen: cada sección es una tabla horizontal — las variables
    #    van como COLUMNAS (encabezado) y los valores en una sola fila ──
    ws = wb.active
    ws.title = "Resumen"
    _titulo_hoja(ws, "ATC — Informe de Cliente", 9)
    inc, pro, bit = data["incidencias"], data["protocolos"], data["bitacora"]

    def _seccion_horizontal(titulo: str, headers: list[str], valores: list) -> None:
        ws.append([])
        ws.append([titulo])
        celda = ws.cell(row=ws.max_row, column=1)
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor=naranja)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
        _encabezados(ws, headers)
        ws.append(valores)
        for cell in ws[ws.max_row]:
            cell.border = borde_full
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)

    _seccion_horizontal(
        "INCIDENCIAS",
        ["Total", "Pendientes", "Cerradas", "Antigüedad prom. (días)", "Resolución prom. (días)",
         "Primera actividad", "Última actividad"],
        [inc["total"], inc["pendientes"], inc["cerradas"], inc["antiguedad_promedio_dias"],
         inc["resolucion_promedio_dias"], inc["primera_fecha"] or "-", inc["ultima_fecha"] or "-"],
    )
    _seccion_horizontal(
        "PROTOCOLOS",
        ["Activaciones", "Exitosos", "No exitosos", "Intrusivos", "Preventivos",
         "Informes", "Informes pendientes", "Primera actividad", "Última actividad"],
        [pro["total_registros"], pro["exitosos"], pro["no_exitosos"], pro["intrusivos"], pro["preventivos"],
         pro["total_informes"], pro["informes_pendientes"], pro["primera_fecha"] or "-", pro["ultima_fecha"] or "-"],
    )
    _seccion_horizontal(
        "BITÁCORA",
        ["Movimientos totales", "Últimos 30 días", "Primera actividad", "Última actividad"],
        [bit["total"], bit["ultimos_30_dias"], bit["primera_fecha"] or "-", bit["ultima_fecha"] or "-"],
    )

    # Desgloses por tipo — también con los tipos como columnas
    if inc["por_tipo"]:
        _seccion_horizontal("INCIDENCIAS POR TIPO",
                            [str(n) for n, _ in inc["por_tipo"]],
                            [c for _, c in inc["por_tipo"]])
    if inc["por_estado"]:
        _seccion_horizontal("INCIDENCIAS POR ESTADO",
                            [str(n) for n, _ in inc["por_estado"]],
                            [c for _, c in inc["por_estado"]])
    if bit["por_tipo"]:
        _seccion_horizontal("BITÁCORA POR TIPO DE MOVIMIENTO",
                            [str(n) for n, _ in bit["por_tipo"]],
                            [c for _, c in bit["por_tipo"]])
    _autoancho(ws, [18] * 12)

    # ── Hoja Incidencias ──
    ws = wb.create_sheet("Incidencias")
    headers_inc = ["ODT", "Fecha registro", "Tipo (problema)", "Derivación", "Estado",
                   "Fecha cierre", "Días resolución", "Días abierta (si pendiente)", "Técnicos"]
    _titulo_hoja(ws, "Detalle de Incidencias", len(headers_inc))
    _encabezados(ws, headers_inc)
    fila_datos = ws.max_row + 1
    for r in inc["_detalle"]:
        ws.append([
            r["odt"], _fmt_fecha(r["fecha_registro"]), r["tipo"], r["derivacion"], r["estado"],
            _fmt_fecha(r["fecha_cierre"]),
            r["dias_resolucion"] if r["dias_resolucion"] is not None else "",
            r["dias_abierta"] if r["dias_abierta"] is not None else "",
            r["tecnicos"],
        ])
    _bordear_datos(ws, fila_datos, len(headers_inc))
    _autoancho(ws, [12, 17, 30, 18, 14, 17, 13, 13, 24])
    ws.freeze_panes = "A4"

    # ── Hoja Protocolos ──
    ws = wb.create_sheet("Protocolos")
    headers_pro = ["Fecha", "Tipo", "Exitoso", "Detectado", "Efectivo", "Sirena", "Voz",
                   "Carabineros", "Alpha3", "Informado", "Puesto", "Operador"]
    _titulo_hoja(ws, "Detalle de Protocolos", len(headers_pro))
    _encabezados(ws, headers_pro)
    fila_datos = ws.max_row + 1
    for r in pro["_detalle"]:
        ws.append([
            _fmt_fecha(r["fecha_registro"]), r["tipo"], r["exitoso"], r["detectado"], r["efectivo"],
            r["sirena"], r["voz"], r["carabineros"], r["alpha3"], r["informado"], r["puesto"], r["operador"],
        ])
    _bordear_datos(ws, fila_datos, len(headers_pro))
    _autoancho(ws, [17, 13, 10, 10, 10, 9, 9, 12, 9, 11, 12, 22])
    ws.freeze_panes = "A4"

    # ── Hoja Bitácora ──
    ws = wb.create_sheet("Bitácora")
    headers_bit = ["Fecha", "Tipo de movimiento", "Operador", "Detalle", "Observación"]
    _titulo_hoja(ws, "Observación de Bitácora", len(headers_bit))
    _encabezados(ws, headers_bit)
    fila_datos = ws.max_row + 1
    for r in bit["_detalle"]:
        ws.append([_fmt_fecha(r["fecha"]), r["tipo"], r["operador"], r["detalle"], r["observacion"]])
    _bordear_datos(ws, fila_datos, len(headers_bit))
    _autoancho(ws, [17, 24, 20, 45, 45])
    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_ascii = re.sub(r"[^A-Za-z0-9_-]+", "_", f"Informe_{data['empresa']}_{data['sucursal']}")[:80]
    filename = f"{nombre_ascii}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/bitacora/informacion-cliente/informe-pdf")
def informacion_cliente_informe_pdf(
    empresa: str = Query(default=""),
    sucursal: str = Query(default=""),
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Informe PDF gerencial: estado completo del cliente (Incidencias /
    Protocolos / Bitácora) con estilo corporativo, para presentar a gerencia."""
    _require_bitacora_access(current_user)
    data = _informacion_cliente_data(incidencias_db, empresa, sucursal, desde, hasta)

    from io import BytesIO

    from ATC.app.services.informe_cliente_service import generar_informe_cliente_pdf

    pdf_bytes = generar_informe_cliente_pdf(data)
    nombre_ascii = re.sub(r"[^A-Za-z0-9_-]+", "_", f"Informe_Gerencial_{data['empresa']}_{data['sucursal']}")[:80]
    filename = f"{nombre_ascii}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _ensure_sucursal_info_extra_campos_pendientes(db: Session) -> None:
    """Agrega a sucursal_info_extra las columnas que registran qué campos quedaron
    marcados como "falta o está mal" en el último Notificar a Comercial — Venta las
    lee (BBDD Sucursales / Información Clientes) para resaltar esos campos y armar
    el resumen de "lo que rellenó" al avisar que ya quedó listo."""
    for columna, tipo in (
        ("campos_pendientes", "NVARCHAR(MAX)"),
        ("campos_pendientes_obs", "NVARCHAR(MAX)"),
        ("campos_pendientes_fecha", "DATETIME"),
        ("campos_pendientes_por", "NVARCHAR(255)"),
    ):
        try:
            db.execute(text(f"""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME = 'sucursal_info_extra' AND COLUMN_NAME = '{columna}'
                )
                BEGIN
                    ALTER TABLE sucursal_info_extra ADD {columna} {tipo}
                END
            """))
            db.commit()
        except Exception:
            db.rollback()


def _ensure_sucursal_info_extra(db: Session) -> None:
    db.execute(text("""
        IF OBJECT_ID('sucursal_info_extra', 'U') IS NULL
        BEGIN
        CREATE TABLE sucursal_info_extra (
            id BIGINT IDENTITY(1,1) PRIMARY KEY,
            sucursal_id BIGINT NOT NULL UNIQUE,
            referencia_ubicacion NVARCHAR(MAX),
            contacto NVARCHAR(MAX),
            correo NVARCHAR(MAX),
            horario_habil NVARCHAR(MAX),
            horario_no_habil NVARCHAR(MAX),
            fecha_inicio NVARCHAR(MAX),
            tipo_vigilancia NVARCHAR(MAX),
            camaras_contratadas NVARCHAR(MAX),
            camaras_televigiladas NVARCHAR(MAX),
            plan_cuadrante NVARCHAR(MAX),
            carabineros NVARCHAR(MAX),
            bomberos NVARCHAR(MAX),
            seguridad_ciudadana NVARCHAR(MAX),
            codigo_p2p NVARCHAR(MAX),
            codigo_dss NVARCHAR(MAX),
            telefono_porton NVARCHAR(MAX),
            telefono_recepcion NVARCHAR(MAX),
            internet_atc NVARCHAR(MAX),
            compania_electricidad NVARCHAR(MAX),
            numero_cliente_electricidad NVARCHAR(MAX),
            proveedor_internet_cliente NVARCHAR(MAX)
        )
        END
    """))
    db.commit()


@router.get("/api/bitacora/sucursal-raw")
def bitacora_sucursal_raw(
    sucursal_id: int = Query(...),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    base = incidencias_db.execute(
        text("""
            SELECT id, nombre_sucursal, direccion_sucursal, referencia_ubicacion,
                   email_facturas, horario_apertura, horario_cierre,
                   proveedor_electricidad, nro_proveedor_electricidad, proveedor_internet,
                   latitud_longitud
            FROM bbdd_sucursales WHERE id = :sid
        """),
        {"sid": sucursal_id},
    ).mappings().first()
    if not base:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")

    extra = None
    try:
        extra = incidencias_db.execute(
            text("SELECT TOP 1 * FROM sucursal_info_extra WHERE sucursal_id = :sid"),
            {"sid": sucursal_id},
        ).mappings().first()
    except Exception:
        extra = None

    def sv(row, key):
        if not row:
            return ""
        return str(row.get(key) or "").strip()

    return {
        "sucursal_id": sucursal_id,
        "nombre_sucursal": sv(base, "nombre_sucursal"),
        "direccion_sucursal": sv(base, "direccion_sucursal"),
        "latitud_longitud": sv(base, "latitud_longitud"),
        "referencia_ubicacion": sv(extra, "referencia_ubicacion") or sv(base, "referencia_ubicacion"),
        "email_facturas": sv(base, "email_facturas"),
        "horario_apertura": sv(base, "horario_apertura"),
        "horario_cierre": sv(base, "horario_cierre"),
        "horario_habil": sv(extra, "horario_habil"),
        "horario_no_habil": sv(extra, "horario_no_habil"),
        "plan_cuadrante": sv(extra, "plan_cuadrante"),
        "carabineros": sv(extra, "carabineros"),
        "bomberos": sv(extra, "bomberos"),
        "seguridad_ciudadana": sv(extra, "seguridad_ciudadana"),
        "camaras_contratadas": sv(extra, "camaras_contratadas"),
        "camaras_televigiladas": sv(extra, "camaras_televigiladas"),
        "codigo_p2p": sv(extra, "codigo_p2p"),
        "codigo_dss": sv(extra, "codigo_dss"),
        "telefono_porton": sv(extra, "telefono_porton"),
        "telefono_recepcion": sv(extra, "telefono_recepcion"),
        "internet_atc": sv(extra, "internet_atc"),
        "compania_electricidad": sv(extra, "compania_electricidad") or sv(base, "proveedor_electricidad"),
        "numero_cliente_electricidad": sv(extra, "numero_cliente_electricidad") or sv(base, "nro_proveedor_electricidad"),
        "proveedor_internet_cliente": sv(extra, "proveedor_internet_cliente") or sv(base, "proveedor_internet"),
    }


@router.post("/api/bitacora/sucursal-editar")
def bitacora_sucursal_editar(
    payload: SucursalEditPayload,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    sid = payload.sucursal_id

    base_row = incidencias_db.execute(
        text("SELECT id FROM bbdd_sucursales WHERE id = :sid"),
        {"sid": sid},
    ).mappings().first()
    if not base_row:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")

    # Mismo formato combinado "lat, lng" que arma venta_service._split_lat_lng
    # al crear la sucursal; acá solo hace falta separar de vuelta para
    # mantener las 3 columnas (latitud, longitud, latitud_longitud) en sync.
    latlng_combinado = payload.latitud_longitud.strip()
    lat = lng = ""
    if latlng_combinado and "," in latlng_combinado:
        izq, der = latlng_combinado.split(",", 1)
        lat, lng = izq.strip(), der.strip()

    incidencias_db.execute(
        text("""
            UPDATE bbdd_sucursales SET
                nombre_sucursal         = :nombre_sucursal,
                direccion_sucursal      = :direccion_sucursal,
                referencia_ubicacion    = :referencia_ubicacion,
                email_facturas          = :email_facturas,
                horario_apertura        = :horario_apertura,
                horario_cierre          = :horario_cierre,
                dias_funcionamiento     = :horario_habil,
                proveedor_electricidad  = :compania_electricidad,
                nro_proveedor_electricidad = :numero_cliente_electricidad,
                proveedor_internet      = :proveedor_internet_cliente,
                latitud                 = :latitud,
                longitud                = :longitud,
                latitud_longitud        = :latitud_longitud
            WHERE id = :sid
        """),
        {
            "sid": sid,
            "nombre_sucursal": payload.nombre_sucursal.strip(),
            "direccion_sucursal": payload.direccion_sucursal.strip(),
            "referencia_ubicacion": payload.referencia_ubicacion.strip(),
            "email_facturas": payload.email_facturas.strip(),
            "horario_apertura": payload.horario_apertura.strip(),
            "horario_cierre": payload.horario_cierre.strip(),
            "horario_habil": payload.horario_habil.strip(),
            "compania_electricidad": payload.compania_electricidad.strip(),
            "numero_cliente_electricidad": payload.numero_cliente_electricidad.strip(),
            "proveedor_internet_cliente": payload.proveedor_internet_cliente.strip(),
            "latitud": lat,
            "longitud": lng,
            "latitud_longitud": latlng_combinado,
        },
    )

    _ensure_sucursal_info_extra(incidencias_db)
    incidencias_db.execute(
        text("""
            MERGE sucursal_info_extra AS t
            USING (VALUES (:sid)) AS s(sucursal_id) ON t.sucursal_id = s.sucursal_id
            WHEN MATCHED THEN UPDATE SET
                referencia_ubicacion    = :referencia_ubicacion,
                contacto                = :contacto,
                horario_habil           = :horario_habil,
                horario_no_habil        = :horario_no_habil,
                plan_cuadrante          = :plan_cuadrante,
                carabineros             = :carabineros,
                bomberos                = :bomberos,
                seguridad_ciudadana     = :seguridad_ciudadana,
                camaras_contratadas     = :camaras_contratadas,
                camaras_televigiladas   = :camaras_televigiladas,
                codigo_p2p              = :codigo_p2p,
                codigo_dss              = :codigo_dss,
                telefono_porton         = :telefono_porton,
                telefono_recepcion      = :telefono_recepcion,
                internet_atc            = :internet_atc,
                compania_electricidad   = :compania_electricidad,
                numero_cliente_electricidad = :numero_cliente_electricidad,
                proveedor_internet_cliente  = :proveedor_internet_cliente
            WHEN NOT MATCHED THEN INSERT (
                sucursal_id, referencia_ubicacion, contacto, horario_habil, horario_no_habil,
                plan_cuadrante, carabineros, bomberos, seguridad_ciudadana,
                camaras_contratadas, camaras_televigiladas, codigo_p2p, codigo_dss,
                telefono_porton, telefono_recepcion, internet_atc,
                compania_electricidad, numero_cliente_electricidad, proveedor_internet_cliente
            ) VALUES (
                :sid, :referencia_ubicacion, :contacto, :horario_habil, :horario_no_habil,
                :plan_cuadrante, :carabineros, :bomberos, :seguridad_ciudadana,
                :camaras_contratadas, :camaras_televigiladas, :codigo_p2p, :codigo_dss,
                :telefono_porton, :telefono_recepcion, :internet_atc,
                :compania_electricidad, :numero_cliente_electricidad, :proveedor_internet_cliente
            );
        """),
        {
            "sid": sid,
            "referencia_ubicacion": payload.referencia_ubicacion.strip(),
            "contacto": payload.contacto.strip(),
            "horario_habil": payload.horario_habil.strip(),
            "horario_no_habil": payload.horario_no_habil.strip(),
            "plan_cuadrante": payload.plan_cuadrante.strip(),
            "carabineros": payload.carabineros.strip(),
            "bomberos": payload.bomberos.strip(),
            "seguridad_ciudadana": payload.seguridad_ciudadana.strip(),
            "camaras_contratadas": payload.camaras_contratadas.strip(),
            "camaras_televigiladas": payload.camaras_televigiladas.strip(),
            "codigo_p2p": payload.codigo_p2p.strip(),
            "codigo_dss": payload.codigo_dss.strip(),
            "telefono_porton": payload.telefono_porton.strip(),
            "telefono_recepcion": payload.telefono_recepcion.strip(),
            "internet_atc": payload.internet_atc.strip(),
            "compania_electricidad": payload.compania_electricidad.strip(),
            "numero_cliente_electricidad": payload.numero_cliente_electricidad.strip(),
            "proveedor_internet_cliente": payload.proveedor_internet_cliente.strip(),
        },
    )
    incidencias_db.commit()
    return {"ok": True}


@router.patch("/api/bitacora/sucursal/{sucursal_id}/deshabilitar")
def bitacora_sucursal_deshabilitar(
    sucursal_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.execute(
        text("SELECT id FROM bbdd_sucursales WHERE id = :sid"),
        {"sid": sucursal_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")
    incidencias_db.execute(
        text("UPDATE bbdd_sucursales SET habilitada = 0 WHERE id = :sid"),
        {"sid": sucursal_id},
    )
    incidencias_db.commit()
    return {"ok": True}


# ──────────────────────────────────────────────
# Información Puestos — cámaras, sucursales, incidencias, protocolos y
# movimientos de bitácora, agrupados por puesto de monitoreo (1-29) y por
# sucursal dentro de cada puesto. Todo el cruce con incidencias/protocolos/
# bitacora_registros es por nombre de texto (no hay FK), igual criterio que
# _informacion_cliente_data.
# ──────────────────────────────────────────────

def _informacion_puestos_data(incidencias_db: Session, desde: str, hasta: str) -> dict:
    inicio, fin_excl = _parse_rango_fechas(desde, hasta)

    def _filtro_fecha(columna: str) -> tuple[str, dict]:
        sql, params = "", {}
        if inicio:
            sql += f" AND {columna} >= :rango_inicio"
            params["rango_inicio"] = inicio
        if fin_excl:
            sql += f" AND {columna} < :rango_fin"
            params["rango_fin"] = fin_excl
        return sql, params

    # ── Mapa sucursal -> puesto, a partir de las cámaras ya asignadas ──
    cam_rows = (
        incidencias_db.query(
            SucursalCamaraMonitoreo.central,
            SucursalBBDD.nombre_sucursal,
            SucursalCamaraMonitoreo.nombre_camara_monitoreo,
        )
        .join(SucursalBBDD, SucursalCamaraMonitoreo.sucursal_id == SucursalBBDD.id)
        .filter(SucursalCamaraMonitoreo.central.isnot(None))
        .all()
    )

    # puesto[n] = {"sucursales": {clave_norm: nombre_real}, "cam_mon": n}
    puestos: dict[int, dict] = {
        p: {"sucursales": {}, "cam_mon": 0} for p in range(1, 30)
    }
    # sucursal_clave -> {"puesto": n, "nombre": str, "cam_mon": n}
    sucursal_info: dict[str, dict] = {}
    for central, nombre_sucursal, cam_mon in cam_rows:
        nombre = (nombre_sucursal or "").strip()
        if not nombre:
            continue
        clave = _normalizar_texto_cam(nombre)
        puesto = puestos[int(central)]
        puesto["sucursales"].setdefault(clave, nombre)
        info = sucursal_info.setdefault(
            clave, {"puesto": int(central), "nombre": nombre, "cam_mon": 0}
        )
        if cam_mon and cam_mon.strip():
            puesto["cam_mon"] += 1
            info["cam_mon"] += 1

    # contadores por sucursal para incidencias/protocolos/bitacora, se
    # inicializan en 0 para que toda sucursal con cámaras aparezca aunque no
    # tenga ningún registro en el período.
    def _contadores_inc() -> dict:
        return {"total": 0, "pendientes": 0}

    def _contadores_proto() -> dict:
        return {"preventivos": 0, "intrusivos": 0}

    incidencias_por_sucursal: dict[str, dict] = defaultdict(_contadores_inc)
    protocolos_por_sucursal: dict[str, dict] = defaultdict(_contadores_proto)
    bitacora_por_sucursal: dict[str, int] = defaultdict(int)

    filtro_inc, params_inc = _filtro_fecha("fecha_registro")
    inc_rows = incidencias_db.execute(
        text(
            "SELECT cliente, estado FROM incidencias WHERE 1=1" + filtro_inc
        ),
        params_inc,
    ).mappings().all()
    for row in inc_rows:
        clave = _normalizar_texto_cam(row.get("cliente"))
        if clave not in sucursal_info:
            continue
        estado_norm = str(row.get("estado") or "").strip().lower()
        contador = incidencias_por_sucursal[clave]
        contador["total"] += 1
        if estado_norm not in _ESTADOS_INCIDENCIA_CERRADA:
            contador["pendientes"] += 1

    filtro_proto, params_proto = _filtro_fecha("fecha_registro")
    proto_rows = incidencias_db.execute(
        text("SELECT sucursal, tipo_protocolo FROM protocolos_registro WHERE 1=1" + filtro_proto),
        params_proto,
    ).mappings().all()
    for row in proto_rows:
        clave = _normalizar_texto_cam(row.get("sucursal"))
        if clave not in sucursal_info:
            continue
        tipo_norm = str(row.get("tipo_protocolo") or "").strip().lower()
        contador = protocolos_por_sucursal[clave]
        if tipo_norm == "intrusivo":
            contador["intrusivos"] += 1
        elif tipo_norm == "preventivo":
            contador["preventivos"] += 1

    # Agrupado en SQL, no en Python: bitacora_registros ya tiene ~1M+ filas
    # (histórico migrado desde BDATC) y traerlas una por una satura el pool
    # de conexiones y afecta a otros usuarios reales del sistema.
    _ensure_bitacora_registros_table(incidencias_db)
    filtro_bit, params_bit = _filtro_fecha("created_at")
    bit_rows = incidencias_db.execute(
        text(
            "SELECT nombre_sucursal, COUNT(*) AS cnt FROM bitacora_registros WHERE 1=1"
            + filtro_bit + " GROUP BY nombre_sucursal"
        ),
        params_bit,
    ).mappings().all()
    for row in bit_rows:
        clave = _normalizar_texto_cam(row.get("nombre_sucursal"))
        if clave in sucursal_info:
            bitacora_por_sucursal[clave] += int(row.get("cnt") or 0)

    # ── Armado de la salida: resumen por puesto + detalle por sucursal ──
    salida_puestos = []
    for n in range(1, 30):
        p = puestos[n]
        detalle = []
        inc_total = inc_pend = proto_prev = proto_intr = bit_total = 0
        for clave, nombre in sorted(p["sucursales"].items(), key=lambda kv: kv[1].casefold()):
            info = sucursal_info[clave]
            inc = incidencias_por_sucursal.get(clave, {"total": 0, "pendientes": 0})
            proto = protocolos_por_sucursal.get(clave, {"preventivos": 0, "intrusivos": 0})
            bit = bitacora_por_sucursal.get(clave, 0)
            inc_total += inc["total"]
            inc_pend += inc["pendientes"]
            proto_prev += proto["preventivos"]
            proto_intr += proto["intrusivos"]
            bit_total += bit
            detalle.append({
                "sucursal": nombre,
                "camaras_monitoreadas": info["cam_mon"],
                "incidencias": inc["total"],
                "incidencias_pendientes": inc["pendientes"],
                "protocolos_preventivos": proto["preventivos"],
                "protocolos_intrusivos": proto["intrusivos"],
                "movimientos_bitacora": bit,
            })
        salida_puestos.append({
            "puesto": n,
            "camaras_monitoreadas": p["cam_mon"],
            "sucursales": len(p["sucursales"]),
            "incidencias": inc_total,
            "incidencias_pendientes": inc_pend,
            "protocolos_preventivos": proto_prev,
            "protocolos_intrusivos": proto_intr,
            "movimientos_bitacora": bit_total,
            "detalle_sucursales": detalle,
        })

    return {
        "desde": desde.strip() if desde else "",
        "hasta": hasta.strip() if hasta else "",
        "puestos": salida_puestos,
    }


_TOP_N_INTRUSIVOS = 6  # debe coincidir con _TOP_N en informe_puestos_service.py


def _pantallas_puesto(puesto: int) -> int:
    """Cantidad de pantallas físicas del puesto: 1-12 tienen 4, 13-29 tienen
    6 (valor fijo validado — la columna "Ubicación de Pantalla" del archivo
    de cámaras tiene variantes de texto inconsistentes, ej. "IZQ ARRIBA" vs
    "IZQUIERDA ARRIBA", así que no se deriva contando valores únicos ahí)."""
    return 4 if puesto <= 12 else 6

_RE_HORA_TEXTO = re.compile(r"\b([01]?\d|2[0-3]):([0-5]\d)\b")

_FRANJAS_HORARIAS = (
    ("Madrugada (00-06h)", 0, 6),
    ("Mañana (06-12h)", 6, 12),
    ("Tarde (12-18h)", 12, 18),
    ("Noche (18-24h)", 18, 24),
)


def _hora_aproximada_evento(observaciones_raw: str | None, fecha_registro: datetime | None) -> tuple[int | None, str]:
    """La hora de fecha_registro es la del REGISTRO posterior del protocolo,
    no necesariamente la hora real en que ocurrió la intrusión — a menudo el
    operador escribe el parte a la mañana siguiente resumiendo la noche. En
    cambio observaciones_raw casi siempre arranca o menciona la hora real
    ("a las 22:22 hrs", "Se recibe alarma... a las 00:47"), así que se usa
    esa hora como aproximación cuando está disponible (solo hora del día,
    sin tocar la fecha — reconstruir la fecha exacta sería adivinar) y solo
    se cae a la hora de fecha_registro si el texto no trae ninguna hora."""
    m = _RE_HORA_TEXTO.search(str(observaciones_raw or ""))
    if m:
        return int(m.group(1)), "texto"
    if fecha_registro:
        return fecha_registro.hour, "registro"
    return None, "sin_dato"


def _franja_horaria(hora: int | None) -> str | None:
    if hora is None:
        return None
    for nombre, ini, fin in _FRANJAS_HORARIAS:
        if ini <= hora < fin:
            return nombre
    return None


def _bitacora_ventana_por_puesto(
    incidencias_db: Session, sucursales_originales: list[str], minutos: int = 30
) -> list[datetime]:
    """Trae SOLO los created_at de bitácora de las sucursales de un puesto
    puntual (acotado por nombre_sucursal, usa el índice ix_nombre_sucursal)
    para poder contar, con bisect, cuántos movimientos cayeron en una
    ventana de +/- `minutos` alrededor de cada evento sin repetir una query
    por evento. Deliberadamente NO se hace esto para los 29 puestos —desde
    el incidente de rendimiento con el histórico completo (ver notas de
    _informacion_puestos_data), solo se llama para los puestos del top que
    de verdad se narran en el PDF."""
    if not sucursales_originales:
        return []
    rows = incidencias_db.execute(
        text(
            "SELECT created_at FROM bitacora_registros WHERE nombre_sucursal IN :sucursales"
        ).bindparams(bindparam("sucursales", expanding=True)),
        {"sucursales": sucursales_originales},
    ).all()
    return sorted(r[0] for r in rows if r[0])


def _contar_en_ventana(timestamps_ordenados: list[datetime], centro: datetime, minutos: int = 15) -> int:
    """Cuenta timestamps en los `minutos` ANTERIORES al evento (no hacia
    adelante): lo que interesa es la carga previa que pudo distraer al
    operador antes de la intrusión, no lo que pasó después."""
    if not timestamps_ordenados or not centro:
        return 0
    ini = centro - timedelta(minutes=minutos)
    izq = bisect.bisect_left(timestamps_ordenados, ini)
    der = bisect.bisect_right(timestamps_ordenados, centro)
    return der - izq


def _movimientos_operador_ventana(incidencias_db: Session, operador: str, centro: datetime, minutos: int = 15) -> int:
    """Cuántos movimientos de bitácora escribió ESE operador puntual en los
    `minutos` ANTERIORES al intrusivo (no hacia adelante). Un operador cubre
    un puesto completo (varias sucursales) como asignación normal, así que
    contar en cuántas sucursales distintas tuvo actividad no dice mucho —
    eso es simplemente su pega de siempre. Lo que sí es una señal real de
    qué tan ocupado/distraído estaba justo antes del evento es el volumen de
    bitácora que él mismo redactó en esa ventana previa."""
    operador = (operador or "").strip()
    if not operador or not centro:
        return 0
    ini = centro - timedelta(minutes=minutos)
    total = incidencias_db.execute(
        text(
            "SELECT COUNT(*) FROM bitacora_registros "
            "WHERE operador = :op AND created_at BETWEEN :ini AND :fin"
        ),
        {"op": operador, "ini": ini, "fin": centro},
    ).scalar()
    return int(total or 0)


def _calcular_senales_carga_alta(puestos: list[dict]) -> None:
    """Por puesto, calcula 4 métricas normalizadas por cámara monitoreada
    (protocolos/cámara, movimientos de bitácora/cámara, incidencias activas
    simultáneas promedio, cámaras/pantalla) y marca cuáles de esas 4 están
    por encima del promedio del grupo — en vez de un índice ponderado 0-100
    (una fórmula que hay que confiar a ciegas), se cuenta cuántas señales de
    carga alta tiene cada puesto, algo que se puede verificar a simple vista
    comparando cada valor contra su propio promedio en la misma tabla. Es
    una comparación relativa entre los puestos de ESTE informe, no una
    escala absoluta.

    Se normaliza por CÁMARAS monitoreadas, no por sucursales: la cantidad
    de sucursales por puesto varía mucho (4 a 35), pero las cámaras están
    parejas entre puestos (~80-120) — dividir por sucursales infla la
    métrica de los puestos con pocas sucursales grandes sin que eso refleje
    más carga real de monitoreo.

    Cámaras/pantalla: cuántas cámaras le tocan repartidas entre las
    pantallas físicas del puesto (4 pantallas en puestos 1-12, 6 en
    puestos 13-29) — a más cámaras por pantalla, más carga visual real para
    el operador aunque el total de cámaras del puesto sea similar."""
    def _prom_incidencias(p: dict) -> float:
        detalle = p.get("intrusivos_detalle") or []
        if not detalle:
            return 0.0
        return sum(len(ev["incidencias_activas"]) for ev in detalle) / len(detalle)

    # Puestos sin sucursales/cámaras asignadas (ej. Puesto 8) quedan fuera
    # del cálculo: no tiene sentido asignarles una carga de monitoreo si no
    # monitorean nada. Igual se les deja el resto de sus datos (sucursales,
    # movimientos) intactos para que sigan apareciendo en la tabla de
    # "Puestos sin intrusiones".
    puestos_validos = [p for p in puestos if p["camaras_monitoreadas"] > 0]

    for p in puestos_validos:
        camaras = p["camaras_monitoreadas"]
        p["protocolos_por_camara"] = round(p["protocolos_preventivos"] + p["protocolos_intrusivos"], 2) / camaras
        p["movimientos_por_camara"] = round(p["movimientos_bitacora"] / camaras, 1)
        p["incidencias_simultaneas_promedio"] = round(_prom_incidencias(p), 2)
        pantallas = _pantallas_puesto(p["puesto"])
        p["camaras_por_pantalla"] = round(camaras / pantallas, 2)

    def _promedio(campo: str) -> float:
        valores = [p[campo] for p in puestos_validos]
        return sum(valores) / len(valores) if valores else 0.0

    prom_proto = _promedio("protocolos_por_camara")
    prom_mov = _promedio("movimientos_por_camara")
    prom_inc = _promedio("incidencias_simultaneas_promedio")
    prom_cam_pant = _promedio("camaras_por_pantalla")

    for p in puestos_validos:
        p["protocolos_camara_alto"] = p["protocolos_por_camara"] > prom_proto
        p["movimientos_camara_alto"] = p["movimientos_por_camara"] > prom_mov
        p["incidencias_simultaneas_alto"] = p["incidencias_simultaneas_promedio"] > prom_inc
        p["camaras_pantalla_alto"] = p["camaras_por_pantalla"] > prom_cam_pant
        p["senales_carga_alta"] = sum([
            p["protocolos_camara_alto"],
            p["movimientos_camara_alto"],
            p["incidencias_simultaneas_alto"],
            p["camaras_pantalla_alto"],
        ])

    for p in puestos:
        if p["camaras_monitoreadas"] <= 0:
            p["senales_carga_alta"] = None


def _informe_puestos_dataset(incidencias_db: Session, desde: str, hasta: str) -> dict:
    """Extiende _informacion_puestos_data con, por puesto, el detalle de sus
    protocolos intrusivos (incidencias activas, operador, desenlace, hora
    aproximada) y métricas normalizadas por sucursal + un índice de carga
    combinado — para el informe PDF de análisis de intrusiones."""
    data = _informacion_puestos_data(incidencias_db, desde, hasta)

    sucursal_a_puesto: dict[str, int] = {}
    sucursales_por_puesto: dict[int, list[str]] = defaultdict(list)
    for p in data["puestos"]:
        for s in p["detalle_sucursales"]:
            clave = _normalizar_texto_cam(s["sucursal"])
            sucursal_a_puesto[clave] = p["puesto"]
            sucursales_por_puesto[p["puesto"]].append(s["sucursal"])

    inicio, fin_excl = _parse_rango_fechas(desde, hasta)
    filtro, params = "", {}
    if inicio:
        filtro += " AND fecha_registro >= :rango_inicio"
        params["rango_inicio"] = inicio
    if fin_excl:
        filtro += " AND fecha_registro < :rango_fin"
        params["rango_fin"] = fin_excl
    proto_rows = incidencias_db.execute(
        text(
            "SELECT sucursal, fecha_registro, tipo_protocolo, operador, protocolo_exitoso, observaciones_raw "
            "FROM protocolos_registro WHERE 1=1" + filtro
        ),
        params,
    ).mappings().all()

    eventos_por_puesto: dict[int, list] = defaultdict(list)
    franja_todos = Counter()
    franja_intrusivos = Counter()
    for row in proto_rows:
        clave = _normalizar_texto_cam(row.get("sucursal"))
        puesto = sucursal_a_puesto.get(clave)
        fecha = row.get("fecha_registro")
        es_intrusivo = str(row.get("tipo_protocolo") or "").strip().lower() == "intrusivo"
        hora, fuente_hora = _hora_aproximada_evento(row.get("observaciones_raw"), fecha)
        franja = _franja_horaria(hora)
        if franja:
            franja_todos[franja] += 1
            if es_intrusivo:
                franja_intrusivos[franja] += 1
        if puesto and fecha and es_intrusivo:
            eventos_por_puesto[puesto].append(
                {
                    "fecha": fecha,
                    "operador": str(row.get("operador") or "").strip() or "Sin operador registrado",
                    "exitoso": str(row.get("protocolo_exitoso") or "").strip().upper(),
                    "hora_aprox": hora,
                    "hora_fuente": fuente_hora,
                    "sucursal": str(row.get("sucursal") or "").strip() or "Sin sucursal registrada",
                    "sucursal_clave": clave,
                }
            )
    data["franja_horaria"] = {
        "todos": dict(franja_todos),
        "intrusivos": dict(franja_intrusivos),
        "orden": [nombre for nombre, _, _ in _FRANJAS_HORARIAS],
    }

    # Incidencias SIN filtro de fecha: una incidencia abierta antes del
    # período (o cerrada después) igual cuenta como "activa" en la fecha
    # puntual del intrusivo, aunque esa fecha de apertura/cierre caiga
    # fuera del rango elegido para el informe.
    #
    # Dato sucio conocido: ~4.6k incidencias quedaron marcadas "terminado"
    # sin que se les cargara fecha_cierre. Si solo mirásemos fecha_cierre,
    # esas quedarían "activas para siempre" e inflarían el promedio. Por
    # eso una incidencia solo cuenta como activa en fecha_evento si:
    # - su estado ACTUAL no es de cierre (sigue realmente pendiente), o
    # - tiene fecha_cierre cargada y esa fecha es posterior al evento.
    #
    # Agrupadas por SUCURSAL (no por puesto): un puesto cubre varias
    # sucursales, y una incidencia en otra sucursal del mismo puesto no
    # explica por qué no se detectó una intrusión en ESTA sucursal — el
    # cruce causal tiene que ser sucursal contra sucursal.
    inc_rows = incidencias_db.execute(
        text("SELECT cliente, fecha_registro, fecha_cierre, estado, problema FROM incidencias")
    ).mappings().all()
    inc_por_sucursal: dict[str, list] = defaultdict(list)
    for row in inc_rows:
        clave = _normalizar_texto_cam(row.get("cliente"))
        if clave and row.get("fecha_registro"):
            estado_norm = str(row.get("estado") or "").strip().lower()
            cerrada_actualmente = estado_norm in _ESTADOS_INCIDENCIA_CERRADA
            problema = str(row.get("problema") or "").strip() or "Sin clasificar"
            inc_por_sucursal[clave].append(
                (row.get("fecha_registro"), row.get("fecha_cierre"), cerrada_actualmente, problema)
            )

    # Puestos "top" por intrusivos: son los únicos que el PDF narra en
    # detalle, así que son los únicos para los que vale la pena pagar el
    # costo extra de consultar ventanas de bitácora/operador por evento.
    top_puestos_ids = {
        p["puesto"]
        for p in sorted(data["puestos"], key=lambda x: x["protocolos_intrusivos"], reverse=True)[:_TOP_N_INTRUSIVOS]
        if p["protocolos_intrusivos"] > 0
    }
    ventana_bitacora_cache: dict[int, list[datetime]] = {}

    for p in data["puestos"]:
        eventos = sorted(eventos_por_puesto.get(p["puesto"], []), key=lambda e: e["fecha"])
        es_top = p["puesto"] in top_puestos_ids
        if es_top:
            ventana_bitacora_cache[p["puesto"]] = _bitacora_ventana_por_puesto(
                incidencias_db, sucursales_por_puesto.get(p["puesto"], [])
            )
        detalle_eventos = []
        for ev in eventos:
            fecha_evento = ev["fecha"]
            incidencias_sucursal = inc_por_sucursal.get(ev["sucursal_clave"], [])
            activas = [
                problema
                for (f_reg, f_cie, cerrada_actualmente, problema) in incidencias_sucursal
                if f_reg <= fecha_evento
                and ((f_cie and f_cie > fecha_evento) or (not f_cie and not cerrada_actualmente))
            ]
            item = {
                "fecha": fecha_evento,
                "incidencias_activas": activas,
                "operador": ev["operador"],
                "exitoso": ev["exitoso"],
                "hora_aprox": ev["hora_aprox"],
                "hora_fuente": ev["hora_fuente"],
                "sucursal": ev["sucursal"],
            }
            if es_top:
                ts = ventana_bitacora_cache[p["puesto"]]
                item["movimientos_ventana_15min"] = _contar_en_ventana(ts, fecha_evento)
                item["movimientos_operador_15min"] = _movimientos_operador_ventana(
                    incidencias_db, ev["operador"], fecha_evento
                )
            detalle_eventos.append(item)
        p["intrusivos_fechas"] = [ev["fecha"] for ev in eventos]
        p["intrusivos_detalle"] = detalle_eventos

        operadores_intrusivos = Counter(ev["operador"] for ev in eventos if ev["operador"] != "Sin operador registrado")
        p["operadores_repetidos"] = [
            {"operador": op, "veces": n} for op, n in operadores_intrusivos.most_common() if n >= 2
        ]
        p["desenlace_no_exitoso"] = sum(1 for ev in eventos if ev["exitoso"] == "NO")

    _calcular_senales_carga_alta(data["puestos"])

    return data


@router.get("/api/bitacora/informacion-puestos")
def informacion_puestos(
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Cámaras, sucursales, incidencias, protocolos y movimientos de bitácora
    agrupados por puesto (1-29) y por sucursal dentro de cada puesto, con
    rango de fechas opcional para incidencias/protocolos/bitácora."""
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    return _informacion_puestos_data(incidencias_db, desde, hasta)


def _puestos_qs(desde: str, hasta: str) -> str:
    from urllib.parse import urlencode

    params = {}
    if desde:
        params["desde"] = desde
    if hasta:
        params["hasta"] = hasta
    return urlencode(params)


def _puestos_periodo_label(data: dict) -> str:
    if data.get("desde") or data.get("hasta"):
        return f"Período: {data.get('desde') or 'inicio'} → {data.get('hasta') or 'hoy'}"
    return "Período: histórico completo"


def _puestos_rankings(puestos: list[dict]) -> dict[str, list[dict]]:
    """Enriquece cada puesto con protocolos_total, y devuelve 4 listas ya
    ordenadas de mayor a menor según cada criterio pedido."""
    enriquecidos = []
    for p in puestos:
        total_proto = p["protocolos_preventivos"] + p["protocolos_intrusivos"]
        enriquecidos.append({**p, "protocolos_total": total_proto})
    return {
        "protocolos_total": sorted(enriquecidos, key=lambda p: -p["protocolos_total"]),
        "protocolos_intrusivos": sorted(enriquecidos, key=lambda p: -p["protocolos_intrusivos"]),
        "protocolos_preventivos": sorted(enriquecidos, key=lambda p: -p["protocolos_preventivos"]),
        "movimientos_bitacora": sorted(enriquecidos, key=lambda p: -p["movimientos_bitacora"]),
    }


_PUESTOS_EXCEL_HEADERS = [
    "Puesto", "Sucursales", "Cámaras monitoreadas", "Incidencias", "Incidencias pendientes",
    "Protocolos preventivos", "Protocolos intrusivos", "Protocolos totales",
    "Movimientos bitácora",
]


def _fila_puesto_valores(p: dict) -> list:
    return [
        f"Puesto {p['puesto']}", p["sucursales"], p["camaras_monitoreadas"],
        p["incidencias"], p["incidencias_pendientes"],
        p["protocolos_preventivos"], p["protocolos_intrusivos"], p["protocolos_total"],
        p["movimientos_bitacora"],
    ]


def _preview_cell_ip(value: object, css: str = "") -> dict:
    return {"value": value if value not in (None, "") else "—", "css": css}


def _preview_puestos_section(titulo: str, rows: list[dict]) -> dict:
    preview_rows = []
    for p in rows:
        preview_rows.append([
            _preview_cell_ip(f"Puesto {p['puesto']}"),
            _preview_cell_ip(p["sucursales"], "num"),
            _preview_cell_ip(p["camaras_monitoreadas"], "num"),
            _preview_cell_ip(p["incidencias"], "num"),
            _preview_cell_ip(p["incidencias_pendientes"], "num" + (" bad" if p["incidencias_pendientes"] else "")),
            _preview_cell_ip(p["protocolos_preventivos"], "num"),
            _preview_cell_ip(p["protocolos_intrusivos"], "num"),
            _preview_cell_ip(p["protocolos_total"], "num"),
            _preview_cell_ip(p["movimientos_bitacora"], "num"),
        ])
    return {
        "title": titulo,
        "subtitle": f"{len(rows)} puesto(s)",
        "headers": _PUESTOS_EXCEL_HEADERS,
        "rows": preview_rows,
    }


def _crear_hoja_ranking_puestos(ws, titulo_hoja: str, periodo: str, rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = _PUESTOS_EXCEL_HEADERS

    ws.append([f"ATC - {titulo_hoja}"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for p in rows:
        ws.append(_fila_puesto_valores(p))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left", vertical="center")

    widths = [12, 12, 18, 12, 16, 18, 16, 16, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


@router.get("/api/bitacora/informacion-puestos/excel/preview", response_class=HTMLResponse)
def informacion_puestos_excel_preview(
    request: Request,
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    data = _informacion_puestos_data(incidencias_db, desde, hasta)
    rankings = _puestos_rankings(data["puestos"])
    sections = [
        _preview_puestos_section("Protocolos (Total)", rankings["protocolos_total"]),
        _preview_puestos_section("Protocolos Intrusivos", rankings["protocolos_intrusivos"]),
        _preview_puestos_section("Protocolos Preventivos", rankings["protocolos_preventivos"]),
        _preview_puestos_section("Movimientos Bitácora", rankings["movimientos_bitacora"]),
    ]
    return templates.TemplateResponse(
        request,
        "guardias_informe_preview.html",
        {
            "request": request,
            "titulo": "Vista previa — Información Puestos (Excel)",
            "periodo": _puestos_periodo_label(data),
            "sections": sections,
            "descarga_url": f"/api/bitacora/informacion-puestos/excel?{_puestos_qs(desde, hasta)}",
        },
    )


@router.get("/api/bitacora/informacion-puestos/excel")
def informacion_puestos_excel(
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Excel con 4 hojas, cada una ordenada de mayor a menor según un
    criterio distinto: protocolos totales, intrusivos, preventivos, y
    movimientos de bitácora."""
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    data = _informacion_puestos_data(incidencias_db, desde, hasta)
    rankings = _puestos_rankings(data["puestos"])
    periodo = _puestos_periodo_label(data)

    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Protocolos Total"
    _crear_hoja_ranking_puestos(ws, "Protocolos (Total)", periodo, rankings["protocolos_total"])
    _crear_hoja_ranking_puestos(wb.create_sheet("Protocolos Intrusivos"), "Protocolos Intrusivos", periodo, rankings["protocolos_intrusivos"])
    _crear_hoja_ranking_puestos(wb.create_sheet("Protocolos Preventivos"), "Protocolos Preventivos", periodo, rankings["protocolos_preventivos"])
    _crear_hoja_ranking_puestos(wb.create_sheet("Movimientos Bitácora"), "Movimientos Bitácora", periodo, rankings["movimientos_bitacora"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"informacion_puestos_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _informacion_puestos_pdf_bytes(incidencias_db: Session, desde: str, hasta: str) -> bytes:
    from ATC.app.services.informe_puestos_service import generar_informe_puestos_pdf

    data = _informe_puestos_dataset(incidencias_db, desde, hasta)
    return generar_informe_puestos_pdf(data)


@router.get("/api/bitacora/informacion-puestos/pdf/preview")
def informacion_puestos_pdf_preview(
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Vista previa: el PDF real, mostrado inline en el visor del navegador
    (no una tabla HTML) — el propio visor trae su botón de descarga."""
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    from io import BytesIO

    pdf_bytes = _informacion_puestos_pdf_bytes(incidencias_db, desde, hasta)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="Informe_Puestos_vista_previa.pdf"'},
    )


@router.get("/api/bitacora/informacion-puestos/pdf")
def informacion_puestos_pdf(
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """PDF de análisis: top puestos por cada métrica y cruce entre
    protocolos intrusivos y movimientos de bitácora, para chequear si la
    cantidad de movimientos de bitácora acompaña a los intrusivos."""
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    from io import BytesIO

    pdf_bytes = _informacion_puestos_pdf_bytes(incidencias_db, desde, hasta)
    filename = f"Informe_Puestos_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_RECINTO_EXCEL_HEADERS = [
    "Sucursal", "Cámaras monitoreadas", "Incidencias", "Incidencias pendientes",
    "Protocolos preventivos", "Protocolos intrusivos", "Movimientos bitácora", "Actividad total",
]


def _actividad_total(s: dict) -> int:
    return s["incidencias"] + s["protocolos_preventivos"] + s["protocolos_intrusivos"] + s["movimientos_bitacora"]


def _detalle_sucursales_top(p: dict) -> list[dict]:
    return sorted(p["detalle_sucursales"], key=lambda s: -_actividad_total(s))


def _fila_recinto_valores(s: dict) -> list:
    return [
        s["sucursal"], s["camaras_monitoreadas"], s["incidencias"], s["incidencias_pendientes"],
        s["protocolos_preventivos"], s["protocolos_intrusivos"], s["movimientos_bitacora"],
        _actividad_total(s),
    ]


def _preview_recinto_section(p: dict) -> dict:
    top = _detalle_sucursales_top(p)
    preview_rows = []
    for s in top:
        preview_rows.append([
            _preview_cell_ip(s["sucursal"]),
            _preview_cell_ip(s["camaras_monitoreadas"], "num"),
            _preview_cell_ip(s["incidencias"], "num"),
            _preview_cell_ip(s["incidencias_pendientes"], "num" + (" bad" if s["incidencias_pendientes"] else "")),
            _preview_cell_ip(s["protocolos_preventivos"], "num"),
            _preview_cell_ip(s["protocolos_intrusivos"], "num"),
            _preview_cell_ip(s["movimientos_bitacora"], "num"),
            _preview_cell_ip(_actividad_total(s), "num"),
        ])
    return {
        "title": f"Puesto {p['puesto']}",
        "subtitle": f"{p['sucursales']} sucursal(es), ordenado por actividad total",
        "headers": _RECINTO_EXCEL_HEADERS,
        "rows": preview_rows,
    }


def _crear_hoja_recinto(ws, puesto_num: int, periodo: str, top: list[dict]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    azul = "0B1424"
    azul_medio = "1E3A5F"
    borde = Side(style="thin", color="CBD5E1")
    headers = _RECINTO_EXCEL_HEADERS

    ws.append([f"ATC - Puesto {puesto_num}"])
    ws.append([periodo])
    ws.append([])
    ws.append(headers)

    for s in top:
        ws.append(_fila_recinto_valores(s))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, color="334155")
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=azul_medio)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    for fila in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in fila:
            cell.border = Border(top=borde, left=borde, right=borde, bottom=borde)
            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left", vertical="center")

    widths = [34, 18, 12, 16, 18, 16, 16, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"


@router.get("/api/bitacora/informacion-puestos/recintos/preview", response_class=HTMLResponse)
def informacion_puestos_recintos_preview(
    request: Request,
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    data = _informacion_puestos_data(incidencias_db, desde, hasta)
    sections = [_preview_recinto_section(p) for p in data["puestos"]]
    return templates.TemplateResponse(
        request,
        "guardias_informe_preview.html",
        {
            "request": request,
            "titulo": "Vista previa — Detalle por Recinto",
            "periodo": _puestos_periodo_label(data),
            "sections": sections,
            "descarga_url": f"/api/bitacora/informacion-puestos/recintos?{_puestos_qs(desde, hasta)}",
        },
    )


@router.get("/api/bitacora/informacion-puestos/recintos")
def informacion_puestos_recintos(
    desde: str = Query(default=""),
    hasta: str = Query(default=""),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Excel con 29 hojas (una por puesto/recinto), cada una con el top
    desglosado por sucursal de movimientos de bitácora, protocolos e
    incidencias, ordenado de mayor a menor actividad total."""
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver Información Puestos.")
    data = _informacion_puestos_data(incidencias_db, desde, hasta)
    periodo = _puestos_periodo_label(data)

    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    primero = True
    for p in data["puestos"]:
        ws = wb.active if primero else wb.create_sheet()
        ws.title = f"Puesto {p['puesto']}"
        primero = False
        _crear_hoja_recinto(ws, p["puesto"], periodo, _detalle_sucursales_top(p))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"informacion_puestos_detalle_recinto_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──────────────────────────────────────────────
# Sucursales pendientes de aceptación
# ──────────────────────────────────────────────

def _mejor_posible_duplicado(db: Session, sucursal_id: int, rut: str, nombre_sucursal: str) -> dict | None:
    """Entre las sucursales YA aceptadas del mismo RUT, busca la más parecida por
    nombre (tildes/mayúsculas ignoradas) para ayudar al revisor a detectar duplicados
    que el chequeo de create_sucursal() no pesca por ser variaciones de tilde/prefijo."""
    import difflib

    def norm(v: str) -> str:
        v = unicodedata.normalize("NFD", (v or "").strip().lower()).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", v)).strip()

    candidatos = db.execute(
        text("""
            SELECT id, nombre_sucursal, direccion_sucursal
            FROM bbdd_sucursales
            WHERE LOWER(TRIM(rut)) = LOWER(TRIM(:rut))
              AND aceptada_bitacora = 1
              AND id <> :sid
        """),
        {"rut": rut, "sid": sucursal_id},
    ).mappings().all()

    objetivo = norm(nombre_sucursal)
    mejor = None
    for c in candidatos:
        ratio = difflib.SequenceMatcher(None, objetivo, norm(c.get("nombre_sucursal"))).ratio()
        if ratio >= 0.75 and (mejor is None or ratio > mejor["similitud"]):
            mejor = {
                "id": c.get("id"),
                "nombre_sucursal": c.get("nombre_sucursal"),
                "direccion_sucursal": c.get("direccion_sucursal"),
                "similitud": ratio,
            }
    if mejor:
        mejor["similitud"] = round(mejor["similitud"], 2)
    return mejor


@router.get("/api/bitacora/sucursales-pendientes")
def bitacora_sucursales_pendientes_api(
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)

    rows = incidencias_db.execute(
        text("""
            SELECT
                id,
                COALESCE(TRIM(rut), '') AS rut,
                COALESCE(TRIM(nombre_empresa), '') AS nombre_empresa,
                COALESCE(TRIM(nombre_sucursal), '') AS nombre_sucursal,
                COALESCE(TRIM(direccion_sucursal), '') AS direccion_sucursal,
                COALESCE(TRIM(comuna), '') AS comuna,
                COALESCE(TRIM(region), '') AS region,
                COALESCE(TRIM(created_by), '') AS created_by,
                created_at
            FROM bbdd_sucursales
            WHERE aceptada_bitacora = 0
            ORDER BY created_at DESC, id DESC
        """)
    ).mappings().all()

    pendientes = []
    for row in rows:
        duplicado = _mejor_posible_duplicado(incidencias_db, row.get("id"), row.get("rut"), row.get("nombre_sucursal"))
        pendientes.append({
            "id": row.get("id"),
            "rut": row.get("rut"),
            "nombre_empresa": row.get("nombre_empresa"),
            "nombre_sucursal": row.get("nombre_sucursal"),
            "direccion_sucursal": row.get("direccion_sucursal"),
            "comuna": row.get("comuna"),
            "region": row.get("region"),
            "created_by": row.get("created_by"),
            "created_at": row.get("created_at").isoformat(sep=" ", timespec="seconds") if isinstance(row.get("created_at"), datetime) else str(row.get("created_at") or ""),
            "posible_duplicado": duplicado,
        })
    return {"total": len(pendientes), "sucursales": pendientes}


@router.post("/api/bitacora/sucursales-pendientes/{sucursal_id}/aceptar")
def bitacora_sucursal_aceptar(
    sucursal_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)

    result = incidencias_db.execute(
        text("UPDATE bbdd_sucursales SET aceptada_bitacora = 1 WHERE id = :sid"),
        {"sid": sucursal_id},
    )
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")
    incidencias_db.commit()
    return {"ok": True}


class RechazarDuplicadoPayload(BaseModel):
    sucursal_aceptada_id: int


@router.post("/api/bitacora/sucursales-pendientes/{sucursal_id}/rechazar-duplicado")
def bitacora_sucursal_rechazar_duplicado(
    sucursal_id: int,
    payload: RechazarDuplicadoPayload,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Único caso en el que sí se descarta una sucursal pendiente: es un duplicado
    de una ya aceptada. Antes de borrarla, re-apunta sus ODS en venta_comercial
    (rut + direccion/nombre) a la identidad de la sucursal aceptada, para que las
    cámaras y demás datos que _ficha_sucursal calcula sumando esas ODS no se
    pierdan — quedan contabilizados en la sucursal que sí queda activa."""
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)

    pendiente = incidencias_db.execute(
        text("""
            SELECT id, COALESCE(TRIM(rut), '') AS rut,
                   COALESCE(TRIM(direccion_sucursal), '') AS direccion_sucursal,
                   COALESCE(TRIM(nombre_sucursal), '') AS nombre_sucursal
            FROM bbdd_sucursales
            WHERE id = :sid AND aceptada_bitacora = 0
        """),
        {"sid": sucursal_id},
    ).mappings().first()
    if not pendiente:
        raise HTTPException(status_code=404, detail="Sucursal pendiente no encontrada.")

    aceptada = incidencias_db.execute(
        text("""
            SELECT id, COALESCE(TRIM(direccion_sucursal), '') AS direccion_sucursal,
                   COALESCE(TRIM(nombre_sucursal), '') AS nombre_sucursal
            FROM bbdd_sucursales
            WHERE id = :sid AND aceptada_bitacora = 1
        """),
        {"sid": payload.sucursal_aceptada_id},
    ).mappings().first()
    if not aceptada:
        raise HTTPException(status_code=404, detail="La sucursal aceptada indicada no existe.")

    ods_reasignadas = incidencias_db.execute(
        text("""
            UPDATE venta_comercial
            SET direccion_sucursal = :direccion_nueva,
                nombre_sucursal = :nombre_nuevo
            WHERE LOWER(TRIM(rut_cliente)) = LOWER(TRIM(:rut))
              AND (
                LOWER(TRIM(direccion_sucursal)) = LOWER(TRIM(:direccion_vieja))
                OR LOWER(TRIM(nombre_sucursal)) = LOWER(TRIM(:nombre_viejo))
              )
        """),
        {
            "direccion_nueva": aceptada["direccion_sucursal"],
            "nombre_nuevo": aceptada["nombre_sucursal"],
            "rut": pendiente["rut"],
            "direccion_vieja": pendiente["direccion_sucursal"],
            "nombre_viejo": pendiente["nombre_sucursal"],
        },
    ).rowcount

    for tabla in (
        "sucursal_guardias",
        "pruebas_sonido",
        "sucursal_personas_autorizadas",
        "sucursal_contactos_emergencia",
        "sucursal_info_extra",
    ):
        incidencias_db.execute(text(f"DELETE FROM {tabla} WHERE sucursal_id = :sid"), {"sid": sucursal_id})

    incidencias_db.execute(text("DELETE FROM bbdd_sucursales WHERE id = :sid"), {"sid": sucursal_id})
    incidencias_db.commit()
    return {
        "ok": True,
        "fusionado_con": aceptada["nombre_sucursal"],
        "ods_reasignadas": ods_reasignadas,
    }


_BITACORA_EMAIL_LOGO_URL = "https://i.imgur.com/VgLG9Ei.png"

_BITACORA_CAMPO_LABELS = {
    "direccion_sucursal": "Dirección",
    "latitud_longitud": "Latitud, Longitud",
    "referencia_ubicacion": "Referencia ubicación",
    "contacto": "Contacto",
    "horario_apertura": "Horario de apertura",
    "horario_cierre": "Horario de cierre",
    "horario_habil": "Días hábiles",
    "telefono_porton": "Teléfono portón",
    "telefono_recepcion": "Teléfono recepción",
    "compania_electricidad": "Compañía electricidad",
    "numero_cliente_electricidad": "N° cliente electricidad",
    "proveedor_internet_cliente": "Proveedor internet cliente",
    "internet_atc": "Internet ATC",
    "contactos_emergencia": "Contacto de emergencia",
    "personas_autorizadas": "Personas autorizadas",
}


def _bitacora_esc(value: object) -> str:
    return html.escape(str(value or "").strip())


def _bitacora_email_html(*, title: str, sections: list[str]) -> str:
    section_html = "".join(
        f"<div style=\"margin-bottom:14px;\">{section}</div>"
        for section in sections
        if section
    )
    return f"""
<div style="background:#f5f6fa;padding:40px 0;font-family:'Segoe UI',Arial,sans-serif;color:#2d3436;">
  <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:8px;
              padding:30px;box-shadow:0 2px 10px rgba(0,0,0,0.05);">
    <div style="text-align:center;margin-bottom:18px;">
      <img src="{_BITACORA_EMAIL_LOGO_URL}" alt="Alguien Te Cuida" style="height:55px;">
    </div>
    <h2 style="text-align:center;color:#2d3436;font-size:18px;margin:0 0 22px 0;">{_bitacora_esc(title)}</h2>
    <div style="font-size:14px;line-height:1.7;">{section_html}</div>
    <p style="margin-top:26px;font-size:14px;line-height:1.6;">
      Saludos cordiales,<br>Alguien Te Cuida
    </p>
    <hr style="border:0;border-top:1px solid #ddd;margin:26px 0;">
    <p style="font-size:12px;color:#999;text-align:center;line-height:1.5;">
      Este correo ha sido generado automáticamente como parte del proceso interno.
    </p>
  </div>
</div>"""


def _bitacora_paragraph_html(text: str) -> str:
    text_value = str(text or "").strip()
    if not text_value:
        return ""
    return (
        '<p style="margin:0 0 16px 0;color:#334155;font-family:Arial,sans-serif;'
        'font-size:14px;line-height:1.72;">'
        f"{_bitacora_esc(text_value).replace(chr(10), '<br>')}"
        "</p>"
    )


def _bitacora_missing_list_html(items: list[str]) -> str:
    if not items:
        return ""
    bullets = "".join(
        '<li style="margin:0 0 7px 0;padding-left:2px;">'
        f"{_bitacora_esc(item)}"
        "</li>"
        for item in items
        if str(item or "").strip()
    )
    return (
        '<ul style="margin:4px 0 18px 22px;padding:0;color:#111827;'
        'font-family:Arial,sans-serif;font-size:14px;line-height:1.65;'
        'list-style-type:disc;">'
        f"{bullets}"
        "</ul>"
    )


def _bitacora_notificar_items(mensaje: str, campos: list[str]) -> list[str]:
    items: list[str] = []
    collecting = False
    for raw_line in str(mensaje or "").splitlines():
        line = raw_line.strip()
        if not line:
            if collecting and items:
                break
            continue
        line_norm = unicodedata.normalize("NFKD", line).encode("ascii", "ignore").decode("ascii").lower()
        if "falta o esta mal" in line_norm:
            collecting = True
            continue
        if collecting:
            item = re.sub(r"^[-•*]\s*", "", line).strip()
            if item:
                items.append(item)
    if items:
        return items
    return [
        _BITACORA_CAMPO_LABELS.get(campo, campo)
        for campo in campos
        if str(campo or "").strip()
    ]


def _enviar_correo_bitacora(to_email: str, subject: str, body_text: str, html_body: str | None = None) -> None:
    """Envio multipart reusando la cuenta contacto@alguientecuida.cl
    ya provisionada (mismo patron que incidencias_service.py/protocolos_service.py
    en esta misma sesion)."""
    import smtplib
    from email.message import EmailMessage
    from ATC.app.routes.inicio_turno import _contacto_smtp_config

    cfg = _contacto_smtp_config()
    if not cfg.get("enabled"):
        raise ValueError(f"SMTP de contacto no disponible: {cfg.get('reason') or 'no configurado'}")

    msg = EmailMessage()
    msg["Subject"] = subject
    from_name = str(cfg.get("from_name") or "Alguien Te Cuida")
    from_email = str(cfg.get("from_email") or cfg.get("username") or "")
    msg["From"] = f"{from_name} <{from_email}>" if from_name else from_email
    msg["To"] = to_email
    msg.set_content(body_text, subtype="plain", charset="utf-8")
    if html_body:
        msg.add_alternative(html_body, subtype="html", charset="utf-8")

    host = str(cfg["host"])
    port = int(cfg["port"])
    username = str(cfg["username"])
    password = str(cfg["password"])
    timeout = int(cfg["timeout"])

    if cfg.get("use_ssl"):
        with smtplib.SMTP_SSL(host, port, timeout=timeout) as smtp:
            smtp.login(username, password)
            smtp.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=timeout) as smtp:
            smtp.ehlo()
            if cfg.get("use_tls"):
                smtp.starttls()
                smtp.ehlo()
            smtp.login(username, password)
            smtp.send_message(msg)


class NotificarComercialPayload(BaseModel):
    mensaje: str
    campos: list[str] = []
    detalle: str = ""


@router.post("/api/bitacora/sucursales-pendientes/{sucursal_id}/notificar-comercial")
def bitacora_sucursal_notificar_comercial(
    sucursal_id: int,
    payload: NotificarComercialPayload,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Avisa por correo a quien registro la sucursal (created_by, resuelto contra
    users.name) que falta o esta mal algo, sin tocar el estado de la sucursal —
    sigue pendiente y editable hasta que quede bien. Reemplaza al viejo Rechazar,
    que eliminaba la fila. payload.campos (claves de campo, no las etiquetas) queda
    guardado en sucursal_info_extra para que Venta (BBDD Sucursales / Información
    Clientes) sepa qué resaltar dinámicamente y arme el resumen de "lo que rellenó"
    cuando avisen que ya quedó listo."""
    _require_bitacora_access(current_user)
    _ensure_sucursal_aceptada_bitacora_column(incidencias_db)
    _ensure_sucursal_info_extra(incidencias_db)
    _ensure_sucursal_info_extra_campos_pendientes(incidencias_db)

    mensaje = payload.mensaje.strip()
    if not mensaje:
        raise HTTPException(status_code=400, detail="Debes indicar qué falta o está mal.")

    row = incidencias_db.execute(
        text("""
            SELECT nombre_sucursal, nombre_empresa, created_by
            FROM bbdd_sucursales WHERE id = :sid AND aceptada_bitacora = 0
        """),
        {"sid": sucursal_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Sucursal pendiente no encontrada.")

    creador_nombre = str(row.get("created_by") or "").strip()
    creador = None
    if creador_nombre:
        creador = incidencias_db.query(User).filter(
            func.lower(func.trim(User.name)) == creador_nombre.lower()
        ).first()
    if not creador or not str(creador.email or "").strip():
        raise HTTPException(
            status_code=400,
            detail=f"No se encontró el correo de quien registró esta sucursal ({creador_nombre or 'desconocido'}).",
        )

    destino = str(creador.email).strip()
    sucursal_txt = str(row.get("nombre_sucursal") or "-")
    empresa_txt = str(row.get("nombre_empresa") or "-")
    revisor_txt = str(current_user.name or current_user.username or "Bitácora").strip()
    creador_txt = str(creador.name or creador_nombre or "Comercial").strip()
    campos_limpios = [c.strip() for c in payload.campos if c.strip()]
    faltantes = _bitacora_notificar_items(mensaje, campos_limpios)
    detalle_extra = payload.detalle.strip()
    detalle_incluido_en_lista = False
    if not faltantes and detalle_extra:
        faltantes = [detalle_extra]
        detalle_incluido_en_lista = True
    asunto = f"Falta completar: {sucursal_txt} — {empresa_txt}"
    faltantes_plain = "\n".join(f"• {item}" for item in faltantes) or "• Información indicada por Bitácora"
    detalle_plain = f"\n\nDetalle adicional:\n{detalle_extra}" if detalle_extra and not detalle_incluido_en_lista else ""
    cuerpo = (
        f"Estimado/a {creador_txt},\n\n"
        f"Al revisar la sucursal {sucursal_txt} en Bitácora, {revisor_txt} indica que falta o está mal:\n\n"
        f"{faltantes_plain}"
        f"{detalle_plain}\n\n"
        "Es importante regularizar esta información antes de que la instalación sea ejecutada y puesta en marcha, "
        "ya que la falta de estos datos puede afectar directamente la correcta prestación del servicio.\n\n"
        f"Alguien Te Cuida"
    )
    html_cuerpo = _bitacora_email_html(
        title="Falta completar información de sucursal",
        sections=[
            _bitacora_paragraph_html(f"Estimado/a {creador_txt},"),
            (
                '<p style="margin:0 0 16px 0;color:#334155;font-family:Arial,sans-serif;'
                'font-size:14px;line-height:1.72;">'
                f"Al revisar la sucursal <strong>{_bitacora_esc(sucursal_txt)}</strong> en Bitácora, "
                f"{_bitacora_esc(revisor_txt)} indica que falta o está mal:"
                "</p>"
            ),
            _bitacora_missing_list_html(faltantes),
            _bitacora_paragraph_html("" if detalle_incluido_en_lista else detalle_extra),
            _bitacora_paragraph_html(
                "Es importante regularizar esta información antes de que la instalación sea ejecutada y puesta en marcha, "
                "ya que la falta de estos datos puede afectar directamente la correcta prestación del servicio."
            ),
        ],
    )
    try:
        _enviar_correo_bitacora(destino, asunto, cuerpo, html_cuerpo)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No se pudo enviar el correo: {exc}") from exc

    incidencias_db.execute(
        text("""
            MERGE sucursal_info_extra AS t
            USING (VALUES (:sid)) AS s(sucursal_id) ON t.sucursal_id = s.sucursal_id
            WHEN MATCHED THEN UPDATE SET
                campos_pendientes = :campos,
                campos_pendientes_obs = :obs,
                campos_pendientes_fecha = GETDATE(),
                campos_pendientes_por = :por
            WHEN NOT MATCHED THEN INSERT (sucursal_id, campos_pendientes, campos_pendientes_obs, campos_pendientes_fecha, campos_pendientes_por)
            VALUES (:sid, :campos, :obs, GETDATE(), :por);
        """),
        {
            "sid": sucursal_id,
            "campos": ",".join(campos_limpios),
            # Solo el detalle libre — el listado de campos ya queda estructurado en
            # "campos"; guardar acá también el "Falta o está mal: ..." generado
            # automáticamente duplicaba lo mismo dos veces en el banner de Venta.
            "obs": payload.detalle.strip(),
            "por": current_user.name or current_user.username,
        },
    )
    incidencias_db.commit()

    return {"ok": True, "enviado_a": destino}


@router.get("/api/bitacora/sucursales-pendientes/{sucursal_id}/preview")
def bitacora_sucursal_pendiente_preview(
    sucursal_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Arma la misma ficha que se ve en 'Buscar por Empresa' (identificación,
    horarios, contactos de emergencia, personas autorizadas, cámaras si ya hay una
    ODS ligada) pero para UNA sucursal puntual buscada por id — funciona esté
    aceptada o pendiente, para poder previsualizar cómo quedaría antes de aceptar."""
    _require_bitacora_access(current_user)

    selected_row = incidencias_db.execute(
        text(
            """
            SELECT
                id,
                COALESCE(TRIM(rut), '') AS rut,
                COALESCE(TRIM(nombre_empresa), '') AS nombre_empresa,
                COALESCE(TRIM(nombre_sucursal), '') AS nombre_sucursal,
                COALESCE(TRIM(direccion_sucursal), '') AS direccion_sucursal,
                COALESCE(TRIM(region), '') AS region,
                COALESCE(TRIM(comuna), '') AS comuna,
                COALESCE(TRIM(referencia_ubicacion), '') AS referencia_ubicacion,
                COALESCE(TRIM(latitud_longitud), '') AS latitud_longitud,
                COALESCE(TRIM(email_facturas), '') AS email_facturas,
                COALESCE(TRIM(proveedor_internet), '') AS proveedor_internet,
                COALESCE(TRIM(proveedor_electricidad), '') AS proveedor_electricidad,
                COALESCE(TRIM(nro_proveedor_electricidad), '') AS nro_proveedor_electricidad,
                COALESCE(TRIM(horario_apertura), '') AS horario_apertura,
                COALESCE(TRIM(horario_cierre), '') AS horario_cierre,
                COALESCE(TRIM(dias_funcionamiento), '') AS dias_funcionamiento,
                created_at
            FROM bbdd_sucursales
            WHERE id = :sid
            """
        ),
        {"sid": sucursal_id},
    ).mappings().first()
    if not selected_row:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")

    ficha = _ficha_sucursal(incidencias_db, selected_row, selected_row.get("nombre_empresa"))
    return {
        "empresa": selected_row.get("nombre_empresa"),
        "sucursal_id": selected_row.get("id"),
        "sucursal_actual": selected_row.get("nombre_sucursal"),
        **ficha,
    }


# ──────────────────────────────────────────────
# Personas autorizadas por sucursal
# ──────────────────────────────────────────────

class PersonaAutorizadaCreate(BaseModel):
    sucursal_id: int
    nombre: str = ""
    rut: str = ""
    telefono: str = ""
    email: str = ""
    clave_verde: str = ""
    clave_roja: str = ""


class PersonaAutorizadaUpdate(BaseModel):
    nombre: str = ""
    rut: str = ""
    telefono: str = ""
    email: str = ""
    clave_verde: str = ""
    clave_roja: str = ""


def _ensure_personas_autorizadas_habilitado(db: Session) -> None:
    """Agrega columna habilitado a sucursal_personas_autorizadas si no existe."""
    try:
        db.execute(text("""
            IF NOT EXISTS (
                SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = 'sucursal_personas_autorizadas'
                  AND COLUMN_NAME = 'habilitado'
            )
            BEGIN
                ALTER TABLE sucursal_personas_autorizadas
                ADD habilitado BIT NOT NULL DEFAULT 1
            END
        """))
        db.commit()
    except Exception:
        db.rollback()


@router.get("/api/bitacora/sucursal/{sucursal_id}/personas-autorizadas")
def api_list_personas_autorizadas(
    sucursal_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_personas_autorizadas_habilitado(incidencias_db)
    rows = (
        incidencias_db.query(SucursalPersonaAutorizada)
        .filter(SucursalPersonaAutorizada.sucursal_id == sucursal_id)
        .order_by(SucursalPersonaAutorizada.nombre)
        .all()
    )
    return [
        {
            "id": r.id,
            "nombre": r.nombre or "",
            "rut": r.rut or "",
            "telefono": r.telefono or "",
            "email": r.email or "",
            "clave_verde": r.clave_verde or "",
            "clave_roja": r.clave_roja or "",
            "habilitado": bool(r.habilitado),
        }
        for r in rows
    ]


@router.post("/api/bitacora/personas-autorizadas")
def api_create_persona_autorizada(
    payload: PersonaAutorizadaCreate,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    _ensure_personas_autorizadas_habilitado(incidencias_db)
    sucursal = incidencias_db.get(SucursalBBDD, payload.sucursal_id)
    if not sucursal:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")
    nueva = SucursalPersonaAutorizada(
        sucursal_id=payload.sucursal_id,
        nombre=payload.nombre.strip() or None,
        rut=payload.rut.strip() or None,
        telefono=payload.telefono.strip() or None,
        email=payload.email.strip() or None,
        clave_verde=payload.clave_verde.strip() or None,
        clave_roja=payload.clave_roja.strip() or None,
        habilitado=True,
    )
    incidencias_db.add(nueva)
    incidencias_db.commit()
    incidencias_db.refresh(nueva)
    return {"ok": True, "id": nueva.id}


@router.put("/api/bitacora/personas-autorizadas/{persona_id}")
def api_update_persona_autorizada(
    persona_id: int,
    payload: PersonaAutorizadaUpdate,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.get(SucursalPersonaAutorizada, persona_id)
    if not row:
        raise HTTPException(status_code=404, detail="Persona no encontrada.")
    row.nombre = payload.nombre.strip() or None
    row.rut = payload.rut.strip() or None
    row.telefono = payload.telefono.strip() or None
    row.email = payload.email.strip() or None
    row.clave_verde = payload.clave_verde.strip() or None
    row.clave_roja = payload.clave_roja.strip() or None
    incidencias_db.commit()
    return {"ok": True}


@router.patch("/api/bitacora/personas-autorizadas/{persona_id}/toggle")
def api_toggle_persona_autorizada(
    persona_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.get(SucursalPersonaAutorizada, persona_id)
    if not row:
        raise HTTPException(status_code=404, detail="Persona no encontrada.")
    row.habilitado = not bool(row.habilitado)
    incidencias_db.commit()
    return {"ok": True, "habilitado": bool(row.habilitado)}


@router.delete("/api/bitacora/personas-autorizadas/{persona_id}")
def api_delete_persona_autorizada(
    persona_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.get(SucursalPersonaAutorizada, persona_id)
    if not row:
        raise HTTPException(status_code=404, detail="Persona no encontrada.")
    incidencias_db.delete(row)
    incidencias_db.commit()
    return {"ok": True}


# ──────────────────────────────────────────────
# Contactos de emergencia por sucursal
# ──────────────────────────────────────────────

class ContactoEmergenciaCreate(BaseModel):
    sucursal_id: int
    nombre: str = ""
    rut: str = ""
    telefono: str = ""
    email: str = ""
    orden: int | None = None


class ContactoEmergenciaUpdate(BaseModel):
    nombre: str = ""
    rut: str = ""
    telefono: str = ""
    email: str = ""
    orden: int | None = None


@router.get("/api/bitacora/sucursal/{sucursal_id}/contactos-emergencia")
def api_list_contactos_emergencia(
    sucursal_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    rows = (
        incidencias_db.query(SucursalContactoEmergencia)
        .filter(SucursalContactoEmergencia.sucursal_id == sucursal_id)
        .order_by(SucursalContactoEmergencia.orden, SucursalContactoEmergencia.id)
        .all()
    )
    return [
        {
            "id": r.id,
            "nombre": r.nombre or "",
            "rut": r.rut or "",
            "telefono": r.telefono or "",
            "email": r.email or "",
            "orden": r.orden,
        }
        for r in rows
    ]


@router.post("/api/bitacora/contactos-emergencia")
def api_create_contacto_emergencia(
    payload: ContactoEmergenciaCreate,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    sucursal = incidencias_db.get(SucursalBBDD, payload.sucursal_id)
    if not sucursal:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada.")
    if payload.orden is not None:
        orden_final = payload.orden
    else:
        orden_final = (
            incidencias_db.query(SucursalContactoEmergencia)
            .filter(SucursalContactoEmergencia.sucursal_id == payload.sucursal_id)
            .count()
        ) + 1
    nuevo = SucursalContactoEmergencia(
        sucursal_id=payload.sucursal_id,
        nombre=payload.nombre.strip() or None,
        rut=payload.rut.strip() or None,
        telefono=payload.telefono.strip() or None,
        email=payload.email.strip() or None,
        orden=orden_final,
    )
    incidencias_db.add(nuevo)
    incidencias_db.commit()
    incidencias_db.refresh(nuevo)
    return {"ok": True, "id": nuevo.id}


@router.put("/api/bitacora/contactos-emergencia/{contacto_id}")
def api_update_contacto_emergencia(
    contacto_id: int,
    payload: ContactoEmergenciaUpdate,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.get(SucursalContactoEmergencia, contacto_id)
    if not row:
        raise HTTPException(status_code=404, detail="Contacto no encontrado.")
    row.nombre = payload.nombre.strip() or None
    row.rut = payload.rut.strip() or None
    row.telefono = payload.telefono.strip() or None
    row.email = payload.email.strip() or None
    if payload.orden is not None:
        row.orden = payload.orden
    incidencias_db.commit()
    return {"ok": True}


@router.delete("/api/bitacora/contactos-emergencia/{contacto_id}")
def api_delete_contacto_emergencia(
    contacto_id: int,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    row = incidencias_db.get(SucursalContactoEmergencia, contacto_id)
    if not row:
        raise HTTPException(status_code=404, detail="Contacto no encontrado.")
    incidencias_db.delete(row)
    incidencias_db.commit()
    return {"ok": True}


class _CreateUserBody(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    username: str = Field(..., min_length=1, max_length=50)
    email: str | None = Field(default=None, max_length=255)
    role: str = Field(..., pattern=r"^(admin|agent)$")


# role -> departamento requerido para tener acceso a Bitacora desde este modal
# (Administrador queda con acceso a Bitacora; Operador con Televigilante).
_ROLE_DEPARTMENT = {"admin": "Bitacora", "agent": "Televigilante"}


@router.post("/api/bitacora/users")
def api_crear_bitacora_user(
    body: _CreateUserBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden registrar usuarios.")

    # La contraseña ya no se pide en el popup: siempre son los primeros 5
    # dígitos del RUT, para agilizar el registro masivo (pedido explícito,
    # ago 2026).
    digitos_rut = "".join(ch for ch in body.username if ch.isdigit())
    password_generada = digitos_rut[:5]

    departamento_requerido = _ROLE_DEPARTMENT[body.role]
    existing = db.query(User).filter(User.username == body.username).first()
    if existing:
        # La lista de "Gestión de Usuarios" de Bitácora solo muestra usuarios
        # con acceso a Bitácora (can_access_bitacora) — un RUT ya registrado
        # en otra área (Guardia, RRHH, etc.) no aparece ahí. En vez de
        # bloquear con un 409 confuso, si a esa cuenta le falta el
        # departamento de Bitácora/Televigilante se lo agregamos ahí mismo.
        departamentos_actuales = _split_departments(getattr(existing, "department", None))
        departamentos_norm = [_normalize(d) for d in departamentos_actuales]
        if _normalize(departamento_requerido) in departamentos_norm:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Ese RUT ya está registrado a nombre de {existing.name} "
                    f"y ya tiene acceso a {departamento_requerido}."
                ),
            )
        departamentos_actuales.append(departamento_requerido)
        existing.department = "; ".join(departamentos_actuales)
        db.commit()
        return {
            "ok": True,
            "id": existing.id,
            "merged": True,
            "detail": f"{existing.name} ya existía — se le agregó acceso a {departamento_requerido}.",
        }

    user = User(
        name=body.name,
        username=body.username,
        hashed_password=hash_password(password_generada),
        email=body.email or None,
        role=body.role,
        department=departamento_requerido,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"ok": True, "id": user.id, "detail": "Usuario registrado correctamente."}


class _EditUserBody(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    username: str = Field(..., min_length=1, max_length=50)
    password: str | None = Field(default=None)
    email: str | None = Field(default=None, max_length=255)
    role: str = Field(..., pattern=r"^(admin|agent)$")
    is_active: bool


@router.put("/api/bitacora/users/{user_id}")
def api_edit_bitacora_user(
    user_id: int,
    body: _EditUserBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    _require_bitacora_access(current_user)
    if not is_bitacora_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo administradores pueden editar usuarios.")
    user = db.get(User, user_id)
    if not user or not can_access_bitacora(user):
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")
    if str(getattr(user, "role", "") or "").strip().lower() == "superadmin" and not getattr(current_user, "is_super_admin", False):
        raise HTTPException(status_code=403, detail="No tenés permiso para editar este usuario.")
    # Check username uniqueness if changed
    if body.username != user.username:
        conflict = db.query(User).filter(User.username == body.username, User.id != user_id).first()
        if conflict:
            raise HTTPException(status_code=409, detail="El nombre de usuario ya está en uso.")
    user.name = body.name
    user.username = body.username
    user.email = body.email or None
    user.role = body.role
    user.is_active = body.is_active
    if body.password:
        user.hashed_password = hash_password(body.password)
    db.commit()
    return {"ok": True, "id": user.id}


# ──────────────────────────────────────────────
# Administrar puestos — mapa real de puestos/pantallas/cámaras
# (desde sucursal_camaras_monitoreo, cargado ago 2026 desde la planilla
# "Cruce de Información Cámaras"). Puestos 1-12: 4 pantallas (2x2).
# Puestos 13-29: 6 pantallas (3x2, agrega columna Centro).
# ──────────────────────────────────────────────

_PANTALLAS_4 = ["Izquierda Arriba", "Derecha Arriba", "Izquierda Abajo", "Derecha Abajo"]
_PANTALLAS_6 = [
    "Izquierda Arriba", "Centro Arriba", "Derecha Arriba",
    "Izquierda Abajo", "Centro Abajo", "Derecha Abajo",
]
_SIN_PANTALLA = "Sin pantalla asignada"

# Variantes/typos vistos en la planilla original -> etiqueta canónica.
_UBICACION_PANTALLA_ALIASES = {
    "IZQUIERDA ARRIBA": "Izquierda Arriba",
    "IZQ ARRIBA": "Izquierda Arriba",
    "IZQUIERDA ABAJO": "Izquierda Abajo",
    "IZQ ABAJO": "Izquierda Abajo",
    "DERECHA ARRIBA": "Derecha Arriba",
    "DERCEHA ARRIBA": "Derecha Arriba",
    "DERECHA ABAJO": "Derecha Abajo",
    "CENTRO ARRIBA": "Centro Arriba",
    "CENTRO ABAJO": "Centro Abajo",
}


def _pantallas_de_puesto(central: int) -> list[str]:
    return _PANTALLAS_4 if central <= 12 else _PANTALLAS_6


def _normalizar_ubicacion_pantalla(raw: object) -> str | None:
    s = " ".join(str(raw or "").strip().upper().split())
    return _UBICACION_PANTALLA_ALIASES.get(s)


@router.get("/api/bitacora/puestos")
def api_bitacora_puestos(
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Arma el contenido real de 'Administrar puestos' (29 puestos, 4 o 6
    pantallas cada uno) a partir de sucursal_camaras_monitoreo, cruzando
    `central` (puesto) y `ubicacion_pantalla` (pantalla). Solo las cámaras
    monitoreadas (nombre_camara_monitoreo) van en la grilla de pantallas —
    las complementarias sin monitoreo (camara_sin_monitoreo) no se asocian
    a una pantalla física; se devuelven aparte, agrupadas por sucursal, para
    el puesto especial "Cámaras sin monitoreo" del front."""
    _require_bitacora_access(current_user)

    rows = (
        incidencias_db.query(
            SucursalCamaraMonitoreo.id,
            SucursalCamaraMonitoreo.central,
            SucursalCamaraMonitoreo.ubicacion_pantalla,
            SucursalCamaraMonitoreo.nombre_camara_monitoreo,
            SucursalCamaraMonitoreo.camara_sin_monitoreo,
            SucursalCamaraMonitoreo.cantidad_equipos,
            SucursalCamaraMonitoreo.sucursal_id,
            SucursalCamaraMonitoreo.slot_index,
            SucursalBBDD.nombre_sucursal,
        )
        .outerjoin(SucursalBBDD, SucursalCamaraMonitoreo.sucursal_id == SucursalBBDD.id)
        .filter(SucursalCamaraMonitoreo.central.isnot(None))
        .order_by(
            SucursalCamaraMonitoreo.central,
            case((SucursalCamaraMonitoreo.slot_index.is_(None), 1), else_=0),
            SucursalCamaraMonitoreo.slot_index,
            SucursalCamaraMonitoreo.id,
        )
        .all()
    )

    incidencias_abiertas = _cargar_incidencias_abiertas_camaras(incidencias_db)

    # grouped[central][etiqueta_pantalla] = [ {id, empresa, camara, slotIndex}, ... ] — solo monitoreadas
    grouped: dict[int, dict[str, list[dict]]] = {}
    camaras_sin_monitoreo: list[dict] = []
    for fila_id, central, ubicacion_pantalla, cam_monitoreada, cam_sin_monitoreo, cantidad_equipos, sucursal_id, slot_index, nombre_sucursal in rows:
        empresa = (nombre_sucursal or "").strip() or "(sucursal sin nombre)"
        try:
            cantidad_equipos_int = max(1, int(cantidad_equipos or 1))
        except (TypeError, ValueError):
            cantidad_equipos_int = 1
        if cam_monitoreada and cam_monitoreada.strip():
            etiqueta = _normalizar_ubicacion_pantalla(ubicacion_pantalla)
            if etiqueta not in _pantallas_de_puesto(central):
                etiqueta = _SIN_PANTALLA
            bucket = grouped.setdefault(int(central), {}).setdefault(etiqueta, [])
            incidencias_camara = _coincidencias_incidencias_camara(
                incidencias_abiertas,
                empresa,
                cam_monitoreada,
                cantidad_equipos=cantidad_equipos_int,
            )
            item = {
                "id": fila_id,
                "empresa": empresa,
                "camara": cam_monitoreada.strip(),
                "cantidad_equipos": cantidad_equipos_int,
                "estado": "problema" if incidencias_camara else "en_linea",
                "sucursalId": sucursal_id,
                "slotIndex": slot_index,
            }
            if incidencias_camara:
                item["incidencias_count"] = len(incidencias_camara)
                item["incidencias"] = incidencias_camara[:4]
            bucket.append(item)
        if cam_sin_monitoreo and cam_sin_monitoreo.strip():
            camaras_sin_monitoreo.append({
                "empresa": empresa,
                "camara": cam_sin_monitoreo.strip(),
                "cantidad_equipos": cantidad_equipos_int,
            })

    def _resolver_slots(items: list[dict], capacidad: int) -> list[dict]:
        # slot_index puede venir nulo/duplicado/fuera de rango en filas viejas
        # o recien importadas — se respeta el valor guardado cuando es valido
        # y se rellenan huecos con los que no tienen uno, en el mismo orden
        # en que llegaron (ya vienen ordenados por slot_index/id).
        ocupados: dict[int, dict] = {}
        sin_slot: list[dict] = []
        for it in items:
            si = it.get("slotIndex")
            if isinstance(si, int) and 0 <= si < capacidad and si not in ocupados:
                ocupados[si] = it
            else:
                sin_slot.append(it)
        libres = [i for i in range(capacidad) if i not in ocupados]
        for it, slot in zip(sin_slot, libres):
            it["slotIndex"] = slot
            ocupados[slot] = it
        # si sobran items sin cupo (mas filas que casillas), se agregan al
        # final igual — el front las muestra fuera de la grilla en vez de
        # perderlas silenciosamente.
        sobrantes = sin_slot[len(libres):]
        return sorted(ocupados.values(), key=lambda x: x["slotIndex"]) + sobrantes

    puestos = {}
    for central in range(1, 30):
        pantallas_puesto = _pantallas_de_puesto(central)
        capacidad = 25 if central <= 12 else 20
        por_pantalla = grouped.get(central, {})
        pantallas = [
            {"nombre": nombre, "capacidad": capacidad, "items": _resolver_slots(por_pantalla.get(nombre, []), capacidad)}
            for nombre in pantallas_puesto
        ]
        sin_pantalla = por_pantalla.get(_SIN_PANTALLA, [])
        if sin_pantalla:
            pantallas.append({"nombre": _SIN_PANTALLA, "capacidad": len(sin_pantalla), "items": sin_pantalla})
        puestos[str(central)] = pantallas

    camaras_sin_monitoreo.sort(key=lambda c: (c["empresa"].lower(), c["camara"].lower()))

    return {
        "puestos": puestos,
        "camaras_sin_monitoreo": camaras_sin_monitoreo,
        "can_manage": can_manage_bitacora_puestos(current_user),
    }


class ColocacionItem(BaseModel):
    camara_id: int
    slot_index: int


class ColocarCamarasPayload(BaseModel):
    puesto_destino: int
    pantalla_destino: str
    colocaciones: list[ColocacionItem]


@router.post("/api/bitacora/puestos/colocar")
def colocar_camaras_puesto(
    payload: ColocarCamarasPayload,
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Guarda la posición exacta (casillero 0..capacidad-1) de una o varias
    cámaras dentro de una pantalla — reemplaza el viejo 'reposicionar' (que
    solo elegía la pantalla destino, sin casillero). El front arma el estado
    final deseado para ESA pantalla completa (arrastrando entre la pila de
    seleccionadas y la grilla) y lo manda de una vez; acá solo se valida y
    se escribe. Todo o nada: si algo no valida, no se mueve ninguna."""
    _require_bitacora_access(current_user)
    if not can_manage_bitacora_puestos(current_user):
        raise HTTPException(status_code=403, detail="Solo usuarios de Soporte pueden reposicionar cámaras.")

    if payload.puesto_destino < 1 or payload.puesto_destino > 29:
        raise HTTPException(status_code=400, detail="Puesto destino inválido.")

    pantallas_validas = _pantallas_de_puesto(payload.puesto_destino)
    if payload.pantalla_destino not in pantallas_validas:
        raise HTTPException(status_code=400, detail="Pantalla destino inválida para ese puesto.")

    colocaciones = payload.colocaciones
    if not colocaciones:
        raise HTTPException(status_code=400, detail="No se indicó ninguna cámara para colocar.")
    if len(colocaciones) > 200:
        raise HTTPException(status_code=400, detail="Demasiadas cámaras en un solo lote.")

    capacidad = 25 if payload.puesto_destino <= 12 else 20
    slots_usados: dict[int, int] = {}
    for c in colocaciones:
        if c.slot_index < 0 or c.slot_index >= capacidad:
            raise HTTPException(status_code=400, detail=f"Casillero {c.slot_index} inválido para esa pantalla.")
        if c.slot_index in slots_usados:
            raise HTTPException(status_code=400, detail="Dos cámaras no pueden quedar en el mismo casillero.")
        slots_usados[c.slot_index] = c.camara_id

    camara_ids = [c.camara_id for c in colocaciones]
    if len(set(camara_ids)) != len(camara_ids):
        raise HTTPException(status_code=400, detail="Una cámara no puede colocarse dos veces en el mismo lote.")

    filas = (
        incidencias_db.query(SucursalCamaraMonitoreo)
        .filter(SucursalCamaraMonitoreo.id.in_(camara_ids))
        .all()
    )
    filas_por_id = {f.id: f for f in filas}
    faltantes = [cid for cid in camara_ids if cid not in filas_por_id]
    if faltantes:
        raise HTTPException(status_code=404, detail=f"No se encontraron {len(faltantes)} de las cámaras seleccionadas.")
    no_monitoreadas = [f.id for f in filas if not f.nombre_camara_monitoreo or not f.nombre_camara_monitoreo.strip()]
    if no_monitoreadas:
        raise HTTPException(status_code=400, detail="Alguna de las filas seleccionadas no es una cámara monitoreada en pantalla.")

    # Cámaras que hoy están en ESTA misma pantalla pero no vienen en el lote
    # quedan desplazadas por el arrastre (drag-and-drop lo resuelve todo del
    # lado del cliente antes de guardar) — se limpian de casilleros que ya
    # no les corresponden para no dejar dos filas "creyendo" tener el mismo
    # casillero. Se compara con _normalizar_ubicacion_pantalla (no como
    # texto exacto) porque ubicacion_pantalla en la BBDD trae variantes
    # ("IZQ ARRIBA" vs "Izquierda Arriba") que igual mapean a la misma
    # pantalla canónica.
    ocupantes_mismo_puesto = (
        incidencias_db.query(SucursalCamaraMonitoreo)
        .filter(
            SucursalCamaraMonitoreo.central == payload.puesto_destino,
            SucursalCamaraMonitoreo.id.notin_(camara_ids),
        )
        .all()
    )
    for ocupante in ocupantes_mismo_puesto:
        if _normalizar_ubicacion_pantalla(ocupante.ubicacion_pantalla) != payload.pantalla_destino:
            continue
        if ocupante.slot_index in slots_usados:
            ocupante.slot_index = None

    movidas = 0
    for c in colocaciones:
        fila = filas_por_id[c.camara_id]
        fila.central = payload.puesto_destino
        fila.ubicacion_pantalla = payload.pantalla_destino
        fila.slot_index = c.slot_index
        movidas += 1
    incidencias_db.commit()

    return {"ok": True, "puesto": payload.puesto_destino, "pantalla": payload.pantalla_destino, "movidas": movidas}


# ──────────────────────────────────────────────
# Administrar puestos — estado de una cámara por incidencias reales
# ──────────────────────────────────────────────

_ESTADOS_INCIDENCIA_ABIERTA_EXCLUIDOS = ("terminado", "repetida", "rechazad")


def _normalizar_texto_cam(valor: object) -> str:
    s = unicodedata.normalize("NFD", str(valor or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.strip().lower().split())


_PREFIJOS_DETALLE_CAMARA = (
    "desconocida de:",
    "debido a internet de:",
    "debido a electricidad de:",
    "camara movida:",
    "camara caida:",
    "intermitencia:",
    "sin senal:",
)


def _incidencia_camara_abierta(row: dict) -> bool:
    estado_norm = _normalizar_texto_cam(row.get("estado"))
    return not any(e in estado_norm for e in _ESTADOS_INCIDENCIA_ABIERTA_EXCLUIDOS)


def _detalle_incidencia_camara(row: dict) -> str:
    detalle = str(row.get("detalle_problema") or row.get("observacion") or "").strip()
    detalle = re.sub(r"^\[[^\]]+\]\s*", "", detalle)
    detalle = re.split(r"\bContacto\s*:", detalle, maxsplit=1, flags=re.IGNORECASE)[0]
    return detalle.strip()


def _tokens_detalle_camara(row: dict) -> list[str]:
    detalle = _normalizar_texto_cam(_detalle_incidencia_camara(row))
    if not detalle:
        return []
    for prefijo in _PREFIJOS_DETALLE_CAMARA:
        if detalle.startswith(prefijo):
            detalle = detalle[len(prefijo):].strip()
            break
    return [token.strip() for token in detalle.split(",") if token.strip()]


def _texto_contiene_camara(texto_norm: str, camara_norm: str) -> bool:
    if not texto_norm or not camara_norm:
        return False
    if texto_norm == camara_norm:
        return True
    return re.search(rf"(^|[^\w]){re.escape(camara_norm)}($|[^\w])", texto_norm) is not None


def _resumen_incidencia_camara(row: dict, coincidencia: str) -> dict:
    return {
        "estado": "problema",
        "odt": row.get("odt"),
        "tipo": row.get("problema"),
        "fecha_registro": _fmt_fecha(row.get("fecha_registro")),
        "detalle": _detalle_incidencia_camara(row),
        "coincidencia": coincidencia,
    }


def _token_equipo_unico_cubre_todo(token: str, cantidad_equipos: int | None) -> bool:
    try:
        equipos = int(cantidad_equipos or 0)
    except (TypeError, ValueError):
        equipos = 0
    if equipos != 1:
        return False
    return re.fullmatch(r"(nvr|equipo)\s*0*1", _normalizar_texto_cam(token or "")) is not None


def _cargar_incidencias_abiertas_camaras(incidencias_db: Session) -> dict[str, list[dict]]:
    rows = incidencias_db.execute(
        text(
            """
            SELECT odt, fecha_registro, problema, detalle_problema, observacion, estado, cliente
            FROM incidencias
            WHERE (detalle_problema IS NOT NULL OR observacion IS NOT NULL)
              AND (
                    estado IS NULL
                 OR (
                        LOWER(CAST(estado AS NVARCHAR(MAX))) NOT LIKE '%terminado%'
                    AND LOWER(CAST(estado AS NVARCHAR(MAX))) NOT LIKE '%repetida%'
                    AND LOWER(CAST(estado AS NVARCHAR(MAX))) NOT LIKE '%rechazad%'
                 )
              )
            ORDER BY fecha_registro DESC
            """
        )
    ).mappings().all()
    por_sucursal: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        if not _incidencia_camara_abierta(row):
            continue
        if not _tokens_detalle_camara(row):
            continue
        clave = _normalizar_texto_cam(row.get("cliente"))
        if clave:
            por_sucursal[clave].append(dict(row))
    return por_sucursal


def _filas_incidencia_sucursal(incidencias_abiertas: dict[str, list[dict]], empresa_norm: str) -> list[dict]:
    if not empresa_norm:
        return []
    exactas = incidencias_abiertas.get(empresa_norm)
    if exactas is not None:
        return exactas
    filas: list[dict] = []
    for sucursal_norm, sucursal_rows in incidencias_abiertas.items():
        if empresa_norm in sucursal_norm or sucursal_norm in empresa_norm:
            filas.extend(sucursal_rows)
    return filas


def _coincidencias_incidencias_camara(
    incidencias_abiertas: dict[str, list[dict]],
    empresa: object,
    camara: object,
    cantidad_equipos: int | None = None,
) -> list[dict]:
    empresa_norm = _normalizar_texto_cam(empresa)
    camara_norm = _normalizar_texto_cam(camara)
    if not empresa_norm:
        return []
    camara_completa_norm = (
        camara_norm if camara_norm.startswith(empresa_norm) else f"{empresa_norm} {camara_norm}".strip()
    )

    coincidencias: list[dict] = []
    for row in _filas_incidencia_sucursal(incidencias_abiertas, empresa_norm):
        for token in _tokens_detalle_camara(row):
            if token == "todo" or re.search(r"\btodo\b", token):
                coincidencias.append(_resumen_incidencia_camara(row, "todo"))
                break
            if _token_equipo_unico_cubre_todo(token, cantidad_equipos):
                coincidencias.append(_resumen_incidencia_camara(row, "equipo_unico"))
                break
            if token.startswith("nvr ") or token.startswith("equipo "):
                continue
            coincide = (
                _texto_contiene_camara(token, camara_norm)
                or _texto_contiene_camara(token, camara_completa_norm)
                or _texto_contiene_camara(camara_norm, token)
            )
            if coincide:
                coincidencias.append(_resumen_incidencia_camara(row, "camara"))
                break
    coincidencias.sort(
        key=lambda c: 0 if "desconex" in _normalizar_texto_cam(c.get("tipo")) else 1
    )
    return coincidencias


@router.get("/api/bitacora/estado-camara")
def estado_camara(
    empresa: str = Query(default=""),
    camara: str = Query(default=""),
    cantidad_equipos: int | None = Query(default=None),
    incidencias_db: Session = Depends(get_incidencias_db),
    current_user: User = Depends(get_current_user_bitacora),
):
    """Cruce best-effort entre una cámara del mapa de 'Administrar puestos' y
    las incidencias reales de la tabla `incidencias` (registradas, entre
    otras vías, desde incidencias_puestos.html con su selector de
    cámaras/Equipo/Todo). No hay FK entre ambos — el cruce es por texto:
    se busca la sucursal (columna `cliente`) y se revisa si la descripción
    de una incidencia abierta menciona "Todo" o el nombre de esta cámara."""
    _require_bitacora_access(current_user)

    empresa_norm = _normalizar_texto_cam(empresa)
    if not empresa_norm:
        raise HTTPException(status_code=400, detail="Falta indicar la empresa/sucursal.")

    coincidencias_encontradas = _coincidencias_incidencias_camara(
        _cargar_incidencias_abiertas_camaras(incidencias_db),
        empresa,
        camara,
        cantidad_equipos=cantidad_equipos,
    )

    if not coincidencias_encontradas:
        return {"estado": "en_linea"}
    return {"estado": "problema", "incidencias": coincidencias_encontradas}
